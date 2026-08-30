"""提取DRUM机械回收来源19个TPUU配方的TGA端点。"""

from __future__ import annotations
import argparse
import hashlib
import json
import tempfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from 提取TGA热稳定端点 import extract_tga_endpoints
from 接入DRUM机械回收 import parse_material_code

ROOT = Path(__file__).resolve().parents[1]
SOURCE = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "DRUM_TPUU_机械回收"
    / "解包内容"
    / "RAW_TPUU_characterization"
    / "TPUU_TGA.xlsx"
)
OUT = ROOT / "结果" / "定向筛选" / "DRUM机械回收TGA端点.csv"
MAN = ROOT / "结果" / "定向筛选" / "DRUM机械回收TGA发布清单.json"


def sha(p):
    d = hashlib.sha256()
    d.update(p.read_bytes())
    return d.hexdigest()


def build_release():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        labels = data[0]
        for col in range(0, ws.max_column - 1, 2):
            if labels[col] is None:
                continue
            material = str(labels[col]).replace("%", "HS")
            curve = pd.DataFrame(
                {
                    "temperature": [r[col] if col < len(r) else None for r in data[2:]],
                    "mass": [
                        r[col + 1] if col + 1 < len(r) else None for r in data[2:]
                    ],
                }
            )
            curve = curve.apply(pd.to_numeric, errors="coerce").dropna()
            ep = extract_tga_endpoints(curve)
            rows.append(
                {
                    "source_id": "source_drum_tpuu_recycling_05ek6k60",
                    "formulation_id": material,
                    **parse_material_code(material),
                    **ep,
                    "chemistry_mapping_status": "family_mn_hard_segment_mapped",
                    "usage_mode": "primary_train",
                    "source_locator": f"{SOURCE.relative_to(ROOT).as_posix()}#sheet={ws.title};material={material}",
                    "license": "CC0-1.0",
                    "citation_keys": "reference-53;reference-54",
                }
            )
    wb.close()
    return pd.DataFrame(rows).sort_values("formulation_id").reset_index(drop=True)


def write(f):
    f.to_csv(OUT, index=False, encoding="utf-8-sig", lineterminator="\n")
    MAN.write_text(
        json.dumps(
            {
                "counts": {
                    "material_count": len(f),
                    "soft_segment_family_count": f.macrodiol_family.nunique(),
                    "t5_count": int(f.T5_degC.notna().sum()),
                    "t10_count": int(f.T10_degC.notna().sum()),
                    "t50_count": int(f.T50_degC.notna().sum()),
                },
                "source_sha256": sha(SOURCE),
                "output_sha256": sha(OUT),
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def check(f):
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / OUT.name
        f.to_csv(p, index=False, encoding="utf-8-sig", lineterminator="\n")
        if sha(p) != sha(OUT):
            raise SystemExit("DRUM TPUU TGA输出不一致")
    print("DRUM TPUU TGA检查通过")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--检查", action="store_true")
    a = p.parse_args()
    f = build_release()
    if a.检查:
        check(f)
    else:
        write(f)
        print(json.dumps({"materials": len(f)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
