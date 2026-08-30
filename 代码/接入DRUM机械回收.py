"""物化DRUM机械回收TPUU来源的独立拉伸端点。"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "DRUM_TPUU_机械回收"
)
AUDIT = SOURCE / "曲线审计清单.tsv"
OUTPUT = ROOT / "结果" / "定向筛选" / "DRUM机械回收拉伸端点.csv"
MANIFEST = ROOT / "结果" / "定向筛选" / "DRUM机械回收发布清单.json"
RELEASE_ID = "tpu-drum-mechanical-recycling-2026-08-30-v1"
SOURCE_ID = "source_drum_tpuu_recycling_05ek6k60"
SOURCE_FAMILY_ID = "family_drum_tpuu_recycling_2024"
CITATION_KEYS = "reference-53;reference-54"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _entry(path: Path) -> dict[str, object]:
    return {
        "path": str(path.relative_to(ROOT)).replace("\\", "/"),
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def parse_material_code(material: str) -> dict[str, object]:
    text = str(material).strip()
    if "14BDO" in text.upper():
        family = "TPU"
        diisocyanate = "IPDI"
        route = "1,4-BDO"
    elif "ELASTOLLAN" in text.upper():
        family = "commercial_TPU"
        diisocyanate = "unknown_commercial"
        route = "unknown_commercial"
    elif "THERMOSET" in text.upper() or "RUBBER" in text.upper():
        family = "adjacent_polyurethane_or_elastomer"
        diisocyanate = "unknown"
        route = "not_applicable_or_unknown"
    else:
        family = "TPUU"
        diisocyanate = "IPDI"
        route = "water_to_urea"
    match = re.search(r"(?P<soft>P4MCL|P4PrCL|PMCL|PCL)-(?P<mn>[0-9.]+)k-(?P<hs>[0-9]+)(?:HS|%)", text, re.I)
    if match:
        macrodiol = match.group("soft")
        mn = float(match.group("mn")) * 1000.0
        hard_segment = float(match.group("hs")) / 100.0
    else:
        macrodiol = pd.NA
        mn = pd.NA
        hard_segment = pd.NA
    return {
        "polymer_family": family,
        "macrodiol_family": macrodiol,
        "macrodiol_nominal_mn_g_mol": mn,
        "hard_segment_mass_fraction": hard_segment,
        "diisocyanate_family": diisocyanate,
        "chain_extension_route": route,
    }


def _sheet_curves(workbook_path: Path, sheet_name: str) -> dict[str, pd.DataFrame]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return {}
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 4:
        workbook.close()
        return {}
    labels = list(rows[0])
    headers = [str(value).strip().lower() if value is not None else "" for value in rows[1]]
    curves: dict[str, pd.DataFrame] = {}
    for stress_index, header in enumerate(headers):
        if header != "stress" or stress_index + 1 >= len(headers):
            continue
        if headers[stress_index + 1] != "strain":
            continue
        group_start = max(0, stress_index - 3)
        label = labels[group_start]
        if label is None:
            for candidate in range(stress_index, max(-1, stress_index - 6), -1):
                if labels[candidate] is not None:
                    label = labels[candidate]
                    break
        if label is None:
            continue
        values = [
            (row[stress_index] if stress_index < len(row) else None,
             row[stress_index + 1] if stress_index + 1 < len(row) else None)
            for row in rows[3:]
        ]
        frame = pd.DataFrame(values, columns=["stress", "strain"])
        frame = frame.apply(pd.to_numeric, errors="coerce").dropna()
        if not frame.empty:
            curves[str(label).strip()] = frame
    workbook.close()
    return curves


def _derive_endpoints(curve: pd.DataFrame) -> dict[str, float | int | str]:
    stress = curve["stress"].to_numpy(dtype=float)
    strain = curve["strain"].to_numpy(dtype=float)
    maximum_index = int(np.nanargmax(strain))
    stress = stress[: maximum_index + 1]
    strain = strain[: maximum_index + 1]
    baseline = float(stress[0])
    corrected = np.maximum(stress - baseline, 0.0)
    relative_strain = np.maximum.accumulate(strain - strain[0]) / 100.0
    toughness = float(np.trapezoid(corrected, relative_strain))
    modulus_mask = (strain >= 0) & (strain <= 5)
    if int(modulus_mask.sum()) >= 3:
        modulus = float(
            np.polyfit(strain[modulus_mask] / 100.0, stress[modulus_mask] - baseline, 1)[0]
        )
    else:
        modulus = float("nan")
    negative_step_fraction = float(np.mean(np.diff(strain) < -0.05)) if len(strain) > 1 else 0.0
    quality = "valid" if negative_step_fraction <= 0.01 and toughness >= 0 else "review_curve_order"
    return {
        "curve_point_count": len(curve),
        "tensile_strength_MPa": float(np.nanmax(stress)),
        "elongation_at_break_percent": float(np.nanmax(strain)),
        "toughness_MJ_m3": toughness,
        "young_modulus_0_5pct_MPa": modulus,
        "negative_strain_step_fraction": negative_step_fraction,
        "endpoint_quality_status": quality,
    }


def _admission(layer: str) -> tuple[str, float]:
    if layer == "核心实验层":
        return "primary_train", 1.0
    if layer.startswith("迁移/"):
        return "auxiliary_train", 0.25 if "机理" in layer else 0.15
    return "reference_only", 0.0


def build_release() -> pd.DataFrame:
    audit = pd.read_csv(AUDIT, sep="\t", low_memory=False)
    tensile = audit[
        audit["试验类型"].eq("单轴拉伸") & audit["准入结论"].eq("准入")
    ].copy()
    records: list[dict[str, object]] = []
    file_hashes: dict[Path, str] = {}
    for relative_path, file_rows in tensile.groupby("文件相对路径", sort=True):
        workbook_path = SOURCE / str(relative_path)
        file_hashes[workbook_path] = _sha256(workbook_path)
        for sheet_name, sheet_rows in file_rows.groupby("工作表", sort=True):
            curves = _sheet_curves(workbook_path, str(sheet_name))
            for row in sheet_rows.itertuples(index=False):
                sample_label = str(getattr(row, "试样标签")).strip()
                if sample_label not in curves:
                    raise ValueError(
                        f"找不到审计曲线：{relative_path}#{sheet_name}#{sample_label}"
                    )
                material = str(getattr(row, "配方键"))
                layer = str(getattr(row, "模型准入层"))
                usage_mode, weight = _admission(layer)
                curve_sha = str(getattr(row, "曲线SHA256"))
                records.append(
                    {
                        "release_id": RELEASE_ID,
                        "source_id": SOURCE_ID,
                        "source_family_id": SOURCE_FAMILY_ID,
                        "observation_id": "drum_tensile_" + curve_sha[:20],
                        "formulation_id": material,
                        "sample_id": str(getattr(row, "试样键")),
                        "batch_id": str(getattr(row, "批次键")),
                        "lifecycle_status": str(getattr(row, "生命周期状态")),
                        "model_admission_layer": layer,
                        "usage_mode": usage_mode,
                        "recommended_loss_weight_ceiling": weight,
                        **parse_material_code(material),
                        **_derive_endpoints(curves[sample_label]),
                        "test_speed": str(getattr(row, "试验速度")),
                        "method_or_test_protocol": str(getattr(row, "方法依据")),
                        "leakage_group": str(getattr(row, "泄漏分组键")),
                        "curve_sha256": curve_sha,
                        "source_file_sha256": file_hashes[workbook_path],
                        "source_locator": (
                            f"{workbook_path.relative_to(ROOT).as_posix()}"
                            f"#sheet={sheet_name};sample={sample_label}"
                        ),
                        "license": "CC0-1.0",
                        "citation_keys": CITATION_KEYS,
                    }
                )
    frame = pd.DataFrame(records).sort_values("observation_id").reset_index(drop=True)
    if len(frame) != len(tensile):
        raise ValueError(f"拉伸曲线物化不完整：{len(frame)} != {len(tensile)}")
    return frame


def _write_csv(frame: pd.DataFrame, path: Path) -> None:
    frame.to_csv(path, index=False, encoding="utf-8-sig", lineterminator="\n")


def _counts(frame: pd.DataFrame) -> dict[str, int]:
    core = frame[frame["model_admission_layer"].eq("核心实验层")]
    return {
        "tensile_curve_rows": len(frame),
        "formulation_count": frame["formulation_id"].nunique(),
        "core_tpuu_curve_rows": len(core),
        "core_tpuu_formulation_count": core["formulation_id"].nunique(),
    }


def _manifest(frame: pd.DataFrame, output: Path) -> dict[str, object]:
    files = sorted(set(frame["source_locator"].str.split("#").str[0]))
    return {
        "release_id": RELEASE_ID,
        "counts": _counts(frame),
        "source": {
            "doi": "10.13020/05ek-6k60",
            "license": "CC0-1.0",
            "audit_file": _entry(AUDIT),
            "input_files": [_entry(ROOT / file) for file in files],
        },
        "output_file": _entry(output),
        "derivation": {
            "toughness": "trapezoid_integral_of_nonnegative_baseline_corrected_stress_vs_relative_strain",
            "elongation_at_break": "maximum_reported_strain",
            "tensile_strength": "maximum_reported_stress",
            "young_modulus": "linear_fit_0_to_5_percent_strain",
        },
    }


def write_release(frame: pd.DataFrame) -> None:
    _write_csv(frame, OUTPUT)
    MANIFEST.write_text(
        json.dumps(_manifest(frame, OUTPUT), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def check_release(frame: pd.DataFrame) -> None:
    if not OUTPUT.is_file() or not MANIFEST.is_file():
        raise SystemExit("缺少DRUM机械回收发布；请先运行生成模式")
    with tempfile.TemporaryDirectory(prefix="drum-recycling-check-") as directory:
        candidate = Path(directory) / OUTPUT.name
        _write_csv(frame, candidate)
        if _sha256(candidate) != _sha256(OUTPUT):
            raise SystemExit("DRUM机械回收端点与当前原件或算法不一致")
    if json.loads(MANIFEST.read_text(encoding="utf-8")) != _manifest(frame, OUTPUT):
        raise SystemExit("DRUM机械回收发布清单不一致")
    print("DRUM机械回收拉伸端点检查通过")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--检查", action="store_true")
    args = parser.parse_args()
    frame = build_release()
    if args.检查:
        check_release(frame)
    else:
        write_release(frame)
        print(json.dumps(_counts(frame), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
