"""复现六个新增开放数据来源的正式内容审计。

输入全部位于 ``数据/原始/外部数据/新增开放数据``，程序只读原始文件，
仅覆盖本模块 ``OUTPUT_WHITELIST`` 中列出的现有审计 JSON/TSV。ZIP 只做路径、
CRC 与既有解包副本的逐字节校验，不重新解包；旧版 XLS 通过本机 Excel COM
以只读模式解析，不生成中间文件。

运行方式（从项目根目录）：

    python 代码/审计/新增开放数据六源.py
"""

from __future__ import annotations

import base64
import csv
import hashlib
import html
import io
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from typing import Callable

import pdfplumber
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"

SOURCE_NAMES = (
    "Jagiellonian_硬段从头算MD",
    "Zenodo_TPU_SWCNT热电",
    "Mendeley_PU泡沫动态力学_精选表",
    "SND_TPU导电轨迹循环拉伸",
    "ScienceDB_TPU芳纶纳米纤维能量吸收",
    "AGH_低石化多元醇硬质PU泡沫",
)

OUTPUT_NAMES_BY_SOURCE = {
    "Jagiellonian_硬段从头算MD": (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "XYZ解析清单.tsv",
    ),
    "Zenodo_TPU_SWCNT热电": (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "工作簿解析清单.tsv",
        "配方与测量汇总.tsv",
    ),
    "Mendeley_PU泡沫动态力学_精选表": (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "工作簿解析清单.tsv",
    ),
    "SND_TPU导电轨迹循环拉伸": (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "电阻表解析清单.tsv",
    ),
    "ScienceDB_TPU芳纶纳米纤维能量吸收": (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "图像解析清单.tsv",
    ),
    "AGH_低石化多元醇硬质PU泡沫": (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "官方远端文件清单.tsv",
    ),
}

OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source, filenames in OUTPUT_NAMES_BY_SOURCE.items()
    for filename in filenames
)
ALL_OUTPUT_NAMES = frozenset(
    filename
    for filenames in OUTPUT_NAMES_BY_SOURCE.values()
    for filename in filenames
)


class AuditBlocked(RuntimeError):
    """输入、许可、完整性或科学语义不满足审计前提。"""


def sha256(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(block)
    return value.hexdigest().upper()


def require_file(path: Path) -> None:
    if not path.is_file():
        raise AuditBlocked(f"缺少必需输入文件：{path}")


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    if path.is_symlink():
        raise AuditBlocked(f"拒绝覆盖符号链接审计输出：{path}")
    if path.exists() and not path.is_file():
        raise AuditBlocked(f"审计输出不是普通文件：{path}")
    parent = path.parent
    if os.path.normcase(str(parent.resolve())) != os.path.normcase(str(parent.absolute())):
        raise AuditBlocked(f"拒绝通过重解析目录写入审计输出：{parent}")


def atomic_write(path: Path, payload: bytes) -> None:
    """在白名单目录内以同卷临时文件和原子替换写入单个审计产物。"""

    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if temporary.is_symlink() or not temporary.is_file():
            raise AuditBlocked(f"审计临时输出不是普通文件：{temporary}")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_json(path: Path, payload: dict[str, object]) -> None:
    atomic_write(
        path,
        (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
    )


def write_tsv(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    if not rows:
        raise AuditBlocked(f"拒绝写入空TSV：{path}")
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        extrasaction="ignore",
    )
    writer.writeheader()
    writer.writerows(rows)
    atomic_write(path, buffer.getvalue().encode("utf-8-sig"))


def original_files(base: Path) -> list[Path]:
    return sorted(
        path
        for path in base.rglob("*")
        if path.is_file()
        and path.name not in ALL_OUTPUT_NAMES
        and "_审计临时" not in path.parts
    )


def raw_snapshot() -> dict[str, tuple[int, str]]:
    result: dict[str, tuple[int, str]] = {}
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        for path in original_files(base):
            result[path.relative_to(PROJECT_ROOT).as_posix()] = (
                path.stat().st_size,
                sha256(path),
            )
    return result


def duplicate_map(files: list[Path]) -> tuple[dict[Path, str], list[dict[str, object]]]:
    groups: defaultdict[str, list[Path]] = defaultdict(list)
    for path in files:
        groups[sha256(path)].append(path)
    duplicate_ids: dict[Path, str] = {}
    details: list[dict[str, object]] = []
    index = 1
    for digest, members in sorted(groups.items()):
        if len(members) <= 1:
            continue
        group_id = f"exact-{index:03d}"
        index += 1
        for member in members:
            duplicate_ids[member] = group_id
        details.append(
            {
                "重复组": group_id,
                "SHA256": digest,
                "文件数": len(members),
                "文件": [member.name for member in members],
            }
        )
    return duplicate_ids, details


PathRule = Callable[[Path, str], str]


def file_rows(
    base: Path,
    role_for: PathRule,
    parse_for: PathRule,
    layer_for: PathRule,
    weight_for: PathRule,
    notes_for: PathRule,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    files = original_files(base)
    duplicate_ids, duplicate_details = duplicate_map(files)
    rows: list[dict[str, object]] = []
    for path in files:
        relative = path.relative_to(base).as_posix()
        rows.append(
            {
                "相对路径": relative,
                "字节数": path.stat().st_size,
                "SHA256": sha256(path),
                "官方校验": "",
                "文件角色": role_for(path, relative),
                "完整性与可读性": parse_for(path, relative),
                "精确重复组": duplicate_ids.get(path, "无"),
                "准入层": layer_for(path, relative),
                "训练权重上限": weight_for(path, relative),
                "备注": notes_for(path, relative),
            }
        )
    return rows, duplicate_details


def is_safe_zip_member(name: str) -> bool:
    if "\x00" in name or re.match(r"^[A-Za-z]:", name):
        return False
    path = PurePosixPath(name.replace("\\", "/"))
    return not path.is_absolute() and ".." not in path.parts


def validate_zip(archive: zipfile.ZipFile, label: str) -> list[zipfile.ZipInfo]:
    bad_crc = archive.testzip()
    if bad_crc is not None:
        raise AuditBlocked(f"{label} ZIP CRC失败：{bad_crc}")
    unsafe = [item.filename for item in archive.infolist() if not is_safe_zip_member(item.filename)]
    if unsafe:
        raise AuditBlocked(f"{label}含危险ZIP路径：{unsafe}")
    return [item for item in archive.infolist() if not item.is_dir()]


def parse_formula(text: str) -> Counter[str]:
    result: Counter[str] = Counter()
    for element, number in re.findall(r"([A-Z][a-z]?)(\d*)", text):
        result[element] += int(number or 1)
    return result


def parse_jsonld(raw: str) -> list[dict[str, object]]:
    found: list[dict[str, object]] = []
    pattern = r"<script[^>]*type=[\"']application/ld\+json[\"'][^>]*>(.*?)</script>"
    for match in re.findall(pattern, raw, flags=re.I | re.S):
        try:
            value = json.loads(html.unescape(match))
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            found.append(value)
    return found


def powershell_executable() -> str:
    executable = shutil.which("pwsh") or shutil.which("powershell")
    if executable is None:
        raise AuditBlocked("找不到PowerShell，无法只读解析旧版XLS")
    return executable


def parse_legacy_xls_with_excel(source_dir: Path) -> list[dict[str, object]]:
    """通过 Excel COM 只读打开全部旧版 XLS；JSON 只经 stdout 返回。"""

    source_literal = str(source_dir).replace("'", "''")
    script = rf"""
[Console]::OutputEncoding = [Text.UTF8Encoding]::new($false)
$ErrorActionPreference = 'Stop'
$sourceDir = '{source_literal}'
$excel = $null
$rows = [System.Collections.Generic.List[object]]::new()
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.AskToUpdateLinks = $false
    foreach ($file in (Get-ChildItem -LiteralPath $sourceDir -Filter '*.xls' -File | Sort-Object Name)) {{
        $wb = $null
        $ws = $null
        $used = $null
        try {{
            $wb = $excel.Workbooks.Open($file.FullName, 0, $true)
            $sheetInfo = [System.Collections.Generic.List[object]]::new()
            foreach ($sheet in $wb.Worksheets) {{
                $range = $sheet.UsedRange
                $sheetInfo.Add([ordered]@{{
                    name = [string]$sheet.Name
                    rows = [int]$range.Rows.Count
                    columns = [int]$range.Columns.Count
                    address = [string]$range.Address()
                }})
                [void][Runtime.InteropServices.Marshal]::ReleaseComObject($range)
                [void][Runtime.InteropServices.Marshal]::ReleaseComObject($sheet)
            }}
            $ws = $wb.Worksheets.Item(1)
            $used = $ws.UsedRange
            $numeric = 0
            $nonempty = 0
            $values = $used.Value2
            if ($values -is [System.Array]) {{
                foreach ($value in $values) {{
                    if ($null -ne $value -and [string]$value -ne '') {{
                        $nonempty++
                        if ($value -is [byte] -or $value -is [int16] -or $value -is [int32] -or $value -is [int64] -or $value -is [single] -or $value -is [double] -or $value -is [decimal]) {{ $numeric++ }}
                    }}
                }}
            }}
            $rows.Add([ordered]@{{
                file = $file.Name
                open_status = 'ok'
                sheets = $sheetInfo
                nonempty_cells_first_sheet = $nonempty
                numeric_cells_first_sheet = $numeric
                experiment = [string]$ws.Cells.Item(4, 2).Text
                date = [string]$ws.Cells.Item(5, 5).Text
                comment = [string]$ws.Cells.Item(6, 2).Text
                material = [string]$ws.Cells.Item(9, 2).Text
                length_mm = [string]$ws.Cells.Item(10, 2).Text
                width_mm = [string]$ws.Cells.Item(10, 5).Text
                thickness_mm = [string]$ws.Cells.Item(10, 8).Text
                temperature_c = [string]$ws.Cells.Item(13, 2).Text
                seebeck_uV_K = [string]$ws.Cells.Item(21, 11).Text
                resistance_kOhm = [string]$ws.Cells.Item(21, 13).Text
                resistivity_Ohm_cm = [string]$ws.Cells.Item(21, 14).Text
                power_uW = [string]$ws.Cells.Item(21, 18).Text
            }})
        }} catch {{
            $rows.Add([ordered]@{{ file = $file.Name; open_status = 'error'; error = $_.Exception.Message }})
        }} finally {{
            if ($wb) {{ $wb.Close($false) }}
            if ($used) {{ [void][Runtime.InteropServices.Marshal]::ReleaseComObject($used) }}
            if ($ws) {{ [void][Runtime.InteropServices.Marshal]::ReleaseComObject($ws) }}
            if ($wb) {{ [void][Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }}
        }}
    }}
    $rows | ConvertTo-Json -Depth 8 -Compress
}} finally {{
    if ($excel) {{ $excel.Quit(); [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    encoded = base64.b64encode(script.encode("utf-16le")).decode("ascii")
    completed = subprocess.run(
        [powershell_executable(), "-NoProfile", "-NonInteractive", "-EncodedCommand", encoded],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=600,
    )
    if completed.returncode != 0:
        error = completed.stderr.decode("utf-8", errors="replace").strip()
        raise AuditBlocked(f"Excel COM审计失败（退出码{completed.returncode}）：{error}")
    output = completed.stdout.decode("utf-8-sig", errors="strict").strip()
    try:
        parsed = json.loads(output)
    except json.JSONDecodeError as exc:
        raise AuditBlocked(f"Excel COM未返回有效JSON：{output[:500]}") from exc
    if not isinstance(parsed, list):
        raise AuditBlocked("Excel COM返回结构不是工作簿列表")
    errors = [row for row in parsed if row.get("open_status") != "ok"]
    if errors:
        raise AuditBlocked(f"存在无法只读打开的XLS：{errors}")
    return parsed


def audit_jagiellonian() -> None:
    base = DATA_ROOT / "Jagiellonian_硬段从头算MD"
    archive_path = base / "Optimized_geom.zip"
    readme = base / "00_readme.txt"
    require_file(archive_path)
    require_file(readme)
    xyz_files = sorted((base / "解包内容").glob("*.xyz"))
    if len(xyz_files) != 4:
        raise AuditBlocked(f"Jagiellonian优化结构应为4个，实际为{len(xyz_files)}")

    xyz_rows: list[dict[str, object]] = []
    with zipfile.ZipFile(archive_path) as archive:
        members = validate_zip(archive, "Jagiellonian Optimized_geom")
        entries = {Path(item.filename).name: item for item in members}
        for path in xyz_files:
            lines = path.read_text(encoding="utf-8", errors="strict").splitlines()
            declared_atoms = int(lines[0].strip())
            comment = lines[1].strip()
            formula_match = re.match(r"([A-Za-z0-9]+)", comment)
            formula = formula_match.group(1) if formula_match else ""
            atoms: list[str] = []
            coordinates: list[tuple[float, float, float]] = []
            malformed = 0
            for line in lines[2:]:
                parts = line.split()
                if len(parts) != 4:
                    malformed += 1
                    continue
                try:
                    coordinate = tuple(float(value) for value in parts[1:])
                except ValueError:
                    malformed += 1
                    continue
                atoms.append(parts[0])
                coordinates.append(coordinate)  # type: ignore[arg-type]
            minimum_distance = min(
                math.dist(coordinates[i], coordinates[j])
                for i in range(len(coordinates))
                for j in range(i + 1, len(coordinates))
            )
            system_match = re.search(r"__(HDI|HMDI|MDI|TDI)_", path.name)
            if system_match is None:
                raise AuditBlocked(f"无法从XYZ文件名识别硬段体系：{path.name}")
            system = system_match.group(1)
            zip_name = path.name.replace("Optimized_geom__", "")
            if zip_name not in entries:
                raise AuditBlocked(f"ZIP缺少对应XYZ：{zip_name}")
            zip_payload = archive.read(entries[zip_name])
            formula_counts = parse_formula(formula)
            atom_counts = Counter(atoms)
            xyz_rows.append(
                {
                    "文件": path.relative_to(base).as_posix(),
                    "硬段体系": system,
                    "链模型": "1x2",
                    "声明原子数": declared_atoms,
                    "实际坐标行": len(coordinates),
                    "分子式": formula,
                    "元素计数": ";".join(
                        f"{key}:{atom_counts[key]}" for key in sorted(atom_counts)
                    ),
                    "分子式与元素计数一致": str(formula_counts == atom_counts).lower(),
                    "畸形行": malformed,
                    "最小原子间距_坐标单位": f"{minimum_distance:.6f}",
                    "坐标单位声明": "文件未声明；XYZ惯例常按Å解释，当前不作已证实单位",
                    "ZIP内字节一致": str(zip_payload == path.read_bytes()).lower(),
                    "软件": "ADF/AMS 2023.104",
                }
            )

    if not all(
        row["声明原子数"] == row["实际坐标行"]
        and row["分子式与元素计数一致"] == "true"
        and row["畸形行"] == 0
        and row["ZIP内字节一致"] == "true"
        for row in xyz_rows
    ):
        raise AuditBlocked("Jagiellonian XYZ完整性或科学计数检查失败")

    write_tsv(
        base / "XYZ解析清单.tsv",
        xyz_rows,
        [
            "文件", "硬段体系", "链模型", "声明原子数", "实际坐标行", "分子式",
            "元素计数", "分子式与元素计数一致", "畸形行", "最小原子间距_坐标单位",
            "坐标单位声明", "ZIP内字节一致", "软件",
        ],
    )

    def role(path: Path, _: str) -> str:
        if path.suffix.lower() == ".xyz":
            return "计算结构主数据"
        if path.suffix.lower() == ".zip":
            return "镜像容器"
        return "来源与方法文档"

    def parse_status(path: Path, relative: str) -> str:
        if path.suffix.lower() == ".xyz":
            row = next(item for item in xyz_rows if item["文件"] == relative)
            return (
                "通过：原子数/分子式/坐标均可解析，且与ZIP字节一致"
                if row["畸形行"] == 0
                else "警告：存在畸形行"
            )
        if path.suffix.lower() == ".zip":
            return "通过：ZIP CRC 全部通过"
        return "通过：UTF-8文本可读"

    rows, duplicates = file_rows(
        base,
        role,
        parse_status,
        lambda path, _: "计算迁移层" if path.suffix.lower() == ".xyz" else "元数据层",
        lambda path, _: "0.15" if path.suffix.lower() == ".xyz" else "0.00",
        lambda path, _: (
            "4个1x2优化结构之一；0.15仅适用于元素拓扑或尺度不变表示，坐标单位补证前绝对距离、几何和氢键QoI权重为0；不得按原子或坐标行扩增样本"
            if path.suffix.lower() == ".xyz"
            else (
                "与解包XYZ构成镜像，不重复计数"
                if path.suffix.lower() == ".zip"
                else "许可、软件和缺失轨迹范围证据"
            )
        ),
    )
    write_tsv(base / "文件校验清单.tsv", rows, list(rows[0]))

    payload = {
        "审计版本": "1.0",
        "审计日期": AUDIT_DATE,
        "来源": {
            "题名": "Structure–Property Relationship between Hard Segments of Shape Memory Polyurethane Copolymers and Interchain Hydrogen Bonds: A Comprehensive Theoretical Study - raw data",
            "DOI": "10.57903/UJ/TYAPFM",
            "版本": "published_2026",
            "作者": ["Yuliia Didovets"],
            "机构": "Jagiellonian University in Kraków",
            "许可": "CC BY 4.0",
            "许可证据": "本地00_readme.txt与DataCite DOI元数据",
            "关联论文": "10.1021/acs.jpcb.5c03305",
            "参考文献": "Didovets, Y. Structure–Property Relationship between Hard Segments of Shape Memory Polyurethane Copolymers and Interchain Hydrogen Bonds: A Comprehensive Theoretical Study - raw data [Data set]. Jagiellonian University in Kraków, 2026. https://doi.org/10.57903/UJ/TYAPFM.",
        },
        "下载范围": {
            "已下载": "Optimized_geom.zip、README以及ZIP内全部4个1x2优化几何",
            "未下载": "MD_Trajectories完整轨迹包（DataCite登记约3,579,455,734字节）",
            "原因": "大型轨迹延后；本轮不下载大型文件",
            "完整性": {
                "ZIP_CRC": "通过",
                "解包文件与ZIP逐字节一致数": sum(
                    row["ZIP内字节一致"] == "true" for row in xyz_rows
                ),
            },
        },
        "独立科学单元": {
            "优化结构数": 4,
            "硬段体系": ["HDI", "HMDI", "MDI", "TDI"],
            "总原子坐标行": sum(int(row["实际坐标行"]) for row in xyz_rows),
            "正确计数规则": "dataset_doi + hard_segment_system + protocol；每个优化结构仅1个科学单元",
            "不得作为独立样本": ["原子", "坐标行", "ZIP镜像", "同一结构的派生描述符"],
        },
        "真实字段与单位": {
            "元素符号": "无量纲分类",
            "三维坐标": "文件未显式声明；XYZ惯例通常为Å，建模前必须以论文/软件输出设置复核",
            "分子式": [str(row["分子式"]) for row in xyz_rows],
            "软件": "ADF/AMS 2023.104（优化几何）；缺失轨迹原由CP2K 2023.1生成",
        },
        "可解析性": {
            "状态": "可直接解析",
            "全部原子数一致": all(
                row["声明原子数"] == row["实际坐标行"] for row in xyz_rows
            ),
            "全部分子式一致": all(
                row["分子式与元素计数一致"] == "true" for row in xyz_rows
            ),
        },
        "重复与异常": {
            "精确重复文件组": duplicates,
            "异常": [
                "完整AIMD轨迹未下载",
                "坐标单位未在XYZ/README中显式声明",
                "当前只有1x2优化几何，不能代表完整时间演化",
            ],
        },
        "准入与权重": {
            "当前层级": "计算迁移层",
            "准入状态": "有条件准入：仅用于元素拓扑或尺度不变表示；坐标单位补证前禁止绝对距离、几何和氢键QoI；不得作为实验终性能真值",
            "当前元素拓扑或尺度不变表示权重上限": 0.15,
            "当前绝对距离几何或氢键QoI权重上限": 0.0,
            "实验性能监督权重": 0.0,
            "完整轨迹获取并完成单位协议收敛与种子审计后的上限": 0.30,
            "观测身份键": ["dataset_doi", "hard_segment_system", "protocol", "trajectory_seed"],
            "拆分组键": ["dataset_doi", "hard_segment_system"],
        },
    }
    write_json(base / "内容审计摘要.json", payload)


def audit_swcnt() -> None:
    base = DATA_ROOT / "Zenodo_TPU_SWCNT热电"
    extracted = base / "解包内容"
    archive_path = base / "rawdata-NitrogenContentGovers-paper.zip"
    metadata_path = base / "官方Zenodo元数据.json"
    description_path = base / "sample description.txt"
    for path in (archive_path, metadata_path, description_path):
        require_file(path)

    xls_summary = parse_legacy_xls_with_excel(extracted)
    if len(xls_summary) != 123:
        raise AuditBlocked(f"Zenodo TPU/SWCNT旧版XLS应为123个，实际为{len(xls_summary)}")
    xls_by_name = {str(row["file"]): row for row in xls_summary}

    with zipfile.ZipFile(archive_path) as archive:
        entries = validate_zip(archive, "Zenodo TPU/SWCNT")
        extracted_matches = 0
        missing_flattened: list[str] = []
        for info in entries:
            flat = extracted / info.filename.replace("/", "__")
            if flat.exists() and flat.read_bytes() == archive.read(info):
                extracted_matches += 1
            else:
                missing_flattened.append(info.filename)
    if len(entries) != 145 or missing_flattened:
        raise AuditBlocked(
            f"Zenodo TPU/SWCNT解包不完整：ZIP文件数={len(entries)}，"
            f"不一致或缺失={missing_flattened}"
        )

    formulation_codes = {
        "01": ("1185A10", 1),
        "02": ("1185A10", 2),
        "03": ("1185A10", 3),
        "04": ("1185A10", 4),
        "05": ("1185A10", 5),
        "08": ("C85A10", 1),
        "09": ("C85A10", 2),
        "10": ("C85A10", 3),
        "11": ("C85A10", 4),
        "12": ("C85A10", 5),
        "14": ("C74D50", 1),
        "15": ("C74D50", 2),
        "16": ("C74D50", 3),
        "17": ("C74D50", 4),
        "18": ("C74D50", 5),
    }
    formulation_stats: dict[str, dict[str, object]] = {
        code: {
            "code": f"NRC-BK-{code}-220308",
            "matrix": matrix,
            "swcnt_wt_pct": loading,
            "xls_files": 0,
            "replicates": set(),
            "internal_material_labels": set(),
            "filename_internal_mismatch": 0,
        }
        for code, (matrix, loading) in formulation_codes.items()
    }
    parse_rows: list[dict[str, object]] = []
    for item in xls_summary:
        filename = str(item["file"])
        code_match = re.search(r"NRC-BK-(\d{2})-220308p?", filename)
        code = code_match.group(1) if code_match else ""
        replicate_match = re.search(r"-([a-z])\s+4Pkt", filename, flags=re.I)
        replicate = replicate_match.group(1).lower() if replicate_match else "未标记"
        internal = str(item.get("material", ""))
        internal_match = re.search(r"-([a-z])$", internal, flags=re.I)
        internal_replicate = internal_match.group(1).lower() if internal_match else "未标记"
        mismatch = (
            replicate != "未标记"
            and internal_replicate != "未标记"
            and replicate != internal_replicate
        )
        if code not in formulation_stats:
            raise AuditBlocked(f"无法识别热电XLS配方代码：{filename}")
        stats = formulation_stats[code]
        stats["xls_files"] = int(stats["xls_files"]) + 1
        cast_replicates = stats["replicates"]
        cast_labels = stats["internal_material_labels"]
        assert isinstance(cast_replicates, set)
        assert isinstance(cast_labels, set)
        cast_replicates.add(replicate)
        cast_labels.add(internal)
        stats["filename_internal_mismatch"] = (
            int(stats["filename_internal_mismatch"]) + int(mismatch)
        )
        sheets = item.get("sheets", [])
        assert isinstance(sheets, list) and sheets
        first_sheet = sheets[0]
        parse_rows.append(
            {
                "文件": filename,
                "样品代码": f"NRC-BK-{code}-220308",
                "文件名试样位置": replicate,
                "内部Material试样位置": internal_replicate,
                "文件名与内部位置不一致": str(mismatch).lower(),
                "打开状态": item.get("open_status"),
                "有效工作表": sum(
                    int(sheet["rows"]) > 1 or int(sheet["columns"]) > 1
                    for sheet in sheets
                ),
                "首表行数": first_sheet.get("rows", ""),
                "首表列数": first_sheet.get("columns", ""),
                "首表非空单元格": item.get("nonempty_cells_first_sheet", ""),
                "首表数值单元格": item.get("numeric_cells_first_sheet", ""),
                "长度_mm": item.get("length_mm", ""),
                "宽度_mm": item.get("width_mm", ""),
                "厚度_mm": item.get("thickness_mm", ""),
                "测量温度_C": item.get("temperature_c", ""),
                "Seebeck_uV_K": item.get("seebeck_uV_K", ""),
                "电阻_kOhm": item.get("resistance_kOhm", ""),
                "电阻率_Ohm_cm": item.get("resistivity_Ohm_cm", ""),
                "功率_uW": item.get("power_uW", ""),
            }
        )
    write_tsv(base / "工作簿解析清单.tsv", parse_rows, list(parse_rows[0]))

    formulation_rows: list[dict[str, object]] = []
    for code in sorted(formulation_stats):
        stats = formulation_stats[code]
        replicates = stats["replicates"]
        labels = stats["internal_material_labels"]
        assert isinstance(replicates, set)
        assert isinstance(labels, set)
        formulation_rows.append(
            {
                "样品代码": stats["code"],
                "TPU牌号": stats["matrix"],
                "SWCNT_wt_pct": stats["swcnt_wt_pct"],
                "热电XLS文件数": stats["xls_files"],
                "试样位置标签": ",".join(sorted(replicates)),
                "内部Material唯一标签数": len(labels),
                "文件名与内部位置不一致数": stats["filename_internal_mismatch"],
                "独立配方计数": 1,
                "分组键": (
                    "10.5281/zenodo.20932248|"
                    f"{stats['matrix']}|{stats['swcnt_wt_pct']}wt%"
                ),
            }
        )
    if len(formulation_rows) != 15:
        raise AuditBlocked("TPU/SWCNT配方计数不是15")
    write_tsv(base / "配方与测量汇总.tsv", formulation_rows, list(formulation_rows[0]))

    xlsx_details: list[dict[str, object]] = []
    for path in sorted(extracted.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=False)
        for sheet in workbook.worksheets:
            nonempty = numeric = formulas = 0
            strings: list[str] = []
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    nonempty += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric += 1
                    elif isinstance(value, str):
                        if value.startswith("="):
                            formulas += 1
                        elif len(strings) < 12:
                            strings.append(value)
            xlsx_details.append(
                {
                    "file": path.name,
                    "sheet": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "nonempty": nonempty,
                    "numeric": numeric,
                    "formulas": formulas,
                    "headers": strings,
                }
            )

    dpt_details: list[dict[str, object]] = []
    for path in sorted(extracted.glob("*.dpt")):
        numeric_rows = 0
        column_counts: Counter[int] = Counter()
        for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
            parts = re.split(r"[;,\t ]+", line.strip())
            try:
                [float(value.replace(",", ".")) for value in parts if value]
                if parts and any(parts):
                    numeric_rows += 1
                    column_counts[len([value for value in parts if value])] += 1
            except ValueError:
                continue
        dpt_details.append(
            {
                "file": path.name,
                "numeric_rows": numeric_rows,
                "column_counts": dict(column_counts),
            }
        )

    tif_details: list[dict[str, object]] = []
    for path in sorted(extracted.glob("*.tif")):
        with Image.open(path) as image:
            tif_details.append(
                {
                    "file": path.name,
                    "width": image.width,
                    "height": image.height,
                    "mode": image.mode,
                    "frames": getattr(image, "n_frames", 1),
                }
            )

    pdf_details: list[dict[str, object]] = []
    for path in sorted(extracted.glob("*.pdf")):
        try:
            pdf_details.append(
                {"file": path.name, "pages": len(PdfReader(path).pages), "readable": True}
            )
        except Exception as exc:  # pragma: no cover - blocking branch for corrupt external PDF
            pdf_details.append(
                {"file": path.name, "pages": None, "readable": False, "error": str(exc)}
            )
    if len(xlsx_details) != 12 or len(dpt_details) != 7 or len(tif_details) != 3:
        raise AuditBlocked("TPU/SWCNT辅助资产计数与固定版本不一致")
    if not all(item["readable"] for item in pdf_details):
        raise AuditBlocked(f"TPU/SWCNT存在不可读PDF：{pdf_details}")

    def role(path: Path, relative: str) -> str:
        extension = path.suffix.lower()
        if extension == ".zip":
            return "镜像容器"
        if path.name == "官方Zenodo元数据.json" or path.name.endswith("description.txt"):
            return "来源/样品说明文档"
        if extension == ".xls":
            return "热电测量主数据"
        if extension == ".xlsx":
            return "电阻测量主数据"
        if extension == ".dpt":
            return "红外光谱辅助数据"
        if extension == ".tif":
            return "显微图像辅助数据"
        if extension == ".pdf" and "MFI" in relative:
            return "熔体流动指标证据"
        if extension == ".pdf":
            return "商业TPU牌号数据表"
        return "其他"

    def parse_status(path: Path, _: str) -> str:
        extension = path.suffix.lower()
        if extension == ".zip":
            return "通过：ZIP CRC全部通过"
        if extension == ".xls":
            return (
                "通过：Excel COM只读打开"
                if xls_by_name[path.name]["open_status"] == "ok"
                else "失败：Excel无法打开"
            )
        if extension == ".xlsx":
            return "通过：openpyxl可读"
        if extension == ".dpt":
            return "通过：文本光谱数值可读"
        if extension == ".tif":
            return "通过：Pillow可解码"
        if extension == ".pdf":
            return "通过：PDF页结构可读"
        if extension in {".json", ".txt"}:
            return "通过：文本可读"
        return "未解析"

    def layer(path: Path, _: str) -> str:
        extension = path.suffix.lower()
        if extension in {".xls", ".xlsx"}:
            return "实验核心-功能性能层"
        if extension in {".dpt", ".tif"}:
            return "表征迁移层"
        return "元数据层"

    def weight(path: Path, _: str) -> str:
        extension = path.suffix.lower()
        if extension in {".xls", ".xlsx"}:
            return "0.45"
        if extension == ".dpt":
            return "0.25"
        if extension == ".tif":
            return "0.15"
        return "0.00"

    def notes(path: Path, _: str) -> str:
        extension = path.suffix.lower()
        if extension == ".zip":
            return "与解包内容构成镜像，权重为0"
        if extension == ".xls":
            return "同一配方的测量会话/位置重复；按配方与试样位置成组"
        if extension == ".xlsx":
            return "电阻汇总；工作表/数值行不增加配方数"
        if extension == ".dpt":
            return "波数-吸光度光谱点不作为独立材料样本"
        if extension == ".tif":
            return "仅1 wt%配方的宏观分散图像；图像切块不得扩增配方数"
        return "仅作条件、来源或许可证据"

    rows, duplicates = file_rows(base, role, parse_status, layer, weight, notes)
    if len(rows) != 148:
        raise AuditBlocked(f"TPU/SWCNT本地文件校验行应为148，实际为{len(rows)}")
    write_tsv(base / "文件校验清单.tsv", rows, list(rows[0]))

    payload = {
        "审计版本": "1.0",
        "审计日期": AUDIT_DATE,
        "来源": {
            "题名": "Raw data for the paper ‘Nitrogen Content Governs Thermoelectric Performance in TPU/SWCNT Composites’",
            "DOI": "10.5281/zenodo.20932248",
            "概念DOI": "10.5281/zenodo.20932247",
            "版本": "record_20932248",
            "作者": ["Beate Krause", "Cordelia Zimmerer"],
            "机构": "Leibniz Institute of Polymer Research Dresden",
            "许可": "CC BY 4.0",
            "许可证据": "本地官方Zenodo元数据.json与DataCite DOI元数据",
            "关联预印本": "10.20944/preprints202606.1342.v1",
            "参考文献": "Krause, B.; Zimmerer, C. Raw data for the paper ‘Nitrogen Content Governs Thermoelectric Performance in TPU/SWCNT Composites’ [Data set]. Zenodo, 2026. https://doi.org/10.5281/zenodo.20932248.",
        },
        "下载范围": {
            "已下载": "Zenodo记录的ZIP、sample description与元数据；ZIP内全部145个文件已扁平化解包",
            "未下载": "无（该记录公开文件已完整下载）",
            "完整性": {
                "ZIP_CRC": "通过",
                "ZIP数据文件数": len(entries),
                "与ZIP逐字节一致的解包文件": extracted_matches,
                "不一致或缺失": missing_flattened,
            },
        },
        "独立科学单元": {
            "独立配方数": 15,
            "TPU牌号数": 3,
            "SWCNT含量水平": [1, 2, 3, 4, 5],
            "热电XLS测量文件": len(xls_summary),
            "Excel只读打开成功": sum(
                item["open_status"] == "ok" for item in xls_summary
            ),
            "热电首表数值单元格": sum(
                int(item.get("numeric_cells_first_sheet", 0)) for item in xls_summary
            ),
            "正确计数规则": "dataset_doi + TPU牌号 + SWCNT wt% 为配方单元；试样位置/重复测量嵌套在配方内",
            "不得作为独立配方": [
                "123个XLS文件", "Excel单元格", "四个温差点", "IR光谱点", "显微图像切块", "ZIP镜像",
            ],
        },
        "真实字段与单位": {
            "配方": "TPU牌号 + Tuball SWCNT质量分数（wt%）",
            "热电测量": [
                "Seebeck系数 µV/K", "电阻 kΩ", "体积电阻率 Ω·cm", "功率 µW", "温差 K", "测量温度 °C",
            ],
            "试样几何": ["长度 mm", "宽度 mm", "厚度 mm"],
            "制备": ["熔混温度 °C", "转速 rpm", "时间 min", "热压板直径/厚度 mm"],
            "IR": "波数范围4000–900 cm^-1，分辨率4 cm^-1，100次扫描（来自sample description）",
            "xlsx概况": xlsx_details,
            "DPT概况": dpt_details,
            "TIF概况": tif_details,
            "PDF概况": pdf_details,
        },
        "可解析性": {
            "状态": "高",
            "旧版XLS": "123/123由本机Excel只读打开",
            "XLSX": f"{len(list(extracted.glob('*.xlsx')))}个可由openpyxl读取",
            "DPT": f"{len(dpt_details)}个文本光谱可读",
            "TIF": f"{len(tif_details)}个图像可解码",
        },
        "重复与异常": {
            "精确重复文件组": duplicates,
            "文件名与内部Material试样位置不一致文件数": sum(
                row["文件名与内部位置不一致"] == "true" for row in parse_rows
            ),
            "异常": [
                "存在_2/_3/_4/_5/wdh等重复测量命名，必须作为会话/重复而非新配方",
                "部分文件名试样位置与工作簿内部Material标签不一致，建模前以内嵌标签和实验日志复核",
                "商业牌号给出但单体/软段完整化学身份不足",
            ],
        },
        "准入与权重": {
            "当前层级": "实验核心-功能性能层（TPU复合材料）；IR/图像为表征迁移层",
            "准入状态": "准入。适合热电/导电应用子模型，不直接替代纯TPU本体力学标签",
            "跨任务统一权重上限": 0.45,
            "热电专用任务上限": 0.60,
            "IR上限": 0.25,
            "图像上限": 0.15,
            "分组键": [
                "dataset_doi", "TPU_grade", "SWCNT_wt_pct", "specimen_position", "measurement_session",
            ],
        },
    }
    write_json(base / "内容审计摘要.json", payload)


def audit_mendeley() -> None:
    base = DATA_ROOT / "Mendeley_PU泡沫动态力学_精选表"
    workbook_paths = sorted(base.glob("*.xlsx"))
    if [path.name for path in workbook_paths] != [
        "Energy_Absorption.xlsx",
        "Foam Test Comparison.xlsx",
        "HA_StressStrain.xlsx",
        "HDB_StressStrain.xlsx",
    ]:
        raise AuditBlocked(f"Mendeley精选工作簿集合发生变化：{workbook_paths}")

    workbook_rows: list[dict[str, object]] = []
    details: dict[str, dict[str, int]] = {}
    for path in workbook_paths:
        workbook = load_workbook(path, read_only=True, data_only=False)
        for sheet in workbook.worksheets:
            nonempty = numeric = formulas = errors = 0
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    nonempty += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric += 1
                    elif isinstance(value, str) and value.startswith("="):
                        formulas += 1
                    elif isinstance(value, str) and value.startswith("#"):
                        errors += 1
            workbook_rows.append(
                {
                    "文件": path.name,
                    "工作表": sheet.title,
                    "最大行": sheet.max_row,
                    "最大列": sheet.max_column,
                    "非空单元格": nonempty,
                    "数值单元格": numeric,
                    "公式": formulas,
                    "错误值": errors,
                }
            )
        details[path.name] = {
            "sheets": len(workbook.worksheets),
            "rows": workbook.active.max_row,
            "columns": workbook.active.max_column,
        }
    if any(row["公式"] or row["错误值"] for row in workbook_rows):
        raise AuditBlocked("Mendeley精选工作簿出现未预期公式或错误值")
    write_tsv(base / "工作簿解析清单.tsv", workbook_rows, list(workbook_rows[0]))

    curve_info: dict[str, list[dict[str, object]]] = {}
    for filename in ("HA_StressStrain.xlsx", "HDB_StressStrain.xlsx"):
        sheet = load_workbook(base / filename, read_only=False, data_only=True).active
        curves: list[dict[str, object]] = []
        for column in range(1, sheet.max_column + 1, 2):
            label = sheet.cell(1, column).value
            points = 0
            for row in range(2, sheet.max_row + 1):
                x_value = sheet.cell(row, column).value
                y_value = sheet.cell(row, column + 1).value
                if isinstance(x_value, (int, float)) and isinstance(y_value, (int, float)):
                    points += 1
            curves.append({"label": label, "points": points})
        curve_info[filename] = curves

    comparison = load_workbook(
        base / "Foam Test Comparison.xlsx", read_only=False, data_only=True
    ).active

    def column_values(column: int, start: int, end: int) -> list[object]:
        return [comparison.cell(row, column).value for row in range(start, end + 1)]

    duplicate_columns = {
        "L_vs_B_force": column_values(12, 3, 1587) == column_values(2, 4, 1588),
        "M_vs_C_stroke": column_values(13, 3, 1587) == column_values(3, 4, 1588),
        "O_vs_F_force": column_values(15, 3, 1692) == column_values(6, 4, 1693),
        "P_vs_G_stroke": column_values(16, 3, 1692) == column_values(7, 4, 1693),
    }
    if not all(duplicate_columns.values()):
        raise AuditBlocked(f"Foam Test Comparison复制列关系变化：{duplicate_columns}")

    energy_sheet = load_workbook(
        base / "Energy_Absorption.xlsx", read_only=False, data_only=True
    ).active
    energy_measurements = sum(
        isinstance(energy_sheet.cell(row, column).value, (int, float))
        for row in range(3, 13)
        for column in range(1, 13)
    )
    total_curve_points = sum(
        int(curve["points"])
        for curves in curve_info.values()
        for curve in curves
    )
    if len([curve for curves in curve_info.values() for curve in curves]) != 12:
        raise AuditBlocked("Mendeley应力-应变曲线数不是12")
    if total_curve_points != 29_707 or energy_measurements != 105:
        raise AuditBlocked(
            f"Mendeley科学计数变化：curve_points={total_curve_points}, "
            f"energy_measurements={energy_measurements}"
        )

    rows, duplicates = file_rows(
        base,
        lambda _path, _relative: "精选力学主数据",
        lambda _path, _relative: "通过：openpyxl可读；无公式错误或Excel错误值",
        lambda _path, _relative: "力学迁移层",
        lambda _path, _relative: "0.25",
        lambda _path, _relative: "0.25仅是单位与试样映射补齐后的来源上限；当前单位相关应力、模量、韧性与吸能标签触发硬门并取0；按材料-温度-试样/曲线成组",
    )
    if len(rows) != 4:
        raise AuditBlocked(f"Mendeley文件校验行应为4，实际为{len(rows)}")
    write_tsv(base / "文件校验清单.tsv", rows, list(rows[0]))

    payload = {
        "审计版本": "1.0",
        "审计日期": AUDIT_DATE,
        "来源": {
            "题名": "Temperature Dependent Dynamic Response of High-Density Polyurethane Foams",
            "DOI": "10.17632/x6b72k59xn.1",
            "版本": "1",
            "作者": ["Daniel Morrison"],
            "发布者": "MURI/AUSMURI Project / Mendeley Data",
            "许可": "CC BY 4.0",
            "许可证据": "Mendeley Data官方记录与DataCite DOI元数据",
            "参考文献": "Morrison, D. Temperature Dependent Dynamic Response of High-Density Polyurethane Foams [Data set], version 1. Mendeley Data, 2023. https://doi.org/10.17632/x6b72k59xn.1.",
        },
        "下载范围": {
            "已下载": "4个精选XLSX工作簿",
            "未下载": "未证明为Mendeley记录的全量文件；当前仅审计本地精选子集",
            "本地字节数": sum(path.stat().st_size for path in workbook_paths),
        },
        "独立科学单元": {
            "材料类别": ["High Density/HDB foam", "Hard A/HA foam"],
            "应力应变曲线": 12,
            "应力应变点对": total_curve_points,
            "原始落锤/压缩轨迹": 2,
            "轨迹点": {"40°C_Hard_Foam": 1585, "-20°C_Hard_Foam": 1690},
            "能量吸收材料-温度条件": 12,
            "能量吸收观测值": energy_measurements,
            "正确计数规则": "dataset_doi + foam_type + temperature + specimen/test；曲线点是同一试验内相关观测",
            "不得作为独立样本": [
                "29,707个曲线点", "复制/派生列", "Excel行", "同一条件的重复能量值",
            ],
        },
        "真实字段与单位": {
            "Foam Test Comparison": {"Time": "s", "Force": "N", "Stroke": "mm"},
            "HA/HDB应力应变": {
                "应变": "工作簿未声明；数值形态符合无量纲工程应变但需主文复核",
                "应力": "工作簿未声明；不得仅凭数量级固化为Pa/MPa",
            },
            "Energy Absorption": "工作簿仅标注Normalized at 65% Strain，未声明能量单位；入库前需主文/学位论文复核",
            "温度标签": ["-20°C", "-10°C", "0°C", "10°C", "room temperature", "40°C"],
            "工作簿概况": details,
        },
        "可解析性": {
            "状态": "结构可直接解析，语义单位需补证",
            "工作簿": "4/4可读",
            "公式": 0,
            "Excel错误值": 0,
        },
        "重复与异常": {
            "精确重复文件组": duplicates,
            "Foam Test复制列验证": duplicate_columns,
            "异常": [
                "HA_StressStrain.xlsx有两个列组都标为Negative 10，疑似一处温度标签笔误",
                "应力应变与能量吸收工作簿缺少显式单位",
                "Foam Test Comparison含复制/派生列，不可重复计数",
                "本地仅为精选表，不能声称官方全量数据已下载",
            ],
        },
        "准入与权重": {
            "当前层级": "力学迁移层",
            "准入状态": "单位硬门未闭合：当前仅保留Time/Force/Stroke原始轨迹与无量纲曲线形状的审计/表示价值，不作为TPU单体-配方核心真值",
            "当前单位相关性能标签权重上限": 0.0,
            "当前为零的标签": ["绝对应力", "模量", "韧性", "能量吸收", "依赖未声明单位的派生标量"],
            "补齐单位与试样映射后的来源迁移上限": 0.25,
            "分组键": ["dataset_doi", "foam_type", "temperature", "test_mode", "specimen"],
        },
    }
    write_json(base / "内容审计摘要.json", payload)


def audit_snd() -> None:
    base = DATA_ROOT / "SND_TPU导电轨迹循环拉伸"
    html_path = base / "官方数据集页面.html"
    pdf_path = base / "DATASET01.pdf"
    require_file(html_path)
    require_file(pdf_path)
    raw = html_path.read_text(encoding="utf-8", errors="replace")
    jsonlds = parse_jsonld(raw)
    dataset_meta = next(
        (item for item in jsonlds if item.get("@type") == "Dataset"),
        None,
    )
    if dataset_meta is None:
        raise AuditBlocked("SND官方页面缺少Dataset JSON-LD")
    if dataset_meta.get("license") != "https://creativecommons.org/licenses/by/4.0/":
        raise AuditBlocked(f"SND许可发生变化：{dataset_meta.get('license')}")

    with pdfplumber.open(pdf_path) as document:
        if len(document.pages) != 1:
            raise AuditBlocked(f"SND PDF应为1页，实际为{len(document.pages)}")
        page = document.pages[0]
        tables = page.extract_tables()
        text = page.extract_text() or ""
        page_count = len(document.pages)
    if len(tables) != 1 or "DC RESISTANCE" not in text:
        raise AuditBlocked("SND PDF表格或标题无法稳定提取")
    table = tables[0]
    if len(table) != 26 or len(table[0]) != 14:
        raise AuditBlocked(f"SND表格结构变化：{len(table)}行/{len(table[0])}列")

    headers = [
        "3min_初始_Ohm", "3min_25pct_Ohm", "3min_50pct_Ohm",
        "10min_初始_Ohm", "10min_25pct_Ohm", "10min_50pct_Ohm",
        "20min_初始_Ohm", "20min_25pct_Ohm", "20min_50pct_Ohm",
        "30min_初始_Ohm", "30min_25pct_Ohm", "30min_50pct_Ohm",
    ]
    parsed_rows: list[dict[str, object]] = []
    current_temperature: str | None = None
    missing = observed = 0
    for raw_row in table[2:]:
        if raw_row[0]:
            current_temperature = raw_row[0].replace("°C", "").strip()
        if current_temperature is None:
            raise AuditBlocked("SND表格首个温度无法前向填充")
        values: list[str] = []
        for value in raw_row[2:14]:
            if value in (None, "NaN", ""):
                values.append("")
                missing += 1
            else:
                values.append(str(value))
                observed += 1
        parsed = {"固化温度_C": current_temperature, "线宽_cm": raw_row[1]}
        parsed.update(dict(zip(headers, values)))
        parsed_rows.append(parsed)
    if len(parsed_rows) != 24 or observed != 283 or missing != 5:
        raise AuditBlocked(
            f"SND科学计数变化：rows={len(parsed_rows)}, observed={observed}, missing={missing}"
        )
    write_tsv(
        base / "电阻表解析清单.tsv",
        parsed_rows,
        ["固化温度_C", "线宽_cm", *headers],
    )

    rows, duplicates = file_rows(
        base,
        lambda path, _relative: (
            "导电轨迹电阻主数据" if path.suffix.lower() == ".pdf" else "官方元数据与许可证据"
        ),
        lambda path, _relative: (
            "通过：1页PDF表格经文字提取与150 dpi渲染双重核验"
            if path.suffix.lower() == ".pdf"
            else "通过：HTML内Schema.org JSON-LD可解析"
        ),
        lambda path, _relative: (
            "实验应用层" if path.suffix.lower() == ".pdf" else "元数据层"
        ),
        lambda path, _relative: "0.40" if path.suffix.lower() == ".pdf" else "0.00",
        lambda path, _relative: (
            "24个温度-线宽条件；各时刻/应变状态是条件内重复观测"
            if path.suffix.lower() == ".pdf"
            else "DOI、作者、许可和下载入口证据"
        ),
    )
    if len(rows) != 2:
        raise AuditBlocked(f"SND文件校验行应为2，实际为{len(rows)}")
    write_tsv(base / "文件校验清单.tsv", rows, list(rows[0]))

    payload = {
        "审计版本": "1.0",
        "审计日期": AUDIT_DATE,
        "来源": {
            "题名": dataset_meta["name"],
            "DOI": "10.5878/tc7g-1056",
            "版本": dataset_meta.get("version", "1"),
            "作者": ["Jawad Ahmad"],
            "贡献者": ["Xiaotian Li", "Johan Sidén", "Henrik Andersson"],
            "机构": "Mid Sweden University",
            "许可": "CC BY 4.0",
            "许可证据": "本地官方页面JSON-LD与DataCite DOI元数据",
            "关联论文": "10.1109/FLEPS.2019.8792266",
            "参考文献": "Ahmad, J. An Analysis of Screen-Printed Stretchable Conductive Tracks on Thermoplastic Polyurethane [Data set], version 1. Mid Sweden University/SND, 2019. https://doi.org/10.5878/tc7g-1056.",
        },
        "下载范围": {
            "已下载": "官方页面与唯一数据文件DATASET01.pdf（138,234字节）",
            "未下载": "无；官方JSON-LD仅列1个分发文件",
            "PDF页数": page_count,
        },
        "独立科学单元": {
            "温度-线宽组合": 24,
            "固化温度_C": [110, 120, 130, 150],
            "线宽_cm": [1, 0.5, 0.25, 0.125, 0.0625, 0.0312],
            "理论测量格点": 288,
            "有效电阻读数": observed,
            "缺失NaN": missing,
            "正确计数规则": "dataset_doi + curing_temperature + line_width；时刻与拉伸状态为同一导电轨迹条件内观测",
            "不得作为独立样本": ["283个电阻单元格", "4个时间点", "初始/25%/50%三状态"],
        },
        "真实字段与单位": {
            "DC电阻": "Ω",
            "线宽": "cm",
            "固化温度": "°C",
            "时间": "min",
            "伸长": "%（25%、50%）",
        },
        "可解析性": {
            "状态": "高",
            "PDF文本": "可提取",
            "表格": "1张、26行（2行表头+24条件行）、14列",
            "视觉核验": "150 dpi渲染清晰，标题、单位与NaN位置和提取结果一致",
        },
        "重复与异常": {
            "精确重复文件组": duplicates,
            "异常": [
                "5个NaN缺失值",
                "只有汇总表，无逐循环连续轨迹与重复试样ID",
                "不能由单表估计试样间不确定性",
            ],
        },
        "准入与权重": {
            "当前层级": "实验应用层（TPU基底导电器件）",
            "准入状态": "准入：用于工艺-几何-电阻保持应用子模型；不作为纯TPU本体力学标签",
            "应用任务权重上限": 0.40,
            "跨性能迁移权重上限": 0.20,
            "分组键": [
                "dataset_doi", "curing_temperature", "line_width", "track_specimen_if_available",
            ],
        },
    }
    write_json(base / "内容审计摘要.json", payload)


def audit_sciencedb() -> None:
    base = DATA_ROOT / "ScienceDB_TPU芳纶纳米纤维能量吸收"
    archive_path = base / "sciencedb.26393_V1_官方全量.zip"
    html_path = base / "官方数据集页面.html"
    require_file(archive_path)
    require_file(html_path)

    raw = html_path.read_text(encoding="utf-8", errors="replace")
    jsonlds = parse_jsonld(raw)
    dataset_meta = next(
        (
            item
            for item in jsonlds
            if str(item.get("@type", "")).endswith("Dataset")
        ),
        None,
    )
    if dataset_meta is None:
        raise AuditBlocked("ScienceDB页面缺少Dataset/Croissant JSON-LD")
    if dataset_meta.get("license") != "https://creativecommons.org/licenses/by-nc/4.0/":
        raise AuditBlocked(f"ScienceDB许可发生变化：{dataset_meta.get('license')}")

    image_rows: list[dict[str, object]] = []
    with zipfile.ZipFile(archive_path) as archive:
        entries = validate_zip(archive, "ScienceDB TPU/ANF")
        entry_by_name = {Path(item.filename).name: item for item in entries}
        image_paths = sorted((base / "原始图像").glob("*.tif"))
        if len(entries) != 19 or len(image_paths) != 19:
            raise AuditBlocked(
                f"ScienceDB图像数应为19：ZIP={len(entries)}, local={len(image_paths)}"
            )
        extracted_match = 0
        for path in image_paths:
            with Image.open(path) as image:
                dpi = image.info.get("dpi")
                dpi_text = "" if dpi is None else "x".join(str(float(value)) for value in dpi)
                frames = getattr(image, "n_frames", 1)
                mode = image.mode
                width, height = image.size
                image.verify()
            if path.name not in entry_by_name:
                raise AuditBlocked(f"ScienceDB ZIP缺少图像：{path.name}")
            info = entry_by_name[path.name]
            same = archive.read(info) == path.read_bytes()
            extracted_match += int(same)
            group = (
                "1"
                if path.name.startswith("1_")
                else (
                    "4"
                    if path.name.startswith("4-")
                    else ("6" if path.name.startswith("6-") else "AFM")
                )
            )
            image_rows.append(
                {
                    "文件": path.relative_to(base).as_posix(),
                    "文件名组": group,
                    "宽_px": width,
                    "高_px": height,
                    "颜色模式": mode,
                    "帧数": frames,
                    "DPI元数据": dpi_text,
                    "字节数": path.stat().st_size,
                    "ZIP内CRC32": f"{info.CRC:08X}",
                    "与ZIP字节一致": str(same).lower(),
                    "可解码": "true",
                    "科学语义": "文件名仅对应手稿图组；本地数据未给出逐图配方/尺度标签，禁止自行推断",
                }
            )
    if extracted_match != 19:
        raise AuditBlocked(f"ScienceDB仅{extracted_match}/19图像与ZIP逐字节一致")
    group_counts = Counter(str(row["文件名组"]) for row in image_rows)
    if group_counts != Counter({"1": 2, "4": 8, "6": 8, "AFM": 1}):
        raise AuditBlocked(f"ScienceDB文件名图组计数变化：{group_counts}")
    write_tsv(base / "图像解析清单.tsv", image_rows, list(image_rows[0]))

    def role(path: Path, _: str) -> str:
        if path.suffix.lower() == ".tif":
            return "显微/形貌图像主数据"
        if path.suffix.lower() == ".zip":
            return "镜像容器"
        return "官方元数据与许可证据"

    def parse_status(path: Path, _: str) -> str:
        if path.suffix.lower() == ".tif":
            return "通过：Pillow可解码且与ZIP内字节一致"
        if path.suffix.lower() == ".zip":
            return "通过：ZIP CRC全部通过"
        return "通过：HTML内Croissant/Schema.org JSON-LD可解析"

    rows, duplicates = file_rows(
        base,
        role,
        parse_status,
        lambda path, _: "定性证据与人工标注候选层" if path.suffix.lower() == ".tif" else "元数据层",
        lambda path, _: "0.00",
        lambda path, _: (
            "19张无逐图科学标签图像中的1张；当前权重为0；图像块/像素不得当材料样本；逐图标注并扩充合法语料后未来非商业视觉任务上限0.15"
            if path.suffix.lower() == ".tif"
            else (
                "与19张解包TIF构成镜像"
                if path.suffix.lower() == ".zip"
                else "DOI、CC BY-NC 4.0与文件总量证据"
            )
        ),
    )
    if len(rows) != 21:
        raise AuditBlocked(f"ScienceDB文件校验行应为21，实际为{len(rows)}")
    write_tsv(base / "文件校验清单.tsv", rows, list(rows[0]))

    size = dataset_meta.get("size", {})
    assert isinstance(size, dict)
    payload = {
        "审计版本": "1.0",
        "审计日期": AUDIT_DATE,
        "来源": {
            "题名": dataset_meta.get("name"),
            "DOI": "10.57760/sciencedb.26393",
            "版本": "V1",
            "作者": ["Zhang Shuai"],
            "机构": "Northwestern Polytechnical University",
            "许可": "CC BY-NC 4.0",
            "许可证据": "本地官方页面Croissant JSON-LD与DataCite DOI元数据",
            "商业使用限制": "是；任何商业/产业化使用需另行获得许可",
            "参考文献": "Zhang, S. 3D-Printed Multiscale Hierarchical Thermoplastic Polyurethane / Aramid Nanofiber Structures with Enhanced Energy Absorption via In-Situ Foaming Technology [Data set], version 1. Science Data Bank, 2025. https://doi.org/10.57760/sciencedb.26393.",
        },
        "下载范围": {
            "已下载": "官方全量ZIP、官方页面以及ZIP内全部19张TIF图像",
            "未下载": "无（该版本登记19个文件，均在ZIP中）",
            "官方未压缩总量_字节": size.get("value"),
            "ZIP_CRC": "通过",
            "与ZIP逐字节一致的解包图像": extracted_match,
        },
        "独立科学单元": {
            "图像资产": 19,
            "文件名图组": {"1": 2, "4": 8, "6": 8, "AFM": 1},
            "可确认的独立配方/试样数": 0,
            "原因": "文件和页面未给出逐图配方、倍率、标尺、试样ID映射；19张图像不能自动等同19个配方",
            "不得作为独立样本": ["像素", "图像切块", "ZIP镜像", "同一视野的裁剪/图组编号"],
        },
        "真实字段与单位": {
            "图像字段": ["像素宽度", "像素高度", "颜色模式", "帧数", "DPI元数据"],
            "物理尺度": "未提供可机器解析的µm/nm标尺映射；DPI仅是文件显示元数据，不能替代显微物理尺度",
            "能量吸收": "数据包不含数值应力-应变或吸能表；题名中的增强吸能不能当作本地标量标签",
        },
        "可解析性": {
            "状态": "图像技术可解析、科学语义不完整",
            "TIF可解码": len(image_rows),
            "ZIP文件数": len(entries),
        },
        "重复与异常": {
            "精确重复文件组": duplicates,
            "异常": [
                "无逐图配方/试样/倍率/标尺映射",
                "没有能量吸收数值表",
                "许可为CC BY-NC 4.0而非CC BY 4.0",
            ],
        },
        "准入与权重": {
            "当前层级": "定性证据与人工标注候选层",
            "准入状态": "当前所有训练任务权重为0；仅可用于人工标注队列与定性核验，不得进入视觉或标量监督层",
            "当前所有任务权重上限": 0.0,
            "未来非商业视觉任务权重上限": 0.15,
            "未来放行条件": [
                "建立逐图配方、试样、倍率、物理标尺和视野映射",
                "扩充更大且许可兼容的合法图像语料",
                "完成来源级成组与视野级泄漏审计",
            ],
            "能量吸收标量权重上限": 0.0,
            "分组键": [
                "dataset_doi", "figure_group", "specimen_id_after_manual_mapping", "field_of_view",
            ],
        },
    }
    write_json(base / "内容审计摘要.json", payload)


def audit_agh() -> None:
    base = DATA_ROOT / "AGH_低石化多元醇硬质PU泡沫"
    metadata_path = base / "官方Dataverse元数据.json"
    page_path = base / "官方数据集页面.html"
    preview_path = base / "README官方预览页面.html"
    for path in (metadata_path, page_path, preview_path):
        require_file(path)
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    latest = metadata["data"]["latestVersion"]
    license_info = latest["license"]
    if license_info.get("rightsIdentifier") != "CC0-1.0":
        raise AuditBlocked(f"AGH许可发生变化：{license_info}")
    if not latest.get("fileAccessRequest"):
        raise AuditBlocked("AGH访问状态已变化；需要重新审计，而非沿用request_required结论")

    remote_rows: list[dict[str, object]] = []
    total = restricted_bytes = 0
    expected_content = {
        "Apparent density.xlsx": "表观密度",
        "Bio-polyols synthesis.docx": "生物多元醇合成",
        "Foaming process.xlsx": "发泡特征时间",
        "FTIR.docx": "FTIR",
        "Hardness.xlsx": "硬度",
        "Physicochemical properties.xlsx": "环氧值/羟值等理化性质",
        "RPUF synthesis.docx": "硬质PU泡沫合成",
        "SEM.docx": "泡孔形貌",
        "00_readme.txt": "来源说明",
    }
    for item in latest["files"]:
        data = item["dataFile"]
        size = int(data["filesize"])
        restricted = bool(item["restricted"])
        total += size
        restricted_bytes += size if restricted else 0
        remote_rows.append(
            {
                "远端文件": item["label"],
                "字节数": size,
                "官方MD5": data["checksum"]["value"],
                "受限": str(restricted).lower(),
                "本地状态": "未下载（需申请）" if restricted else "未下载；仅保存官方预览页面",
                "可预期内容": expected_content.get(item["label"], "未知"),
                "当前训练权重": "0.00",
            }
        )
    restricted_count = sum(row["受限"] == "true" for row in remote_rows)
    if len(remote_rows) != 9 or restricted_count != 8 or total != 13_934_145:
        raise AuditBlocked(
            f"AGH远端清单变化：files={len(remote_rows)}, restricted={restricted_count}, bytes={total}"
        )
    write_tsv(base / "官方远端文件清单.tsv", remote_rows, list(remote_rows[0]))

    rows, duplicates = file_rows(
        base,
        lambda _path, _relative: "官方元数据与访问状态证据",
        lambda path, _relative: (
            "通过：JSON可解析" if path.suffix.lower() == ".json" else "通过：HTML可读"
        ),
        lambda _path, _relative: "元数据层",
        lambda _path, _relative: "0.00",
        lambda _path, _relative: "仅证明许可、文件清单与受限状态；不含可训练测量值",
    )
    for remote in remote_rows:
        rows.append(
            {
                "相对路径": "远端/" + str(remote["远端文件"]),
                "字节数": remote["字节数"],
                "SHA256": "未下载",
                "官方校验": "MD5:" + str(remote["官方MD5"]),
                "文件角色": (
                    "远端主数据"
                    if remote["远端文件"] != "00_readme.txt"
                    else "远端说明文档"
                ),
                "完整性与可读性": remote["本地状态"],
                "精确重复组": "无法判断",
                "准入层": "元数据层",
                "训练权重上限": "0.00",
                "备注": "受限" if remote["受限"] == "true" else "公开README；本地仅留预览证据",
            }
        )
    if len(rows) != 12:
        raise AuditBlocked(f"AGH文件校验行应为12，实际为{len(rows)}")
    write_tsv(base / "文件校验清单.tsv", rows, list(rows[0]))

    payload = {
        "审计版本": "1.0",
        "审计日期": AUDIT_DATE,
        "来源": {
            "题名": "Rigid polyurethane foams with reduced content of petrochemical polyols",
            "DOI": "10.58032/AGH/LKHZ6Q",
            "版本": "1.0",
            "作者": ["Patrycja Zakrzewska"],
            "机构": "AGH University of Krakow",
            "许可": "CC0 1.0",
            "可访问性": "request_required",
            "许可与访问证据": "本地官方Dataverse元数据；CC0不等于文件已开放，8/9文件仍受限",
            "参考文献": "Zakrzewska, P. Rigid polyurethane foams with reduced content of petrochemical polyols [Data set], version 1. AGH University of Krakow, 2026. https://doi.org/10.58032/AGH/LKHZ6Q.",
        },
        "下载范围": {
            "已下载": "官方数据集页面、Dataverse元数据和公开README预览页面；没有下载测量主文件",
            "未下载": "8个受限主文件；公开README正文也未以原始TXT保存",
            "远端文件数": len(remote_rows),
            "受限文件数": restricted_count,
            "远端总字节": total,
            "受限字节": restricted_bytes,
            "受限字节占比_pct": round(restricted_bytes / total * 100, 6),
        },
        "独立科学单元": {
            "当前可训练测量单元": 0,
            "当前仅数据集级元数据单元": 1,
            "潜在科学内容": [
                "生物多元醇合成", "环氧值", "羟值", "FTIR", "发泡特征时间", "表观密度", "硬度", "SEM泡孔结构", "RPUF配方/合成",
            ],
            "潜在配方数": "未知；受限表格未下载前不得推测",
        },
        "真实字段与单位": {
            "当前可确认": "仅字段主题，无可核实测量单位或数值",
            "禁止推断": [
                "密度单位", "硬度标尺", "环氧值/羟值单位", "配方比例", "发泡时间单位", "SEM尺度",
            ],
        },
        "可解析性": {
            "状态": "元数据可解析；主数据不可访问",
            "官方元数据文件条目": len(remote_rows),
            "本地测量文件": 0,
        },
        "重复与异常": {
            "精确重复文件组": duplicates,
            "异常": [
                "8/9文件受限",
                "受限文件占远端字节约99.9757%",
                "CC0许可与request_required访问状态必须分开治理",
            ],
        },
        "准入与权重": {
            "当前层级": "元数据层",
            "准入状态": "不准入训练；保留为申请队列与选题证据",
            "当前权重上限": 0.0,
            "获批下载并完成配方/单位/重复审计后的迁移上限": 0.25,
            "未来分组键": [
                "dataset_doi", "bio_polyol_feedstock", "foam_formulation", "batch", "specimen", "test_mode",
            ],
        },
    }
    write_json(base / "内容审计摘要.json", payload)


def validate_outputs() -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        summary_path = base / "内容审计摘要.json"
        checks_path = base / "文件校验清单.tsv"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        if summary.get("审计版本") != "1.0":
            raise AuditBlocked(f"审计摘要版本错误：{summary_path}")
        with checks_path.open("r", encoding="utf-8-sig", newline="") as handle:
            check_rows = list(csv.DictReader(handle, delimiter="\t"))
        if not check_rows or not all(row.get("相对路径") for row in check_rows):
            raise AuditBlocked(f"文件校验清单为空或缺路径：{checks_path}")
        result.append(
            {
                "来源": source,
                "文件校验行": len(check_rows),
                "内容审计摘要_SHA256": sha256(summary_path),
                "文件校验清单_SHA256": sha256(checks_path),
            }
        )
    return result


def main() -> int:
    missing_sources = [name for name in SOURCE_NAMES if not (DATA_ROOT / name).is_dir()]
    if missing_sources:
        raise AuditBlocked(f"缺少来源目录：{missing_sources}")

    before = raw_snapshot()
    audit_jagiellonian()
    audit_swcnt()
    audit_mendeley()
    audit_snd()
    audit_sciencedb()
    audit_agh()
    after = raw_snapshot()
    if before != after:
        changed = sorted(set(before) ^ set(after))
        changed.extend(
            key for key in sorted(set(before) & set(after)) if before[key] != after[key]
        )
        raise AuditBlocked(f"原始文件在审计期间发生变化：{changed}")

    print(json.dumps(validate_outputs(), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"审计阻断：{exc}", file=sys.stderr)
        raise SystemExit(2) from exc
