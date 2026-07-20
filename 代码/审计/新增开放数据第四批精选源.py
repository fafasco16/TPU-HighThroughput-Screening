"""只读深审第四批一个实验数据集与八个 ACS 支持信息来源。

脚本不联网、不提取 ZIP 到磁盘、不修改科学原件，也不创建训练集。Mendeley
来源按物理试样逐行复算单位换算并隔离外链公式；ACS PDF 只登记为文献证据层，
验证文件、页数、安全目录键、文本锚点及已识别表/图组，不把 PDF 文字或图线冒充
规范训练记录。每个来源只原子写入三个固定审计输出。

运行：

    python 代码/审计/新增开放数据第四批精选源.py
"""

from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import BinaryIO
from urllib.parse import urlsplit

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "01_原始数据" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"
AUDIT_VERSION = "1.2"

MENDELEY = "Mendeley_TPU压缩打印DOE"
PDF_SOURCES = (
    "ACS_Figshare_TPU退火硬段聚集",
    "ACS_Figshare_双相演化聚氨酯",
    "ACS_Figshare_PLA立构复合TPU",
    "ACS_Figshare_呋喃高强聚氨酯",
    "ACS_Figshare_聚酰亚胺回收链扩剂PU",
    "ACS_Figshare_二氧化碳共聚酯聚氨酯",
    "ACS_Figshare_聚碳酸酯大分子二醇TPU",
    "ACS_Figshare_氢键纳米结构TPU",
)
SOURCE_NAMES = (MENDELEY, *PDF_SOURCES)
OUTPUT_NAMES = ("内容审计摘要.json", "文件校验清单.tsv", "曲线审计清单.tsv")
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source in SOURCE_NAMES
    for filename in OUTPUT_NAMES
)

MENDELEY_ARCHIVE = (
    "7zcd9bmmg5-1.zip",
    1_717_731,
    "0b26707846f5cd23d2f843eb30d90ad24e548fce277a2cbffa5555348d226397",
)
MENDELEY_PREFIX = (
    "Impact of Infill and Shell Design Features on Compression Stiffness "
    "in Material Extrusion of Thermoplastic Urethane/"
)
NINJA_WORKBOOK = MENDELEY_PREFIX + "NinjaTek Data.xlsx"
POLY_WORKBOOK = MENDELEY_PREFIX + "PolyFlex Data.xlsx"
POLY_MPX = MENDELEY_PREFIX + "PolyFlex DOE_3x2 with k.mpx"
NINJA_MPX = MENDELEY_PREFIX + "2level Ninjatek DOE analysis_RVD.mpx"
MENDELEY_MEMBERS = {
    NINJA_WORKBOOK: (
        155_449,
        "d2916bcd7ff431c01aa9f738fc245570e1743cf176373206adc3a33e5bdfad74",
    ),
    POLY_WORKBOOK: (
        61_705,
        "668fc97bdde2e660aa7fa61cb334ab451046625b6ecccb8275814e414c1a2527",
    ),
    POLY_MPX: (
        640_525,
        "e1985c9f03af545fb2594333b5722ef43f4ed847708fbcf660270782d92daa85",
    ),
    NINJA_MPX: (
        927_107,
        "2c8da608c45ad723fc280d7a1afab43ba3d76beafbe21a43ab1a7cf61a659c26",
    ),
}

NINJA_SHAPES = {
    "Raw Data": (73, 27),
    "Peak stresses": (14, 9),
    "NinjaFlex Cubes - Do Not Sort": (73, 16),
    "NinjaFlex Cylinders - Do Not S.": (37, 9),
    "Averages and Graphs": (40, 21),
    "Cylinder vs Cube Comparison": (28, 13),
    "Cylinder Graphs including lines": (19, 11),
    "Solid Cylinder Analysis v1": (33, 18),
    "std devs": (73, 16),
    "cylinder std devs": (17, 15),
    "Infill Percentage investigation": (49, 9),
}
POLY_SHAPES = {
    "Raw data - Do not sort": (90, 30),
    "Data for DOE": (25, 16),
    "Averages": (13, 8),
    "Comparison of polyflex_ninjafle": (25, 15),
}


@dataclass(frozen=True)
class EvidenceGroup:
    group_id: str
    evidence_object: str
    reported_items: str
    variables: str
    manual_action: str
    anomaly: str = ""


@dataclass(frozen=True)
class PdfSpec:
    directory: str
    article_id: int
    supplement_doi: str
    resource_doi: str
    filename: str
    size: int
    sha256: str
    pages: int
    min_text_chars: int
    anchors: tuple[str, ...]
    groups: tuple[EvidenceGroup, ...]


PDF_SPECS = (
    PdfSpec(
        "ACS_Figshare_TPU退火硬段聚集",
        28_906_446,
        "10.1021/acs.macromol.5c00142.s001",
        "10.1021/acs.macromol.5c00142",
        "ma5c00142_si_001.pdf",
        1_527_655,
        "06baadbcd7cdf81a2e6d66dd9cc06b087dc2eb86dbe18ec65827402d0dfd7983",
        10,
        7_000,
        (
            "Table S1. Total melting enthalpy",
            "Table S2. Morphological parameters obtained from one-dimensional correlation function analysis",
            "Table S3. Statistical results about the width of hard blocks",
        ),
        (
            EvidenceGroup(
                "table_s1_annealing_thermal",
                "Table S1",
                "21 temperature-time conditions",
                "annealing_temperature;annealing_time;Tm;melting_enthalpy",
                "人工双录并逐格核对 PDF 表格；绑定退火协议后再规范化",
            ),
            EvidenceGroup(
                "table_s2_saxs_morphology",
                "Table S2",
                "temperature-time condition matrix",
                "long_period;hard_domain_length;soft_domain_length;interface_width;hard_segment_fraction",
                "人工双录并检查表头跨页；不得从文本顺序直接推断列映射",
            ),
            EvidenceGroup(
                "table_s3_afm_width",
                "Table S3",
                "4 material/process states",
                "hard_block_width_min;median;max;mean;standard_deviation",
                "人工核对单位与图像统计口径",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_双相演化聚氨酯",
        29_074_233,
        "10.1021/acsmaterialslett.5c00732.s001",
        "10.1021/acsmaterialslett.5c00732",
        "tz5c00732_si_001.pdf",
        1_176_571,
        "5bdedae10fcaff85da215a98a5dadfe7b0608ea6d14ba7dcc1adcbbc468938c9",
        17,
        13_000,
        (
            "Table S1. The molar ratio of each chemical.",
            "Table S2. Molecular weight and its dispersion index",
            "D4C0 4 0 4",
        ),
        (
            EvidenceGroup(
                "table_s1_formulation",
                "Table S1",
                "5 formulations",
                "D400_molar_ratio;PCL_molar_ratio;HMDI_molar_ratio",
                "人工双录并与实验章节原料分子量交叉核对",
            ),
            EvidenceGroup(
                "table_s2_molecular_weight",
                "Table S2",
                "5 elastomers",
                "Mn_kDa;Mw_kDa;dispersity",
                "人工双录；核对异常宽分布是否为原文值",
            ),
            EvidenceGroup(
                "table_s3_hydrogen_bond",
                "Table S3",
                "5 elastomers",
                "carbonyl_assignment;wavenumber;peak_area_fraction",
                "人工解析多级表头并绑定样品",
            ),
            EvidenceGroup(
                "mechanical_figures",
                "stress-strain figures",
                "figure curves",
                "strain;stress",
                "只可从矢量图人工/半自动数字化，需两人复核并保存图页坐标",
                "支持信息未提供原始曲线文件",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_PLA立构复合TPU",
        31_333_274,
        "10.1021/acs.macromol.5c03502.s001",
        "10.1021/acs.macromol.5c03502",
        "ma5c03502_si_001.pdf",
        5_618_558,
        "c4d5ec8522eaccd52a2a208809efa9a6f4fccb3555aaeb733f0119398cfc9ec6",
        42,
        35_000,
        (
            "Table S1. Degrees of polymerization of PCL",
            "Table S2. Degrees of polymerization of PLLA",
            "Table S3. Summary of characteristics of the triblock copolymers.",
            "Table S4. Thermal characteristics and crystallinity of the triblock-based TPU.",
        ),
        (
            EvidenceGroup(
                "table_s1_pcl",
                "Table S1",
                "3 PCL homopolymers",
                "NMR_integrals;PCL_degree_of_polymerization",
                "人工双录并核对引发剂脚注",
            ),
            EvidenceGroup(
                "table_s2_triblock",
                "Table S2",
                "10 triblocks",
                "NMR_integrals;block_ratio;PLA_or_PDLA_degree;Mn_NMR",
                "人工双录并保留 L/D 立构身份",
            ),
            EvidenceGroup(
                "table_s3_molecular_weight",
                "Table S3",
                "10 triblocks",
                "Mn_NMR;Mn_GPC;Mw_GPC;dispersity",
                "人工双录；N/A 不得填零",
            ),
            EvidenceGroup(
                "table_s4_thermal",
                "Table S4",
                "triblock-based TPU series",
                "Tg_soft_segment;Tm_PCL;Tm_scPLA;enthalpy;crystallinity",
                "人工处理跨页表格并与样品命名表联结",
            ),
            EvidenceGroup(
                "mechanical_figures",
                "tensile and cyclic figures",
                "multiple TPU curves",
                "Young_modulus;ultimate_stress;elongation;toughness;cyclic_response",
                "SI 只给图线；数字化前不得创建数值标签",
                "机械指标无逐样品数值表",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_呋喃高强聚氨酯",
        31_429_142,
        "10.1021/acs.macromol.5c03627.s001",
        "10.1021/acs.macromol.5c03627",
        "ma5c03627_si_001.pdf",
        2_261_687,
        "1b85a8294ce375e9b7f7cf314df369eaf7edfa9713a2ff6031aae74330df9108",
        35,
        30_000,
        (
            "Table S1. Formulations for FPUs",
            "Table S5. Energy dissipation value",
            "Table S6. Elastic recovery value",
            "FPU-3 2.22 0.91 23.8%",
        ),
        (
            EvidenceGroup(
                "table_s1_s2_formulations",
                "Tables S1-S2",
                "6 formulations",
                "PTMG_2000_mass;IPDI_mass;chain_extender_mass;hard_segment_content",
                "人工双录；PDF 合并单元格导致 PTMG=10 g 需按原表视觉复核",
                "文本抽取会移动合并单元格值",
            ),
            EvidenceGroup(
                "table_s5_dissipation",
                "Table S5",
                "6 materials across stepped strain",
                "maximum_strain;energy_dissipation_mean;standard_deviation",
                "人工双录并记录缺失的高应变单元格",
            ),
            EvidenceGroup(
                "table_s6_recovery",
                "Table S6",
                "6 materials across stepped strain",
                "maximum_strain;elastic_recovery_mean;standard_deviation",
                "人工双录并核对列数",
            ),
            EvidenceGroup(
                "table_s7_residual_strain",
                "Table S7",
                "5 materials at 3 strain levels",
                "maximum_strain;first_loop_residual;last_loop_residual",
                "人工双录",
            ),
            EvidenceGroup(
                "tensile_figures",
                "tensile figures",
                "multiple triplicate curves",
                "strain;stress;tensile_strength;elongation;toughness",
                "仅图线，数字化需保留重复曲线和坐标校准证据",
                "摘要性能值不能替代逐样品曲线",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_聚酰亚胺回收链扩剂PU",
        31_614_502,
        "10.1021/acsapm.5c04872.s001",
        "10.1021/acsapm.5c04872",
        "ap5c04872_si_001.pdf",
        922_002,
        "c18bb54c66f7182cff03508067f7def63ee417fbdbdbfe29067ba568849bedea",
        11,
        2_500,
        (
            "Table S1. The reaction temperature and reaction time for different samples",
            "Table S2. The components of different samples",
            "Table S3. Summary of the mechanical properties of PU-4HPA-X samples",
            "PU-4HPA-4 78.97 1267",
        ),
        (
            EvidenceGroup(
                "table_s1_chain_extender_process",
                "Table S1",
                "5 polyimide upcycling conditions",
                "reaction_temperature;reaction_time",
                "人工双录并与主文产率定义交叉核对",
            ),
            EvidenceGroup(
                "table_s2_formulation",
                "Table S2",
                "5 PU-4HPA formulations",
                "PTMEG_mmol;IPDI_mmol;DBTDL_mass;4HPA_mmol;4HPA_fraction",
                "人工双录；明确 4HPA 百分比的分母定义",
            ),
            EvidenceGroup(
                "table_s3_mechanics",
                "Table S3",
                "5 PU-4HPA samples",
                "tensile_strength_MPa;elongation_at_break_percent",
                "人工双录并保留样品序列；主文测试协议需绑定后才可规范化",
            ),
            EvidenceGroup(
                "figure_s19_tensile",
                "Figure S19",
                "PU-BDO and PU-4HPA-4 curves",
                "strain;stress",
                "只可从图线数字化并校准坐标；不得用 Table S3 端点重建曲线",
                "无原始应力-应变曲线文件",
            ),
            EvidenceGroup(
                "table_s4_literature_comparison",
                "Table S4",
                "secondary literature comparison",
                "reported_tensile_strength;reported_elongation;citation",
                "只保留为检索线索，回到各原始论文后才能入库",
                "二手汇总，不属于本来源的独立实验样本",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_二氧化碳共聚酯聚氨酯",
        31_989_433,
        "10.1021/acsmacrolett.6c00123.s001",
        "10.1021/acsmacrolett.6c00123",
        "mz6c00123_si_001.pdf",
        1_392_127,
        "a8770b0aee18e63efe119807c745378d060ae81fbcd131f9734dee55f2e4406e",
        20,
        18_000,
        (
            "Table S1. ROP of EVP and CL catalyzed by the TBD/Urea binary system.",
            "Stress–strain curves of",
            "Chain extension of poly(EVP-co-CL) diols to PU11",
        ),
        (
            EvidenceGroup(
                "table_s1_copolyester",
                "Table S1",
                "4 polymerization runs",
                "feed_ratio;conversion;selectivity;incorporation;Mn_theory;Mn_NMR;Mn_SEC;dispersity",
                "人工双录并保留脚注中的测量方法",
            ),
            EvidenceGroup(
                "figure_s6_tensile",
                "Figure S6",
                "4 materials x 3 replicates",
                "strain;stress",
                "图线数字化后按材料分组拆分；不得将图例均值当单试样",
                "无原始曲线文件",
            ),
            EvidenceGroup(
                "synthesis_protocol",
                "experimental section",
                "PU0/PU6/PU11/PU19 series",
                "polyol_identity;IPDI;ADH;catalyst;solvent;temperature;time",
                "结构化实验步骤需人工复核化学计量和单位",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_聚碳酸酯大分子二醇TPU",
        32_256_977,
        "10.1021/acsapm.6c00646.s001",
        "10.1021/acsapm.6c00646",
        "ap6c00646_si_001.pdf",
        970_780,
        "2bd71204aa0807379e5f27a1e30de02ca0e1d4b981b3251c1e1567c1ad0109ec",
        11,
        7_000,
        (
            "Table S1. Immortal copolymerization of PO/EO/CO2",
            "Table S2. Repetitive synthesis of PEPCDLs",
            "Table S4. Properties of terminal hydroxyl",
            "DMA curves of TPUs based on PPCDL",
        ),
        (
            EvidenceGroup(
                "table_s1_macrodiol_screen",
                "Table S1",
                "feed-composition screen",
                "PO_feed;EO_feed;polymer_composition;Mn;PDI;conversion",
                "人工双录并校验 NMR/GPC 脚注",
            ),
            EvidenceGroup(
                "table_s2_reproducibility",
                "Table S2",
                "replicate synthesis series",
                "feed_ratio;CO2_pressure;time;temperature;composition;Mn;conversion",
                "人工按重复实验编号录入，不得先做平均",
            ),
            EvidenceGroup(
                "table_s4_hydroxyl",
                "Table S4",
                "5 macrodiols",
                "primary_OH_fraction;secondary_OH_fraction;carbonate_or_ether_context",
                "人工解析多级列头",
            ),
            EvidenceGroup(
                "tpu_property_figures",
                "Figures S12-S13",
                "TPU thermal/mechanical curves",
                "DMA_temperature_sweep;tensile_recovery",
                "只作证据索引，数值化前需坐标校准与重复性标注",
                "无原始曲线文件",
            ),
        ),
    ),
    PdfSpec(
        "ACS_Figshare_氢键纳米结构TPU",
        32_567_339,
        "10.1021/acs.macromol.6c00352.s001",
        "10.1021/acs.macromol.6c00352",
        "ma6c00352_si_001.pdf",
        2_077_355,
        "29d8b451025a86acb3c075a6cb7c29428b725e7c40083a70d271500444b93765",
        28,
        22_000,
        (
            "Table S1. Feeding compositions for synthesis of HTPUs.",
            "Table S2. Molecular weights of HTPUs.",
            "Table S3. Glass transition temperatures",
            "HTPU-P1 -32.00 / 47.00 79.21 1475 365.72",
        ),
        (
            EvidenceGroup(
                "table_s1_formulation",
                "Table S1",
                "7 HTPU formulations",
                "diol_identity;diol_mass;IPDI_mass;chain_extender_identity;chain_extender_mass;molar_ratio",
                "人工双录并建立原料身份字典",
            ),
            EvidenceGroup(
                "table_s2_molecular_weight",
                "Table S2",
                "7 HTPU samples",
                "Mn_kDa;Mw_kDa;dispersity",
                "人工双录",
            ),
            EvidenceGroup(
                "table_s3_mechanics",
                "Table S3",
                "7 HTPU samples",
                "Tg_soft;Tg_hard;tensile_strength;elongation_at_break;toughness",
                "人工双录；破折号必须保留为缺失而非零",
            ),
            EvidenceGroup(
                "table_s4_hydrogen_bond",
                "Table S4",
                "3 selected HTPU samples",
                "carbonyl_assignment;wavenumber;area_fraction;total_hydrogen_bond_degree",
                "人工解析多级表头",
            ),
        ),
    ),
)


class AuditBlocked(RuntimeError):
    """输入、容器或科学语义不满足固定审计协议。"""


def same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(flag and attributes & flag) or bool(is_junction())


def assert_plain_chain(path: Path, stop: Path) -> None:
    if path != stop and stop not in path.parents:
        raise AuditBlocked(f"路径越出审计根：{path}")
    cursor = path
    while True:
        if is_reparse_point(cursor):
            raise AuditBlocked(f"拒绝符号链接或重解析点：{cursor}")
        if cursor == stop:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not same_path(resolved, path.absolute()):
        raise AuditBlocked(f"目录缺失或经链接解析：{path}")
    assert_plain_chain(path, PROJECT_ROOT)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not same_path(resolved, path.absolute()):
        raise AuditBlocked(f"文件缺失或经链接解析：{path}")
    assert_plain_chain(path, PROJECT_ROOT)


def assert_output(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    require_directory(path.parent)
    if path.exists() and (not path.is_file() or is_reparse_point(path)):
        raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    assert_output(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if not temporary.is_file() or is_reparse_point(temporary):
            raise AuditBlocked(f"审计临时文件异常：{temporary}")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_json(path: Path, payload: dict[str, object]) -> None:
    rendered = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False
    )
    atomic_write(path, (rendered + "\n").encode("utf-8"))


def render_tsv(rows: list[dict[str, object]], columns: list[str]) -> bytes:
    if not rows:
        raise AuditBlocked("拒绝渲染空 TSV")
    normalized_rows = [
        {
            key: ("true" if value else "false") if isinstance(value, bool) else value
            for key, value in row.items()
        }
        for row in rows
    ]
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(normalized_rows)
    return buffer.getvalue().encode("utf-8")


def hash_stream(handle: BinaryIO, algorithm: str = "sha256") -> str:
    value = hashlib.new(algorithm)
    for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
        value.update(block)
    return value.hexdigest()


def file_hash(path: Path, algorithm: str = "sha256") -> str:
    require_file(path)
    with path.open("rb") as handle:
        return hash_stream(handle, algorithm)


def load_json(path: Path) -> dict[str, object]:
    require_file(path)
    raw = path.read_bytes()
    lowered = raw.lower()
    if b"x-amz-" in lowered or b"x-goog-" in lowered or b"signature=" in lowered:
        raise AuditBlocked(f"元数据持久化了临时签名：{path}")
    try:
        value = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise AuditBlocked(f"JSON 无效：{path}: {exc}") from exc
    if not isinstance(value, dict):
        raise AuditBlocked(f"JSON 顶层不是对象：{path}")
    return value


def validate_raw_api_captures(metadata: dict[str, object], expected_urls: set[str]) -> None:
    if metadata.get("raw_api_capture_format") != "exact_response_bytes_base64_with_sha256":
        raise AuditBlocked("官方 API 未保存精确响应字节快照")
    captures = metadata.get("raw_api_captures")
    if not isinstance(captures, list) or len(captures) != len(expected_urls):
        raise AuditBlocked("官方 API 响应快照数量不符")
    seen: set[str] = set()
    for capture in captures:
        if not isinstance(capture, dict):
            raise AuditBlocked("官方 API 响应快照结构错误")
        request_url = str(capture.get("request_url", ""))
        final_url = str(capture.get("final_url", ""))
        if request_url not in expected_urls or request_url in seen:
            raise AuditBlocked(f"官方 API 快照端点重复或越界：{request_url}")
        seen.add(request_url)
        request_parsed = urlsplit(request_url)
        final_parsed = urlsplit(final_url)
        if (
            request_parsed.scheme != "https"
            or final_parsed.scheme != "https"
            or request_parsed.hostname != final_parsed.hostname
            or int(capture.get("status", -1)) != 200
        ):
            raise AuditBlocked(f"官方 API 快照重定向或状态异常：{request_url}")
        try:
            payload = base64.b64decode(str(capture.get("payload_base64", "")), validate=True)
            json.loads(payload.decode("utf-8"))
        except (ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise AuditBlocked(f"官方 API 快照载荷不可复核：{request_url}") from exc
        if (
            len(payload) != int(capture.get("payload_bytes", -1))
            or hashlib.sha256(payload).hexdigest()
            != str(capture.get("payload_sha256", "")).lower()
        ):
            raise AuditBlocked(f"官方 API 快照字节或SHA256不符：{request_url}")
    if seen != expected_urls:
        raise AuditBlocked("官方 API 快照端点集合不闭合")


def load_manifest(path: Path) -> list[dict[str, str]]:
    require_file(path)
    raw = path.read_bytes()
    lowered = raw.lower()
    if b"x-amz-" in lowered or b"x-goog-" in lowered or b"signature=" in lowered:
        raise AuditBlocked(f"清单持久化了临时签名：{path}")
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise AuditBlocked(f"TSV 不是 UTF-8：{path}") from exc
    rows = list(csv.DictReader(io.StringIO(text), delimiter="\t"))
    if len(rows) != 1:
        raise AuditBlocked(f"官方清单不再是单文件记录：{path}")
    return rows


def validate_local_governance(
    directory: Path,
    *,
    provider: str,
    record_id: str,
    supplement_doi: str,
    resource_doi: str,
    filename: str,
    size: int,
    sha256: str,
) -> tuple[dict[str, object], list[dict[str, str]]]:
    metadata_path = directory / "官方API元数据.json"
    manifest_path = directory / "官方文件清单.tsv"
    metadata = load_json(metadata_path)
    manifest = load_manifest(manifest_path)
    row = manifest[0]
    endpoints = metadata.get("official_endpoints") or {}
    if not isinstance(endpoints, dict):
        raise AuditBlocked(f"官方端点结构错误：{directory.name}")
    capture_urls = {
        str(value) for key, value in endpoints.items() if key != "stable_download"
    }
    expected_capture_count = 3 if provider == "Mendeley Data" else 2
    if len(capture_urls) != expected_capture_count:
        raise AuditBlocked(f"官方 API 捕获端点数量错误：{directory.name}")
    validate_raw_api_captures(metadata, capture_urls)
    if str(metadata.get("provider")) != provider:
        raise AuditBlocked(f"来源提供方漂移：{directory.name}")
    if provider == "Mendeley Data":
        if (
            str(metadata.get("dataset_id")) != record_id
            or str(metadata.get("doi", "")).casefold() != supplement_doi.casefold()
            or str((metadata.get("license") or {}).get("short_name", "")).casefold()
            != "cc by 4.0"
        ):
            raise AuditBlocked("Mendeley 本地元数据身份或许可错误")
    else:
        primary_article = metadata.get("primary_article_crossref") or {}
        citation_notes = metadata.get("citation_quality_notes") or []
        if (
            str(metadata.get("article_id")) != record_id
            or str(metadata.get("supplement_doi", "")).casefold()
            != supplement_doi.casefold()
            or str(metadata.get("resource_doi", "")).casefold()
            != resource_doi.casefold()
            or str((metadata.get("license") or {}).get("name", "")).casefold()
            != "cc by-nc 4.0"
            or str(primary_article.get("doi", "")).casefold() != resource_doi.casefold()
            or str(primary_article.get("type", "")) != "journal-article"
            or not primary_article.get("published_date_parts")
        ):
            raise AuditBlocked(
                f"ACS Figshare 本地元数据身份或许可错误：{directory.name}"
            )
        published_year = int(str(metadata.get("published_date", ""))[:4])
        if int(primary_article["published_date_parts"][0]) != published_year:
            raise AuditBlocked(f"主论文与 SI 发布年份冲突：{directory.name}")
        if record_id == "31614502" and not any(
            "1753" in str(note) and "2026" in str(note) for note in citation_notes
        ):
            raise AuditBlocked("聚酰亚胺链扩剂来源未保留 Figshare 错误年份说明")
    if (
        row.get("provider") != provider
        or row.get("record_id") != record_id
        or row.get("supplement_doi", "").casefold() != supplement_doi.casefold()
        or row.get("resource_doi", "").casefold() != resource_doi.casefold()
        or row.get("filename") != filename
        or int(row.get("bytes", "-1")) != size
        or row.get("local_sha256", "").lower() != sha256
        or row.get("local_state") != "verified_present"
        or row.get("signed_redirect_url_persisted", "").casefold() != "false"
    ):
        raise AuditBlocked(f"本地官方文件清单漂移：{directory.name}")
    stable_url = row.get("stable_download_url", "")
    parsed = urlsplit(stable_url)
    if parsed.scheme != "https" or parsed.query or parsed.fragment:
        raise AuditBlocked(f"清单下载 URL 非稳定端点：{stable_url}")
    return metadata, manifest


def safe_outer_member(name: str) -> None:
    if "\\" in name or not name or name.startswith("/"):
        raise AuditBlocked(f"ZIP 成员路径不安全：{name!r}")
    parts = PurePosixPath(name).parts
    if any(part in {"", ".", ".."} or ":" in part for part in parts):
        raise AuditBlocked(f"ZIP 成员路径越界：{name!r}")


def zip_member_sha256(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> str:
    with archive.open(info) as handle:
        return hash_stream(handle)


def workbook_formula_counts(workbook) -> tuple[int, int, list[tuple[str, str]]]:
    formula_count = 0
    external_count = 0
    external_cells: list[tuple[str, str]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if "[" in value or "]" in value:
                        external_count += 1
                        external_cells.append((sheet.title, cell.coordinate))
    return formula_count, external_count, external_cells


RECORD_COLUMNS = [
    "source_directory",
    "record_id",
    "specimen_family_id",
    "source_file",
    "source_location",
    "material_grade",
    "record_kind",
    "geometry_or_evidence",
    "process_or_protocol",
    "reported_items",
    "variables",
    "direct_numeric_count",
    "derived_formula_count",
    "invalid_cached_formula_count",
    "missing_numeric_count",
    "parse_state",
    "anomaly",
    "manual_action",
    "physical_specimen",
    "training_split_materialized",
    "training_weight_materialized",
]
FILE_COLUMNS = [
    "source_directory",
    "path",
    "role",
    "bytes",
    "sha256",
    "integrity",
    "parser_state",
    "redistribution_note",
    "training_split_materialized",
    "training_weight_materialized",
]


def audit_mendeley() -> tuple[
    dict[str, object], list[dict[str, object]], list[dict[str, object]]
]:
    directory = DATA_ROOT / MENDELEY
    require_directory(directory)
    filename, expected_size, expected_sha256 = MENDELEY_ARCHIVE
    archive_path = directory / filename
    require_file(archive_path)
    if (
        archive_path.stat().st_size != expected_size
        or file_hash(archive_path) != expected_sha256
    ):
        raise AuditBlocked("Mendeley ZIP 大小或 SHA256 错误")
    metadata, _ = validate_local_governance(
        directory,
        provider="Mendeley Data",
        record_id="7zcd9bmmg5",
        supplement_doi="10.17632/7zcd9bmmg5.1",
        resource_doi="",
        filename=filename,
        size=expected_size,
        sha256=expected_sha256,
    )

    file_rows: list[dict[str, object]] = []
    record_rows: list[dict[str, object]] = []
    file_rows.append(
        {
            "source_directory": MENDELEY,
            "path": filename,
            "role": "official_archive",
            "bytes": expected_size,
            "sha256": expected_sha256,
            "integrity": "verified",
            "parser_state": "zip_container_audited",
            "redistribution_note": "CC BY 4.0",
            "training_split_materialized": False,
            "training_weight_materialized": False,
        }
    )

    with zipfile.ZipFile(archive_path) as archive:
        infos = archive.infolist()
        if archive.testzip() is not None:
            raise AuditBlocked("Mendeley ZIP CRC 校验失败")
        names = [info.filename for info in infos]
        if set(names) != set(MENDELEY_MEMBERS) or len(names) != len(MENDELEY_MEMBERS):
            raise AuditBlocked("Mendeley ZIP 成员清单漂移")
        if len({name.casefold() for name in names}) != len(names):
            raise AuditBlocked("Mendeley ZIP 存在大小写冲突成员")
        member_bytes: dict[str, bytes] = {}
        for info in infos:
            safe_outer_member(info.filename)
            expected_member_size, expected_member_hash = MENDELEY_MEMBERS[info.filename]
            actual_hash = zip_member_sha256(archive, info)
            if (
                info.file_size != expected_member_size
                or actual_hash != expected_member_hash
            ):
                raise AuditBlocked(f"Mendeley ZIP 成员漂移：{info.filename}")
            payload = archive.read(info)
            member_bytes[info.filename] = payload
            parser_state = (
                "xlsx_parsed"
                if info.filename.endswith(".xlsx")
                else "mpx_zip_audited_not_parsed"
            )
            file_rows.append(
                {
                    "source_directory": MENDELEY,
                    "path": info.filename,
                    "role": "archive_member",
                    "bytes": info.file_size,
                    "sha256": actual_hash,
                    "integrity": "verified",
                    "parser_state": parser_state,
                    "redistribution_note": "CC BY 4.0; MPX kept opaque",
                    "training_split_materialized": False,
                    "training_weight_materialized": False,
                }
            )

    nested_stats = {}
    for name in (POLY_MPX, NINJA_MPX):
        try:
            with zipfile.ZipFile(BytesIO(member_bytes[name])) as nested:
                if nested.testzip() is not None:
                    raise AuditBlocked(f"MPX 内部 ZIP CRC 错误：{name}")
                nested_stats[name] = {
                    "entry_count": len(nested.infolist()),
                    "uncompressed_bytes": sum(
                        info.file_size for info in nested.infolist()
                    ),
                    "has_leading_slash_members": any(
                        info.filename.startswith("/") for info in nested.infolist()
                    ),
                    "extraction_allowed": False,
                    "scientific_values_parsed": False,
                }
        except zipfile.BadZipFile as exc:
            raise AuditBlocked(f"MPX 不是可读 ZIP 容器：{name}") from exc
    if (
        nested_stats[POLY_MPX]["entry_count"] != 162
        or nested_stats[NINJA_MPX]["entry_count"] != 118
    ):
        raise AuditBlocked("MPX 内部条目数量漂移")

    ninja_formula = load_workbook(
        BytesIO(member_bytes[NINJA_WORKBOOK]),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    ninja_values = load_workbook(
        BytesIO(member_bytes[NINJA_WORKBOOK]),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    poly_formula = load_workbook(
        BytesIO(member_bytes[POLY_WORKBOOK]),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    poly_values = load_workbook(
        BytesIO(member_bytes[POLY_WORKBOOK]),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if {
            name: (ninja_formula[name].max_row, ninja_formula[name].max_column)
            for name in ninja_formula.sheetnames
        } != NINJA_SHAPES:
            raise AuditBlocked("NinjaFlex 工作簿形状漂移")
        if {
            name: (poly_formula[name].max_row, poly_formula[name].max_column)
            for name in poly_formula.sheetnames
        } != POLY_SHAPES:
            raise AuditBlocked("PolyFlex 工作簿形状漂移")

        ninja_formula_count, ninja_external_count, ninja_external_cells = (
            workbook_formula_counts(ninja_formula)
        )
        poly_formula_count, poly_external_count, poly_external_cells = (
            workbook_formula_counts(poly_formula)
        )
        if (ninja_formula_count, ninja_external_count) != (1_203, 10):
            raise AuditBlocked("NinjaFlex 公式或外链公式计数漂移")
        if (poly_formula_count, poly_external_count) != (831, 0):
            raise AuditBlocked("PolyFlex 公式计数漂移")
        if {sheet for sheet, _ in ninja_external_cells} != {
            "Peak stresses"
        } or poly_external_cells:
            raise AuditBlocked("外链公式不再只位于 NinjaFlex Peak stresses")

        ninja_raw = ninja_formula["Raw Data"]
        ninja_cached = ninja_values["Raw Data"]
        cube_rows = []
        for row_number in range(2, ninja_raw.max_row + 1):
            values = [
                ninja_raw.cell(row_number, column).value for column in range(1, 28)
            ]
            if values[4] in {"A", "B", "C", "D"}:
                cube_rows.append((row_number, values))
        if len(cube_rows) != 72 or len({tuple(row[1][:4]) for row in cube_rows}) != 18:
            raise AuditBlocked("NinjaFlex 立方体试样或设计计数漂移")
        for row_number, values in cube_rows:
            loads = values[5:13]
            area = values[13]
            if not all(isinstance(value, (int, float)) for value in (*loads, area)):
                raise AuditBlocked(f"NinjaFlex 直接载荷或面积缺失：第 {row_number} 行")
            for offset, load in enumerate(loads):
                formula = ninja_raw.cell(row_number, 20 + offset).value
                observed = ninja_cached.cell(row_number, 20 + offset).value
                expected = load / area * 6.89476
                if not isinstance(formula, str) or not formula.startswith("="):
                    raise AuditBlocked(f"NinjaFlex 应力公式缺失：第 {row_number} 行")
                if not isinstance(observed, (int, float)) or not math.isclose(
                    observed, expected, rel_tol=1e-12, abs_tol=1e-8
                ):
                    raise AuditBlocked(f"NinjaFlex 应力复算失败：第 {row_number} 行")
            record_rows.append(
                {
                    "source_directory": MENDELEY,
                    "record_id": f"ninjaflex_cube_r{row_number}",
                    "specimen_family_id": (
                        f"ninjaflex|cube|{values[0]}|{values[1]}|"
                        f"{values[2]}|{values[3]}"
                    ),
                    "source_file": NINJA_WORKBOOK,
                    "source_location": f"Raw Data!{row_number}",
                    "material_grade": "NinjaFlex (NinjaTek; exact chemistry/hardness unavailable)",
                    "record_kind": "compression_discrete_strain",
                    "geometry_or_evidence": "cube",
                    "process_or_protocol": f"{values[1]}; infill={values[2]}%; shells={values[3]}; directions=2; strain=5/10/15/20%",
                    "reported_items": "one physical specimen",
                    "variables": "load_lbf;area_in2;derived_stress_kPa",
                    "direct_numeric_count": 9,
                    "derived_formula_count": 8,
                    "invalid_cached_formula_count": 0,
                    "missing_numeric_count": 0,
                    "parse_state": "direct_and_recomputed",
                    "anomaly": "",
                    "manual_action": "训练前按 material+geometry+DOE configuration 分组切分",
                    "physical_specimen": True,
                    "training_split_materialized": False,
                    "training_weight_materialized": False,
                }
            )

        cylinder_formula = ninja_formula["NinjaFlex Cylinders - Do Not S."]
        cylinder_cached = ninja_values["NinjaFlex Cylinders - Do Not S."]
        cylinder_count = 0
        cylinder_direct = 0
        cylinder_specimens = 0
        solid_cube_control_specimens = 0
        solid_cube_control_direct = 0
        solid_cube_control_derived = 0
        if "cube psi data" not in str(cylinder_formula.cell(34, 3).value).casefold():
            raise AuditBlocked("NinjaFlex 9999-bottom-layer 原始行不再标识为 cube psi data")
        for row_number in range(2, cylinder_formula.max_row + 1):
            values = [
                cylinder_formula.cell(row_number, column).value
                for column in range(1, 10)
            ]
            if values[4] not in {"A", "B", "C", "D"}:
                continue
            cylinder_count += 1
            direct_count = sum(isinstance(value, (int, float)) for value in values[5:9])
            formula_count = sum(
                isinstance(value, str) and value.startswith("=")
                for value in values[5:9]
            )
            cached_values = [
                cylinder_cached.cell(row_number, column).value
                for column in range(6, 10)
            ]
            if direct_count + formula_count != 4 or not all(
                isinstance(value, (int, float)) for value in cached_values
            ):
                raise AuditBlocked(f"NinjaFlex 圆柱应力字段异常：第 {row_number} 行")
            is_solid_cube_control = formula_count > 0
            source_values: list[object] = []
            source_row: int | None = None
            if is_solid_cube_control:
                if (
                    direct_count != 0
                    or formula_count != 4
                    or values[0] is not None
                    or values[1] != "9999 bottom layers"
                    or values[2:4] != [1, 2]
                    or row_number not in range(14, 18)
                ):
                    raise AuditBlocked(
                        f"NinjaFlex 实心立方体控制组身份漂移：第 {row_number} 行"
                    )
                source_row = row_number + 20
                source_values = [
                    cylinder_formula.cell(source_row, column).value
                    for column in range(6, 10)
                ]
                if not all(isinstance(value, (int, float)) for value in source_values):
                    raise AuditBlocked(
                        f"NinjaFlex 实心立方体控制组原始 psi 缺失：第 {source_row} 行"
                    )
                for column, (formula, source_value, observed) in enumerate(
                    zip(values[5:9], source_values, cached_values, strict=True), start=6
                ):
                    letter = get_column_letter(column)
                    expected_formula = f"={letter}{source_row}*6.89476"
                    if str(formula).replace(" ", "") != expected_formula:
                        raise AuditBlocked(
                            f"NinjaFlex 实心立方体控制组公式血缘漂移："
                            f"第 {row_number} 行/{letter}"
                        )
                    if not math.isclose(
                        float(observed),
                        float(source_value) * 6.89476,
                        rel_tol=1e-12,
                        abs_tol=1e-8,
                    ):
                        raise AuditBlocked(
                            f"NinjaFlex 实心立方体控制组单位换算失败："
                            f"第 {row_number} 行/{letter}"
                        )
                solid_cube_control_specimens += 1
                solid_cube_control_direct += 4
                solid_cube_control_derived += 4
                geometry = "solid_cube_control"
                family_id = "ninjaflex|solid_cube_control|9999_bottom_layers|1|2"
                record_id = f"ninjaflex_solid_cube_control_r{row_number}"
                source_location = (
                    f"NinjaFlex Cylinders - Do Not S.!{row_number};"
                    f"raw_source_F{source_row}:I{source_row}"
                )
                variables = "source_stress_psi;derived_stress_kPa"
                parse_state = "direct_and_recomputed_from_source_row"
                anomaly = "stored in cylinder worksheet but source note identifies cube psi data"
                manual_action = "按实心立方体控制组聚类；不得与圆柱或四个重复试样随机拆分"
                record_direct_count = 4
                record_derived_count = 4
            else:
                if direct_count != 4:
                    raise AuditBlocked(f"NinjaFlex 圆柱直接应力数量漂移：第 {row_number} 行")
                cylinder_specimens += 1
                cylinder_direct += 4
                geometry = "cylinder"
                record_id = f"ninjaflex_cylinder_r{row_number}"
                family_id = (
                    f"ninjaflex|cylinder|{values[0]}|{values[1]}|"
                    f"{values[2]}|{values[3]}"
                )
                source_location = f"NinjaFlex Cylinders - Do Not S.!{row_number}"
                variables = "stress_kPa"
                parse_state = "direct_numeric"
                anomaly = ""
                manual_action = "按材料+圆柱DOE组合聚类；四个重复试样不得随机拆分"
                record_direct_count = 4
                record_derived_count = 0
            record_rows.append(
                {
                    "source_directory": MENDELEY,
                    "record_id": record_id,
                    "specimen_family_id": family_id,
                    "source_file": NINJA_WORKBOOK,
                    "source_location": source_location,
                    "material_grade": "NinjaFlex (NinjaTek; exact chemistry/hardness unavailable)",
                    "record_kind": "compression_discrete_strain",
                    "geometry_or_evidence": geometry,
                    "process_or_protocol": f"{values[1]}; infill={values[2]}; shells={values[3]}; direction=1; strain=5/10/15/20%",
                    "reported_items": "one physical specimen",
                    "variables": variables,
                    "direct_numeric_count": record_direct_count,
                    "derived_formula_count": record_derived_count,
                    "invalid_cached_formula_count": 0,
                    "missing_numeric_count": 0,
                    "parse_state": parse_state,
                    "anomaly": anomaly,
                    "manual_action": manual_action,
                    "physical_specimen": True,
                    "training_split_materialized": False,
                    "training_weight_materialized": False,
                }
            )
        if (
            cylinder_count,
            cylinder_specimens,
            cylinder_direct,
            solid_cube_control_specimens,
            solid_cube_control_direct,
            solid_cube_control_derived,
        ) != (24, 20, 80, 4, 16, 16):
            raise AuditBlocked("NinjaFlex 圆柱/实心立方体控制组计数或血缘漂移")

        poly_raw = poly_formula["Raw data - Do not sort"]
        poly_cached = poly_values["Raw data - Do not sort"]
        poly_count = 0
        poly_direct = 0
        poly_missing = 0
        poly_weights = 0
        missing_cells: list[str] = []
        for row_number in range(3, poly_raw.max_row + 1):
            values = [
                poly_raw.cell(row_number, column).value for column in range(1, 31)
            ]
            if values[4] not in {"A", "B", "C", "D"}:
                continue
            poly_count += 1
            direct_values = values[6:14]
            direct_count = sum(
                isinstance(value, (int, float)) for value in direct_values
            )
            missing_count = 8 - direct_count
            poly_direct += direct_count
            poly_missing += missing_count
            poly_weights += int(isinstance(values[17], (int, float)))
            for offset, psi in enumerate(direct_values):
                formula = poly_raw.cell(row_number, 23 + offset).value
                observed = poly_cached.cell(row_number, 23 + offset).value
                source_letter = get_column_letter(7 + offset)
                expected_formula = f"={source_letter}{row_number}*6.89476"
                if (
                    not isinstance(formula, str)
                    or formula.replace(" ", "") != expected_formula
                ):
                    raise AuditBlocked(f"PolyFlex kPa 公式缺失：第 {row_number} 行")
                if isinstance(psi, (int, float)):
                    expected = psi * 6.89476
                    if not isinstance(observed, (int, float)) or not math.isclose(
                        observed, expected, rel_tol=1e-12, abs_tol=1e-8
                    ):
                        raise AuditBlocked(
                            f"PolyFlex 单位换算复算失败：第 {row_number} 行"
                        )
                else:
                    missing_cells.append(f"{source_letter}{row_number}")
                    if not isinstance(observed, (int, float)) or not math.isclose(
                        float(observed), 0.0, rel_tol=0.0, abs_tol=0.0
                    ):
                        raise AuditBlocked(
                            f"PolyFlex 缺失输入的缓存不再是待隔离伪零："
                            f"{source_letter}{row_number}"
                        )
            anomaly = ""
            if missing_count:
                anomaly = "horizontal direction missing at 5/10/15/20% strain; derived kPa caches are zero and must be ignored"
            record_rows.append(
                {
                    "source_directory": MENDELEY,
                    "record_id": f"polyflex_cube_r{row_number}",
                    "specimen_family_id": (
                        f"polyflex|cube|{values[0]}|{values[1]}|"
                        f"{values[2]}|{values[3]}"
                    ),
                    "source_file": POLY_WORKBOOK,
                    "source_location": f"Raw data - Do not sort!{row_number}",
                    "material_grade": "PolyFlex (exact chemistry/hardness unavailable)",
                    "record_kind": "compression_discrete_strain",
                    "geometry_or_evidence": "cube",
                    "process_or_protocol": f"{values[1]}; infill={values[2]}%; shells={values[3]}; directions=2; strain=5/10/15/20%",
                    "reported_items": "one physical specimen",
                    "variables": "stress_psi;derived_stress_kPa;weight_g",
                    "direct_numeric_count": direct_count
                    + int(isinstance(values[17], (int, float))),
                    "derived_formula_count": 8 - missing_count,
                    "invalid_cached_formula_count": missing_count,
                    "missing_numeric_count": missing_count,
                    "parse_state": "direct_and_recomputed"
                    if not missing_count
                    else "partial_hold",
                    "anomaly": anomaly,
                    "manual_action": "含 Old/New 标签的重复配置是独立试样批次；组间切分不得随机拆散",
                    "physical_specimen": True,
                    "training_split_materialized": False,
                    "training_weight_materialized": False,
                }
            )
        if (poly_count, poly_direct, poly_missing, poly_weights) != (88, 700, 4, 56):
            raise AuditBlocked("PolyFlex 试样、直接值、缺失值或重量计数漂移")
        if missing_cells != ["K19", "L19", "M19", "N19"]:
            raise AuditBlocked(f"PolyFlex 缺失位置漂移：{missing_cells}")

        for name, label in ((POLY_MPX, "polyflex_mpx"), (NINJA_MPX, "ninjaflex_mpx")):
            record_rows.append(
                {
                    "source_directory": MENDELEY,
                    "record_id": label,
                    "specimen_family_id": "",
                    "source_file": name,
                    "source_location": "nested ZIP project",
                    "material_grade": "PolyFlex" if name == POLY_MPX else "NinjaFlex",
                    "record_kind": "proprietary_DOE_project",
                    "geometry_or_evidence": "Minitab MPX",
                    "process_or_protocol": "DOE analysis project",
                    "reported_items": str(nested_stats[name]["entry_count"]),
                    "variables": "not parsed",
                    "direct_numeric_count": 0,
                    "derived_formula_count": 0,
                    "invalid_cached_formula_count": 0,
                    "missing_numeric_count": 0,
                    "parse_state": "container_only_hold",
                    "anomaly": "leading-slash internal member names; extraction prohibited",
                    "manual_action": "仅在隔离环境用受信 Minitab 导出纯文本后再审计",
                    "physical_specimen": False,
                    "training_split_materialized": False,
                    "training_weight_materialized": False,
                }
            )

        physical_rows = [row for row in record_rows if row["physical_specimen"]]
        family_ids = {
            str(row["specimen_family_id"])
            for row in physical_rows
            if str(row["specimen_family_id"])
        }
        if len(physical_rows) != 184 or len(family_ids) != 46:
            raise AuditBlocked("物理试样或 DOE/重复试样家族计数漂移")

        summary = {
            "audit_date": AUDIT_DATE,
            "audit_version": AUDIT_VERSION,
            "source_directory": MENDELEY,
            "provider": "Mendeley Data",
            "dataset_id": "7zcd9bmmg5",
            "doi": "10.17632/7zcd9bmmg5.1",
            "license": "CC BY 4.0",
            "title": metadata.get("title"),
            "scientific_layer": "direct_experiment_application_process",
            "materials": ["NinjaFlex", "PolyFlex"],
            "physical_specimens": 184,
            "physical_specimen_family_count": 46,
            "ninjaflex_doe_cube_specimens": 72,
            "ninjaflex_cylinder_specimens": 20,
            "ninjaflex_solid_cube_control_specimens": 4,
            "polyflex_cube_specimens": 88,
            "specimen_family_counts": {
                "ninjaflex_doe_cube": 18,
                "ninjaflex_cylinder": 5,
                "ninjaflex_solid_cube_control": 1,
                "polyflex_cube": 22,
            },
            "complete_direct_measurements": {
                "ninjaflex_cube_load_values": 576,
                "ninjaflex_cylinder_stress_values": 80,
                "ninjaflex_solid_cube_control_stress_psi_values": 16,
                "polyflex_stress_psi_values": 700,
            },
            "valid_derived_formula_values": {
                "ninjaflex_cube_stress_kPa": 576,
                "ninjaflex_solid_cube_control_stress_kPa": 16,
                "polyflex_stress_kPa": 700,
                "total": 1_292,
            },
            "invalid_cached_formula_values": {
                "polyflex_missing_input_pseudo_zero": 4,
            },
            "known_missing_direct_values": 4,
            "known_missing_cells": missing_cells,
            "workbook_formula_counts": {
                "ninjaflex_total": ninja_formula_count,
                "ninjaflex_external": ninja_external_count,
                "polyflex_total": poly_formula_count,
                "polyflex_external": poly_external_count,
            },
            "external_formula_cells_quarantined": [
                f"{sheet}!{cell}" for sheet, cell in ninja_external_cells
            ],
            "external_formula_policy": "10 external formulas in Peak stresses are quarantined; cached outputs are not supervision",
            "mpx_projects": nested_stats,
            "protocol_from_official_metadata": {
                "extruder_temperature_C": 225,
                "bed_temperature_C": {"NinjaFlex": 55, "PolyFlex": 30},
                "layer_height_mm": 0.2,
                "travel_speed_mm_per_s": 175,
                "test_temperature_C": 21,
                "deflection_rate_mm_per_min": 12.65,
                "acquisition_rate_Hz": 10,
                "peak_deflection_percent": 20,
                "adapted_standard": "ASTM D575-91",
            },
            "limitations": [
                "commercial grade chemistry and nominal hardness are absent",
                "dataset does not contain continuous stress-displacement histories",
                "MPX analysis projects are proprietary and intentionally not deserialized",
                "Old/New labels denote distinct batches but exact chronology is incomplete",
            ],
            "training_split_materialized": False,
            "training_weight_materialized": False,
            "training_state": "held_pending_source_policy_and_group_split",
        }
    finally:
        ninja_formula.close()
        ninja_values.close()
        poly_formula.close()
        poly_values.close()

    for governance_name, role in (
        ("官方API元数据.json", "official_metadata"),
        ("官方文件清单.tsv", "official_manifest"),
    ):
        path = directory / governance_name
        file_rows.append(
            {
                "source_directory": MENDELEY,
                "path": governance_name,
                "role": role,
                "bytes": path.stat().st_size,
                "sha256": file_hash(path),
                "integrity": "parsed",
                "parser_state": "governance_validated",
                "redistribution_note": "locally generated normalized metadata",
                "training_split_materialized": False,
                "training_weight_materialized": False,
            }
        )
    return summary, file_rows, record_rows


def normalized_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def audit_pdf(
    spec: PdfSpec,
) -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    directory = DATA_ROOT / spec.directory
    require_directory(directory)
    pdf_path = directory / spec.filename
    require_file(pdf_path)
    if pdf_path.stat().st_size != spec.size or file_hash(pdf_path) != spec.sha256:
        raise AuditBlocked(f"PDF 大小或 SHA256 错误：{spec.directory}")
    metadata, _ = validate_local_governance(
        directory,
        provider="ACS Figshare",
        record_id=str(spec.article_id),
        supplement_doi=spec.supplement_doi,
        resource_doi=spec.resource_doi,
        filename=spec.filename,
        size=spec.size,
        sha256=spec.sha256,
    )

    try:
        reader = PdfReader(pdf_path, strict=True)
    except Exception as exc:  # noqa: BLE001
        raise AuditBlocked(f"PDF 严格解析失败：{spec.filename}: {exc}") from exc
    if reader.is_encrypted or len(reader.pages) != spec.pages:
        raise AuditBlocked(f"PDF 加密状态或页数漂移：{spec.filename}")
    root = reader.trailer["/Root"]
    dangerous_catalog_keys = [
        key for key in ("/Names", "/OpenAction", "/AA") if root.get(key) is not None
    ]
    if dangerous_catalog_keys:
        raise AuditBlocked(
            f"PDF 含活动目录键：{spec.filename}: {dangerous_catalog_keys}"
        )
    try:
        page_texts = [page.extract_text() or "" for page in reader.pages]
    except Exception as exc:  # noqa: BLE001
        raise AuditBlocked(f"PDF 文本提取失败：{spec.filename}: {exc}") from exc
    text = normalized_text("\n".join(page_texts))
    if len(text) < spec.min_text_chars:
        raise AuditBlocked(f"PDF 可提取文本异常减少：{spec.filename}")
    missing_anchors = [
        anchor for anchor in spec.anchors if normalized_text(anchor) not in text
    ]
    if missing_anchors:
        raise AuditBlocked(f"PDF 内容锚点缺失：{spec.filename}: {missing_anchors}")

    file_rows = []
    for name, role in (
        (spec.filename, "scientific_supporting_information"),
        ("官方API元数据.json", "official_metadata"),
        ("官方文件清单.tsv", "official_manifest"),
    ):
        path = directory / name
        file_rows.append(
            {
                "source_directory": spec.directory,
                "path": name,
                "role": role,
                "bytes": path.stat().st_size,
                "sha256": file_hash(path),
                "integrity": "verified" if name == spec.filename else "parsed",
                "parser_state": "pdf_evidence_indexed"
                if name == spec.filename
                else "governance_validated",
                "redistribution_note": "CC BY-NC 4.0; preserve attribution and non-commercial restriction",
                "training_split_materialized": False,
                "training_weight_materialized": False,
            }
        )
    record_rows = [
        {
            "source_directory": spec.directory,
            "record_id": group.group_id,
            "specimen_family_id": "",
            "source_file": spec.filename,
            "source_location": group.evidence_object,
            "material_grade": "paper-defined PU/TPU series",
            "record_kind": "literature_evidence_group",
            "geometry_or_evidence": group.evidence_object,
            "process_or_protocol": "see supporting information and primary article",
            "reported_items": group.reported_items,
            "variables": group.variables,
            "direct_numeric_count": 0,
            "derived_formula_count": 0,
            "invalid_cached_formula_count": 0,
            "missing_numeric_count": 0,
            "parse_state": "evidence_only_not_materialized",
            "anomaly": group.anomaly,
            "manual_action": group.manual_action,
            "physical_specimen": False,
            "training_split_materialized": False,
            "training_weight_materialized": False,
        }
        for group in spec.groups
    ]
    summary = {
        "audit_date": AUDIT_DATE,
        "audit_version": AUDIT_VERSION,
        "source_directory": spec.directory,
        "provider": "ACS Figshare",
        "article_id": spec.article_id,
        "supplement_doi": spec.supplement_doi,
        "resource_doi": spec.resource_doi,
        "citation": metadata.get("citation"),
        "primary_article_crossref": metadata.get("primary_article_crossref"),
        "citation_quality_notes": metadata.get("citation_quality_notes"),
        "license": "CC BY-NC 4.0",
        "title": metadata.get("title"),
        "scientific_layer": "literature_supporting_information_evidence",
        "file": {
            "filename": spec.filename,
            "bytes": spec.size,
            "sha256": spec.sha256,
            "pages": spec.pages,
            "extracted_text_chars_normalized": len(text),
            "encrypted": False,
            "active_catalog_keys": [],
        },
        "evidence_group_count": len(spec.groups),
        "evidence_groups": [group.group_id for group in spec.groups],
        "record_materialization": "none",
        "reason_not_direct_training": (
            "PDF tables and figures require page-aware human verification; text order is not a "
            "lossless table representation and plotted curves are not raw instrument data"
        ),
        "required_next_step": (
            "dual-entry extraction with page/table/figure coordinates, unit normalization, "
            "sample identity reconciliation, and second-person verification"
        ),
        "training_split_materialized": False,
        "training_weight_materialized": False,
        "training_state": "evidence_only_hold",
    }
    return summary, file_rows, record_rows


def write_source_outputs(
    source: str,
    summary: dict[str, object],
    file_rows: list[dict[str, object]],
    record_rows: list[dict[str, object]],
) -> None:
    directory = DATA_ROOT / source
    write_json(directory / "内容审计摘要.json", summary)
    atomic_write(
        directory / "文件校验清单.tsv",
        render_tsv(
            sorted(file_rows, key=lambda row: str(row["path"]).casefold()), FILE_COLUMNS
        ),
    )
    atomic_write(
        directory / "曲线审计清单.tsv",
        render_tsv(
            sorted(record_rows, key=lambda row: str(row["record_id"]).casefold()),
            RECORD_COLUMNS,
        ),
    )


def main() -> int:
    if len(PDF_SPECS) != 8 or {spec.directory for spec in PDF_SPECS} != set(
        PDF_SOURCES
    ):
        raise AuditBlocked("固定 PDF 来源清单错误")
    summary, file_rows, record_rows = audit_mendeley()
    write_source_outputs(MENDELEY, summary, file_rows, record_rows)
    print(
        f"{MENDELEY}: specimens={summary['physical_specimens']}; records={len(record_rows)}",
        flush=True,
    )
    result = [{"source": MENDELEY, "records": len(record_rows)}]
    for spec in PDF_SPECS:
        summary, file_rows, record_rows = audit_pdf(spec)
        write_source_outputs(spec.directory, summary, file_rows, record_rows)
        print(
            f"{spec.directory}: pages={spec.pages}; evidence_groups={len(record_rows)}",
            flush=True,
        )
        result.append({"source": spec.directory, "records": len(record_rows)})
    print(json.dumps(result, ensure_ascii=False, sort_keys=True, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"BLOCKED: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc
