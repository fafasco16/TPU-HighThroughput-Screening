"""审计 Sheffield Figshare 的 PU 泡沫理性设计数据。

数据集 DOI：10.15131/shef.data.21510876.v1
论文 DOI：10.3390/polym14235111

脚本从冻结的 90 MB 官方 ZIP 中只读解析 CSV、XLSX 和嵌套 DOCX，绝不
把原件解压到工作区。补充表 S1 有 40 个样品批次；F05-13 与 F05-34
配方完全相同，因此是 39 个唯一配方（23 个筛选配方 + 16 个最终 DoE
配方）。同配方数据共享 split_group，约 40 个动力学文件不会被误计为
40 个独立材料。

原文件名 ATRsumm 中的 ATR 是 adiabatic temperature rise（绝热温升），
不是 ATR-FTIR。输出用于 Gold-E 配方/工艺—动力学—形貌—流体输运参考；
商业原料没有 SMILES，不能直接作为单体结构到性能的高权重监督。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第九批实验_Sheffield_PU理性设计"
)
SOURCE_ZIP = SOURCE_DIR / "Raw Data.zip"
SOURCE_SNAPSHOT = SOURCE_DIR / "来源快照.json"
SOURCE_CROSSREF = SOURCE_DIR / "论文元数据_Crossref.json"
SOURCE_FULLTEXT = SOURCE_DIR / "论文全文_PMC.xml"
SOURCE_LICENSE = SOURCE_DIR / "许可_CC_BY_4.0.txt"

DATASET_DOI = "10.15131/shef.data.21510876.v1"
PAPER_DOI = "10.3390/polym14235111"
LICENSE = "CC BY 4.0"
DATASET_URL = "https://figshare.shef.ac.uk/articles/dataset/PU_Rational_Design_Data/21510876"
API_URL = "https://api.figshare.com/v2/articles/21510876"
DOWNLOAD_URL = "https://ndownloader.figshare.com/files/38125788"
AUDIT_VERSION = "batch9-sheffield-pu-rational-design-v1"

FROZEN_FILES = {
    "Raw Data.zip": (
        90_452_519,
        "76a7c5bd1672b19b5130a3bd55aec1cea82f45c1118dd16acc97a90294e98318",
        "29f7b22b6aa292c60e6b8712c9108a22",
    ),
    "来源快照.json": (
        6_953,
        "d7bdf9ef319fee3c34831d7005e5e535d03483821c9be9eb81c484a2788cabbe",
        "43937b28171ecb14c74ac231fd58abd2",
    ),
    "论文元数据_Crossref.json": (
        14_530,
        "926a137eede543182311644c9a05b06c0b2b15d9737b2c481b0cf36c0f7fa8f2",
        "ff80fad368d7b3d98e7115eeb2f35fa8",
    ),
    "论文全文_PMC.xml": (
        134_839,
        "ea6f7929e7ec32715fc4e7818830ca19747a3a19877ce00e005b7ac4729e3def",
        "5cd606fe5b79d5f3674e6d9c47a4dcc6",
    ),
    "许可_CC_BY_4.0.txt": (
        18_657,
        "9ba9550ad48438d0836ddab3da480b3b69ffa0aac7b7878b5a0039e7ab429411",
        "2ab724713fdaf49e4523c4503bfd068d",
    ),
}

EXPECTED_ZIP_MEMBERS = 231
EXPECTED_ZIP_FILES = 214
EXPECTED_ZIP_DIRECTORIES = 17
EXPECTED_ZIP_UNCOMPRESSED_BYTES = 357_901_903
EXPECTED_FORMULATION_BATCHES = 40
EXPECTED_UNIQUE_FORMULATIONS = 39
EXPECTED_SCREEN_UNIQUE = 23
EXPECTED_FINAL_UNIQUE = 16
EXPECTED_KINETIC_CURVES = 40
EXPECTED_KINETIC_SUMMARIES = 39
EXPECTED_CELL_DISTRIBUTIONS = 47
EXPECTED_CURVES = 155
EXPECTED_SCALARS = 764

OUTPUT_NAMES = (
    "内容审计摘要.json",
    "配方审计清单.tsv",
    "曲线审计清单.tsv",
    "标量审计清单.tsv",
    "ZIP成员审计清单.tsv",
    "文件校验清单.tsv",
)

FORMULATION_COLUMNS = (
    "sample_id",
    "experiment",
    "sample_role",
    "formulation_id",
    "split_group",
    "is_unique_formulation_representative",
    "duplicate_of_sample_id",
    "voranol_1447_pphp",
    "voranol_3322_pphp",
    "water_pphp",
    "vorasurf_5906_pphp",
    "surfactant_2_pphp",
    "surfactant_2_publication_identity",
    "surfactant_2_supplement_header",
    "dabco_t_pphp",
    "cloisite_ne_116_pphp",
    "specflex_ne_112_pphp",
    "total_pphp",
    "isocyanate_index",
    "chemistry_resolution",
    "gold_layer",
    "gold_admission_status",
    "future_weight_ceiling",
    "source_location",
    "notes",
)

CURVE_COLUMNS = (
    "curve_id",
    "sample_id",
    "experiment",
    "formulation_id",
    "split_group",
    "curve_type",
    "replicate_index",
    "source_location",
    "point_count",
    "x_name",
    "x_unit",
    "y_names",
    "y_units",
    "x_min",
    "x_max",
    "x_values_json",
    "y_values_json",
    "raw_headers",
    "target_origin",
    "data_origin",
    "record_granularity",
    "gold_layer",
    "gold_admission_status",
    "future_weight_ceiling",
    "is_external_control",
    "notes",
)

SCALAR_COLUMNS = (
    "scalar_id",
    "sample_id",
    "experiment",
    "formulation_id",
    "split_group",
    "observable",
    "value",
    "unit",
    "replicate_index",
    "source_location",
    "target_origin",
    "data_origin",
    "record_granularity",
    "derivation",
    "gold_layer",
    "gold_admission_status",
    "future_weight_ceiling",
    "is_external_control",
    "chemistry_resolution",
    "notes",
)

MEMBER_COLUMNS = (
    "member_index",
    "member_name",
    "member_type",
    "experiment",
    "sample_id",
    "file_size",
    "compressed_size",
    "compression_ratio",
    "crc32",
    "zip_path_safe",
    "encrypted",
    "symlink",
)


class AuditBlocked(RuntimeError):
    """原件身份、压缩包安全、结构或冻结计数发生漂移。"""


def _digest(path: Path, algorithm: str) -> str:
    digest = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number(value: object, *, label: str) -> float:
    if value is None or isinstance(value, bool):
        raise AuditBlocked(f"{label} 不是数值：{value!r}")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise AuditBlocked(f"{label} 无法解析：{value!r}") from exc
    if not math.isfinite(result):
        raise AuditBlocked(f"{label} 不是有限数：{value!r}")
    return result


def _optional_number(value: object, *, label: str) -> float | None:
    if value is None or str(value).strip() in {"", "-"}:
        return None
    return _number(value, label=label)


def _close(left: float, right: float, *, tolerance: float = 1e-8) -> bool:
    return math.isclose(left, right, rel_tol=tolerance, abs_tol=tolerance)


def _sample_id(value: object) -> str:
    text = str(value).strip()
    match = re.search(r"(?i)F0?5[-_ ]?(\d{1,2})", text)
    if not match:
        if text.upper() == "RW":
            return "RW"
        raise AuditBlocked(f"无法识别样品编号：{value!r}")
    return f"F05-{int(match.group(1)):02d}"


def _experiment_from_member(name: str) -> int:
    match = re.search(r"/Experiment (\d)/", name)
    return int(match.group(1)) if match else 0


def verify_files() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for name, expected in FROZEN_FILES.items():
        path = SOURCE_DIR / name
        if not path.is_file():
            raise AuditBlocked(f"缺少冻结原件：{path}")
        actual = (path.stat().st_size, _digest(path, "sha256"), _digest(path, "md5"))
        if actual != expected:
            raise AuditBlocked(
                f"原件漂移：{name} bytes={actual[0]}, sha256={actual[1]}, md5={actual[2]}"
            )
        rows.append(
            {
                "file": name,
                "bytes": actual[0],
                "sha256": actual[1],
                "md5": actual[2],
                "verification": "matched_frozen_identity",
            }
        )

    snapshot = json.loads(SOURCE_SNAPSHOT.read_text(encoding="utf-8"))
    if (
        snapshot.get("id") != 21_510_876
        or snapshot.get("doi") != DATASET_DOI
        or snapshot.get("version") != 1
        or snapshot.get("status") != "public"
        or snapshot.get("download_disabled") is not False
        or snapshot.get("license", {}).get("name") != LICENSE
    ):
        raise AuditBlocked("Figshare 数据集身份、版本、许可或下载状态漂移")
    official_files = snapshot.get("files", [])
    if len(official_files) != 1:
        raise AuditBlocked("Figshare 官方文件数量漂移")
    official = official_files[0]
    if (
        official.get("id") != 38_125_788
        or official.get("name") != "Raw Data.zip"
        or official.get("size") != FROZEN_FILES["Raw Data.zip"][0]
        or official.get("supplied_md5") != FROZEN_FILES["Raw Data.zip"][2]
        or official.get("download_url") != DOWNLOAD_URL
    ):
        raise AuditBlocked("Figshare 官方文件身份漂移")

    crossref = json.loads(SOURCE_CROSSREF.read_text(encoding="utf-8"))["message"]
    if crossref.get("DOI", "").lower() != PAPER_DOI:
        raise AuditBlocked("Crossref 论文 DOI 漂移")
    if crossref.get("title") != ["Rational Design of a Polyurethane Foam"]:
        raise AuditBlocked("Crossref 论文标题漂移")

    fulltext = ElementTree.parse(SOURCE_FULLTEXT).getroot()
    article_dois = {
        "".join(node.itertext()).strip().lower()
        for node in fulltext.findall(".//article-id[@pub-id-type='doi']")
    }
    if PAPER_DOI not in article_dois:
        raise AuditBlocked("PMC 全文 DOI 漂移")
    license_text = SOURCE_LICENSE.read_text(encoding="utf-8")
    if "Attribution 4.0 International" not in license_text:
        raise AuditBlocked("CC BY 4.0 法律文本漂移")
    return rows


def _member_type(name: str, is_dir: bool) -> str:
    lower = name.casefold()
    if is_dir:
        return "directory"
    if lower.endswith("supplementary data.docx"):
        return "formulation_supplement"
    if lower.endswith("_corrected.csv"):
        return "foampi_kinetic_curve"
    if lower.endswith(("atrsumm.csv", "atrsummfin.csv")):
        return "foampi_kinetic_summary"
    if lower.endswith(".jpg.csv"):
        return "cell_area_distribution"
    if lower.endswith(("cellsize_r.csv", "cellsizefin.csv", "/cellsize.csv")):
        return "cell_size_summary"
    if lower.endswith("af_all.csv") or lower.endswith("airflow.xlsx"):
        return "airflow_summary"
    if lower.endswith("density.xlsx"):
        return "density_summary"
    if lower.endswith("whc.xlsx"):
        return "water_holding_summary"
    if lower.endswith("/cap.csv"):
        return "capillary_curve_table"
    if lower.endswith("/wdpt.csv"):
        return "wdpt_replicate_table"
    if lower.endswith(".avi"):
        return "wdpt_video"
    if lower.endswith(".tif"):
        return "cell_micrograph_tif"
    if lower.endswith(".jpg"):
        return "scale_image" if "scale" in lower or "ruler" in lower else "cell_micrograph_jpg"
    return "unclassified"


def audit_zip_members() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    seen: set[str] = set()
    seen_casefold: set[str] = set()
    total_uncompressed = 0
    with zipfile.ZipFile(SOURCE_ZIP) as archive:
        infos = archive.infolist()
        if len(infos) != EXPECTED_ZIP_MEMBERS:
            raise AuditBlocked(f"ZIP成员数漂移：{len(infos)}")
        for index, info in enumerate(infos, start=1):
            name = info.filename
            path = PurePosixPath(name)
            unsafe = (
                not name
                or "\\" in name
                or name.startswith("/")
                or bool(re.match(r"^[A-Za-z]:", name))
                or any(part in {"", ".", ".."} for part in path.parts)
            )
            folded = name.casefold()
            duplicate = name in seen or folded in seen_casefold
            seen.add(name)
            seen_casefold.add(folded)
            unix_mode = (info.external_attr >> 16) & 0xFFFF
            symlink = stat.S_ISLNK(unix_mode)
            encrypted = bool(info.flag_bits & 0x1)
            ratio = (
                float(info.file_size) / float(info.compress_size)
                if info.compress_size
                else (0.0 if info.file_size == 0 else math.inf)
            )
            if unsafe or duplicate or symlink or encrypted:
                raise AuditBlocked(
                    f"ZIP安全门禁失败：{name!r}, unsafe={unsafe}, duplicate={duplicate}, "
                    f"symlink={symlink}, encrypted={encrypted}"
                )
            if info.file_size > 10_000_000 or ratio > 200:
                raise AuditBlocked(f"ZIP成员异常膨胀：{name}, bytes={info.file_size}, ratio={ratio}")
            total_uncompressed += info.file_size
            sample = ""
            if not info.is_dir():
                try:
                    sample = _sample_id(Path(name).name)
                except AuditBlocked:
                    numbered = re.search(r"\b007(2[5-9]|30)\b", Path(name).name)
                    if numbered:
                        sample = f"F05-{int(numbered.group(1)) - 15:02d}"
            rows.append(
                {
                    "member_index": index,
                    "member_name": name,
                    "member_type": _member_type(name, info.is_dir()),
                    "experiment": _experiment_from_member(name),
                    "sample_id": sample,
                    "file_size": info.file_size,
                    "compressed_size": info.compress_size,
                    "compression_ratio": round(ratio, 8),
                    "crc32": f"{info.CRC:08x}",
                    "zip_path_safe": True,
                    "encrypted": False,
                    "symlink": False,
                }
            )
    file_count = sum(row["member_type"] != "directory" for row in rows)
    directory_count = len(rows) - file_count
    if (file_count, directory_count, total_uncompressed) != (
        EXPECTED_ZIP_FILES,
        EXPECTED_ZIP_DIRECTORIES,
        EXPECTED_ZIP_UNCOMPRESSED_BYTES,
    ):
        raise AuditBlocked(
            "ZIP文件/目录/未压缩字节数漂移："
            f"{file_count}, {directory_count}, {total_uncompressed}"
        )
    if any(row["member_type"] == "unclassified" for row in rows):
        unknown = [row["member_name"] for row in rows if row["member_type"] == "unclassified"]
        raise AuditBlocked(f"出现未分类ZIP成员：{unknown}")
    return rows


def _read_member(archive: zipfile.ZipFile, name: str, *, limit: int = 10_000_000) -> bytes:
    info = archive.getinfo(name)
    if info.file_size > limit:
        raise AuditBlocked(f"成员超过内存解析上限：{name}")
    payload = archive.read(info)
    if len(payload) != info.file_size:
        raise AuditBlocked(f"成员读取不完整：{name}")
    return payload


def _nested_document_xml(payload: bytes) -> ElementTree.Element:
    with zipfile.ZipFile(io.BytesIO(payload)) as nested:
        infos = nested.infolist()
        if len(infos) > 100 or sum(info.file_size for info in infos) > 5_000_000:
            raise AuditBlocked("嵌套DOCX结构异常")
        for info in infos:
            path = PurePosixPath(info.filename)
            if info.filename.startswith("/") or ".." in path.parts or "\\" in info.filename:
                raise AuditBlocked("嵌套DOCX存在不安全路径")
        return ElementTree.fromstring(nested.read("word/document.xml"))


def parse_formulations() -> tuple[list[dict[str, object]], dict[str, dict[str, object]]]:
    member = "Raw Data/Supplementary Data.docx"
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(SOURCE_ZIP) as archive:
        root = _nested_document_xml(_read_member(archive, member))
    tables = root.findall(".//w:tbl", namespace)
    if len(tables) != 1:
        raise AuditBlocked(f"补充材料表格数漂移：{len(tables)}")
    matrix: list[list[str]] = []
    for table_row in tables[0].findall("./w:tr", namespace):
        cells = []
        for cell in table_row.findall("./w:tc", namespace):
            text = "".join(
                node.text or "" for node in cell.findall(".//w:t", namespace)
            )
            cells.append(" ".join(text.split()))
        matrix.append(cells)
    expected_header = [
        "Exp", "Sample", "Voranol 1447", "Voranol 3322", "Water",
        "Vorasurf 5906", "Vorasurf 5959", "Dabco T", "Cloisite",
        "SpecFlex NE 112", "Total",
    ]
    if len(matrix) != 42 or matrix[1] != expected_header:
        raise AuditBlocked("补充表S1结构或表头漂移")

    numeric_keys = (
        "voranol_1447_pphp", "voranol_3322_pphp", "water_pphp",
        "vorasurf_5906_pphp", "surfactant_2_pphp", "dabco_t_pphp",
        "cloisite_ne_116_pphp", "specflex_ne_112_pphp", "total_pphp",
    )
    rows: list[dict[str, object]] = []
    current_experiment = 0
    first_by_signature: dict[str, str] = {}
    for source_row, cells in enumerate(matrix[2:], start=3):
        if len(cells) != 11:
            raise AuditBlocked(f"补充表S1第{source_row}行列数漂移")
        if cells[0]:
            current_experiment = int(cells[0])
        sample = _sample_id(cells[1])
        values = [_number(value, label=f"S1第{source_row}行") for value in cells[2:]]
        if not _close(sum(values[:-1]), values[-1], tolerance=1e-4):
            raise AuditBlocked(f"S1第{source_row}行Total不能复算")
        signature = "|".join(format(value, ".12g") for value in values[:-1])
        digest = hashlib.sha256(signature.encode("ascii")).hexdigest()[:16]
        formulation_id = f"sheffield_pu_{digest}"
        duplicate_of = first_by_signature.get(signature, "")
        first_by_signature.setdefault(signature, sample)
        row = {
            "sample_id": sample,
            "experiment": current_experiment,
            "sample_role": "doe_final" if current_experiment == 4 else (
                "screening_control_repeat" if duplicate_of else "screening"
            ),
            "formulation_id": formulation_id,
            "split_group": f"sheffield:{formulation_id}",
            "is_unique_formulation_representative": not bool(duplicate_of),
            "duplicate_of_sample_id": duplicate_of,
            **dict(zip(numeric_keys, values, strict=True)),
            "surfactant_2_publication_identity": "Tegostab 8476",
            "surfactant_2_supplement_header": "Vorasurf 5959",
            "isocyanate_index": 1.15,
            "chemistry_resolution": "commercial_product_and_pphp_no_smiles",
            "gold_layer": "Gold-E",
            "gold_admission_status": "admitted_reference",
            "future_weight_ceiling": 0.45 if current_experiment == 4 else 0.35,
            "source_location": f"Raw Data.zip!/{member}#Table S1 row {source_row}",
            "notes": (
                "论文正文将第二表面活性剂标为Tegostab 8476；补充表表头写Vorasurf 5959，"
                "保留该出版物内部命名冲突。商业组分无SMILES，结构任务权重为0。"
            ),
        }
        rows.append(row)

    by_sample = {str(row["sample_id"]): row for row in rows}
    unique_ids = {str(row["formulation_id"]) for row in rows}
    screen_ids = {
        str(row["formulation_id"]) for row in rows if int(row["experiment"]) < 4
    }
    final_ids = {
        str(row["formulation_id"]) for row in rows if int(row["experiment"]) == 4
    }
    duplicates = [row for row in rows if row["duplicate_of_sample_id"]]
    if (
        len(rows) != EXPECTED_FORMULATION_BATCHES
        or len(by_sample) != EXPECTED_FORMULATION_BATCHES
        or len(unique_ids) != EXPECTED_UNIQUE_FORMULATIONS
        or len(screen_ids) != EXPECTED_SCREEN_UNIQUE
        or len(final_ids) != EXPECTED_FINAL_UNIQUE
        or Counter(int(row["experiment"]) for row in rows) != {1: 8, 2: 7, 3: 9, 4: 16}
        or [(row["sample_id"], row["duplicate_of_sample_id"]) for row in duplicates]
        != [("F05-34", "F05-13")]
    ):
        raise AuditBlocked("配方批次、唯一配方或对照重复映射漂移")
    return rows, by_sample


def _csv_rows(archive: zipfile.ZipFile, member: str) -> list[list[str]]:
    text = _read_member(archive, member).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise AuditBlocked(f"空CSV：{member}")
    return rows


def _entity(sample: str, formulations: dict[str, dict[str, object]]) -> dict[str, object]:
    if sample == "RW":
        return {
            "sample_id": "RW",
            "experiment": 4,
            "formulation_id": "external_control_rockwool",
            "split_group": "external_control:rockwool",
            "is_external_control": True,
            "chemistry_resolution": "non_pu_reference_material",
        }
    formulation = formulations.get(sample)
    if formulation is None:
        raise AuditBlocked(f"样品没有配方映射：{sample}")
    return {
        "sample_id": sample,
        "experiment": int(formulation["experiment"]),
        "formulation_id": formulation["formulation_id"],
        "split_group": formulation["split_group"],
        "is_external_control": False,
        "chemistry_resolution": formulation["chemistry_resolution"],
    }


def _curve_common(
    sample: str,
    formulations: dict[str, dict[str, object]],
    *,
    admission: str = "admitted_reference",
    ceiling: float = 0.40,
) -> dict[str, object]:
    entity = _entity(sample, formulations)
    if entity["is_external_control"]:
        admission, ceiling = "evidence_only", 0.05
    return {
        **entity,
        "target_origin": "experimental",
        "gold_layer": "Gold-E",
        "gold_admission_status": admission,
        "future_weight_ceiling": ceiling,
    }


def _add_scalar(
    target: list[dict[str, object]],
    formulations: dict[str, dict[str, object]],
    *,
    sample: str,
    observable: str,
    value: float,
    unit: str,
    source_location: str,
    replicate_index: int | str = "",
    granularity: str,
    derivation: str,
    admission: str = "admitted_reference",
    ceiling: float = 0.40,
    notes: str = "",
) -> None:
    common = _curve_common(sample, formulations, admission=admission, ceiling=ceiling)
    scalar_id = hashlib.sha256(
        f"{sample}|{observable}|{source_location}|{replicate_index}|{granularity}".encode("utf-8")
    ).hexdigest()[:20]
    target.append(
        {
            "scalar_id": f"sheffield_{scalar_id}",
            **common,
            "observable": observable,
            "value": value,
            "unit": unit,
            "replicate_index": replicate_index,
            "source_location": source_location,
            "data_origin": "experimental_source_summary" if derivation == "source_value" else "experimental_derived",
            "record_granularity": granularity,
            "derivation": derivation,
            "notes": notes,
        }
    )


def parse_kinetics(
    formulations: dict[str, dict[str, object]],
) -> tuple[list[dict[str, object]], list[dict[str, object]], dict[str, object]]:
    curves: list[dict[str, object]] = []
    scalars: list[dict[str, object]] = []
    point_total = 0
    with zipfile.ZipFile(SOURCE_ZIP) as archive:
        members = sorted(
            name for name in archive.namelist() if name.casefold().endswith("_corrected.csv")
        )
        if len(members) != EXPECTED_KINETIC_CURVES:
            raise AuditBlocked(f"FoamPi动力学曲线数漂移：{len(members)}")
        curve_samples: list[str] = []
        for member in members:
            rows = _csv_rows(archive, member)
            header, data = rows[0], rows[1:]
            expected_short = ["t", "Tj", "Traw", "Tcorr", "Hraw", "Hcorr", "dTdt", "dHdt"]
            expected_long = ["t", "Tj", "Traw", "Tcorr", "dTdt", "Hraw", "Hcorr", "dHdt", "mraw", "mcorr", "dmdt"]
            if header not in (expected_short, expected_long):
                raise AuditBlocked(f"动力学表头漂移：{member} {header}")
            numeric: list[list[float]] = []
            for line_no, row in enumerate(data, start=2):
                if len(row) != len(header):
                    raise AuditBlocked(f"动力学列数漂移：{member}:{line_no}")
                numeric.append([_number(value, label=f"{member}:{line_no}") for value in row])
            times = [row[0] for row in numeric]
            if len(times) < 100 or any(right <= left for left, right in zip(times, times[1:])):
                raise AuditBlocked(f"动力学时间轴异常：{member}")
            sample = _sample_id(Path(member).name)
            curve_samples.append(sample)
            common = _curve_common(sample, formulations, ceiling=0.40)
            curve_id = f"sheffield_kinetic_{sample.lower()}"
            curves.append(
                {
                    "curve_id": curve_id,
                    **common,
                    "curve_type": "foampi_adiabatic_temperature_rise_and_foam_rise",
                    "replicate_index": 1,
                    "source_location": f"Raw Data.zip!/{member}",
                    "point_count": len(numeric),
                    "x_name": "time",
                    "x_unit": "s",
                    "y_names": ";".join(header[1:]),
                    "y_units": "degC;degC;degC;mm;mm;degC_per_s;mm_per_s" if len(header) == 8 else "degC;degC;degC;degC_per_s;mm;mm;mm_per_s;g;g;g_per_s",
                    "x_min": times[0],
                    "x_max": times[-1],
                    "x_values_json": "",
                    "y_values_json": "",
                    "raw_headers": ";".join(header),
                    "data_origin": "experimental_raw_curve",
                    "record_granularity": "reaction_time_series",
                    "notes": "ATR=adiabatic temperature rise，不是ATR-FTIR；曲线文件不等于独立材料。",
                }
            )
            point_total += len(numeric)
        if Counter(curve_samples)["F05-34"] != 1 or len(set(curve_samples)) != 40:
            raise AuditBlocked("动力学曲线样品映射漂移")

        summaries = (
            "Raw Data/Experiment 1/ATRsumm.csv",
            "Raw Data/Experiment 2/Kinetic/ATRsumm.csv",
            "Raw Data/Experiment 3/Kinetic/ATRsummFin.csv",
            "Raw Data/Experiment 4/Kinetics/ATRsumm.csv",
        )
        summary_samples: list[str] = []
        max_sag_error = 0.0
        max_normalized_error = 0.0
        for member in summaries:
            rows = _csv_rows(archive, member)
            header = rows[0]
            for source_row, values in enumerate(rows[1:], start=2):
                record = dict(zip(header, values, strict=True))
                sample = _sample_id(record["Sample"])
                summary_samples.append(sample)
                source = f"Raw Data.zip!/{member}#row {source_row}"
                def get(*keys: str) -> float | None:
                    for key in keys:
                        if key in record:
                            return _optional_number(record[key], label=f"{source}:{key}")
                    return None
                raw_dt = get("DelTraw")
                max_dt = get("DelTmax")
                t_tmax = get("tTmax")
                hmax = get("Hmax")
                t_hmax = get("tHmax")
                hfin = get("Hfin")
                sag_h = get("SagH")
                conversion = get("Conversion", "Conv")
                mass = get("Mass")
                required = (raw_dt, t_tmax, hmax, t_hmax, hfin, sag_h)
                if any(value is None for value in required):
                    raise AuditBlocked(f"动力学摘要核心字段缺失：{source}")
                core = (
                    ("temperature_rise_raw", raw_dt, "degC"),
                    ("time_to_max_temperature", t_tmax, "s"),
                    ("maximum_foam_height", hmax, "mm"),
                    ("time_to_max_foam_height", t_hmax, "s"),
                    ("final_foam_height", hfin, "mm"),
                    ("foam_sag_height", sag_h, "mm"),
                )
                for observable, value, unit in core:
                    assert value is not None
                    _add_scalar(
                        scalars, formulations, sample=sample, observable=observable,
                        value=value, unit=unit, source_location=source,
                        granularity="reaction_batch_summary", derivation="source_value",
                        ceiling=0.40, notes="FoamPi绝热温升/泡高动力学作者汇总。",
                    )
                if max_dt is not None:
                    _add_scalar(
                        scalars, formulations, sample=sample,
                        observable="temperature_rise_adiabatic_corrected_max",
                        value=max_dt, unit="degC", source_location=source,
                        granularity="reaction_batch_summary", derivation="source_value",
                        ceiling=0.40,
                    )
                if conversion is not None:
                    _add_scalar(
                        scalars, formulations, sample=sample,
                        observable="isocyanate_conversion", value=conversion, unit="1",
                        source_location=source, granularity="reaction_batch_summary",
                        derivation="source_value", ceiling=0.45,
                    )
                if mass is not None:
                    _add_scalar(
                        scalars, formulations, sample=sample, observable="reaction_mass",
                        value=mass, unit="g", source_location=source,
                        granularity="reaction_batch_summary", derivation="source_value",
                        ceiling=0.35,
                    )
                    normalized = float(hmax) / mass
                    source_normalized = get("Maxh_m", "Hmax_m", "Hdm")
                    if source_normalized is not None:
                        max_normalized_error = max(max_normalized_error, abs(normalized - source_normalized))
                    _add_scalar(
                        scalars, formulations, sample=sample,
                        observable="maximum_foam_height_normalized", value=normalized,
                        unit="mm/g", source_location=source,
                        granularity="reaction_batch_summary", derivation="Hmax/Mass",
                        ceiling=0.35,
                    )
                sag_percent = 100.0 * (float(hmax) - float(hfin)) / float(hmax)
                source_sag = get("Sag_m", "Sag_mp", "Sagper")
                if source_sag is not None:
                    max_sag_error = max(max_sag_error, abs(sag_percent - source_sag))
                sag_value = source_sag if source_sag is not None else sag_percent
                sag_derivation = "source_value" if source_sag is not None else "100*(Hmax-Hfin)/Hmax"
                sag_note = ""
                if source_sag is not None and abs(sag_percent - source_sag) > 1e-6:
                    sag_note = (
                        f"作者Sag%={source_sag:.9g}，按同一摘要Hmax/Hfin复算={sag_percent:.9g}；"
                        "保留作者值并在审计摘要记录差异，不擅自覆盖。"
                    )
                _add_scalar(
                    scalars, formulations, sample=sample, observable="foam_sag_percent",
                    value=sag_value, unit="%", source_location=source,
                    granularity="reaction_batch_summary", derivation=sag_derivation,
                    ceiling=0.40, notes=sag_note,
                )

    if (
        len(summary_samples) != EXPECTED_KINETIC_SUMMARIES
        or len(set(summary_samples)) != EXPECTED_KINETIC_SUMMARIES
        or "F05-34" in summary_samples
        or max_normalized_error > 1e-6
        or len(scalars) != 389
    ):
        raise AuditBlocked(
            "动力学摘要计数或复算漂移："
            f"rows={len(summary_samples)}, scalars={len(scalars)}, "
            f"sag_error={max_sag_error}, normalized_error={max_normalized_error}"
        )
    return curves, scalars, {
        "curve_points": point_total,
        "max_abs_sag_percent_recalculation_error": max_sag_error,
        "max_abs_normalized_height_recalculation_error": max_normalized_error,
    }


def parse_cell_size(
    formulations: dict[str, dict[str, object]],
) -> tuple[list[dict[str, object]], list[dict[str, object]], dict[str, object]]:
    curves: list[dict[str, object]] = []
    scalars: list[dict[str, object]] = []
    cell_points = 0
    numbered_images = {
        "00725.jpg.csv": "F05-10",
        "00726.jpg.csv": "F05-11",
        "00727.jpg.csv": "F05-12",
        "00728.jpg.csv": "F05-13",
        "00729.jpg.csv": "F05-14",
        "00730.jpg.csv": "F05-15",
    }
    with zipfile.ZipFile(SOURCE_ZIP) as archive:
        members = sorted(
            name for name in archive.namelist() if name.casefold().endswith(".jpg.csv")
        )
        if len(members) != EXPECTED_CELL_DISTRIBUTIONS:
            raise AuditBlocked(f"泡孔面积分布文件数漂移：{len(members)}")
        distribution_counts: Counter[str] = Counter()
        for member in members:
            rows = _csv_rows(archive, member)
            header, data = rows[0], rows[1:]
            if len(header) not in {5, 9} or header[1] != "Area":
                raise AuditBlocked(f"泡孔面积表头漂移：{member} {header}")
            areas: list[float] = []
            indices: list[int] = []
            for source_row, row in enumerate(data, start=2):
                if len(row) != len(header):
                    raise AuditBlocked(f"泡孔面积列数漂移：{member}:{source_row}")
                index = int(_number(row[0], label=f"{member}:{source_row}:cell"))
                area = _number(row[1], label=f"{member}:{source_row}:area")
                if area <= 0:
                    raise AuditBlocked(f"泡孔面积非正：{member}:{source_row}")
                indices.append(index)
                areas.append(area)
            if indices != list(range(1, len(indices) + 1)) or len(areas) < 50:
                raise AuditBlocked(f"泡孔编号或计数异常：{member}")
            filename = Path(member).name
            sample = numbered_images.get(filename) or _sample_id(filename)
            experiment = _experiment_from_member(member)
            replicate = 1
            replicate_match = re.search(r"-(\d)\.jpg\.csv$", filename, re.IGNORECASE)
            if experiment == 4 and replicate_match:
                replicate = int(replicate_match.group(1))
            common = _curve_common(sample, formulations, ceiling=0.35)
            curves.append(
                {
                    "curve_id": f"sheffield_cell_distribution_{sample.lower()}_{experiment}_{replicate}",
                    **common,
                    "curve_type": "individual_cell_area_distribution",
                    "replicate_index": replicate,
                    "source_location": f"Raw Data.zip!/{member}",
                    "point_count": len(areas),
                    "x_name": "cell_index",
                    "x_unit": "1",
                    "y_names": "cell_area",
                    "y_units": "mm^2",
                    "x_min": 1,
                    "x_max": len(areas),
                    "x_values_json": "",
                    "y_values_json": "",
                    "raw_headers": ";".join(header),
                    "data_origin": "experimental_raw_distribution",
                    "record_granularity": "microscopy_image",
                    "notes": "ImageJ逐泡孔面积；原图/TIF留在同一官方ZIP中。",
                }
            )
            distribution_counts[sample] += 1
            cell_points += len(areas)

        member = "Raw Data/Experiment 2/Cell Size/CellSize_R.csv"
        rows = _csv_rows(archive, member)
        header = rows[0]
        if header != ["Filename", "Sample", "Surfactant", "AvgArea", "AVG Diameter", "SE Diameter"]:
            raise AuditBlocked("实验2泡孔汇总表头漂移")
        exp2_count = 0
        for source_row, values in enumerate(rows[1:], start=2):
            record = dict(zip(header, values, strict=True))
            mean_mm = _optional_number(record["AVG Diameter"], label=f"{member}:{source_row}")
            se_mm = _optional_number(record["SE Diameter"], label=f"{member}:{source_row}")
            if mean_mm is None and se_mm is None:
                continue
            if mean_mm is None or se_mm is None:
                raise AuditBlocked("实验2泡孔均值/标准误缺失不对称")
            sample = _sample_id(record["Sample"])
            source = f"Raw Data.zip!/{member}#row {source_row}"
            for observable, value in (
                ("mean_cell_diameter", mean_mm * 1000.0),
                ("cell_diameter_standard_error", se_mm * 1000.0),
            ):
                _add_scalar(
                    scalars, formulations, sample=sample, observable=observable,
                    value=value, unit="um", source_location=source,
                    granularity="formulation_microscopy_summary", derivation="source_value_converted_mm_to_um",
                    ceiling=0.40,
                )
            exp2_count += 1

        member = "Raw Data/Experiment 3/Cell Size/CellSizeFin.csv"
        rows = _csv_rows(archive, member)
        if rows[0][:3] != ["Sample", "Surf", "AvgArea"] or len(rows[0]) != 7:
            raise AuditBlocked("实验3泡孔汇总表头漂移")
        exp3_count = 0
        for source_row, values in enumerate(rows[1:], start=2):
            if len(values) != 7:
                raise AuditBlocked("实验3泡孔汇总列数漂移")
            sample = _sample_id(values[0])
            mean_um = _number(values[4], label=f"{member}:{source_row}:mean_um")
            se_um = _number(values[6], label=f"{member}:{source_row}:se_um")
            if not _close(_number(values[3], label="mean_mm") * 1000.0, mean_um, tolerance=1e-6):
                raise AuditBlocked("实验3泡孔直径mm/um换算漂移")
            if not _close(_number(values[5], label="se_mm") * 1000.0, se_um, tolerance=1e-6):
                raise AuditBlocked("实验3泡孔标准误mm/um换算漂移")
            source = f"Raw Data.zip!/{member}#row {source_row}"
            for observable, value in (
                ("mean_cell_diameter", mean_um),
                ("cell_diameter_standard_error", se_um),
            ):
                _add_scalar(
                    scalars, formulations, sample=sample, observable=observable,
                    value=value, unit="um", source_location=source,
                    granularity="formulation_microscopy_summary", derivation="source_value",
                    ceiling=0.40,
                )
            exp3_count += 1

        member = "Raw Data/Experiment 4/Cell Size/CellSize.csv"
        rows = _csv_rows(archive, member)
        if rows[0][:4] != ["Sample", "AvgArea", "AVG Diameter", "SE Diameter"]:
            raise AuditBlocked("实验4泡孔汇总表头漂移")
        exp4_by_sample: dict[str, list[tuple[float, float]]] = defaultdict(list)
        for source_row, values in enumerate(rows[1:], start=2):
            if len(values) != 10:
                raise AuditBlocked(f"实验4泡孔汇总列数漂移：{source_row}")
            sample = _sample_id(values[0])
            replicate_match = re.search(r"-(\d)\.jpg\.csv$", values[0], re.IGNORECASE)
            if not replicate_match:
                raise AuditBlocked(f"实验4泡孔重复编号缺失：{values[0]}")
            replicate = int(replicate_match.group(1))
            mean_um = _number(values[2], label=f"{member}:{source_row}:mean") * 1000.0
            se_um = _number(values[3], label=f"{member}:{source_row}:se") * 1000.0
            source = f"Raw Data.zip!/{member}#row {source_row}"
            for observable, value in (
                ("mean_cell_diameter", mean_um),
                ("cell_diameter_standard_error", se_um),
            ):
                _add_scalar(
                    scalars, formulations, sample=sample, observable=observable,
                    value=value, unit="um", source_location=source,
                    replicate_index=replicate, granularity="microscopy_image",
                    derivation="source_value_converted_mm_to_um", ceiling=0.35,
                )
            exp4_by_sample[sample].append((mean_um, se_um))
        if len(exp4_by_sample) != 16 or any(len(values) != 2 for values in exp4_by_sample.values()):
            raise AuditBlocked("实验4泡孔重复或配方数漂移")
        for sample, values in sorted(exp4_by_sample.items()):
            mean_um = sum(value[0] for value in values) / 2.0
            source = f"Raw Data.zip!/{member}#two-image mean for {sample}"
            _add_scalar(
                scalars, formulations, sample=sample, observable="mean_cell_diameter",
                value=mean_um, unit="um", source_location=source,
                granularity="formulation_two_image_mean", derivation="mean(two image AVG Diameter)",
                ceiling=0.45, notes="最终DoE配方的两张显微图均值；两图同split_group。",
            )

    if (
        exp2_count != 6
        or exp3_count != 8
        or len(exp4_by_sample) != 16
        or len(scalars) != 108
        or distribution_counts["F05-34"] != 1
        or sum(distribution_counts.values()) != EXPECTED_CELL_DISTRIBUTIONS
    ):
        raise AuditBlocked("泡孔分布或标量计数漂移")
    return curves, scalars, {
        "individual_cell_area_points": cell_points,
        "distribution_files": len(curves),
        "experiment_2_formulation_summaries": exp2_count,
        "experiment_3_formulation_summaries": exp3_count,
        "experiment_4_microscopy_images": sum(len(values) for values in exp4_by_sample.values()),
    }


def _xlsx_pair(archive: zipfile.ZipFile, member: str):
    payload = _read_member(archive, member, limit=2_000_000)
    formulas = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    values = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if formulas.sheetnames != ["Sheet1"] or values.sheetnames != ["Sheet1"]:
        formulas.close()
        values.close()
        raise AuditBlocked(f"工作簿工作表漂移：{member}")
    return formulas, values


def _property_admission(sample: str) -> tuple[str, float]:
    return ("evidence_only", 0.05) if sample == "RW" else ("admitted_reference", 0.45)


def parse_airflow_density_whc(
    formulations: dict[str, dict[str, object]],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    scalars: list[dict[str, object]] = []
    max_formula_error = 0.0
    with zipfile.ZipFile(SOURCE_ZIP) as archive:
        member = "Raw Data/Experiment 3/Airflow/AF_All.csv"
        rows = _csv_rows(archive, member)
        if rows[0] != ["Sample", "SurfCo", "P", "Q", "AdjQ", "OpenCell"] or len(rows) != 10:
            raise AuditBlocked("实验3气流表结构漂移")
        for source_row, values in enumerate(rows[1:], start=2):
            sample = _sample_id(values[0])
            pressure = _number(values[2], label="airflow pressure")
            measured = _number(values[3], label="airflow measured")
            adjusted = _number(values[4], label="airflow adjusted")
            calibrated_raw = _number(values[5], label="open cell raw")
            expected_adjusted = 125.0 / pressure * measured
            expected_raw = adjusted * 0.0064 + 0.1194
            max_formula_error = max(
                max_formula_error, abs(adjusted - expected_adjusted), abs(calibrated_raw - expected_raw)
            )
            source = f"Raw Data.zip!/{member}#row {source_row}"
            admission, ceiling = _property_admission(sample)
            _add_scalar(
                scalars, formulations, sample=sample, observable="airflow_at_125_pa",
                value=adjusted, unit="L/min", source_location=source,
                granularity="foam_specimen", derivation="125/P*Q",
                admission=admission, ceiling=ceiling,
                notes=f"原测压力={pressure:g} Pa；校准开孔估计={calibrated_raw:.9g}。",
            )
            _add_scalar(
                scalars, formulations, sample=sample, observable="effective_open_cell_fraction",
                value=min(1.0, calibrated_raw), unit="1", source_location=source,
                granularity="foam_specimen", derivation="min(1,0.0064*airflow_125Pa+0.1194)",
                admission=admission, ceiling=ceiling,
                notes="校准式超过1时按原工作簿语义截断为1；未截断值写入气流记录注释。",
            )

        member = "Raw Data/Experiment 4/Airflow/Airflow.xlsx"
        formulas, values = _xlsx_pair(archive, member)
        try:
            formula_sheet = formulas["Sheet1"]
            value_sheet = values["Sheet1"]
            if tuple(value_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0][:6] != (
                "Sample", "P", "Q", "PredQ", "OpenCell", "OpenCellA"
            ):
                raise AuditBlocked("实验4气流工作簿表头漂移")
            for excel_row in range(2, 19):
                sample = _sample_id(value_sheet.cell(excel_row, 1).value)
                pressure = _number(value_sheet.cell(excel_row, 2).value, label="airflow P")
                measured = _number(value_sheet.cell(excel_row, 3).value, label="airflow Q")
                adjusted = 125.0 / pressure * measured
                calibrated_raw = adjusted * 0.0064 + 0.1194
                capped = min(1.0, calibrated_raw)
                cached = (
                    _number(value_sheet.cell(excel_row, 4).value, label="PredQ"),
                    _number(value_sheet.cell(excel_row, 5).value, label="OpenCell"),
                    _number(value_sheet.cell(excel_row, 6).value, label="OpenCellA"),
                )
                expected = (adjusted, calibrated_raw, capped)
                max_formula_error = max(max_formula_error, *(abs(a - b) for a, b in zip(cached, expected)))
                if not all(isinstance(formula_sheet.cell(excel_row, col).value, str) for col in (4, 5, 6)):
                    raise AuditBlocked("实验4气流公式单元格漂移")
                source = f"Raw Data.zip!/{member}#Sheet1 row {excel_row}"
                admission, ceiling = _property_admission(sample)
                for observable, value, unit, derivation in (
                    ("airflow_at_125_pa", adjusted, "L/min", "125/P*Q"),
                    ("effective_open_cell_fraction", capped, "1", "min(1,0.0064*airflow_125Pa+0.1194)"),
                ):
                    _add_scalar(
                        scalars, formulations, sample=sample, observable=observable,
                        value=value, unit=unit, source_location=source,
                        granularity="foam_specimen", derivation=derivation,
                        admission=admission, ceiling=ceiling,
                        notes=f"原测压力={pressure:g} Pa；校准未截断值={calibrated_raw:.9g}。",
                    )
        finally:
            formulas.close()
            values.close()

        member = "Raw Data/Experiment 4/Density/Density.xlsx"
        formulas, values = _xlsx_pair(archive, member)
        try:
            fs, vs = formulas["Sheet1"], values["Sheet1"]
            if tuple(vs.iter_rows(min_row=1, max_row=1, values_only=True))[0] != (
                "Sample", "L", "B", "H", "M", "rou"
            ):
                raise AuditBlocked("密度工作簿表头漂移")
            for excel_row in range(2, 19):
                sample = _sample_id(vs.cell(excel_row, 1).value)
                length, breadth, height, mass = (
                    _number(vs.cell(excel_row, col).value, label="density input") for col in range(2, 6)
                )
                density = 1_000_000.0 * mass / (length * breadth * height)
                cached = _number(vs.cell(excel_row, 6).value, label="density cached")
                max_formula_error = max(max_formula_error, abs(density - cached))
                if not isinstance(fs.cell(excel_row, 6).value, str):
                    raise AuditBlocked("密度公式单元格漂移")
                source = f"Raw Data.zip!/{member}#Sheet1 row {excel_row}"
                admission, ceiling = _property_admission(sample)
                _add_scalar(
                    scalars, formulations, sample=sample, observable="foam_density",
                    value=density, unit="kg/m^3", source_location=source,
                    granularity="foam_specimen", derivation="1e6*M/(L*B*H)",
                    admission=admission, ceiling=ceiling,
                    notes="ASTM D3574-11 Test A；L/B/H为mm，M为g。",
                )
        finally:
            formulas.close()
            values.close()

        member = "Raw Data/Experiment 4/WHC/WHC.xlsx"
        formulas, values = _xlsx_pair(archive, member)
        try:
            fs, vs = formulas["Sheet1"], values["Sheet1"]
            expected = ("Sample", "L", "B", "H", "M", "rou", "Msat", "Mvvert", "WHC", "WHCv", "Wloss")
            if tuple(vs.iter_rows(min_row=1, max_row=1, values_only=True))[0] != expected:
                raise AuditBlocked("WHC工作簿表头漂移")
            for excel_row in range(2, 19):
                sample = _sample_id(vs.cell(excel_row, 1).value)
                length, breadth, height, dry, saturated, drained = (
                    _number(vs.cell(excel_row, col).value, label="WHC input")
                    for col in (2, 3, 4, 5, 7, 8)
                )
                volume = length * breadth * height
                whc_sat = 1_000_000.0 * (saturated - dry) / volume
                whc_drained = 1_000_000.0 * (drained - dry) / volume
                water_loss = whc_sat - whc_drained
                cached = tuple(
                    _number(vs.cell(excel_row, col).value, label="WHC cached") for col in (9, 10, 11)
                )
                max_formula_error = max(
                    max_formula_error, *(abs(a - b) for a, b in zip(cached, (whc_sat, whc_drained, water_loss)))
                )
                if not all(isinstance(fs.cell(excel_row, col).value, str) for col in (9, 10, 11)):
                    raise AuditBlocked("WHC公式单元格漂移")
                source = f"Raw Data.zip!/{member}#Sheet1 row {excel_row}"
                admission, ceiling = _property_admission(sample)
                for observable, value, derivation in (
                    ("water_holding_capacity_saturated", whc_sat, "1e6*(Msat-Mdry)/volume"),
                    ("water_holding_capacity_after_15min_drain", whc_drained, "1e6*(Mdrained-Mdry)/volume"),
                    ("water_loss_during_15min_drain", water_loss, "WHC_saturated-WHC_drained"),
                ):
                    _add_scalar(
                        scalars, formulations, sample=sample, observable=observable,
                        value=value, unit="g/dm^3", source_location=source,
                        granularity="foam_specimen", derivation=derivation,
                        admission=admission, ceiling=ceiling,
                        notes="25×50×50 mm量级试样；浸水24 h，随后自由排水15 min。",
                    )
        finally:
            formulas.close()
            values.close()

    if len(scalars) != 120 or max_formula_error > 1e-6:
        raise AuditBlocked(
            f"气流/密度/WHC标量数或公式复算漂移：{len(scalars)}, {max_formula_error}"
        )
    return scalars, {"max_abs_workbook_or_csv_formula_error": max_formula_error}


def _fit_capillary(times: list[float], values: list[float]) -> tuple[float, float, float]:
    """拟合 y=a*(1-exp(-k*t))；只做一维有界搜索，不依赖SciPy。"""

    def evaluate(log_k: float) -> tuple[float, float]:
        k = math.exp(log_k)
        basis = [1.0 - math.exp(-k * time) for time in times]
        denominator = sum(value * value for value in basis)
        if denominator <= 0:
            return math.inf, 0.0
        amplitude = sum(y * g for y, g in zip(values, basis)) / denominator
        if amplitude <= 0:
            return math.inf, amplitude
        residual = sum((y - amplitude * g) ** 2 for y, g in zip(values, basis))
        return residual, amplitude

    left, right = math.log(1e-6), math.log(1.0)
    phi = (1.0 + math.sqrt(5.0)) / 2.0
    c = right - (right - left) / phi
    d = left + (right - left) / phi
    fc, _ = evaluate(c)
    fd, _ = evaluate(d)
    for _ in range(120):
        if fc < fd:
            right, d, fd = d, c, fc
            c = right - (right - left) / phi
            fc, _ = evaluate(c)
        else:
            left, c, fc = c, d, fd
            d = left + (right - left) / phi
            fd, _ = evaluate(d)
    log_k = (left + right) / 2.0
    residual, amplitude = evaluate(log_k)
    mean = sum(values) / len(values)
    total = sum((value - mean) ** 2 for value in values)
    r_squared = 1.0 - residual / total if total > 0 else 1.0
    return amplitude, math.exp(log_k), r_squared


def parse_wdpt_capillary(
    formulations: dict[str, dict[str, object]],
) -> tuple[list[dict[str, object]], list[dict[str, object]], dict[str, object]]:
    curves: list[dict[str, object]] = []
    scalars: list[dict[str, object]] = []
    max_wdpt_mean_error = 0.0
    capillary_fit_r2: list[float] = []
    with zipfile.ZipFile(SOURCE_ZIP) as archive:
        member = "Raw Data/Experiment 4/WDPT/WDPT.csv"
        rows = _csv_rows(archive, member)
        if rows[0][:4] != ["Sample", "DropStart", "DropFinish", "WDPT"]:
            raise AuditBlocked("WDPT表头漂移")
        wdpt_by_sample: dict[str, list[float]] = defaultdict(list)
        reported_means: dict[str, float] = {}
        for source_row, values in enumerate(rows[1:], start=2):
            if len(values) != 9:
                raise AuditBlocked(f"WDPT列数漂移：{source_row}")
            sample = _sample_id(values[0])
            start = _number(values[1], label="WDPT start")
            finish = _number(values[2], label="WDPT finish")
            wdpt = _number(values[3], label="WDPT")
            if not _close(finish - start, wdpt, tolerance=1e-9):
                raise AuditBlocked(f"WDPT开始/结束时间不能复算：{source_row}")
            repeat = len(wdpt_by_sample[sample]) + 1
            wdpt_by_sample[sample].append(wdpt)
            if values[4].strip():
                reported_means[sample] = _number(values[4], label="WDPT mean")
            _add_scalar(
                scalars, formulations, sample=sample, observable="water_drop_penetration_time",
                value=wdpt, unit="s", source_location=f"Raw Data.zip!/{member}#row {source_row}",
                replicate_index=repeat, granularity="water_drop_replicate",
                derivation="DropFinish-DropStart", ceiling=0.40,
                notes="每个PU配方5滴重复；原AVI保留在ZIP中，未做视频二次推断。",
            )
        if len(wdpt_by_sample) != 16 or any(len(values) != 5 for values in wdpt_by_sample.values()):
            raise AuditBlocked("WDPT配方数或五次重复漂移")
        for sample, values in sorted(wdpt_by_sample.items()):
            mean = sum(values) / len(values)
            if sample not in reported_means:
                raise AuditBlocked(f"WDPT作者均值缺失：{sample}")
            max_wdpt_mean_error = max(max_wdpt_mean_error, abs(mean - reported_means[sample]))
            _add_scalar(
                scalars, formulations, sample=sample,
                observable="water_drop_penetration_time_mean", value=mean, unit="s",
                source_location=f"Raw Data.zip!/{member}#five-drop mean for {sample}",
                granularity="formulation_five_drop_mean", derivation="mean(5 WDPT replicates)",
                ceiling=0.45,
            )

        member = "Raw Data/Experiment 4/Capillary/Cap.csv"
        rows = _csv_rows(archive, member)
        expected_header = ["Sample", "Repeat", "0", "30", "60", "90", "120", "150", "210", "330", "570"]
        if rows[0] != expected_header:
            raise AuditBlocked("毛细上升表头漂移")
        times = [_number(value, label="capillary time") for value in expected_header[2:]]
        capillary_by_sample: dict[str, list[list[float]]] = defaultdict(list)
        for source_row, values in enumerate(rows[1:], start=2):
            if len(values) != len(expected_header):
                raise AuditBlocked(f"毛细曲线列数漂移：{source_row}")
            sample = _sample_id(values[0])
            repeat = int(_number(values[1], label="capillary repeat"))
            heights = [_number(value, label=f"capillary row {source_row}") for value in values[2:]]
            if repeat != len(capillary_by_sample[sample]) + 1 or abs(heights[0]) > 1e-12:
                raise AuditBlocked(f"毛细重复编号或零时点漂移：{source_row}")
            capillary_by_sample[sample].append(heights)
            common = _curve_common(sample, formulations, ceiling=0.45)
            curves.append(
                {
                    "curve_id": f"sheffield_capillary_{sample.lower()}_r{repeat}",
                    **common,
                    "curve_type": "capillary_rise_height",
                    "replicate_index": repeat,
                    "source_location": f"Raw Data.zip!/{member}#row {source_row}",
                    "point_count": len(times),
                    "x_name": "cumulative_sub_irrigation_time",
                    "x_unit": "s",
                    "y_names": "capillary_rise_height",
                    "y_units": "cm",
                    "x_min": times[0],
                    "x_max": times[-1],
                    "x_values_json": json.dumps(times, separators=(",", ":")),
                    "y_values_json": json.dumps(heights, separators=(",", ":")),
                    "raw_headers": ";".join(expected_header),
                    "data_origin": "experimental_raw_curve",
                    "record_granularity": "foam_specimen_capillary_repeat",
                    "notes": "20×20×50 mm试样；每配方3个独立试件。",
                }
            )
        if len(capillary_by_sample) != 17 or any(len(values) != 3 for values in capillary_by_sample.values()):
            raise AuditBlocked("毛细曲线配方/外部对照或三次重复漂移")
        for sample, repeats in sorted(capillary_by_sample.items()):
            means = [sum(repeat[index] for repeat in repeats) / 3.0 for index in range(len(times))]
            amplitude, rate, r_squared = _fit_capillary(times, means)
            capillary_fit_r2.append(r_squared)
            common = _curve_common(sample, formulations, ceiling=0.45)
            curves.append(
                {
                    "curve_id": f"sheffield_capillary_{sample.lower()}_mean",
                    **common,
                    "curve_type": "capillary_rise_height_triplicate_mean",
                    "replicate_index": "mean_of_3",
                    "source_location": f"Raw Data.zip!/{member}#triplicate mean for {sample}",
                    "point_count": len(times),
                    "x_name": "cumulative_sub_irrigation_time",
                    "x_unit": "s",
                    "y_names": "mean_capillary_rise_height",
                    "y_units": "cm",
                    "x_min": times[0],
                    "x_max": times[-1],
                    "x_values_json": json.dumps(times, separators=(",", ":")),
                    "y_values_json": json.dumps(means, separators=(",", ":")),
                    "raw_headers": ";".join(expected_header),
                    "data_origin": "experimental_derived_curve",
                    "record_granularity": "formulation_triplicate_mean",
                    "notes": "逐时点对3个独立试件取均值；与原始重复共享split_group。",
                }
            )
            admission, ceiling = _property_admission(sample)
            _add_scalar(
                scalars, formulations, sample=sample,
                observable="capillary_rise_height_at_570s", value=means[-1], unit="cm",
                source_location=f"Raw Data.zip!/{member}#triplicate mean at 570 s for {sample}",
                granularity="formulation_triplicate_mean", derivation="mean(3 repeats at 570 s)",
                admission=admission, ceiling=ceiling,
            )
            fit_admission = "evidence_only" if sample == "RW" else "conditional_reference"
            fit_ceiling = 0.05 if sample == "RW" else 0.30
            fit_note = (
                f"拟合y=a*(1-exp(-k*t))，R2={r_squared:.6g}；论文把α2写成cm/s，"
                "但指数项按量纲应为s^-1，故拟合参数仅条件参考。"
            )
            _add_scalar(
                scalars, formulations, sample=sample,
                observable="capillary_asymptotic_rise_alpha1", value=amplitude, unit="cm",
                source_location=f"Raw Data.zip!/{member}#bounded fit for {sample}",
                granularity="formulation_triplicate_mean_fit", derivation="bounded_least_squares_grid_free",
                admission=fit_admission, ceiling=fit_ceiling, notes=fit_note,
            )
            _add_scalar(
                scalars, formulations, sample=sample,
                observable="capillary_uptake_rate_alpha2", value=rate, unit="1/s",
                source_location=f"Raw Data.zip!/{member}#bounded fit for {sample}",
                granularity="formulation_triplicate_mean_fit", derivation="bounded_least_squares_grid_free",
                admission=fit_admission, ceiling=fit_ceiling, notes=fit_note,
            )

    if len(scalars) != 147 or len(curves) != 68 or max_wdpt_mean_error > 1e-9:
        raise AuditBlocked(
            f"WDPT/毛细计数或均值复算漂移：scalars={len(scalars)}, curves={len(curves)}, "
            f"mean_error={max_wdpt_mean_error}"
        )
    return curves, scalars, {
        "wdpt_pu_formulations": len(wdpt_by_sample),
        "wdpt_drop_replicates": sum(len(values) for values in wdpt_by_sample.values()),
        "capillary_materials_including_rockwool": len(capillary_by_sample),
        "capillary_raw_repeat_curves": sum(len(values) for values in capillary_by_sample.values()),
        "capillary_mean_curves": len(capillary_by_sample),
        "max_abs_reported_wdpt_mean_error": max_wdpt_mean_error,
        "capillary_fit_r2_min": min(capillary_fit_r2),
        "capillary_fit_r2_max": max(capillary_fit_r2),
    }


def _tsv(rows: list[dict[str, object]], columns: tuple[str, ...]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=columns, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in columns})
    return stream.getvalue().encode("utf-8")


def _json(value: object) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8")


def atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def run_audit(*, write_outputs: bool = True) -> dict[str, object]:
    frozen_files = verify_files()
    members = audit_zip_members()
    formulations, formulations_by_sample = parse_formulations()
    kinetic_curves, kinetic_scalars, kinetic_checks = parse_kinetics(formulations_by_sample)
    cell_curves, cell_scalars, cell_checks = parse_cell_size(formulations_by_sample)
    property_scalars, property_checks = parse_airflow_density_whc(formulations_by_sample)
    capillary_curves, wdpt_capillary_scalars, water_checks = parse_wdpt_capillary(
        formulations_by_sample
    )
    curves = kinetic_curves + cell_curves + capillary_curves
    scalars = kinetic_scalars + cell_scalars + property_scalars + wdpt_capillary_scalars

    if len(curves) != EXPECTED_CURVES or len(scalars) != EXPECTED_SCALARS:
        raise AuditBlocked(f"总曲线/标量计数漂移：{len(curves)}, {len(scalars)}")
    if len({row["curve_id"] for row in curves}) != len(curves):
        raise AuditBlocked("曲线ID不唯一")
    if len({row["scalar_id"] for row in scalars}) != len(scalars):
        raise AuditBlocked("标量ID不唯一")
    if {
        row["split_group"] for row in scalars if row["sample_id"] in {"F05-13", "F05-34"}
    } != {formulations_by_sample["F05-13"]["split_group"]}:
        raise AuditBlocked("F05-13/F05-34没有共享同配方split_group")

    scalar_statuses = Counter(str(row["gold_admission_status"]) for row in scalars)
    curve_statuses = Counter(str(row["gold_admission_status"]) for row in curves)
    if scalar_statuses != {
        "admitted_reference": 723,
        "conditional_reference": 32,
        "evidence_only": 9,
    }:
        raise AuditBlocked(f"标量准入计数漂移：{dict(scalar_statuses)}")
    if curve_statuses != {"admitted_reference": 151, "evidence_only": 4}:
        raise AuditBlocked(f"曲线准入计数漂移：{dict(curve_statuses)}")

    member_counts = Counter(str(row["member_type"]) for row in members)
    observable_counts = Counter(str(row["observable"]) for row in scalars)
    summary = {
        "audit_version": AUDIT_VERSION,
        "source": {
            "title": "PU Rational Design Data",
            "dataset_doi": DATASET_DOI,
            "paper_doi": PAPER_DOI,
            "dataset_url": DATASET_URL,
            "api_url": API_URL,
            "download_url": DOWNLOAD_URL,
            "license": LICENSE,
            "source_reliability": "R1",
        },
        "counts": {
            "zip_members": len(members),
            "zip_files": sum(row["member_type"] != "directory" for row in members),
            "zip_directories": sum(row["member_type"] == "directory" for row in members),
            "experimental_sample_batches": len(formulations),
            "unique_formulations": len({row["formulation_id"] for row in formulations}),
            "screening_sample_batches": sum(int(row["experiment"]) < 4 for row in formulations),
            "screening_unique_formulations": len({row["formulation_id"] for row in formulations if int(row["experiment"]) < 4}),
            "final_doe_unique_formulations": len({row["formulation_id"] for row in formulations if int(row["experiment"]) == 4}),
            "duplicate_formula_control_batches": sum(bool(row["duplicate_of_sample_id"]) for row in formulations),
            "numeric_curves": len(curves),
            "scalar_records": len(scalars),
            "pu_scalar_records": sum(not bool(row["is_external_control"]) for row in scalars),
            "external_rockwool_control_scalars": sum(bool(row["is_external_control"]) for row in scalars),
        },
        "zip_member_type_counts": dict(sorted(member_counts.items())),
        "curve_type_counts": dict(sorted(Counter(str(row["curve_type"]) for row in curves).items())),
        "scalar_observable_counts": dict(sorted(observable_counts.items())),
        "gold_admission_counts": {
            "curves": dict(sorted(curve_statuses.items())),
            "scalars": dict(sorted(scalar_statuses.items())),
        },
        "split_group_evidence": {
            "unit": "unique PPHP formulation signature",
            "unique_pu_split_groups": EXPECTED_UNIQUE_FORMULATIONS,
            "duplicate_control_pair": ["F05-13", "F05-34"],
            "duplicate_pair_shared_split_group": formulations_by_sample["F05-13"]["split_group"],
        },
        "reconciliation_checks": {
            **kinetic_checks,
            **cell_checks,
            **property_checks,
            **water_checks,
        },
        "scientific_classification": {
            "gold_layer": "Gold-E",
            "scientific_role": "PU泡沫配方-反应动力学-泡孔形貌-流体输运多任务参考",
            "maximum_future_weight": 0.45,
            "direct_monomer_structure_supervision": False,
            "recommended_grouping": "formulation split_group; never random row/file split",
            "recommended_uses": [
                "PPHP配方到绝热温升、泡高、异氰酸酯转化率的多任务预训练",
                "催化剂/表面活性剂到泡孔直径、气流、开孔率的配方级代理模型",
                "最终16配方的密度、持水、WDPT和毛细曲线多任务参考",
                "主动学习中对新泡沫配方的实验优先级排序",
            ],
        },
        "limitations": [
            "商业原料只有产品级身份和PPHP，没有单体SMILES、批次Mn/Mw或结构分布。",
            "40个实验样品编号对应39个唯一配方；F05-34是F05-13的配方对照重复。",
            "ATRsumm中的ATR是绝热温升，不是ATR-FTIR光谱。",
            "最终DoE多数性能只有单个配方批次；泡孔、WDPT和毛细有图像/滴液/试件重复，但不能把文件数当独立材料数。",
            "补充表把第二表面活性剂写作Vorasurf 5959，论文正文写作Tegostab 8476；两种名称都保留。",
            "论文把毛细拟合α2描述为cm/s，但指数函数按量纲应为1/s；重新拟合参数仅条件参考。",
            "Rockwool记录是非PU外部基准，只保留为evidence_only。",
        ],
    }

    outputs = {
        "内容审计摘要.json": _json(summary),
        "配方审计清单.tsv": _tsv(formulations, FORMULATION_COLUMNS),
        "曲线审计清单.tsv": _tsv(curves, CURVE_COLUMNS),
        "标量审计清单.tsv": _tsv(scalars, SCALAR_COLUMNS),
        "ZIP成员审计清单.tsv": _tsv(members, MEMBER_COLUMNS),
        "文件校验清单.tsv": _tsv(
            frozen_files, ("file", "bytes", "sha256", "md5", "verification")
        ),
    }
    if tuple(outputs) != OUTPUT_NAMES:
        raise AuditBlocked("输出文件集合漂移")
    if write_outputs:
        for name, payload in outputs.items():
            atomic_write(SOURCE_DIR / name, payload)
    return {
        "summary": summary,
        "formulations": formulations,
        "curves": curves,
        "scalars": scalars,
        "members": members,
        "files": frozen_files,
        "outputs": outputs,
    }


if __name__ == "__main__":
    result = run_audit(write_outputs=True)
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
