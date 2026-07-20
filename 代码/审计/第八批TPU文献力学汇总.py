"""审计 Mendeley TPU 文献力学汇总数据（10.17632/ftntxg4zdz.1）。

该来源是文献二次汇总，不是 62 个原始试样。脚本保留每一行的生产方式、
Shore 硬度、拉伸强度、断裂伸长率和原始引用，并按引用分组防止同一论文
跨数据折。输出只用于 Gold-E 条件参考层，当前不会创建训练划分或权重。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import tempfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第八批实验_TPU文献力学汇总"
)
SOURCE_XLSX = SOURCE_DIR / "原始数据.xlsx"
SOURCE_RIS = SOURCE_DIR / "参考文献.ris"
SOURCE_SNAPSHOT = SOURCE_DIR / "来源快照.json"
ZIP_METADATA = SOURCE_DIR / "压缩包元数据.json"

DOI = "10.17632/ftntxg4zdz.1"
LICENSE = "CC BY 4.0"
SOURCE_URL = "https://data.mendeley.com/datasets/ftntxg4zdz/1"
AUDIT_VERSION = "batch8-tpu-literature-mechanics-v1"
SHEET_NAME = "Fig18 and Fig19"

FROZEN_FILES = {
    "原始数据.xlsx": (
        16_585,
        "7bbb8f5bf02c3f1c37ff0f1cddee8423ae65ee54f13d2d8255d19aad3a7b271b",
    ),
    "参考文献.ris": (
        61_745,
        "a04a262afbebce099a9217b2e0b13dd5ba33b0664c8018840cd1ec8ad626ac47",
    ),
    "来源快照.json": (
        2_165,
        "1c2c4804c628c177ef26dfacf54f0ceba69fd7c0a68c7bffdf33f5b2def427fc",
    ),
    "压缩包元数据.json": (
        291,
        "71926990e26b1893774ec96b5f68d6364ce5de1f6800316d373e34bbb01476e3",
    ),
}

EXPECTED_ROWS = 62
EXPECTED_SCALARS = 186
EXPECTED_REFERENCE_GROUPS = 45
EXPECTED_RIS_RECORDS = 37
EXPECTED_REFERENCE_DOIS = 30
EXPECTED_PROCESSES = {"FDM": 31, "IM": 15, "SLS": 9, "MJF": 7}
EXPECTED_HEADERS = (
    "Production technique",
    "Shore A/D",
    "UTS [MPa]",
    "A [%]",
    "Reference",
)
OUTPUT_NAMES = ("内容审计摘要.json", "标量审计清单.tsv", "文件校验清单.tsv")

SCALAR_COLUMNS = (
    "scalar_id",
    "record_id",
    "source_row",
    "source_location",
    "split_group",
    "production_technique",
    "shore_hardness_raw",
    "shore_hardness_value",
    "shore_hardness_scale",
    "observable",
    "value",
    "unit",
    "target_origin",
    "data_origin",
    "record_granularity",
    "quality_gate",
    "future_weight_ceiling",
    "reference_group_id",
    "reference_doi",
    "reference_text",
    "notes",
)


class AuditBlocked(RuntimeError):
    """原件身份、工作簿结构或冻结计数发生漂移。"""


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_files() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for name, (expected_bytes, expected_sha256) in FROZEN_FILES.items():
        path = SOURCE_DIR / name
        if not path.is_file():
            raise AuditBlocked(f"缺少原件：{path}")
        actual_bytes = path.stat().st_size
        actual_sha256 = sha256(path)
        if (actual_bytes, actual_sha256) != (expected_bytes, expected_sha256):
            raise AuditBlocked(
                f"原件漂移：{name} bytes={actual_bytes}, sha256={actual_sha256}"
            )
        rows.append(
            {
                "file": name,
                "bytes": actual_bytes,
                "sha256": actual_sha256,
                "verification": "matched_frozen_identity",
            }
        )

    snapshot = json.loads(SOURCE_SNAPSHOT.read_text(encoding="utf-8"))
    if snapshot.get("doi") != DOI or snapshot.get("version") != 1:
        raise AuditBlocked("Mendeley DOI 或版本漂移")
    if snapshot.get("licence", {}).get("short_name") != LICENSE:
        raise AuditBlocked("Mendeley 许可漂移")
    zip_metadata = json.loads(ZIP_METADATA.read_text(encoding="utf-8"))
    if zip_metadata.get("size") != 34_727 or zip_metadata.get("sha256_hash") != (
        "d89a3e38bb4e9ebef232f4e6ee802b0134a97e1fdf545c09b4c9fd8736a7eb67"
    ):
        raise AuditBlocked("Mendeley 压缩包身份漂移")
    return rows


def _normalize_reference(text: str) -> str:
    return " ".join(text.split()).strip().rstrip(".")


def _reference_group(text: str) -> str:
    digest = hashlib.sha256(_normalize_reference(text).casefold().encode("utf-8"))
    return f"ref_{digest.hexdigest()[:16]}"


def _extract_doi(text: str) -> str:
    match = re.search(r"(?i)10\.\d{4,9}/[^\s]+", text)
    return match.group(0).rstrip(".,;:)]}").lower() if match else ""


def _number(value: object, *, label: str, excel_row: int) -> float:
    if isinstance(value, bool):
        raise AuditBlocked(f"{label} 第{excel_row}行不是数值")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise AuditBlocked(f"{label} 第{excel_row}行无法解析：{value!r}") from exc
    if not math.isfinite(result):
        raise AuditBlocked(f"{label} 第{excel_row}行不是有限数")
    return result


def _hardness(value: object, excel_row: int) -> tuple[str, float, str]:
    raw = str(value).strip()
    scale = "D" if raw.upper().endswith("D") else "A"
    numeric_text = raw[:-1] if scale == "D" else raw
    numeric = _number(numeric_text, label="Shore硬度", excel_row=excel_row)
    return raw, numeric, scale


def parse_workbook() -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise AuditBlocked(f"工作表漂移：{workbook.sheetnames}")
        sheet = workbook[SHEET_NAME]
        header = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        if header != EXPECTED_HEADERS:
            raise AuditBlocked(f"表头漂移：{header}")

        records: list[dict[str, object]] = []
        exact_rows: list[tuple[object, ...]] = []
        for excel_row, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=5, values_only=True), start=2
        ):
            if all(value is None for value in values):
                continue
            if any(value is None or str(value).strip() == "" for value in values):
                raise AuditBlocked(f"第{excel_row}行存在缺字段：{values}")
            process, hardness_source, uts_source, elongation_source, reference_source = values
            process = str(process).strip()
            if process not in EXPECTED_PROCESSES:
                raise AuditBlocked(f"未知生产方式：{process}")
            hardness_raw, hardness_value, hardness_scale = _hardness(
                hardness_source, excel_row
            )
            uts = _number(uts_source, label="UTS", excel_row=excel_row)
            elongation = _number(
                elongation_source, label="断裂伸长率", excel_row=excel_row
            )
            reference = _normalize_reference(str(reference_source))
            reference_group = _reference_group(reference)
            reference_doi = _extract_doi(reference)
            record_id = f"ftntxg4zdz_v1_row_{excel_row:03d}"
            exact_rows.append((process, hardness_raw, uts, elongation, reference))

            common = {
                "record_id": record_id,
                "source_row": excel_row,
                "source_location": f"原始数据.xlsx#{SHEET_NAME}!A{excel_row}:E{excel_row}",
                "split_group": f"doi:{DOI}|{reference_group}",
                "production_technique": process,
                "shore_hardness_raw": hardness_raw,
                "shore_hardness_value": hardness_value,
                "shore_hardness_scale": hardness_scale,
                "target_origin": "experimental",
                "data_origin": "experimental_literature_aggregate",
                "record_granularity": "literature_aggregate",
                "quality_gate": "conditional_reference",
                "future_weight_ceiling": 0.20 if reference_doi else 0.10,
                "reference_group_id": reference_group,
                "reference_doi": reference_doi,
                "reference_text": reference,
                "notes": (
                    "文献或数据表汇总值；无原始试样、重复、牌号化学组成和完整测试协议，"
                    "同一引用全部同折，化学结构任务权重为0。"
                ),
            }
            for suffix, observable, value, unit in (
                ("shore", "shore_hardness", hardness_value, f"Shore {hardness_scale}"),
                ("uts", "ultimate_tensile_strength", uts, "MPa"),
                ("elong", "elongation_at_break", elongation, "%"),
            ):
                records.append(
                    {
                        "scalar_id": f"{record_id}_{suffix}",
                        **common,
                        "observable": observable,
                        "value": value,
                        "unit": unit,
                    }
                )
    finally:
        workbook.close()

    literature_rows = len(exact_rows)
    reference_groups = {row["reference_group_id"] for row in records}
    reference_dois = {row["reference_doi"] for row in records if row["reference_doi"]}
    processes = Counter(row[0] for row in exact_rows)
    if literature_rows != EXPECTED_ROWS or len(records) != EXPECTED_SCALARS:
        raise AuditBlocked("文献行数或标量数漂移")
    if len(set(exact_rows)) != EXPECTED_ROWS:
        raise AuditBlocked("工作簿出现精确重复行")
    if dict(processes) != EXPECTED_PROCESSES:
        raise AuditBlocked(f"生产方式计数漂移：{dict(processes)}")
    if len(reference_groups) != EXPECTED_REFERENCE_GROUPS:
        raise AuditBlocked("引用分组数漂移")
    if len(reference_dois) != EXPECTED_REFERENCE_DOIS:
        raise AuditBlocked("引用 DOI 数漂移")

    ris_text = SOURCE_RIS.read_text(encoding="utf-8-sig")
    ris_records = sum(1 for line in ris_text.splitlines() if line.startswith("TY  -"))
    if ris_records != EXPECTED_RIS_RECORDS:
        raise AuditBlocked(f"RIS记录数漂移：{ris_records}")

    values_by_property = {
        observable: [float(row["value"]) for row in records if row["observable"] == observable]
        for observable in ("shore_hardness", "ultimate_tensile_strength", "elongation_at_break")
    }
    statistics = {
        observable: {
            "count": len(values),
            "min": min(values),
            "max": max(values),
            "unique": len(set(values)),
        }
        for observable, values in values_by_property.items()
    }
    summary = {
        "audit_version": AUDIT_VERSION,
        "source": {
            "title": "Literature database in mechanical characteristic of Thermoplastic Polyurethane",
            "doi": DOI,
            "version": 1,
            "url": SOURCE_URL,
            "license": LICENSE,
            "source_reliability": "R1",
        },
        "counts": {
            "literature_aggregate_rows": literature_rows,
            "scalar_records": len(records),
            "numeric_values": len(records),
            "exact_duplicate_rows": 0,
            "reference_groups": len(reference_groups),
            "reference_dois": len(reference_dois),
            "ris_records": ris_records,
            "independent_specimens": 0,
            "resolved_formulations": 0,
        },
        "production_technique_counts": dict(processes),
        "property_statistics": statistics,
        "scientific_classification": {
            "gold_layer": "Gold-E",
            "gold_admission_status": "conditional_reference",
            "scientific_role": "TPU加工-硬度-力学迁移参考",
            "independent_weight_unit": "reference_group",
            "maximum_future_weight": 0.20,
            "direct_chemistry_property_supervision": False,
        },
        "limitations": [
            "62行是文献或商业数据表汇总，不是62个原始独立试样。",
            "同一引用的多行共享泄漏组；曲线、重复数、测试标准和完整加工参数未提供。",
            "工作簿没有稳定牌号、配方、SMILES、分子量、硬段含量或NCO/OH。",
            "未标D的硬度按工作簿语义保存为Shore A；3行显式70D单独保留为Shore D。",
            "CC BY 4.0适用于数据集；引用条目中的第三方内容仍需按原论文权利核对。",
        ],
    }
    return records, summary


def _tsv(rows: list[dict[str, object]], columns: tuple[str, ...]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=columns, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in columns})
    return stream.getvalue().encode("utf-8")


def _json(value: object) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8")


def atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def run_audit(*, write_outputs: bool = True) -> dict[str, object]:
    files = verify_files()
    scalar_rows, summary = parse_workbook()
    outputs = {
        "内容审计摘要.json": _json(summary),
        "标量审计清单.tsv": _tsv(scalar_rows, SCALAR_COLUMNS),
        "文件校验清单.tsv": _tsv(
            files, ("file", "bytes", "sha256", "verification")
        ),
    }
    if write_outputs:
        for name, payload in outputs.items():
            atomic_write(SOURCE_DIR / name, payload)
    return {"summary": summary, "scalars": scalar_rows, "files": files, "outputs": outputs}


if __name__ == "__main__":
    result = run_audit(write_outputs=True)
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
