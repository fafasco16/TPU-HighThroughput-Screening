"""只读复算三个标准力学开放数据来源的来源级审计。

覆盖来源：

* ``MaterialsCloud_商用PU泡沫多轴断裂力学``
* ``ScienceDB_微孔PU动态力学``
* ``Texas_湿干单根电纺PU纤维力学``

脚本只读取科学输入，只能原子覆盖 ``OUTPUT_WHITELIST`` 中已有名称的审计
JSON/TSV。ZIP 仅做 CRC、路径安全和既有解压副本逐文件 SHA-256 核验，不会
重新解压。审计基准日为协议常量；同一科学输入的输出应逐字节稳定。

运行方式（可从任意当前目录调用）：

    python 代码/审计/新增开放数据标准力学三源.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Iterable

from openpyxl import load_workbook
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"
AUDIT_VERSION = "1.0"

MATERIALS_CLOUD = "MaterialsCloud_商用PU泡沫多轴断裂力学"
SCIENCE_DB = "ScienceDB_微孔PU动态力学"
TEXAS = "Texas_湿干单根电纺PU纤维力学"
SOURCE_NAMES = (MATERIALS_CLOUD, SCIENCE_DB, TEXAS)

OUTPUT_NAMES_BY_SOURCE = {
    MATERIALS_CLOUD: (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "曲线审计清单.tsv",
    ),
    SCIENCE_DB: (
        "内容审计摘要.json",
        "文件校验清单.tsv",
    ),
    TEXAS: (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "曲线审计清单.tsv",
    ),
}
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source, filenames in OUTPUT_NAMES_BY_SOURCE.items()
    for filename in filenames
)


class AuditBlocked(RuntimeError):
    """输入、完整性或科学语义不满足固定审计协议。"""


def _same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def _is_reparse_point(path: Path) -> bool:
    """同时识别符号链接和 Windows junction/其他重解析点。"""

    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & reparse_flag) or bool(is_junction())


def _assert_plain_chain(path: Path, stop: Path) -> None:
    if path != stop and stop not in path.parents:
        raise AuditBlocked(f"路径越出审计根目录：{path}")
    cursor = path
    while True:
        if _is_reparse_point(cursor):
            raise AuditBlocked(f"拒绝符号链接或重解析点：{cursor}")
        if cursor == stop:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"目录缺失、不是普通目录或经链接解析：{path}")
    _assert_plain_chain(path, PROJECT_ROOT)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    _assert_plain_chain(path, PROJECT_ROOT)


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")

    # 保持本函数自包含：公共安全测试会仅抽取本定义、AuditBlocked 与
    # atomic_write，而不会执行模块里的其他辅助函数。
    def is_reparse(candidate: Path) -> bool:
        try:
            details = candidate.lstat()
        except FileNotFoundError:
            return False
        attributes = getattr(details, "st_file_attributes", 0)
        is_junction = getattr(candidate, "is_junction", lambda: False)
        return (
            candidate.is_symlink()
            or bool(attributes & 0x400)  # FILE_ATTRIBUTE_REPARSE_POINT
            or bool(is_junction())
        )

    parent = path.parent
    try:
        resolved_parent = parent.resolve(strict=True)
    except FileNotFoundError as exc:
        raise AuditBlocked(f"审计输出目录不存在：{parent}") from exc
    if (
        not parent.is_dir()
        or is_reparse(parent)
        or os.path.normcase(os.path.abspath(str(resolved_parent)))
        != os.path.normcase(os.path.abspath(str(parent)))
    ):
        raise AuditBlocked(f"拒绝通过符号链接或重解析目录写入：{parent}")

    if path.exists() or path.is_symlink():
        if is_reparse(path):
            raise AuditBlocked(f"拒绝覆盖符号链接或重解析输出：{path}")
        try:
            resolved = path.resolve(strict=True)
        except FileNotFoundError as exc:
            raise AuditBlocked(f"审计输出链接目标缺失：{path}") from exc
        if (
            not path.is_file()
            or os.path.normcase(os.path.abspath(str(resolved)))
            != os.path.normcase(os.path.abspath(str(path)))
        ):
            raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    """同目录普通临时文件落盘同步后原子替换单个白名单输出。"""

    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        details = temporary.lstat()
        attributes = getattr(details, "st_file_attributes", 0)
        is_junction = getattr(temporary, "is_junction", lambda: False)
        if (
            temporary.is_symlink()
            or bool(attributes & 0x400)
            or bool(is_junction())
            or not temporary.is_file()
            or os.path.normcase(os.path.abspath(str(temporary.resolve(strict=True))))
            != os.path.normcase(os.path.abspath(str(temporary)))
        ):
            raise AuditBlocked(f"审计临时输出不是普通文件：{temporary}")
        assert_output_allowed(path)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_json(path: Path, payload: dict[str, object]) -> None:
    rendered = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        indent=2,
        allow_nan=False,
    )
    atomic_write(path, (rendered + "\n").encode("utf-8"))


def write_tsv(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    if not rows:
        raise AuditBlocked(f"拒绝写入空TSV：{path}")
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    atomic_write(path, buffer.getvalue().encode("utf-8"))


def _hash_stream(handle: BinaryIO, algorithm: str = "sha256") -> str:
    digest = hashlib.new(algorithm)
    for block in iter(lambda: handle.read(1024 * 1024), b""):
        digest.update(block)
    return digest.hexdigest()


def file_hash(path: Path, algorithm: str = "sha256") -> str:
    require_file(path)
    with path.open("rb") as handle:
        return _hash_stream(handle, algorithm)


def load_json(path: Path) -> dict[str, object]:
    require_file(path)
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise AuditBlocked(f"JSON根节点不是对象：{path}")
    return value


def _manifest_digest(items: Iterable[tuple[str, int, str]]) -> str:
    digest = hashlib.sha256()
    for relative, size, checksum in sorted(items):
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(str(size).encode("ascii"))
        digest.update(b"\0")
        digest.update(checksum.encode("ascii"))
        digest.update(b"\n")
    return digest.hexdigest()


def scientific_input_snapshot() -> dict[str, tuple[int, str]]:
    """对三个来源的全部非审计产物建立路径、大小和 SHA-256 快照。"""

    snapshot: dict[str, tuple[int, str]] = {}
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        require_directory(base)
        for path in sorted(base.rglob("*")):
            if _is_reparse_point(path):
                raise AuditBlocked(f"科学输入树含符号链接或重解析点：{path}")
            if not path.is_file() or path in OUTPUT_WHITELIST:
                continue
            if path.name.endswith(".audit.tmp"):
                raise AuditBlocked(f"发现遗留审计临时文件：{path}")
            require_file(path)
            relative = path.relative_to(PROJECT_ROOT).as_posix()
            snapshot[relative] = (path.stat().st_size, file_hash(path))
    return snapshot


def snapshot_by_source(
    snapshot: dict[str, tuple[int, str]],
) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for source in SOURCE_NAMES:
        marker = f"/新增开放数据/{source}/"
        items = [
            (path, size, checksum)
            for path, (size, checksum) in snapshot.items()
            if marker in f"/{path}"
        ]
        result[source] = {
            "文件数": len(items),
            "总字节数": sum(size for _, size, _ in items),
            "清单SHA256": _manifest_digest(items),
        }
    return result


def _safe_zip_name(name: str) -> str:
    if "\x00" in name or re.match(r"^[A-Za-z]:", name):
        raise AuditBlocked(f"ZIP含危险成员路径：{name!r}")
    candidate = PurePosixPath(name.replace("\\", "/"))
    if candidate.is_absolute() or ".." in candidate.parts:
        raise AuditBlocked(f"ZIP含危险成员路径：{name!r}")
    normalized = candidate.as_posix()
    if normalized in {"", "."}:
        raise AuditBlocked(f"ZIP含空成员路径：{name!r}")
    return normalized


def audit_zip_mirror(archive_path: Path, extracted_root: Path) -> dict[str, object]:
    """核验 ZIP CRC，并把每个文件成员与既有只读解压副本逐哈希比对。"""

    require_file(archive_path)
    require_directory(extracted_root)
    local_files = {
        path.relative_to(extracted_root).as_posix(): path
        for path in sorted(extracted_root.rglob("*"))
        if path.is_file()
    }
    for path in extracted_root.rglob("*"):
        if _is_reparse_point(path):
            raise AuditBlocked(f"解压副本含符号链接或重解析点：{path}")

    with zipfile.ZipFile(archive_path) as archive:
        bad_crc = archive.testzip()
        if bad_crc is not None:
            raise AuditBlocked(f"ZIP CRC失败：{archive_path.name}/{bad_crc}")
        infos = archive.infolist()
        normalized = [_safe_zip_name(info.filename) for info in infos]
        duplicates = [name for name, count in Counter(normalized).items() if count > 1]
        if duplicates:
            raise AuditBlocked(f"ZIP含重复成员路径：{duplicates}")

        file_infos: list[tuple[str, zipfile.ZipInfo]] = []
        for name, info in zip(normalized, infos):
            mode = (info.external_attr >> 16) & 0o170000
            if mode == stat.S_IFLNK:
                raise AuditBlocked(f"ZIP含符号链接成员：{name}")
            if not info.is_dir():
                file_infos.append((name, info))

        expected = {name for name, _ in file_infos}
        actual = set(local_files)
        if expected != actual:
            raise AuditBlocked(
                "ZIP与解压副本路径集合不一致："
                f"缺失={sorted(expected - actual)}, 多余={sorted(actual - expected)}"
            )

        manifest_items: list[tuple[str, int, str]] = []
        extensions: Counter[str] = Counter()
        for name, info in sorted(file_infos):
            local = local_files[name]
            require_file(local)
            if local.stat().st_size != info.file_size:
                raise AuditBlocked(f"ZIP成员与解压文件大小不一致：{name}")
            with archive.open(info, "r") as member:
                member_hash = _hash_stream(member)
            local_hash = file_hash(local)
            if member_hash != local_hash:
                raise AuditBlocked(f"ZIP成员与解压文件SHA256不一致：{name}")
            manifest_items.append((name, info.file_size, member_hash))
            extension = PurePosixPath(name).suffix.upper().lstrip(".") or "无扩展名"
            extensions[extension] += 1

    return {
        "ZIP_CRC_testzip": "通过；无坏成员",
        "ZIP条目数": len(infos),
        "目录数": sum(info.is_dir() for info in infos),
        "文件数": len(file_infos),
        "解压后字节数": sum(info.file_size for _, info in file_infos),
        "危险路径数": 0,
        "重复成员路径数": 0,
        "解压逐文件SHA256一致数": len(manifest_items),
        "解压逐文件清单SHA256": _manifest_digest(manifest_items),
        "扩展名盘点": dict(sorted(extensions.items())),
    }


def verify_docx(path: Path) -> None:
    require_file(path)
    with zipfile.ZipFile(path) as archive:
        bad = archive.testzip()
        names = {_safe_zip_name(info.filename) for info in archive.infolist()}
    if bad is not None or "word/document.xml" not in names:
        raise AuditBlocked(f"README.docx OOXML完整性失败：{path}")


def _decode_csv(path: Path) -> str:
    payload = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise AuditBlocked(f"CSV编码无法按固定协议解析：{path}")


def _finite_number(value: object, *, decimal_comma: bool = False) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace("−", "-")
    if not text:
        return None
    if decimal_comma and "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _asset_row(
    path: Path,
    official_md5: str,
    source_url: str,
    status: str,
) -> dict[str, object]:
    actual_md5 = file_hash(path, "md5")
    if official_md5 and actual_md5.casefold() != official_md5.casefold():
        raise AuditBlocked(f"官方MD5不匹配：{path.name}")
    return {
        "文件名": path.name,
        "字节数": path.stat().st_size,
        "MD5": actual_md5,
        "SHA256": file_hash(path),
        "官方MD5": official_md5,
        "官方MD5匹配": "是" if official_md5 else "不适用",
        "来源或直链": source_url,
        "状态": status,
    }


def _materials_group(path: Path, puf_root: Path) -> tuple[str, str, str]:
    parts = path.relative_to(puf_root).parts
    test = parts[0]
    if test == "Toughness" and len(parts) >= 3:
        return test, "未分方向", parts[1]
    if test in {"Compression", "Shear", "Tension"} and len(parts) >= 4:
        return test, parts[1], parts[2]
    raise AuditBlocked(f"MaterialsCloud试样路径无法分组：{path.relative_to(puf_root)}")


def _materials_csv_counts(path: Path, dic: bool) -> tuple[int, int]:
    reader = csv.reader(io.StringIO(_decode_csv(path), newline=""), delimiter=";")
    data_rows = 0
    missing_dic_rows = 0
    for row in reader:
        if not row or _finite_number(row[0], decimal_comma=True) is None:
            continue
        data_rows += 1
        if dic and (
            len(row) < 2
            or any(
                _finite_number(value, decimal_comma=True) is None for value in row[1:]
            )
        ):
            missing_dic_rows += 1
    return data_rows, missing_dic_rows


def audit_materials_cloud() -> dict[str, object]:
    base = DATA_ROOT / MATERIALS_CLOUD
    archive_path = base / "PUF.zip"
    readme_path = base / "README.txt"
    metadata_path = base / "官方MaterialsCloud元数据.json"
    datacite_path = base / "官方DataCite元数据.json"
    for path in (archive_path, readme_path, metadata_path, datacite_path):
        require_file(path)

    metadata = load_json(metadata_path)
    datacite = load_json(datacite_path)
    doi = str(metadata["pids"]["doi"]["identifier"])
    datacite_doi = str(datacite["data"]["attributes"]["doi"])
    if doi.casefold() != "10.24435/materialscloud:vf-ry" or (
        datacite_doi.casefold() != doi.casefold()
    ):
        raise AuditBlocked("MaterialsCloud DOI元数据不一致")
    rights = metadata["metadata"]["rights"]
    if not rights or rights[0]["id"] != "cc-by-4.0":
        raise AuditBlocked("MaterialsCloud许可不再是CC BY 4.0")
    official_entries = metadata["files"]["entries"]
    if metadata["files"]["count"] != 3 or metadata["files"]["total_bytes"] != 342_207_768:
        raise AuditBlocked("MaterialsCloud官方文件总量发生变化")
    if set(official_entries) != {"PUF.zip", "README.txt", "SFFE.zip"}:
        raise AuditBlocked("MaterialsCloud官方文件名集合发生变化")
    local_root_archives = {
        path.name for path in base.glob("*.zip") if path.is_file() or path.is_symlink()
    }
    if local_root_archives != {"PUF.zip"}:
        raise AuditBlocked(
            "MaterialsCloud本地ZIP范围与审计子集不一致："
            f"{sorted(local_root_archives)}"
        )
    if (base / "SFFE.zip").exists() or (base / "SFFE.zip").is_symlink():
        raise AuditBlocked("SFFE.zip 已出现，必须先建立独立范围和审计规则")

    zip_audit = audit_zip_mirror(archive_path, base / "解压内容")
    if zip_audit["ZIP条目数"] != 204 or zip_audit["文件数"] != 157:
        raise AuditBlocked(f"MaterialsCloud PUF.zip固定内容发生变化：{zip_audit}")

    puf_root = base / "解压内容" / "PUF"
    require_directory(puf_root)
    groups: defaultdict[tuple[str, str, str], dict[str, list[Path]]] = defaultdict(
        lambda: {"raw": [], "dic": [], "images": []}
    )
    aggregate: defaultdict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    csv_paths = sorted(puf_root.rglob("*.csv"))
    for path in csv_paths:
        group = _materials_group(path, puf_root)
        kind = "dic" if "Correlation" in path.name else "raw"
        if kind == "raw" and "RawData" not in path.name:
            raise AuditBlocked(f"MaterialsCloud未知CSV角色：{path}")
        groups[group][kind].append(path)
        rows, missing = _materials_csv_counts(path, kind == "dic")
        aggregate[group[:2]][kind] += rows
        aggregate[group[:2]]["missing"] += missing

    image_suffixes = {".png", ".tif", ".tiff"}
    image_paths = sorted(
        path for path in puf_root.rglob("*") if path.suffix.lower() in image_suffixes
    )
    for path in image_paths:
        require_file(path)
        with Image.open(path) as image:
            image.verify()
        if path.parent.name.startswith(("Sample ", "Specimen ")):
            groups[_materials_group(path, puf_root)]["images"].append(path)

    bad_groups = {
        "|".join(group): {kind: len(paths) for kind, paths in roles.items()}
        for group, roles in sorted(groups.items())
        if len(roles["raw"]) != 1
        or len(roles["dic"]) != 1
        or len(roles["images"]) != 2
    }
    if bad_groups:
        raise AuditBlocked(f"MaterialsCloud试样多模态绑定不完整：{bad_groups}")

    expected_layout = [
        ("Compression", "ASTM D695", "Direction 11", 6),
        ("Compression", "ASTM D695", "Direction 22", 5),
        ("Shear", "modified SCA;k=1.3", "Direction 12", 8),
        ("Tension", "ASTM D638", "Direction 11", 6),
        ("Tension", "ASTM D638", "Direction 22", 6),
        ("Toughness", "ASTM E399;SENB", "未分方向", 6),
    ]
    curve_rows: list[dict[str, object]] = []
    test_details: list[dict[str, object]] = []
    for test, standard, direction, expected_specimens in expected_layout:
        specimen_count = sum(
            group[:2] == (test, direction) for group in groups
        )
        counts = aggregate[(test, direction)]
        if specimen_count != expected_specimens:
            raise AuditBlocked(
                f"MaterialsCloud {test}/{direction}试样数={specimen_count}，"
                f"预期={expected_specimens}"
            )
        complete = counts["raw"] + counts["dic"] - counts["missing"]
        note = (
            "Sample 1在156.4-179.6 s缺epsXY；隔离59行，禁止插值冒充实测"
            if counts["missing"]
            else "完整"
        )
        direction_key = direction.replace(" ", "") if direction != "未分方向" else "NA"
        curve_rows.append(
            {
                "试验": test,
                "标准": standard,
                "方向": direction,
                "独立试样数": specimen_count,
                "机器原始曲线数": len(
                    [group for group in groups if group[:2] == (test, direction)]
                ),
                "DIC曲线数": len(
                    [group for group in groups if group[:2] == (test, direction)]
                ),
                "机器原始点": counts["raw"],
                "DIC点": counts["dic"],
                "完整点": complete,
                "缺失DIC点": counts["missing"],
                "质量备注": note,
                "数据库层级": "迁移学习或FEA标定",
                "泄漏分组建议": (
                    f"doi|PCF20|{test}|{direction_key}|specimen_id"
                ),
            }
        )
        test_details.append(
            {
                "试验": test,
                "标准": standard,
                "方向": direction,
                "独立试样数": specimen_count,
                "机器原始点": counts["raw"],
                "DIC点": counts["dic"],
                "完整点": complete,
                "缺失点": counts["missing"],
            }
        )

    totals = {
        "specimens": len(groups),
        "raw": sum(row["机器原始点"] for row in curve_rows),
        "dic": sum(row["DIC点"] for row in curve_rows),
        "complete": sum(row["完整点"] for row in curve_rows),
        "missing": sum(row["缺失DIC点"] for row in curve_rows),
    }
    if totals != {
        "specimens": 37,
        "raw": 42_051,
        "dic": 12_365,
        "complete": 54_357,
        "missing": 59,
    }:
        raise AuditBlocked(f"MaterialsCloud科学计数发生变化：{totals}")
    if len(csv_paths) != 74 or len(image_paths) != 79:
        raise AuditBlocked("MaterialsCloud CSV或图像数发生变化")
    calibration_count = len(list(puf_root.rglob("*.txt")))
    if calibration_count != 4:
        raise AuditBlocked("MaterialsCloud标定TXT数发生变化")

    curve_rows.append(
        {
            "试验": "合计",
            "标准": "",
            "方向": "",
            "独立试样数": totals["specimens"],
            "机器原始曲线数": totals["specimens"],
            "DIC曲线数": totals["specimens"],
            "机器原始点": totals["raw"],
            "DIC点": totals["dic"],
            "完整点": totals["complete"],
            "缺失DIC点": totals["missing"],
            "质量备注": "74个CSV；机器曲线、DIC和图像已按37个物理试样同组",
            "数据库层级": "迁移学习或FEA标定",
            "泄漏分组建议": "材料泛化时整个DOI同组",
        }
    )

    puf_entry = official_entries["PUF.zip"]
    readme_entry = official_entries["README.txt"]
    file_rows = [
        _asset_row(
            archive_path,
            str(puf_entry["checksum"]).split(":", 1)[-1],
            str(puf_entry["links"]["content"]),
            "ZIP CRC通过；157个文件与解压副本逐SHA256一致",
        ),
        _asset_row(
            readme_path,
            str(readme_entry["checksum"]).split(":", 1)[-1],
            str(readme_entry["links"]["content"]),
            "已下载并核验",
        ),
        _asset_row(
            metadata_path,
            "",
            str(metadata["links"]["self"]),
            "官方API快照",
        ),
        _asset_row(
            datacite_path,
            "",
            f"https://api.datacite.org/dois/{doi}",
            "官方API快照",
        ),
    ]
    write_tsv(
        base / "文件校验清单.tsv",
        file_rows,
        [
            "文件名", "字节数", "MD5", "SHA256", "官方MD5", "官方MD5匹配",
            "来源或直链", "状态",
        ],
    )
    write_tsv(
        base / "曲线审计清单.tsv",
        curve_rows,
        [
            "试验", "标准", "方向", "独立试样数", "机器原始曲线数", "DIC曲线数",
            "机器原始点", "DIC点", "完整点", "缺失DIC点", "质量备注",
            "数据库层级", "泄漏分组建议",
        ],
    )

    creators = [
        creator["person_or_org"]["name"]
        for creator in metadata["metadata"]["creators"]
    ]
    summary = {
        "审计版本": AUDIT_VERSION,
        "审计日期": AUDIT_DATE,
        "来源": {
            "仓储": "Materials Cloud Archive",
            "记录ID": metadata["id"],
            "DOI": "10.24435/materialscloud:VF-RY",
            "版本": f"v{metadata['versions']['index']}",
            "发布日期": metadata["metadata"]["publication_date"],
            "题名": metadata["metadata"]["title"],
            "作者": creators,
            "许可证": {
                "名称": rights[0]["title"]["en"],
                "SPDX": "CC-BY-4.0",
                "链接": rights[0]["props"]["url"],
            },
        },
        "下载范围": {
            "官方记录文件数": metadata["files"]["count"],
            "官方记录总字节数": metadata["files"]["total_bytes"],
            "已下载": ["PUF.zip", "README.txt"],
            "未下载": {
                "文件": "SFFE.zip",
                "字节数": official_entries["SFFE.zip"]["size"],
                "原因": "短纤维填充环氧不属于本轮聚氨酯子集",
            },
        },
        "压缩包与解压审计": zip_audit,
        "材料身份与实验单位": {
            "仓储可核实材料": "Fourth-generation Sawbones PCF20 rigid cellular polyurethane foam，标称密度20 lb/ft³",
            "独立样本定义": "一个试样文件夹是一件物理试样；机器原始曲线、DIC曲线和试样图像同组",
            "独立试样数": totals["specimens"],
            "机器原始曲线数": totals["specimens"],
            "DIC相关曲线数": totals["specimens"],
            "同组多模态验证通过数": len(groups),
            "CSV总数": len(csv_paths),
            "图像数": len(image_paths),
            "标定TXT数": calibration_count,
            "机器原始数据点行数": totals["raw"],
            "DIC数据点行数": totals["dic"],
            "总数据点行数": totals["raw"] + totals["dic"],
            "完整数据点行数": totals["complete"],
            "缺失DIC值行数": totals["missing"],
            "不可推断项": [
                "商业牌号的完整化学配方未公开",
                "37个试样不是37种材料化学",
                "硬质多孔PUF不能直接等同致密热塑性TPU",
            ],
        },
        "试验统计": test_details,
        "质量审计": {
            "缺失值处理": "59行仅报告并隔离，不插值、不伪装为实测",
            "试样级多模态绑定": "37/37通过",
            "方向级尺寸模板图": "不计作独立物理试样",
        },
        "数据库判定": {
            "层级": "迁移学习、FEA本构标定或外部验证层",
            "是否进入TPU配方核心训练": False,
            "禁止用法": [
                "把37个试样当成37种化学配方",
                "把机器曲线、DIC曲线或图像分开随机拆分",
                "对59个缺失DIC值插值后宣称为实测",
            ],
        },
        "泄漏控制建议": {
            "试样级分组键": "dataset_doi|material_grade|test_type|direction|specimen_id",
            "材料泛化分组键": "dataset_doi|material_grade|assumed_manufacturing_batch",
            "规则": "同一试样的机器曲线、DIC和全部试样图像必须在同一折",
        },
    }
    write_json(base / "内容审计摘要.json", summary)
    return {
        "独立物理试样": totals["specimens"],
        "机器曲线": totals["specimens"],
        "DIC曲线": totals["specimens"],
        "数据行": totals["raw"] + totals["dic"],
        "DIC缺失行": totals["missing"],
    }


def _count_finite_csv(rows: list[list[str]]) -> int:
    return sum(
        _finite_number(value) is not None for row in rows for value in row
    )


def _drop_channel_counts(rows: list[list[str]]) -> dict[str, int]:
    if len(rows) < 3 or len(rows[0]) % 2:
        raise AuditBlocked("ScienceDB跌落CSV不是成对的时间-加速度列")
    width = len(rows[0])
    if len(rows[1]) != width:
        raise AuditBlocked("ScienceDB跌落CSV双层表头列数不一致")
    result = {
        "simulation_curves": 0,
        "experiment_curves": 0,
        "simulation_finite": 0,
        "experiment_finite": 0,
    }
    kinds: list[str] = []
    for column in range(0, width, 2):
        label = " ".join(
            value.strip().lower()
            for value in rows[1][column : column + 2]
            if value.strip()
        )
        if "simulation" in label:
            kind = "simulation"
        elif "experiment" in label:
            kind = "experiment"
        else:
            raise AuditBlocked(f"ScienceDB跌落通道无法分类：{label!r}")
        kinds.append(kind)
        result[f"{kind}_curves"] += 1
    for row in rows[2:]:
        padded = row + [""] * (width - len(row))
        for column, kind in zip(range(0, width, 2), kinds):
            result[f"{kind}_finite"] += sum(
                _finite_number(value) is not None
                for value in padded[column : column + 2]
            )
    return result


def audit_science_db() -> dict[str, object]:
    base = DATA_ROOT / SCIENCE_DB
    metadata_path = base / "官方元数据.json"
    page_path = base / "官方数据集页面.html"
    raw_root = base / "原始文件"
    require_file(metadata_path)
    require_file(page_path)
    require_directory(raw_root)
    metadata = load_json(metadata_path)
    if str(metadata["@id"]).casefold() != (
        "https://doi.org/10.57760/sciencedb.j00189.00045"
    ).casefold():
        raise AuditBlocked("ScienceDB DOI发生变化")
    if metadata["license"] != "https://creativecommons.org/licenses/by/4.0/":
        raise AuditBlocked("ScienceDB许可不再是CC BY 4.0")

    distribution = metadata["distribution"]
    if not isinstance(distribution, list) or len(distribution) != 24:
        raise AuditBlocked("ScienceDB官方文件数不再是24")
    raw_files = {path.name: path for path in raw_root.iterdir() if path.is_file()}
    official_names = {str(item["name"]) for item in distribution}
    if set(raw_files) != official_names:
        raise AuditBlocked(
            f"ScienceDB本地与官方文件集合不一致："
            f"缺失={sorted(official_names - set(raw_files))}，"
            f"多余={sorted(set(raw_files) - official_names)}"
        )

    file_rows: list[dict[str, object]] = []
    total_bytes = 0
    for index, item in enumerate(distribution, start=1):
        name = str(item["name"])
        path = raw_files[name]
        require_file(path)
        official_size = int(str(item["contentSize"]).split()[0])
        actual_md5 = file_hash(path, "md5")
        if path.stat().st_size != official_size or actual_md5 != item["md5"]:
            raise AuditBlocked(f"ScienceDB官方校验不一致：{name}")
        total_bytes += official_size
        file_rows.append(
            {
                "序号": index,
                "文件名": name,
                "内容类型": item["encodingFormat"],
                "来源URL": item["contentUrl"],
                "许可": "CC BY 4.0",
                "许可URL": metadata["license"],
                "官方字节数": official_size,
                "实际字节数": path.stat().st_size,
                "官方MD5": item["md5"],
                "实际MD5": actual_md5,
                "实际SHA256": file_hash(path),
                "校验状态": "通过",
            }
        )
    if total_bytes != int(metadata["size"]["value"]) or total_bytes != 7_371_062:
        raise AuditBlocked(f"ScienceDB总字节数变化：{total_bytes}")

    table_details: list[dict[str, object]] = []
    csv_matrices: dict[str, list[list[str]]] = {}
    structured_files = sorted(
        path for path in raw_files.values() if path.suffix.lower() in {".csv", ".xlsx"}
    )
    table_count = 0
    finite_total = 0
    for path in structured_files:
        if path.suffix.lower() == ".csv":
            rows = list(
                csv.reader(io.StringIO(_decode_csv(path), newline=""), delimiter=",")
            )
            finite = _count_finite_csv(rows)
            width = max((len(row) for row in rows), default=0)
            csv_matrices[path.name] = rows
            table_count += 1
            finite_total += finite
            table_details.append(
                {
                    "文件": path.name,
                    "表或工作表": path.stem,
                    "行数": len(rows),
                    "最大列数": width,
                    "有限数值": finite,
                }
            )
        else:
            with zipfile.ZipFile(path) as archive:
                if archive.testzip() is not None:
                    raise AuditBlocked(f"ScienceDB XLSX容器损坏：{path.name}")
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    finite = sum(
                        _finite_number(value) is not None
                        for row in sheet.iter_rows(values_only=True)
                        for value in row
                    )
                    table_count += 1
                    finite_total += finite
                    table_details.append(
                        {
                            "文件": path.name,
                            "表或工作表": sheet.title,
                            "行数": sheet.max_row,
                            "最大列数": sheet.max_column,
                            "有限数值": finite,
                        }
                    )
            finally:
                workbook.close()

    if len(structured_files) != 10 or table_count != 10 or finite_total != 183_689:
        raise AuditBlocked(
            "ScienceDB结构化计数变化："
            f"files={len(structured_files)}, tables={table_count}, finite={finite_total}"
        )

    image_files = sorted(
        path
        for path in raw_files.values()
        if path.suffix.lower() in {".tif", ".tiff", ".jpg", ".jpeg", ".png"}
    )
    for path in image_files:
        with Image.open(path) as image:
            image.verify()
    if len(image_files) != 14:
        raise AuditBlocked(f"ScienceDB图像数变化：{len(image_files)}")

    dma_files = ["Figure 7a-20240329.csv", "Figure 7b-20240329.csv"]
    dma_finite = sum(_count_finite_csv(csv_matrices[name]) for name in dma_files)
    dma_response_series = 0
    for name in dma_files:
        width = len(csv_matrices[name][0])
        if name.startswith("Figure 7a"):
            if width != 9:
                raise AuditBlocked("ScienceDB DMA E'/E''表列数变化")
            dma_response_series += (width // 3) * 2
        else:
            if width != 6:
                raise AuditBlocked("ScienceDB DMA tanδ表列数变化")
            dma_response_series += width // 2

    shpb_mapping = {
        "Figure 8a-20290330.xlsx": "400M",
        "Figure 8b-20290330.csv": "600M",
        "Figure 8c-20290330.csv": "800M",
    }
    shpb_details = [
        detail for detail in table_details if detail["文件"] in shpb_mapping
    ]
    shpb_sensor_curves = 0
    for detail in shpb_details:
        columns = int(detail["最大列数"])
        if columns != 12:
            raise AuditBlocked(f"ScienceDB SHPB列数变化：{detail}")
        shpb_sensor_curves += columns // 2
    shpb_finite = sum(int(detail["有限数值"]) for detail in shpb_details)

    drop_totals = Counter()
    drop_files = sorted(name for name in csv_matrices if name.startswith("Figure 10"))
    for name in drop_files:
        drop_totals.update(_drop_channel_counts(csv_matrices[name]))
    expected_drop = {
        "simulation_curves": 16,
        "experiment_curves": 11,
        "simulation_finite": 118_786,
        "experiment_finite": 4_318,
    }
    if dict(drop_totals) != expected_drop:
        raise AuditBlocked(f"ScienceDB跌落实验/仿真计数变化：{dict(drop_totals)}")

    densities = ["400M", "600M", "800M"]
    summary = {
        "审计版本": AUDIT_VERSION,
        "审计日期": AUDIT_DATE,
        # 保留本来源既有摘要的顶层计数字段，避免下游读取方因集中脚本化而
        # 改读第二套路径；更细的通道独立性证据记录在后续嵌套字段中。
        "数据集DOI": "10.57760/sciencedb.j00189.00045",
        "许可": "CC BY 4.0",
        "官方文件数": len(distribution),
        "下载文件数": len(raw_files),
        "下载总字节数": total_bytes,
        "结构化文件数": len(structured_files),
        "结构化表或工作表数": table_count,
        "有限数值单元格数": finite_total,
        "真实材料或配方数": len(densities),
        "材料标识": densities,
        "表观密度_kg_m3": [400, 600, 800],
        "动态测试设计": {
            "DMA": "3种密度；剪切模式10 Hz；-100至60 °C；3 °C/min；E′、E″、tanδ",
            "轻气炮_SHPB": "3种密度×3种冲击速度×2传感器通道；9个物理条件、18条通道曲线",
            "跌落": "600 kg/m3防护件的3/5/10 m设计；实验与仿真通道分别计数",
        },
        "定位": "微孔聚氨酯弹性体动态冲击与多保真辅助层；不是线性TPU配方核心层",
        "源数据问题": [
            "部分CSV单位或中文表头存在源文件乱码，字段需按论文图注和方法统一"
        ],
        "来源": {
            "数据集DOI": "10.57760/sciencedb.j00189.00045",
            "题名": metadata["name"],
            "版本": metadata["version"],
            "作者": [person["name"] for person in metadata["creator"]],
            "许可": "CC BY 4.0",
            "许可URL": metadata["license"],
        },
        "下载与完整性": {
            "官方文件数": len(distribution),
            "下载文件数": len(raw_files),
            "下载总字节数": total_bytes,
            "官方大小和MD5通过数": len(file_rows),
            "可解码图像数": len(image_files),
        },
        "结构化复算": {
            "结构化文件数": len(structured_files),
            "结构化表或工作表数": table_count,
            "有限数值单元格数": finite_total,
            "逐表统计": table_details,
        },
        "材料与科学计数": {
            "真实材料或密度级数": len(densities),
            "材料标识": densities,
            "表观密度_kg_m3": [400, 600, 800],
            "DMA": {
                "密度条件数": 3,
                "响应序列数": dma_response_series,
                "有限数值数": dma_finite,
                "协议": "剪切模式10 Hz；-100至60 °C；3 °C/min；E′、E″、tanδ",
                "独立性说明": "响应通道不是新增材料或物理试样",
            },
            "SHPB": {
                "密度条件数": 3,
                "冲击速度条件数": 3,
                "物理密度速度条件数": 9,
                "每条件传感器通道数": 2,
                "传感器应力时间曲线数": shpb_sensor_curves,
                "有限数值数": shpb_finite,
                "文件到密度映射": shpb_mapping,
                "独立性说明": "双传感器通道嵌套在同一冲击条件，18条通道曲线不是18个独立物理试样",
            },
            "跌落实验与仿真": {
                "数据表数": len(drop_files),
                "仓储说明高度_m": [3, 5, 10],
                "实验测量通道曲线数": drop_totals["experiment_curves"],
                "仿真通道曲线数": drop_totals["simulation_curves"],
                "实验通道有限数值数": drop_totals["experiment_finite"],
                "仿真通道有限数值数": drop_totals["simulation_finite"],
                "独立性说明": "仿真曲线和同一跌落的多测点通道均不得冒充独立物理试样",
            },
        },
        "质量与准入": {
            "定位": "微孔聚氨酯弹性体动态冲击与多保真辅助层；不是线性TPU配方核心层",
            "源数据问题": "部分CSV单位或中文表头存在源文件乱码，字段需按论文图注和方法统一",
            "默认分组键": "dataset_doi|density_grade|test_mode|impact_condition",
            "禁止计数": [
                "把DMA响应通道当独立材料",
                "把SHPB双传感器通道当独立物理试样",
                "把跌落仿真或多测点曲线当新增物理试样",
            ],
        },
    }
    write_tsv(
        base / "文件校验清单.tsv",
        file_rows,
        [
            "序号", "文件名", "内容类型", "来源URL", "许可", "许可URL",
            "官方字节数", "实际字节数", "官方MD5", "实际MD5", "实际SHA256",
            "校验状态",
        ],
    )
    write_json(base / "内容审计摘要.json", summary)
    return {
        "原始文件": len(raw_files),
        "密度级": len(densities),
        "结构化表": table_count,
        "有限数值": finite_total,
        "SHPB物理条件": 9,
        "SHPB传感器曲线": shpb_sensor_curves,
        "跌落实验曲线": drop_totals["experiment_curves"],
        "跌落仿真曲线": drop_totals["simulation_curves"],
    }


def _texas_sem_key(path: Path, data_root: Path) -> tuple[str, str, int]:
    relative = path.relative_to(data_root)
    condition = relative.parts[0]
    match = re.search(r"(\d{2}_\d{2}_\d{2})_(\d+)\.jpg$", path.name, flags=re.I)
    if match is None:
        raise AuditBlocked(f"Texas SEM文件名无法解析：{relative}")
    token = match.group(2)
    fiber_number = int(token[:-4]) if len(token) > 3 else int(token)
    return condition, match.group(1), fiber_number


def audit_texas() -> dict[str, object]:
    base = DATA_ROOT / TEXAS
    archive_path = base / "Data.zip"
    readme_path = base / "README.docx"
    metadata_path = base / "官方Dataverse元数据.json"
    datacite_path = base / "官方DataCite元数据.json"
    for path in (archive_path, readme_path, metadata_path, datacite_path):
        require_file(path)
    verify_docx(readme_path)

    metadata = load_json(metadata_path)
    datacite = load_json(datacite_path)
    doi = str(metadata["data"]["latestVersion"]["datasetPersistentId"]).split(":", 1)[-1]
    datacite_doi = str(datacite["data"]["attributes"]["doi"])
    if doi.casefold() != "10.18738/t8/zyq5z1" or datacite_doi.casefold() != doi.casefold():
        raise AuditBlocked("Texas DOI元数据不一致")
    version = metadata["data"]["latestVersion"]
    license_info = version["license"]
    if license_info["rightsIdentifier"] != "CC0-1.0":
        raise AuditBlocked("Texas许可不再是CC0 1.0")

    zip_audit = audit_zip_mirror(archive_path, base / "解压内容")
    if zip_audit["ZIP条目数"] != 143 or zip_audit["文件数"] != 123:
        raise AuditBlocked(f"Texas Data.zip固定内容发生变化：{zip_audit}")

    data_root = base / "解压内容" / "Data"
    require_directory(data_root)
    conditions = ("Dry", "Soaked", "Submerged")
    fiber_details: list[dict[str, object]] = []
    fiber_keys: dict[tuple[str, str, int], dict[str, object]] = {}
    condition_stats: dict[str, dict[str, object]] = {}
    for condition in conditions:
        csv_paths = sorted((data_root / condition).glob("*.csv"))
        batches: Counter[str] = Counter()
        diameters: list[float] = []
        temperatures: list[float] = []
        point_rows = finite_values = segment_count = 0
        for path in csv_paths:
            match = re.fullmatch(
                r"(?P<material>PCU85)_(?P<date>\d{2}_\d{2}_\d{2})_"
                r"(?P<fiber>\d{3})Data\.csv",
                path.name,
            )
            if match is None:
                raise AuditBlocked(f"Texas机械CSV文件名无法解析：{path.name}")
            date_batch = match.group("date")
            fiber_number = int(match.group("fiber"))
            key = (condition, date_batch, fiber_number)
            if key in fiber_keys:
                raise AuditBlocked(f"Texas重复fiber_csv_id：{path.name}")
            batches[date_batch] += 1

            rows = list(
                csv.reader(io.StringIO(_decode_csv(path), newline=""), delimiter=",")
            )
            if not rows or len(rows[0]) != 17 or rows[0][0] != "Set Name":
                raise AuditBlocked(f"Texas CSV表头变化：{path.name}")
            data_rows = rows[1:]
            if any(len(row) != 17 for row in data_rows):
                raise AuditBlocked(f"Texas CSV存在非17列数据行：{path.name}")
            previous: tuple[str, str] | None = None
            file_segments = 0
            file_finite = 0
            file_diameters: list[float] = []
            file_temperatures: list[float] = []
            for row in data_rows:
                segment = (row[0].strip(), row[1].strip())
                if not all(segment):
                    raise AuditBlocked(f"Texas CSV分段标签为空：{path.name}")
                if segment != previous:
                    file_segments += 1
                    previous = segment
                file_finite += sum(_finite_number(value) is not None for value in row)
                diameter = _finite_number(row[12])
                temperature = _finite_number(row[10])
                if diameter is not None:
                    file_diameters.append(diameter)
                if temperature is not None:
                    file_temperatures.append(temperature)
            if len(file_diameters) != 1 or not file_temperatures:
                raise AuditBlocked(f"Texas直径或温度字段计数异常：{path.name}")
            if file_segments != len({(row[0], row[1]) for row in data_rows}):
                raise AuditBlocked(f"Texas同一曲线分段非连续出现：{path.name}")

            detail = {
                "fiber_csv_id": path.stem,
                "material_code": match.group("material"),
                "hydration_condition": condition,
                "test_date_batch": date_batch,
                "fiber_number": fiber_number,
                "机械数据点行数": len(data_rows),
                "有限数值单元格数": file_finite,
                "加载或恢复曲线段数": file_segments,
                "直径_um": file_diameters[0],
                "温度最小_C": min(file_temperatures),
                "温度最大_C": max(file_temperatures),
                "SEM图像数": 0,
            }
            fiber_details.append(detail)
            fiber_keys[key] = detail
            point_rows += len(data_rows)
            finite_values += file_finite
            segment_count += file_segments
            diameters.extend(file_diameters)
            temperatures.extend(file_temperatures)

        condition_stats[condition] = {
            "独立纤维数": len(csv_paths),
            "机械数据点行数": point_rows,
            "有限数值单元格数": finite_values,
            "曲线段数": segment_count,
            "直径样本数": len(diameters),
            "直径最小_um": min(diameters),
            "直径均值_um": round(sum(diameters) / len(diameters), 4),
            "直径最大_um": max(diameters),
            "温度最小_C": min(temperatures),
            "温度最大_C": max(temperatures),
            "测试日期批次": dict(sorted(batches.items())),
        }

    image_paths = sorted(data_root.rglob("*.jpg"))
    unmatched_sem: Counter[tuple[str, str, int]] = Counter()
    for path in image_paths:
        require_file(path)
        with Image.open(path) as image:
            image.verify()
        key = _texas_sem_key(path, data_root)
        if key in fiber_keys:
            fiber_keys[key]["SEM图像数"] = int(fiber_keys[key]["SEM图像数"]) + 1
        else:
            unmatched_sem[key] += 1

    fibers_without_sem = [
        detail["fiber_csv_id"]
        for detail in fiber_details
        if detail["SEM图像数"] == 0
    ]
    totals = {
        "fibers": len(fiber_details),
        "rows": sum(int(stats["机械数据点行数"]) for stats in condition_stats.values()),
        "finite": sum(
            int(stats["有限数值单元格数"]) for stats in condition_stats.values()
        ),
        "segments": sum(int(stats["曲线段数"]) for stats in condition_stats.values()),
        "images": len(image_paths),
        "mapped_images": sum(int(detail["SEM图像数"]) for detail in fiber_details),
    }
    if totals != {
        "fibers": 38,
        "rows": 53_846,
        "finite": 485_108,
        "segments": 646,
        "images": 85,
        "mapped_images": 83,
    }:
        raise AuditBlocked(f"Texas科学计数发生变化：{totals}")
    expected_conditions = {
        "Dry": (15, 21_255, 191_490, 255),
        "Soaked": (11, 15_587, 140_426, 187),
        "Submerged": (12, 17_004, 153_192, 204),
    }
    for condition, expected in expected_conditions.items():
        stats = condition_stats[condition]
        actual = (
            stats["独立纤维数"],
            stats["机械数据点行数"],
            stats["有限数值单元格数"],
            stats["曲线段数"],
        )
        if actual != expected:
            raise AuditBlocked(f"Texas {condition}计数变化：{actual}")
    if fibers_without_sem:
        raise AuditBlocked(f"Texas存在无SEM的机械纤维：{fibers_without_sem}")
    unmatched_sem_rows = [
        {
            "hydration_condition": condition,
            "test_date_batch": date,
            "fiber_number": fiber,
            "SEM图像数": count,
        }
        for (condition, date, fiber), count in sorted(unmatched_sem.items())
    ]
    if unmatched_sem_rows != [
        {
            "hydration_condition": "Submerged",
            "test_date_batch": "09_23_24",
            "fiber_number": 2,
            "SEM图像数": 2,
        }
    ]:
        raise AuditBlocked(f"Texas无机械CSV的SEM映射变化：{unmatched_sem_rows}")

    curve_rows: list[dict[str, object]] = []
    for condition in conditions:
        stats = condition_stats[condition]
        batches = stats["测试日期批次"]
        curve_rows.append(
            {
                "条件": condition,
                "独立纤维数": stats["独立纤维数"],
                "机械CSV数": stats["独立纤维数"],
                "机械点行数": stats["机械数据点行数"],
                "有限数值单元格数": stats["有限数值单元格数"],
                "加载恢复曲线段数": stats["曲线段数"],
                "直径样本数": stats["直径样本数"],
                "直径最小_um": f"{float(stats['直径最小_um']):.3f}",
                "直径均值_um": f"{float(stats['直径均值_um']):.4f}",
                "直径最大_um": f"{float(stats['直径最大_um']):.3f}",
                "温度最小_C": f"{float(stats['温度最小_C']):.1f}",
                "温度最大_C": f"{float(stats['温度最大_C']):.1f}",
                "测试日期批次及样本数": ";".join(
                    f"{batch}:{count}" for batch, count in batches.items()
                ),
                "数据库层级": "迁移学习或外部验证",
                "泄漏分组建议": (
                    f"dataset_doi|material_code|{condition}|test_date_batch；"
                    "同一fiber_csv_id全部阶段与SEM绑定"
                ),
            }
        )
    curve_rows.append(
        {
            "条件": "合计",
            "独立纤维数": totals["fibers"],
            "机械CSV数": totals["fibers"],
            "机械点行数": totals["rows"],
            "有限数值单元格数": totals["finite"],
            "加载恢复曲线段数": totals["segments"],
            "直径样本数": totals["fibers"],
            "直径最小_um": "",
            "直径均值_um": "",
            "直径最大_um": "",
            "温度最小_C": "",
            "温度最大_C": "",
            "测试日期批次及样本数": "",
            "数据库层级": "迁移学习或外部验证",
            "泄漏分组建议": "禁止按点或646个曲线段随机拆分",
        }
    )

    official_files = {item["label"]: item for item in version["files"]}
    if set(official_files) != {"Data.zip", "README.docx"}:
        raise AuditBlocked("Texas官方文件清单发生变化")
    data_entry = official_files["Data.zip"]["dataFile"]
    readme_entry = official_files["README.docx"]["dataFile"]
    file_rows = [
        _asset_row(
            archive_path,
            data_entry["checksum"]["value"],
            f"https://dataverse.tdl.org/api/access/datafile/{data_entry['id']}",
            "ZIP CRC通过；123个文件与解压副本逐SHA256一致",
        ),
        _asset_row(
            readme_path,
            readme_entry["checksum"]["value"],
            f"https://dataverse.tdl.org/api/access/datafile/{readme_entry['id']}",
            "已下载、MD5匹配且OOXML完整",
        ),
        _asset_row(
            metadata_path,
            "",
            "https://dataverse.tdl.org/api/datasets/:persistentId/?persistentId=doi:10.18738/T8/ZYQ5Z1",
            "官方API快照",
        ),
        _asset_row(
            datacite_path,
            "",
            "https://api.datacite.org/dois/10.18738/T8/ZYQ5Z1",
            "官方API快照",
        ),
    ]
    write_tsv(
        base / "文件校验清单.tsv",
        file_rows,
        [
            "文件名", "字节数", "MD5", "SHA256", "官方MD5", "官方MD5匹配",
            "来源或直链", "状态",
        ],
    )
    write_tsv(
        base / "曲线审计清单.tsv",
        curve_rows,
        [
            "条件", "独立纤维数", "机械CSV数", "机械点行数", "有限数值单元格数",
            "加载恢复曲线段数", "直径样本数", "直径最小_um", "直径均值_um",
            "直径最大_um", "温度最小_C", "温度最大_C", "测试日期批次及样本数",
            "数据库层级", "泄漏分组建议",
        ],
    )

    summary_condition_stats = [
        {"条件": condition, **condition_stats[condition]} for condition in conditions
    ]
    summary = {
        "审计版本": AUDIT_VERSION,
        "审计日期": AUDIT_DATE,
        "来源": {
            "仓储": metadata["data"]["publisher"],
            "DOI": "10.18738/T8/ZYQ5Z1",
            "版本": f"V{version['versionNumber']}.{version['versionMinorNumber']}",
            "发布日期": metadata["data"]["publicationDate"],
            "题名": datacite["data"]["attributes"]["titles"][0]["title"],
            "许可证": {
                "名称": license_info["name"],
                "SPDX": license_info["rightsIdentifier"],
                "链接": license_info["uri"],
            },
        },
        "压缩包与解压审计": zip_audit,
        "材料身份与实验单位": {
            "仓储可核实材料": "单根电纺聚氨酯纤维；机械CSV材料代码PCU85",
            "独立样本定义": "每个机械CSV对应一根实际测试纤维；文件内阶段和加载/恢复段嵌套于fiber_csv_id",
            "观测身份字段": "fiber_csv_id",
            "默认拆分键": "dataset_doi|material_code|hydration_condition|test_date_batch",
            "独立纤维数": totals["fibers"],
            "机械CSV数": totals["fibers"],
            "SEM图像数": totals["images"],
            "映射到机械fiber_csv_id的SEM图像数": totals["mapped_images"],
            "无机械CSV匹配的SEM图像": unmatched_sem_rows,
            "机械数据点行数": totals["rows"],
            "有限数值单元格数": totals["finite"],
            "加载或恢复曲线段数": totals["segments"],
        },
        "条件统计": summary_condition_stats,
        "纤维观测": sorted(fiber_details, key=lambda item: str(item["fiber_csv_id"])),
        "质量审计": {
            "CSV列数一致": True,
            "每个CSV列数": 17,
            "畸形数据行数": 0,
            "全部机械fiber_csv_id至少有一张SEM": True,
            "SEM映射异常": "Submerged 09_23_24 fiber 2有2张SEM，但无对应机械CSV；只作SEM辅助证据，不增加机械样本数",
            "独立性说明": "646个曲线段与85张SEM都不是新增材料或独立力学样本",
        },
        "数据库判定": {
            "层级": "迁移学习或外部验证层",
            "是否进入块体TPU配方核心训练": False,
            "禁止用法": [
                "把646条曲线段当成646个独立材料样本",
                "把85张SEM当成85个独立力学样本",
                "把单纤维结果直接当块体TPU性能",
            ],
        },
        "泄漏控制建议": {
            "观测身份键": "dataset_doi|material_code|hydration_condition|fiber_csv_id",
            "默认拆分键": "dataset_doi|material_code|hydration_condition|test_date_batch",
            "规则": "同一fiber_csv_id的全部应变阶段、加载/恢复段和已匹配SEM必须同折",
        },
    }
    write_json(base / "内容审计摘要.json", summary)
    return {
        "独立纤维": totals["fibers"],
        "Dry": condition_stats["Dry"]["独立纤维数"],
        "Soaked": condition_stats["Soaked"]["独立纤维数"],
        "Submerged": condition_stats["Submerged"]["独立纤维数"],
        "机械数据行": totals["rows"],
        "加载恢复分段": totals["segments"],
        "SEM图像": totals["images"],
    }


def validate_outputs() -> list[dict[str, object]]:
    expected_rows = {
        (MATERIALS_CLOUD, "文件校验清单.tsv"): 4,
        (MATERIALS_CLOUD, "曲线审计清单.tsv"): 7,
        (SCIENCE_DB, "文件校验清单.tsv"): 24,
        (TEXAS, "文件校验清单.tsv"): 4,
        (TEXAS, "曲线审计清单.tsv"): 4,
    }
    result: list[dict[str, object]] = []
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        for filename in OUTPUT_NAMES_BY_SOURCE[source]:
            path = base / filename
            require_file(path)
            row_count: int | str = "JSON"
            if path.suffix == ".json":
                payload = load_json(path)
                if payload.get("审计版本") != AUDIT_VERSION:
                    raise AuditBlocked(f"审计版本错误：{path}")
            else:
                with path.open("r", encoding="utf-8", newline="") as handle:
                    rows = list(csv.DictReader(handle, delimiter="\t"))
                row_count = len(rows)
                expected = expected_rows[(source, filename)]
                if row_count != expected:
                    raise AuditBlocked(
                        f"输出TSV行数错误：{source}/{filename}={row_count}，预期={expected}"
                    )
            result.append(
                {
                    "来源": source,
                    "文件": filename,
                    "字节数": path.stat().st_size,
                    "数据行数": row_count,
                    "SHA256": file_hash(path),
                }
            )
    return result


def main() -> int:
    before = scientific_input_snapshot()
    counts = {
        MATERIALS_CLOUD: audit_materials_cloud(),
        SCIENCE_DB: audit_science_db(),
        TEXAS: audit_texas(),
    }
    after = scientific_input_snapshot()
    if before != after:
        changed = sorted(set(before) ^ set(after))
        changed.extend(
            path
            for path in sorted(set(before) & set(after))
            if before[path] != after[path]
        )
        raise AuditBlocked(f"科学输入在审计期间发生变化：{changed}")

    result = {
        "审计日期": AUDIT_DATE,
        "科学输入哈希不变": True,
        "科学输入清单": snapshot_by_source(before),
        "科学计数": counts,
        "输出": validate_outputs(),
    }
    print(json.dumps(result, ensure_ascii=False, sort_keys=True, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (AuditBlocked, KeyError, TypeError, ValueError, zipfile.BadZipFile) as exc:
        print(f"审计阻断：{exc}", file=sys.stderr)
        raise SystemExit(2) from exc
