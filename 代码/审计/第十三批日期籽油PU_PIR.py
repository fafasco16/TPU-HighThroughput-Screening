"""审计日期籽油基低密度刚性 PU-PIR 的 Mendeley Data v3 原始数据。

数据集 DOI 为 10.17632/xs78ch5jb7.3，许可为 CC BY 4.0。本模块只读
官方附件，RAR 中的 XLSX 通过 7-Zip 标准输出在内存中读取，不解压到磁盘。
所有曲线点保留来源原值；重复呈现只登记一次，标签冲突则降为条件参考。

机械工作簿没有轴标题，支持信息也没有机械测试协议。因此机械曲线虽然来源
可靠、样品标签明确，但轴单位和协议在本批中不作猜测，进入 Gold-E 条件层。
曲线点始终是配方内观测，不是独立材料样本，也不生成训练权重或数据划分。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第十三批实验_日期籽油PU-PIR"
)

OUTPUT_LONG_TABLE = SOURCE_DIR / "Gold_E_实验观测长表.tsv"
OUTPUT_AUDIT = SOURCE_DIR / "内容审计摘要.json"
OUTPUT_CHECKSUMS = SOURCE_DIR / "文件校验清单.tsv"
OUTPUT_README = SOURCE_DIR / "来源说明.md"

SOURCE_ID = "source_mendeley_xs78ch5jb7_v3"
SOURCE_DOI = "10.17632/xs78ch5jb7.3"
ARTICLE_DOI = "10.1016/j.indcrop.2024.120152"
SOURCE_LICENSE = "CC BY 4.0"
ARTICLE_LICENSE = "CC BY-NC 4.0"
DATASET_URL = "https://data.mendeley.com/datasets/xs78ch5jb7/3"
METADATA_URL = (
    "https://data.mendeley.com/public-api/datasets/xs78ch5jb7?version=3"
)
FILES_URL = (
    "https://data.mendeley.com/public-api/datasets/xs78ch5jb7/files"
    "?folder_id=root&version=3"
)
VERSIONS_URL = "https://data.mendeley.com/public-api/datasets/xs78ch5jb7/versions"
ARTICLE_URL = (
    "https://www.sciencedirect.com/science/article/pii/S0926669024021290"
)
CITATION_KEYS = f"doi:{SOURCE_DOI};doi:{ARTICLE_DOI}"

AUDIT_VERSION = "batch13-date-seed-oil-rigid-pu-pir-v1"
FAMILY_KEY = "family_date_seed_oil_rigid_pu_pir_2025"
FAMILY_LEAKAGE_GROUP = "family_date_seed_oil_rigid_pu_pir_2025"
FORMULATIONS = {
    "S1": 0,
    "S2": 30,
    "S3": 50,
    "S4": 70,
    "S5": 100,
}

OFFICIAL_ATTACHMENTS: dict[str, tuple[int, str]] = {
    "Supplemental data 1.xlsx": (
        155_964,
        "62c1a42c2f0a7948b88b5a2f0e19a9f57dc01fdba132ccc941de66a3ceb2e19e",
    ),
    "Supplemental data 2.rar": (
        14_231,
        "1c50a33431f289f2dc4751bb36f634adf7eacbbcfbee8c144314304d6298a1aa",
    ),
    "Supplemental data 3.xlsx": (
        199_570,
        "d5f7c380d5e7316b6b3841f268676cd4f3ef7e8b5577e3130bd8591dc33b0034",
    ),
    "Supplemental data 4.xlsx": (
        265_470,
        "5f5ad0c4f510962deada8f96360480189369183828129f6fc3d8fb49be04c700",
    ),
    "Supplemental data 5.xlsx": (
        257_708,
        "627374459dabeee633cbe401e22799f963e17d6b4558cb2ba4a189e20ca486e2",
    ),
    "Supplemental data 6.xlsx": (
        902_298,
        "f3cf0a7058ffec498a9321ba2b5a93ccd13fe394bce40becac62e929e4b49ba2",
    ),
    "Supporting Information.pdf": (
        1_433_624,
        "b1b21246555a02b1620356a748a908d936d1e02a1701acbc86c38f26799d18a3",
    ),
}

LOCAL_METADATA_SNAPSHOTS: dict[str, tuple[int, str]] = {
    "Mendeley_元数据_v3.json": (
        3_001,
        "fe3ae38dbc9770b2b1f2279763cfd550a4b9292673d1915ca008b553de55d7d2",
    ),
    "Mendeley_文件清单_v3.json": (
        5_287,
        "1adc9b263048c8782280ee345a6516db2f3651da6628382bd5c818af277cfbd2",
    ),
    "Mendeley_版本清单.json": (
        178,
        "7ad0760bceb197bbc8ccbc1a0cd754ee0b10d87bc7590e012c03b4f1f8812667",
    ),
    "论文Crossref元数据.json": (
        21_891,
        "65af1c1d7eafa24a522ee932e066fbc018714bf1d1135ae692fca9d8844a680a",
    ),
}

EXPECTED_WORKBOOK_LAYOUT = {
    "Supplemental data 1.xlsx": [
        ("FT-IR data of Date seed oil", 1867, 2),
        ("FT-IR data of DSO polyol ", 1867, 2),
        ("Monitoring using FT-IR", 1867, 5),
    ],
    "Supplemental data 2.rar::Supplemental data 2.xlsx": [
        ("HNMR OF Date seed oil", 40, 14),
        ("CNMR OF Date seed oil", 52, 15),
        ("HNMR OF DSOpolyol", 39, 15),
        ("CNMR OF DSOpolyol", 54, 15),
    ],
    "Supplemental data 3.xlsx": [
        ("TGA-DTG-DTA- DSC OF DSOpolyol", 4341, 10),
        ("Sheet2", 1, 1),
        ("Sheet3", 1, 1),
    ],
    "Supplemental data 4.xlsx": [
        ("FT-IR of S1 (0%DSOpolyol)", 1867, 2),
        ("FT-IR OF S2 (30%DSOpolyol)", 1867, 2),
        ("FT-IR OF S3 (50%DSOpolyol)", 1867, 2),
        ("FT-IR OF S4 (70%DSOpolyol)", 1867, 2),
        ("FT-IR OF S5 (100%DSOpolyol)", 1867, 13),
    ],
    "Supplemental data 5.xlsx": [
        ("S1 (0% DSO polyol)", 796, 3),
        ("S2 (30% DSO polyol)", 996, 5),
        ("S3 (50% DSO polyol)", 896, 5),
        ("S4 (70% DSO polyol)", 796, 5),
        ("S5 (100% DSO polyol)", 857, 3),
    ],
    "Supplemental data 6.xlsx": [
        ("Tensile stress-strain", 1585, 6),
        ("Compressive stress-strain", 14983, 6),
    ],
}

INNER_XLSX_MEMBER = "Supplemental data 2\\Supplemental data 2.xlsx"
INNER_XLSX_BYTES = 16_853
INNER_XLSX_SHA256 = (
    "e2f1d116a3f2546a01ab85c418a92769eca726b122a7c1b11ec2fd7fab11ae34"
)

RECORD_COLUMNS = (
    "source_directory",
    "source_record_id",
    "observation_id",
    "formulation_id",
    "sample_id",
    "record_kind",
    "component_name",
    "component_role",
    "property_name",
    "value",
    "unit",
    "uncertainty_value",
    "uncertainty_type",
    "condition_name",
    "condition_value",
    "condition_unit",
    "target_origin",
    "data_origin",
    "reduction_level",
    "method_or_test_protocol",
    "fidelity_level",
    "gold_admission_status",
    "mapping_status",
    "protocol_status",
    "potential_weight_ceiling",
    "current_weight_materialized",
    "training_weight",
    "split_group",
    "source_locator",
    "file_sha256",
    "license",
    "citation_keys",
    "notes",
    "curve_id",
    "point_index",
    "secondary_condition_name",
    "secondary_condition_value",
    "secondary_condition_unit",
    "auxiliary_value_name",
    "auxiliary_value",
    "auxiliary_unit",
    "sample_identity_status",
    "global_structure_family_key",
    "family_leakage_group",
    "curve_points_are_independent_samples",
    "duplicate_status",
)

FILE_COLUMNS = (
    "file",
    "role",
    "bytes",
    "sha256",
    "official_sha256",
    "official_sha256_match",
    "verification",
    "notes",
)


class AuditBlocked(RuntimeError):
    """来源身份、原件或冻结工作簿结构发生漂移。"""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _number_text(value: Any, *, context: str) -> str:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise AuditBlocked(f"非数值单元格：{context}={value!r}")
    numeric = float(value)
    if not math.isfinite(numeric):
        raise AuditBlocked(f"非有限数值：{context}={value!r}")
    return str(value) if isinstance(value, int) else repr(numeric)


def _load_json(name: str) -> Any:
    return json.loads((SOURCE_DIR / name).read_text(encoding="utf-8"))


def _seven_zip() -> str:
    executable = shutil.which("7z") or shutil.which("7za")
    if not executable:
        raise AuditBlocked("读取 RAR 需要 7z/7za，但当前环境未找到")
    return executable


def _read_inner_xlsx() -> bytes:
    archive = SOURCE_DIR / "Supplemental data 2.rar"
    completed = subprocess.run(
        [_seven_zip(), "e", "-so", str(archive), INNER_XLSX_MEMBER],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    if completed.returncode != 0:
        raise AuditBlocked(
            "RAR 内 XLSX 读取失败："
            + completed.stderr.decode("utf-8", errors="replace")[-500:]
        )
    payload = completed.stdout
    if (len(payload), _sha256_bytes(payload)) != (
        INNER_XLSX_BYTES,
        INNER_XLSX_SHA256,
    ):
        raise AuditBlocked("RAR 内 XLSX 大小或 SHA256 漂移")
    return payload


def _assert_safe_xlsx(payload: bytes | Path, label: str) -> None:
    source: Any = io.BytesIO(payload) if isinstance(payload, bytes) else payload
    with zipfile.ZipFile(source) as archive:
        if archive.testzip() is not None:
            raise AuditBlocked(f"XLSX ZIP CRC 失败：{label}")
        for name in archive.namelist():
            normalized = name.replace("\\", "/")
            if normalized.startswith("/") or "../" in f"/{normalized}":
                raise AuditBlocked(f"XLSX 含不安全成员：{label}::{name}")


def _sheet_inventory(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows.append(
            {
                "sheet": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": sum(
                    cell.value is not None
                    for row in sheet.iter_rows()
                    for cell in row
                ),
                "formula_cells": [
                    f"{cell.coordinate}:{cell.value}"
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.data_type == "f"
                ],
                "chart_count": len(sheet._charts),
                "image_count": len(sheet._images),
            }
        )
    return rows


def _verify_workbook_layout(
    label: str, workbook: Any
) -> list[dict[str, Any]]:
    actual = [
        (sheet.title, sheet.max_row, sheet.max_column)
        for sheet in workbook.worksheets
    ]
    if actual != EXPECTED_WORKBOOK_LAYOUT[label]:
        raise AuditBlocked(f"工作簿结构漂移：{label}: {actual}")
    return _sheet_inventory(workbook)


def verify_sources() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    file_rows: list[dict[str, Any]] = []
    for name, (expected_bytes, expected_hash) in {
        **OFFICIAL_ATTACHMENTS,
        **LOCAL_METADATA_SNAPSHOTS,
    }.items():
        path = SOURCE_DIR / name
        if not path.is_file():
            raise AuditBlocked(f"缺少冻结来源文件：{path}")
        actual = (path.stat().st_size, _sha256(path))
        if actual != (expected_bytes, expected_hash):
            raise AuditBlocked(
                f"来源文件漂移：{name}; bytes={actual[0]}; sha256={actual[1]}"
            )
        official_hash = OFFICIAL_ATTACHMENTS.get(name, (0, ""))[1]
        file_rows.append(
            {
                "file": name,
                "role": (
                    "official_attachment"
                    if name in OFFICIAL_ATTACHMENTS
                    else "captured_metadata_snapshot"
                ),
                "bytes": actual[0],
                "sha256": actual[1],
                "official_sha256": official_hash,
                "official_sha256_match": (
                    "true" if official_hash else "not_applicable"
                ),
                "verification": "size_and_sha256_verified",
                "notes": "",
            }
        )

    metadata = _load_json("Mendeley_元数据_v3.json")
    if metadata.get("id") != "xs78ch5jb7" or metadata.get("version") != 3:
        raise AuditBlocked("Mendeley 数据集 ID 或版本漂移")
    if metadata.get("doi", {}).get("id") != SOURCE_DOI:
        raise AuditBlocked("Mendeley DOI 漂移")
    if metadata.get("data_licence", {}).get("short_name") != SOURCE_LICENSE:
        raise AuditBlocked("Mendeley 许可漂移")
    if metadata.get("available") is not True:
        raise AuditBlocked("Mendeley v3 不再公开可用")
    contributors = [
        f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
        for row in metadata.get("contributors", [])
    ]
    if contributors != ["Samaneh Taghvaei Nia", "Mir Mohammad Alavi Nikje"]:
        raise AuditBlocked(f"贡献者漂移：{contributors}")

    versions = _load_json("Mendeley_版本清单.json")
    expected_versions = [
        {"version": 1, "publish_date": "2024-07-10", "available": True},
        {"version": 2, "publish_date": "2024-10-29", "available": True},
        {"version": 3, "publish_date": "2024-11-07", "available": True},
    ]
    if versions != expected_versions:
        raise AuditBlocked(f"Mendeley 版本历史漂移：{versions}")

    official_files = _load_json("Mendeley_文件清单_v3.json")
    if len(official_files) != len(OFFICIAL_ATTACHMENTS):
        raise AuditBlocked("Mendeley v3 附件数量漂移")
    indexed_files = {row["filename"]: row for row in official_files}
    if set(indexed_files) != set(OFFICIAL_ATTACHMENTS):
        raise AuditBlocked("Mendeley v3 附件文件名漂移")
    for name, (expected_bytes, expected_hash) in OFFICIAL_ATTACHMENTS.items():
        details = indexed_files[name]["content_details"]
        if (details["size"], details["sha256_hash"]) != (
            expected_bytes,
            expected_hash,
        ):
            raise AuditBlocked(f"Mendeley 官方附件身份漂移：{name}")

    crossref = _load_json("论文Crossref元数据.json").get("message", {})
    if crossref.get("DOI", "").lower() != ARTICLE_DOI:
        raise AuditBlocked("关联论文 DOI 漂移")
    title = " ".join((crossref.get("title") or [""])[0].split())
    expected_title = (
        "Novel bio-polyol synthesis based on date seed oil for low-density "
        "rigid polyurethane-polyisocyanurate foams"
    )
    if title != expected_title:
        raise AuditBlocked(f"关联论文标题漂移：{title}")

    workbook_inventory: dict[str, list[dict[str, Any]]] = {}
    for name in (
        "Supplemental data 1.xlsx",
        "Supplemental data 3.xlsx",
        "Supplemental data 4.xlsx",
        "Supplemental data 5.xlsx",
        "Supplemental data 6.xlsx",
    ):
        path = SOURCE_DIR / name
        _assert_safe_xlsx(path, name)
        workbook = load_workbook(path, data_only=False, read_only=False)
        workbook_inventory[name] = _verify_workbook_layout(name, workbook)

    inner_payload = _read_inner_xlsx()
    inner_label = "Supplemental data 2.rar::Supplemental data 2.xlsx"
    _assert_safe_xlsx(inner_payload, inner_label)
    inner_workbook = load_workbook(
        io.BytesIO(inner_payload), data_only=False, read_only=False
    )
    workbook_inventory[inner_label] = _verify_workbook_layout(
        inner_label, inner_workbook
    )
    formulas = [
        (label, sheet["sheet"], formula)
        for label, sheets in workbook_inventory.items()
        for sheet in sheets
        for formula in sheet["formula_cells"]
    ]
    if formulas != [
        (
            inner_label,
            "CNMR OF Date seed oil",
            "O6:=---R19",
        )
    ]:
        raise AuditBlocked(f"工作簿公式清单漂移：{formulas}")

    pdf = PdfReader(SOURCE_DIR / "Supporting Information.pdf")
    if len(pdf.pages) != 19:
        raise AuditBlocked("支持信息 PDF 页数漂移")
    page18 = " ".join((pdf.pages[17].extract_text() or "").split())
    for anchor in (
        "Table S1 Formulations for low-density RPU-PIR foams",
        "119.6616 127.9249 133.4338 138.9427 147.2060",
        "120 128 133 139 147",
    ):
        if anchor not in page18:
            raise AuditBlocked(f"支持信息表 S1 文本锚点漂移：{anchor}")

    return file_rows, {
        "metadata": metadata,
        "versions": versions,
        "official_files": official_files,
        "workbook_inventory": workbook_inventory,
        "inner_xlsx": {
            "member": INNER_XLSX_MEMBER,
            "bytes": INNER_XLSX_BYTES,
            "sha256": INNER_XLSX_SHA256,
            "read_mode": "7z_stdout_to_memory_no_disk_extraction",
        },
        "supporting_information_pages": len(pdf.pages),
        "stray_formula_excluded": {
            "source": f"{inner_label}#CNMR OF Date seed oil!O6",
            "formula": "=---R19",
            "reason": "旁栏孤立公式，不属于 NMR 峰表",
        },
    }


def _file_hash(name: str) -> str:
    return OFFICIAL_ATTACHMENTS[name][1]


def _split_group(material_id: str) -> str:
    return f"date_seed_oil_pu_pir::{material_id}"


def _base_row(
    *,
    record_id: str,
    formulation_id: str,
    sample_id: str,
    record_kind: str,
    property_name: str,
    value: Any,
    unit: str,
    source_file: str,
    source_locator: str,
    data_origin: str,
    reduction_level: str,
    method: str,
    fidelity: str,
    admission: str,
    mapping_status: str,
    protocol_status: str,
    weight_ceiling: str,
    material_id: str,
    sample_identity_status: str,
    notes: str,
    component_name: str = "",
    component_role: str = "",
    condition_name: str = "",
    condition_value: Any = "",
    condition_unit: str = "",
    curve_id: str = "",
    point_index: int | str = "",
    secondary_condition_name: str = "",
    secondary_condition_value: Any = "",
    secondary_condition_unit: str = "",
    auxiliary_value_name: str = "",
    auxiliary_value: Any = "",
    auxiliary_unit: str = "",
    duplicate_status: str = "unique_canonical_observation",
) -> dict[str, Any]:
    row = {column: "" for column in RECORD_COLUMNS}
    row.update(
        {
            "source_directory": SOURCE_DIR.name,
            "source_record_id": record_id,
            "observation_id": record_id,
            "formulation_id": formulation_id,
            "sample_id": sample_id,
            "record_kind": record_kind,
            "component_name": component_name,
            "component_role": component_role,
            "property_name": property_name,
            "value": _number_text(value, context=record_id),
            "unit": unit,
            "condition_name": condition_name,
            "condition_value": (
                _number_text(condition_value, context=f"{record_id}:condition")
                if condition_value != ""
                else ""
            ),
            "condition_unit": condition_unit,
            "target_origin": "experimental",
            "data_origin": data_origin,
            "reduction_level": reduction_level,
            "method_or_test_protocol": method,
            "fidelity_level": fidelity,
            "gold_admission_status": admission,
            "mapping_status": mapping_status,
            "protocol_status": protocol_status,
            "potential_weight_ceiling": weight_ceiling,
            "current_weight_materialized": "false",
            "training_weight": "",
            "split_group": _split_group(material_id),
            "source_locator": source_locator,
            "file_sha256": _file_hash(source_file),
            "license": SOURCE_LICENSE,
            "citation_keys": CITATION_KEYS,
            "notes": notes,
            "curve_id": curve_id,
            "point_index": point_index,
            "secondary_condition_name": secondary_condition_name,
            "secondary_condition_value": (
                _number_text(
                    secondary_condition_value,
                    context=f"{record_id}:secondary_condition",
                )
                if secondary_condition_value != ""
                else ""
            ),
            "secondary_condition_unit": secondary_condition_unit,
            "auxiliary_value_name": auxiliary_value_name,
            "auxiliary_value": (
                _number_text(auxiliary_value, context=f"{record_id}:auxiliary")
                if auxiliary_value != ""
                else ""
            ),
            "auxiliary_unit": auxiliary_unit,
            "sample_identity_status": sample_identity_status,
            "global_structure_family_key": FAMILY_KEY,
            "family_leakage_group": FAMILY_LEAKAGE_GROUP,
            "curve_points_are_independent_samples": "false",
            "duplicate_status": duplicate_status,
        }
    )
    return row


def _curve_row(
    *,
    curve_id: str,
    point_index: int,
    formulation_id: str,
    material_id: str,
    sample_identity_status: str,
    modality: str,
    value: Any,
    unit: str,
    x_name: str,
    x_value: Any,
    x_unit: str,
    source_file: str,
    source_sheet: str,
    source_row: int,
    source_column: int,
    admission: str,
    mapping_status: str,
    protocol_status: str,
    weight_ceiling: str,
    method: str,
    notes: str,
    secondary_name: str = "",
    secondary_value: Any = "",
    secondary_unit: str = "",
    auxiliary_name: str = "",
    auxiliary_value: Any = "",
    auxiliary_unit: str = "",
    duplicate_status: str = "unique_canonical_observation",
) -> dict[str, Any]:
    record_id = f"xs78ch5jb7_v3|{curve_id}|point={point_index}"
    cell = f"{get_column_letter(source_column)}{source_row}"
    return _base_row(
        record_id=record_id,
        formulation_id=formulation_id,
        sample_id=material_id,
        record_kind=f"{modality}_curve_point",
        property_name=modality,
        value=value,
        unit=unit,
        source_file=source_file,
        source_locator=f"{source_file}#sheet={source_sheet};cell={cell}",
        data_origin="source_native_experimental_curve",
        reduction_level="within_curve_point",
        method=method,
        fidelity="raw_deposited_curve_point",
        admission=admission,
        mapping_status=mapping_status,
        protocol_status=protocol_status,
        weight_ceiling=weight_ceiling,
        material_id=material_id,
        sample_identity_status=sample_identity_status,
        notes=notes,
        condition_name=x_name,
        condition_value=x_value,
        condition_unit=x_unit,
        curve_id=curve_id,
        point_index=point_index,
        secondary_condition_name=secondary_name,
        secondary_condition_value=secondary_value,
        secondary_condition_unit=secondary_unit,
        auxiliary_value_name=auxiliary_name,
        auxiliary_value=auxiliary_value,
        auxiliary_unit=auxiliary_unit,
        duplicate_status=duplicate_status,
    )


def _numeric_pairs(
    sheet: Any,
    *,
    x_column: int,
    y_column: int,
    first_row: int,
    last_row: int,
) -> Iterable[tuple[int, Any, Any]]:
    for row_index in range(first_row, last_row + 1):
        x_value = sheet.cell(row_index, x_column).value
        y_value = sheet.cell(row_index, y_column).value
        _number_text(x_value, context=f"{sheet.title}!{row_index}:x")
        _number_text(y_value, context=f"{sheet.title}!{row_index}:y")
        yield row_index, x_value, y_value


def build_ftir_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook1 = load_workbook(
        SOURCE_DIR / "Supplemental data 1.xlsx", data_only=True, read_only=False
    )
    oil, polyol, monitoring = workbook1.worksheets
    polyol_values = [
        (polyol.cell(row, 1).value, polyol.cell(row, 2).value)
        for row in range(2, 1868)
    ]
    monitor_4h_values = [
        (monitoring.cell(row, 1).value, monitoring.cell(row, 5).value)
        for row in range(2, 1868)
    ]
    if polyol_values != monitor_4h_values:
        raise AuditBlocked("DSO polyol FT-IR 与 4 h 监测重复关系漂移")

    data1_curves = (
        (
            "ftir_date_seed_oil",
            oil,
            2,
            "date_seed_oil",
            "date_seed_oil",
            "exact_material_state_label",
            "single_presentation",
        ),
        (
            "ftir_dso_polyol_monitor_4h",
            polyol,
            2,
            "dso_polyol",
            "dso_polyol",
            "exact_material_state_label",
            "canonicalized_exact_duplicate_with_monitoring_4h",
        ),
        (
            "ftir_dso_reaction_0h",
            monitoring,
            2,
            "dso_reaction_0h",
            "dso_reaction_0h",
            "reaction_time_label_exact",
            "single_presentation",
        ),
        (
            "ftir_dso_reaction_2h",
            monitoring,
            3,
            "dso_reaction_2h",
            "dso_reaction_2h",
            "reaction_time_label_exact",
            "single_presentation",
        ),
        (
            "ftir_dso_reaction_3h",
            monitoring,
            4,
            "dso_reaction_3h",
            "dso_reaction_3h",
            "reaction_time_label_exact",
            "single_presentation",
        ),
    )
    for curve_id, sheet, y_column, material_id, sample_id, identity, duplicate in data1_curves:
        for point_index, (source_row, x_value, y_value) in enumerate(
            _numeric_pairs(
                sheet,
                x_column=1,
                y_column=y_column,
                first_row=2,
                last_row=1867,
            ),
            start=1,
        ):
            rows.append(
                _curve_row(
                    curve_id=curve_id,
                    point_index=point_index,
                    formulation_id="",
                    material_id=material_id,
                    sample_identity_status=identity,
                    modality="ftir_source_native_signal",
                    value=y_value,
                    unit="source_native_signal_unit_unresolved",
                    x_name="wavenumber",
                    x_value=x_value,
                    x_unit="cm^-1",
                    source_file="Supplemental data 1.xlsx",
                    source_sheet=sheet.title,
                    source_row=source_row,
                    source_column=y_column,
                    admission="conditional_reference",
                    mapping_status="material_or_reaction_time_exact_signal_semantics_unresolved",
                    protocol_status="wavenumber_resolved_signal_axis_unlabelled",
                    weight_ceiling="0.25",
                    method="FT-IR; deposited worksheet",
                    notes="signal_axis_semantics_not_declared",
                    duplicate_status=duplicate,
                )
            )

    workbook4 = load_workbook(
        SOURCE_DIR / "Supplemental data 4.xlsx", data_only=True, read_only=False
    )
    combined = workbook4.worksheets[4]
    combined_matches: dict[str, bool] = {}
    for index, sheet in enumerate(workbook4.worksheets, start=1):
        formulation = f"S{index}"
        original = [
            (sheet.cell(row, 1).value, sheet.cell(row, 2).value)
            for row in range(2, 1868)
        ]
        aggregate = [
            (combined.cell(row, 8).value, combined.cell(row, 8 + index).value)
            for row in range(2, 1868)
        ]
        combined_matches[formulation] = original == aggregate
        if not combined_matches[formulation]:
            raise AuditBlocked(f"泡沫 FT-IR 汇总表与独立表不一致：{formulation}")
        curve_id = f"ftir_foam_{formulation.lower()}"
        for point_index, (source_row, x_value, y_value) in enumerate(
            _numeric_pairs(
                sheet,
                x_column=1,
                y_column=2,
                first_row=2,
                last_row=1867,
            ),
            start=1,
        ):
            rows.append(
                _curve_row(
                    curve_id=curve_id,
                    point_index=point_index,
                    formulation_id=formulation,
                    material_id=formulation,
                    sample_identity_status="exact_formulation_label_specimen_unresolved",
                    modality="ftir_source_native_signal",
                    value=y_value,
                    unit="source_native_signal_unit_unresolved",
                    x_name="wavenumber",
                    x_value=x_value,
                    x_unit="cm^-1",
                    source_file="Supplemental data 4.xlsx",
                    source_sheet=sheet.title,
                    source_row=source_row,
                    source_column=2,
                    admission="conditional_reference",
                    mapping_status="formulation_exact_signal_semantics_unresolved",
                    protocol_status="wavenumber_resolved_signal_axis_unlabelled",
                    weight_ceiling="0.25",
                    method="FT-IR; deposited worksheet",
                    notes="combined_table_H_to_M_is_exact_duplicate_and_excluded",
                    duplicate_status="canonicalized_from_exact_duplicate_summary_table",
                )
            )

    if len(rows) != 18_660:
        raise AuditBlocked(f"FT-IR 去重观测数漂移：{len(rows)}")
    return rows, {
        "unique_curve_count": 10,
        "unique_point_count": len(rows),
        "dso_polyol_equals_monitoring_4h": True,
        "foam_combined_table_exact_matches": combined_matches,
        "duplicate_series_presentations_excluded": 6,
        "duplicate_numeric_values_excluded": 11_196,
    }


NMR_PEAK_PATTERN = re.compile(
    r"^\s*(\d+)\s+([-+0-9.]+)\s+([-+0-9.]+)\s+"
    r"([-+0-9.]+)\s+([-+0-9.]+)\s*$"
)


def _parse_nmr_peaks(sheet: Any) -> list[tuple[int, float, float, float, float, int]]:
    parsed: list[tuple[int, float, float, float, float, int]] = []
    for row_index in range(1, sheet.max_row + 1):
        cell = sheet.cell(row_index, 1).value
        match = NMR_PEAK_PATTERN.match(str(cell))
        if match:
            ordinal, address, frequency, ppm, intensity = match.groups()
            parsed.append(
                (
                    int(ordinal),
                    float(address),
                    float(frequency),
                    float(ppm),
                    float(intensity),
                    row_index,
                )
            )
            continue
        values = [sheet.cell(row_index, column).value for column in range(1, 6)]
        if all(
            isinstance(value, (int, float)) and not isinstance(value, bool)
            for value in values
        ):
            parsed.append(
                (
                    int(values[0]),
                    float(values[1]),
                    float(values[2]),
                    float(values[3]),
                    float(values[4]),
                    row_index,
                )
            )
    return parsed


def build_nmr_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    payload = _read_inner_xlsx()
    workbook = load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    specs = (
        ("HNMR OF Date seed oil", "date_seed_oil", "nmr_1H", 25),
        ("CNMR OF Date seed oil", "date_seed_oil", "nmr_13C", 34),
        ("HNMR OF DSOpolyol", "dso_polyol", "nmr_1H", 19),
        ("CNMR OF DSOpolyol", "dso_polyol", "nmr_13C", 31),
    )
    rows: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    for sheet_name, material_id, modality, expected_count in specs:
        sheet = workbook[sheet_name]
        peaks = _parse_nmr_peaks(sheet)
        if len(peaks) != expected_count:
            raise AuditBlocked(f"NMR 峰数漂移：{sheet_name}={len(peaks)}")
        curve_id = f"{modality}_{material_id}"
        counts[curve_id] = len(peaks)
        for ordinal, address, frequency, ppm, intensity, source_row in peaks:
            record_id = f"xs78ch5jb7_v3|{curve_id}|peak={ordinal}"
            row = _base_row(
                record_id=record_id,
                formulation_id="",
                sample_id=material_id,
                record_kind=f"{modality}_peak",
                property_name="nmr_peak_chemical_shift",
                value=ppm,
                unit="ppm",
                source_file="Supplemental data 2.rar",
                source_locator=(
                    "Supplemental data 2.rar::Supplemental data 2/"
                    f"Supplemental data 2.xlsx#sheet={sheet_name};row={source_row}"
                ),
                data_origin="source_native_experimental_peak_list",
                reduction_level="within_spectrum_peak",
                method=f"Bruker {modality} NMR peak list; CDCl3; 300 K",
                fidelity="raw_deposited_peak_list",
                admission="admitted_reference",
                mapping_status="material_state_and_peak_position_exact",
                protocol_status="acquisition_parameters_embedded_in_workbook",
                weight_ceiling="0.35",
                material_id=material_id,
                sample_identity_status="exact_material_state_label",
                notes=f"address={address};full_spectrum_array_not_deposited",
                condition_name="peak_ordinal",
                condition_value=ordinal,
                condition_unit="index",
                curve_id=curve_id,
                point_index=ordinal,
                secondary_condition_name="frequency",
                secondary_condition_value=frequency,
                secondary_condition_unit="Hz",
                auxiliary_value_name="source_native_peak_intensity",
                auxiliary_value=intensity,
                auxiliary_unit="source_native_relative_unit_unresolved",
            )
            rows.append(row)
    if len(rows) != 109:
        raise AuditBlocked(f"NMR 峰观测数漂移：{len(rows)}")
    return rows, {
        "spectrum_peak_list_count": 4,
        "peak_count": len(rows),
        "counts_by_spectrum": counts,
        "full_spectrum_arrays_deposited": False,
        "stray_formula_excluded": "CNMR OF Date seed oil!O6:=---R19",
    }


def _add_thermal_channel(
    rows: list[dict[str, Any]],
    *,
    sheet: Any,
    source_file: str,
    first_row: int,
    last_row: int,
    time_column: int,
    temperature_column: int,
    value_column: int,
    curve_id: str,
    formulation_id: str,
    material_id: str,
    identity_status: str,
    modality: str,
    unit: str,
    admission: str,
    mapping_status: str,
    protocol_status: str,
    weight: str,
    notes: str,
    duplicate_status: str = "unique_canonical_observation",
) -> None:
    for point_index, source_row in enumerate(range(first_row, last_row + 1), start=1):
        rows.append(
            _curve_row(
                curve_id=curve_id,
                point_index=point_index,
                formulation_id=formulation_id,
                material_id=material_id,
                sample_identity_status=identity_status,
                modality=modality,
                value=sheet.cell(source_row, value_column).value,
                unit=unit,
                x_name="time",
                x_value=sheet.cell(source_row, time_column).value,
                x_unit="min",
                source_file=source_file,
                source_sheet=sheet.title,
                source_row=source_row,
                source_column=value_column,
                admission=admission,
                mapping_status=mapping_status,
                protocol_status=protocol_status,
                weight_ceiling=weight,
                method="thermal analysis; deposited time-temperature trace",
                notes=notes,
                secondary_name="temperature",
                secondary_value=sheet.cell(source_row, temperature_column).value,
                secondary_unit="degC",
                duplicate_status=duplicate_status,
            )
        )


def build_thermal_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook3 = load_workbook(
        SOURCE_DIR / "Supplemental data 3.xlsx", data_only=True, read_only=False
    )
    workbook5 = load_workbook(
        SOURCE_DIR / "Supplemental data 5.xlsx", data_only=True, read_only=False
    )
    source3 = workbook3.worksheets[0]
    source5_s2 = workbook5.worksheets[1]
    data3_tga = [
        tuple(source3.cell(row, column).value for column in range(1, 6))
        for row in range(16, 1010)
    ]
    data5_s2 = [
        tuple(source5_s2.cell(row, column).value for column in range(1, 6))
        for row in range(3, 997)
    ]
    if data3_tga != data5_s2:
        raise AuditBlocked("DSO polyol / S2 TGA 跨工作簿重复关系漂移")

    rows: list[dict[str, Any]] = []
    conflict_common = {
        "sheet": source3,
        "source_file": "Supplemental data 3.xlsx",
        "first_row": 16,
        "last_row": 1009,
        "time_column": 1,
        "temperature_column": 2,
        "formulation_id": "",
        "material_id": "dso_polyol_or_S2_identity_conflict",
        "identity_status": "conflicting_labels_dso_polyol_vs_S2_30pct_foam",
        "admission": "conditional_reference",
        "mapping_status": "exact_numeric_duplicate_conflicting_sample_labels",
        "protocol_status": "workbook_protocol_partial_identity_unresolved",
        "weight": "0.10",
        "notes": "identical_to_Supplemental_data_5_S2;not_assigned_to_either_material",
        "duplicate_status": "canonicalized_cross_workbook_identity_conflict",
    }
    for value_column, curve_suffix, modality, unit in (
        (3, "dta", "dta_signal", "uV"),
        (4, "mass", "tga_mass_signal", "ug"),
        (5, "dtg", "dtg_mass_rate", "ug/min"),
    ):
        _add_thermal_channel(
            rows,
            value_column=value_column,
            curve_id=f"thermal_dso_polyol_or_s2_conflict_{curve_suffix}",
            modality=modality,
            unit=unit,
            **conflict_common,
        )

    _add_thermal_channel(
        rows,
        sheet=source3,
        source_file="Supplemental data 3.xlsx",
        first_row=16,
        last_row=4341,
        time_column=8,
        temperature_column=9,
        value_column=10,
        curve_id="dsc_dso_polyol_heat_flow",
        formulation_id="",
        material_id="dso_polyol",
        identity_status="exact_material_state_label",
        modality="dsc_heat_flow",
        unit="W/g",
        admission="admitted_reference",
        mapping_status="material_state_axis_and_unit_exact",
        protocol_status="time_temperature_trace_and_unit_present",
        weight="0.40",
        notes="DSC_protocol_standard_not_present_in_supporting_information",
    )

    thermal_specs = (
        (0, "S1", 3, 796, ((3, "mass", "tga_mass_signal", "%"),)),
        (
            2,
            "S3",
            3,
            896,
            (
                (3, "dta", "dta_signal", "uV"),
                (4, "mass", "tga_mass_signal", "ug"),
                (5, "dtg", "dtg_mass_rate", "ug/min"),
            ),
        ),
        (
            3,
            "S4",
            3,
            796,
            (
                (3, "dta", "dta_signal", "uV"),
                (4, "mass", "tga_mass_signal", "ug"),
                (5, "dtg", "dtg_mass_rate", "ug/min"),
            ),
        ),
        (4, "S5", 3, 857, ((3, "mass", "tga_mass_signal", "%"),)),
    )
    for sheet_index, formulation, first_row, last_row, channels in thermal_specs:
        sheet = workbook5.worksheets[sheet_index]
        for value_column, suffix, modality, unit in channels:
            _add_thermal_channel(
                rows,
                sheet=sheet,
                source_file="Supplemental data 5.xlsx",
                first_row=first_row,
                last_row=last_row,
                time_column=1,
                temperature_column=2,
                value_column=value_column,
                curve_id=f"thermal_{formulation.lower()}_{suffix}",
                formulation_id=formulation,
                material_id=formulation,
                identity_status="exact_formulation_label_specimen_unresolved",
                modality=modality,
                unit=unit,
                admission="admitted_reference",
                mapping_status="formulation_axis_and_unit_exact",
                protocol_status="time_temperature_trace_and_unit_present_standard_unreported",
                weight="0.45",
                notes="single_deposited_trace_replicate_count_unreported",
            )

    modality_counts = Counter(row["property_name"] for row in rows)
    expected = {
        "dta_signal": 2_682,
        "tga_mass_signal": 4_331,
        "dtg_mass_rate": 2_682,
        "dsc_heat_flow": 4_326,
    }
    if dict(modality_counts) != expected:
        raise AuditBlocked(f"热分析观测计数漂移：{dict(modality_counts)}")
    return rows, {
        "unique_curve_count": 12,
        "unique_observation_count": len(rows),
        "counts_by_modality": expected,
        "dso_polyol_tga_equals_s2_tga": True,
        "identity_conflict_observation_count": 2_982,
        "duplicate_series_presentations_excluded": 3,
        "duplicate_numeric_values_excluded": 2_982,
        "source_native_mass_units": ["%", "ug"],
        "mass_values_normalized_or_converted": False,
    }


def build_mechanical_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(
        SOURCE_DIR / "Supplemental data 6.xlsx", data_only=True, read_only=False
    )
    rows: list[dict[str, Any]] = []
    curve_counts: dict[str, int] = {}
    ragged_tail_blank_count = 0
    specs = (
        ("Tensile stress-strain", "tensile_stress_signal", "tensile"),
        ("Compressive stress-strain", "compressive_stress_signal", "compressive"),
    )
    for sheet_name, modality, prefix in specs:
        sheet = workbook[sheet_name]
        x_values = [sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)]
        if not all(isinstance(value, (int, float)) for value in x_values):
            raise AuditBlocked(f"机械曲线公共横轴含非数值：{sheet_name}")
        for column in range(2, 7):
            formulation = str(sheet.cell(1, column).value)
            if formulation not in FORMULATIONS:
                raise AuditBlocked(f"机械曲线配方标签漂移：{sheet_name}!{column}")
            values = [
                sheet.cell(row, column).value
                for row in range(2, sheet.max_row + 1)
            ]
            nonempty = [index for index, value in enumerate(values) if value is not None]
            if not nonempty or nonempty != list(range(nonempty[-1] + 1)):
                raise AuditBlocked(f"机械曲线含内部缺失：{sheet_name}/{formulation}")
            point_count = len(nonempty)
            ragged_tail_blank_count += len(values) - point_count
            curve_id = f"{prefix}_{formulation.lower()}"
            curve_counts[curve_id] = point_count
            for point_index in range(1, point_count + 1):
                source_row = point_index + 1
                rows.append(
                    _curve_row(
                        curve_id=curve_id,
                        point_index=point_index,
                        formulation_id=formulation,
                        material_id=formulation,
                        sample_identity_status="exact_formulation_label_specimen_unresolved",
                        modality=modality,
                        value=sheet.cell(source_row, column).value,
                        unit="source_axis_unit_unresolved",
                        x_name="strain_signal",
                        x_value=sheet.cell(source_row, 1).value,
                        x_unit="source_axis_unit_unresolved",
                        source_file="Supplemental data 6.xlsx",
                        source_sheet=sheet.title,
                        source_row=source_row,
                        source_column=column,
                        admission="conditional_reference",
                        mapping_status="formulation_exact_axis_units_unresolved",
                        protocol_status="unresolved_in_deposit_workbooks_and_supporting_information",
                        weight_ceiling="0.30",
                        method=f"{sheet_name}; deposited worksheet",
                        notes="MPa_and_percent_are_plausible_but_not_asserted_without_article_protocol",
                    )
                )
    expected_counts = {
        "tensile_s1": 1_584,
        "tensile_s2": 1_244,
        "tensile_s3": 1_260,
        "tensile_s4": 1_026,
        "tensile_s5": 907,
        "compressive_s1": 14_982,
        "compressive_s2": 12_588,
        "compressive_s3": 9_829,
        "compressive_s4": 12_656,
        "compressive_s5": 12_743,
    }
    if curve_counts != expected_counts:
        raise AuditBlocked(f"机械曲线点数漂移：{curve_counts}")
    if len(rows) != 68_819 or ragged_tail_blank_count != 14_011:
        raise AuditBlocked(
            f"机械观测/尾部填充漂移：{len(rows)}/{ragged_tail_blank_count}"
        )
    return rows, {
        "curve_count": len(curve_counts),
        "paired_stress_strain_point_count": len(rows),
        "curve_point_counts": curve_counts,
        "internal_missing_value_count": 0,
        "ragged_terminal_padding_blank_cells": ragged_tail_blank_count,
        "x_axis_semantics": "strain_from_sheet_title",
        "x_axis_unit": "unresolved",
        "y_axis_semantics": "stress_from_sheet_title",
        "y_axis_unit": "unresolved",
        "protocol_status": "not_present_in_deposited_workbooks_or_supporting_information",
        "replicate_status": "not_reported",
    }


def build_formulation_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    polyester = (100, 70, 50, 30, 0)
    dso_polyol = (0, 30, 50, 70, 100)
    water = (0.75, 0.75, 0.75, 0.75, 0.75)
    pmdi_a = (119.6616, 127.9249, 133.4338, 138.9427, 147.2060)
    pmdi_b = (119.6494, 127.9103, 133.4173, 138.9243, 147.1848)
    pmdi_used = (120, 128, 133, 139, 147)
    rows: list[dict[str, Any]] = []
    for index, formulation in enumerate(FORMULATIONS):
        material_id = formulation
        specs = (
            (
                "polyester_polyol_share",
                polyester[index],
                "parts_per_100_polyol_parts",
                "polyester_polyol",
                "polyol_component",
                "admitted_reference",
                "0.65",
                "source_reported_formulation",
                "polyol_blend_basis_closed_by_0_30_50_70_100_percent_labels",
            ),
            (
                "dso_polyol_share",
                dso_polyol[index],
                "parts_per_100_polyol_parts",
                "date_seed_oil_polyol",
                "bio_polyol_component",
                "admitted_reference",
                "0.65",
                "source_reported_formulation",
                "polyol_blend_basis_closed_by_0_30_50_70_100_percent_labels",
            ),
            (
                "water_amount",
                water[index],
                "source_table_amount_basis_unresolved",
                "water",
                "blowing_agent",
                "conditional_reference",
                "0.30",
                "source_reported_formulation",
                "table_has_no_explicit_mass_unit",
            ),
            (
                "pmdi_amount_used",
                pmdi_used[index],
                "source_table_amount_basis_unresolved",
                "PMDI",
                "isocyanate_component",
                "conditional_reference",
                "0.35",
                "source_reported_formulation",
                "average_value_used_but_table_has_no_explicit_mass_unit",
            ),
            (
                "pmdi_required_method_a",
                pmdi_a[index],
                "source_table_amount_basis_unresolved",
                "PMDI",
                "calculated_isocyanate_amount",
                "conditional_reference",
                "0.20",
                "source_reported_calculation",
                "calculated_by_method_A_not_final_used_amount",
            ),
            (
                "pmdi_required_method_b",
                pmdi_b[index],
                "source_table_amount_basis_unresolved",
                "PMDI",
                "calculated_isocyanate_amount",
                "conditional_reference",
                "0.20",
                "source_reported_calculation",
                "calculated_by_method_B_not_final_used_amount",
            ),
            (
                "isocyanate_index",
                1.7,
                "dimensionless",
                "PMDI",
                "process_ratio",
                "admitted_reference",
                "0.60",
                "source_reported_process_condition",
                "reported_for_all_five_samples",
            ),
        )
        for ordinal, (
            property_name,
            value,
            unit,
            component,
            role,
            admission,
            weight,
            data_origin,
            notes,
        ) in enumerate(specs, start=1):
            record_id = (
                f"xs78ch5jb7_v3|formulation={formulation}|"
                f"property={property_name}"
            )
            rows.append(
                _base_row(
                    record_id=record_id,
                    formulation_id=formulation,
                    sample_id=formulation,
                    record_kind="formulation_or_process_scalar",
                    component_name=component,
                    component_role=role,
                    property_name=property_name,
                    value=value,
                    unit=unit,
                    source_file="Supporting Information.pdf",
                    source_locator=f"Supporting Information.pdf#page=18;Table=S1;row={ordinal}",
                    data_origin=data_origin,
                    reduction_level="source_reported_scalar",
                    method="Supporting Information Table S1",
                    fidelity="source_reported_formulation_or_process_value",
                    admission=admission,
                    mapping_status="exact_formulation_mapping",
                    protocol_status="source_reported",
                    weight_ceiling=weight,
                    material_id=material_id,
                    sample_identity_status="exact_formulation_label",
                    notes=notes,
                    duplicate_status="source_reported_single_value",
                )
            )

    record_id = "xs78ch5jb7_v3|reagent=PMDI|property=nco_content"
    rows.append(
        _base_row(
            record_id=record_id,
            formulation_id="",
            sample_id="PMDI",
            record_kind="reagent_scalar",
            component_name="PMDI",
            component_role="isocyanate_component",
            property_name="nco_content",
            value=30.5,
            unit="%",
            source_file="Supporting Information.pdf",
            source_locator="Supporting Information.pdf#page=17;equation=6",
            data_origin="source_reported_reagent_property",
            reduction_level="source_reported_scalar",
            method="Supporting Information equation 6",
            fidelity="source_reported_reagent_value",
            admission="admitted_reference",
            mapping_status="exact_reagent_mapping",
            protocol_status="source_reported",
            weight_ceiling="0.60",
            material_id="PMDI",
            sample_identity_status="reagent_family_identity_only",
            notes="PMDI_trade_grade_not_reported_in_supporting_information",
            duplicate_status="source_reported_single_value",
        )
    )
    if len(rows) != 36:
        raise AuditBlocked(f"配方/过程标量数漂移：{len(rows)}")
    return rows, {
        "final_formulation_count": 5,
        "source_reported_scalar_count": len(rows),
        "admitted_reference_count": sum(
            row["gold_admission_status"] == "admitted_reference" for row in rows
        ),
        "conditional_reference_count": sum(
            row["gold_admission_status"] == "conditional_reference" for row in rows
        ),
        "actual_preparation_amount_unit_status": "water_and_pmdi_basis_unresolved",
        "method_a_b_values_are_final_used_values": False,
    }


def build_gold_e_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    ftir_rows, ftir_summary = build_ftir_rows()
    nmr_rows, nmr_summary = build_nmr_rows()
    thermal_rows, thermal_summary = build_thermal_rows()
    mechanical_rows, mechanical_summary = build_mechanical_rows()
    formulation_rows, formulation_summary = build_formulation_rows()
    rows = [
        *ftir_rows,
        *nmr_rows,
        *thermal_rows,
        *mechanical_rows,
        *formulation_rows,
    ]
    if len(rows) != 101_645:
        raise AuditBlocked(f"Gold-E 来源内长表计数漂移：{len(rows)}")
    observation_ids = [row["observation_id"] for row in rows]
    if len(set(observation_ids)) != len(observation_ids):
        duplicates = Counter(observation_ids)
        raise AuditBlocked(
            "Gold-E observation_id 重复："
            + ",".join(key for key, count in duplicates.items() if count > 1)
        )
    if any(tuple(row) != RECORD_COLUMNS for row in rows):
        raise AuditBlocked("Gold-E 来源内长表字段顺序漂移")

    admission_counts = Counter(row["gold_admission_status"] for row in rows)
    if admission_counts != {
        "admitted_reference": 11_164,
        "conditional_reference": 90_481,
    }:
        raise AuditBlocked(f"Gold-E 准入计数漂移：{dict(admission_counts)}")
    curve_rows = [row for row in rows if row["curve_id"]]
    curve_points = len(curve_rows)
    if curve_points != 101_609:
        raise AuditBlocked(f"曲线/峰点计数漂移：{curve_points}")
    if any(row["training_weight"] for row in rows):
        raise AuditBlocked("本批不得物化训练权重")
    if {row["current_weight_materialized"] for row in rows} != {"false"}:
        raise AuditBlocked("当前权重状态漂移")

    modality_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        modality_rows[row["property_name"]].append(row)
    counts_by_property = {
        property_name: {
            "observations": len(property_rows),
            "series": len(
                {row["curve_id"] for row in property_rows if row["curve_id"]}
            ),
        }
        for property_name, property_rows in sorted(modality_rows.items())
    }

    return rows, {
        "ftir": ftir_summary,
        "nmr": nmr_summary,
        "thermal": thermal_summary,
        "mechanical": mechanical_summary,
        "formulation": formulation_summary,
        "gold_e_observation_count": len(rows),
        "curve_and_peak_observation_count": curve_points,
        "source_reported_scalar_count": len(formulation_rows),
        "unique_numeric_series_count": 36,
        "raw_numeric_series_presentations_before_dedup": 45,
        "duplicate_series_presentations_excluded": 9,
        "duplicate_numeric_values_excluded": 14_178,
        "gold_admission_status_counts": dict(admission_counts),
        "counts_by_property": counts_by_property,
        "independent_final_formulations": 5,
        "independent_material_families": 1,
        "explicit_physical_specimen_count": 0,
        "explicit_replicate_group_count": 0,
        "curve_points_counted_as_independent_samples": 0,
        "training_weight_materialized": False,
    }


def _render_tsv(columns: tuple[str, ...], rows: Iterable[dict[str, Any]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _render_readme(summary: dict[str, Any]) -> bytes:
    counts = summary["materialization"]
    mechanical = counts["mechanical"]
    admission = counts["gold_admission_status_counts"]
    text = f"""# 日期籽油基低密度刚性 PU-PIR 数据来源说明

## 来源身份

- 数据集：Taghvaei Nia, S.; Alavi Nikje, M. M. *Novel bio-polyol synthesis based on date seed oil for low-density rigid polyurethane-polyisocyanurate foams* (Version 3) [Data set]. Mendeley Data, 2024. DOI: [{SOURCE_DOI}](https://doi.org/{SOURCE_DOI}).
- 关联论文：Taghvaei Nia, S.; Alavi Nikje, M. M. Novel bio-polyol synthesis based on date seed oil for low-density rigid polyurethane-polyisocyanurate foams. *Industrial Crops and Products* **2025**, *223*, 120152. DOI: [{ARTICLE_DOI}](https://doi.org/{ARTICLE_DOI}).
- 数据集发布日期：2024-11-07；版本：3；许可：[CC BY 4.0](https://creativecommons.org/licenses/by/4.0/)。论文页面标注开放获取，Crossref 记录的论文许可为 CC BY-NC 4.0；本目录物化值来自 CC BY 4.0 数据集附件。
- 官方入口：[数据集]({DATASET_URL})、[元数据]({METADATA_URL})、[附件清单]({FILES_URL})、[版本历史]({VERSIONS_URL})、[论文]({ARTICLE_URL})。

## 数据内容与精确计数

- 7 个官方附件：5 个直接 XLSX、1 个含单一 XLSX 的 RAR、1 个 19 页支持信息 PDF；合计 6 个逻辑工作簿、22 个工作表，其中 20 个非空。
- 5 个最终刚性 PU-PIR 配方：S1/S2/S3/S4/S5，对应 DSO polyol 0/30/50/70/100%。没有可区分的物理试样编号或重复组，因此不能把曲线数或点数当作独立材料数。
- 来源内 Gold-E 长表 {counts['gold_e_observation_count']:,} 行：曲线/峰点 {counts['curve_and_peak_observation_count']:,} 行，配方/过程标量 {counts['source_reported_scalar_count']} 行。
- 机械：10 条曲线，{mechanical['paired_stress_strain_point_count']:,} 个应力—应变配对点；无曲线内部缺失，14,011 个空单元格只是不同曲线长度造成的尾部填充。
- FT-IR：10 条去重曲线，18,660 点；NMR：4 份峰表，109 个峰；TGA/DTA/DTG：9,695 个去重观测；DSC：4,326 点。
- 准入：正式参考 {admission['admitted_reference']:,} 行，条件参考 {admission['conditional_reference']:,} 行；当前训练权重和训练/验证划分均未物化。

## 重复、冲突与门禁

1. `Supplemental data 1.xlsx` 的 DSO polyol FT-IR 与 4 h 监测曲线逐点完全相同，保留一个规范观测。
2. `Supplemental data 4.xlsx` 的 H:M 汇总区逐点重复 S1—S5 五个独立工作表，汇总区不重复物化。
3. `Supplemental data 3.xlsx` 的“DSO polyol”TGA 五列与 `Supplemental data 5.xlsx` 的“S2”逐点完全相同但样品标签冲突。该组 2,982 个 DTA/TG/DTG 数值以冲突身份保留在条件层，不强行归入任一材料。
4. 机械工作簿只有 `Tensile stress-strain` / `Compressive stress-strain` 与 S1—S5 标签，没有轴标题；支持信息也没有机械测试标准、试样尺寸、加载速率或重复数。因此本批不把 MPa/% 当作已确认单位，机械 68,819 点全部为条件参考。
5. FT-IR 波数单位可确认，但纵轴信号语义/单位未声明，因此保留原值并进入条件层。TGA 的 `%` 与 `ug` 原单位不换算、不归一化。
6. RAR 内 `CNMR OF Date seed oil!O6` 有孤立公式 `=---R19`，不属于峰表，未物化为科学数据。

## 文件说明

- `Gold_E_实验观测长表.tsv`：来源内去重后的实验曲线、峰和配方/过程标量；每行保留文件、工作表/单元格、样品映射、单位状态、准入状态和泄漏分组。
- `内容审计摘要.json`：机器可读计数、工作簿结构、重复/冲突和门禁。
- `文件校验清单.tsv`：官方附件及本地元数据快照的大小与 SHA-256。

## 论文写作参考文献

1. Taghvaei Nia, S.; Alavi Nikje, M. M. *Novel bio-polyol synthesis based on date seed oil for low-density rigid polyurethane-polyisocyanurate foams* (Version 3) [Data set]. Mendeley Data, 2024. https://doi.org/{SOURCE_DOI}.
2. Taghvaei Nia, S.; Alavi Nikje, M. M. Novel bio-polyol synthesis based on date seed oil for low-density rigid polyurethane-polyisocyanurate foams. *Industrial Crops and Products* **2025**, *223*, 120152. https://doi.org/{ARTICLE_DOI}.
3. Creative Commons. *Attribution 4.0 International (CC BY 4.0).* https://creativecommons.org/licenses/by/4.0/.
"""
    return text.encode("utf-8")


def _atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def run_audit(*, write_outputs: bool = True) -> dict[str, Any]:
    file_rows, verification = verify_sources()
    rows, materialization = build_gold_e_rows()
    summary = {
        "audit_version": AUDIT_VERSION,
        "source": {
            "source_id": SOURCE_ID,
            "dataset_doi": SOURCE_DOI,
            "version": 3,
            "publish_date": "2024-11-07",
            "license": SOURCE_LICENSE,
            "article_doi": ARTICLE_DOI,
            "article_license": ARTICLE_LICENSE,
            "repository": "Mendeley Data",
            "source_reliability": "R1",
            "dataset_url": DATASET_URL,
        },
        "verification": verification,
        "materialization": materialization,
        "gold_policy": {
            "layer": "Gold-E multi-fidelity reference",
            "formal_rule": "source identity, material mapping, value and unit semantics closed",
            "conditional_rule": "reliable source but unit, protocol, or sample identity has an explicit unresolved field",
            "point_level_split_forbidden": True,
            "formulation_group_split_required": True,
            "family_leakage_group": FAMILY_LEAKAGE_GROUP,
            "current_weight_materialized": False,
        },
        "risk_register": [
            {
                "risk": "mechanical_axis_units_and_protocol_unresolved",
                "affected_observations": 68_819,
                "action": "obtain article methods/figure axis or author clarification before formal promotion",
            },
            {
                "risk": "dso_polyol_vs_s2_tga_identity_conflict",
                "affected_observations": 2_982,
                "action": "retain under neutral conflict identity; never assign to S2 or DSO polyol without external evidence",
            },
            {
                "risk": "ftir_signal_axis_semantics_unresolved",
                "affected_observations": 18_660,
                "action": "retain raw values as auxiliary fingerprint only",
            },
            {
                "risk": "replicate_and_physical_specimen_ids_unreported",
                "affected_observations": 101_609,
                "action": "group all points by formulation and family; do not point-split",
            },
        ],
    }

    outputs = {
        OUTPUT_LONG_TABLE.name: _render_tsv(RECORD_COLUMNS, rows),
        OUTPUT_CHECKSUMS.name: _render_tsv(FILE_COLUMNS, file_rows),
        OUTPUT_AUDIT.name: (
            json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True)
            + "\n"
        ).encode("utf-8"),
        OUTPUT_README.name: _render_readme(summary),
    }
    if write_outputs:
        for name, payload in outputs.items():
            _atomic_write(SOURCE_DIR / name, payload)
    return {
        "summary": summary,
        "files": file_rows,
        "rows": rows,
        "outputs": outputs,
    }


def main() -> None:
    result = run_audit(write_outputs=True)
    materialization = result["summary"]["materialization"]
    print(
        json.dumps(
            {
                "source": SOURCE_ID,
                "gold_e_observations": materialization["gold_e_observation_count"],
                "formal": materialization["gold_admission_status_counts"][
                    "admitted_reference"
                ],
                "conditional": materialization["gold_admission_status_counts"][
                    "conditional_reference"
                ],
                "final_formulations": materialization[
                    "independent_final_formulations"
                ],
                "mechanical_points": materialization["mechanical"][
                    "paired_stress_strain_point_count"
                ],
            },
            ensure_ascii=True,
        )
    )


if __name__ == "__main__":
    main()
