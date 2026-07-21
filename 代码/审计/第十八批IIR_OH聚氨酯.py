"""审计 Mendeley IIR-OH 低渗透抗钙化聚氨酯固定版本。

只从已校验 ZIP 流读取，不解包原件。输出来源目录下的内容审计摘要、
文件校验清单和曲线审计清单；不创建训练划分或训练权重。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from 审计.第十批ACS表格物化 import RECORD_COLUMNS
except ModuleNotFoundError as error:
    if error.name != "审计":
        raise
    from 第十批ACS表格物化 import RECORD_COLUMNS


ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "第十八批实验_IIR-OH聚氨酯"
)
ARCHIVE = SOURCE_DIR / "wg3znh66bv-1.zip"
OUTPUT_GOLD_E_SCALARS = SOURCE_DIR / "Gold_E_实验指标.tsv"
ARCHIVE_BYTES = 38_825_776
ARCHIVE_SHA256 = "5224ae550be0c04022f0de02f03bd08f6ae39f2e6f84044fb8ec9e1cd5872e76"
DATASET_DOI = "10.17632/wg3znh66bv.1"
LICENSE = "CC-BY-4.0"
SOURCE_FAMILY_KEY = "family_iir_oh_barrier_pu_2026"
CITATION_KEYS = (
    "ledger-172-zhang-2026-iir-oh-barrier-pu-data;"
    "ledger-173-zhang-2026-iir-oh-barrier-pu-preprint"
)
EXPECTED_MEMBER_COUNT = 207
EXPECTED_UNCOMPRESSED_BYTES = 157_552_695
EXPECTED_SUFFIX_COUNTS = {
    ".arw": 2,
    ".csv": 53,
    ".opju": 1,
    ".pdf": 2,
    ".spa": 47,
    ".txt": 99,
    ".xlsx": 3,
}
EXPECTED_CATEGORY_COUNTS = {
    "barrier_processed_project": 1,
    "cyclic_tensile_raw": 6,
    "ftir_curve_csv": 47,
    "ftir_vendor_binary": 47,
    "gpc_report": 2,
    "gpc_text_curve": 2,
    "gpc_workbook": 1,
    "hydrodynamic_curve": 3,
    "hydrolytic_aging_tensile_raw": 12,
    "nmr_curve": 6,
    "network_calculation_workbook": 1,
    "processed_tensile_workbook": 1,
    "tensile_raw": 45,
    "uv_vis_curve": 33,
}
NETWORK_FORMULATION_CODES = (
    "HDI-2",
    "HDI-4",
    "HDI-6",
    "HDI-8",
    "HDI-10",
    "HMDI-2",
    "HMDI-4",
    "HMDI-6",
    "HMDI-8",
    "HMDI-10",
    "MDI-1",
    "MDI-2",
    "MDI-4",
    "MDI-6",
    "MDI-8",
    "MDI-10",
)
GOLD_E_OPTIONAL_COLUMNS = (
    "batch_id",
    "curve_id",
    "point_index",
    "secondary_condition_name",
    "secondary_condition_value",
    "secondary_condition_unit",
    "auxiliary_value_name",
    "auxiliary_value",
    "auxiliary_unit",
    "sample_identity_status",
    "global_structure_family_key",
    "family_leakage_group",
    "curve_points_are_independent_samples",
    "duplicate_status",
)
GOLD_E_SCALAR_COLUMNS = (*RECORD_COLUMNS, *GOLD_E_OPTIONAL_COLUMNS)


class AuditBlocked(RuntimeError):
    """固定原件、归档边界或表结构发生漂移。"""


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _sha256_member(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> str:
    digest = hashlib.sha256()
    with archive.open(info) as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _safe_member(info: zipfile.ZipInfo) -> None:
    if "\\" in info.filename or "\x00" in info.filename:
        raise AuditBlocked(f"ZIP成员路径格式不安全: {info.filename!r}")
    path = PurePosixPath(info.filename)
    if path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        raise AuditBlocked(f"ZIP成员路径逃逸: {info.filename!r}")
    unix_mode = (info.external_attr >> 16) & 0xFFFF
    if unix_mode and stat.S_ISLNK(unix_mode):
        raise AuditBlocked(f"ZIP成员是符号链接: {info.filename!r}")


def _category(name: str) -> str:
    lower = name.lower()
    suffix = PurePosixPath(name).suffix.lower()
    if "representative raw ¹h nmr" in lower and suffix == ".csv":
        return "nmr_curve"
    if "representative gpc export" in lower:
        if suffix == ".xlsx":
            return "gpc_workbook"
        if suffix == ".arw":
            return "gpc_text_curve"
        if suffix == ".pdf":
            return "gpc_report"
    if "/tensile raw data/" in lower and suffix == ".txt":
        return "tensile_raw"
    if "/tensile raw data/" in lower and suffix == ".xlsx":
        return "processed_tensile_workbook"
    if "cyclic tensile raw data" in lower and suffix == ".txt":
        return "cyclic_tensile_raw"
    if "swelling raw data and crosslink density" in lower and suffix == ".xlsx":
        return "network_calculation_workbook"
    if "oxygen and water vapor permeability" in lower and suffix == ".opju":
        return "barrier_processed_project"
    if "hydrolytic aging mechanical retention" in lower and suffix == ".txt":
        return "hydrolytic_aging_tensile_raw"
    if "water uptake ftir raw data" in lower:
        if suffix == ".csv":
            return "ftir_curve_csv"
        if suffix == ".spa":
            return "ftir_vendor_binary"
    if "uv–vis calibration raw data" in lower and suffix == ".txt":
        return "uv_vis_curve"
    if "hydrodynamic testing data" in lower and suffix == ".txt":
        return "hydrodynamic_curve"
    raise AuditBlocked(f"未治理的ZIP成员类型: {name}")


def _finite_number(value: str, context: str) -> float:
    try:
        parsed = float(value)
    except ValueError as error:
        raise AuditBlocked(f"非数值字段: {context}={value!r}") from error
    if not math.isfinite(parsed):
        raise AuditBlocked(f"非有限数值: {context}={value!r}")
    return parsed


def _text_lines(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> list[str]:
    payload = archive.read(info)
    encodings = ("utf-16",) if payload.startswith((b"\xff\xfe", b"\xfe\xff")) else (
        "utf-8-sig",
        "gb18030",
    )
    for encoding in encodings:
        try:
            return payload.decode(encoding).splitlines()
        except UnicodeError:
            continue
    raise AuditBlocked(f"文本编码无法解析: {info.filename}")


def _two_column_curve(lines: Iterable[str], context: str) -> int:
    count = 0
    delimiter: str | None = None
    for line_number, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        if delimiter is None:
            delimiter = "\t" if "\t" in line else "," if "," in line else None
            if delimiter is None:
                raise AuditBlocked(f"双列曲线分隔符未知: {context}#L{line_number}")
        fields = [part.strip() for part in line.split(delimiter)]
        while fields and not fields[-1]:
            fields.pop()
        if len(fields) != 2:
            raise AuditBlocked(f"双列曲线字段漂移: {context}#L{line_number}")
        _finite_number(fields[0], f"{context}#L{line_number}:x")
        _finite_number(fields[1], f"{context}#L{line_number}:y")
        count += 1
    if not count:
        raise AuditBlocked(f"双列曲线为空: {context}")
    return count


def _two_column_values(lines: Iterable[str], context: str) -> list[tuple[float, float]]:
    values: list[tuple[float, float]] = []
    delimiter: str | None = None
    for line_number, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        if delimiter is None:
            delimiter = "\t" if "\t" in line else "," if "," in line else None
            if delimiter is None:
                raise AuditBlocked(f"双列曲线分隔符未知: {context}#L{line_number}")
        fields = [part.strip() for part in line.split(delimiter)]
        while fields and not fields[-1]:
            fields.pop()
        if len(fields) != 2:
            raise AuditBlocked(f"双列曲线字段漂移: {context}#L{line_number}")
        values.append(
            (
                _finite_number(fields[0], f"{context}#L{line_number}:x"),
                _finite_number(fields[1], f"{context}#L{line_number}:y"),
            )
        )
    if not values:
        raise AuditBlocked(f"双列曲线为空: {context}")
    return values


def _instrument_tensile_values(
    lines: list[str], context: str
) -> tuple[list[float], list[float]]:
    if len(lines) < 3 or not lines[0].startswith("DataNumber:"):
        raise AuditBlocked(f"拉伸文件头漂移: {context}")
    expected = int(lines[0].split(":", 1)[1])
    header = lines[1].split()
    if header != ["Load", "Exten.", "B-Exten.", "Displ.", "Stress", "Strain", "Time"]:
        raise AuditBlocked(f"拉伸字段漂移: {context}={header}")
    stress: list[float] = []
    strain: list[float] = []
    for line_number, line in enumerate(lines[2:], start=3):
        if not line.strip():
            continue
        fields = line.split()
        if len(fields) != 7:
            raise AuditBlocked(f"拉伸数据列数漂移: {context}#L{line_number}")
        parsed = [
            _finite_number(value, f"{context}#L{line_number}:c{index + 1}")
            for index, value in enumerate(fields)
        ]
        stress.append(parsed[4])
        strain.append(parsed[5])
    if len(stress) != expected:
        raise AuditBlocked(
            f"拉伸DataNumber不一致: {context}={expected}/{len(stress)}"
        )
    return stress, strain


def _instrument_tensile(lines: list[str], context: str) -> int:
    stress, _ = _instrument_tensile_values(lines, context)
    return len(stress)


def _cyclic_tensile(lines: list[str], context: str) -> int:
    if not lines or lines[0].split() != ["应变(%)", "应力(MPa)"]:
        raise AuditBlocked(f"循环拉伸字段漂移: {context}")
    count = 0
    for line_number, line in enumerate(lines[1:], start=2):
        if not line.strip():
            continue
        fields = line.split()
        if len(fields) != 2:
            raise AuditBlocked(f"循环拉伸列数漂移: {context}#L{line_number}")
        _finite_number(fields[0], f"{context}#L{line_number}:strain")
        _finite_number(fields[1], f"{context}#L{line_number}:stress")
        count += 1
    return count


def _uv_vis(lines: list[str], context: str) -> int:
    try:
        marker = lines.index("Wavelength Scan Data Record")
    except ValueError as error:
        raise AuditBlocked(f"UV-vis扫描段缺失: {context}") from error
    if marker + 1 >= len(lines) or not lines[marker + 1].startswith("No."):
        raise AuditBlocked(f"UV-vis扫描表头漂移: {context}")
    count = 0
    for line_number, line in enumerate(lines[marker + 2 :], start=marker + 3):
        if not line.strip():
            continue
        fields = line.split()
        if len(fields) != 7 or not fields[0].isdigit():
            raise AuditBlocked(f"UV-vis扫描行漂移: {context}#L{line_number}")
        for index, value in enumerate(fields[1:]):
            _finite_number(value, f"{context}#L{line_number}:c{index + 2}")
        count += 1
    return count


def _hydrodynamic(lines: list[str], context: str) -> int:
    if len(lines) < 3 or lines[0].split()[0] != "Times" or lines[1].split()[0] != "s":
        raise AuditBlocked(f"流体测试表头漂移: {context}")
    count = 0
    for line_number, line in enumerate(lines[2:], start=3):
        if not line.strip():
            continue
        fields = line.split()
        if len(fields) != 4:
            raise AuditBlocked(f"流体测试列数漂移: {context}#L{line_number}")
        for index, value in enumerate(fields):
            _finite_number(value, f"{context}#L{line_number}:c{index + 1}")
        count += 1
    return count


def _curve_row(
    *,
    info: zipfile.ZipInfo,
    category: str,
    curve_id: str,
    point_count: int,
    x_name: str,
    x_unit: str,
    y_name: str,
    y_unit: str,
    source_sha256: str,
    status: str = "candidate",
    notes: str = "",
) -> dict[str, Any]:
    return {
        "curve_id": curve_id,
        "category": category,
        "member_path": info.filename,
        "member_sha256": source_sha256,
        "point_count": point_count,
        "x_name": x_name,
        "x_unit": x_unit,
        "y_name": y_name,
        "y_unit": y_unit,
        "audit_status": status,
        "notes": notes,
    }


def _workbook_curves(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    category: str,
    source_sha256: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    payload = archive.read(info)
    formulas = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    values = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    workbook_summary: dict[str, Any] = {
        "member_path": info.filename,
        "sheets": [],
    }
    curves: list[dict[str, Any]] = []
    for formula_sheet, value_sheet in zip(formulas.worksheets, values.worksheets, strict=True):
        formula_count = 0
        cached_formula_count = 0
        for formula_row, value_row in zip(
            formula_sheet.iter_rows(), value_sheet.iter_rows(), strict=True
        ):
            for formula_cell, value_cell in zip(formula_row, value_row, strict=True):
                if formula_cell.data_type == "f":
                    formula_count += 1
                    cached_formula_count += value_cell.value is not None
        workbook_summary["sheets"].append(
            {
                "sheet": formula_sheet.title,
                "max_row": formula_sheet.max_row,
                "max_column": formula_sheet.max_column,
                "formula_count": formula_count,
                "cached_formula_count": cached_formula_count,
            }
        )

    if category == "gpc_workbook":
        sheet = values["GPC"]
        header = next(sheet.iter_rows(min_row=1, max_row=1, max_col=4, values_only=True))
        labels = [str(header[0]), str(header[2])]
        point_counts = [0, 0]
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=3, max_col=4, values_only=True), start=3
        ):
            for pair_index, start_index in enumerate((0, 2)):
                left = row[start_index]
                right = row[start_index + 1]
                if left is None and right is None:
                    continue
                if not isinstance(left, (int, float)) or not isinstance(right, (int, float)):
                    raise AuditBlocked(
                        f"GPC数值漂移: {info.filename}#{labels[pair_index]}:row={row_number}"
                    )
                point_counts[pair_index] += 1
        for label, point_count in zip(labels, point_counts, strict=True):
            curves.append(
                _curve_row(
                    info=info,
                    category=category,
                    curve_id=f"gpc:{label}",
                    point_count=point_count,
                    x_name="retention_time",
                    x_unit="min",
                    y_name="source_reported_Mv_signal",
                    y_unit="source_native_unresolved",
                    source_sha256=source_sha256,
                    status="semantic_duplicate_representation",
                    notes="与同目录N-1/M-1文本ARW曲线语义重复；工作簿仅作交叉核验，不重复计曲线。",
                )
            )
    elif category == "processed_tensile_workbook":
        sheet = values["MDI"]
        header = next(sheet.iter_rows(min_row=3, max_row=3, max_col=4, values_only=True))
        labels = [str(header[1]), str(header[3])]
        point_counts = [0, 0]
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=5, max_col=4, values_only=True), start=5
        ):
            for pair_index, start_index in enumerate((0, 2)):
                left = row[start_index]
                right = row[start_index + 1]
                if left is None and right in {None, "--"}:
                    continue
                if not isinstance(left, (int, float)) or not isinstance(right, (int, float)):
                    raise AuditBlocked(
                        "处理后拉伸数值漂移: "
                        f"{info.filename}#{labels[pair_index]}:row={row_number}"
                    )
                point_counts[pair_index] += 1
        for pair_index, (label, point_count) in enumerate(
            zip(labels, point_counts, strict=True)
        ):
            curves.append(
                _curve_row(
                    info=info,
                    category=category,
                    curve_id=f"processed-tensile:{label}:column={pair_index * 2 + 1}",
                    point_count=point_count,
                    x_name="strain",
                    x_unit="%",
                    y_name="stress",
                    y_unit="MPa",
                    source_sha256=source_sha256,
                    status="semantic_duplicate_representation",
                    notes="两组曲线分别是MDI-2-1/2原始TXT的舍入副本；工作簿不重复计曲线。",
                )
            )
        workbook_summary["processed_tensile_label_nonunique"] = (
            len(set(labels)) != len(labels)
        )
    elif category == "network_calculation_workbook":
        workbook_summary["calculation_semantics"] = (
            "48个三重复试样槽、16个配方代码的m0/m1/m2、凝胶含量和"
            "Flory-Rehner交联密度；公式及缓存值完整，单位和1/2/4/6/8/10"
            "代码语义仍需论文/预印本闭合。"
        )
    return curves, workbook_summary


def _workbook_pair_values(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    sheet_name: str,
    min_row: int,
) -> list[list[tuple[float, float]]]:
    workbook = load_workbook(
        io.BytesIO(archive.read(info)), read_only=True, data_only=True
    )
    try:
        sheet = workbook[sheet_name]
        pairs: list[list[tuple[float, float]]] = [[], []]
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=min_row, max_col=4, values_only=True),
            start=min_row,
        ):
            for pair_index, start_index in enumerate((0, 2)):
                left = row[start_index]
                right = row[start_index + 1]
                if left is None and right in {None, "--"}:
                    continue
                if not isinstance(left, (int, float)) or not isinstance(
                    right, (int, float)
                ):
                    raise AuditBlocked(
                        f"工作簿曲线值漂移: {info.filename}#row={row_number}"
                    )
                pairs[pair_index].append((float(left), float(right)))
        return pairs
    finally:
        workbook.close()


def _maximum_pair_difference(
    left: list[tuple[float, float]], right: list[tuple[float, float]]
) -> tuple[float, float]:
    if len(left) != len(right):
        raise AuditBlocked(f"语义重复曲线长度不一致: {len(left)}/{len(right)}")
    return (
        max(abs(a[0] - b[0]) for a, b in zip(left, right, strict=True)),
        max(abs(a[1] - b[1]) for a, b in zip(left, right, strict=True)),
    )


def _semantic_duplicate_groups(
    archive: zipfile.ZipFile, infos: list[zipfile.ZipInfo]
) -> list[dict[str, Any]]:
    by_name = {PurePosixPath(info.filename).name: info for info in infos}
    required = {
        "GPC-Data.xlsx",
        "N-1.arw",
        "M-1.arw",
        "Tensile.xlsx",
        "MDI-2-1.txt",
        "MDI-2-2.txt",
    }
    if not required <= by_name.keys():
        raise AuditBlocked(
            f"语义重复核验资产缺失: {sorted(required - by_name.keys())}"
        )

    gpc_pairs = _workbook_pair_values(
        archive, by_name["GPC-Data.xlsx"], sheet_name="GPC", min_row=3
    )
    groups: list[dict[str, Any]] = []
    for pair_index, raw_name in enumerate(("N-1.arw", "M-1.arw")):
        raw_values = _two_column_values(
            _text_lines(archive, by_name[raw_name]), by_name[raw_name].filename
        )
        max_x, max_y = _maximum_pair_difference(raw_values, gpc_pairs[pair_index])
        if max_x > 1e-12 or max_y > 1e-12:
            raise AuditBlocked(f"GPC文本与工作簿曲线不一致: {raw_name}")
        groups.append(
            {
                "canonical_member_path": by_name[raw_name].filename,
                "duplicate_locator": (
                    f"{by_name['GPC-Data.xlsx'].filename}#column_pair={pair_index + 1}"
                ),
                "relation": "same_numeric_curve",
                "point_count": len(raw_values),
                "maximum_x_difference": max_x,
                "maximum_y_difference": max_y,
            }
        )

    processed_pairs = _workbook_pair_values(
        archive, by_name["Tensile.xlsx"], sheet_name="MDI", min_row=5
    )
    for pair_index, raw_name in enumerate(("MDI-2-1.txt", "MDI-2-2.txt")):
        stress, strain = _instrument_tensile_values(
            _text_lines(archive, by_name[raw_name]), by_name[raw_name].filename
        )
        raw_values = list(zip(strain, stress, strict=True))
        max_x, max_y = _maximum_pair_difference(
            raw_values, processed_pairs[pair_index]
        )
        if max_x > 1e-12 or max_y > 0.0005 + 1e-12:
            raise AuditBlocked(f"处理曲线不是原始TXT的预期舍入副本: {raw_name}")
        groups.append(
            {
                "canonical_member_path": by_name[raw_name].filename,
                "duplicate_locator": (
                    f"{by_name['Tensile.xlsx'].filename}#column_pair={pair_index + 1}"
                ),
                "relation": "rounded_numeric_curve_copy",
                "point_count": len(raw_values),
                "maximum_x_difference": max_x,
                "maximum_y_difference": max_y,
            }
        )
    return groups


def _network_measurements(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
) -> list[dict[str, Any]]:
    payload = archive.read(info)
    formulas = load_workbook(io.BytesIO(payload), read_only=False, data_only=False)
    values = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    try:
        formula_sheet = formulas["Crosslink density calculation"]
        value_sheet = values["Crosslink density calculation"]
        records: list[dict[str, Any]] = []
        for formulation_index, expected_code in enumerate(NETWORK_FORMULATION_CODES):
            start_column = 3 + formulation_index * 3
            source_code = str(value_sheet.cell(3, start_column).value).strip()
            if source_code != expected_code:
                raise AuditBlocked(
                    f"网络工作簿配方代码漂移: column={start_column}, "
                    f"expected={expected_code}, actual={source_code}"
                )
            for replicate_offset in range(3):
                column_index = start_column + replicate_offset
                column = get_column_letter(column_index)
                replicate = value_sheet.cell(4, column_index).value
                if replicate != replicate_offset + 1:
                    raise AuditBlocked(
                        f"网络工作簿重复编号漂移: {column}4={replicate}"
                    )
                numeric = {
                    name: value_sheet.cell(row, column_index).value
                    for name, row in {
                        "chi": 2,
                        "m0": 5,
                        "m1": 6,
                        "m2": 7,
                        "vr": 8,
                        "n_raw": 9,
                        "crosslink_density_reported_scaled": 10,
                        "gel_content": 13,
                    }.items()
                }
                if not all(isinstance(value, (int, float)) for value in numeric.values()):
                    raise AuditBlocked(
                        f"网络工作簿缓存数值缺失: {expected_code}::{replicate}"
                    )
                parsed = {key: float(value) for key, value in numeric.items()}
                m0, m1, m2 = parsed["m0"], parsed["m1"], parsed["m2"]
                vr = (m2 / 0.975) / (m2 / 0.975 + (m1 - m2) / 0.872)
                chi = 0.49 + 0.25 * vr
                n_raw = -(
                    math.log(1.0 - vr) + vr + chi * vr**2
                ) / (106.5 * (vr ** (1.0 / 3.0) - vr / 2.0))
                recalculated = {
                    "chi": chi,
                    "vr": vr,
                    "n_raw": n_raw,
                    "crosslink_density_reported_scaled": n_raw * 10_000.0,
                    "gel_content": m2 / m0,
                }
                for name, expected in recalculated.items():
                    if not math.isclose(
                        parsed[name], expected, rel_tol=1e-12, abs_tol=1e-15
                    ):
                        raise AuditBlocked(
                            f"网络工作簿公式缓存无法复算: {column}{name}="
                            f"{parsed[name]}/{expected}"
                        )
                expected_formulas = {
                    2: f"=0.49+0.25*{column}8",
                    8: (
                        f"=({column}7/0.975)/({column}7/0.975+"
                        f"({column}6-{column}7)/0.872)"
                    ),
                    9: (
                        f"=-(LN(1-{column}8)+{column}8+{column}2*{column}8^2)"
                        f"/(106.5*({column}8^(1/3)-{column}8/2))"
                    ),
                    10: f"={column}9*10000",
                    11: f"=-(LN(1-{column}8)+{column}8+{column}2*{column}8^2)",
                    12: (
                        f"=106.5*({column}8^(1/3)-{column}8/2)",
                        f"=106.5*({column}8^(1/3)-{column}8*0.5)",
                    ),
                    13: f"={column}7/{column}5",
                }
                for row, expected_formula in expected_formulas.items():
                    actual_formula = formula_sheet.cell(row, column_index).value
                    allowed_formulas = (
                        expected_formula
                        if isinstance(expected_formula, tuple)
                        else (expected_formula,)
                    )
                    if actual_formula not in allowed_formulas:
                        raise AuditBlocked(
                            f"网络工作簿公式漂移: {column}{row}="
                            f"{actual_formula!r}/{expected_formula!r}"
                        )
                records.append(
                    {
                        "formulation_code": expected_code,
                        "replicate_id": int(replicate),
                        "sample_id": f"{expected_code}::rep{int(replicate)}",
                        "source_column": column,
                        **parsed,
                        "normalized_swollen_mass_ratio": m1 / m0,
                        "reported_scale_factor": 10_000,
                    }
                )
        if len(records) != 48:
            raise AuditBlocked(f"网络工作簿试样数漂移: {len(records)}")
        return records
    finally:
        formulas.close()
        values.close()


def audit_archive() -> dict[str, Any]:
    if not ARCHIVE.is_file():
        raise AuditBlocked(f"固定ZIP不存在: {ARCHIVE}")
    if ARCHIVE.stat().st_size != ARCHIVE_BYTES or _sha256_file(ARCHIVE) != ARCHIVE_SHA256:
        raise AuditBlocked("固定ZIP字节数或SHA-256漂移")

    file_rows: list[dict[str, Any]] = []
    curve_rows: list[dict[str, Any]] = []
    workbook_summaries: list[dict[str, Any]] = []
    member_hashes: dict[str, list[str]] = defaultdict(list)
    semantic_duplicate_groups: list[dict[str, Any]] = []
    network_measurements: list[dict[str, Any]] = []
    with zipfile.ZipFile(ARCHIVE) as archive:
        if archive.testzip() is not None:
            raise AuditBlocked("ZIP CRC校验失败")
        infos = [info for info in archive.infolist() if not info.is_dir()]
        if len(infos) != EXPECTED_MEMBER_COUNT:
            raise AuditBlocked(f"ZIP文件数漂移: {len(infos)}")
        if len({info.filename for info in infos}) != len(infos):
            raise AuditBlocked("ZIP存在重复成员路径")
        if sum(info.file_size for info in infos) != EXPECTED_UNCOMPRESSED_BYTES:
            raise AuditBlocked("ZIP解压总字节数漂移")

        for info in infos:
            _safe_member(info)
            category = _category(info.filename)
            digest = _sha256_member(archive, info)
            member_hashes[digest].append(info.filename)
            point_count: int | str = ""
            name = PurePosixPath(info.filename).name
            if category in {"nmr_curve", "ftir_curve_csv", "gpc_text_curve"}:
                point_count = _two_column_curve(_text_lines(archive, info), info.filename)
                if category == "nmr_curve":
                    x_name, x_unit, y_name = "chemical_shift", "ppm", "source_signal"
                elif category == "ftir_curve_csv":
                    x_name, x_unit, y_name = "wavenumber", "cm^-1", "source_signal"
                else:
                    x_name, x_unit, y_name = (
                        "retention_time",
                        "min",
                        "source_reported_Mv_signal",
                    )
                curve_rows.append(
                    _curve_row(
                        info=info,
                        category=category,
                        curve_id=f"{category}:{name}",
                        point_count=point_count,
                        x_name=x_name,
                        x_unit=x_unit,
                        y_name=y_name,
                        y_unit="source_native",
                        source_sha256=digest,
                        status="conditional" if category == "gpc_text_curve" else "candidate",
                        notes=(
                            "ARW实际为双列文本曲线；Mv纵轴物理语义与校准尚待正文闭合。"
                            if category == "gpc_text_curve"
                            else ""
                        ),
                    )
                )
            elif category in {"tensile_raw", "hydrolytic_aging_tensile_raw"}:
                point_count = _instrument_tensile(_text_lines(archive, info), info.filename)
                curve_rows.append(
                    _curve_row(
                        info=info,
                        category=category,
                        curve_id=f"{category}:{name}",
                        point_count=point_count,
                        x_name="strain",
                        x_unit="%",
                        y_name="stress",
                        y_unit="MPa",
                        source_sha256=digest,
                    )
                )
            elif category == "cyclic_tensile_raw":
                point_count = _cyclic_tensile(_text_lines(archive, info), info.filename)
                curve_rows.append(
                    _curve_row(
                        info=info,
                        category=category,
                        curve_id=f"{category}:{name}",
                        point_count=point_count,
                        x_name="strain",
                        x_unit="%",
                        y_name="stress",
                        y_unit="MPa",
                        source_sha256=digest,
                    )
                )
            elif category == "uv_vis_curve":
                point_count = _uv_vis(_text_lines(archive, info), info.filename)
                curve_rows.append(
                    _curve_row(
                        info=info,
                        category=category,
                        curve_id=f"{category}:{name}",
                        point_count=point_count,
                        x_name="wavelength",
                        x_unit="nm",
                        y_name="absorbance",
                        y_unit="dimensionless",
                        source_sha256=digest,
                    )
                )
            elif category == "hydrodynamic_curve":
                point_count = _hydrodynamic(_text_lines(archive, info), info.filename)
                curve_rows.append(
                    _curve_row(
                        info=info,
                        category=category,
                        curve_id=f"{category}:{name}",
                        point_count=point_count,
                        x_name="time",
                        x_unit="s",
                        y_name="multi_channel_pressure_or_flow",
                        y_unit="mmHg_or_mL_per_s",
                        source_sha256=digest,
                    )
                )
            elif category in {
                "gpc_workbook",
                "processed_tensile_workbook",
                "network_calculation_workbook",
            }:
                workbook_curves, workbook_summary = _workbook_curves(
                    archive, info, category, digest
                )
                curve_rows.extend(workbook_curves)
                workbook_summaries.append(workbook_summary)

            file_rows.append(
                {
                    "member_path": info.filename,
                    "category": category,
                    "suffix": PurePosixPath(info.filename).suffix.lower(),
                    "uncompressed_bytes": info.file_size,
                    "compressed_bytes": info.compress_size,
                    "crc32": f"{info.CRC:08x}",
                    "sha256": digest,
                    "point_count": point_count,
                }
            )

        semantic_duplicate_groups = _semantic_duplicate_groups(archive, infos)
        network_info = next(
            info
            for info in infos
            if _category(info.filename) == "network_calculation_workbook"
        )
        network_measurements = _network_measurements(archive, network_info)

    suffix_counts = Counter(row["suffix"] for row in file_rows)
    category_counts = Counter(row["category"] for row in file_rows)
    if dict(sorted(suffix_counts.items())) != EXPECTED_SUFFIX_COUNTS:
        raise AuditBlocked(f"文件后缀数量漂移: {dict(suffix_counts)}")
    if dict(sorted(category_counts.items())) != EXPECTED_CATEGORY_COUNTS:
        raise AuditBlocked(f"科学类别数量漂移: {dict(category_counts)}")

    category_by_path = {row["member_path"]: row["category"] for row in file_rows}
    duplicate_groups: list[dict[str, Any]] = []
    for digest, paths in sorted(member_hashes.items()):
        if len(paths) <= 1:
            continue
        ordered_paths = sorted(
            paths,
            key=lambda path: (
                category_by_path[path] != "tensile_raw",
                path.casefold(),
            ),
        )
        canonical = ordered_paths[0]
        duplicate_groups.append(
            {
                "sha256": digest,
                "canonical_member_path": canonical,
                "member_paths": sorted(paths),
            }
        )
        for curve in curve_rows:
            if curve["member_path"] in ordered_paths[1:]:
                curve["audit_status"] = "exact_duplicate_representation"
                curve["notes"] = f"与{canonical}字节级完全重复；只保留血缘。"

    tensile_formulation_codes = sorted(
        {
            match.group(0)
            for row in file_rows
            if row["category"] == "tensile_raw"
            and (match := re.match(r"(?:HDI|HMDI|MDI)-(?:2|4|6|8|10)", PurePosixPath(row["member_path"]).name))
        }
    )
    candidate_curve_rows = [
        row
        for row in curve_rows
        if row["audit_status"] in {"candidate", "conditional"}
    ]
    gold_standard = next(
        row
        for row in network_measurements
        if row["formulation_code"] == "HDI-2" and row["replicate_id"] == 1
    )
    summary = {
        "audit_version": "batch18-iir-oh-pu-inventory-v2",
        "dataset_doi": DATASET_DOI,
        "license": LICENSE,
        "archive": {
            "filename": ARCHIVE.name,
            "bytes": ARCHIVE_BYTES,
            "sha256": ARCHIVE_SHA256,
            "member_count": len(file_rows),
            "uncompressed_bytes": sum(int(row["uncompressed_bytes"]) for row in file_rows),
            "crc_status": "pass",
            "path_safety_status": "pass",
        },
        "category_file_counts": dict(sorted(category_counts.items())),
        "suffix_counts": dict(sorted(suffix_counts.items())),
        "curve_count_audited": len(curve_rows),
        "curve_point_count_audited": sum(int(row["point_count"]) for row in curve_rows),
        "curve_count_candidate_after_dedup": len(candidate_curve_rows),
        "curve_point_count_candidate_after_dedup": sum(
            int(row["point_count"]) for row in candidate_curve_rows
        ),
        "curve_counts_by_category": dict(
            sorted(Counter(row["category"] for row in curve_rows).items())
        ),
        "curve_points_by_category": dict(
            sorted(
                {
                    category: sum(
                        int(row["point_count"])
                        for row in curve_rows
                        if row["category"] == category
                    )
                    for category in {row["category"] for row in curve_rows}
                }.items()
            )
        ),
        "tensile_formulation_codes": tensile_formulation_codes,
        "tensile_formulation_code_count": len(tensile_formulation_codes),
        "network_formulation_codes": list(NETWORK_FORMULATION_CODES),
        "network_formulation_code_count": len(NETWORK_FORMULATION_CODES),
        "all_formulation_code_count": len(NETWORK_FORMULATION_CODES),
        "duplicate_groups": duplicate_groups,
        "semantic_duplicate_groups": semantic_duplicate_groups,
        "duplicate_representation_curve_count": sum(
            row["audit_status"].endswith("duplicate_representation")
            for row in curve_rows
        ),
        "workbooks": workbook_summaries,
        "network_scalar_audit": {
            "sample_count": len(network_measurements),
            "resolved_final_scalar_count": len(network_measurements) * 2,
            "conditional_m1_over_m0_count": len(network_measurements),
            "gold_standard_hdi_2_rep1": gold_standard,
            "unit_status": "not_stated_in_workbook",
            "reported_scale_factor": 10_000,
        },
        "scientific_boundaries": [
            "2/4/6/8/10的配方计量语义尚待预印本或配方表闭合。",
            "Tensile.xlsx两组MDI-2曲线是原始TXT的舍入副本，GPC工作簿与文本ARW重复；均不重复计数。",
            "ARW已确认是可解析双列文本；SPA、PDF和OPJU保留为原始/证据资产，不从专有格式伪造数值。",
            "精确重复曲线保留血缘但物化时只计一次；曲线点不作为独立材料数。",
            "SSRN预印本关联是高置信推断，不是Mendeley仓库官方关联论文。",
        ],
        "training_split_created": False,
        "training_weight_materialized": False,
        "model_ready_record_count": 0,
    }
    return {"summary": summary, "files": file_rows, "curves": curve_rows}


def _gold_e_base_row(**updates: Any) -> dict[str, Any]:
    row: dict[str, Any] = {column: "" for column in RECORD_COLUMNS}
    row.update(
        {
            "source_directory": SOURCE_DIR.name,
            "target_origin": "experimental",
            "fidelity_level": "experimental",
            "current_weight_materialized": "false",
            "training_weight": "",
            "license": "CC BY 4.0",
            "citation_keys": CITATION_KEYS,
            "batch_id": "",
            "curve_id": "",
            "point_index": "",
            "secondary_condition_name": "",
            "secondary_condition_value": "",
            "secondary_condition_unit": "",
            "auxiliary_value_name": "",
            "auxiliary_value": "",
            "auxiliary_unit": "",
            "sample_identity_status": "source_filename_or_workbook_column",
            "global_structure_family_key": SOURCE_FAMILY_KEY,
            "family_leakage_group": SOURCE_FAMILY_KEY,
            "curve_points_are_independent_samples": "false",
            "duplicate_status": "unique_after_source_level_dedup",
        }
    )
    row.update(updates)
    return row


def _tensile_identity(name: str) -> tuple[str, str, int]:
    match = re.fullmatch(
        r"(?P<formulation>(?:HDI|HMDI|MDI)-(?:1|2|4|6|8|10))"
        r"(?:-(?P<state>Before|After))?-(?P<replicate>[123])\.txt",
        name,
    )
    if not match:
        raise AuditBlocked(f"拉伸文件身份无法解析: {name}")
    state = (match.group("state") or "baseline").lower()
    return match.group("formulation"), state, int(match.group("replicate"))


def _format_number(value: float) -> str:
    return format(value, ".15g")


def _build_gold_e_scalar_rows(
    audited: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    file_by_path = {row["member_path"]: row for row in audited["files"]}
    duplicate_paths = {
        path
        for group in audited["summary"]["duplicate_groups"]
        for path in group["member_paths"]
        if path != group["canonical_member_path"]
    }
    rows: list[dict[str, Any]] = []
    tensile_curve_count = 0
    network_info: zipfile.ZipInfo | None = None
    with zipfile.ZipFile(ARCHIVE) as archive:
        infos = sorted(
            (info for info in archive.infolist() if not info.is_dir()),
            key=lambda info: info.filename.casefold(),
        )
        for info in infos:
            category = _category(info.filename)
            if category == "network_calculation_workbook":
                network_info = info
            if category not in {"tensile_raw", "hydrolytic_aging_tensile_raw"}:
                continue
            if info.filename in duplicate_paths:
                continue
            name = PurePosixPath(info.filename).name
            formulation, state, replicate = _tensile_identity(name)
            stress, strain = _instrument_tensile_values(
                _text_lines(archive, info), info.filename
            )
            if any(
                right + 1e-12 < left
                for left, right in zip(strain, strain[1:])
            ):
                raise AuditBlocked(f"拉伸应变非单调，无法确定性积分: {info.filename}")
            tensile_curve_count += 1
            observed_area = sum(
                (left_stress + right_stress)
                * 0.5
                * (right_strain - left_strain)
                / 100.0
                for left_stress, right_stress, left_strain, right_strain in zip(
                    stress, stress[1:], strain, strain[1:]
                )
            )
            curve_id = (
                f"mendeley-wg3znh66bv:{formulation}:{state}:rep{replicate}:tensile"
            )
            sample_id = f"{formulation}|{state}|rep{replicate}"
            admission = (
                "admitted_reference"
                if category == "tensile_raw"
                else "conditional_reference"
            )
            endpoints = (
                ("maximum_observed_tensile_stress", max(stress), "MPa"),
                ("maximum_observed_tensile_strain", max(strain), "%"),
                (
                    "observed_stress_strain_area_to_last_point",
                    observed_area,
                    "MJ/m3",
                ),
            )
            for property_name, value, unit in endpoints:
                observation_id = f"{curve_id}:{property_name}"
                rows.append(
                    _gold_e_base_row(
                        source_record_id=observation_id,
                        observation_id=observation_id,
                        formulation_id=formulation,
                        sample_id=sample_id,
                        record_kind="curve_derived_scalar",
                        property_name=property_name,
                        value=_format_number(value),
                        unit=unit,
                        condition_name="source_curve_state",
                        condition_value=state,
                        data_origin="experimental_derived_scalar",
                        reduction_level="deterministic_curve_endpoint",
                        method_or_test_protocol=(
                            "source instrument tensile export; protocol details "
                            "pending preprint closure"
                        ),
                        gold_admission_status=admission,
                        mapping_status="formulation_code_only",
                        protocol_status="partial",
                        potential_weight_ceiling="0.45",
                        split_group=f"{DATASET_DOI}|{formulation}",
                        source_locator=info.filename,
                        file_sha256=file_by_path[info.filename]["sha256"],
                        curve_id=curve_id,
                        sample_identity_status=(
                            "filename_formulation_state_replicate"
                        ),
                        notes=(
                            "由唯一原始拉伸曲线确定性复算；峰值应力、最大观测应变和记录末点前曲线面积"
                            "均不冒充协议已闭合的拉伸强度、断裂伸长率或完整断裂韧性。"
                        ),
                    )
                )

        if network_info is None:
            raise AuditBlocked("网络计算工作簿缺失")
        network_sha256 = file_by_path[network_info.filename]["sha256"]
        network_measurements = _network_measurements(archive, network_info)
        property_specs = (
            (
                "gel_content",
                "gel_content",
                "1",
                "admitted_reference",
                "0.45",
                13,
                "工作簿缓存分数值并独立复算；Excel显示为百分数。",
            ),
            (
                "crosslink_density_reported_scaled",
                "crosslink_density_reported_scaled",
                "source_scale_unit_unresolved",
                "conditional_reference",
                "0.30",
                10,
                "来源按n×10000报告，但工作簿未声明物理单位；保留原尺度名称。",
            ),
            (
                "normalized_swollen_mass_ratio",
                "normalized_swollen_mass_ratio",
                "1",
                "conditional_reference",
                "0.15",
                7,
                "由m1/m0确定性复算；文件未声明m1语义，暂作条件参考。",
            ),
        )
        for measurement in network_measurements:
            formulation = str(measurement["formulation_code"])
            sample_id = str(measurement["sample_id"])
            column = str(measurement["source_column"])
            for property_name, value_key, unit, admission, ceiling, row_number, notes in (
                property_specs
            ):
                observation_id = (
                    f"mendeley-wg3znh66bv:{sample_id}:network:{property_name}"
                )
                if property_name == "normalized_swollen_mass_ratio":
                    locator = (
                        f"{network_info.filename}#sheet=Crosslink density calculation;"
                        f"derived_from={column}6/{column}5"
                    )
                else:
                    locator = (
                        f"{network_info.filename}#sheet=Crosslink density calculation;"
                        f"cell={column}{row_number}"
                    )
                rows.append(
                    _gold_e_base_row(
                        source_record_id=observation_id,
                        observation_id=observation_id,
                        formulation_id=formulation,
                        sample_id=sample_id,
                        record_kind="source_formula_scalar",
                        property_name=property_name,
                        value=_format_number(float(measurement[value_key])),
                        unit=unit,
                        data_origin="experimental_derived_scalar",
                        reduction_level="source_workbook_formula_scalar",
                        method_or_test_protocol=(
                            "source cached Excel formula independently recomputed"
                        ),
                        gold_admission_status=admission,
                        mapping_status="formulation_code_and_replicate",
                        protocol_status="partial",
                        potential_weight_ceiling=ceiling,
                        split_group=f"{DATASET_DOI}|{formulation}",
                        source_locator=locator,
                        file_sha256=network_sha256,
                        sample_identity_status="workbook_formulation_replicate_column",
                        notes=notes,
                    )
                )

    if tensile_curve_count != 54:
        raise AuditBlocked(f"唯一拉伸曲线数漂移: {tensile_curve_count}")
    if len(rows) != 306 or len({row["observation_id"] for row in rows}) != 306:
        raise AuditBlocked(f"Gold-E标量数量或身份漂移: {len(rows)}")
    admission_counts = Counter(row["gold_admission_status"] for row in rows)
    if admission_counts != {
        "admitted_reference": 183,
        "conditional_reference": 123,
    }:
        raise AuditBlocked(f"Gold-E标量准入计数漂移: {dict(admission_counts)}")
    summary = {
        "tensile_curve_count": tensile_curve_count,
        "derived_tensile_endpoint_count": tensile_curve_count * 3,
        "network_sample_count": 48,
        "network_scalar_count": 144,
        "gold_e_scalar_row_count": len(rows),
        "admission_counts": dict(admission_counts),
        "training_split_created": False,
        "training_weight_materialized": False,
        "model_ready_record_count": 0,
    }
    return rows, summary


def build_gold_e_scalar_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    audited = audit_archive()
    rows, materialization = _build_gold_e_scalar_rows(audited)
    audited["summary"]["gold_e_scalar_materialization"] = materialization
    return rows, audited["summary"]


def _atomic_write(path: Path, payload: bytes) -> None:
    if not path.parent.is_dir() or path.parent.is_symlink():
        raise AuditBlocked(f"输出目录不是普通目录: {path.parent}")
    if path.exists() and (not path.is_file() or path.is_symlink()):
        raise AuditBlocked(f"拒绝覆盖非普通文件: {path}")
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _tsv_bytes(rows: list[dict[str, Any]], columns: tuple[str, ...]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(
        output,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def write_outputs(
    bundle: dict[str, Any], gold_e_scalar_rows: list[dict[str, Any]] | None = None
) -> None:
    _atomic_write(
        SOURCE_DIR / "内容审计摘要.json",
        (
            json.dumps(bundle["summary"], ensure_ascii=False, indent=2, sort_keys=True)
            + "\n"
        ).encode("utf-8"),
    )
    file_columns = (
        "member_path",
        "category",
        "suffix",
        "uncompressed_bytes",
        "compressed_bytes",
        "crc32",
        "sha256",
        "point_count",
    )
    curve_columns = (
        "curve_id",
        "category",
        "member_path",
        "member_sha256",
        "point_count",
        "x_name",
        "x_unit",
        "y_name",
        "y_unit",
        "audit_status",
        "notes",
    )
    _atomic_write(SOURCE_DIR / "文件校验清单.tsv", _tsv_bytes(bundle["files"], file_columns))
    _atomic_write(SOURCE_DIR / "曲线审计清单.tsv", _tsv_bytes(bundle["curves"], curve_columns))
    if gold_e_scalar_rows is not None:
        _atomic_write(
            OUTPUT_GOLD_E_SCALARS,
            _tsv_bytes(gold_e_scalar_rows, GOLD_E_SCALAR_COLUMNS),
        )


def main() -> int:
    bundle = audit_archive()
    gold_e_scalar_rows, materialization = _build_gold_e_scalar_rows(bundle)
    bundle["summary"]["gold_e_scalar_materialization"] = materialization
    write_outputs(bundle, gold_e_scalar_rows)
    print(json.dumps(bundle["summary"], ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
