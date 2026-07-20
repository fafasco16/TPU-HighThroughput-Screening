"""只读复算 TPU 数据库 v0.2 第二批四个开放来源。

覆盖来源：

* ``Zenodo_标准化弹性体表征``
* ``Zenodo_PU微球复合材料拉伸``
* ``Figshare_PU高低速变形后应力松弛``
* ``Zenodo_TPU1301热黏弹黏塑本构``

本脚本不联网、不解压到原始数据树、不创建训练集。它对固定本地文件及 ZIP
执行哈希、CRC、成员路径、重复名、加密、符号链接、压缩比和总解压量硬门，
再按材料—试样—通道—曲线—点复算科学语义。只有全部输入通过后，才原子
覆盖各来源的三个白名单审计输出。

运行：

    python 代码/审计/新增开放数据第二批四源.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import itertools
import json
import math
import os
import pickletools
import re
import shutil
import stat
import subprocess
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "01_原始数据" / "外部数据" / "新增开放数据"
XLS_READER = Path(__file__).with_name("读取标准弹性体旧版XLS.ps1")
AUDIT_DATE = "2026-07-20"
AUDIT_VERSION = "1.0"

STANDARD = "Zenodo_标准化弹性体表征"
MICROSPHERE = "Zenodo_PU微球复合材料拉伸"
FIGSHARE = "Figshare_PU高低速变形后应力松弛"
TPU1301 = "Zenodo_TPU1301热黏弹黏塑本构"
SOURCE_NAMES = (STANDARD, MICROSPHERE, FIGSHARE, TPU1301)

OUTPUT_NAMES = ("内容审计摘要.json", "文件校验清单.tsv", "曲线审计清单.tsv")
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source in SOURCE_NAMES
    for filename in OUTPUT_NAMES
)

# ZIP 硬门是防御上限；每个固定归档还会核对精确条目数和精确解压字节数。
MAX_COMPRESSION_RATIO = 1_000.0
MAX_UNCOMPRESSED_BYTES = 1_500_000_000

EXPECTED_SCIENTIFIC_FILES: dict[str, dict[str, tuple[int, str]]] = {
    STANDARD: {
        "Curing.zip": (4_052_971, "70c6959eaa3dadbbf7b63332f72b8607d484f927c78d091d92fe630f1755cfbb"),
        "Glass transition.zip": (1_346_354, "098c85c35d1c41ea376393ebb8cb69138e94f90b1f7f2563f60f5f3f2f0729d3"),
        "Melting.zip": (308_782, "9d902b31027a36f9e6a38e5fa5873dce289d0d0cad0088ab49458aa8b21adda4"),
        "Stress relaxation.zip": (80_183_980, "79b9887b989bc8b420b53cd07b0ea9301b5f00def955febedf451932ee2fd3ed"),
        "Thermal degradation.zip": (1_158_247, "3a1c0b13d178db76bc6e1d2b600bcecd01b51a775673bc44f16e9b4674200c0d"),
        "Uniaxial compression.zip": (73_253, "7acde7e17883150264a528a55bced8d1828c8f70179066df81ad4f6545c5b856"),
        "Uniaxial tension.zip": (1_138_881, "b4f328793d248dd908a0097ca9ed091472278fd49514da91a6d0076d64503ac3"),
    },
    MICROSPHERE: {
        "Data_csv.zip": (780_946, "dfbc16ae5369eff16588fd8b2f45d47c4327c3a063aad4819e31eb61a572df92"),
        "readme.md": (3_480, "9e14e1c8d43dfbcb79a9186ec45e747f82a88cd510efe20eb3091ac7d95b68e6"),
    },
    FIGSHARE: {
        "rspa20220830_si_002.zip": (8_831_991, "146bfae71237b203d34b1fa9db46ff07e711dfdff86265cee7a940dac79d1194"),
    },
    TPU1301: {
        "ijss_2025_vevp_ScriptsForTestsImages.zip": (
            450_879_687,
            "988c4d2f972582b98be2d40e3ebc0d76538330ff9059aaff3f885d322cfec7ee",
        ),
    },
}

EXPECTED_ZIP_SHAPES = {
    "Curing.zip": (15, 14, 10_215_424),
    "Glass transition.zip": (11, 10, 7_497_696),
    "Melting.zip": (6, 3, 979_419),
    "Stress relaxation.zip": (11, 10, 297_074_920),
    "Thermal degradation.zip": (11, 10, 3_294_740),
    "Uniaxial compression.zip": (41, 40, 243_482),
    "Uniaxial tension.zip": (53, 52, 3_502_774),
    "Data_csv.zip": (100, 96, 3_316_400),
    "rspa20220830_si_002.zip": (53, 53, 26_105_915),
    "ijss_2025_vevp_ScriptsForTestsImages.zip": (614, 549, 1_381_087_479),
}

STANDARD_CORE_LABELS = {
    "Cheetah",
    "Dragon Skin 20",
    "Dragon Skin 30",
    "Ecoflex 00-20",
    "Ecoflex 00-30",
    "Ecoflex 00-50",
    "Filaflex 60A",
    "FsCO-BMI689",
    "Mold Max 14NV",
    "Mold Star 30",
}
STANDARD_ALL_LABELS = STANDARD_CORE_LABELS | {"NinjaFlex 90A"}
TARGET_STANDARD_LABELS = {"Cheetah", "Filaflex 60A"}

MICRO_CONDITIONS = ("00", "05", "10", "15", "20", "25")
MICRO_SPECIMENS = {
    "poro_00_spec_02b",
    "poro_00_spec_03b",
    *(f"poro_{condition}_spec_{replicate}" for condition in MICRO_CONDITIONS[1:] for replicate in ("02", "03")),
}

TPU_PA12_EXCLUSION = "Uniaxial_compression_2CV_2p78E-3_RT_PA12.csv"
TPU_IDENTITY_CONFLICT = {
    "filename": "Relaxation_7H_1E-1_RT_TPU.csv",
    "header_label": "6V",
    "decision": "quarantine_identity_conflict",
}


class AuditBlocked(RuntimeError):
    """输入、完整性或科学语义不满足固定审计协议。"""


def _same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def _is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & flag) or bool(is_junction())


def _assert_plain_chain(path: Path, stop: Path) -> None:
    if path != stop and stop not in path.parents:
        raise AuditBlocked(f"路径越出审计根：{path}")
    cursor = path
    while True:
        if _is_reparse_point(cursor):
            raise AuditBlocked(f"拒绝符号链接或重解析点：{cursor}")
        if cursor == stop:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"目录缺失、不是普通目录或经链接解析：{path}")
    _assert_plain_chain(path, PROJECT_ROOT)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    _assert_plain_chain(path, PROJECT_ROOT)


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")

    # 保持函数自包含，安全测试会只抽取本定义、AuditBlocked 与 atomic_write。
    def is_reparse(candidate: Path) -> bool:
        try:
            details = candidate.lstat()
        except FileNotFoundError:
            return False
        attributes = getattr(details, "st_file_attributes", 0)
        is_junction = getattr(candidate, "is_junction", lambda: False)
        return (
            candidate.is_symlink()
            or bool(attributes & 0x400)
            or bool(is_junction())
        )

    parent = path.parent
    try:
        resolved_parent = parent.resolve(strict=True)
    except FileNotFoundError as exc:
        raise AuditBlocked(f"审计输出目录不存在：{parent}") from exc
    if (
        not parent.is_dir()
        or is_reparse(parent)
        or os.path.normcase(os.path.abspath(str(resolved_parent)))
        != os.path.normcase(os.path.abspath(str(parent)))
    ):
        raise AuditBlocked(f"拒绝通过符号链接或重解析目录写入：{parent}")
    if path.exists() or path.is_symlink():
        if is_reparse(path):
            raise AuditBlocked(f"拒绝覆盖符号链接或重解析输出：{path}")
        resolved = path.resolve(strict=True)
        if (
            not path.is_file()
            or os.path.normcase(os.path.abspath(str(resolved)))
            != os.path.normcase(os.path.abspath(str(path)))
        ):
            raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        details = temporary.lstat()
        attributes = getattr(details, "st_file_attributes", 0)
        is_junction = getattr(temporary, "is_junction", lambda: False)
        if (
            temporary.is_symlink()
            or bool(attributes & 0x400)
            or bool(is_junction())
            or not temporary.is_file()
            or os.path.normcase(os.path.abspath(str(temporary.resolve(strict=True))))
            != os.path.normcase(os.path.abspath(str(temporary)))
        ):
            raise AuditBlocked(f"审计临时输出不是普通文件：{temporary}")
        assert_output_allowed(path)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_json(path: Path, payload: dict[str, object]) -> None:
    rendered = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False
    )
    atomic_write(path, (rendered + "\n").encode("utf-8"))


def render_tsv(rows: list[dict[str, object]], columns: list[str]) -> bytes:
    if not rows:
        raise AuditBlocked("拒绝渲染空 TSV")
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _hash_stream(handle: BinaryIO, algorithm: str = "sha256") -> str:
    digest = hashlib.new(algorithm)
    for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
        digest.update(block)
    return digest.hexdigest()


def file_hash(path: Path, algorithm: str = "sha256") -> str:
    require_file(path)
    with path.open("rb") as handle:
        return _hash_stream(handle, algorithm)


def _manifest_digest(items: Iterable[tuple[str, int, str]]) -> str:
    digest = hashlib.sha256()
    for relative, size, checksum in sorted(items):
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(str(size).encode("ascii"))
        digest.update(b"\0")
        digest.update(checksum.encode("ascii"))
        digest.update(b"\n")
    return digest.hexdigest()


def scientific_input_snapshot() -> dict[str, tuple[int, str]]:
    snapshot: dict[str, tuple[int, str]] = {}
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        require_directory(base)
        expected_names = set(EXPECTED_SCIENTIFIC_FILES[source]) | {
            "官方API元数据.json",
            "官方文件清单.tsv",
        }
        actual_names: set[str] = set()
        for path in sorted(base.iterdir(), key=lambda item: item.name.casefold()):
            if _is_reparse_point(path):
                raise AuditBlocked(f"来源根含符号链接或重解析点：{path}")
            if path in OUTPUT_WHITELIST:
                continue
            if not path.is_file():
                raise AuditBlocked(f"来源根出现未登记目录或特殊对象：{path}")
            if path.name.endswith((".part", ".audit.tmp")):
                raise AuditBlocked(f"来源根出现未完成临时文件：{path}")
            actual_names.add(path.name)
            relative = path.relative_to(PROJECT_ROOT).as_posix()
            snapshot[relative] = (path.stat().st_size, file_hash(path))
        if actual_names != expected_names:
            raise AuditBlocked(
                f"{source}本地输入集合漂移：缺失={sorted(expected_names-actual_names)}，"
                f"多余={sorted(actual_names-expected_names)}"
            )
        for filename, (size, sha256) in EXPECTED_SCIENTIFIC_FILES[source].items():
            path = base / filename
            if path.stat().st_size != size or file_hash(path) != sha256:
                raise AuditBlocked(f"固定科学文件大小或SHA256不匹配：{source}/{filename}")
    return snapshot


def snapshot_by_source(snapshot: dict[str, tuple[int, str]]) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for source in SOURCE_NAMES:
        marker = f"/新增开放数据/{source}/"
        items = [
            (path, size, checksum)
            for path, (size, checksum) in snapshot.items()
            if marker in f"/{path}"
        ]
        result[source] = {
            "文件数": len(items),
            "总字节数": sum(size for _, size, _ in items),
            "清单SHA256": _manifest_digest(items),
        }
    return result


def _safe_zip_name(name: str) -> str:
    if "\x00" in name or re.match(r"^[A-Za-z]:", name):
        raise AuditBlocked(f"ZIP含危险成员路径：{name!r}")
    candidate = PurePosixPath(name.replace("\\", "/"))
    if candidate.is_absolute() or ".." in candidate.parts:
        raise AuditBlocked(f"ZIP含危险成员路径：{name!r}")
    normalized = candidate.as_posix()
    if normalized in {"", "."}:
        raise AuditBlocked(f"ZIP含空成员路径：{name!r}")
    return normalized


def audit_zip(path: Path) -> tuple[dict[str, object], list[dict[str, object]]]:
    require_file(path)
    expected_shape = EXPECTED_ZIP_SHAPES.get(path.name)
    if expected_shape is None:
        raise AuditBlocked(f"未冻结ZIP形状：{path.name}")
    rows: list[dict[str, object]] = []
    with zipfile.ZipFile(path) as archive:
        bad = archive.testzip()
        if bad is not None:
            raise AuditBlocked(f"ZIP CRC失败：{path.name}/{bad}")
        infos = archive.infolist()
        normalized = [_safe_zip_name(info.filename) for info in infos]
        duplicate_names = [
            name for name, count in Counter(normalized).items() if count > 1
        ]
        if duplicate_names:
            raise AuditBlocked(f"ZIP含重复成员名：{path.name}/{duplicate_names}")
        total_uncompressed = 0
        file_count = 0
        max_ratio = 0.0
        extensions: Counter[str] = Counter()
        for name, info in zip(normalized, infos):
            mode = (info.external_attr >> 16) & 0o170000
            if mode == stat.S_IFLNK:
                raise AuditBlocked(f"ZIP含符号链接成员：{path.name}/{name}")
            if info.flag_bits & 0x1:
                raise AuditBlocked(f"ZIP含加密成员：{path.name}/{name}")
            if info.is_dir():
                continue
            file_count += 1
            total_uncompressed += info.file_size
            ratio = (
                float("inf")
                if info.compress_size == 0 and info.file_size > 0
                else info.file_size / max(info.compress_size, 1)
            )
            max_ratio = max(max_ratio, ratio)
            if ratio > MAX_COMPRESSION_RATIO:
                raise AuditBlocked(f"ZIP成员压缩比超限：{path.name}/{name}={ratio:.2f}")
            suffix = PurePosixPath(name).suffix.lower() or "[none]"
            extensions[suffix] += 1
            rows.append(
                {
                    "归档": path.name,
                    "成员": name,
                    "未压缩字节": info.file_size,
                    "压缩字节": info.compress_size,
                    "CRC32": f"{info.CRC:08x}",
                    "扩展名": suffix,
                    "角色": "待来源语义分类",
                }
            )
        if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
            raise AuditBlocked(f"ZIP总解压量超限：{path.name}={total_uncompressed}")
        shape = (len(infos), file_count, total_uncompressed)
        if shape != expected_shape:
            raise AuditBlocked(f"ZIP固定形状漂移：{path.name}={shape}/{expected_shape}")
    return (
        {
            "归档": path.name,
            "ZIP_CRC_testzip": "通过；无坏成员",
            "条目数": expected_shape[0],
            "文件数": expected_shape[1],
            "目录数": expected_shape[0] - expected_shape[1],
            "解压后字节数": expected_shape[2],
            "最大成员压缩比": round(max_ratio, 6),
            "危险路径数": 0,
            "重复成员路径数": 0,
            "加密成员数": 0,
            "符号链接成员数": 0,
            "扩展名盘点": dict(sorted(extensions.items())),
        },
        rows,
    )


def _finite(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip().replace("−", "-")
    if not text or text.casefold() in {"nan", "inf", "-inf", "--"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _text_encoding(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> str:
    with archive.open(info) as handle:
        sample = handle.read(64 * 1024)
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise AuditBlocked(f"文本编码不在固定白名单：{info.filename}")


def _physical_line_count(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> int:
    count = 0
    last = b""
    with archive.open(info) as handle:
        for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
            count += block.count(b"\n")
            last = block[-1:]
    return count + (1 if info.file_size and last != b"\n" else 0)


def _first_lines(
    archive: zipfile.ZipFile, info: zipfile.ZipInfo, count: int
) -> list[str]:
    encoding = _text_encoding(archive, info)
    result: list[str] = []
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as handle:
        for _ in range(count):
            line = handle.readline()
            if line == "":
                break
            result.append(line.rstrip("\r\n"))
    return result


def _count_numeric_csv(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    header_rows: int,
    delimiter: str = ",",
    minimum_columns: int,
) -> tuple[int, int]:
    encoding = _text_encoding(archive, info)
    data_rows = 0
    invalid_rows = 0
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as text:
        reader = csv.reader(text, delimiter=delimiter)
        for index, row in enumerate(reader):
            if index < header_rows or not row or not any(cell.strip() for cell in row):
                continue
            data_rows += 1
            if len(row) < minimum_columns or any(
                _finite(cell) is None for cell in row[:minimum_columns]
            ):
                invalid_rows += 1
    return data_rows, invalid_rows


def _governance(
    *,
    fidelity: str,
    split_group_key: str,
    source_weight_ceiling: float,
    candidate_eligible: bool,
    note: str,
) -> dict[str, object]:
    return {
        "data_fidelity": fidelity,
        "split_group_key": split_group_key,
        "source_weight_ceiling": source_weight_ceiling,
        "candidate_eligible_after_governance_materialization": candidate_eligible,
        "training_split_created": False,
        "training_weight_materialized": False,
        "note": note,
    }


def _count_csv_selected(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    header_rows: int,
    delimiter: str,
    numeric_columns: tuple[int, ...],
) -> tuple[int, int]:
    encoding = _text_encoding(archive, info)
    data_rows = 0
    invalid_rows = 0
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as text:
        reader = csv.reader(text, delimiter=delimiter)
        for index, row in enumerate(reader):
            if index < header_rows or not row or not any(cell.strip() for cell in row):
                continue
            data_rows += 1
            if not numeric_columns or max(numeric_columns) >= len(row) or any(
                _finite(row[column]) is None for column in numeric_columns
            ):
                invalid_rows += 1
    return data_rows, invalid_rows


def _standard_material_label(archive_name: str, member: str) -> str:
    stem = PurePosixPath(member).stem
    if archive_name == "Curing.zip":
        stem = re.sub(r"-\d+°C$", "", stem)
    elif archive_name == "Stress relaxation.zip":
        stem = re.sub(r"-relaxation$", "", stem)
    elif archive_name in {"Uniaxial compression.zip", "Uniaxial tension.zip"}:
        stem = re.sub(r"\s+-\s+\d+$", "", stem)
    return {
        "Dragon skin 30": "Dragon Skin 30",
        "Mold Max 14 NV": "Mold Max 14NV",
    }.get(stem, stem)


def _column_finite_counts(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    header_rows: int,
) -> tuple[int, list[int]]:
    encoding = _text_encoding(archive, info)
    rows = 0
    counts: list[int] = []
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as text:
        reader = csv.reader(text)
        for index, row in enumerate(reader):
            if index < header_rows or not row or not any(cell.strip() for cell in row):
                continue
            rows += 1
            if len(counts) < len(row):
                counts.extend([0] * (len(row) - len(counts)))
            for column, value in enumerate(row):
                if _finite(value) is not None:
                    counts[column] += 1
    return rows, counts


def _count_exact_pair(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    pair: tuple[float, float],
) -> int:
    encoding = _text_encoding(archive, info)
    matches = 0
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as text:
        reader = csv.reader(text)
        next(reader, None)
        for row in reader:
            if len(row) >= 2 and (_finite(row[0]), _finite(row[1])) == pair:
                matches += 1
    return matches


def _find_powershell_for_xls() -> str:
    """优先选择能按 UTF-8 读取中文脚本路径的 PowerShell 7。"""
    for executable in ("pwsh.exe", "powershell.exe"):
        resolved = shutil.which(executable)
        if resolved is None:
            continue
        try:
            probe = subprocess.run(
                [resolved, "-NoProfile", "-Command", "$PSVersionTable.PSVersion.ToString()"],
                check=False,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=20,
            )
        except (OSError, subprocess.TimeoutExpired):
            continue
        if probe.returncode == 0:
            return resolved
    raise AuditBlocked("找不到可用 PowerShell，无法只读解析标准弹性体旧版 XLS")


def _audit_standard_legacy_xls() -> dict[str, object]:
    require_file(XLS_READER)
    command = [
        _find_powershell_for_xls(),
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(XLS_READER),
    ]
    try:
        completed = subprocess.run(
            command,
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=300,
        )
    except subprocess.TimeoutExpired as exc:
        raise AuditBlocked("标准弹性体旧版 XLS Excel COM 审计超时") from exc
    stderr = completed.stderr.decode("utf-8", errors="replace").strip()
    if completed.returncode != 0:
        raise AuditBlocked(
            f"标准弹性体旧版 XLS Excel COM 审计失败（{completed.returncode}）：{stderr}"
        )
    stdout = completed.stdout.decode("utf-8-sig", errors="strict").strip()
    try:
        profile = json.loads(stdout)
    except json.JSONDecodeError as exc:
        raise AuditBlocked(f"标准弹性体 XLS 辅助脚本未返回合法 JSON：{stdout[:500]}") from exc
    if not isinstance(profile, dict):
        raise AuditBlocked(f"标准弹性体 XLS 审计结果类型错误：{type(profile)}")

    exact_scalars = {
        "source_archive": "Melting.zip",
        "source_archive_sha256": EXPECTED_SCIENTIFIC_FILES[STANDARD]["Melting.zip"][1],
        "member": "Melting/viscosity/Filaflex 60A.xls",
        "member_sha256": "3fc855fb76b452a1768df8b18e9edc270843bb9b370fd59213f6b9e2e3dc0295",
        "read_only": True,
        "workbook_sheet_count": 17,
        "curve_count": 16,
        "curve_point_count": 2_094,
    }
    for key, expected in exact_scalars.items():
        if profile.get(key) != expected:
            raise AuditBlocked(f"标准弹性体 XLS 字段漂移：{key}={profile.get(key)!r}/{expected!r}")
    expected_group_curves = {
        "Temperature ramp 1": 5,
        "Flow ramp 6": 1,
        "Temperature ramp 3": 5,
        "Temperature ramp 4": 5,
    }
    expected_group_points = {
        "Temperature ramp 1": 462,
        "Flow ramp 6": 22,
        "Temperature ramp 3": 1_194,
        "Temperature ramp 4": 416,
    }
    if profile.get("group_curve_counts") != expected_group_curves:
        raise AuditBlocked(f"标准弹性体 XLS 曲线组数漂移：{profile.get('group_curve_counts')}")
    if profile.get("group_point_counts") != expected_group_points:
        raise AuditBlocked(f"标准弹性体 XLS 曲线组点数漂移：{profile.get('group_point_counts')}")

    expected_curve_points = {
        "Temperature ramp - 1 - 0,312599 Hz": 93,
        "Temperature ramp - 1 - 0,562301 Hz": 93,
        "Temperature ramp - 1 - 1,0 Hz": 92,
        "Temperature ramp - 1 - 1,778 Hz": 92,
        "Temperature ramp - 1 - 3,12599 Hz": 92,
        "Flow ramp - 6": 22,
        "Temperature ramp - 3 - 0,312599 Hz": 239,
        "Temperature ramp - 3 - 0,562301 Hz": 239,
        "Temperature ramp - 3 - 1,0 Hz": 239,
        "Temperature ramp - 3 - 1,778 Hz": 239,
        "Temperature ramp - 3 - 3,12599 Hz": 238,
        "Temperature ramp - 4 - 0,312599 Hz": 84,
        "Temperature ramp - 4 - 0,562301 Hz": 83,
        "Temperature ramp - 4 - 1,0 Hz": 83,
        "Temperature ramp - 4 - 1,778 Hz": 83,
        "Temperature ramp - 4 - 3,12599 Hz": 83,
    }
    curves = profile.get("curves")
    if not isinstance(curves, list) or len(curves) != 16:
        raise AuditBlocked("标准弹性体 XLS 曲线记录数漂移")
    actual_curve_points: dict[str, int] = {}
    for curve in curves:
        if not isinstance(curve, dict):
            raise AuditBlocked("标准弹性体 XLS 曲线记录不是对象")
        curve_id = str(curve.get("curve_id", ""))
        digest = str(curve.get("data_sha256", ""))
        if not re.fullmatch(r"[0-9a-f]{64}", digest):
            raise AuditBlocked(f"标准弹性体 XLS 曲线哈希异常：{curve_id}/{digest}")
        if curve_id in actual_curve_points:
            raise AuditBlocked(f"标准弹性体 XLS 曲线ID重复：{curve_id}")
        actual_curve_points[curve_id] = int(curve.get("point_count", -1))
    if actual_curve_points != expected_curve_points:
        raise AuditBlocked(f"标准弹性体 XLS 逐曲线点数漂移：{actual_curve_points}")
    return profile


def audit_standard() -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    base = DATA_ROOT / STANDARD
    xls_profile = _audit_standard_legacy_xls()
    zip_summaries: list[dict[str, object]] = []
    file_rows: list[dict[str, object]] = []
    material_modalities: defaultdict[str, set[str]] = defaultdict(set)
    curve_rows: list[dict[str, object]] = []
    target_points: Counter[str] = Counter()
    target_usable_points: Counter[str] = Counter()
    target_csv_count: Counter[str] = Counter()
    target_xls_curve_count: Counter[str] = Counter()
    target_xls_points: Counter[str] = Counter()
    identified_mechanical_specimens = 0
    ninjaflex_empty_response: dict[str, object] | None = None

    for archive_name in EXPECTED_SCIENTIFIC_FILES[STANDARD]:
        archive_path = base / archive_name
        zip_summary, member_rows = audit_zip(archive_path)
        zip_summaries.append(zip_summary)
        for row in member_rows:
            label = _standard_material_label(archive_name, str(row["成员"]))
            if label not in STANDARD_ALL_LABELS:
                raise AuditBlocked(
                    f"标准弹性体出现未冻结材料标签：{archive_name}/{row['成员']} -> {label}"
                )
            material_modalities[label].add(archive_name.removesuffix(".zip"))
            row["角色"] = "全材料模态证据"
            if label in TARGET_STANDARD_LABELS:
                row["角色"] = "目标TPE实验测量"
            file_rows.append(row)

        with zipfile.ZipFile(archive_path) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                label = _standard_material_label(archive_name, info.filename)
                suffix = PurePosixPath(info.filename).suffix.lower()
                if label == "NinjaFlex 90A":
                    if info.filename != "Melting/DSC/NinjaFlex 90A.csv":
                        raise AuditBlocked(f"NinjaFlex伴随文件角色漂移：{info.filename}")
                    rows, finite_counts = _column_finite_counts(
                        archive, info, header_rows=1
                    )
                    if rows != 12_237 or finite_counts[:3] != [12_237, 12_236, 0]:
                        raise AuditBlocked(
                            f"NinjaFlex空响应形态漂移：rows={rows}, finite={finite_counts}"
                        )
                    ninjaflex_empty_response = {
                        "文件": info.filename,
                        "数据行": rows,
                        "Time有限": finite_counts[0],
                        "Temperature有限": finite_counts[1],
                        "Heat Flow有限": finite_counts[2],
                        "完整性质点": 0,
                        "状态": "quarantine_empty_response_column",
                    }
                    continue
                if label not in TARGET_STANDARD_LABELS:
                    continue
                if suffix == ".xls":
                    if info.filename != "Melting/viscosity/Filaflex 60A.xls":
                        # Curing 中没有两个目标热塑弹性体，命中其他目标 XLS 即范围漂移。
                        raise AuditBlocked(f"目标TPE出现未知XLS角色：{info.filename}")
                    for curve in xls_profile["curves"]:
                        curve_id = str(curve["curve_id"])
                        points = int(curve["point_count"])
                        target_xls_curve_count[label] += 1
                        target_xls_points[label] += points
                        curve_rows.append(
                            {
                                "来源": STANDARD,
                                "材料": label,
                                "工况或试样": str(curve["group"]),
                                "数据角色": "实验熔体黏弹多变量曲线",
                                "曲线或文件": f"{info.filename}::{curve_id}",
                                "点数": points,
                                "试样组键": (
                                    f"10.5281/zenodo.14983287|{label}|viscosity|{curve_id}"
                                ),
                                "准入状态": "candidate_after_grouped_split",
                                "备注": (
                                    "Excel COM只读复算；一个工作表是一条同步多变量曲线；"
                                    "点数按同步行计，不按A/B/C响应通道重复计数；"
                                    f"data_sha256={curve['data_sha256']}"
                                ),
                            }
                        )
                    continue
                if suffix != ".csv":
                    raise AuditBlocked(f"目标TPE出现未知文件角色：{info.filename}")

                expected_header, columns = {
                    "Glass transition.zip": (
                        "Time (min),Temperature (°C),Heat Flow (W/g)",
                        3,
                    ),
                    "Melting.zip": (
                        "Time (min),Temperature (°C),Heat Flow (W/g)",
                        3,
                    ),
                    "Stress relaxation.zip": (
                        "Force (N),Strain (%),Stress (MPa),Time (s)",
                        4,
                    ),
                    "Thermal degradation.zip": ("Temperature (C),Mass (%)", 2),
                    "Uniaxial compression.zip": ("Strain (%),Stress (MPa)", 2),
                    "Uniaxial tension.zip": ("Strain (%),Stress (MPa)", 2),
                }.get(archive_name, ("", 0))
                if not expected_header:
                    raise AuditBlocked(f"目标TPE CSV处于未允许归档：{info.filename}")
                if _first_lines(archive, info, 1) != [expected_header]:
                    raise AuditBlocked(f"目标TPE表头漂移：{info.filename}")
                points, invalid = _count_numeric_csv(
                    archive,
                    info,
                    header_rows=1,
                    minimum_columns=columns,
                )
                expected_invalid = {
                    "Glass transition/Cheetah.csv": 7,
                    "Glass transition/Filaflex 60A.csv": 7,
                    "Melting/DSC/Filaflex 60A.csv": 7,
                }.get(info.filename, 0)
                if invalid != expected_invalid:
                    raise AuditBlocked(
                        f"目标TPE非完整行数漂移：{info.filename}={invalid}/{expected_invalid}"
                    )
                sentinel_rows = 0
                if archive_name == "Thermal degradation.zip":
                    sentinel_rows = _count_exact_pair(archive, info, (2.0, 0.0))
                    if sentinel_rows != 1:
                        raise AuditBlocked(f"TGA哨兵行数漂移：{info.filename}={sentinel_rows}")
                target_points[label] += points
                target_usable_points[label] += points - invalid - sentinel_rows
                target_csv_count[label] += 1
                modality = archive_name.removesuffix(".zip")
                replicate = re.search(r"\s+-\s+(\d+)\.csv$", info.filename)
                if replicate:
                    identified_mechanical_specimens += 1
                curve_rows.append(
                    {
                        "来源": STANDARD,
                        "材料": label,
                        "工况或试样": modality,
                        "数据角色": "实验原始曲线",
                        "曲线或文件": info.filename,
                        "点数": points - invalid - sentinel_rows,
                        "试样组键": f"10.5281/zenodo.14983287|{label}|{PurePosixPath(info.filename).stem}",
                        "准入状态": "candidate_after_grouped_split",
                        "备注": (
                            f"原始行{points}；缺列尾行{invalid}；TGA哨兵{sentinel_rows}；"
                            "单文件采样点不增加材料权重"
                        ),
                    }
                )

    if set(material_modalities) != STANDARD_ALL_LABELS:
        raise AuditBlocked(
            f"标准弹性体材料集合漂移：{sorted(material_modalities)}/{sorted(STANDARD_ALL_LABELS)}"
        )
    if material_modalities["NinjaFlex 90A"] != {"Melting"}:
        raise AuditBlocked("NinjaFlex 90A不再是仅熔融模态的伴随材料")
    if ninjaflex_empty_response is None:
        raise AuditBlocked("未复算NinjaFlex 90A空响应伴随文件")
    if target_csv_count != Counter({"Cheetah": 12, "Filaflex 60A": 14}):
        raise AuditBlocked(f"目标TPE CSV曲线数漂移：{target_csv_count}")
    if target_xls_curve_count != Counter({"Filaflex 60A": 16}):
        raise AuditBlocked(f"目标TPE XLS曲线数漂移：{target_xls_curve_count}")
    if target_xls_points != Counter({"Filaflex 60A": 2_094}):
        raise AuditBlocked(f"目标TPE XLS同步点数漂移：{target_xls_points}")
    expected_points = Counter({"Cheetah": 655_323, "Filaflex 60A": 684_446})
    if target_points != expected_points:
        raise AuditBlocked(f"目标TPE点数漂移：{target_points}/{expected_points}")
    expected_usable = Counter({"Cheetah": 655_315, "Filaflex 60A": 684_431})
    if target_usable_points != expected_usable:
        raise AuditBlocked(
            f"目标TPE可用完整点数漂移：{target_usable_points}/{expected_usable}"
        )
    if identified_mechanical_specimens != 19:
        raise AuditBlocked(f"目标TPE显式编号拉伸/压缩试样数漂移：{identified_mechanical_specimens}")

    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": STANDARD,
        "DOI": "10.5281/zenodo.14983287",
        "许可": "CC BY 4.0",
        "ZIP审计": zip_summaries,
        "材料边界": {
            "官方描述核心材料数": 10,
            "实际唯一材料标签数": 11,
            "核心材料": sorted(STANDARD_CORE_LABELS),
            "伴随材料": ["NinjaFlex 90A"],
            "伴随材料空响应审计": ninjaflex_empty_response,
            "目标热塑弹性体": sorted(TARGET_STANDARD_LABELS),
            "说明": "实际文件多出仅在Melting出现的NinjaFlex 90A；不得隐去或冒充完整多模态材料。",
        },
        "目标TPE实测": {
            "CSV曲线数": sum(target_csv_count.values()),
            "旧版XLS黏度文件数": 1,
            "旧版XLS黏度多变量曲线数": sum(target_xls_curve_count.values()),
            "旧版XLS黏度同步点数": sum(target_xls_points.values()),
            "旧版XLS只读审计": {
                "workbook_sheet_count": xls_profile["workbook_sheet_count"],
                "curve_count": xls_profile["curve_count"],
                "curve_point_count": 2_094,
                "group_curve_counts": xls_profile["group_curve_counts"],
                "group_point_counts": xls_profile["group_point_counts"],
                "member_sha256": xls_profile["member_sha256"],
                "read_only": xls_profile["read_only"],
            },
            "CSV原始数据行数": sum(target_points.values()),
            "CSV可用完整点数": sum(target_usable_points.values()),
            "各材料CSV原始行数": dict(sorted(target_points.items())),
            "各材料CSV可用完整点数": dict(sorted(target_usable_points.items())),
            "各材料实验原始行或同步点": {
                "Cheetah": target_points["Cheetah"],
                "Filaflex 60A": target_points["Filaflex 60A"] + target_xls_points["Filaflex 60A"],
            },
            "各材料实验可用行或同步点": {
                "Cheetah": target_usable_points["Cheetah"],
                "Filaflex 60A": (
                    target_usable_points["Filaflex 60A"] + target_xls_points["Filaflex 60A"]
                ),
            },
            "实验曲线总数": sum(target_csv_count.values()) + sum(target_xls_curve_count.values()),
            "显式编号拉伸压缩重复曲线数": identified_mechanical_specimens,
            "physical_specimen_count": None,
            "跨模态物理试样身份": "未提供；单次DSC/TGA/松弛/黏度测量不得与编号试样擅自合并或拆分",
        },
        "治理": _governance(
            fidelity="direct_experiment_commercial_TPE",
            split_group_key="dataset_doi|material_grade|test_mode|specimen_or_run",
            source_weight_ceiling=0.35,
            candidate_eligible=True,
            note="牌号清楚但化学配方不公开；旧版XLS已通过固定成员哈希与Excel COM只读复算。",
        ),
    }
    return summary, file_rows, curve_rows


def _micro_specimen(member: str) -> str:
    match = re.search(r"(poro_\d{2}_spec_(?:02b|03b|02|03))", member)
    if not match:
        raise AuditBlocked(f"微球数据文件名缺试样身份：{member}")
    return match.group(1)


def _csv_index_sequence(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    header_rows: int,
) -> tuple[int, ...]:
    encoding = _text_encoding(archive, info)
    values: list[int] = []
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as text:
        reader = csv.reader(text)
        for index, row in enumerate(reader):
            if index < header_rows or not row or not any(cell.strip() for cell in row):
                continue
            number = _finite(row[0])
            if number is None or number != int(number):
                raise AuditBlocked(f"微球索引不是有限整数：{info.filename}/{row[:1]}")
            values.append(int(number))
    if tuple(values) != tuple(range(len(values))):
        raise AuditBlocked(f"微球索引不再是0..n-1：{info.filename}")
    return tuple(values)


def _micro_xlsx_anomaly(payload: bytes) -> dict[str, object]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Feuil1", "Feuil2"]:
            raise AuditBlocked(f"MinMax_Jp.xlsx工作表集合漂移：{workbook.sheetnames}")
        sheets: list[list[tuple[object, ...]]] = []
        finite_values: list[float] = []
        for sheet in workbook.worksheets:
            logical_rows: list[tuple[object, ...]] = []
            for row in sheet.iter_rows(values_only=True):
                values = list(row)
                while values and values[-1] is None:
                    values.pop()
                if not values or not any(value is not None and str(value) != "" for value in values):
                    continue
                logical_rows.append(tuple(values))
                finite_values.extend(
                    number
                    for value in values
                    if (number := _finite(value)) is not None
                )
            sheets.append(logical_rows)
        if sheets[0] != sheets[1]:
            raise AuditBlocked("MinMax_Jp.xlsx两张工作表不再是逻辑重复")
        anomaly_count = sum(value == 98_998_646.0 for value in finite_values)
        if anomaly_count != 2:
            raise AuditBlocked(f"MinMax_Jp.xlsx数量级异常出现次数漂移：{anomaly_count}")
        return {
            "工作表数": 2,
            "逻辑重复工作表数": 2,
            "异常值": 98_998_646,
            "CSV对应值": 0.98998646,
            "状态": "quarantine_semantic_duplicate_and_magnitude_error",
        }
    finally:
        workbook.close()


def audit_microsphere() -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    base = DATA_ROOT / MICROSPHERE
    archive_path = base / "Data_csv.zip"
    zip_summary, file_rows = audit_zip(archive_path)
    curve_rows: list[dict[str, object]] = []
    specimen_rows: defaultdict[str, dict[str, int]] = defaultdict(dict)
    specimen_indices: defaultdict[str, dict[str, tuple[int, ...]]] = defaultdict(dict)
    role_totals: Counter[str] = Counter()
    role_usable_totals: Counter[str] = Counter()

    macosx_rows = [row for row in file_rows if str(row["成员"]).startswith("__MACOSX/")]
    data_rows = [row for row in file_rows if str(row["成员"]).startswith("Data_csv/")]
    if len(macosx_rows) != 50:
        raise AuditBlocked(f"微球ZIP __MACOSX成员数漂移：{len(macosx_rows)}")
    for row in macosx_rows:
        row["角色"] = "Apple资源分叉噪声；权重0"

    with zipfile.ZipFile(archive_path) as archive:
        infos = {info.filename: info for info in archive.infolist() if not info.is_dir()}
        actual_csv = {
            name for name in infos if name.startswith("Data_csv/") and name.endswith(".csv")
        }
        if len(actual_csv) != 43:
            raise AuditBlocked(f"微球实际数据CSV数漂移：{len(actual_csv)}")
        ordinary_ds_store = {
            name
            for name in infos
            if name.startswith("Data_csv/") and name.endswith("/.DS_Store")
        }
        resource_fork_ds_store = {
            name
            for name in infos
            if name.startswith("__MACOSX/") and name.endswith("/._.DS_Store")
        }
        if len(ordinary_ds_store) != 2 or len(resource_fork_ds_store) != 2:
            raise AuditBlocked(
                "微球ZIP .DS_Store噪声数漂移："
                f"普通={len(ordinary_ds_store)}，资源分叉={len(resource_fork_ds_store)}"
            )
        if "Data_csv/Post/MinMax_Jp.xlsx" not in infos:
            raise AuditBlocked("微球ZIP缺少派生MinMax_Jp.xlsx")

        for name in sorted(actual_csv):
            info = infos[name]
            if name.startswith("Data_csv/Machine/"):
                specimen = _micro_specimen(name)
                points, invalid = _count_csv_selected(
                    archive,
                    info,
                    header_rows=1,
                    delimiter=",",
                    numeric_columns=(0, 3, 4, 5, 6, 7, 8),
                )
                role = "Machine"
                specimen_rows[specimen][role] = points
                specimen_indices[specimen][role] = _csv_index_sequence(
                    archive, info, header_rows=1
                )
                status = "candidate_after_channel_binding"
                material = f"PU+热塑微球 {specimen.split('_')[1]} vol%"
            elif name.startswith("Data_csv/DIC/"):
                specimen = _micro_specimen(name)
                channel = "DIC_xy" if name.endswith("_strain_xy.csv") else "DIC_yz"
                points, invalid = _count_csv_selected(
                    archive,
                    info,
                    header_rows=2,
                    delimiter=",",
                    numeric_columns=(0, 1, 2, 3, 4, 5, 6),
                )
                role = channel
                specimen_rows[specimen][role] = points
                specimen_indices[specimen][role] = _csv_index_sequence(
                    archive, info, header_rows=2
                )
                status = "candidate_after_channel_binding"
                material = f"PU+热塑微球 {specimen.split('_')[1]} vol%"
            elif re.fullmatch(r"Data_csv/Post/poro_\d{2}_moyenne\.csv", name):
                points, invalid = _count_csv_selected(
                    archive,
                    info,
                    header_rows=0,
                    delimiter=",",
                    numeric_columns=(0, 1, 2),
                )
                role = "Post_condition_mean"
                specimen = PurePosixPath(name).stem
                status = "derived_weight_zero"
                material = f"PU+热塑微球 {specimen.split('_')[1]} vol%"
            elif name == "Data_csv/Post/MinMax_Jp.csv":
                points, invalid = _count_csv_selected(
                    archive,
                    info,
                    header_rows=0,
                    delimiter=";",
                    numeric_columns=(0, 1, 2),
                )
                role = "Post_MinMax_summary"
                specimen = "all_conditions"
                status = "derived_weight_zero"
                material = "PU+热塑微球 六条件汇总"
            else:
                raise AuditBlocked(f"微球CSV角色未覆盖：{name}")
            expected_invalid = {
                "Data_csv/DIC/poro_05_spec_02_strain_yz.csv": 1,
                "Data_csv/DIC/poro_10_spec_02_strain_yz.csv": 3,
                "Data_csv/DIC/poro_15_spec_02_strain_yz.csv": 5,
            }.get(name, 0)
            if invalid != expected_invalid:
                raise AuditBlocked(
                    f"微球CSV非完整行数漂移：{name}={invalid}/{expected_invalid}"
                )
            role_totals[role] += points
            role_usable_totals[role] += points - invalid
            curve_rows.append(
                {
                    "来源": MICROSPHERE,
                    "材料": material,
                    "工况或试样": specimen,
                    "数据角色": role,
                    "曲线或文件": name,
                    "点数": points - invalid,
                    "试样组键": f"10.5281/zenodo.6390478|{specimen}",
                    "准入状态": status,
                    "备注": (
                        f"原始索引行{points}；缺失掩码行{invalid}；Machine、DIC_xy、DIC_yz"
                        "按同一物理试样绑定；Post不新增试样"
                    ),
                }
            )

    if set(specimen_rows) != set(MICRO_SPECIMENS):
        raise AuditBlocked(f"微球12个试样集合漂移：{sorted(specimen_rows)}")
    for specimen, channels in specimen_rows.items():
        if set(channels) != {"Machine", "DIC_xy", "DIC_yz"}:
            raise AuditBlocked(f"微球试样通道不完整：{specimen}/{channels}")
        if len(set(channels.values())) != 1:
            raise AuditBlocked(f"微球试样跨通道点数不一致：{specimen}/{channels}")
        if len(set(specimen_indices[specimen].values())) != 1:
            raise AuditBlocked(f"微球试样跨通道索引序列不一致：{specimen}")
    expected_totals = Counter(
        {
            "Machine": 7_974,
            "DIC_xy": 7_974,
            "DIC_yz": 7_974,
            "Post_condition_mean": 3_000,
            "Post_MinMax_summary": 10,
        }
    )
    if role_totals != expected_totals:
        raise AuditBlocked(f"微球角色点数漂移：{role_totals}/{expected_totals}")
    expected_usable_totals = expected_totals.copy()
    expected_usable_totals["DIC_yz"] = 7_965
    if role_usable_totals != expected_usable_totals:
        raise AuditBlocked(
            f"微球可用完整点数漂移：{role_usable_totals}/{expected_usable_totals}"
        )
    condition_counts = Counter(specimen.split("_")[1] for specimen in specimen_rows)
    if condition_counts != Counter({condition: 2 for condition in MICRO_CONDITIONS}):
        raise AuditBlocked(f"微球六条件重复数漂移：{condition_counts}")

    with zipfile.ZipFile(archive_path) as archive:
        with archive.open("Data_csv/Post/MinMax_Jp.xlsx") as handle:
            xlsx_anomaly = _micro_xlsx_anomaly(handle.read())

    for row in data_rows:
        name = str(row["成员"])
        if name.endswith(".csv"):
            row["角色"] = "数值主数据或派生视图"
        elif name.endswith(".xlsx"):
            row["角色"] = "派生工作簿；权重0"
        else:
            row["角色"] = "系统噪声；权重0"

    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": MICROSPHERE,
        "DOI": "10.5281/zenodo.6390478",
        "许可": "CC BY 4.0",
        "ZIP审计": zip_summary,
        "最小充分下载": {
            "下载数值包字节": 780_946,
            "排除原始图像ZIP数": 12,
            "排除原始图像总量说明": "约26.9 GB；不增加六条件、12个试样或数值通道身份",
        },
        "实验结构": {
            "体积分数条件": [0, 5, 10, 15, 20, 25],
            "每条件物理试样数": 2,
            "物理试样总数": 12,
            "Machine曲线数": 12,
            "DIC曲线数": 24,
            "Machine点数": role_totals["Machine"],
            "DIC点数": role_totals["DIC_xy"] + role_totals["DIC_yz"],
            "DIC完整可用点数": role_usable_totals["DIC_xy"] + role_usable_totals["DIC_yz"],
            "YZ末端缺失掩码行数": 9,
            "条件均值曲线数": 6,
            "条件均值点数": role_totals["Post_condition_mean"],
            "MinMax派生行数": role_totals["Post_MinMax_summary"],
            "MinMax工作簿冲突": xlsx_anomaly,
            "Apple资源分叉成员数": len(macosx_rows),
        },
        "治理": _governance(
            fidelity="direct_experiment_PU_microsphere_composite",
            split_group_key="dataset_doi|porosity|specimen_id",
            source_weight_ceiling=0.30,
            candidate_eligible=True,
            note="Machine/DIC同试样总权重归一；Post均值、MinMax、XLSX与__MACOSX权重为0。",
        ),
    }
    return summary, file_rows, curve_rows


FIG_PAIR2_H0 = {
    "Figure_10", "Figure_11", "Figure_12", "Figure_13",
    "Figure_16a", "Figure_16b", "Figure_18a", "Figure_18b",
    "Figure_20b", "Figure_21a", "Figure_21b",
    *{f"Figure_{figure}{panel}" for figure in range(22, 30) for panel in "ab"},
    "Figure_33", "Figure_34",
}
FIG_PAIR2_H1 = {"Figure_5a", "Figure_5b", "Figure_32a", "Figure_32b"}
FIG_PAIR2_H2 = {"Figure_7", "Figure_8a", "Figure_8b"}
FIG_PAIR2_H3 = {"Figure_14"}
FIG_PAIR3_H2 = {
    "Figure_6a", "Figure_6b", "Figure_9a", "Figure_9b",
    "Figure_15a", "Figure_15b", "Figure_19a", "Figure_19b",
}
FIG_PAIR3_H3 = {"Figure_17a", "Figure_17b"}
FIG_SHARED_X = {
    "Figure_4a": (2, 0, tuple(range(1, 29))),
    # 原文件列错位：B列才是log(time)，A列是-85 °C模量。
    "Figure_4b": (2, 1, (0, *range(2, 29))),
    "Figure_31a": (2, 0, tuple(range(1, 73))),
    "Figure_31b": (2, 0, tuple(range(1, 58))),
}
EXPECTED_FIG_CSV_PROFILE = {
    "Figure_4a": (26, 672), "Figure_4b": (26, 672),
    "Figure_5a": (30, 29), "Figure_5b": (30, 29),
    "Figure_6a": (698, 936), "Figure_6b": (704, 881),
    "Figure_7": (704, 1_442), "Figure_8a": (928, 6_294),
    "Figure_8b": (928, 7_523), "Figure_9a": (121, 999),
    "Figure_9b": (121, 1_010), "Figure_10": (3, 15),
    "Figure_11": (3, 15), "Figure_12": (3, 15), "Figure_13": (3, 15),
    "Figure_14": (929, 7_445), "Figure_15a": (928, 6_330),
    "Figure_15b": (928, 7_492), "Figure_16a": (129, 3_495),
    "Figure_16b": (128, 3_435), "Figure_17a": (194, 3_495),
    "Figure_17b": (31_235, 472_714), "Figure_18a": (129, 3_495),
    "Figure_18b": (129, 3_435), "Figure_19a": (698, 2_447),
    "Figure_19b": (704, 2_381), "Figure_20b": (275, 2_836),
    "Figure_21a": (32_288, 289_750), "Figure_21b": (32_288, 289_324),
    "Figure_22a": (876, 3_873), "Figure_22b": (875, 3_882),
    "Figure_23a": (845, 3_780), "Figure_23b": (875, 3_882),
    "Figure_24a": (851, 3_753), "Figure_24b": (875, 3_882),
    "Figure_25a": (876, 1_931), "Figure_25b": (875, 1_942),
    "Figure_26a": (876, 3_258), "Figure_26b": (875, 2_598),
    "Figure_27a": (876, 3_258), "Figure_27b": (1_173, 2_174),
    "Figure_28a": (1_184, 3_258), "Figure_28b": (1_173, 2_174),
    "Figure_29a": (876, 2_177), "Figure_29b": (875, 2_180),
    "Figure_31a": (7, 360), "Figure_31b": (7, 285),
    "Figure_32a": (73, 72), "Figure_32b": (58, 57),
    "Figure_33": (702, 1_398), "Figure_34": (56, 112),
}


def _fig_layout(stem: str) -> tuple[int, str, int, int | None, tuple[int, ...]]:
    if stem == "Figure_20a":
        return 0, "paired", 2, None, ()
    if stem in FIG_SHARED_X:
        header_rows, x_column, y_columns = FIG_SHARED_X[stem]
        return header_rows, "shared_x", 0, x_column, y_columns
    for members, header_rows, stride in (
        (FIG_PAIR2_H0, 0, 2), (FIG_PAIR2_H1, 1, 2),
        (FIG_PAIR2_H2, 2, 2), (FIG_PAIR2_H3, 3, 2),
        (FIG_PAIR3_H2, 2, 3), (FIG_PAIR3_H3, 3, 3),
    ):
        if stem in members:
            return header_rows, "paired", stride, None, ()
    raise AuditBlocked(f"未定义Figshare坐标布局：{stem}")


def _fig_count_complete_pairs(
    rows: list[list[object]], stem: str
) -> tuple[int, int]:
    header_rows, kind, stride, x_column, y_columns = _fig_layout(stem)
    data_rows = rows[header_rows:]
    complete = 0
    orphan_numeric = 0
    if kind == "shared_x":
        if x_column is None:
            raise AuditBlocked(f"Figshare共享横坐标布局缺x列：{stem}")
        for row in data_rows:
            x = _finite(row[x_column]) if x_column < len(row) else None
            for y_column in y_columns:
                y = _finite(row[y_column]) if y_column < len(row) else None
                if x is not None and y is not None:
                    complete += 1
                elif x is not None and y is None:
                    orphan_numeric += 1
        return complete, orphan_numeric

    width = max((len(row) for row in data_rows), default=0)
    for row in data_rows:
        for start in range(0, width - 1, stride):
            x = _finite(row[start]) if start < len(row) else None
            y = _finite(row[start + 1]) if start + 1 < len(row) else None
            if x is not None and y is not None:
                complete += 1
            elif (x is None) != (y is None):
                orphan_numeric += 1
    return complete, orphan_numeric


def _fig_csv_rows(
    archive: zipfile.ZipFile, info: zipfile.ZipInfo
) -> list[list[object]]:
    encoding = _text_encoding(archive, info)
    with archive.open(info) as raw, io.TextIOWrapper(
        raw, encoding=encoding, newline=""
    ) as text:
        return [list(row) for row in csv.reader(text)]


def _fig_csv_profile(
    archive: zipfile.ZipFile, info: zipfile.ZipInfo
) -> tuple[dict[str, int], list[list[object]]]:
    rows = _fig_csv_rows(archive, info)
    stem = PurePosixPath(info.filename).stem
    complete, orphan = _fig_count_complete_pairs(rows, stem)
    expected = EXPECTED_FIG_CSV_PROFILE.get(stem)
    if expected is None or (len(rows), complete) != expected:
        raise AuditBlocked(
            f"Figshare逐文件行/坐标对漂移：{stem}={(len(rows), complete)}/{expected}"
        )
    return {
        "raw_rows": len(rows),
        "coordinate_pair_points": complete,
        "orphan_numeric_count": orphan,
    }, rows


def _fig_xlsx_profile(payload: bytes, stem: str) -> tuple[dict[str, int], list[list[object]]]:
    expected = {
        "Figure_10": (3, 12, 15, 0),
        "Figure_20a": (32_303, 60, 290_588, 1),
    }.get(stem)
    if expected is None:
        raise AuditBlocked(f"Figshare XLSX布局未冻结：{stem}")
    with zipfile.ZipFile(io.BytesIO(payload)) as container:
        names = [_safe_zip_name(name) for name in container.namelist()]
        if any("vbaproject.bin" in name.casefold() for name in names):
            raise AuditBlocked(f"Figshare XLSX含宏：{stem}")
        if any(name.startswith("xl/externalLinks/") for name in names):
            raise AuditBlocked(f"Figshare XLSX含外部链接：{stem}")

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    try:
        if len(workbook.worksheets) != 1 or workbook.worksheets[0].sheet_state != "visible":
            raise AuditBlocked(f"Figshare XLSX工作表结构漂移：{stem}")
        sheet = workbook.worksheets[0]
        rows: list[list[object]] = []
        formula_count = 0
        error_count = 0
        for row in sheet.iter_rows(values_only=False):
            values: list[object] = []
            for cell in row:
                if cell.data_type == "f":
                    formula_count += 1
                if cell.data_type == "e":
                    error_count += 1
                values.append(cell.value)
            rows.append(values)
        complete, orphan = _fig_count_complete_pairs(rows, stem)
        expected_rows, expected_width, expected_pairs, expected_orphan = expected
        if (
            len(rows) != expected_rows
            or max(map(len, rows), default=0) != expected_width
            or complete != expected_pairs
            or orphan != expected_orphan
            or formula_count != 0
            or error_count != 0
        ):
            raise AuditBlocked(
                f"Figshare XLSX语义漂移：{stem}/rows={len(rows)}/pairs={complete}/"
                f"orphan={orphan}/formula={formula_count}/error={error_count}"
            )
        return {
            "raw_rows": len(rows),
            "coordinate_pair_points": complete,
            "orphan_numeric_count": orphan,
            "formula_count": formula_count,
            "error_count": error_count,
            "sheet_count": 1,
        }, rows
    finally:
        workbook.close()


def audit_figshare() -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    base = DATA_ROOT / FIGSHARE
    archive_path = base / "rspa20220830_si_002.zip"
    zip_summary, file_rows = audit_zip(archive_path)
    curve_rows: list[dict[str, object]] = []
    csv_profiles: dict[str, dict[str, int]] = {}
    xlsx_profiles: dict[str, dict[str, int]] = {}
    csv_rows_by_stem: dict[str, list[list[object]]] = {}
    xlsx_rows_by_stem: dict[str, list[list[object]]] = {}

    expected_csv_names = {
        f"RSPA-2022-0830 Data/Figure_{figure}{suffix}.csv"
        for figure, suffixes in {
            4: ("a", "b"), 5: ("a", "b"), 6: ("a", "b"), 7: ("",),
            8: ("a", "b"), 9: ("a", "b"), 10: ("",), 11: ("",),
            12: ("",), 13: ("",), 14: ("",), 15: ("a", "b"),
            16: ("a", "b"), 17: ("a", "b"), 18: ("a", "b"),
            19: ("a", "b"), 20: ("b",), 21: ("a", "b"),
            22: ("a", "b"), 23: ("a", "b"), 24: ("a", "b"),
            25: ("a", "b"), 26: ("a", "b"), 27: ("a", "b"),
            28: ("a", "b"), 29: ("a", "b"), 31: ("a", "b"),
            32: ("a", "b"), 33: ("",), 34: ("",),
        }.items()
        for suffix in suffixes
    }
    if len(expected_csv_names) != 51:
        raise AuditBlocked("内部Figshare固定CSV清单定义错误")

    with zipfile.ZipFile(archive_path) as archive:
        infos = {info.filename: info for info in archive.infolist() if not info.is_dir()}
        actual_csv = {name for name in infos if name.endswith(".csv")}
        actual_xlsx = {name for name in infos if name.endswith(".xlsx")}
        if actual_csv != expected_csv_names:
            raise AuditBlocked(
                f"Figshare CSV集合漂移：缺失={sorted(expected_csv_names-actual_csv)}，"
                f"多余={sorted(actual_csv-expected_csv_names)}"
            )
        expected_xlsx = {
            "RSPA-2022-0830 Data/Figure_10.xlsx",
            "RSPA-2022-0830 Data/Figure_20a.xlsx",
        }
        if actual_xlsx != expected_xlsx:
            raise AuditBlocked(f"Figshare XLSX集合漂移：{sorted(actual_xlsx)}")
        defined_csv_layouts = (
            FIG_PAIR2_H0
            | FIG_PAIR2_H1
            | FIG_PAIR2_H2
            | FIG_PAIR2_H3
            | FIG_PAIR3_H2
            | FIG_PAIR3_H3
            | set(FIG_SHARED_X)
        )
        if defined_csv_layouts != {PurePosixPath(name).stem for name in actual_csv}:
            raise AuditBlocked("Figshare固定CSV布局集合与成员集合不一致")

        for name in sorted(actual_csv):
            profile, rows = _fig_csv_profile(archive, infos[name])
            csv_profiles[name] = profile
            figure = PurePosixPath(name).stem
            csv_rows_by_stem[figure] = rows
            material = "Task 3" if figure.endswith("a") else "Task 11" if figure.endswith("b") else "Task 3/Task 11或跨材料"
            status = "PU_auxiliary_series_split_required"
            note = "反应固化商用PU；需按系列拆分实验/变换/Prony/Abaqus且与父曲线同折"
            if figure in {"Figure_31a", "Figure_31b"}:
                status = "quarantine_protocol_conflict"
                note = "工作簿温度网格与论文协议不一致；protocol_consistency=false"
            elif figure in {
                "Figure_5a", "Figure_5b", "Figure_10", "Figure_11", "Figure_12",
                "Figure_13", "Figure_14", "Figure_15a", "Figure_15b", "Figure_17a",
                "Figure_17b", "Figure_18a", "Figure_18b", "Figure_19a", "Figure_19b",
                "Figure_21a", "Figure_21b", "Figure_24a", "Figure_24b", "Figure_28a",
                "Figure_28b", "Figure_29a", "Figure_29b", "Figure_32a", "Figure_32b",
                "Figure_33", "Figure_34",
            }:
                status = "lineage_or_duplicate_view_weight_zero"
                note = "主曲线/摘要/变换坐标/精确重复视图；保留血缘，独立训练权重0"
            curve_rows.append(
                {
                    "来源": FIGSHARE,
                    "材料": material,
                    "工况或试样": figure,
                    "数据角色": "论文图数据；实验/模型/派生混合",
                    "曲线或文件": name,
                    "点数": profile["coordinate_pair_points"],
                    "试样组键": f"10.6084/m9.figshare.23635998.v1|{figure}",
                    "准入状态": status,
                    "备注": note,
                }
            )
        for name in sorted(actual_xlsx):
            with archive.open(infos[name]) as handle:
                stem = PurePosixPath(name).stem
                profile, rows = _fig_xlsx_profile(handle.read(), stem)
            xlsx_profiles[name] = profile
            xlsx_rows_by_stem[stem] = rows
            is_duplicate = stem == "Figure_10"
            curve_rows.append(
                {
                    "来源": FIGSHARE,
                    "材料": "Task 3/Task 11或跨材料",
                    "工况或试样": PurePosixPath(name).stem,
                    "数据角色": "论文图工作簿",
                    "曲线或文件": name,
                    "点数": profile["coordinate_pair_points"],
                    "试样组键": f"10.6084/m9.figshare.23635998.v1|{PurePosixPath(name).stem}",
                    "准入状态": (
                        "exact_duplicate_weight_zero"
                        if is_duplicate
                        else "PU_auxiliary_series_split_required"
                    ),
                    "备注": (
                        "与Figure_10.csv的30个数值完全一致；独立权重0"
                        if is_duplicate
                        else "低速大应变松弛实验与模型混合；需按系列/视图拆分"
                    ),
                }
            )

    def normalized_matrix(rows: list[list[object]]) -> list[list[float | None]]:
        return [
            [_finite(value) if value is not None and str(value).strip() else None for value in row]
            for row in rows
        ]

    if normalized_matrix(csv_rows_by_stem["Figure_10"]) != normalized_matrix(
        xlsx_rows_by_stem["Figure_10"]
    ):
        raise AuditBlocked("Figure_10 CSV/XLSX不再是数值精确重复")

    figure4b = csv_rows_by_stem["Figure_4b"]
    if not (
        len(figure4b) == 26
        and len(figure4b[0]) == len(figure4b[1]) == 29
        and figure4b[0][0] == ""
        and figure4b[0][1] == "-85"
        and figure4b[0][2] == "-80"
        and figure4b[1][0] == "Log(time)"
        and _finite(figure4b[2][0]) == 6.52853
        and _finite(figure4b[2][1]) == -0.77629
        and _finite(figure4b[-1][1]) == -2.77806
    ):
        raise AuditBlocked("Figure_4b已知log(time)/-85C列错位形态漂移")

    figure31a = csv_rows_by_stem["Figure_31a"]
    figure31b = csv_rows_by_stem["Figure_31b"]
    temperatures31a = [_finite(value) for value in figure31a[0][1:]]
    temperatures31b = [_finite(value) for value in figure31b[0][1:]]
    if temperatures31a != [float(value) for value in range(-70, 74, 2)]:
        raise AuditBlocked("Figure_31a温度网格漂移")
    if temperatures31b != [float(value) for value in range(-70, 44, 2)]:
        raise AuditBlocked("Figure_31b温度网格漂移")
    expected_log_times = [-1.0, -0.49485, 0.0, 0.50515, 1.0]
    if [_finite(row[0]) for row in figure31a[2:]] != expected_log_times:
        raise AuditBlocked("Figure_31a log(time)截面漂移")
    if [_finite(row[0]) for row in figure31b[2:]] != expected_log_times:
        raise AuditBlocked("Figure_31b log(time)截面漂移")

    figure34 = csv_rows_by_stem["Figure_34"]
    row31a_zero = next(row for row in figure31a[2:] if _finite(row[0]) == 0.0)
    row31b_zero = next(row for row in figure31b[2:] if _finite(row[0]) == 0.0)
    lookup31a = {str(key): str(value) for key, value in zip(figure31a[0][1:], row31a_zero[1:])}
    lookup31b = {str(key): str(value) for key, value in zip(figure31b[0][1:], row31b_zero[1:])}
    for row in figure34:
        if len(row) != 4 or lookup31a.get(str(row[0])) != str(row[1]) or lookup31b.get(str(row[2])) != str(row[3]):
            raise AuditBlocked("Figure_34不再精确复制Figure_31的log(time)=0截面")

    csv_rows_total = sum(item["raw_rows"] for item in csv_profiles.values())
    xlsx_rows_total = sum(item["raw_rows"] for item in xlsx_profiles.values())
    csv_pair_total = sum(item["coordinate_pair_points"] for item in csv_profiles.values())
    xlsx_pair_total = sum(item["coordinate_pair_points"] for item in xlsx_profiles.values())
    csv_orphan_total = sum(item["orphan_numeric_count"] for item in csv_profiles.values())
    xlsx_orphan_total = sum(item["orphan_numeric_count"] for item in xlsx_profiles.values())
    if (
        csv_rows_total != 121_069
        or xlsx_rows_total != 32_306
        or csv_pair_total != 1_168_907
        or xlsx_pair_total != 290_603
        or csv_pair_total + xlsx_pair_total != 1_459_510
        or csv_orphan_total != 1_069
        or xlsx_orphan_total != 1
    ):
        raise AuditBlocked(
            "Figshare全量行/坐标对/孤立数值总数漂移："
            f"{csv_rows_total}/{xlsx_rows_total}/{csv_pair_total}/{xlsx_pair_total}/"
            f"{csv_orphan_total}/{xlsx_orphan_total}"
        )

    for row in file_rows:
        name = str(row["成员"])
        if name.endswith("Figure_10.xlsx"):
            row["角色"] = "Figure_10.csv精确重复；权重0"
        elif name.endswith(("Figure_31a.csv", "Figure_31b.csv")):
            row["角色"] = "单悬臂DMA协议冲突；隔离"
        else:
            row["角色"] = "反应固化PU论文图级实验/模型/派生数据"
    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": FIGSHARE,
        "数据DOI": "10.6084/m9.figshare.23635998.v1",
        "论文DOI": "10.1098/rspa.2022.0830",
        "许可": "CC BY 4.0",
        "ZIP审计": zip_summary,
        "材料与条件": {
            "材料体系数": 2,
            "材料": ["Bentley TASK 3; DSC Tg约55°C", "Bentley TASK 11; DSC Tg约15°C"],
            "材料分类": "双组分混合、脱泡、室温固化24 h并65°C后固化4 h的商用反应固化PU",
            "TPU确认状态": False,
            "结构或配方已知": False,
            "低速应变率_s-1": [0.001, 0.01, 0.1],
            "高速应变率_s-1": [1000, 1500, 2000],
            "低温_C": [-60, -40, -20, 0, 20],
            "测试": ["DMA", "低/高速压缩", "低速松弛", "SHPB动态松弛", "Prony模型"],
        },
        "图数据盘点": {
            "CSV文件数": len(csv_profiles),
            "XLSX文件数": len(xlsx_profiles),
            "CSV原始行数": csv_rows_total,
            "XLSX原始行数": xlsx_rows_total,
            "总原始行数": csv_rows_total + xlsx_rows_total,
            "CSV完整坐标对数": csv_pair_total,
            "XLSX完整坐标对数": xlsx_pair_total,
            "所有视图完整坐标对数": csv_pair_total + xlsx_pair_total,
            "CSV孤立数值数": csv_orphan_total,
            "XLSX孤立数值数": xlsx_orphan_total,
            "四个高密度视图坐标对数": 1_342_376,
            "四个高密度视图占比": 0.9197,
            "材料工况单元数": 38,
            "可恢复机械实验曲线实例数": 108,
            "物理试样总数": None,
            "试样身份状态": "108仅为曲线/重复试验实例；图级源数据无稳定sample_id，禁止宣称108个独立物理试样",
            "Figure_10_CSV_XLSX": "30个数值完全一致；XLSX为精确重复",
            "Figure_4b列错位": "B列为log(time)，A列为-85°C模量；已用固定共享x布局复算",
            "Figure_31协议一致性": {
                "protocol_consistency": False,
                "论文": "-70至70°C，2.5°C间隔",
                "Figure_31a": "-70至72°C，2°C网格",
                "Figure_31b": "-70至42°C，2°C网格",
            },
            "Figure_34血缘": "两材料列分别精确复制Figure_31a/31b的log(time)=0截面；权重0",
        },
        "跨Figure血缘": {
            "14": "全部来自Figure 8",
            "15": "低温复制Figure 9，高速复用Figure 8并含横坐标变换",
            "16_17_18": "同一低速松弛试验的时间/归一化/应力-应变视图，含Prony模型",
            "20_21": "同一大应变松弛试验的时间与应力-应变视图",
            "22_23_24": "同一10 mm SHPB试验的多通道/坐标视图",
            "26_27_28": "同一6 mm SHPB试验的多通道/坐标视图",
            "29": "Figure 26代表曲线精确重复",
            "33": "Figure 6/7三点弯曲主曲线精确重复",
            "34": "Figure 31的log(time)=0精确截面",
        },
        "未来任务权重上限_尚未物化": {
            "Task3_11原始实验力学_PU辅助": 0.25,
            "Figure31协议冲突_当前": 0.0,
            "Figure31协议解决后": 0.10,
            "归一化_主曲线_摘要标量": 0.10,
            "Prony_Abaqus模拟曲线": 0.05,
            "精确重复及替代绘图视图": 0.0,
            "TPU化学结构到性能主任务": 0.0,
            "整源批次占比": 0.05,
        },
        "治理": _governance(
            fidelity="mixed_figure_source_experiment_and_model",
            split_group_key="dataset_doi|material|physical_or_condition_family",
            source_weight_ceiling=0.0,
            candidate_eligible=False,
            note=(
                "不是已知化学结构TPU；当前只作PU力学辅助与人工验证证据。"
                "未来须按曲线内→工况→材料分层平均，禁止按1459510个采样点加权。"
            ),
        ),
    }
    return summary, file_rows, curve_rows


TPU_ARCHIVE = "ijss_2025_vevp_ScriptsForTestsImages.zip"
TPU_ROOT_PREFIX = "ijss_2025_vevp_ScriptsForTestsImages/"
TPU_EXPERIMENT_PREFIX = TPU_ROOT_PREFIX + "Experiments/TPU/"
TPU_INVALID_STRAIN_FILENAME = "Cyclic_compression_1V_2p78E-3_RT_TPU.csv"
TPU_RESOLVED_REGIME_TYPO_FILENAME = "Hysteresis_2V_0p5_1E-2_RT_TPU.csv"
EXPECTED_CALIBRATION_CASES = {
    "10_3",
    "10_4",
    "10_5",
    "10_7",
    "10_8",
    "10_9",
    "10_10",
}
EXPECTED_VALIDATION_RUNS = {
    "Compression/viii",
    "Compression/ix",
    "Compression/x",
    "Compression/xi",
    "Tension/iii",
    "Tension/iv",
    "Tension/v",
    "Tension/vi",
    "Torsion/F22",
    "Torsion/F23",
    "Torsion/F24",
    "Torsion/F25",
    "Torsion/F26",
    "Torsion/F27",
    "TorsionRotationCorrection/F26",
    "TorsionRotationCorrection/F43",
}
EXPECTED_CALIBRATION_PROFILE = {
    "10_3": {"subrun_points": [100, 100, 100], "point_count": 300},
    "10_4": {"subrun_points": [114, 111, 100], "point_count": 325},
    "10_5": {"subrun_points": [100, 3_742], "point_count": 3_842},
    "10_7": {"subrun_points": [1_231], "point_count": 1_231},
    "10_8": {"subrun_points": [1_000], "point_count": 1_000},
    "10_9": {"subrun_points": [100, 100, 100, 100, 100], "point_count": 500},
    "10_10": {"subrun_points": [141, 114, 112, 113, 114], "point_count": 594},
}
EXPECTED_VALIDATION_PROFILE = {
    "Compression/viii": (5, 2, 1_204, 2_408),
    "Compression/ix": (5, 2, 840, 1_680),
    "Compression/x": (9, 2, 1_165, 2_330),
    "Compression/xi": (7, 2, 588, 1_176),
    "Tension/iii": (5, 2, 2_561, 5_122),
    "Tension/iv": (5, 2, 1_151, 2_302),
    "Tension/v": (5, 2, 1_135, 2_270),
    "Tension/vi": (6, 2, 1_132, 2_264),
    "Torsion/F22": (10, 6, 843, 5_058),
    "Torsion/F23": (9, 6, 827, 4_962),
    "Torsion/F24": (9, 6, 788, 4_728),
    "Torsion/F25": (10, 7, 1_422, 9_954),
    "Torsion/F26": (9, 5, 935, 4_675),
    "Torsion/F27": (8, 5, 4_028, 20_140),
    "TorsionRotationCorrection/F26": (13, 9, 935, 8_415),
    "TorsionRotationCorrection/F43": (36, 32, 2_665, 85_280),
}


def _tpu_experiment_category(relative: str) -> tuple[str, int]:
    if relative.startswith("BulkCharacterisation/") and relative.endswith(".csv"):
        return "bulk", 10
    if relative.startswith("DumbbellCompression/") and (
        relative.endswith("Std_Data.csv") or relative.endswith("HS_Data.dat")
    ):
        return "dumbbell_compression", 5
    if relative.startswith("DumbbellTension/") and relative.endswith(".csv"):
        return "dumbbell_tension", 1
    if relative.startswith("DumbbellTorsion/") and relative.endswith(".csv"):
        return "dumbbell_torsion", 2
    raise AuditBlocked(f"TPU实验测量文件角色未覆盖：{relative}")


def _tpu_specimen_group(relative: str, category: str) -> str:
    if category == "bulk":
        return PurePosixPath(relative).stem
    if category == "dumbbell_compression":
        parts = PurePosixPath(relative).parts
        if len(parts) < 3:
            raise AuditBlocked(f"空心哑铃压缩路径缺试样目录：{relative}")
        return parts[1]
    return PurePosixPath(relative).stem


def _validation_run(relative: str) -> str | None:
    parts = PurePosixPath(relative).parts
    if len(parts) < 3 or parts[0] != "Validation":
        return None
    if parts[1] not in {
        "Compression",
        "Tension",
        "Torsion",
        "TorsionRotationCorrection",
    }:
        return None
    return f"{parts[1]}/{parts[2]}"


def _restricted_numpy_scalar_pickle_profile(payload: bytes) -> dict[str, int]:
    """只解析 pickle opcode；绝不反序列化或执行其中 GLOBAL/REDUCE。"""
    try:
        operations = list(pickletools.genops(payload))
    except ValueError as exc:  # pragma: no cover - defensive
        raise AuditBlocked(f"标定pickle opcode损坏：{exc}") from exc
    if not operations or operations[-1][0].name != "STOP":
        raise AuditBlocked("标定pickle缺少终止STOP")
    allowed_opcodes = {
        "APPEND", "APPENDS", "BINGET", "BININT", "BININT1", "BUILD",
        "EMPTY_LIST", "FRAME", "MARK", "MEMOIZE", "NEWFALSE", "NEWTRUE",
        "NONE", "PROTO", "REDUCE", "SHORT_BINBYTES", "SHORT_BINUNICODE",
        "STACK_GLOBAL", "STOP", "TUPLE", "TUPLE2", "TUPLE3",
    }
    opcode_counts = Counter(operation.name for operation, _arg, _position in operations)
    unexpected = set(opcode_counts) - allowed_opcodes
    if unexpected:
        raise AuditBlocked(f"标定pickle出现未允许opcode：{sorted(unexpected)}")
    protocols = [arg for operation, arg, _position in operations if operation.name == "PROTO"]
    if protocols != [4]:
        raise AuditBlocked(f"标定pickle协议漂移：{protocols}")

    static_globals: list[tuple[str, str]] = []
    for index, (operation, _arg, _position) in enumerate(operations):
        if operation.name != "STACK_GLOBAL":
            continue
        strings: list[str] = []
        for previous, previous_arg, _previous_position in reversed(operations[:index]):
            if previous.name == "SHORT_BINUNICODE":
                strings.append(str(previous_arg))
                if len(strings) == 2:
                    break
        if len(strings) != 2:
            raise AuditBlocked("标定pickle无法静态解析STACK_GLOBAL")
        static_globals.append((strings[1], strings[0]))
    if static_globals != [
        ("numpy.core.multiarray", "scalar"),
        ("numpy", "dtype"),
    ]:
        raise AuditBlocked(f"标定pickle GLOBAL 白名单漂移：{static_globals}")
    if opcode_counts["BUILD"] != 1 or opcode_counts["STACK_GLOBAL"] != 2:
        raise AuditBlocked(f"标定pickle numpy标量骨架漂移：{opcode_counts}")
    point_count = opcode_counts["REDUCE"] - 1
    subrun_count = opcode_counts["EMPTY_LIST"] - 1
    if point_count <= 0 or subrun_count <= 0:
        raise AuditBlocked(
            f"标定pickle同步点或子运行数非法：points={point_count}, runs={subrun_count}"
        )
    return {"point_count": point_count, "subrun_count": subrun_count}


def _audit_torsion_numeric_layout(
    archive: zipfile.ZipFile, info: zipfile.ZipInfo
) -> dict[str, int]:
    expected_headers = [
        "Drehmoment", "Frequenz T absolut", "Kraft", "Kraft (Aufnehmerwert)",
        "Kraft absolut", "Kraft absolut (Aufnehmerwert)", "Weg", "Winkel",
        "Winkel absolut", "Zeit", "Key",
    ]
    mandatory_columns = (0, 2, 4, 6, 7, 8, 9, 10)
    declared_nan_columns = (1, 3, 5)
    rows = 0
    with archive.open(info) as raw:
        first = raw.readline().decode("utf-8-sig", errors="strict").rstrip("\r\n").split(";")
        second = raw.readline().decode("utf-8", errors="strict").rstrip("\r\n").split(";")
        if first != expected_headers or len(second) != 11:
            raise AuditBlocked(f"TPU扭转11列表头漂移：{info.filename}")
        for physical_line, raw_line in enumerate(raw, 3):
            cells = raw_line.rstrip(b"\r\n").split(b";")
            if len(cells) != 11:
                raise AuditBlocked(
                    f"TPU扭转错列：{info.filename}/line={physical_line}/cols={len(cells)}"
                )
            for column in mandatory_columns:
                try:
                    number = float(cells[column])
                except ValueError as exc:
                    raise AuditBlocked(
                        f"TPU扭转必需通道非数值：{info.filename}/{physical_line}/{column}"
                    ) from exc
                if not math.isfinite(number):
                    raise AuditBlocked(
                        f"TPU扭转必需通道非有限：{info.filename}/{physical_line}/{column}"
                    )
            if any(cells[column] != b"NaN" for column in declared_nan_columns):
                raise AuditBlocked(
                    f"TPU扭转声明空辅助通道漂移：{info.filename}/{physical_line}"
                )
            rows += 1
    return {
        "rows": rows,
        "mandatory_finite_columns": len(mandatory_columns),
        "declared_all_nan_columns": len(declared_nan_columns),
    }


_VALIDATION_NUMERIC_BYTES = b"0123456789+-.eE"


def _validation_csv_profile(
    archive: zipfile.ZipFile, info: zipfile.ZipInfo
) -> dict[str, int]:
    """线性扫描验证输出；逐行校验列宽和数值字符，不加载整表。"""
    data_rows = 0
    numeric_cells = 0
    width: int | None = None
    with archive.open(info) as raw:
        first = raw.readline()
        if b";" not in first:
            raise AuditBlocked(f"TPU验证CSV分隔符漂移：{info.filename}")
        first_cells = first.rstrip(b"\r\n").split(b";")
        try:
            first_number = float(first_cells[0])
            first_is_data = math.isfinite(first_number)
        except ValueError:
            first_is_data = False

        pending = [first] if first_is_data else []
        for raw_line in itertools.chain(pending, raw):
            line = raw_line.rstrip(b"\r\n")
            if not line:
                raise AuditBlocked(f"TPU验证CSV出现空数据行：{info.filename}")
            cells = line.split(b";")
            if width is None:
                width = len(cells)
            if len(cells) != width or any(cell == b"" for cell in cells):
                raise AuditBlocked(f"TPU验证CSV列宽或空单元格漂移：{info.filename}")
            for cell in cells:
                if cell.translate(None, _VALIDATION_NUMERIC_BYTES):
                    raise AuditBlocked(f"TPU验证CSV出现非数值字符：{info.filename}")
            try:
                first_value = float(cells[0])
            except ValueError as exc:
                raise AuditBlocked(f"TPU验证CSV时间列非数值：{info.filename}") from exc
            if not math.isfinite(first_value):
                raise AuditBlocked(f"TPU验证CSV时间列非有限：{info.filename}")
            data_rows += 1
            numeric_cells += len(cells)
    if data_rows <= 0 or width is None:
        raise AuditBlocked(f"TPU验证CSV无数据：{info.filename}")
    return {"data_rows": data_rows, "numeric_cells": numeric_cells, "columns": width}


def audit_tpu1301() -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    base = DATA_ROOT / TPU1301
    archive_path = base / TPU_ARCHIVE
    zip_summary, file_rows = audit_zip(archive_path)
    curve_rows: list[dict[str, object]] = []
    experimental_points: Counter[str] = Counter()
    experimental_files: Counter[str] = Counter()
    specimen_groups: defaultdict[str, set[str]] = defaultdict(set)
    validation_rows = 0
    validation_csv_count = 0
    validation_numeric_cells = 0
    validation_run_file_counts: Counter[str] = Counter()
    validation_run_csv_counts: Counter[str] = Counter()
    validation_run_rows: Counter[str] = Counter()
    validation_run_grids: defaultdict[str, set[int]] = defaultdict(set)
    validation_run_hashes: defaultdict[str, list[str]] = defaultdict(list)
    validation_member_hashes: dict[str, str] = {}
    validation_csv_hashes: list[str] = []
    validation_root_files: list[str] = []
    compression_channel_points: defaultdict[str, dict[str, int]] = defaultdict(dict)
    torsion_layout_rows = 0

    with zipfile.ZipFile(archive_path) as archive:
        infos = {info.filename: info for info in archive.infolist() if not info.is_dir()}
        experiment_measurements = {
            name: info
            for name, info in infos.items()
            if name.startswith(TPU_EXPERIMENT_PREFIX)
            and (
                name.endswith(".csv")
                or name.endswith("HS_Data.dat")
            )
        }
        bulk = {
            name
            for name in experiment_measurements
            if name.startswith(TPU_EXPERIMENT_PREFIX + "BulkCharacterisation/")
        }
        dumbbell_compression = {
            name
            for name in experiment_measurements
            if name.startswith(TPU_EXPERIMENT_PREFIX + "DumbbellCompression/")
        }
        dumbbell_tension = {
            name
            for name in experiment_measurements
            if name.startswith(TPU_EXPERIMENT_PREFIX + "DumbbellTension/")
        }
        dumbbell_torsion = {
            name
            for name in experiment_measurements
            if name.startswith(TPU_EXPERIMENT_PREFIX + "DumbbellTorsion/")
        }
        if (len(bulk), len(dumbbell_compression), len(dumbbell_tension), len(dumbbell_torsion)) != (
            68,
            16,
            4,
            6,
        ):
            raise AuditBlocked(
                "TPU实验测量文件数漂移："
                f"bulk={len(bulk)}, dumbbell_compression={len(dumbbell_compression)}, "
                f"tension={len(dumbbell_tension)}, torsion={len(dumbbell_torsion)}"
            )

        conflict_points = 0
        malformed_numeric_points = 0
        for name, info in sorted(experiment_measurements.items()):
            relative = name.removeprefix(TPU_EXPERIMENT_PREFIX)
            category, header_lines = _tpu_experiment_category(relative)
            if category == "dumbbell_torsion":
                torsion_profile = _audit_torsion_numeric_layout(archive, info)
                points = torsion_profile["rows"]
                torsion_layout_rows += points
            else:
                lines = _physical_line_count(archive, info)
                points = lines - header_lines
            if points <= 0:
                raise AuditBlocked(f"TPU实验文件没有正数据行：{relative}={points}")
            filename = PurePosixPath(relative).name
            specimen = _tpu_specimen_group(relative, category)
            experimental_points[category] += points
            experimental_files[category] += 1
            specimen_groups[category].add(specimen)
            if category == "dumbbell_compression":
                channel = "HS" if relative.endswith("HS_Data.dat") else "Std"
                compression_channel_points[specimen][channel] = points

            status = "candidate_after_grouped_split"
            material = "EOS TPU 1301"
            note = "单一材料内按试样/打印方向/速率成组；采样点不增加材料权重"
            usable_points = points
            if category == "dumbbell_tension":
                status = "candidate_digitized_curve_after_grouped_split"
                note = (
                    "论文PNG手工数字化曲线；仅曲线实例可恢复、物理sample_id未知；"
                    "与对应Validation运行同折且降权"
                )
            if filename == TPU_PA12_EXCLUSION:
                status = "excluded_non_TPU_PA12"
                material = "PA12"
                note = "位于TPU目录但文件名明确为PA12；不可纳入TPU计数"
            elif filename == str(TPU_IDENTITY_CONFLICT["filename"]):
                header = _first_lines(archive, info, 8)
                label_lines = [line for line in header if line.startswith("Specimen label:")]
                if len(label_lines) != 1 or label_lines[0].split(":", 1)[1].strip() != "6V":
                    raise AuditBlocked(f"已知松弛试样身份冲突形态漂移：{label_lines}")
                status = str(TPU_IDENTITY_CONFLICT["decision"])
                note = "文件名7H但内嵌Specimen label为6V；文件非重复，身份解决前权重0"
                conflict_points = points
            elif filename == TPU_INVALID_STRAIN_FILENAME:
                first_data_line = _first_lines(archive, info, 11)
                if len(first_data_line) != 11:
                    raise AuditBlocked("TPU已知坏应变行位置漂移")
                cells = [cell.strip() for cell in first_data_line[10].split("\t")]
                if len(cells) != 5 or cells[3] != "V" or any(
                    _finite(cells[index]) is None for index in (0, 1, 2, 4)
                ):
                    raise AuditBlocked(f"TPU已知坏应变行形态漂移：{cells}")
                malformed_numeric_points += 1
                usable_points -= 1
                note += "；首数据行Eng.strain误写V，该单点保留缺失掩码且监督权重0"
            elif filename == TPU_RESOLVED_REGIME_TYPO_FILENAME:
                header = _first_lines(archive, info, 10)
                label_lines = [line for line in header if line.startswith("Specimen label:")]
                if len(label_lines) != 1 or label_lines[0].split(":", 1)[1].strip() != "2V_100%":
                    raise AuditBlocked(f"TPU已知滞回工况头部错写形态漂移：{label_lines}")
                note += (
                    "；头部2V_100%与文件名0.5 true-strain及曲线幅值冲突，"
                    "按文件名/物理曲线解析为50%并登记resolved_header_regime_typo"
                )

            role = {
                "bulk": "SLS块体实验曲线",
                "dumbbell_compression": (
                    "空心哑铃压缩高速通道" if relative.endswith("HS_Data.dat") else "空心哑铃压缩标准通道"
                ),
                "dumbbell_tension": "空心哑铃拉伸实验曲线",
                "dumbbell_torsion": "空心哑铃扭转实验曲线",
            }[category]
            curve_rows.append(
                {
                    "来源": TPU1301,
                    "材料": material,
                    "工况或试样": specimen,
                    "数据角色": role,
                    "曲线或文件": relative,
                    "点数": usable_points,
                    "试样组键": f"10.5281/zenodo.15370425|{specimen}",
                    "准入状态": status,
                    "备注": note,
                }
            )

        expected_experimental_points = Counter(
            {
                "bulk": 748_209,
                "dumbbell_compression": 868_597,
                "dumbbell_tension": 98,
                "dumbbell_torsion": 4_209_030,
            }
        )
        if experimental_points != expected_experimental_points:
            raise AuditBlocked(
                f"TPU实验点数漂移：{experimental_points}/{expected_experimental_points}"
            )
        # bulk包含一个明确PA12；TPU块体本身67个文件、其中一个身份冲突。
        pa12_points = next(
            int(row["点数"])
            for row in curve_rows
            if row["准入状态"] == "excluded_non_TPU_PA12"
        )
        if pa12_points != 2_402:
            raise AuditBlocked(f"PA12误入文件点数漂移：{pa12_points}")
        if malformed_numeric_points != 1:
            raise AuditBlocked(f"TPU已知坏数值点数漂移：{malformed_numeric_points}")
        if torsion_layout_rows != 4_209_030:
            raise AuditBlocked(f"TPU扭转11列数值布局行数漂移：{torsion_layout_rows}")

        compression_groups = {
            PurePosixPath(name.removeprefix(TPU_EXPERIMENT_PREFIX)).parts[1]
            for name in dumbbell_compression
        }
        if len(compression_groups) != 8:
            raise AuditBlocked(f"空心哑铃压缩物理试样目录数漂移：{compression_groups}")
        for group in compression_groups:
            members = [
                name for name in dumbbell_compression if f"DumbbellCompression/{group}/" in name
            ]
            if len(members) != 2 or not any(name.endswith("Std_Data.csv") for name in members) or not any(
                name.endswith("HS_Data.dat") for name in members
            ):
                raise AuditBlocked(f"空心哑铃压缩双通道绑定不完整：{group}/{members}")
        expected_compression_channels = {
            "TPUalt_X_1mms_1": {"Std": 2_134, "HS": 318_255},
            "TPUalt_X_10mms_1": {"Std": 415, "HS": 60_585},
            "TPUalt_X_100mms_1": {"Std": 157, "HS": 21_750},
            "TPUalt_X_1000mms_1": {"Std": 113, "HS": 14_790},
            "TPUalt_Z_1mms_1": {"Std": 2_134, "HS": 318_615},
            "TPUalt_Z_10mms_1": {"Std": 619, "HS": 91_245},
            "TPUalt_Z_100mms_1": {"Std": 157, "HS": 22_260},
            "TPUalt_Z_1000mms_1": {"Std": 113, "HS": 15_255},
        }
        if dict(compression_channel_points) != expected_compression_channels:
            raise AuditBlocked(
                f"空心哑铃压缩双通道点数漂移：{dict(compression_channel_points)}"
            )

        calibration_prefix = TPU_ROOT_PREFIX + "Calibration/TPU/"
        calibration_fields: defaultdict[str, set[str]] = defaultdict(set)
        for name in infos:
            if not name.startswith(calibration_prefix):
                continue
            match = re.fullmatch(
                re.escape(calibration_prefix)
                + r"(F|M|P|T|epl|L)_Case_(10_(?:3|4|5|7|8|9|10))\.dat",
                name,
            )
            if match:
                calibration_fields[match.group(2)].add(match.group(1))
        if set(calibration_fields) != EXPECTED_CALIBRATION_CASES or any(
            fields != {"F", "M", "P", "T", "epl", "L"}
            for fields in calibration_fields.values()
        ):
            raise AuditBlocked(f"TPU标定case或六字段集合漂移：{calibration_fields}")

        calibration_profiles: dict[str, dict[str, object]] = {}
        for case, expected in sorted(EXPECTED_CALIBRATION_PROFILE.items()):
            channel_profiles: dict[str, dict[str, int]] = {}
            for channel in ("F", "P", "T", "epl"):
                member = f"{calibration_prefix}{channel}_Case_{case}.dat"
                if member not in infos:
                    raise AuditBlocked(f"TPU标定缺少固定pickle通道：{member}")
                channel_profiles[channel] = _restricted_numpy_scalar_pickle_profile(
                    archive.read(infos[member])
                )
            point_counts = {profile["point_count"] for profile in channel_profiles.values()}
            subrun_counts = {profile["subrun_count"] for profile in channel_profiles.values()}
            if point_counts != {int(expected["point_count"])}:
                raise AuditBlocked(f"TPU标定四通道同步点漂移：{case}/{point_counts}")
            if subrun_counts != {len(expected["subrun_points"])}:
                raise AuditBlocked(f"TPU标定子运行数漂移：{case}/{subrun_counts}")
            calibration_profiles[case] = {
                "subrun_points": list(expected["subrun_points"]),
                "subrun_count": len(expected["subrun_points"]),
                "synchronous_point_count": int(expected["point_count"]),
                "finite_scalar_count_F_P_T_epl": int(expected["point_count"]) * 4,
                "static_pickle_audit": "pickletools.genops_only_no_deserialization",
            }
            curve_rows.append(
                {
                    "来源": TPU1301,
                    "材料": "EOS TPU 1301",
                    "工况或试样": f"Calibration/{case}",
                    "数据角色": "本构标定同步序列",
                    "曲线或文件": f"Calibration/TPU/{{F,P,T,epl}}_Case_{case}.dat",
                    "点数": int(expected["point_count"]),
                    "试样组键": f"10.5281/zenodo.15370425|simulation|calibration|{case}",
                    "准入状态": "calibration_fit_weight_zero_for_property_prediction",
                    "备注": (
                        "仅用pickletools静态opcode复算，未执行pickle；F/P/T/epl四通道同步；"
                        "拟合内输出不作独立性能监督"
                    ),
                }
            )
        if sum(
            int(profile["synchronous_point_count"])
            for profile in calibration_profiles.values()
        ) != 7_792:
            raise AuditBlocked("TPU标定同步点总数漂移")

        validation_prefix = TPU_ROOT_PREFIX + "Validation/"
        validation_files = {
            name: info for name, info in infos.items() if name.startswith(validation_prefix)
        }
        for name, info in validation_files.items():
            run = _validation_run(name.removeprefix(TPU_ROOT_PREFIX))
            if run is None:
                validation_root_files.append(name)
                continue
            validation_run_file_counts[run] += 1
            with archive.open(info) as handle:
                digest = _hash_stream(handle)
            validation_member_hashes[name] = digest
            validation_run_hashes[run].append(digest)
        validation_runs = {
            run
            for name in validation_files
            if (run := _validation_run(name.removeprefix(TPU_ROOT_PREFIX))) is not None
        }
        if validation_runs != EXPECTED_VALIDATION_RUNS:
            raise AuditBlocked(
                f"TPU验证运行集合漂移：缺失={sorted(EXPECTED_VALIDATION_RUNS-validation_runs)}，"
                f"多余={sorted(validation_runs-EXPECTED_VALIDATION_RUNS)}"
            )
        expected_validation_root_files = {
            validation_prefix + "validationPlots.py",
            *{
                validation_prefix + "Images_Paper1/" + filename
                for filename in (
                    "TPU_Tension_HollowDumbbellZoltan_Numerical.pdf",
                    "energyComparison.pdf",
                    "axialForceSax0_Numerical.pdf",
                    "torsion_Fax0_Numerical.pdf",
                    "verticalDisplacementFax0_Numerical.pdf",
                    "torque.pdf",
                    "torsion_Sax0_Numerical.pdf",
                    "TPU_Compression_HollowDumbbellZoltan_Numerical.pdf",
                )
            },
        }
        if set(validation_root_files) != expected_validation_root_files:
            raise AuditBlocked(f"TPU验证根级辅助文件集合漂移：{validation_root_files}")
        for name, info in sorted(validation_files.items()):
            if not name.endswith(".csv"):
                continue
            validation_csv_count += 1
            profile = _validation_csv_profile(archive, info)
            data_rows = profile["data_rows"]
            validation_rows += data_rows
            validation_numeric_cells += profile["numeric_cells"]
            run = _validation_run(name.removeprefix(TPU_ROOT_PREFIX))
            if run is None:
                raise AuditBlocked(f"TPU验证CSV无法映射运行：{name}")
            validation_run_csv_counts[run] += 1
            validation_run_rows[run] += data_rows
            validation_run_grids[run].add(data_rows)
            validation_csv_hashes.append(validation_member_hashes[name])
            curve_rows.append(
                {
                    "来源": TPU1301,
                    "材料": "EOS TPU 1301",
                    "工况或试样": run or "validation_root",
                    "数据角色": "有限元验证输出",
                    "曲线或文件": name.removeprefix(TPU_ROOT_PREFIX),
                    "点数": data_rows,
                    "试样组键": f"10.5281/zenodo.15370425|simulation|{run}",
                    "准入状态": "candidate_calibrated_simulation_after_family_cap",
                    "备注": "同一运行的分片、节点、时间步总权重归一；不计为新增材料或试样",
                }
            )
        if validation_csv_count != 92:
            raise AuditBlocked(f"TPU验证CSV数漂移：{validation_csv_count}")
        validation_profile: dict[str, dict[str, int]] = {}
        for run, (expected_files, expected_csv, expected_grid, expected_rows) in sorted(
            EXPECTED_VALIDATION_PROFILE.items()
        ):
            actual = (
                validation_run_file_counts[run],
                validation_run_csv_counts[run],
                validation_run_grids[run],
                validation_run_rows[run],
            )
            if actual != (expected_files, expected_csv, {expected_grid}, expected_rows):
                raise AuditBlocked(f"TPU验证逐运行统计漂移：{run}/{actual}")
            validation_profile[run] = {
                "file_count": expected_files,
                "csv_count": expected_csv,
                "time_grid_points": expected_grid,
                "csv_data_rows": expected_rows,
            }
        if set(validation_run_file_counts) != set(EXPECTED_VALIDATION_PROFILE):
            raise AuditBlocked(f"TPU验证运行文件集合漂移：{validation_run_file_counts}")
        if validation_rows != 162_764:
            raise AuditBlocked(f"TPU验证CSV数据行总数漂移：{validation_rows}")
        if validation_numeric_cells != 112_358_792:
            raise AuditBlocked(f"TPU验证有限数值单元总数漂移：{validation_numeric_cells}")
        grid_total = sum(next(iter(points)) for points in validation_run_grids.values())
        if grid_total != 22_219:
            raise AuditBlocked(f"TPU验证目录运行时间网格总数漂移：{grid_total}")
        validation_hash_counts = Counter(validation_csv_hashes)
        if (
            len(validation_hash_counts) != 70
            or sum(count > 1 for count in validation_hash_counts.values()) != 11
            or sum(count - 1 for count in validation_hash_counts.values()) != 22
        ):
            raise AuditBlocked(f"TPU验证CSV内容重复谱漂移：{validation_hash_counts}")
        f26_overlap = sum(
            (
                Counter(validation_run_hashes["Torsion/F26"])
                & Counter(validation_run_hashes["TorsionRotationCorrection/F26"])
            ).values()
        )
        if f26_overlap != 9:
            raise AuditBlocked(f"TPU验证F26基线复用文件数漂移：{f26_overlap}")

    for row in file_rows:
        name = str(row["成员"])
        if name.startswith(TPU_ROOT_PREFIX + "Experiments/TPU/"):
            row["角色"] = "TPU实验输入或仪器辅助"
        elif name.startswith(TPU_ROOT_PREFIX + "Experiments/PP/"):
            row["角色"] = "PP外域实验；不纳入TPU"
        elif name.startswith(TPU_ROOT_PREFIX + "Calibration/TPU/"):
            row["角色"] = "TPU本构标定输入/输出"
        elif name.startswith(TPU_ROOT_PREFIX + "Calibration/PP/"):
            row["角色"] = "PP外域标定；不纳入TPU"
        elif name.startswith(TPU_ROOT_PREFIX + "Validation/"):
            row["角色"] = "TPU有限元验证脚本/输出/图像"
        else:
            row["角色"] = "许可或总说明"

    tpu_experimental_points = (
        experimental_points["bulk"]
        - 2_402
        + experimental_points["dumbbell_compression"]
        + experimental_points["dumbbell_tension"]
        + experimental_points["dumbbell_torsion"]
    )
    finite_tpu_experimental_points = tpu_experimental_points - malformed_numeric_points
    clean_candidate_points = finite_tpu_experimental_points - conflict_points
    if (
        tpu_experimental_points != 5_823_532
        or finite_tpu_experimental_points != 5_823_531
        or clean_candidate_points != 5_818_564
    ):
        raise AuditBlocked(
            "TPU排除PA12、坏点和身份冲突后的分层计数漂移："
            f"{tpu_experimental_points}/{finite_tpu_experimental_points}/{clean_candidate_points}"
        )
    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": TPU1301,
        "DOI": "10.5281/zenodo.15370425",
        "论文题名": "A consistent finite-strain thermomechanical quasi-nonlinear-viscoelastic viscoplastic constitutive model for thermoplastic polymers",
        "许可": "CC BY 4.0",
        "ZIP审计": zip_summary,
        "材料边界": {
            "目标材料": "EOS TPU 1301；SLS块体与空心哑铃",
            "排除PP目录": True,
            "排除误入PA12文件": TPU_PA12_EXCLUSION,
            "PA12排除点数": 2_402,
        },
        "实验层": {
            "实验采集或曲线单元数": 85,
            "可明确物理运行数": 81,
            "身份闭合直接实验运行数": 80,
            "身份冲突隔离直接实验运行数": 1,
            "手工数字化曲线实例数_物理试样ID未知": 4,
            "身份冲突": TPU_IDENTITY_CONFLICT,
            "已解决头部工况笔误": {
                "文件": TPU_RESOLVED_REGIME_TYPO_FILENAME,
                "头部值": "2V_100%",
                "解析工况": "0.5 true strain（约39.4%工程应变）",
                "状态": "resolved_header_regime_typo",
            },
            "SLS块体TPU文件数": 67,
            "空心哑铃压缩物理试样数": 8,
            "空心哑铃压缩测量通道文件数": 16,
            "空心哑铃压缩双通道点数": dict(sorted(compression_channel_points.items())),
            "空心哑铃压缩通道关系": "Std与HS同一运行但时间原点/采样率不同；绑定分组但禁止逐行join",
            "空心哑铃拉伸手工数字化曲线数": 4,
            "空心哑铃扭转试样数": 6,
            "空心哑铃扭转11列布局": {
                "数据行": torsion_layout_rows,
                "逐行有限必需通道数": 8,
                "逐行声明NaN辅助通道数": 3,
                "错列或Inf数": 0,
            },
            "TPU实验数据行数_含隔离": tpu_experimental_points,
            "非数值应变坏点数": malformed_numeric_points,
            "TPU实验有限有效行数_含身份隔离": finite_tpu_experimental_points,
            "当前身份闭合候选数据行数": clean_candidate_points,
        },
        "模拟层": {
            "标定case数": len(EXPECTED_CALIBRATION_CASES),
            "每个标定case字段": ["F", "L", "M", "P", "T", "epl"],
            "标定输出文件数": 42,
            "标定pickle安全解析": "仅pickletools.genops静态检查；禁止普通反序列化和执行归档脚本",
            "标定子运行数": sum(
                int(profile["subrun_count"]) for profile in calibration_profiles.values()
            ),
            "标定同步点数": sum(
                int(profile["synchronous_point_count"])
                for profile in calibration_profiles.values()
            ),
            "标定F_P_T_epl有限标量数": sum(
                int(profile["finite_scalar_count_F_P_T_epl"])
                for profile in calibration_profiles.values()
            ),
            "标定实际唯一加载条件数": 8,
            "逐case标定复算": calibration_profiles,
            "验证目录运行数": len(EXPECTED_VALIDATION_RUNS),
            "验证唯一模拟运行数": 15,
            "验证逐运行复算": validation_profile,
            "验证CSV输出数": validation_csv_count,
            "验证CSV数据行数": validation_rows,
            "验证CSV有限数值单元格数": validation_numeric_cells,
            "验证目录时间网格点数": grid_total,
            "验证唯一模拟时间网格点数": 21_284,
            "验证CSV唯一内容数": len(validation_hash_counts),
            "验证CSV重复内容组数": 11,
            "验证CSV额外副本数": 22,
            "F26基线跨目录复用文件数": f26_overlap,
            "独立材料增量": 0,
            "实验映射": {
                "Tension/iii_iv_v_vi": "手工数字化0.1/1/10/100 mm/s拉伸曲线",
                "Compression/viii_ix_x_xi": "X向Std通道1/10/100/1000 mm/s；HS不用于论文映射",
                "Torsion/F22_F23_F24": "位移约束weg0/S0的0.1/1/10 deg/s",
                "Torsion/F25_F26_F27": "无轴向约束F0的0.1/1/10 deg/s",
                "TorsionRotationCorrection/F43": "F26同工况coaxial-correction算法变体",
            },
            "独立性边界": "节点、网格分片、时间步、F26复制及正则化扫描均不增加独立样本权重",
        },
        "未来任务权重上限_尚未物化": {
            "直接原始TPU实验": 1.0,
            "手工数字化拉伸": 0.65,
            "映射Validation_FE物理代理辅助": 0.35,
            "Calibration拟合内输出_性能监督": 0.0,
            "Calibration单独FE_emulator": 0.25,
            "时间步_网格分片_精确副本_PA12_身份冲突_坏点": 0.0,
        },
        "治理": _governance(
            fidelity="direct_experiment_plus_experiment_calibrated_FE",
            split_group_key="dataset_doi|material_grade|specimen_or_simulation_family",
            source_weight_ceiling=1.0,
            candidate_eligible=True,
            note=(
                "base ceiling为1.0，任务映射质量另行乘且不得重复衰减；实验与有限元分层；"
                "PA12排除、7H/6V冲突隔离；每个模拟家族总权重封顶。"
            ),
        ),
    }
    return summary, file_rows, curve_rows


FILE_COLUMNS = ["归档", "成员", "未压缩字节", "压缩字节", "CRC32", "扩展名", "角色"]
CURVE_COLUMNS = [
    "来源",
    "材料",
    "工况或试样",
    "数据角色",
    "曲线或文件",
    "点数",
    "试样组键",
    "准入状态",
    "备注",
]


def main() -> int:
    before = scientific_input_snapshot()
    source_snapshots = snapshot_by_source(before)
    audits = {
        STANDARD: audit_standard(),
        MICROSPHERE: audit_microsphere(),
        FIGSHARE: audit_figshare(),
        TPU1301: audit_tpu1301(),
    }
    after = scientific_input_snapshot()
    if before != after:
        raise AuditBlocked("审计期间科学输入发生变化，拒绝写出")

    rendered_outputs: list[tuple[Path, bytes]] = []
    for source in SOURCE_NAMES:
        summary, file_rows, curve_rows = audits[source]
        summary["科学输入快照"] = source_snapshots[source]
        summary["科学输入运行前后不变"] = True
        summary["训练状态"] = {
            "training_split_created": False,
            "training_weight_materialized": False,
            "simulation_timestep_materialized_as_observation": False,
        }
        rendered_outputs.extend(
            [
                (
                    DATA_ROOT / source / "内容审计摘要.json",
                    (
                        json.dumps(
                            summary,
                            ensure_ascii=False,
                            sort_keys=True,
                            indent=2,
                            allow_nan=False,
                        )
                        + "\n"
                    ).encode("utf-8"),
                ),
                (
                    DATA_ROOT / source / "文件校验清单.tsv",
                    render_tsv(sorted(file_rows, key=lambda row: (str(row["归档"]), str(row["成员"]))), FILE_COLUMNS),
                ),
                (
                    DATA_ROOT / source / "曲线审计清单.tsv",
                    render_tsv(sorted(curve_rows, key=lambda row: str(row["曲线或文件"])), CURVE_COLUMNS),
                ),
            ]
        )

    if {path for path, _ in rendered_outputs} != set(OUTPUT_WHITELIST):
        raise AuditBlocked("渲染输出集合与白名单不一致")
    for path, payload in rendered_outputs:
        atomic_write(path, payload)

    output_hashes = {
        path.relative_to(PROJECT_ROOT).as_posix(): hashlib.sha256(payload).hexdigest()
        for path, payload in rendered_outputs
    }
    print(
        json.dumps(
            {
                "status": "pass",
                "sources": list(SOURCE_NAMES),
                "outputs": len(rendered_outputs),
                "output_hashes": dict(sorted(output_hashes.items())),
                "training_split_created": False,
                "training_weight_materialized": False,
            },
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"BLOCKED: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc
