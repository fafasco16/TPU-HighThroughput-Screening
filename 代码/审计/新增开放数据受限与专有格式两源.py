"""复算受限 Bath 元数据和专有格式 Mendeley 超分子 PU 来源。

本脚本没有网络能力，不下载数据，也不尝试猜读 OPJ、OPJU 或 MNOVA。它只读取
项目中已经保存的官方元数据、官方文件和权利证据，原子覆盖两个来源目录内现有的
审计 JSON/TSV。未解析专有格式和 Bath 未获许可工作簿的当前权重始终为零。

从项目根目录运行：

    python 代码/审计/新增开放数据受限与专有格式两源.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import tempfile
import zipfile
from collections import Counter
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"

MENDELEY_NAME = "Mendeley_热可逆超分子PU宽应变率"
BATH_NAME = "Bath_多牌号PU泡沫多模态表征"
SOURCE_NAMES = (MENDELEY_NAME, BATH_NAME)

OUTPUT_NAMES_BY_SOURCE = {
    MENDELEY_NAME: (
        "内容审计摘要.json",
        "文件校验清单.tsv",
        "解析状态清单.tsv",
    ),
    BATH_NAME: (
        "内容审计摘要.json",
        "文件校验清单.tsv",
    ),
}
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source_name / output_name
    for source_name, output_names in OUTPUT_NAMES_BY_SOURCE.items()
    for output_name in output_names
)
ALL_OUTPUT_NAMES = frozenset(
    output_name
    for output_names in OUTPUT_NAMES_BY_SOURCE.values()
    for output_name in output_names
)


class AuditBlocked(RuntimeError):
    """审计输入、完整性或权利边界不满足冻结协议。"""


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def md5(path: Path) -> str:
    digest = hashlib.md5(usedforsecurity=False)
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().lower()


def require_file(path: Path) -> None:
    if path.is_symlink() or not path.is_file():
        raise AuditBlocked(f"缺少普通输入文件或输入为符号链接：{path}")


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    if path.is_symlink():
        raise AuditBlocked(f"拒绝覆盖符号链接审计输出：{path}")
    if path.exists() and not path.is_file():
        raise AuditBlocked(f"审计输出不是普通文件：{path}")
    resolved_parent = path.parent.resolve(strict=True)
    absolute_parent = path.parent.absolute()
    if os.path.normcase(str(resolved_parent)) != os.path.normcase(str(absolute_parent)):
        raise AuditBlocked(f"拒绝通过重解析目录写入审计输出：{path.parent}")


def atomic_write(path: Path, payload: bytes) -> None:
    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if temporary.is_symlink() or not temporary.is_file():
            raise AuditBlocked(f"审计临时输出不是普通文件：{temporary}")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    atomic_write(
        path,
        (json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode(
            "utf-8"
        ),
    )


def write_tsv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    if not rows:
        raise AuditBlocked(f"拒绝写入空 TSV：{path}")
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="ignore",
    )
    writer.writeheader()
    writer.writerows(rows)
    atomic_write(path, buffer.getvalue().encode("utf-8"))


def original_files(base: Path) -> list[Path]:
    return sorted(
        path
        for path in base.rglob("*")
        if path.is_file()
        and path.name not in ALL_OUTPUT_NAMES
        and not path.name.endswith(".audit.tmp")
    )


def raw_snapshot() -> dict[str, tuple[int, str]]:
    snapshot: dict[str, tuple[int, str]] = {}
    for source_name in SOURCE_NAMES:
        base = DATA_ROOT / source_name
        for path in original_files(base):
            snapshot[path.relative_to(PROJECT_ROOT).as_posix()] = (
                path.stat().st_size,
                sha256(path),
            )
    return snapshot


def require_equal(label: str, actual: Any, expected: Any) -> None:
    if actual != expected:
        raise AuditBlocked(f"{label} 漂移：实际 {actual!r}，协议期望 {expected!r}")


def validate_zip(path: Path) -> None:
    require_file(path)
    with zipfile.ZipFile(path) as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise AuditBlocked(f"ZIP CRC 失败：{path.name} / {bad_member}")
        for item in archive.infolist():
            normalized = item.filename.replace("\\", "/")
            member = PurePosixPath(normalized)
            has_drive_prefix = bool(member.parts and member.parts[0].endswith(":"))
            if member.is_absolute() or has_drive_prefix or ".." in member.parts:
                raise AuditBlocked(f"ZIP 包含不安全成员：{path.name} / {item.filename}")


def is_finite_number(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    )


def workbook_stats(path: Path) -> dict[str, int]:
    validate_zip(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    numeric = 0
    formulas = 0
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if is_finite_number(value):
                        numeric += 1
                    elif isinstance(value, str) and value.startswith("="):
                        formulas += 1
        return {
            "sheet_count": len(workbook.sheetnames),
            "finite_numeric_cells": numeric,
            "formula_cells": formulas,
        }
    finally:
        workbook.close()


def count_gpc_pairs(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=False)
    count = 0
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == "Results":
                continue
            in_curve_table = False
            for row in worksheet.iter_rows(values_only=True):
                if any(value == "RT (mins)" for value in row):
                    in_curve_table = True
                    continue
                if (
                    in_curve_table
                    and len(row) >= 3
                    and is_finite_number(row[1])
                    and is_finite_number(row[2])
                ):
                    count += 1
        return count
    finally:
        workbook.close()


def count_saxs_points(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=False)
    count = 0
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == "Plot":
                continue
            for row in worksheet.iter_rows(values_only=True):
                if len(row) >= 3 and all(is_finite_number(value) for value in row[:3]):
                    count += 1
        return count
    finally:
        workbook.close()


def count_cooling_pairs(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=False)
    count = 0
    try:
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows(values_only=True):
            if len(row) >= 2 and all(is_finite_number(value) for value in row[:2]):
                count += 1
            if len(row) >= 5 and all(is_finite_number(value) for value in row[3:5]):
                count += 1
        return count
    finally:
        workbook.close()


def dsc_stats(path: Path) -> dict[str, int]:
    require_file(path)
    measurement_rows = 0
    marker_rows = 0
    finite_values = 0
    for line in path.read_text(encoding="utf-16le").splitlines():
        parts = line.split("\t")
        if len(parts) != 3:
            continue
        try:
            values = [float(part) for part in parts]
        except ValueError:
            continue
        if not all(math.isfinite(value) for value in values):
            raise AuditBlocked(f"DSC 出现非有限数值：{line!r}")
        if values[0] < 0:
            marker_rows += 1
        else:
            measurement_rows += 1
            finite_values += 3
    return {
        "measurement_rows": measurement_rows,
        "marker_rows": marker_rows,
        "finite_measurement_values": finite_values,
    }


def mendeley_status(extension: str) -> dict[str, str]:
    if extension == ".xlsx":
        return {
            "parser_status": "parsed_read_only_openpyxl",
            "content_role": "机器可读实验辅助数据",
            "quality_tier": "gold",
            "admissibility": "parsed_auxiliary_only",
            "observation_unit": "同一SPU配方的条件或采集",
            "note": "只读解析；按条件/曲线计，不按单元格计独立样本。",
        }
    if extension == ".txt":
        return {
            "parser_status": "parsed_utf16le_tabular",
            "content_role": "DSC原始数据",
            "quality_tier": "gold",
            "admissibility": "parsed_auxiliary_only",
            "observation_unit": "一个DSC采集",
            "note": "控制标记行与测量行分开，十个程序段不计十个独立采集。",
        }
    if extension == ".pdf":
        return {
            "parser_status": "visual_evidence_only",
            "content_role": "NMR视觉证据",
            "quality_tier": "silver",
            "admissibility": "no_numeric_target",
            "observation_unit": "谱图图层（非数值曲线）",
            "note": "不从PDF图像数字化或生成机械标签。",
        }
    if extension == ".slx":
        return {
            "parser_status": "inspected_zip_xml",
            "content_role": "Simscape模型资产",
            "quality_tier": "silver",
            "admissibility": "simulation_auxiliary_only",
            "observation_unit": "模型资产（非实验样本）",
            "note": "容器完整但不是实验真值。",
        }
    if extension == ".opj":
        return {
            "parser_status": "pending_parser",
            "content_role": "Origin经典项目",
            "quality_tier": "pending",
            "admissibility": "not_admitted_until_reliable_parse",
            "observation_unit": "未知；禁止按文件或图号推断",
            "note": "没有固定且验证通过的OPJ解析器，不做二进制猜读。",
        }
    if extension == ".opju":
        return {
            "parser_status": "quarantined_unsupported_opju",
            "content_role": "Origin Unicode项目",
            "quality_tier": "quarantine",
            "admissibility": "not_admitted",
            "observation_unit": "未知；禁止推断",
            "note": "专有OPJU容器，当前无验证通过的解析路径。",
        }
    if extension == ".mnova":
        return {
            "parser_status": "quarantined_proprietary_mnova",
            "content_role": "MestReNova项目",
            "quality_tier": "quarantine",
            "admissibility": "not_admitted",
            "observation_unit": "未知；禁止推断",
            "note": "专有MNOVA容器，仅保留文件级证据。",
        }
    raise AuditBlocked(f"Mendeley 出现未登记格式：{extension}")


def audit_mendeley() -> dict[str, Any]:
    base = DATA_ROOT / MENDELEY_NAME
    official_manifest_path = base / "官方完整文件清单.json"
    official_files_dir = base / "官方文件"
    require_file(official_manifest_path)
    if official_files_dir.is_symlink() or not official_files_dir.is_dir():
        raise AuditBlocked("Mendeley 官方文件目录缺失或为重解析链接")

    official_rows = json.loads(official_manifest_path.read_text(encoding="utf-8"))
    if not isinstance(official_rows, list):
        raise AuditBlocked("Mendeley 官方文件清单不是列表")
    require_equal("Mendeley 官方文件数", len(official_rows), 34)
    if len({row["filename"] for row in official_rows}) != len(official_rows):
        raise AuditBlocked("Mendeley 官方清单存在重复文件名")

    actual_files = sorted(path for path in official_files_dir.iterdir() if path.is_file())
    require_equal("Mendeley 本地官方文件数", len(actual_files), 34)
    actual_by_name = {path.name: path for path in actual_files}
    if set(actual_by_name) != {row["filename"] for row in official_rows}:
        raise AuditBlocked("Mendeley 本地文件名与官方清单不一致")

    manifest_rows: list[dict[str, Any]] = []
    extension_counts: Counter[str] = Counter()
    total_bytes = 0
    for row in sorted(official_rows, key=lambda item: item["filename"].casefold()):
        filename = row["filename"]
        details = row["content_details"]
        path = actual_by_name[filename]
        extension = path.suffix.lower()
        extension_counts[extension] += 1
        total_bytes += path.stat().st_size
        expected_size = int(details["size"])
        expected_sha = str(details["sha256_hash"]).upper()
        local_sha = sha256(path)
        size_match = path.stat().st_size == expected_size
        hash_match = local_sha == expected_sha
        if not size_match or not hash_match:
            raise AuditBlocked(f"Mendeley 文件完整性不匹配：{filename}")
        if extension in {".xlsx", ".slx"}:
            validate_zip(path)
        status = mendeley_status(extension)
        manifest_rows.append(
            {
                "filename": filename,
                "extension": extension,
                "local_size_bytes": path.stat().st_size,
                "official_size_bytes": expected_size,
                "local_sha256": local_sha,
                "official_sha256": expected_sha,
                "size_match": str(size_match).lower(),
                "sha256_match": str(hash_match).lower(),
                **status,
            }
        )

    require_equal("Mendeley 官方总字节", total_bytes, 44_711_485)
    require_equal(
        "Mendeley 格式分布",
        dict(sorted(extension_counts.items())),
        {
            ".mnova": 1,
            ".opj": 24,
            ".opju": 3,
            ".pdf": 1,
            ".slx": 1,
            ".txt": 1,
            ".xlsx": 3,
        },
    )

    gpc_path = official_files_dir / "FIG3 - GPC-RawData.xlsx"
    saxs_path = official_files_dir / "FIG5 - SAXS.xlsx"
    cooling_path = official_files_dir / "FIGSI1 - from thermometer.xlsx"
    dsc_path = official_files_dir / "FIGSI3-Example of DSC raw data.txt"
    gpc = workbook_stats(gpc_path)
    saxs = workbook_stats(saxs_path)
    cooling = workbook_stats(cooling_path)
    dsc = dsc_stats(dsc_path)
    gpc_pairs = count_gpc_pairs(gpc_path)
    saxs_points = count_saxs_points(saxs_path)
    cooling_pairs = count_cooling_pairs(cooling_path)

    require_equal("GPC 工作表数", gpc["sheet_count"], 10)
    require_equal("GPC 有限数值", gpc["finite_numeric_cells"], 13_081)
    # 9 个条件页从 ``RT (mins)`` 表头后逐行复算为 6,446 对；旧摘要的
    # 6,395 少计 51 行，与 13,081 个有限数值单元格的闭合关系不成立。
    require_equal("GPC 曲线点对", gpc_pairs, 6_446)
    require_equal("SAXS 工作表数", saxs["sheet_count"], 8)
    require_equal("SAXS 有限数值", saxs["finite_numeric_cells"], 30_933)
    require_equal("SAXS 公式数", saxs["formula_cells"], 2)
    require_equal("SAXS 曲线点", saxs_points, 10_311)
    require_equal("冷却曲线点对", cooling_pairs, 162)
    require_equal("冷却有限数值", cooling["finite_numeric_cells"], 324)
    require_equal("DSC 测量行", dsc["measurement_rows"], 19_000)
    require_equal("DSC 标记行", dsc["marker_rows"], 10)
    require_equal("DSC 有限测量值", dsc["finite_measurement_values"], 57_000)

    synchronized_rows = gpc_pairs + saxs_points + cooling_pairs + dsc["measurement_rows"]
    require_equal("机器可读同步点行", synchronized_rows, 35_919)

    status_rows = [
        {
            "format": ".xlsx",
            "file_count": 3,
            "reliably_inspected_or_parsed": 3,
            "status": "parsed_read_only_openpyxl",
            "ratio": "100.00%",
            "method_or_attempt": "openpyxl read_only=True；OOXML CRC",
            "admission_rule": "实验辅助层；按条件/曲线计，不按单元格计样本",
        },
        {
            "format": ".txt",
            "file_count": 1,
            "reliably_inspected_or_parsed": 1,
            "status": "parsed_utf16le_tabular",
            "ratio": "100.00%",
            "method_or_attempt": "UTF-16LE + tab parser",
            "admission_rule": "DSC原始采集；排除程序控制标记",
        },
        {
            "format": ".pdf",
            "file_count": 1,
            "reliably_inspected_or_parsed": 1,
            "status": "visual_evidence_only",
            "ratio": "视觉100%; 数值0%",
            "method_or_attempt": "既有人工视觉证据；本脚本只校验文件",
            "admission_rule": "不数字化，不生成数值标签",
        },
        {
            "format": ".slx",
            "file_count": 1,
            "reliably_inspected_or_parsed": 1,
            "status": "inspected_zip_xml",
            "ratio": "结构100%; 实验数值0%",
            "method_or_attempt": "ZIP CRC/结构检查",
            "admission_rule": "模型辅助资产，不计实验样本",
        },
        {
            "format": ".opj",
            "file_count": 24,
            "reliably_inspected_or_parsed": 0,
            "status": "pending_parser",
            "ratio": "0.00%",
            "method_or_attempt": "无固定且验证通过的Ropj/liborigin解析器",
            "admission_rule": "不得二进制猜读；当前权重0",
        },
        {
            "format": ".opju",
            "file_count": 3,
            "reliably_inspected_or_parsed": 0,
            "status": "quarantined_unsupported_opju",
            "ratio": "0.00%",
            "method_or_attempt": "无验证通过的OPJU解析路径",
            "admission_rule": "隔离；当前权重0",
        },
        {
            "format": ".mnova",
            "file_count": 1,
            "reliably_inspected_or_parsed": 0,
            "status": "quarantined_proprietary_mnova",
            "ratio": "0.00%",
            "method_or_attempt": "专有MestReNova容器",
            "admission_rule": "隔离；当前权重0",
        },
    ]

    summary = {
        "audit_date": AUDIT_DATE,
        "audit_schema_version": "tpu-proprietary-format-audit-v1.1",
        "source": {
            "doi": "10.17632/byjbmymyhh.5",
            "repository": "Mendeley Data",
            "version": 5,
            "license": "CC BY-NC 3.0",
            "independent_chemistry_count": 1,
            "specimen_count": None,
        },
        "integrity": {
            "official_file_count": 34,
            "local_file_count": 34,
            "official_and_local_total_bytes": total_bytes,
            "sha256_and_size_match_count": 34,
            "extension_counts": dict(sorted(extension_counts.items())),
            "all_verified": True,
        },
        "parsed_evidence": {
            "reliably_parsed_experiment_assets": 4,
            "direct_numeric_curve_acquisitions": 19,
            "definition": "9 GPC + 7 SAXS + 2 cooling + 1 DSC run",
            "synchronized_point_rows": synchronized_rows,
            "gpc": {**gpc, "curve_count": 9, "curve_point_pairs": gpc_pairs},
            "saxs": {**saxs, "curve_count": 7, "curve_points": saxs_points},
            "cooling": {**cooling, "curve_count": 2, "curve_point_pairs": cooling_pairs},
            "dsc": {**dsc, "run_count": 1},
        },
        "unparsed_or_quarantined": {
            "opj": {"file_count": 24, "current_weight_ceiling": 0.0},
            "opju": {"file_count": 3, "current_weight_ceiling": 0.0},
            "mnova": {"file_count": 1, "current_weight_ceiling": 0.0},
            "rule": "文件数和图号不能推断曲线、试样或标签；缺固定解析器时保持零权重。",
        },
        "admission": {
            "parsed_process_thermal_morphology_transfer_ceiling": 0.35,
            "core_tpu_chemistry_property_model": False,
            "reason": "只有一个SPU化学体系，不能按点数或条件数扩大化学空间。",
            "split_group_key": "dataset_doi|material_chemistry_id|source_batch_or_unknown",
        },
        "formal_citations": [
            "Chen, H.; Hart, L. R.; Hayes, W.; Siviour, C. R. (2021). MECHANICAL CHARACTERISATION AND MODELLING OF A THERMOREVERSIBLE SUPERAMOLECULAR POLYURETHANE OVER A WIDE RANGE OF RATES. Mendeley Data, V5. https://doi.org/10.17632/byjbmymyhh.5",
            "Chen, H.; Hart, L. R.; Hayes, W.; Siviour, C. R. Mechanical characterisation and modelling of a thermoreversible superamolecular polyurethane over a wide range of rates. Polymer 221 (2021) 123607. https://doi.org/10.1016/j.polymer.2021.123607",
        ],
    }

    write_tsv(
        base / "文件校验清单.tsv",
        manifest_rows,
        [
            "filename",
            "extension",
            "local_size_bytes",
            "official_size_bytes",
            "local_sha256",
            "official_sha256",
            "size_match",
            "sha256_match",
            "parser_status",
            "content_role",
            "quality_tier",
            "admissibility",
            "observation_unit",
            "note",
        ],
    )
    write_tsv(
        base / "解析状态清单.tsv",
        status_rows,
        [
            "format",
            "file_count",
            "reliably_inspected_or_parsed",
            "status",
            "ratio",
            "method_or_attempt",
            "admission_rule",
        ],
    )
    write_json(base / "内容审计摘要.json", summary)
    return summary


def read_tsv(path: Path) -> list[dict[str, str]]:
    require_file(path)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def audit_bath() -> dict[str, Any]:
    base = DATA_ROOT / BATH_NAME
    official_inventory_path = base / "官方文件清单.tsv"
    datacite_path = base / "官方DataCite元数据.json"
    eprints_path = base / "官方EPrints元数据.json"
    for path in (official_inventory_path, datacite_path, eprints_path):
        require_file(path)

    official_rows = read_tsv(official_inventory_path)
    require_equal("Bath 官方工作簿数", len(official_rows), 10)
    require_equal(
        "Bath 官方文件总字节",
        sum(int(row["bytes"]) for row in official_rows),
        27_150_386,
    )
    if len({row["filename"] for row in official_rows}) != 10:
        raise AuditBlocked("Bath 官方文件清单存在重复文件名")
    if not all(len(row["official_md5"]) == 32 for row in official_rows):
        raise AuditBlocked("Bath 官方文件清单缺少 MD5")
    if any(row["license_in_official_record"] != "未注明" for row in official_rows):
        raise AuditBlocked("Bath 官方文件许可状态与冻结协议不一致")
    if any(row["local_file_downloaded"] != "否" for row in official_rows):
        raise AuditBlocked("Bath 出现未经许可下载的原始工作簿")

    local_workbooks = sorted(base.rglob("*.xlsx"))
    require_equal("Bath 本地原始工作簿数", len(local_workbooks), 0)

    datacite = json.loads(datacite_path.read_text(encoding="utf-8"))
    eprints = json.loads(eprints_path.read_text(encoding="utf-8"))
    require_equal("Bath DataCite DOI", str(datacite.get("doi", "")).upper(), "10.15125/BATH-00385")
    if datacite.get("rightsList"):
        raise AuditBlocked("Bath DataCite 新出现 rightsList，必须重新做权利裁决")
    documents = eprints.get("documents", [])
    if not isinstance(documents, list) or len(documents) != 10:
        raise AuditBlocked("Bath EPrints document 清单漂移")
    if any(document.get("licence") for document in documents):
        raise AuditBlocked("Bath EPrints 新出现 licence，必须重新做权利裁决")
    eprints_files: dict[str, dict[str, Any]] = {}
    for document in documents:
        files = document.get("files", [])
        if not isinstance(files, list) or len(files) != 1:
            raise AuditBlocked("Bath EPrints 每个 document 必须恰含一个文件")
        file_record = files[0]
        filename = str(file_record.get("filename", ""))
        if not filename or document.get("main") != filename:
            raise AuditBlocked("Bath EPrints main 与文件名不一致")
        if filename in eprints_files:
            raise AuditBlocked(f"Bath EPrints 出现重复文件名：{filename}")
        if str(file_record.get("hash_type", "")).upper() != "MD5":
            raise AuditBlocked(f"Bath EPrints 哈希算法不是 MD5：{filename}")
        eprints_files[filename] = {
            "bytes": int(file_record["filesize"]),
            "md5": str(file_record["hash"]).lower(),
            "mime_type": str(file_record["mime_type"]),
            "document_revision": str(document["rev_number"]),
            "security": str(document["security"]),
            "position": str(document["pos"]),
        }
    if set(eprints_files) != {row["filename"] for row in official_rows}:
        raise AuditBlocked("Bath TSV 与 EPrints 的文件名集合不一致")
    for row in official_rows:
        expected = eprints_files[row["filename"]]
        observed = {
            "bytes": int(row["bytes"]),
            "md5": row["official_md5"].lower(),
            "mime_type": row["mime_type"],
            "document_revision": row["document_revision"],
            "security": row["security"],
            "position": row["position"],
        }
        if observed != expected:
            raise AuditBlocked(
                f"Bath TSV 与 EPrints 逐文件证据不一致：{row['filename']}"
            )

    local_rows: list[dict[str, Any]] = []
    for path in original_files(base):
        relative_path = path.relative_to(base).as_posix()
        role = (
            "official_evidence"
            if path.name.startswith("官方")
            else "local_rights_or_request_record"
        )
        local_rows.append(
            {
                "relative_path": relative_path,
                "bytes": path.stat().st_size,
                "md5": md5(path),
                "sha256": sha256(path).lower(),
                "role": role,
            }
        )

    modalities = [
        "compression",
        "flexure",
        "shear",
        "creep",
        "drilling_resistance",
        "XRD",
        "FTIR",
        "DMA",
        "STA",
        "rheology",
    ]
    summary = {
        "audit_date": AUDIT_DATE,
        "audit_schema_version": "bath-pu-metadata-audit-v1.1",
        "source": {
            "doi": "10.15125/BATH-00385",
            "repository": "University of Bath Research Data Archive",
            "version": 1,
            "rights_holder": "University of Bath",
        },
        "license_and_access_decision": {
            "metadata_license": "CC0",
            "raw_data_license_in_record": None,
            "rights_evidence_state": "evidence_missing",
            "raw_data_downloaded": False,
            "decision": "仅保存元数据和条款证据；未获书面复用许可前不下载、不打开十个Excel。",
        },
        "official_file_inventory": {
            "file_count": 10,
            "format_counts": {"xlsx": 10},
            "total_bytes": 27_150_386,
            "all_files_have_official_md5": True,
            "raw_files_locally_present": 0,
        },
        "scientific_scope_from_official_metadata_only": {
            "independent_material_grades_identified": 3,
            "material_grades": ["Reprocell 500", "Reprocell 300", "LD40"],
            "modalities_claimed_by_record": modalities,
            "specimen_count": None,
            "curve_count": None,
            "numeric_point_count": None,
            "unit_audit_completed": False,
        },
        "database_admission": {
            "current_layer": "metadata_only_candidate",
            "current_weight_ceiling": 0.0,
            "core_tpu_dataset": False,
            "future_mechanical_transfer_ceiling_after_written_permission_and_full_audit": 0.25,
            "future_representation_ceiling_after_written_permission_and_full_audit": 0.35,
            "split_group_key": "dataset_doi|material_grade|batch",
        },
        "hard_blocks": [
            "record_has_no_raw_data_license",
            "written_reuse_permission_not_obtained",
            "raw_workbooks_not_downloaded",
            "specimen_curve_unit_and_duplicate_audit_not_possible",
        ],
        "formal_citation": "Dams, B. (2017). Reprocell 500, Reprocell 300 and LD40 Polyurethane foam mechanical and characterisation tests October 2016 - April 2017. University of Bath Research Data Archive. https://doi.org/10.15125/BATH-00385.",
    }

    write_tsv(
        base / "文件校验清单.tsv",
        local_rows,
        ["relative_path", "bytes", "md5", "sha256", "role"],
    )
    write_json(base / "内容审计摘要.json", summary)
    return summary


def main() -> None:
    before = raw_snapshot()
    mendeley = audit_mendeley()
    bath = audit_bath()
    after = raw_snapshot()
    if before != after:
        raise AuditBlocked("科学输入在审计运行期间发生变化")
    print(
        json.dumps(
            {
                "audit_date": AUDIT_DATE,
                "sources": list(SOURCE_NAMES),
                "scientific_input_count": len(before),
                "mendeley_verified_files": mendeley["integrity"]["local_file_count"],
                "mendeley_direct_acquisitions": mendeley["parsed_evidence"][
                    "direct_numeric_curve_acquisitions"
                ],
                "bath_local_raw_workbooks": bath["official_file_inventory"][
                    "raw_files_locally_present"
                ],
                "training_created": False,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
