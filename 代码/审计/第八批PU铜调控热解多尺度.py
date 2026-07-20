"""审计 Zenodo 18414263 的 PU 铜调控热解多尺度工作簿。

输出采用曲线通道、实验条件和计算物理体系级粒度。工作簿中的采样点、
路径坐标点与 ESP 网格分箱只计为观测点，绝不据此扩增独立样本数。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import tempfile
import zipfile
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第八批混合_PU铜调控热解多尺度"
)
SOURCE_XLSX = SOURCE_DIR / "原始数据.xlsx"

DOI = "10.5281/zenodo.18414263"
CONCEPT_DOI = "10.5281/zenodo.18414262"
LICENSE = "CC BY 4.0"
SOURCE_URL = "https://zenodo.org/records/18414263"
ARTICLE_DOI = "10.1038/s43247-026-03339-9"
AUDIT_VERSION = "batch8-pu-copper-pyrolysis-multiscale-v1"

EXPECTED_BYTES = 3_140_743
EXPECTED_MD5 = "b9bf18264a6338e9ef1d7ffe23ef9b91"
EXPECTED_SHA256 = (
    "7f63eba865fced2234dd073fad230936bb5159cd8e24bcf7ed8a63a700a8022f"
)
EXPECTED_ZIP_ENTRIES = 25
EXPECTED_UNCOMPRESSED_BYTES = 16_432_122
EXPECTED_SHEETS = (
    "Figure 1a-d",
    "Figure 1e activation energy",
    "Figure1e Cu-free TG DTG",
    "Figure2",
    "Figure 3",
    "Figure4",
    "Figure5",
    "Figure 6 and Figure 7",
    "Supplementary Figure3",
)
EXPECTED_DIMENSIONS = {
    "Figure 1a-d": (19_002, 24),
    "Figure 1e activation energy": (24, 17),
    "Figure1e Cu-free TG DTG": (18_997, 3),
    "Figure2": (6, 10),
    "Figure 3": (46, 17),
    "Figure4": (50, 8),
    "Figure5": (35, 8),
    "Figure 6 and Figure 7": (89, 25),
    "Supplementary Figure3": (7_458, 4),
}
EXPECTED_NONEMPTY_BY_SHEET = {
    "Figure 1a-d": 175_664,
    "Figure 1e activation energy": 261,
    "Figure1e Cu-free TG DTG": 56_991,
    "Figure2": 42,
    "Figure 3": 339,
    "Figure4": 132,
    "Figure5": 162,
    "Figure 6 and Figure 7": 1_002,
    "Supplementary Figure3": 29_830,
}
OUTPUT_NAMES = (
    "内容审计摘要.json",
    "实验曲线审计清单.tsv",
    "计算观测清单.tsv",
    "文件校验清单.tsv",
)

CURVE_COLUMNS = (
    "curve_id",
    "sheet",
    "source_columns",
    "material_state",
    "copper_status",
    "heating_rate_c_per_min",
    "response_type",
    "axis_unit",
    "response_unit",
    "observed_response_points",
    "paired_response_points",
    "orphan_axis_points",
    "orphan_response_points",
    "internal_blank_rows",
    "numeric_text_cells",
    "unique_axis_values",
    "independent_condition_id",
    "reported_replicates",
    "replicate_values_available",
    "independent_specimen_count",
    "record_granularity",
    "target_origin",
    "gold_layer",
    "admission_status",
    "future_weight_ceiling",
    "qc_flags",
    "notes",
)

COMPUTE_COLUMNS = (
    "observation_id",
    "sheet",
    "observation_kind",
    "system_id",
    "copper_status",
    "method",
    "protocol",
    "point_count",
    "surface_minima_count",
    "surface_maxima_count",
    "esp_area_bin_count",
    "path_range",
    "target_origin",
    "record_granularity",
    "independent_weight_unit",
    "gold_layer",
    "admission_status",
    "direct_target_weight_ceiling",
    "future_weight_ceiling",
    "qc_flags",
    "notes",
)


class AuditBlocked(RuntimeError):
    """原件身份、结构或冻结科学计数发生漂移。"""


def _hash(path: Path, algorithm: str) -> str:
    digest = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        result = float(value)
    elif isinstance(value, str):
        try:
            result = float(value.strip())
        except ValueError:
            return None
    else:
        return None
    return result if math.isfinite(result) else None


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise AuditBlocked(message)


def _verify_xlsx(path: Path) -> tuple[dict[str, object], dict[str, object]]:
    if not path.is_file():
        raise AuditBlocked(f"缺少原始工作簿：{path}")
    size = path.stat().st_size
    md5 = _hash(path, "md5")
    sha256 = _hash(path, "sha256")
    _require(size == EXPECTED_BYTES, f"原件字节数漂移：{size}")
    _require(md5 == EXPECTED_MD5, f"原件 MD5 漂移：{md5}")
    _require(sha256 == EXPECTED_SHA256, f"原件 SHA256 漂移：{sha256}")

    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        names = [item.filename for item in infos]
        unsafe = []
        for name in names:
            pure = PurePosixPath(name)
            if pure.is_absolute() or ".." in pure.parts or "\\" in name:
                unsafe.append(name)
        encrypted = [item.filename for item in infos if item.flag_bits & 0x1]
        lowered = [name.casefold() for name in names]
        active = [
            name
            for name in names
            if name.casefold().endswith("vbaproject.bin")
            or "/externallinks/" in name.casefold()
            or "/embeddings/" in name.casefold()
            or "/activex/" in name.casefold()
        ]
        uncompressed = sum(item.file_size for item in infos)
    _require(not unsafe, f"OOXML 包含不安全路径：{unsafe}")
    _require(not encrypted, f"OOXML 包含加密条目：{encrypted}")
    _require(not active, f"OOXML 包含活动或外部内容：{active}")
    _require(len(lowered) == EXPECTED_ZIP_ENTRIES, "OOXML 条目数漂移")
    _require(
        uncompressed == EXPECTED_UNCOMPRESSED_BYTES,
        f"OOXML 解压字节数漂移：{uncompressed}",
    )
    file_row = {
        "file": path.name,
        "bytes": size,
        "md5": md5,
        "sha256": sha256,
        "verification": "matched_frozen_identity",
        "license": LICENSE,
        "doi": DOI,
    }
    archive_summary = {
        "zip_entry_count": len(names),
        "uncompressed_bytes": uncompressed,
        "unsafe_path_count": 0,
        "encrypted_entry_count": 0,
        "active_or_external_entry_count": 0,
    }
    return file_row, archive_summary


def _scan_workbook(workbook) -> tuple[dict[str, dict[str, int]], dict[str, int]]:
    _require(tuple(workbook.sheetnames) == EXPECTED_SHEETS, "工作表清单或顺序漂移")
    inventory: dict[str, dict[str, int]] = {}
    totals = {
        "nonempty": 0,
        "native_numeric": 0,
        "strings": 0,
        "numeric_text": 0,
        "formulas": 0,
        "errors": 0,
    }
    for name in EXPECTED_SHEETS:
        sheet = workbook[name]
        _require(
            (sheet.max_row, sheet.max_column) == EXPECTED_DIMENSIONS[name],
            f"{name} 维度漂移：{sheet.max_row}x{sheet.max_column}",
        )
        local = {key: 0 for key in totals}
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                local["nonempty"] += 1
                if cell.data_type == "f":
                    local["formulas"] += 1
                if cell.data_type == "e":
                    local["errors"] += 1
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    local["native_numeric"] += 1
                elif isinstance(value, str):
                    local["strings"] += 1
                    if _number(value) is not None:
                        local["numeric_text"] += 1
        _require(
            local["nonempty"] == EXPECTED_NONEMPTY_BY_SHEET[name],
            f"{name} 非空单元格计数漂移：{local['nonempty']}",
        )
        inventory[name] = {
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            **local,
        }
        for key in totals:
            totals[key] += local[key]
    _require(totals["formulas"] == 0, "工作簿出现公式")
    _require(totals["errors"] == 0, "工作簿出现 Excel 错误值")
    _require(totals["nonempty"] == 264_423, "全簿非空单元格计数漂移")
    _require(totals["native_numeric"] == 263_503, "全簿原生数值计数漂移")
    _require(totals["strings"] == 920, "全簿字符串计数漂移")
    _require(totals["numeric_text"] == 610, "全簿数值型字符串计数漂移")
    return inventory, totals


def _curve_row(
    sheet,
    *,
    curve_id: str,
    x_col: int,
    y_col: int,
    material_state: str,
    copper_status: str,
    heating_rate: int | str,
    response_type: str,
    axis_unit: str,
    response_unit: str,
    condition_id: str,
    reported_replicates: int | str,
    ceiling: float,
    status: str,
    qc_flags: str,
    notes: str,
) -> dict[str, object]:
    x_values: list[float | None] = []
    y_values: list[float | None] = []
    numeric_text = 0
    for row in sheet.iter_rows(
        min_row=2, max_row=sheet.max_row, min_col=min(x_col, y_col), max_col=max(x_col, y_col)
    ):
        x_cell = row[x_col - min(x_col, y_col)]
        y_cell = row[y_col - min(x_col, y_col)]
        x = _number(x_cell.value)
        y = _number(y_cell.value)
        x_values.append(x)
        y_values.append(y)
        numeric_text += int(isinstance(x_cell.value, str) and x is not None)
        numeric_text += int(isinstance(y_cell.value, str) and y is not None)
    active = [i for i, (x, y) in enumerate(zip(x_values, y_values)) if x is not None or y is not None]
    internal_blanks = 0
    if active:
        internal_blanks = sum(
            x_values[i] is None and y_values[i] is None
            for i in range(active[0], active[-1] + 1)
        )
    paired = sum(x is not None and y is not None for x, y in zip(x_values, y_values))
    observed = sum(y is not None for y in y_values)
    row = {
        "curve_id": curve_id,
        "sheet": sheet.title,
        "source_columns": f"{get_column_letter(x_col)}:{get_column_letter(y_col)}",
        "material_state": material_state,
        "copper_status": copper_status,
        "heating_rate_c_per_min": heating_rate,
        "response_type": response_type,
        "axis_unit": axis_unit,
        "response_unit": response_unit,
        "observed_response_points": observed,
        "paired_response_points": paired,
        "orphan_axis_points": sum(x is not None and y is None for x, y in zip(x_values, y_values)),
        "orphan_response_points": sum(y is not None and x is None for x, y in zip(x_values, y_values)),
        "internal_blank_rows": internal_blanks,
        "numeric_text_cells": numeric_text,
        "unique_axis_values": len({x for x in x_values if x is not None}),
        "independent_condition_id": condition_id,
        "reported_replicates": reported_replicates,
        "replicate_values_available": False,
        "independent_specimen_count": 0,
        "record_granularity": "response_curve",
        "target_origin": "experimental",
        "gold_layer": "Gold-E",
        "admission_status": status,
        "future_weight_ceiling": ceiling,
        "qc_flags": qc_flags,
        "notes": notes,
        "_axis_max": max((x for x in x_values if x is not None), default=None),
    }
    return row


def _experimental_curves(workbook) -> list[dict[str, object]]:
    sheet = workbook["Figure 1a-d"]
    specs = (
        (1, 2, 5, "DTG"),
        (3, 4, 5, "TG"),
        (6, 7, 10, "TG"),
        (8, 9, 10, "DTG"),
        (11, 12, 15, "TG"),
        (13, 14, 15, "DTG"),
        (16, 17, 25, "TG"),
        (18, 19, 25, "DTG"),
        (21, 22, 20, "TG"),
        (23, 24, 20, "DTG"),
    )
    rows: list[dict[str, object]] = []
    for x_col, y_col, rate, response in specs:
        conflict = rate == 20
        rows.append(
            _curve_row(
                sheet,
                curve_id=f"fig1_{rate}cpm_{response.lower()}",
                x_col=x_col,
                y_col=y_col,
                material_state="commercial_PU_enamelled_copper_wire",
                copper_status="Cu-containing",
                heating_rate=rate,
                response_type=response,
                axis_unit="degC",
                response_unit="percent" if response == "TG" else "percent_per_min",
                condition_id=f"tga_cu_{rate}cpm",
                reported_replicates=3,
                ceiling=0.10 if conflict else 0.25,
                status="conditional_reference" if conflict else "admitted_reference",
                qc_flags=(
                    "article_protocol_lists_5_10_15_25_not_20;averaged_curve_only"
                    if conflict
                    else "averaged_curve_only"
                ),
                notes="论文称条件做三次，但工作簿未给逐重复曲线。",
            )
        )

    cu_free = workbook["Figure1e Cu-free TG DTG"]
    for y_col, response in ((2, "TG"), (3, "DTG")):
        rows.append(
            _curve_row(
                cu_free,
                curve_id=f"fig1_cu_free_5cpm_{response.lower()}",
                x_col=1,
                y_col=y_col,
                material_state="PU_enamel_Cu-free_reference",
                copper_status="Cu-free",
                heating_rate=5,
                response_type=response,
                axis_unit="degC",
                response_unit="percent" if response == "TG" else "percent_per_min",
                condition_id="tga_cu_free_5cpm",
                reported_replicates="not_resolved",
                ceiling=0.20,
                status="admitted_reference",
                qc_flags="averaged_curve_only",
                notes="无配方、分子量与逐重复信息。",
            )
        )

    ftir = workbook["Supplementary Figure3"]
    for x_col, y_col, state in ((1, 2, "before_pyrolysis"), (3, 4, "after_pyrolysis")):
        rows.append(
            _curve_row(
                ftir,
                curve_id=f"supp_fig3_ftir_{state}",
                x_col=x_col,
                y_col=y_col,
                material_state=state,
                copper_status="Cu-containing",
                heating_rate="not_applicable",
                response_type="FTIR_intensity",
                axis_unit="cm^-1",
                response_unit="unspecified",
                condition_id=f"ftir_{state}",
                reported_replicates="not_resolved",
                ceiling=0.10,
                status="conditional_reference",
                qc_flags="rounded_repeated_axis;response_unit_unspecified",
                notes="7457 行仅含 900 个唯一波数，不能把行数当独立频率样本。",
            )
        )
    _require(len(rows) == 14, "实验曲线通道数漂移")
    _require(sum(int(row["observed_response_points"]) for row in rows) == 140_680, "实验响应点计数漂移")
    _require(sum(int(row["paired_response_points"]) for row in rows) == 140_675, "实验配对点计数漂移")
    return rows


def _compute_observations(workbook) -> list[dict[str, object]]:
    common = {
        "target_origin": "computed",
        "record_granularity": "physical_system_observation",
        "gold_layer": "Gold-C",
        # 来源级仍正式准入 Gold-E+Gold-C；但逐条计算记录存在路径、基组、
        # 几何/输出或面积单位缺口，只能作为 Gold-C 条件参考。
        "admission_status": "conditional_reference",
    }
    rows: list[dict[str, object]] = []
    topology = workbook["Figure4"]
    path_blocks = (("S1", 4, 24), ("M1", 29, 50))
    all_path_ids: list[int] = []
    for system, start, end in path_blocks:
        path_ids = [int(topology.cell(row, 1).value) for row in range(start, end + 1)]
        all_path_ids.extend(path_ids)
        rows.append(
            {
                "observation_id": f"fig4_{system.lower()}_reaction_topology",
                "sheet": "Figure4",
                "observation_kind": "reaction_topology",
                "system_id": system,
                "copper_status": "unspecified",
                "method": "author_curated_pathway_topology",
                "protocol": "textual reaction network",
                "point_count": len(path_ids),
                "surface_minima_count": 0,
                "surface_maxima_count": 0,
                "esp_area_bin_count": 0,
                "path_range": f"{min(path_ids)}-{max(path_ids)}",
                "independent_weight_unit": system,
                "direct_target_weight_ceiling": 0.0,
                "future_weight_ceiling": 0.05,
                "qc_flags": "article_claims_45_paths;workbook_has_43;identity_typo_risk",
                "notes": "仅作路径拓扑表征；不作为连续物性直接监督。",
                **common,
            }
        )
    _require(all_path_ids == list(range(1, 44)), "Figure4 路径编号漂移")
    _require(topology["E37"].value == 22, "Figure4 E37 异常数值漂移")

    energy = workbook["Figure5"]
    energy_specs = (
        ("S1_pathway_A", 2, 14, 2, "S1"),
        ("S1_pathway_B", 2, 14, 5, "S1"),
        ("S1_pathway_C", 2, 14, 8, "S1"),
        ("M1_pathway_A", 21, 32, 2, "M1"),
        ("M1_pathway_B", 21, 32, 5, "M1"),
        ("M3_workbook_or_M1_article", 21, 35, 8, "M3/M1_conflict"),
    )
    for observation_id, start, end, y_col, system in energy_specs:
        point_count = sum(_number(energy.cell(row, y_col).value) is not None for row in range(start, end + 1))
        rows.append(
            {
                "observation_id": f"fig5_{observation_id.lower()}",
                "sheet": "Figure5",
                "observation_kind": "reaction_energy_curve",
                "system_id": system,
                "copper_status": "unspecified",
                "method": "DFT",
                "protocol": "Gaussian09;M06-2X;basis_set_conflict;298.15K;1atm",
                "point_count": point_count,
                "surface_minima_count": 0,
                "surface_maxima_count": 0,
                "esp_area_bin_count": 0,
                "path_range": "not_labeled_in_workbook",
                "independent_weight_unit": observation_id,
                "direct_target_weight_ceiling": 0.15,
                "future_weight_ceiling": 0.15,
                "qc_flags": "missing_path_id;basis_set_conflict" + (";system_identity_conflict" if "conflict" in system else ""),
                "notes": "反应坐标点属于同一条计算曲线，不是独立材料样本。",
                **common,
            }
        )

    esp = workbook["Figure 6 and Figure 7"]
    esp_specs = (
        ("EPU", "Cu-free", (4, 18), (22, 36), (4, 18)),
        ("EPU+Cu", "Cu-containing", (4, 19), (24, 38), (4, 18)),
        ("S1", "Cu-free", (48, 51), (60, 62), (48, 62)),
        ("S1+Cu", "Cu-containing", (48, 51), (58, 62), (48, 62)),
        ("M1", "Cu-free", (71, 77), (82, 89), (71, 85)),
        ("M1+Cu", "Cu-containing", (71, 75), (79, 85), (70, 84)),
    )
    for index, (system, copper, minima_rows, maxima_rows, area_rows) in enumerate(esp_specs):
        left = index % 2 == 0
        extrema_col = 1 if left else 14
        area_col = 9 if left else (24 if system.startswith("S1") else 22)
        minima = sum(
            _number(esp.cell(row, extrema_col).value) is not None
            for row in range(minima_rows[0], minima_rows[1] + 1)
        )
        maxima = sum(
            _number(esp.cell(row, extrema_col).value) is not None
            for row in range(maxima_rows[0], maxima_rows[1] + 1)
        )
        bins = sum(
            _number(esp.cell(row, area_col).value) is not None
            for row in range(area_rows[0], area_rows[1] + 1)
        )
        rows.append(
            {
                "observation_id": f"fig6_7_{system.lower().replace('+', '_plus_')}",
                "sheet": "Figure 6 and Figure 7",
                "observation_kind": "esp_surface_and_area_distribution",
                "system_id": system,
                "copper_status": copper,
                "method": "DFT_ESP",
                "protocol": "Gaussian09;M06-2X;basis_set_conflict;single_point_uncertainty_0.05eV",
                "point_count": minima + maxima + bins,
                "surface_minima_count": minima,
                "surface_maxima_count": maxima,
                "esp_area_bin_count": bins,
                "path_range": "not_applicable",
                "independent_weight_unit": system,
                "direct_target_weight_ceiling": 0.15,
                "future_weight_ceiling": 0.15,
                "qc_flags": "missing_geometry_and_output;area_unit_unspecified;basis_set_conflict",
                "notes": "极值与分箱点按同一分子态聚合，不计为独立体系。",
                **common,
            }
        )
    _require(len(rows) == 14, "计算观测记录数漂移")
    _require(sum(int(row["point_count"]) for row in rows if row["observation_kind"] == "reaction_energy_curve") == 78, "Figure5 能量点计数漂移")
    _require(sum(int(row["surface_minima_count"]) + int(row["surface_maxima_count"]) for row in rows) == 104, "ESP 极值点计数漂移")
    _require(sum(int(row["esp_area_bin_count"]) for row in rows) == 90, "ESP 面积分箱计数漂移")
    return rows


def _scalar_and_qc_checks(workbook, curves: list[dict[str, object]]) -> tuple[dict[str, int], dict[str, object]]:
    activation = workbook["Figure 1e activation energy"]
    accumulated_ok = all(
        math.isclose(
            float(activation.cell(row, 8).value),
            sum(float(activation.cell(row, col).value) for col in range(2, 8)),
            rel_tol=0,
            abs_tol=1e-9,
        )
        for row in range(3, 11)
    )
    _require(accumulated_ok, "活化能 Accumulated 不再是六方法之和")

    eds = workbook["Figure2"]
    figure3 = workbook["Figure 3"]
    duplicate = all(
        figure3.cell(row, col).value == figure3.cell(row + 19, col).value
        for row in range(3, 9)
        for col in range(1, 9)
    )
    delta_reversed = all(
        math.isclose(
            float(figure3.cell(row, col + 7).value),
            float(figure3.cell(row, col).value)
            - float(figure3.cell(row - 9, col + 1).value),
            rel_tol=0,
            abs_tol=0.011,
        )
        for row in range(12, 18)
        for col in range(1, 8)
    )
    _require(duplicate, "Figure3 重复展示块不再完全相同")
    _require(delta_reversed, "Figure3 差值方向或数值漂移")
    ftir_curves = [row for row in curves if row["response_type"] == "FTIR_intensity"]
    counts = {
        "activation_energy_target_scalars": 8 * 6,
        "talpha_unique_target_scalars": 9 * 5,
        "thermodynamic_target_scalars": 8 * 3,
        "activation_candidate_scalar_count": 8 * 6 + 9 * 5 + 8 * 3,
        "activation_quality_r2_count": 8 * 5,
        "activation_derived_sum_count": 8,
        "eds_scalar_observed": 2 * 4 * 3,
        "eds_scalar_candidate": 2 * 4 * 2,
        "py_gc_ms_scalar_observed": 42 + 42 + 42 + 42 + 30 + 72,
        "py_gc_ms_scalar_candidate": 42 + 42 + 30 + 72,
        "figure3_duplicate_display_scalars": 42,
        "figure3_derived_delta_scalars": 42,
    }
    qc = {
        "figure1_20c_protocol_conflict": True,
        "figure1_temperature_exceeds_reported_950c": max(row["_axis_max"] for row in curves if row["_axis_max"] is not None) > 950,
        "figure1_replicates_reported_but_not_resolved": True,
        "figure2_last_element_label_should_be_n": eds["A6"].value == "C" and eds["G6"].value == "C",
        "figure2_after_carbon_atomic_percent_conflict": float(eds["J3"].value) == 94.32,
        "figure3_delta_header_sign_reversed": delta_reversed,
        "figure3_exact_duplicate_display_block": duplicate,
        "figure4_missing_path_ids": [44, 45],
        "figure4_stray_numeric_22_in_step3": workbook["Figure4"]["E37"].value == 22,
        "figure5_identity_or_caption_conflict": workbook["Figure5"]["G20"].value == "M3",
        "dft_basis_set_text_conflict": True,
        "supplementary_ftir_axis_is_rounded_and_repeated": all(int(row["unique_axis_values"]) == 900 for row in ftir_curves),
    }
    return counts, qc


def _tsv(rows: list[dict[str, object]], columns: tuple[str, ...]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=columns, delimiter="\t", lineterminator="\n", extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in columns})
    return stream.getvalue().encode("utf-8")


def _json(payload: object) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")


def _atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_symlink():
        raise AuditBlocked(f"拒绝覆盖符号链接：{path}")
    fd, temporary = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def run_audit(
    *,
    source_xlsx: Path | None = None,
    output_dir: Path | None = None,
    write_outputs: bool = False,
) -> dict[str, object]:
    """校验原件并返回结构化审计；默认不落盘。"""

    source = Path(source_xlsx) if source_xlsx is not None else SOURCE_XLSX
    destination = Path(output_dir) if output_dir is not None else source.parent
    file_row, archive_summary = _verify_xlsx(source)
    workbook = load_workbook(source, read_only=False, data_only=False, keep_links=False)
    try:
        inventory, cell_totals = _scan_workbook(workbook)
        curves = _experimental_curves(workbook)
        computations = _compute_observations(workbook)
        scalar_counts, qc = _scalar_and_qc_checks(workbook, curves)
    finally:
        workbook.close()

    public_curves = [{key: value for key, value in row.items() if not key.startswith("_")} for row in curves]
    counts = {
        "sheet_count": len(EXPECTED_SHEETS),
        "nonempty_cell_count": cell_totals["nonempty"],
        "native_numeric_cell_count": cell_totals["native_numeric"],
        "string_cell_count": cell_totals["strings"],
        "numeric_text_cell_count": cell_totals["numeric_text"],
        "formula_cell_count": cell_totals["formulas"],
        "error_cell_count": cell_totals["errors"],
        "base_material_count": 1,
        "independent_formulation_count": 0,
        "independent_specimens": 0,
        "experimental_curve_records": len(curves),
        "experimental_curve_independent_conditions": len({row["independent_condition_id"] for row in curves}),
        "experimental_response_points_observed": sum(int(row["observed_response_points"]) for row in curves),
        "experimental_response_points_paired": sum(int(row["paired_response_points"]) for row in curves),
        **scalar_counts,
        "computational_observation_records": len(computations),
        "computational_system_count": 6,
        "dft_topology_family_count": 2,
        "dft_path_id_count_workbook": 43,
        "dft_path_id_count_article": 45,
        "dft_pathway_family_count": 6,
        "dft_energy_curve_count": 6,
        "dft_energy_point_count": sum(int(row["point_count"]) for row in computations if row["observation_kind"] == "reaction_energy_curve"),
        "esp_surface_extrema_point_count": sum(int(row["surface_minima_count"]) + int(row["surface_maxima_count"]) for row in computations),
        "esp_area_distribution_count": 6,
        "esp_area_bin_point_count": sum(int(row["esp_area_bin_count"]) for row in computations),
    }
    summary = {
        "audit_version": AUDIT_VERSION,
        "source": {
            "title": "Reversible copper coordination redirects pyrolysis products in waste polyurethane enamelled copper wire",
            "doi": DOI,
            "concept_doi": CONCEPT_DOI,
            "url": SOURCE_URL,
            "license": LICENSE,
            "published": "2026-01-29",
            "primary_article_doi": ARTICLE_DOI,
            "source_scope": "one commercial polyurethane enamelled copper-wire system; formulation unresolved",
        },
        "file": file_row,
        "archive_safety": archive_summary,
        "workbook_inventory": inventory,
        "counts": counts,
        "qc": qc,
        "scientific_classification": {
            "gold_layers": ["Gold-E", "Gold-C"],
            "gold_admission_status": "admitted_reference",
            "intended_use": "mechanism_transfer_and_multifidelity_reference",
            "direct_tpu_mechanics_supervision": False,
            "source_weight_ceiling": 0.25,
            "mechanics_weight_ceiling": 0.0,
            "reason": "来源可靠且含实验与DFT多尺度观测，但只有一个未知配方的商业PU漆包线体系。",
        },
        "sample_semantics": {
            "independent_material_samples": 0,
            "experimental_curve_channels": len(curves),
            "computational_physical_systems": 6,
            "rule": "点是曲线或物理体系内部观测，不能作为独立材料样本。",
        },
    }
    outputs = {
        "内容审计摘要.json": _json(summary),
        "实验曲线审计清单.tsv": _tsv(public_curves, CURVE_COLUMNS),
        "计算观测清单.tsv": _tsv(computations, COMPUTE_COLUMNS),
        "文件校验清单.tsv": _tsv([file_row], ("file", "bytes", "md5", "sha256", "verification", "license", "doi")),
    }
    if write_outputs:
        for name, payload in outputs.items():
            _atomic_write(destination / name, payload)
    return {
        "summary": summary,
        "experimental_curves": public_curves,
        "computational_observations": computations,
        "files": [file_row],
        "outputs": outputs,
    }


if __name__ == "__main__":
    run_audit(write_outputs=True)
