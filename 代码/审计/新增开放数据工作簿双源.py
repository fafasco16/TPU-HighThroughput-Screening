"""只读复算 Mendeley SLS 与 Figshare 热固 PU 工作簿来源审计。

覆盖来源：

* ``Mendeley_SLS_TPU工艺力学``：155 个 XLSX 由 openpyxl 只读解析，
  59 个旧版 XLS 交给同目录 PowerShell/Excel COM 辅助脚本以 ReadOnly
  模式读取；原始工作簿从不保存、转换或改写。
* ``Figshare_热固PU原子经济升级回收``：5 个 XLSX 先核验官方大小、
  MD5 与 ZIP CRC，再用 ``zipfile.ZipFile`` 和 XML ``iterparse`` 逐行审计，
  包括 31.87 MB 补充工作簿，绝不整体载入工作表 XML。

脚本只允许原子覆盖 ``OUTPUT_WHITELIST`` 中的三个既有审计产物。JSON
与 TSV 均使用固定排序/列序和 LF；审计日期是协议常量，不读取当前时间。
所有科学输入在解析前后复核 SHA-256，路径链出现符号链接、junction 或其他
Windows 重解析点时失败关闭。
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import os
import posixpath
import re
import stat
import subprocess
import sys
import tempfile
import time
import tracemalloc
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from typing import Any, BinaryIO, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"
SLS_NAME = "Mendeley_SLS_TPU工艺力学"
FIGSHARE_NAME = "Figshare_热固PU原子经济升级回收"
SLS_ROOT = DATA_ROOT / SLS_NAME
FIGSHARE_ROOT = DATA_ROOT / FIGSHARE_NAME
SLS_SUMMARY = SLS_ROOT / "内容审计摘要.json"
SLS_MANIFEST_OUTPUT = SLS_ROOT / "文件取舍与校验清单.tsv"
FIGSHARE_SUMMARY = FIGSHARE_ROOT / "数据审计摘要.json"
SLS_OFFICIAL_MANIFEST = SLS_ROOT / "官方完整文件清单.json"
SLS_OFFICIAL_METADATA = SLS_ROOT / "官方元数据.json"
SLS_TABLES = SLS_ROOT / "结构化表格"
SLS_REPORTS = SLS_ROOT / "说明报告"
FIGSHARE_METADATA = FIGSHARE_ROOT / "官方API元数据.json"
XLS_READER = Path(__file__).with_name("读取SLS旧版XLS.ps1")

OUTPUT_WHITELIST = frozenset(
    {SLS_SUMMARY, SLS_MANIFEST_OUTPUT, FIGSHARE_SUMMARY}
)

SLS_GOLD_FILES = frozenset(
    {
        "TPU_1_edge_v3.xls",
        "T_TPU_1.xls",
        "TPU_1_edge_v2.xls",
        "T_TPU_2.xls",
        "T_TPU_4.xls",
        "TPU_1_1.xls",
        "T_TPU_3.xls",
        "T_TPU_5.xls",
        "TPU_4_1.xls",
        "TPU_2_1.xls",
        "TPU_10sem5.xls",
        "TPU_1_edge.xls",
        "TPU_6.xls",
        "TPU_3_1.xls",
        "TPU_5_1.xls",
        "TPU_11.xls",
        "TPU_6_05.xls",
        "TPU_12.xls",
        "TPU_13.xls",
        "TPU_1 (45°).xlsx",
        "TPU_1 (on edge).xlsx",
        "TPU_1 (upright).xlsx",
        "TPU_3 (45°).xlsx",
        "TPU_3 (on edge).xlsx",
        "TPU_3 (upright).xlsx",
        "edge_v3 (45°).xlsx",
        "edge_v3 (on edge).xlsx",
        "edge_v3 (upright).xlsx",
        "double_contour_v3 (45°).xlsx",
        "double_contour_v3 (on edge).xlsx",
        "double_contour_v3 (upright).xlsx",
    }
)

SLS_EXPECTED = {
    "official_files": 338,
    "official_bytes": 264_634_356,
    "local_files": 232,
    "local_bytes": 128_018_356,
    "xlsx": 155,
    "xls": 59,
    "pdf": 18,
    "workbooks": 214,
    "sheets": 1_065,
    "nonempty": 6_074_843,
    "finite": 5_971_548,
    "formulas": 15_909,
    "sequences": 75,
    "specimens": 350,
    "points": 1_787_452,
    "gold_sequences": 31,
    "gold_specimens": 140,
    "silver_sequences": 44,
    "silver_specimens": 210,
}

FIGSHARE_EXPECTED = {
    "workbooks": 5,
    "sheets": 28,
    "cells": 5_423_399,
    "finite": 5_423_004,
    "text": 264,
    "blank": 131,
    "formulas": 0,
    "errors": 0,
}

TSV_COLUMNS = [
    "序号",
    "文件ID",
    "文件名",
    "扩展名",
    "内容类型",
    "来源URL",
    "许可",
    "许可URL",
    "官方字节数",
    "取舍",
    "本地分层",
    "本地相对路径",
    "实际字节数",
    "官方SHA256",
    "实际SHA256",
    "校验状态",
    "未下载原因",
]

OOXML_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
PACKAGE_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"


class AuditBlocked(RuntimeError):
    """输入、完整性或科学语义不满足固定审计协议。"""


def assert_output_allowed(path: Path) -> None:
    """验证输出是白名单内普通路径；本定义刻意保持可独立抽取测试。"""

    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")

    def is_reparse(candidate: Path) -> bool:
        try:
            details = candidate.lstat()
        except FileNotFoundError:
            return False
        attributes = getattr(details, "st_file_attributes", 0)
        is_junction = getattr(candidate, "is_junction", lambda: False)
        return (
            candidate.is_symlink()
            or bool(attributes & 0x400)
            or bool(is_junction())
        )

    parent = path.parent
    try:
        resolved_parent = parent.resolve(strict=True)
    except FileNotFoundError as exc:
        raise AuditBlocked(f"审计输出目录不存在：{parent}") from exc
    if (
        not parent.is_dir()
        or is_reparse(parent)
        or os.path.normcase(os.path.abspath(str(resolved_parent)))
        != os.path.normcase(os.path.abspath(str(parent)))
    ):
        raise AuditBlocked(f"拒绝通过符号链接或重解析目录写入：{parent}")

    if path.exists() or path.is_symlink():
        if is_reparse(path):
            raise AuditBlocked(f"拒绝覆盖符号链接或重解析输出：{path}")
        try:
            resolved = path.resolve(strict=True)
        except FileNotFoundError as exc:
            raise AuditBlocked(f"审计输出链接目标缺失：{path}") from exc
        if (
            not path.is_file()
            or os.path.normcase(os.path.abspath(str(resolved)))
            != os.path.normcase(os.path.abspath(str(path)))
        ):
            raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    """在目标同目录 flush+fsync 后，以 os.replace 原子覆盖白名单输出。"""

    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        details = temporary.lstat()
        attributes = getattr(details, "st_file_attributes", 0)
        is_junction = getattr(temporary, "is_junction", lambda: False)
        if (
            temporary.is_symlink()
            or bool(attributes & 0x400)
            or bool(is_junction())
            or not temporary.is_file()
            or os.path.normcase(os.path.abspath(str(temporary.resolve(strict=True))))
            != os.path.normcase(os.path.abspath(str(temporary)))
        ):
            raise AuditBlocked(f"审计临时输出不是普通文件：{temporary}")
        assert_output_allowed(path)
        try:
            os.replace(temporary, path)
        except PermissionError:
            # Windows 的下载文件可能带 ReadOnly 属性；只清理白名单目标属性。
            os.chmod(path, 0o666)
            assert_output_allowed(path)
            os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def _is_reparse(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & flag) or bool(is_junction())


def _assert_plain_chain(path: Path, stop: Path = PROJECT_ROOT) -> None:
    absolute = path.absolute()
    stop_absolute = stop.absolute()
    if absolute != stop_absolute and stop_absolute not in absolute.parents:
        raise AuditBlocked(f"路径越出项目根目录：{path}")
    cursor = absolute
    while True:
        if _is_reparse(cursor):
            raise AuditBlocked(f"拒绝符号链接、junction 或重解析点：{cursor}")
        if cursor == stop_absolute:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"目录缺失、不是普通目录或经链接解析：{path}")
    _assert_plain_chain(path)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    _assert_plain_chain(path)


def hash_stream(handle: BinaryIO, algorithm: str = "sha256") -> str:
    digest = hashlib.new(algorithm)
    for block in iter(lambda: handle.read(1024 * 1024), b""):
        digest.update(block)
    return digest.hexdigest()


def file_hash(path: Path, algorithm: str = "sha256") -> str:
    require_file(path)
    with path.open("rb") as handle:
        return hash_stream(handle, algorithm)


def read_json(path: Path) -> Any:
    require_file(path)
    return json.loads(path.read_text(encoding="utf-8"))


def render_json(payload: dict[str, Any]) -> bytes:
    return (
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
            allow_nan=False,
        )
        + "\n"
    ).encode("utf-8")


def render_tsv(rows: list[dict[str, Any]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=TSV_COLUMNS,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def finite_number(value: Any) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    converted = float(value)
    return converted if math.isfinite(converted) else None


def stable_float(value: float) -> str:
    return format(float(value), ".17g")


def scientific_input_paths() -> list[Path]:
    paths = [SLS_OFFICIAL_MANIFEST, SLS_OFFICIAL_METADATA, FIGSHARE_METADATA]
    paths.extend(sorted(SLS_TABLES.glob("*"), key=lambda item: item.name.casefold()))
    paths.extend(sorted(SLS_REPORTS.glob("*"), key=lambda item: item.name.casefold()))
    paths.extend(
        sorted(FIGSHARE_ROOT.glob("*.xlsx"), key=lambda item: item.name.casefold())
    )
    for path in paths:
        require_file(path)
    return paths


def input_hashes(paths: Iterable[Path]) -> dict[str, str]:
    return {
        path.relative_to(PROJECT_ROOT).as_posix(): file_hash(path)
        for path in sorted(paths, key=lambda item: item.as_posix().casefold())
    }


def classify_missing(filename: str) -> tuple[str, str]:
    lower = filename.lower()
    if filename.startswith("~$"):
        return "Office锁文件", "Office临时锁文件，非数据"
    suffix = Path(filename).suffix.lower()
    if suffix == ".jpg":
        return "JPG", "低分辨率预览图，表格/PDF已保留可审计内容"
    mapping = {
        ".zs2": ("ZS2", "专有仪器二进制格式，当前固定解析器不可用"),
        ".da2": ("DA2", "专有仪器二进制格式，当前固定解析器不可用"),
        ".zp2": ("ZP2", "专有仪器二进制格式，当前固定解析器不可用"),
        ".0": ("点0文件", "专有仪器二进制分片，当前固定解析器不可用"),
        ".1": ("点1文件", "专有仪器二进制分片，当前固定解析器不可用"),
    }
    if suffix in mapping:
        return mapping[suffix]
    raise AuditBlocked(f"未登记的 SLS 未下载文件类型：{filename} ({lower})")


def build_sls_manifest() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    manifest = read_json(SLS_OFFICIAL_MANIFEST)
    if not isinstance(manifest, list) or len(manifest) != SLS_EXPECTED["official_files"]:
        raise AuditBlocked("SLS 官方文件清单不是预期的 338 条记录")

    local_by_name: dict[str, tuple[str, Path]] = {}
    for layer, directory in (("结构化表格", SLS_TABLES), ("说明报告", SLS_REPORTS)):
        require_directory(directory)
        for path in sorted(directory.iterdir(), key=lambda item: item.name.casefold()):
            require_file(path)
            if path.name in local_by_name:
                raise AuditBlocked(f"SLS 本地文件名重复：{path.name}")
            local_by_name[path.name] = (layer, path)

    if len(local_by_name) != SLS_EXPECTED["local_files"]:
        raise AuditBlocked(f"SLS 本地文件数异常：{len(local_by_name)}")

    official_names = {str(item["filename"]) for item in manifest}
    unexpected = sorted(set(local_by_name) - official_names)
    if unexpected:
        raise AuditBlocked(f"SLS 存在官方清单外文件：{unexpected}")

    rows: list[dict[str, Any]] = []
    local_counts: Counter[str] = Counter()
    missing_counts: Counter[str] = Counter()
    local_total = 0
    official_total = 0
    for index, item in enumerate(manifest, 1):
        name = str(item["filename"])
        details = item["content_details"]
        official_size = int(details["size"])
        official_sha = str(details["sha256_hash"]).lower()
        official_total += official_size
        suffix = Path(name).suffix.lower()
        if name in local_by_name:
            layer, path = local_by_name[name]
            actual_size = path.stat().st_size
            actual_sha = file_hash(path)
            if actual_size != official_size or actual_sha != official_sha:
                raise AuditBlocked(f"SLS 官方哈希/大小不一致：{name}")
            if suffix == ".xlsx":
                category = "有效XLSX"
            elif suffix == ".xls":
                category = "XLS"
            elif suffix == ".pdf":
                category = "PDF"
            else:
                raise AuditBlocked(f"SLS 已下载文件类型未登记：{name}")
            local_counts[category] += 1
            local_total += actual_size
            decision = "已下载"
            relative = f"{layer}/{name}"
            reason = ""
            verification = "通过"
        else:
            category, reason = classify_missing(name)
            missing_counts[category] += 1
            layer = ""
            relative = ""
            actual_size = ""
            actual_sha = ""
            decision = "未下载"
            verification = "不适用"
        rows.append(
            {
                "序号": index,
                "文件ID": item["id"],
                "文件名": name,
                "扩展名": suffix,
                "内容类型": details["content_type"],
                "来源URL": details["download_url"],
                "许可": "CC BY 4.0",
                "许可URL": "https://creativecommons.org/licenses/by/4.0/",
                "官方字节数": official_size,
                "取舍": decision,
                "本地分层": layer,
                "本地相对路径": relative,
                "实际字节数": actual_size,
                "官方SHA256": official_sha,
                "实际SHA256": actual_sha,
                "校验状态": verification,
                "未下载原因": reason,
            }
        )

    expected_local = Counter({"有效XLSX": 155, "XLS": 59, "PDF": 18})
    expected_missing = Counter(
        {
            "ZS2": 78,
            "JPG": 9,
            "DA2": 4,
            "ZP2": 3,
            "点0文件": 9,
            "点1文件": 1,
            "Office锁文件": 2,
        }
    )
    if local_counts != expected_local or missing_counts != expected_missing:
        raise AuditBlocked(
            f"SLS 文件分类异常：local={local_counts}, missing={missing_counts}"
        )
    if official_total != SLS_EXPECTED["official_bytes"]:
        raise AuditBlocked(f"SLS 官方总字节异常：{official_total}")
    if local_total != SLS_EXPECTED["local_bytes"]:
        raise AuditBlocked(f"SLS 本地总字节异常：{local_total}")
    return rows, {
        "local_counts": dict(local_counts),
        "missing_counts": dict(missing_counts),
        "official_total": official_total,
        "local_total": local_total,
    }


def workbook_byte_duplicate_groups(paths: Iterable[Path]) -> list[list[str]]:
    by_hash: defaultdict[str, list[str]] = defaultdict(list)
    for path in paths:
        by_hash[file_hash(path)].append(path.name)
    groups = [sorted(names, key=str.casefold) for names in by_hash.values() if len(names) > 1]
    groups.sort(key=lambda names: tuple(name.casefold() for name in names))
    return groups


def audit_sls_xlsx() -> dict[str, Any]:
    paths = sorted(SLS_TABLES.glob("*.xlsx"), key=lambda item: item.name.casefold())
    if len(paths) != SLS_EXPECTED["xlsx"]:
        raise AuditBlocked(f"SLS XLSX 数量异常：{len(paths)}")

    totals = Counter()
    sequences: list[dict[str, Any]] = []
    negative_modules: list[dict[str, Any]] = []
    for path in paths:
        require_file(path)
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise AuditBlocked(f"SLS XLSX 无法只读解析：{path.name}") from exc
        try:
            totals["workbooks"] += 1
            workbook_curves: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                totals["sheets"] += 1
                is_values = sheet.title == "Valeurs Série"
                triple_headers: list[str | None] | None = None
                sequence_hashers: list[Any] = []
                point_counts: list[int] = []
                if is_values:
                    triple_headers = [
                        sheet.cell(row=2, column=column).value
                        for column in range(1, sheet.max_column + 1)
                    ]
                    if (
                        len(triple_headers) % 3 == 0
                        and triple_headers
                        and all(
                            triple_headers[offset : offset + 3]
                            == ["Allongement", "Force standard", "Course standard"]
                            for offset in range(0, len(triple_headers), 3)
                        )
                    ):
                        sequence_hashers = [
                            hashlib.sha256() for _ in range(len(triple_headers) // 3)
                        ]
                        point_counts = [0] * len(sequence_hashers)

                for row_number, row in enumerate(sheet.iter_rows(), 1):
                    values = [cell.value for cell in row]
                    for cell, value in zip(row, values):
                        if value is not None and value != "":
                            totals["nonempty"] += 1
                            if finite_number(value) is not None:
                                totals["finite"] += 1
                        if cell.data_type == "f":
                            totals["formulas"] += 1

                    if sequence_hashers and row_number >= 4:
                        for specimen_index, digest in enumerate(sequence_hashers):
                            first = specimen_index * 3
                            x = finite_number(values[first]) if first < len(values) else None
                            y = (
                                finite_number(values[first + 1])
                                if first + 1 < len(values)
                                else None
                            )
                            if x is not None and y is not None:
                                digest.update(f"{stable_float(x)},{stable_float(y)}\n".encode())
                                point_counts[specimen_index] += 1

                    if sheet.title.startswith("Résultats") and row_number >= 3:
                        module = finite_number(values[5]) if len(values) >= 6 else None
                        if module is not None and module < 0:
                            negative_modules.append(
                                {
                                    "file": path.name,
                                    "sheet": sheet.title,
                                    "row": row_number,
                                    "module_e_mpa": module,
                                }
                            )

                if sequence_hashers:
                    for index, (digest, count) in enumerate(
                        zip(sequence_hashers, point_counts), 1
                    ):
                        if count <= 0:
                            raise AuditBlocked(f"SLS 曲线为空：{path.name}/specimen-{index}")
                        workbook_curves.append(
                            {
                                "specimen": index,
                                "point_count": count,
                                "curve_sha256": digest.hexdigest(),
                            }
                        )
            if workbook_curves:
                sequence_digest = hashlib.sha256()
                for curve in workbook_curves:
                    sequence_digest.update((curve["curve_sha256"] + "\n").encode())
                sequences.append(
                    {
                        "file": path.name,
                        "format": "xlsx",
                        "specimen_count": len(workbook_curves),
                        "curve_point_count": sum(
                            int(curve["point_count"]) for curve in workbook_curves
                        ),
                        "sequence_sha256": sequence_digest.hexdigest(),
                    }
                )
        finally:
            workbook.close()

    expected = Counter(
        {
            "workbooks": 155,
            "sheets": 617,
            "nonempty": 5_125_499,
            "finite": 5_027_038,
            "formulas": 15_909,
        }
    )
    if totals != expected:
        raise AuditBlocked(f"SLS XLSX 单元格复算异常：{totals} != {expected}")
    if len(sequences) != 18:
        raise AuditBlocked(f"SLS 完整三通道 XLSX 序列数异常：{len(sequences)}")
    if len(negative_modules) != 2 or {item["file"] for item in negative_modules} != {
        "TPU_2 (45°).xlsx"
    }:
        raise AuditBlocked(f"SLS 负模量隔离记录异常：{negative_modules}")
    return {
        "totals": dict(totals),
        "sequences": sequences,
        "negative_modules": negative_modules,
    }


def find_powershell() -> str:
    for executable in ("pwsh.exe", "powershell.exe"):
        try:
            probe = subprocess.run(
                [executable, "-NoProfile", "-Command", "$PSVersionTable.PSVersion.ToString()"],
                check=False,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=20,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue
        if probe.returncode == 0:
            return executable
    raise AuditBlocked("找不到可用 PowerShell，无法调用 SLS 旧版 XLS 只读辅助脚本")


def audit_sls_xls() -> dict[str, Any]:
    require_file(XLS_READER)
    command = [
        find_powershell(),
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(XLS_READER),
        "-目录",
        str(SLS_TABLES),
    ]
    try:
        completed = subprocess.run(
            command,
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=1_800,
        )
    except subprocess.TimeoutExpired as exc:
        raise AuditBlocked("SLS 旧版 XLS Excel COM 审计超时") from exc
    stderr = completed.stderr.decode("utf-8", errors="replace").strip()
    if completed.returncode != 0:
        raise AuditBlocked(
            f"SLS 旧版 XLS Excel COM 审计失败（{completed.returncode}）：{stderr}"
        )
    stdout = completed.stdout.decode("utf-8-sig", errors="strict").strip()
    try:
        records = json.loads(stdout)
    except json.JSONDecodeError as exc:
        raise AuditBlocked(f"SLS XLS 辅助脚本未返回合法 JSON：{stdout[:500]}") from exc
    if not isinstance(records, list) or len(records) != SLS_EXPECTED["xls"]:
        raise AuditBlocked(f"SLS XLS 解析记录数异常：{len(records) if isinstance(records, list) else type(records)}")
    if not all(record.get("read_only") is True for record in records):
        raise AuditBlocked("SLS XLS 存在非 ReadOnly 工作簿")
    totals = {
        "workbooks": len(records),
        "sheets": sum(int(record["sheet_count"]) for record in records),
        "nonempty": sum(int(record["nonempty_cells"]) for record in records),
        "finite": sum(int(record["finite_numeric_cells"]) for record in records),
        "formulas": sum(int(record["formula_cells"]) for record in records),
    }
    expected = {
        "workbooks": 59,
        "sheets": 448,
        "nonempty": 949_344,
        "finite": 944_510,
        "formulas": 0,
    }
    if totals != expected:
        raise AuditBlocked(f"SLS XLS 单元格复算异常：{totals} != {expected}")
    return {"totals": totals, "sequences": records, "stderr": stderr}


def preferred_sequence(records: list[dict[str, Any]]) -> dict[str, Any]:
    random_suffix = re.compile(r"-[A-Za-z0-9]{6}\.xls$", re.IGNORECASE)
    return sorted(
        records,
        key=lambda item: (
            bool(random_suffix.search(str(item["file"]))),
            str(item["file"]).casefold(),
        ),
    )[0]


def build_sls_summary() -> tuple[dict[str, Any], bytes]:
    manifest_rows, manifest_stats = build_sls_manifest()
    xlsx = audit_sls_xlsx()
    xls = audit_sls_xls()

    combined_totals = {
        key: int(xlsx["totals"][key]) + int(xls["totals"][key])
        for key in ("workbooks", "sheets", "nonempty", "finite", "formulas")
    }
    expected_combined = {
        "workbooks": SLS_EXPECTED["workbooks"],
        "sheets": SLS_EXPECTED["sheets"],
        "nonempty": SLS_EXPECTED["nonempty"],
        "finite": SLS_EXPECTED["finite"],
        "formulas": SLS_EXPECTED["formulas"],
    }
    if combined_totals != expected_combined:
        raise AuditBlocked(f"SLS 新旧工作簿合计异常：{combined_totals}")

    raw_sequences = list(xlsx["sequences"]) + list(xls["sequences"])
    by_sequence: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in raw_sequences:
        by_sequence[str(record["sequence_sha256"])].append(record)
    deduplicated = [preferred_sequence(group) for group in by_sequence.values()]
    deduplicated.sort(key=lambda item: str(item["file"]).casefold())
    duplicate_scientific = [
        sorted((str(item["file"]) for item in group), key=str.casefold)
        for group in by_sequence.values()
        if len(group) > 1
    ]
    duplicate_scientific.sort(key=lambda group: tuple(name.casefold() for name in group))

    sequence_count = len(deduplicated)
    specimen_count = sum(int(item["specimen_count"]) for item in deduplicated)
    point_count = sum(int(item["curve_point_count"]) for item in deduplicated)
    gold = [item for item in deduplicated if item["file"] in SLS_GOLD_FILES]
    silver = [item for item in deduplicated if item["file"] not in SLS_GOLD_FILES]
    counts = {
        "sequences": sequence_count,
        "specimens": specimen_count,
        "points": point_count,
        "gold_sequences": len(gold),
        "gold_specimens": sum(int(item["specimen_count"]) for item in gold),
        "silver_sequences": len(silver),
        "silver_specimens": sum(int(item["specimen_count"]) for item in silver),
    }
    expected_counts = {key: SLS_EXPECTED[key] for key in counts}
    if counts != expected_counts:
        raise AuditBlocked(f"SLS 科学计数异常：{counts} != {expected_counts}")
    if {str(item["file"]) for item in gold} != SLS_GOLD_FILES:
        missing = sorted(SLS_GOLD_FILES - {str(item["file"]) for item in gold})
        raise AuditBlocked(f"SLS 论文金标准映射缺失：{missing}")
    if len(duplicate_scientific) != 2:
        raise AuditBlocked(f"SLS 科学序列重复组异常：{duplicate_scientific}")

    all_workbooks = list(SLS_TABLES.glob("*.xlsx")) + list(SLS_TABLES.glob("*.xls"))
    byte_duplicates = workbook_byte_duplicate_groups(all_workbooks)
    if len(byte_duplicates) != 4 or sum(len(group) - 1 for group in byte_duplicates) != 4:
        raise AuditBlocked(f"SLS 工作簿字节重复组异常：{byte_duplicates}")

    summary: dict[str, Any] = {
        "schema_version": "open-data-audit-v2",
        "audited_at": AUDIT_DATE,
        "数据集DOI": "10.17632/wfsm6f9rbn.1",
        "许可": "CC BY 4.0",
        "官方文件数": 338,
        "官方总字节数": manifest_stats["official_total"],
        "已下载文件数": 232,
        "已下载总字节数": manifest_stats["local_total"],
        "未下载文件数": 106,
        "未下载总字节数": manifest_stats["official_total"] - manifest_stats["local_total"],
        "已下载分类": manifest_stats["local_counts"],
        "未下载分类": manifest_stats["missing_counts"],
        "工作簿解析": {
            "工作簿数": combined_totals["workbooks"],
            "解析失败数": 0,
            "工作表数": combined_totals["sheets"],
            "非空单元格数": combined_totals["nonempty"],
            "有限数值单元格数": combined_totals["finite"],
            "公式单元格数": combined_totals["formulas"],
            "XLSX只读": {
                "工作簿数": xlsx["totals"]["workbooks"],
                "工作表数": xlsx["totals"]["sheets"],
                "有限数值单元格数": xlsx["totals"]["finite"],
            },
            "XLS_Excel_COM只读": {
                "工作簿数": xls["totals"]["workbooks"],
                "工作表数": xls["totals"]["sheets"],
                "有限数值单元格数": xls["totals"]["finite"],
                "全部ReadOnly": True,
                "保存或转换": False,
            },
        },
        "精确内容重复": {
            "重复组数": len(byte_duplicates),
            "重复额外文件数": sum(len(group) - 1 for group in byte_duplicates),
            "文件组": byte_duplicates,
        },
        "科学序列重复": {
            "重复组数": len(duplicate_scientific),
            "文件组": duplicate_scientific,
            "规则": "同一工作簿内按曲线顺序组合点对SHA-256；重复序列仅保留无随机后缀文件",
        },
        "真实材料数": 1,
        "材料": "EOS TPU 1301粉末（D10=22 µm, D50=72 µm, D90=138 µm）",
        "拉伸数据去重后": {
            "独立试验序列": sequence_count,
            "独立试样": specimen_count,
            "完整应力-应变曲线": specimen_count,
            "曲线点对": point_count,
            "端点指标数_每试样": 5,
            "端点标量总数": specimen_count * 5,
        },
        "发表论文可映射金标准": {
            "工艺条件": len(gold),
            "试样": sum(int(item["specimen_count"]) for item in gold),
            "文件": sorted(SLS_GOLD_FILES, key=str.casefold),
            "说明": "论文Table 3的19项标准试样条件与Table 4的12项0.5 mm薄膜条件",
        },
        "探索性银标准": {
            "独立试验序列": len(silver),
            "试样": sum(int(item["specimen_count"]) for item in silver),
            "说明": "原始曲线与端点保留，但工艺代码/版本映射尚不完整，不应与金标准等权",
        },
        "隔离质控": [
            {
                "系列": "TPU_2 (45°).xlsx",
                "原因": "探索性系列出现负杨氏模量且变异极高，需回看原始仪器/拟合窗口",
                "负模量记录": xlsx["negative_modules"],
            }
        ],
        "过程维度": {
            "层厚_mm": 0.1,
            "粉床温度_C": 110,
            "冷却腔温度_C": 80,
            "扫描间距_mm": [0.2, 0.25],
            "扫描速度_mm_s": [2000, 2500, 3000],
            "激光功率_W": [12, 16, 18, 19, 22, 25],
            "体积能量密度_J_mm3": [0.16, 0.5],
            "取向": ["flat", "on-edge", "45°", "upright"],
            "轮廓策略": ["simple", "edge", "double contour"],
        },
        "审计协议": {
            "科学输入只读": True,
            "科学输入运行内哈希复核": True,
            "输出路径白名单": [
                SLS_SUMMARY.relative_to(PROJECT_ROOT).as_posix(),
                SLS_MANIFEST_OUTPUT.relative_to(PROJECT_ROOT).as_posix(),
            ],
            "训练拆分权重物化": False,
        },
        "统一准入权重策略": {
            "policy_authority": "multi-fidelity-admission-weight-v0.2.9",
            "source_weight_ceiling": 0.35,
            "exact_duplicate_or_derived_record_weight": 0.0,
            "note": "这里只记录准入上限；本审计不物化训练权重。",
        },
        "定位": "单一商业TPU的SLS工艺-结构-力学数据；适合工艺迁移、各向异性和完整曲线模型，不增加化学空间覆盖。",
    }
    return summary, render_tsv(manifest_rows)


def safe_zip_member(name: str) -> str:
    normalized = name.replace("\\", "/")
    pure = PurePosixPath(normalized)
    if pure.is_absolute() or ".." in pure.parts or normalized.startswith("/"):
        raise AuditBlocked(f"OOXML ZIP 含不安全成员路径：{name}")
    return normalized


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    member = "xl/sharedStrings.xml"
    if member not in archive.namelist():
        return []
    strings: list[str] = []
    with archive.open(member) as handle:
        for event, element in ElementTree.iterparse(handle, events=("end",)):
            if element.tag == OOXML_NS + "si":
                strings.append("".join(node.text or "" for node in element.iter(OOXML_NS + "t")))
                element.clear()
    return strings


def workbook_sheet_map(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    with archive.open("xl/workbook.xml") as handle:
        workbook_root = ElementTree.parse(handle).getroot()
    with archive.open("xl/_rels/workbook.xml.rels") as handle:
        relations_root = ElementTree.parse(handle).getroot()
    relation_targets = {
        relation.attrib["Id"]: relation.attrib["Target"]
        for relation in relations_root.findall(PACKAGE_REL_NS + "Relationship")
    }
    result: list[tuple[str, str]] = []
    for sheet in workbook_root.findall(".//" + OOXML_NS + "sheet"):
        relation_id = sheet.attrib[REL_NS + "id"]
        target = relation_targets.get(relation_id)
        if target is None:
            raise AuditBlocked(f"OOXML 工作表关系缺失：{relation_id}")
        if target.startswith("/"):
            member = target.lstrip("/")
        else:
            member = posixpath.normpath(posixpath.join("xl", target))
        result.append((sheet.attrib["name"], safe_zip_member(member)))
    return result


def column_number(cell_reference: str) -> int:
    match = re.match(r"^([A-Z]+)", cell_reference.upper())
    if match is None:
        raise AuditBlocked(f"非法 OOXML 单元格引用：{cell_reference}")
    number = 0
    for character in match.group(1):
        number = number * 26 + ord(character) - ord("A") + 1
    return number


def decode_ooxml_cell(
    cell: ElementTree.Element, shared_strings: list[str]
) -> tuple[str, Any]:
    cell_type = cell.attrib.get("t", "n")
    formula = cell.find(OOXML_NS + "f")
    if formula is not None:
        return "formula", formula.text or ""
    value_element = cell.find(OOXML_NS + "v")
    if cell_type == "inlineStr":
        text = "".join(node.text or "" for node in cell.iter(OOXML_NS + "t"))
        return "text", text
    if value_element is None or value_element.text is None:
        return "blank", None
    raw = value_element.text
    if cell_type == "s":
        try:
            return "text", shared_strings[int(raw)]
        except (ValueError, IndexError) as exc:
            raise AuditBlocked(f"OOXML shared string 索引非法：{raw}") from exc
    if cell_type in {"str", "b", "d"}:
        return "text", raw
    if cell_type == "e":
        return "error", raw
    try:
        number = float(raw)
    except ValueError:
        return "text", raw
    if math.isfinite(number):
        return "finite", number
    return "error", raw


def audit_ooxml_sheet(
    archive: zipfile.ZipFile,
    member: str,
    shared_strings: list[str],
    pair_starts: tuple[int, ...] = (),
) -> dict[str, Any]:
    counts: Counter[str] = Counter()
    captured: dict[str, Any] = {}
    pair_hashers = {start: hashlib.sha256() for start in pair_starts}
    pair_counts: Counter[int] = Counter()
    current_values: dict[int, float] = {}
    max_active_row_value_buffers = 0
    resident_rows = 0
    released_row_elements = 0
    element_stack: list[ElementTree.Element] = []
    with archive.open(member) as handle:
        for event, element in ElementTree.iterparse(handle, events=("start", "end")):
            if event == "start":
                element_stack.append(element)
                if element.tag == OOXML_NS + "row":
                    resident_rows += 1
                    max_active_row_value_buffers = max(
                        max_active_row_value_buffers, resident_rows
                    )
                continue
            if element.tag == OOXML_NS + "c":
                reference = element.attrib.get("r", "")
                column = column_number(reference)
                kind, value = decode_ooxml_cell(element, shared_strings)
                counts["cells"] += 1
                counts[kind] += 1
                if kind == "finite":
                    current_values[column] = float(value)
                match = re.search(r"(\d+)$", reference)
                row_number = int(match.group(1)) if match else 0
                if row_number <= 16 and column <= 24:
                    captured[reference] = value
                element.clear()
            elif element.tag == OOXML_NS + "row":
                for start, digest in pair_hashers.items():
                    if start in current_values and start + 1 in current_values:
                        digest.update(
                            (
                                stable_float(current_values[start])
                                + ","
                                + stable_float(current_values[start + 1])
                                + "\n"
                            ).encode()
                        )
                        pair_counts[start] += 1
                current_values.clear()
                if len(element_stack) < 2:
                    raise AuditBlocked(f"OOXML 行缺少父节点：{member}")
                parent = element_stack[-2]
                parent.remove(element)
                element.clear()
                resident_rows -= 1
                released_row_elements += 1
                if resident_rows < 0:
                    raise AuditBlocked(f"OOXML 行驻留计数失衡：{member}")
            if not element_stack or element_stack[-1] is not element:
                raise AuditBlocked(f"OOXML 解析栈失衡：{member}")
            element_stack.pop()
    if resident_rows != 0 or element_stack:
        raise AuditBlocked(f"OOXML 解析结束后仍有驻留节点：{member}")
    return {
        "cells_present": counts["cells"],
        "finite_numeric_cells": counts["finite"],
        "text_cells": counts["text"],
        "explicit_structural_blank_cells": counts["blank"],
        "formula_cells": counts["formula"],
        "error_cells": counts["error"],
        "captured": captured,
        "pair_hashes": {str(start): digest.hexdigest() for start, digest in pair_hashers.items()},
        "pair_counts": {str(start): pair_counts[start] for start in pair_starts},
        "max_active_row_value_buffers": max_active_row_value_buffers,
        "released_row_elements": released_row_elements,
        "completed_row_elements_retained_after_parse": 0,
    }


def audit_figshare_workbooks() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    metadata = read_json(FIGSHARE_METADATA)
    if not isinstance(metadata, dict):
        raise AuditBlocked("Figshare 官方 API 元数据根节点异常")
    official_files = metadata.get("files")
    if not isinstance(official_files, list) or len(official_files) != 5:
        raise AuditBlocked("Figshare 官方文件列表不是 5 个工作簿")

    results: list[dict[str, Any]] = []
    sheet_lookup: dict[str, dict[str, Any]] = {}
    for official in sorted(official_files, key=lambda item: str(item["name"]).casefold()):
        name = str(official["name"])
        path = FIGSHARE_ROOT / name
        require_file(path)
        if path.suffix.lower() != ".xlsx":
            raise AuditBlocked(f"Figshare 官方资产不是 XLSX：{name}")
        actual_size = path.stat().st_size
        actual_md5 = file_hash(path, "md5")
        if actual_size != int(official["size"]) or actual_md5 != official["supplied_md5"]:
            raise AuditBlocked(f"Figshare 文件大小/MD5 不一致：{name}")

        try:
            with zipfile.ZipFile(path, "r") as archive:
                bad_member = archive.testzip()
                if bad_member is not None:
                    raise AuditBlocked(f"Figshare OOXML CRC 失败：{name}/{bad_member}")
                names = archive.namelist()
                if len(names) != len(set(names)):
                    raise AuditBlocked(f"Figshare OOXML ZIP 成员名重复：{name}")
                for member_name in names:
                    safe_zip_member(member_name)
                shared_strings = read_shared_strings(archive)
                sheets = workbook_sheet_map(archive)
                workbook_sheets: list[dict[str, Any]] = []
                for sheet_name, member in sheets:
                    pair_starts: tuple[int, ...] = ()
                    if name == "Huang_Source_Date_Figure_4.xlsx" and sheet_name == "Figure 4c":
                        pair_starts = (2, 4, 6, 8, 10)
                    elif name == "Supplementary Data   Set.xlsx" and sheet_name == "SI-2":
                        pair_starts = (2, 5)
                    stats = audit_ooxml_sheet(
                        archive,
                        member,
                        shared_strings,
                        pair_starts,
                    )
                    stats["name"] = sheet_name
                    workbook_sheets.append(stats)
                    sheet_lookup[f"{name}|{sheet_name}"] = stats
        except zipfile.BadZipFile as exc:
            raise AuditBlocked(f"Figshare XLSX 不是有效 OOXML ZIP：{name}") from exc

        results.append(
            {
                "name": name,
                "figshare_file_id": int(official["id"]),
                "bytes": actual_size,
                "official_md5": str(official["supplied_md5"]),
                "local_sha256": file_hash(path),
                "sheet_count": len(workbook_sheets),
                "sheets": workbook_sheets,
            }
        )

    totals = Counter()
    for workbook in results:
        totals["workbooks"] += 1
        totals["sheets"] += int(workbook["sheet_count"])
        for sheet in workbook["sheets"]:
            totals["cells"] += int(sheet["cells_present"])
            totals["finite"] += int(sheet["finite_numeric_cells"])
            totals["text"] += int(sheet["text_cells"])
            totals["blank"] += int(sheet["explicit_structural_blank_cells"])
            totals["formulas"] += int(sheet["formula_cells"])
            totals["errors"] += int(sheet["error_cells"])
            if int(sheet["max_active_row_value_buffers"]) > 1:
                raise AuditBlocked("Figshare OOXML 同时物化超过一个行值缓冲区")
            if int(sheet["completed_row_elements_retained_after_parse"]) != 0:
                raise AuditBlocked("Figshare OOXML 解析结束后仍保留已完成行节点")
    if dict(totals) != FIGSHARE_EXPECTED:
        raise AuditBlocked(f"Figshare 工作簿复算异常：{dict(totals)} != {FIGSHARE_EXPECTED}")
    return results, sheet_lookup


def captured_values(sheet: dict[str, Any], references: Iterable[str]) -> list[Any]:
    captured = sheet["captured"]
    return [captured.get(reference) for reference in references]


def build_figshare_summary() -> dict[str, Any]:
    metadata = read_json(FIGSHARE_METADATA)
    workbooks, lookup = audit_figshare_workbooks()

    temperatures = captured_values(
        lookup["Huang_Source_Date_Figure_2.xlsx|Figure 2b"],
        ("B3", "B4", "B5", "B6", "B7"),
    )
    ratios = captured_values(
        lookup["Huang_Source_Date_Figure_2.xlsx|Figure 2c"],
        ("B3", "B4", "B5", "B6", "B7"),
    )
    formulation_headers = captured_values(
        lookup["Huang_Source_Date_Figure_3.xlsx|Figure 3g"],
        ("B2", "D2", "F2"),
    )
    substrates = captured_values(
        lookup["Huang_Source_Date_Figure_5.xlsx|Figure 5d"],
        ("B4", "B5", "B6", "B7"),
    )
    cycles = captured_values(
        lookup["Huang_Source_Date_Figure_5.xlsx|Figure 5e"],
        ("B4", "B5", "B6"),
    )
    if temperatures != [140.0, 150.0, 160.0, 170.0, 180.0]:
        raise AuditBlocked(f"Figshare 解构温度条件异常：{temperatures}")
    if ratios != [0.5, 0.75, 1.0, 2.0, 3.0]:
        raise AuditBlocked(f"Figshare EAA/PUF 条件异常：{ratios}")
    if formulation_headers != ["89% PUF Waste", "80% PUF Waste", "67% PUF Waste"]:
        raise AuditBlocked(f"Figshare 再生配方表头异常：{formulation_headers}")
    if substrates != ["Steel", "Aluminum", "Paper", "Teflon"]:
        raise AuditBlocked(f"Figshare 剥离基底异常：{substrates}")
    if cycles != [1.0, 2.0, 3.0]:
        raise AuditBlocked(f"Figshare 重粘接循环异常：{cycles}")

    fig4c = lookup["Huang_Source_Date_Figure_4.xlsx|Figure 4c"]
    si2 = lookup["Supplementary Data   Set.xlsx|SI-2"]
    if not (
        fig4c["pair_hashes"]["4"]
        == si2["pair_hashes"]["2"]
        == si2["pair_hashes"]["5"]
    ):
        raise AuditBlocked("Figshare original PUF FTIR 跨文件重复指纹不一致")
    if not (
        fig4c["pair_counts"]["4"]
        == si2["pair_counts"]["2"]
        == si2["pair_counts"]["5"]
        == 1765
    ):
        raise AuditBlocked("Figshare FTIR 重复曲线点数异常")

    totals = {
        "workbook_count": len(workbooks),
        "sheet_count": sum(item["sheet_count"] for item in workbooks),
        "cells_present": sum(
            sheet["cells_present"] for item in workbooks for sheet in item["sheets"]
        ),
        "finite_numeric_cells": sum(
            sheet["finite_numeric_cells"] for item in workbooks for sheet in item["sheets"]
        ),
        "text_cells": sum(
            sheet["text_cells"] for item in workbooks for sheet in item["sheets"]
        ),
        "explicit_structural_blank_cells": sum(
            sheet["explicit_structural_blank_cells"]
            for item in workbooks
            for sheet in item["sheets"]
        ),
        "formula_cells": sum(
            sheet["formula_cells"] for item in workbooks for sheet in item["sheets"]
        ),
        "error_cells": sum(
            sheet["error_cells"] for item in workbooks for sheet in item["sheets"]
        ),
    }

    contents = {
        "Huang_Source_Date_Figure_2.xlsx": "5个温度条件、5个EAA/PUF质量比条件、8组模型化合物NMR谱、3组键断裂动力学序列",
        "Huang_Source_Date_Figure_3.xlsx": "3个回收PUF含量配方的应力-应变、tan delta及89%配方循环拉伸轨迹",
        "Huang_Source_Date_Figure_4.xlsx": "EAA回收率汇总、初始/回收EAA的NMR、5组FTIR、3组TGA",
        "Huang_Source_Date_Figure_5.xlsx": "4种基材剥离强度汇总和3次重粘循环汇总",
        "Supplementary Data   Set.xlsx": "FTIR、GPC、模型化合物NMR与动力学、凝胶含量、DMA储能/损耗模量和应力松弛；共约46个曲线块及3个凝胶含量汇总条件",
    }
    workbook_summaries = []
    for item in workbooks:
        workbook_summaries.append(
            {
                "name": item["name"],
                "sheet_count": item["sheet_count"],
                "sheets": [sheet["name"] for sheet in item["sheets"]],
                "finite_numeric_cells": sum(
                    sheet["finite_numeric_cells"] for sheet in item["sheets"]
                ),
                "content": contents[item["name"]],
            }
        )

    integrity_files = [
        {
            "name": item["name"],
            "figshare_file_id": item["figshare_file_id"],
            "bytes": item["bytes"],
            "official_md5": item["official_md5"],
            "local_sha256": item["local_sha256"],
        }
        for item in workbooks
    ]
    integrity_files.append(
        {
            "name": FIGSHARE_METADATA.name,
            "bytes": FIGSHARE_METADATA.stat().st_size,
            "local_md5": file_hash(FIGSHARE_METADATA, "md5"),
            "local_sha256": file_hash(FIGSHARE_METADATA),
        }
    )

    return {
        "schema_version": "open-data-audit-v2",
        "audited_at": AUDIT_DATE,
        "dataset": {
            "title": metadata["title"],
            "platform": "Springer Nature figshare",
            "article_id": metadata["id"],
            "version": metadata["version"],
            "dataset_doi": metadata["doi"],
            "related_article_doi": metadata["resource_doi"],
            "official_api_url": metadata["url_public_api"],
            "official_html_url": metadata["url_public_html"],
            "published_date": metadata["published_date"],
            "status": metadata["status"],
            "download_disabled": metadata["download_disabled"],
            "license": metadata["license"]["name"],
            "license_url": metadata["license"]["url"],
            "authors_in_official_dataset_order": [
                author["full_name"] for author in metadata["authors"]
            ],
            "official_citation": metadata["citation"],
        },
        "download_integrity": {
            "official_data_file_count": len(workbooks),
            "official_data_total_bytes": sum(item["bytes"] for item in workbooks),
            "all_official_sizes_match": True,
            "all_official_md5_match": True,
            "all_xlsx_ooxml_crc_pass": True,
            "files": integrity_files,
        },
        "workbook_audit": {
            **totals,
            "note": "数值单元主要是NMR、FTIR、TGA、DMA和应力-应变曲线的采样点，不是独立材料样本。结构性空白来自多级表头，不应解释为缺失观测。",
            "streaming_ooxml": {
                "all_five_workbooks_streamed": True,
                "maximum_active_row_value_buffers": 1,
                "completed_row_elements_retained_after_parse": 0,
                "xml_parser_chunk_prefetch_note": "ElementTree可预读未完成XML节点；本门禁只声明行值缓冲为1，并保证已完成row从父树移除。",
                "supplementary_workbook_loaded_whole": False,
                "parser": "zipfile.ZipFile + xml.etree.ElementTree.iterparse",
            },
            "workbooks": workbook_summaries,
        },
        "semantic_units": {
            "material_system_count": 1,
            "material_system": "单一来源的工业级商品热固性聚氨酯泡沫，精确异氰酸酯/多元醇配方未公开；论文给出脲/氨基甲酸酯/缩二脲键摩尔比约3:1:0.2。",
            "deconstruction_screen_rows": len(temperatures) + len(ratios),
            "deconstruction_unique_process_conditions": len(temperatures) + len(ratios) - 1,
            "deconstruction_note": "5个温度条件与5个EAA/PUF质量比条件在180摄氏度、质量比1处重合；不能把该重合点当独立重复实验。",
            "model_compound_classes": 3,
            "model_compound_classes_list": ["biuret", "urethane", "urea"],
            "upcycled_formulation_count": len(formulation_headers),
            "upcycled_formulations_puf_waste_wt_percent": [67, 80, 89],
            "upcycled_formulations_mai_wt_percent_approx": [33, 20, 11],
            "eaa_recovery_parallel_batches": 3,
            "eaa_recovery_batch_id_status": "missing_from_public_workbook; aggregate mean and spread only",
            "peel_substrate_count": len(substrates),
            "peel_substrates": [str(value).lower() if value != "Teflon" else value for value in substrates],
            "rebonding_cycle_summary_count": len(cycles),
            "representative_curve_replicate_ids": "missing; do not invent specimen IDs",
            "important_warning": "文件数、工作表数、行数、数值点数和曲线采样点均不得作为材料样本量。核心独立化学体系仍只有1个热固PUF来源和3个再生配方。",
        },
        "cross_file_lineage": [
            "Figure 2d的8组NMR谱均与Supplementary Data中的SI-9、SI-11或SI-13曲线精确重复，不能重复计数。",
            "Figure 2e的3组conversion序列由SI-9、SI-11、SI-13中的residual content序列按100减去残余含量得到，是派生表。",
            "Figure 3h的温度轴与SI-19精确重复，tan delta来自SI-19的G''/G'关系，应标为派生曲线而非新增试样。",
            "Figure 4c中的original PUF FTIR曲线与SI-2两份展示曲线点级SHA-256一致（各1765点）；三者只能计作一条曲线。",
            "跨文件列指纹审计发现大量重复坐标轴；坐标轴重复本身不等于材料重复，但在点级拼表时必须按curve_id隔离。",
        ],
        "admission_decision": {
            "candidate_eligible_after_governance_materialization": True,
            "training_weight_materialized": False,
            "split_materialized": False,
            "split_group_key": "dataset_doi|feedstock",
            "tier": "auxiliary_transfer_only",
            "core_tpu_chemistry_training_eligible": False,
            "core_tpu_mechanics_training_eligible": False,
            "eligible_tasks": [
                "PU选择性断键与化学回收工艺知识",
                "工艺条件到解构时间的辅助建模",
                "回收物比例到光固化再生网络力学/DMA的跨域迁移",
                "循环恢复、粘接、FTIR、TGA、NMR和应力松弛的表征预训练",
                "论文可持续性与闭环利用路线设计",
            ],
            "domain_boundary": "研究对象是化学交联的热固性PU泡沫及其光固化再生网络，不是热塑性聚氨酯TPU；不存在可直接用于TPU候选单体筛选的完整单体结构、软硬段配方、分子量和热塑加工标签。",
            "recommended_weights_relative_to_one_verified_core_tpu_experimental_sample": {
                "policy_authority": "multi-fidelity-admission-weight-v0.2.9",
                "source_weight_ceiling": 0.25,
                "core_tpu_structure_property": 0.0,
                "exact_duplicate_or_derived_record_weight": 0.0,
                "auxiliary_process_summary_per_unique_condition": "0.20-0.25",
                "auxiliary_mechanical_or_dma_curve_total_per_curve": "0.15-0.25",
                "spectroscopy_curve_total_per_curve_for_general_structure_property_model": "0.03-0.10",
                "point_weight_rule": "每条曲线总权重归一为1，再按该曲线有限点数分摊；禁止每个采样点获得一个完整样本权重。",
            },
            "split_rule": "默认泄漏/拆分键固定为dataset_doi|feedstock；原始观测中的feedstock_id映射到策略字段feedstock。deconstruction_condition、formulation_id与curve_id仅作观测身份，曲线点必须同折且不得跨折。最保守做法是把整个数据集作为单一外部迁移/验证组。",
        },
        "quality_caveats": [
            "工业PUF的精确多元醇、异氰酸酯、催化剂及商业配方未公开，不能构造可靠的TPU单体到性能映射。",
            "主文报告的拉伸强度、断裂伸长率、模量和韧性含均值及不确定度，但公开XLSX主要提供代表性曲线，没有原始逐重复试样标识。",
            "Figure 5工作簿列名写standard deviation，而论文图注写s.e.m.且n不小于3；不确定度类型存在来源间冲突，入库前应标记uncertainty_type_ambiguous。",
            "大量SI工作表只写SI编号和坐标轴，具体化合物/试样身份需要借助论文及补充信息映射，机器可读性并非完美。",
            "所有XLSX均无公式和Excel错误值；这只说明数值为静态发布值，不代表不存在实验偏差或选择性报告。",
        ],
        "审计协议": {
            "科学输入只读": True,
            "科学输入运行内哈希复核": True,
            "输出路径白名单": [FIGSHARE_SUMMARY.relative_to(PROJECT_ROOT).as_posix()],
            "训练拆分权重物化": False,
        },
        "references": [
            "Yutian Zhu; Yan Huang; Siyi Ye; Yi Deng; Jianwen Chen; Zenghe Liu; Xiaoxiao Guo; Yanling Zhu. Atom-economy upcycling of commodity thermoset polyurethane into photocuring 3D printing resins based on selective cleavage—crosslink strategy [Data set]. figshare, Version 1, 2026. https://doi.org/10.6084/m9.figshare.31552786.v1",
            "Huang, Y.; Guo, X.; Deng, Y.; Ye, S.; Zhu, Y.; Liu, Z.; Chen, J.; Zhu, Y. Atom-economy upcycling of commodity thermoset polyurethane into photocuring 3D printing resins based on selective cleavage—crosslink strategy. Nature Communications 17, 4151 (2026). https://doi.org/10.1038/s41467-026-70951-w",
        ],
    }


def validate_rendered_outputs(
    sls_summary_payload: bytes,
    sls_manifest_payload: bytes,
    figshare_payload: bytes,
) -> None:
    sls = json.loads(sls_summary_payload.decode("utf-8"))
    figshare = json.loads(figshare_payload.decode("utf-8"))
    rows = list(csv.DictReader(io.StringIO(sls_manifest_payload.decode("utf-8")), delimiter="\t"))
    if len(rows) != 338 or list(rows[0]) != TSV_COLUMNS:
        raise AuditBlocked("SLS TSV 渲染后回读失败")
    if sls["拉伸数据去重后"]["曲线点对"] != 1_787_452:
        raise AuditBlocked("SLS JSON 渲染后回读失败")
    if figshare["workbook_audit"]["finite_numeric_cells"] != 5_423_004:
        raise AuditBlocked("Figshare JSON 渲染后回读失败")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check-only",
        action="store_true",
        help="完成全部只读解析与断言，但不覆盖白名单审计产物",
    )
    arguments = parser.parse_args(argv)

    start = time.perf_counter()
    tracemalloc.start()
    for directory in (PROJECT_ROOT, DATA_ROOT, SLS_ROOT, FIGSHARE_ROOT, SLS_TABLES, SLS_REPORTS):
        require_directory(directory)
    inputs = scientific_input_paths()
    before = input_hashes(inputs)

    sls_summary, sls_manifest_payload = build_sls_summary()
    figshare_summary = build_figshare_summary()
    sls_summary_payload = render_json(sls_summary)
    figshare_payload = render_json(figshare_summary)
    validate_rendered_outputs(
        sls_summary_payload,
        sls_manifest_payload,
        figshare_payload,
    )

    after = input_hashes(inputs)
    if before != after:
        changed = sorted(set(before) | set(after))
        changed = [name for name in changed if before.get(name) != after.get(name)]
        raise AuditBlocked(f"科学输入在审计期间发生变化：{changed}")

    if not arguments.check_only:
        atomic_write(SLS_SUMMARY, sls_summary_payload)
        atomic_write(SLS_MANIFEST_OUTPUT, sls_manifest_payload)
        atomic_write(FIGSHARE_SUMMARY, figshare_payload)
        # 对真实落盘结果做语法与行数回读，防止半截/编码错误产物被接受。
        validate_rendered_outputs(
            SLS_SUMMARY.read_bytes(),
            SLS_MANIFEST_OUTPUT.read_bytes(),
            FIGSHARE_SUMMARY.read_bytes(),
        )

    current, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    result = {
        "status": "ok",
        "check_only": arguments.check_only,
        "scientific_input_file_count": len(inputs),
        "scientific_inputs_unchanged": True,
        "sls": {
            "workbooks": sls_summary["工作簿解析"]["工作簿数"],
            "sequences": sls_summary["拉伸数据去重后"]["独立试验序列"],
            "specimens": sls_summary["拉伸数据去重后"]["独立试样"],
            "curve_points": sls_summary["拉伸数据去重后"]["曲线点对"],
        },
        "figshare": {
            "workbooks": figshare_summary["workbook_audit"]["workbook_count"],
            "sheets": figshare_summary["workbook_audit"]["sheet_count"],
            "finite_numeric_cells": figshare_summary["workbook_audit"]["finite_numeric_cells"],
            "maximum_active_row_value_buffers": 1,
            "completed_row_elements_retained_after_parse": 0,
        },
        "resource_observation": {
            "elapsed_seconds": round(time.perf_counter() - start, 3),
            "python_tracemalloc_current_bytes": current,
            "python_tracemalloc_peak_bytes": peak,
        },
    }
    print(json.dumps(result, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"审计阻断：{exc}", file=sys.stderr)
        raise SystemExit(2) from exc
