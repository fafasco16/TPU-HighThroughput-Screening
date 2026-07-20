"""只读深审四个已在本地、但尚未进入总账的 TPU/PU 力学数据源。

本脚本只读取固定字节的原始 ZIP、XLSX 与 CSV，不联网、不改写科学原件、
不创建训练划分，也不物化训练权重。输出仅为可复现的来源审计摘要和逐记录
TSV，用于 Gold-E 参考层的后续总账生成。

运行：

    python 代码/审计/既有力学数据精选审计.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import tempfile
import zipfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始/外部数据/力学曲线"
AUDIT_DATE = "2026-07-21"
AUDIT_VERSION = "1.0"

SELF_HEALING = "SelfHealingTPU_4TU"
SCHWARZ = "Schwarz2022_EPU40"
ZENODO_4156 = "Zenodo4156000"
ZENODO_1098 = "Zenodo1098206"
SOURCE_NAMES = (SELF_HEALING, SCHWARZ, ZENODO_4156, ZENODO_1098)

OUTPUT_NAMES = (
    "内容审计摘要.json",
    "文件校验清单.tsv",
    "曲线审计清单.tsv",
    "标量审计清单.tsv",
    "配方审计清单.tsv",
)
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source in SOURCE_NAMES
    for filename in OUTPUT_NAMES
)

SELF_HEALING_FILE = (
    "source_data.zip",
    546_535,
    "9d563b8389686530a1a73e62a0244c57a1c19b8a039b60ec63f0753b2ff034a8",
)
SCHWARZ_FILE = (
    "Raw_Data.xlsx",
    2_145_753,
    "2e782dd443b5f8b09eab3d5a4ebc78e7071d02b474f90048c1683c2eeb01c9f9",
)
ZENODO_1098_FILE = (
    "Supronics_Porous-TPU-Nanocomposites Dataset.xlsx",
    750_592,
    "11967acf1deec0ce05ad2d1e63b70738c4357345fae9dd708952a75630304bc8",
)
ZENODO_4156_MANIFEST_SHA256 = (
    "098a04ef06e3f3b7650267e71c50c807fe49b3a8f2721c5f3d1e861996b3890c"
)

FILE_COLUMNS = (
    "source_directory",
    "path",
    "role",
    "bytes",
    "sha256",
    "integrity",
    "parser_state",
    "license",
    "dedup_status",
    "duplicate_of",
    "training_split_materialized",
    "training_weight_materialized",
)
CURVE_COLUMNS = (
    "source_directory",
    "record_id",
    "source_file",
    "source_location",
    "formulation_id",
    "material_scope",
    "modality",
    "test_type",
    "point_count",
    "usable_point_count",
    "partial_point_count",
    "quality_status",
    "axis_fields",
    "unit_status",
    "sample_mapping_status",
    "instance_key",
    "split_group",
    "dedup_status",
    "duplicate_of",
    "decision",
    "future_weight_ceiling",
    "training_split",
    "training_weight",
    "notes",
)
SCALAR_COLUMNS = (
    "source_directory",
    "record_id",
    "source_file",
    "source_location",
    "formulation_id",
    "task_role",
    "material_state",
    "result_names",
    "direct_numeric_result_count",
    "derived_numeric_result_count",
    "candidate_numeric_result_count",
    "unit_status",
    "decision",
    "future_weight_ceiling",
    "split_group",
    "training_split",
    "training_weight",
    "notes",
)
FORMULATION_COLUMNS = (
    "source_directory",
    "formulation_id",
    "material_family",
    "component_1",
    "component_1_fraction",
    "component_2",
    "component_2_fraction",
    "fraction_basis",
    "identity_mapping_status",
    "evidence",
    "split_group",
    "future_weight_ceiling",
    "training_split",
    "training_weight",
    "notes",
)


class AuditBlocked(RuntimeError):
    """原件、解析结构或输出安全门禁失败。"""


@dataclass(frozen=True)
class AuditBundle:
    source_directory: str
    summary: dict[str, Any]
    files: list[dict[str, Any]]
    curves: list[dict[str, Any]]
    scalars: list[dict[str, Any]]
    formulations: list[dict[str, Any]]


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(flag and attributes & flag) or bool(is_junction())


def _require_plain_directory(path: Path) -> None:
    if not path.is_dir() or _is_reparse_point(path):
        raise AuditBlocked(f"来源目录不是普通目录：{path}")
    if path.resolve(strict=True) != path.absolute():
        raise AuditBlocked(f"来源目录解析发生漂移：{path}")


def _require_file(path: Path, expected_bytes: int, expected_sha256: str) -> None:
    if not path.is_file() or _is_reparse_point(path):
        raise AuditBlocked(f"原件不是普通文件：{path}")
    actual_bytes = path.stat().st_size
    actual_sha256 = _sha256_file(path)
    if actual_bytes != expected_bytes or actual_sha256 != expected_sha256:
        raise AuditBlocked(
            f"原件字节漂移：{path.name} bytes={actual_bytes} sha256={actual_sha256}"
        )


def _safe_zip_member(info: zipfile.ZipInfo) -> None:
    path = PurePosixPath(info.filename.replace("\\", "/"))
    if path.is_absolute() or ".." in path.parts or not path.parts:
        raise AuditBlocked(f"ZIP成员路径不安全：{info.filename!r}")
    unix_mode = info.external_attr >> 16
    if stat.S_ISLNK(unix_mode):
        raise AuditBlocked(f"ZIP成员为符号链接：{info.filename!r}")
    if info.file_size > 10_000_000:
        raise AuditBlocked(f"ZIP单成员过大：{info.filename!r}")
    if info.compress_size and info.file_size / info.compress_size > 250:
        raise AuditBlocked(f"ZIP压缩比异常：{info.filename!r}")


def _finite(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _sequence_sha256(points: Iterable[tuple[float, ...]]) -> str:
    digest = hashlib.sha256()
    for point in points:
        digest.update("\t".join(format(value, ".17g") for value in point).encode("ascii"))
        digest.update(b"\n")
    return digest.hexdigest()


def _record_id(prefix: str, *parts: str) -> str:
    raw = "|".join(parts).encode("utf-8")
    return f"{prefix}_{hashlib.sha256(raw).hexdigest()[:16]}"


def _read_semicolon(payload: bytes) -> list[list[str]]:
    text = payload.decode("cp1252")
    return list(csv.reader(io.StringIO(text), delimiter=";"))


def _numeric_points(
    rows: Iterable[list[Any]], columns: tuple[int, ...], start_row: int = 1
) -> tuple[list[tuple[float, ...]], int]:
    points: list[tuple[float, ...]] = []
    partial = 0
    for row_number, row in enumerate(rows):
        if row_number < start_row:
            continue
        values = [
            _finite(row[column]) if column < len(row) else None for column in columns
        ]
        if all(value is not None for value in values):
            points.append(tuple(value for value in values if value is not None))
        elif any(value is not None for value in values):
            partial += 1
    return points, partial


def _file_row(
    source: str,
    path: str,
    role: str,
    payload: bytes,
    license_name: str,
    *,
    dedup_status: str = "unique",
    duplicate_of: str = "",
) -> dict[str, Any]:
    return {
        "source_directory": source,
        "path": path,
        "role": role,
        "bytes": len(payload),
        "sha256": _sha256_bytes(payload),
        "integrity": "verified",
        "parser_state": "parsed",
        "license": license_name,
        "dedup_status": dedup_status,
        "duplicate_of": duplicate_of,
        "training_split_materialized": "false",
        "training_weight_materialized": "false",
    }


def _curve_row(
    source: str,
    record_id: str,
    source_file: str,
    source_location: str,
    formulation_id: str,
    material_scope: str,
    modality: str,
    test_type: str,
    points: list[tuple[float, ...]],
    partial: int,
    axis_fields: str,
    unit_status: str,
    instance_key: str,
    split_group: str,
    decision: str,
    ceiling: float,
    *,
    quality_status: str = "accepted_reference",
    dedup_status: str = "unique",
    duplicate_of: str = "",
    notes: str = "",
) -> dict[str, Any]:
    usable = len(points) if decision.endswith("candidate") else 0
    return {
        "source_directory": source,
        "record_id": record_id,
        "source_file": source_file,
        "source_location": source_location,
        "formulation_id": formulation_id,
        "material_scope": material_scope,
        "modality": modality,
        "test_type": test_type,
        "point_count": len(points),
        "usable_point_count": usable,
        "partial_point_count": partial,
        "quality_status": quality_status,
        "axis_fields": axis_fields,
        "unit_status": unit_status,
        "sample_mapping_status": "explicit_record_group",
        "instance_key": instance_key,
        "split_group": split_group,
        "dedup_status": dedup_status,
        "duplicate_of": duplicate_of,
        "decision": decision,
        "future_weight_ceiling": f"{ceiling:.2f}",
        "training_split": "false",
        "training_weight": "false",
        "notes": notes,
    }


def _scalar_row(
    source: str,
    record_id: str,
    source_file: str,
    source_location: str,
    formulation_id: str,
    task_role: str,
    material_state: str,
    result_names: str,
    direct_count: int,
    derived_count: int,
    candidate_count: int,
    unit_status: str,
    decision: str,
    ceiling: float,
    split_group: str,
    notes: str,
) -> dict[str, Any]:
    return {
        "source_directory": source,
        "record_id": record_id,
        "source_file": source_file,
        "source_location": source_location,
        "formulation_id": formulation_id,
        "task_role": task_role,
        "material_state": material_state,
        "result_names": result_names,
        "direct_numeric_result_count": direct_count,
        "derived_numeric_result_count": derived_count,
        "candidate_numeric_result_count": candidate_count,
        "unit_status": unit_status,
        "decision": decision,
        "future_weight_ceiling": f"{ceiling:.2f}",
        "split_group": split_group,
        "training_split": "false",
        "training_weight": "false",
        "notes": notes,
    }


def _formulation_row(
    source: str,
    formulation_id: str,
    material_family: str,
    component_1: str,
    fraction_1: Any,
    component_2: str,
    fraction_2: Any,
    fraction_basis: str,
    mapping: str,
    evidence: str,
    split_group: str,
    ceiling: float,
    notes: str,
) -> dict[str, Any]:
    return {
        "source_directory": source,
        "formulation_id": formulation_id,
        "material_family": material_family,
        "component_1": component_1,
        "component_1_fraction": fraction_1,
        "component_2": component_2,
        "component_2_fraction": fraction_2,
        "fraction_basis": fraction_basis,
        "identity_mapping_status": mapping,
        "evidence": evidence,
        "split_group": split_group,
        "future_weight_ceiling": f"{ceiling:.2f}",
        "training_split": "false",
        "training_weight": "false",
        "notes": notes,
    }


def audit_self_healing() -> AuditBundle:
    source_dir = DATA_ROOT / SELF_HEALING
    _require_plain_directory(source_dir)
    filename, size, digest = SELF_HEALING_FILE
    path = source_dir / filename
    _require_file(path, size, digest)
    archive_payload = path.read_bytes()
    files = [
        _file_row(
            SELF_HEALING,
            filename,
            "primary_archive",
            archive_payload,
            "CC-BY-4.0",
        )
    ]
    curves: list[dict[str, Any]] = []
    scalars: list[dict[str, Any]] = []
    formulation_groups = {
        "Ninjaflex": "doi:10.4121/13603775.v1|Ninjaflex",
        "SH-TPU": "doi:10.4121/13603775.v1|SH-TPU",
    }
    csv_count = 0
    txt_count = 0
    uncompressed_bytes = 0
    specimen_keys: set[str] = set()
    with zipfile.ZipFile(io.BytesIO(archive_payload)) as archive:
        infos = [info for info in archive.infolist() if not info.is_dir()]
        if len(infos) != 71:
            raise AuditBlocked(f"4TU ZIP成员数漂移：{len(infos)}")
        for info in infos:
            _safe_zip_member(info)
            payload = archive.read(info)
            if len(payload) != info.file_size:
                raise AuditBlocked(f"4TU ZIP成员读取不完整：{info.filename}")
            uncompressed_bytes += len(payload)
            suffix = PurePosixPath(info.filename).suffix.lower()
            csv_count += suffix == ".csv"
            txt_count += suffix == ".txt"
            files.append(
                _file_row(
                    SELF_HEALING,
                    f"{filename}!/{info.filename}",
                    "archive_member",
                    payload,
                    "CC-BY-4.0",
                )
            )
            if suffix != ".csv":
                continue
            rows = _read_semicolon(payload)
            inner = info.filename
            if "/Filament thickness/" in inner:
                labels = rows[0][1:5]
                values_by_label: list[list[float]] = [[] for _ in labels]
                for row in rows[2:]:
                    for index, raw in enumerate(row[1:5]):
                        value = _finite(raw)
                        if value is not None:
                            values_by_label[index].append(value)
                if [len(values) for values in values_by_label] != [8, 8, 8, 8]:
                    raise AuditBlocked("4TU丝材直径表结构漂移")
                for label, values in zip(labels, values_by_label):
                    formulation = "Ninjaflex" if label == "Ninjaflex" else "SH-TPU"
                    scalars.append(
                        _scalar_row(
                            SELF_HEALING,
                            _record_id("4tu_thickness", label),
                            filename,
                            inner,
                            formulation,
                            "filament_diameter_uniformity",
                            label,
                            "diameter_mm_at_8_positions",
                            len(values),
                            0,
                            len(values),
                            "explicit:mm",
                            "low_weight_candidate",
                            0.35,
                            formulation_groups[formulation],
                            "沿丝材每5 cm测量；同一丝材内8点不当作8个独立配方。",
                        )
                    )
                continue

            material = "Ninjaflex" if "/Ninjaflex/" in inner else "SH-TPU"
            group = formulation_groups[material]
            decision = "low_weight_candidate"
            ceiling = 0.30
            quality = "accepted_reference"
            dedup = "unique"
            duplicate_of = ""
            notes = ""
            if "/Mechanical testing/" in inner:
                modality = "mechanical"
                test_type = "instrumented_blade_cut_force_displacement"
                columns = (0, 1)
                axis = "displacement_mm;load_N"
                ceiling = 0.55 if material == "SH-TPU" else 0.35
                basename = PurePosixPath(inner).name
                instance = re.sub(r"_healed(?=\.csv$)", "", basename, flags=re.I)
                specimen_keys.add(f"{material}|{instance}")
                notes = (
                    "4×4×4 mm立方体，18°刀尖，10 mm/s，20°C；healed文件与其原始文件共享试样键。"
                )
            elif "/DSC/" in inner:
                modality = "thermal"
                test_type = "DSC_heat_flow_temperature"
                columns = (1, 2)
                axis = "temperature_C;normalized_heat_flow_W_g"
                instance = PurePosixPath(inner).stem
            elif "/FTIR/" in inner and inner.endswith("Normalised_data.csv"):
                rows_for_curve = rows
                for column, state in enumerate(rows[0][1:6], start=1):
                    points, partial = _numeric_points(rows_for_curve, (0, column), 1)
                    curves.append(
                        _curve_row(
                            SELF_HEALING,
                            _record_id("4tu_ftir_norm", inner, state),
                            filename,
                            f"{inner}#{state}",
                            material,
                            "TPU",
                            "spectroscopy",
                            "FTIR_normalised_dependent_view",
                            points,
                            partial,
                            "wavenumber_cm-1;normalised_signal",
                            "explicit:cm-1;dimensionless",
                            state,
                            group,
                            "hold_dependent_view",
                            0.0,
                            quality_status="derived_dependent",
                            dedup_status="derived_from_raw_spectra",
                            notes="与5个原始吸收光谱同源的归一化视图；保留审计但不重复计权。",
                        )
                    )
                continue
            elif "/FTIR/" in inner:
                modality = "spectroscopy"
                test_type = "FTIR_absorbance"
                columns = (0, 1)
                axis = "wavenumber_cm-1;absorbance_A"
                instance = PurePosixPath(inner).stem
                ceiling = 0.25
            elif "/Rheology/Shear rate analyses/" in inner:
                modality = "rheology"
                test_type = "viscosity_shear_rate"
                columns = (0, 1)
                axis = "shear_rate_s-1;viscosity_Pa_s"
                instance = PurePosixPath(inner).stem
                ceiling = 0.35
            elif "/Rheology/Temperature sweep/" in inner:
                modality = "rheology"
                test_type = "temperature_sweep_multichannel"
                columns = (2, 3, 4, 5, 6, 7, 8)
                axis = (
                    "temperature_C;storage_modulus_Pa;loss_modulus_Pa;complex_viscosity_Pa_s;"
                    "phase_angle_deg;strain;torque_Ncm"
                )
                instance = PurePosixPath(inner).stem
                ceiling = 0.35
            elif "/TGA/" in inner:
                modality = "thermal"
                test_type = "TGA_weight_temperature"
                columns = (4, 1)
                axis = "sample_temperature_C;unsubtracted_weight_g"
                instance = PurePosixPath(inner).stem
            else:
                raise AuditBlocked(f"4TU未知CSV角色：{inner}")
            points, partial = _numeric_points(rows, columns, 1)
            if test_type == "viscosity_shear_rate":
                negative_viscosity = sum(point[1] < 0 for point in points)
                if negative_viscosity:
                    decision = "hold_negative_viscosity"
                    ceiling = 0.0
                    quality = "physically_suspect"
                    notes = f"发现{negative_viscosity}个负黏度点；原值保留，主论文/仪器导出复核前硬零。"
            curves.append(
                _curve_row(
                    SELF_HEALING,
                    _record_id("4tu_curve", inner),
                    filename,
                    inner,
                    material,
                    "TPU",
                    modality,
                    test_type,
                    points,
                    partial,
                    axis,
                    "explicit",
                    instance,
                    group,
                    decision,
                    ceiling,
                    quality_status=quality,
                    dedup_status=dedup,
                    duplicate_of=duplicate_of,
                    notes=notes,
                )
            )
    if csv_count != 65 or txt_count != 6 or uncompressed_bytes != 3_115_377:
        raise AuditBlocked(
            f"4TU归档结构漂移：csv={csv_count}, txt={txt_count}, bytes={uncompressed_bytes}"
        )
    if len(curves) != 68 or sum(int(row["point_count"]) for row in curves) != 148_379:
        raise AuditBlocked("4TU曲线或点数漂移")
    if len(specimen_keys) != 26 or sum(int(row["direct_numeric_result_count"]) for row in scalars) != 32:
        raise AuditBlocked("4TU试样键或标量数漂移")
    candidate_curves = [row for row in curves if row["decision"].endswith("candidate")]
    formulations = [
        _formulation_row(
            SELF_HEALING,
            "SH-TPU",
            "self_healing_thermoplastic_polyurethane",
            "CroHeal 2000 based MDI-p SH-TPU",
            "",
            "",
            "",
            "identity_only",
            "article_identity;exact_stoichiometry_not_in_deposit",
            "Polymers 2021 DOI 10.3390/polym13020305",
            formulation_groups["SH-TPU"],
            0.55,
            "正文确认长链二醇CroHeal 2000与先前命名MDI-p；当前原始包不能恢复完整摩尔配比。",
        ),
        _formulation_row(
            SELF_HEALING,
            "Ninjaflex",
            "commercial_TPU_control",
            "Ninjaflex commercial TPU",
            "",
            "",
            "",
            "commercial_grade_only",
            "archive_and_article",
            "Polymers 2021 DOI 10.3390/polym13020305",
            formulation_groups["Ninjaflex"],
            0.35,
            "商业对照的精确单体/配比未公开。",
        ),
    ]
    summary = {
        "audit_version": AUDIT_VERSION,
        "audit_as_of": AUDIT_DATE,
        "source_directory": SELF_HEALING,
        "canonical_identifier": "doi:10.4121/13603775.v1",
        "publication_identifier": "doi:10.3390/polym13020305",
        "license": "CC-BY-4.0",
        "gold_layer": "Gold-E_transfer_reference",
        "thermoplastic_tpu_core": True,
        "archive_member_count": 71,
        "csv_member_count": 65,
        "text_metadata_member_count": 6,
        "material_count": 2,
        "formulation_count": 2,
        "material_state_count": 8,
        "confirmed_mechanical_specimen_key_count": 26,
        "curve_count_observed": len(curves),
        "curve_count_candidate": len(candidate_curves),
        "point_count_observed": sum(int(row["point_count"]) for row in curves),
        "point_count_candidate": sum(int(row["usable_point_count"]) for row in curves),
        "scalar_count_observed": 32,
        "scalar_count_candidate": 32,
        "derived_dependent_curve_count": 5,
        "negative_viscosity_curve_hold_count": 2,
        "candidate_weight_ceiling": 0.55,
        "known_metadata_defects": [
            "归档TGA_info.txt错误列出filament_thickness.CSV，实际两条TGA文件存在",
            "Mechanical_testing_info.txt把SH-TPU_230C_XZ_2错写为SH-TPU_2330_XZ_2",
            "归档内论文DOI仍为投稿占位符10.3390/xxxxx；正式引用以10.3390/polym13020305为准",
        ],
        "training_split_materialized": False,
        "training_weight_materialized": False,
    }
    return AuditBundle(SELF_HEALING, summary, files, curves, scalars, formulations)


def audit_schwarz() -> AuditBundle:
    source_dir = DATA_ROOT / SCHWARZ
    _require_plain_directory(source_dir)
    filename, size, digest = SCHWARZ_FILE
    path = source_dir / filename
    _require_file(path, size, digest)
    payload = path.read_bytes()
    files = [_file_row(SCHWARZ, filename, "primary_workbook", payload, "CC-BY-4.0")]
    workbook = load_workbook(path, read_only=True, data_only=False)
    if workbook.sheetnames != ["Stress-Stretch", "Failure stress,stretch, E0", "Coef. Diffusion"]:
        raise AuditBlocked(f"Schwarz工作表漂移：{workbook.sheetnames}")
    sheet = workbook["Stress-Stretch"]
    iterator = sheet.iter_rows(values_only=True)
    labels = list(next(iterator))
    headers = list(next(iterator))
    active_pairs = [index for index in range(0, len(labels), 2) if labels[index] is not None]
    if len(active_pairs) != 45:
        raise AuditBlocked(f"Schwarz有效曲线列对漂移：{len(active_pairs)}")
    points_by_pair: dict[int, list[tuple[float, float]]] = {index: [] for index in active_pairs}
    partial_by_pair: Counter[int] = Counter()
    for row in iterator:
        for index in active_pairs:
            x = _finite(row[index] if index < len(row) else None)
            y = _finite(row[index + 1] if index + 1 < len(row) else None)
            if x is not None and y is not None:
                points_by_pair[index].append((x, y))
            elif x is not None or y is not None:
                partial_by_pair[index] += 1
    curves: list[dict[str, Any]] = []
    seen_hashes: dict[str, str] = {}
    for index in active_pairs:
        label = str(labels[index])
        if headers[index] != "λ1" or headers[index + 1] != "P [MPa]":
            raise AuditBlocked(f"Schwarz轴表头漂移：{label}")
        points = points_by_pair[index]
        curve_hash = _sequence_sha256(points)
        duplicate_of = seen_hashes.get(curve_hash, "")
        seen_hashes[curve_hash] = label
        decision = "hold_exact_duplicate" if duplicate_of else "high_value_candidate"
        curves.append(
            _curve_row(
                SCHWARZ,
                _record_id("schwarz_curve", label),
                filename,
                f"Stress-Stretch:{label}",
                "EPU40",
                "polyurethane_elastomer",
                "mechanical",
                "uniaxial_tensile_stress_stretch",
                points,
                partial_by_pair[index],
                "stretch_ratio_lambda1;nominal_stress_MPa",
                "explicit:dimensionless;MPa",
                label,
                "doi:10.17632/wcwtjrkfsm.1|EPU40",
                decision,
                0.65 if not duplicate_of else 0.0,
                quality_status="accepted_reference" if not duplicate_of else "exact_duplicate",
                dedup_status="unique" if not duplicate_of else "exact_duplicate",
                duplicate_of=duplicate_of,
                notes="EPU40光固化打印聚氨酯；同一材料的干燥/水浸时间条件必须同折。",
            )
        )
    if sum(int(row["point_count"]) for row in curves) != 73_500:
        raise AuditBlocked("Schwarz曲线点数漂移")

    scalars: list[dict[str, Any]] = []
    failure = workbook["Failure stress,stretch, E0"]
    failure_rows = [
        row
        for row in failure.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    ]
    if len(failure_rows) != 45:
        raise AuditBlocked(f"Schwarz失效标量行数漂移：{len(failure_rows)}")
    for row in failure_rows:
        label = str(row[0])
        values = [_finite(value) for value in row[1:4]]
        if any(value is None for value in values):
            raise AuditBlocked(f"Schwarz失效标量非数值：{label}")
        scalars.append(
            _scalar_row(
                SCHWARZ,
                _record_id("schwarz_failure", label),
                filename,
                f"Failure stress,stretch, E0:{label}",
                "EPU40",
                "failure_and_initial_modulus",
                label,
                "failure_nominal_stress_MPa;failure_stretch_ratio;initial_modulus_MPa",
                3,
                0,
                3,
                "explicit:MPa;dimensionless;MPa",
                "high_value_candidate",
                0.65,
                "doi:10.17632/wcwtjrkfsm.1|EPU40",
                "与同名拉伸曲线为同一试样的作者派生端点，不与母曲线重复计权。",
            )
        )
    diffusion = workbook["Coef. Diffusion"]
    diffusion_rows = [
        row
        for row in diffusion.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    ]
    if len(diffusion_rows) != 35:
        raise AuditBlocked(f"Schwarz吸水支持行数漂移：{len(diffusion_rows)}")
    for row in diffusion_rows:
        label = str(row[0])
        values = [_finite(value) for value in row[1:3]]
        if any(value is None for value in values):
            raise AuditBlocked(f"Schwarz吸水支持标量非数值：{label}")
        scalars.append(
            _scalar_row(
                SCHWARZ,
                _record_id("schwarz_diffusion", label),
                filename,
                f"Coef. Diffusion:{label}",
                "EPU40",
                "water_uptake_support",
                label,
                "thickness_mm;mass_g",
                2,
                0,
                2,
                "explicit:mm;g",
                "low_weight_candidate",
                0.35,
                "doi:10.17632/wcwtjrkfsm.1|EPU40",
                "35个条件槽的厚度和质量；与45个拉伸试样是否同一物理件不能由工作簿证明。",
            )
        )
    formulation = _formulation_row(
        SCHWARZ,
        "EPU40",
        "commercial_photopolymerized_elastomeric_polyurethane",
        "EPU40 commercial resin",
        "",
        "",
        "",
        "identity_only",
        "commercial_grade_only",
        "Mendeley workbook and Polymers 2022 article",
        "doi:10.17632/wcwtjrkfsm.1|EPU40",
        0.65,
        "EPU40是商业光固化聚氨酯树脂；单体组成未公开，不构造伪SMILES。",
    )
    summary = {
        "audit_version": AUDIT_VERSION,
        "audit_as_of": AUDIT_DATE,
        "source_directory": SCHWARZ,
        "canonical_identifier": "doi:10.17632/wcwtjrkfsm.1",
        "publication_identifier": "doi:10.3390/polym14245496",
        "license": "CC-BY-4.0",
        "gold_layer": "Gold-E_transfer_reference",
        "thermoplastic_tpu_core": False,
        "material_count": 1,
        "formulation_count": 1,
        "confirmed_tensile_specimen_count": 45,
        "additional_mass_thickness_condition_slot_count": 35,
        "curve_count_observed": len(curves),
        "curve_count_candidate": sum(row["decision"].endswith("candidate") for row in curves),
        "point_count_observed": sum(int(row["point_count"]) for row in curves),
        "point_count_candidate": sum(int(row["usable_point_count"]) for row in curves),
        "scalar_count_observed": sum(
            int(row["direct_numeric_result_count"]) + int(row["derived_numeric_result_count"])
            for row in scalars
        ),
        "scalar_count_candidate": sum(int(row["candidate_numeric_result_count"]) for row in scalars),
        "exact_curve_duplicate_count": 0,
        "formatted_but_empty_curve_pair_count": 10,
        "candidate_weight_ceiling": 0.65,
        "training_split_materialized": False,
        "training_weight_materialized": False,
    }
    return AuditBundle(SCHWARZ, summary, files, curves, scalars, [formulation])


def _zenodo_4156_manifest(source_dir: Path) -> tuple[list[Path], str]:
    paths = sorted(source_dir.glob("*.csv"), key=lambda item: item.name)
    lines: list[str] = []
    for path in paths:
        if not path.is_file() or _is_reparse_point(path):
            raise AuditBlocked(f"Zenodo4156000原件不是普通文件：{path}")
        lines.append(f"{path.name}\t{path.stat().st_size}\t{_sha256_file(path)}")
    manifest = ("\n".join(lines) + "\n").encode("utf-8")
    return paths, _sha256_bytes(manifest)


def _zenodo_4156_material(filename: str) -> str:
    match = re.search(r"_(Eel0\.5|Empa0\.3|Empa0\.5)(?:_|\.)", filename, re.I)
    if not match:
        match = re.search(r"_(eel0\.5|empa0\.3|empa0\.5)\.csv$", filename, re.I)
    if not match:
        raise AuditBlocked(f"Zenodo4156000材料标签无法解析：{filename}")
    raw = match.group(1).lower()
    return {"eel0.5": "Eel0.5", "empa0.3": "Empa0.3", "empa0.5": "Empa0.5"}[raw]


def _zenodo_4156_formulation(material_condition: str) -> str:
    if material_condition == "Eel0.5":
        return "Eel_TPU_CB18"
    if material_condition in {"Empa0.3", "Empa0.5"}:
        return "Empa_SEBS_CB_1to1"
    raise AuditBlocked(f"Zenodo4156000未知材料条件：{material_condition}")


def audit_zenodo_4156() -> AuditBundle:
    source_dir = DATA_ROOT / ZENODO_4156
    _require_plain_directory(source_dir)
    paths, manifest_sha = _zenodo_4156_manifest(source_dir)
    if len(paths) != 15 or manifest_sha != ZENODO_4156_MANIFEST_SHA256:
        raise AuditBlocked(
            f"Zenodo4156000目录清单漂移：files={len(paths)} sha256={manifest_sha}"
        )
    files: list[dict[str, Any]] = []
    curves: list[dict[str, Any]] = []
    file_hash_owner: dict[str, str] = {}
    curve_hash_owner: dict[str, str] = {}
    for path in paths:
        payload = path.read_bytes()
        file_hash = _sha256_bytes(payload)
        duplicate_file = file_hash_owner.get(file_hash, "")
        file_hash_owner.setdefault(file_hash, path.name)
        files.append(
            _file_row(
                ZENODO_4156,
                path.name,
                "primary_multichannel_csv",
                payload,
                "CC-BY-4.0",
                dedup_status="exact_duplicate" if duplicate_file else "unique",
                duplicate_of=duplicate_file,
            )
        )
        rows = _read_semicolon(payload)
        material_condition = _zenodo_4156_material(path.name)
        formulation = _zenodo_4156_formulation(material_condition)
        material_scope = (
            "TPU_CB_composite"
            if formulation == "Eel_TPU_CB18"
            else "TPS_SEBS_CB_composite"
        )
        candidate_ceiling = 0.35 if formulation == "Eel_TPU_CB18" else 0.15
        if path.name.startswith("Dynamic_"):
            channel_specs = (
                ("strain_time", (0, 1), "time_s;strain_percent", "protocol_only"),
                (
                    "relative_resistance_time",
                    (3, 4),
                    "time_s;relative_resistance",
                    "response",
                ),
            )
            protocol = "dynamic_tensile_0_70"
        elif path.name.startswith("Quasi_"):
            wide = len(rows[0]) >= 9
            channel_specs = (
                ("strain_time", (0, 1), "time_s;strain_percent", "protocol_only"),
                (
                    "stress_time",
                    (4, 5) if wide else (3, 4),
                    "time_s;stress_MPa",
                    "response",
                ),
                (
                    "relative_resistance_time",
                    (7, 8) if wide else (6, 7),
                    "time_s;relative_resistance",
                    "response",
                ),
            )
            protocol = "quasi_static"
        else:
            channel_specs = (
                ("stress_strain", (0, 1), "strain_percent;stress_MPa", "response"),
                (
                    "relative_resistance_strain",
                    (3, 4),
                    "strain_percent;relative_resistance",
                    "response",
                ),
            )
            speed_match = re.search(r"_(50|100|200)_mm_min", path.name)
            if not speed_match:
                raise AuditBlocked(f"Zenodo4156000断裂速度无法解析：{path.name}")
            protocol = f"tensile_to_breakpoint_{speed_match.group(1)}_mm_min"
        for modality, columns, axis, role in channel_specs:
            points, partial = _numeric_points(rows, columns, 1)
            curve_hash = _sequence_sha256(points)
            duplicate_curve = curve_hash_owner.get(curve_hash, "")
            curve_hash_owner.setdefault(curve_hash, f"{path.name}#{modality}")
            if role == "protocol_only":
                decision = "hold_protocol_coordinate_only"
                ceiling = 0.0
                quality = "protocol_context"
                dedup = "shared_protocol_channel" if duplicate_curve else "protocol_channel"
                notes = "施加的时间—应变程序只作输入条件，不作为材料响应标签。"
            elif duplicate_file:
                decision = "hold_exact_duplicate_file"
                ceiling = 0.0
                quality = "exact_duplicate"
                dedup = "exact_duplicate"
                notes = "不同材料文件名字节完全一致；保留审计，重复件不增加样本。"
            else:
                decision = "low_weight_candidate"
                ceiling = candidate_ceiling
                quality = "accepted_reference"
                dedup = "shared_protocol_only" if duplicate_curve else "unique"
                notes = (
                    "力学与电阻通道来自同一文件/运行，必须共享拆分键；"
                    + (
                        "Eel为导电TPU应用迁移。"
                        if formulation == "Eel_TPU_CB18"
                        else "Empa为SEBS/TPS跨材料迁移，不得作为TPU本征标签。"
                    )
                )
            curves.append(
                _curve_row(
                    ZENODO_4156,
                    _record_id("zen4156_curve", path.name, modality),
                    path.name,
                    f"{path.name}#{modality}",
                    formulation,
                    material_scope,
                    "mechanical_electrical_linked",
                    modality,
                    points,
                    partial,
                    axis,
                    "explicit",
                    path.stem,
                    f"zenodo:4156000|{formulation}",
                    decision,
                    ceiling,
                    quality_status=quality,
                    dedup_status=dedup,
                    duplicate_of=duplicate_curve if duplicate_file else "",
                    notes=notes,
                )
            )
    if len(curves) != 33 or sum(int(row["point_count"]) for row in curves) != 377_353:
        raise AuditBlocked("Zenodo4156000曲线或点数漂移")
    candidate_curves = [row for row in curves if row["decision"].endswith("candidate")]
    if len(candidate_curves) != 25 or sum(int(row["usable_point_count"]) for row in curves) != 152_271:
        raise AuditBlocked("Zenodo4156000候选曲线或点数漂移")
    formulations = [
        _formulation_row(
            ZENODO_4156,
            "Eel_TPU_CB18",
            "conductive_thermoplastic_elastomer_composite",
            "Ninjatek Eel conductive TPU",
            82,
            "carbon black",
            18,
            "approximate_mass_percent",
            "article_resolved",
            "accepted manuscript methods 2.2; approximately 18 wt% carbon black",
            "zenodo:4156000|Eel_TPU_CB18",
            0.35,
            "Eel0.5是约0.5 mm丝材/喷嘴条件，不是独立化学配方。",
        ),
        _formulation_row(
            ZENODO_4156,
            "Empa_SEBS_CB_1to1",
            "conductive_sebs_thermoplastic_elastomer_composite",
            "SEBS-based thermoplastic elastomer",
            1,
            "ENASCO 250 carbon black",
            1,
            "mass_ratio",
            "article_resolved",
            "accepted manuscript methods 2.1-2.2; SEBS and carbon black mixed 1:1",
            "zenodo:4156000|Empa_SEBS_CB_1to1",
            0.15,
            "Empa0.3和Empa0.5是同一配方的0.3/0.5 mm工艺条件，必须共享配方与泄漏组。",
        ),
    ]
    summary = {
        "audit_version": AUDIT_VERSION,
        "audit_as_of": AUDIT_DATE,
        "source_directory": ZENODO_4156,
        "canonical_identifier": "zenodo:4156000",
        "publication_identifier": "doi:10.1088/2058-8585/ab9a22",
        "license": "CC-BY-4.0",
        "gold_layer": "Gold-E_application_transfer_reference",
        "thermoplastic_tpu_core": False,
        "material_count": 2,
        "formulation_count": 2,
        "material_condition_count": 3,
        "file_count": 15,
        "unique_file_payload_count": 14,
        "confirmed_independent_specimen_count": None,
        "run_payload_count": 14,
        "curve_count_observed": len(curves),
        "curve_count_candidate": len(candidate_curves),
        "point_count_observed": sum(int(row["point_count"]) for row in curves),
        "point_count_candidate": sum(int(row["usable_point_count"]) for row in curves),
        "protocol_coordinate_curve_count": 6,
        "response_curve_duplicate_count": 2,
        "partial_axis_or_response_row_count": sum(int(row["partial_point_count"]) for row in curves),
        "candidate_weight_ceiling": 0.35,
        "training_split_materialized": False,
        "training_weight_materialized": False,
    }
    return AuditBundle(ZENODO_4156, summary, files, curves, [], formulations)


def audit_zenodo_1098() -> AuditBundle:
    source_dir = DATA_ROOT / ZENODO_1098
    _require_plain_directory(source_dir)
    filename, size, digest = ZENODO_1098_FILE
    path = source_dir / filename
    _require_file(path, size, digest)
    payload = path.read_bytes()
    files = [_file_row(ZENODO_1098, filename, "primary_workbook", payload, "CC-BY-4.0")]
    workbook = load_workbook(path, read_only=True, data_only=False)
    expected_sheets = [
        "Content",
        "1. FTIR",
        "2. Tensile Properties",
        "3. Conductivity ",
        "4. Piezoresistive Properties ",
        "5. Resistance vs Strain",
        " 6. Porosity",
    ]
    if workbook.sheetnames != expected_sheets:
        raise AuditBlocked(f"Zenodo1098206工作表漂移：{workbook.sheetnames}")
    curves: list[dict[str, Any]] = []
    group_root = "doi:10.1038/s41598-017-17647-w"

    ftir = workbook["1. FTIR"]
    for material, start in zip(
        ("TPU", "TPU10", "TPU20", "TPU30", "TPU40", "CNFs"),
        (5, 7, 9, 11, 13, 15),
    ):
        points: list[tuple[float, float]] = []
        partial = 0
        for row in ftir.iter_rows(min_row=5, min_col=start, max_col=start + 1, values_only=True):
            x, y = (_finite(value) for value in row)
            if x is not None and y is not None:
                points.append((x, y))
            elif x is not None or y is not None:
                partial += 1
        curves.append(
            _curve_row(
                ZENODO_1098,
                _record_id("zen1098_ftir", material),
                filename,
                f"1. FTIR:{material}",
                material,
                "TPU_CNF_composite" if material != "CNFs" else "CNF_reference",
                "spectroscopy",
                "FTIR_transmittance",
                points,
                partial,
                "wavenumber_cm-1;transmittance_percent",
                "explicit:cm-1;percent",
                material,
                f"{group_root}|{material}",
                "low_weight_candidate",
                0.25 if material != "CNFs" else 0.10,
                notes="CNFs单独光谱仅作组分参考；所有复合配方必须按材料家族同折。",
            )
        )

    tensile = workbook["2. Tensile Properties"]
    for material, start in zip(
        ("TPU", "TPU10", "TPU20", "TPU30", "TPU40"),
        (5, 16, 27, 38, 49),
    ):
        for replicate in range(5):
            points: list[tuple[float, float]] = []
            partial = 0
            for row in tensile.iter_rows(
                min_row=6,
                min_col=start + replicate * 2,
                max_col=start + replicate * 2 + 1,
                values_only=True,
            ):
                x, y = (_finite(value) for value in row)
                if x is not None and y is not None:
                    points.append((x, y))
                elif x is not None or y is not None:
                    partial += 1
            specimen = f"{material}-S{replicate + 1}"
            curves.append(
                _curve_row(
                    ZENODO_1098,
                    _record_id("zen1098_tensile", specimen),
                    filename,
                    f"2. Tensile Properties:{specimen}",
                    material,
                    "TPU_CNF_composite",
                    "mechanical",
                    "uniaxial_tensile_stress_strain",
                    points,
                    partial,
                    "strain_percent;stress_MPa",
                    "explicit:percent;MPa",
                    specimen,
                    f"{group_root}|{material}",
                    "high_value_candidate",
                    0.60 if material == "TPU" else 0.50,
                    notes="矩形50×2×2 mm，25°C，100 mm/min；每配方5个明确重复试样。",
                )
            )

    piezo = workbook["4. Piezoresistive Properties "]
    for material, first_row in zip(("TPU10", "TPU20", "TPU30", "TPU40"), (4, 12, 20, 28)):
        pressures = [_finite(piezo.cell(row, 6).value) for row in range(first_row, first_row + 6)]
        if any(value is None for value in pressures):
            raise AuditBlocked(f"Zenodo1098206压力轴漂移：{material}")
        for replicate, column in enumerate(range(7, 12), start=1):
            resistance = [_finite(piezo.cell(row, column).value) for row in range(first_row, first_row + 6)]
            points = [
                (float(x), float(y))
                for x, y in zip(pressures, resistance)
                if x is not None and y is not None
            ]
            specimen = f"{material}-P{replicate}"
            curves.append(
                _curve_row(
                    ZENODO_1098,
                    _record_id("zen1098_piezo", specimen),
                    filename,
                    f"4. Piezoresistive Properties:{specimen}",
                    material,
                    "TPU_CNF_composite",
                    "electrical",
                    "resistance_pressure",
                    points,
                    0,
                    "pressure_MPa;resistance_Ohm",
                    "explicit:MPa;Ohm",
                    specimen,
                    f"{group_root}|{material}",
                    "low_weight_candidate",
                    0.35,
                    notes="5个重复通道×6个压力点；同配方重复不能跨训练/测试折。",
                )
            )

    resistance = workbook["5. Resistance vs Strain"]
    for material, start in zip(("TPU10", "TPU20", "TPU30", "TPU40"), (5, 9, 13, 17)):
        points: list[tuple[float, float, float]] = []
        partial = 0
        for row in resistance.iter_rows(min_row=5, min_col=start, max_col=start + 2, values_only=True):
            values = tuple(_finite(value) for value in row)
            if all(value is not None for value in values):
                points.append(tuple(float(value) for value in values if value is not None))
            elif any(value is not None for value in values):
                partial += 1
        curves.append(
            _curve_row(
                ZENODO_1098,
                _record_id("zen1098_resistance_strain", material),
                filename,
                f"5. Resistance vs Strain:{material}",
                material,
                "TPU_CNF_composite",
                "electromechanical",
                "resistance_deformation_time",
                points,
                partial,
                "time_s;deformation_mm;resistance_Ohm",
                "explicit:s;mm;Ohm",
                material,
                f"{group_root}|{material}",
                "low_weight_candidate",
                0.35,
                notes="时间、变形、电阻三通道同一运行；不拆成独立样本。",
            )
        )
    if len(curves) != 55 or sum(int(row["point_count"]) for row in curves) != 43_032:
        raise AuditBlocked("Zenodo1098206曲线或点数漂移")
    scalars: list[dict[str, Any]] = []
    conductivity = workbook["3. Conductivity "]
    materials = ("TPU10", "TPU20", "TPU30", "TPU40")
    columns = (6, 7, 8, 9)
    valid_conductivity_formula_counts = {"TPU10": 5, "TPU20": 4, "TPU30": 0, "TPU40": 5}
    for material, source_row, column in zip(materials, range(5, 9), columns):
        thickness = _finite(conductivity.cell(source_row, 6).value)
        radius = _finite(conductivity.cell(source_row, 7).value)
        resistances = [_finite(conductivity.cell(row, column).value) for row in range(12, 17)]
        if thickness is None or radius is None or any(value is None for value in resistances):
            raise AuditBlocked(f"Zenodo1098206电导原始标量漂移：{material}")
        scalars.append(
            _scalar_row(
                ZENODO_1098,
                _record_id("zen1098_conductivity_raw", material),
                filename,
                f"3. Conductivity:{material}:raw",
                material,
                "conductivity_raw_support",
                "disk_specimen",
                "thickness_mm;radius_mm;resistance_Ohm_x5",
                7,
                0,
                7,
                "explicit:mm;mm;Ohm",
                "low_weight_candidate",
                0.35,
                f"{group_root}|{material}",
                "几何量与5个电阻重复共享同一派生关系；不得与派生电导率重复计权。",
            )
        )
        valid = valid_conductivity_formula_counts[material]
        decision = "low_weight_candidate" if valid else "hold_formula_reference_error"
        scalars.append(
            _scalar_row(
                ZENODO_1098,
                _record_id("zen1098_conductivity_derived", material),
                filename,
                f"3. Conductivity:{material}:derived",
                material,
                "conductivity_formula_derived",
                "disk_specimen",
                "conductivity_S_cm_x5",
                0,
                5,
                valid,
                "formula_derived:S/cm",
                decision,
                0.30 if valid else 0.0,
                f"{group_root}|{material}",
                (
                    f"5个工作簿电导率公式中{valid}个厚度引用正确；其余引用了其他配方厚度。"
                    "均值/标准差公式仅作依赖汇总，不增加独立标签。"
                ),
            )
        )
    porosity = workbook[" 6. Porosity"]
    for row in range(9, 14):
        material = str(porosity.cell(row, 2).value)
        density_values = [_finite(porosity.cell(row, column).value) for column in (3, 4)]
        formula = porosity.cell(row, 5).value
        if any(value is None for value in density_values) or not str(formula).startswith("=1-"):
            raise AuditBlocked(f"Zenodo1098206孔隙率表漂移：{material}")
        scalars.append(
            _scalar_row(
                ZENODO_1098,
                _record_id("zen1098_porosity", material),
                filename,
                f"6. Porosity:{material}",
                material,
                "porosity_and_density",
                "porous_film",
                "true_density_g_cm3;bulk_density_g_cm3;porosity_fraction_derived",
                2,
                1,
                3,
                "explicit:g/cm3;g/cm3;dimensionless",
                "low_weight_candidate",
                0.40,
                f"{group_root}|{material}",
                "孔隙率为工作簿公式1-true_density/bulk_density；保留公式血缘，不与两项密度重复计权。",
            )
        )
    observed_scalars = sum(
        int(row["direct_numeric_result_count"]) + int(row["derived_numeric_result_count"])
        for row in scalars
    )
    candidate_scalars = sum(int(row["candidate_numeric_result_count"]) for row in scalars)
    if observed_scalars != 63 or candidate_scalars != 57:
        raise AuditBlocked(
            f"Zenodo1098206标量计数漂移：observed={observed_scalars}, candidate={candidate_scalars}"
        )
    formulations = []
    for loading in (0, 10, 20, 30, 40):
        material = "TPU" if loading == 0 else f"TPU{loading}"
        formulations.append(
            _formulation_row(
                ZENODO_1098,
                material,
                "porous_CNF_TPU_nanocomposite",
                "IROGRAN PS 455-203 TPU",
                100 - loading,
                "carbon nanofibers",
                loading,
                "mass_percent",
                "article_resolved",
                "Scientific Reports 2017 methods",
                f"{group_root}|{material}",
                0.60 if loading == 0 else 0.50,
                "TPU为Huntsman IROGRAN PS 455-203；CNF质量分数由正文明确为0/10/20/30/40 wt%。",
            )
        )
    formulations.append(
        _formulation_row(
            ZENODO_1098,
            "CNFs",
            "component_reference",
            "carbon nanofibers",
            100,
            "",
            "",
            "mass_percent",
            "component_only",
            "Scientific Reports 2017 methods",
            f"{group_root}|CNFs",
            0.10,
            "仅有FTIR组分参考，不是TPU配方。",
        )
    )
    summary = {
        "audit_version": AUDIT_VERSION,
        "audit_as_of": AUDIT_DATE,
        "source_directory": ZENODO_1098,
        "canonical_identifier": "zenodo:1098206",
        "publication_identifier": "doi:10.1038/s41598-017-17647-w",
        "license": "CC-BY-4.0",
        "gold_layer": "Gold-E_composite_transfer_reference",
        "thermoplastic_tpu_core": True,
        "commercial_tpu_grade": "IROGRAN PS 455-203",
        "material_or_component_identity_count": 6,
        "formulation_count": 5,
        "explicit_tensile_specimen_count": 25,
        "curve_count_observed": len(curves),
        "curve_count_candidate": len(curves),
        "point_count_observed": sum(int(row["point_count"]) for row in curves),
        "point_count_candidate": sum(int(row["usable_point_count"]) for row in curves),
        "scalar_count_observed": observed_scalars,
        "scalar_count_candidate": candidate_scalars,
        "conductivity_individual_formula_count": 20,
        "conductivity_formula_reference_error_count": 6,
        "dependent_mean_and_standard_deviation_formula_count": 8,
        "porosity_formula_count": 5,
        "exact_curve_duplicate_count": 0,
        "candidate_weight_ceiling": 0.60,
        "training_split_materialized": False,
        "training_weight_materialized": False,
    }
    return AuditBundle(ZENODO_1098, summary, files, curves, scalars, formulations)


def _tsv_bytes(rows: list[dict[str, Any]], columns: tuple[str, ...]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=list(columns),
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in columns})
    return buffer.getvalue().encode("utf-8-sig")


def render_outputs(bundle: AuditBundle) -> dict[str, bytes]:
    return {
        "内容审计摘要.json": (
            json.dumps(bundle.summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
        ).encode("utf-8"),
        "文件校验清单.tsv": _tsv_bytes(bundle.files, FILE_COLUMNS),
        "曲线审计清单.tsv": _tsv_bytes(bundle.curves, CURVE_COLUMNS),
        "标量审计清单.tsv": _tsv_bytes(bundle.scalars, SCALAR_COLUMNS),
        "配方审计清单.tsv": _tsv_bytes(bundle.formulations, FORMULATION_COLUMNS),
    }


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"审计输出不在白名单：{path}")
    if _is_reparse_point(path) or _is_reparse_point(path.parent):
        raise AuditBlocked(f"审计输出路径为符号链接或联接点：{path}")
    if path.parent.resolve(strict=True) != path.parent.absolute():
        raise AuditBlocked(f"审计输出父目录解析发生漂移：{path.parent}")


def atomic_write(path: Path, payload: bytes) -> None:
    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def main() -> None:
    bundles = (
        audit_self_healing(),
        audit_schwarz(),
        audit_zenodo_4156(),
        audit_zenodo_1098(),
    )
    for bundle in bundles:
        source_dir = DATA_ROOT / bundle.source_directory
        for filename, payload in render_outputs(bundle).items():
            atomic_write(source_dir / filename, payload)
        print(
            f"{bundle.source_directory}: curves={bundle.summary['curve_count_observed']} "
            f"candidate={bundle.summary['curve_count_candidate']} "
            f"points={bundle.summary['point_count_observed']}"
        )


if __name__ == "__main__":
    main()
