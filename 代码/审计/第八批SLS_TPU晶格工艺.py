"""审计 Figshare SLS-TPU 晶格工艺数据（10.6084/m9.figshare.31550614.v1）。

该来源包含一个 L25（5^3）工艺设计：25 个工况、每工况 3 个独立试件，
共 75 个试件。脚本把 25%载荷、65%载荷、SAG、HLR 和 Weight 展开为
375 条逐试件标量记录。Weight 的原文件没有单位，因此保留为条件参考，
不猜测单位；其余 300 条记录单位闭合。

原补充材料将 LP/(SS*HD) 称为“volumetric ED”，但公式和单位 J/mm²
实际对应面能量密度。脚本同时保存官方面能量密度，以及用固定层厚 0.1 mm
推导的体积能量密度；后者是可复算过程特征，不冒充额外实验观测。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import statistics
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第八批实验_SLS_TPU晶格工艺"
)
SOURCE_XLSX = SOURCE_DIR / "原始数据.xlsx"
SOURCE_DOCX = SOURCE_DIR / "补充材料.docx"
SOURCE_SNAPSHOT = SOURCE_DIR / "来源快照.json"

DOI = "10.6084/m9.figshare.31550614.v1"
LICENSE = "CC BY 4.0"
SOURCE_URL = (
    "https://figshare.com/articles/dataset/"
    "Dataset_and_Supplementary_Files_for_Influence_of_SLS_Process_Parameters_"
    "on_Cushion_Properties_and_Sustainability_of_TPU_Lattice_Structures_for_"
    "Automotive_Seating_/31550614"
)
AUDIT_VERSION = "batch8-sls-tpu-lattice-process-v1"
RAW_SHEET = "L25_OA table"
PROCESSED_SHEET = "Processed_Metrics"

FROZEN_FILES = {
    "原始数据.xlsx": (
        63_965,
        "ed327722f815fbf171a5a9b126c3438a8a85fb492429b0fb7016c0c286a7747a",
        "93176aff87ae5a15d9889ca0cfdb8918",
    ),
    "补充材料.docx": (
        714_893,
        "62614bab5fb27c1f7987d323c6e1e4acbfc28edbdb1e449be2c1187215b659db",
        "65dfd162ddbaa3d9f5ab0da1891ea02a",
    ),
    "来源快照.json": (
        6_750,
        "369b0dab05b2f76cf444dea674db593f08e41ce6ba5eb2f382f6e95782c080e9",
        "2980fee02a3f0c01802dce267bc52276",
    ),
}

EXPECTED_SPECIMENS = 75
EXPECTED_CONDITIONS = 25
EXPECTED_REPLICATES = 3
EXPECTED_SCALARS = 375
EXPECTED_UNIT_CLOSED = 300
EXPECTED_UNIT_UNRESOLVED = 75
OUTPUT_NAMES = ("内容审计摘要.json", "标量审计清单.tsv", "文件校验清单.tsv")

RAW_HEADERS = (
    "Specimen #",
    "LP",
    "SS",
    "HD",
    "Load@25%",
    "Load@65%",
    "SAG(1)",
    "HLR(1)",
    "set#",
)

SCALAR_COLUMNS = (
    "scalar_id",
    "specimen_id",
    "source_row",
    "source_location",
    "condition_id",
    "replicate_index",
    "split_group",
    "repeat_leakage_group",
    "material_geometry_leakage_group",
    "observable",
    "value",
    "unit",
    "unit_status",
    "target_origin",
    "data_origin",
    "record_granularity",
    "quality_gate",
    "gold_layer",
    "gold_admission_status",
    "future_weight_ceiling",
    "laser_power_w",
    "scan_speed_mm_s",
    "hatch_distance_mm",
    "layer_thickness_mm",
    "energy_density_areal_j_mm2",
    "energy_density_volumetric_j_mm3",
    "energy_density_areal_provenance",
    "energy_density_volumetric_provenance",
    "chemistry_resolution",
    "geometry_resolution",
    "notes",
)


class AuditBlocked(RuntimeError):
    """原件身份、工作簿结构或冻结计数发生漂移。"""


def _digest(path: Path, algorithm: str) -> str:
    digest = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_files() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for name, (expected_bytes, expected_sha256, expected_md5) in FROZEN_FILES.items():
        path = SOURCE_DIR / name
        if not path.is_file():
            raise AuditBlocked(f"缺少原件：{path}")
        actual = (path.stat().st_size, _digest(path, "sha256"), _digest(path, "md5"))
        expected = (expected_bytes, expected_sha256, expected_md5)
        if actual != expected:
            raise AuditBlocked(
                f"原件漂移：{name} bytes={actual[0]}, sha256={actual[1]}, md5={actual[2]}"
            )
        rows.append(
            {
                "file": name,
                "bytes": actual[0],
                "sha256": actual[1],
                "md5": actual[2],
                "verification": "matched_frozen_identity",
            }
        )

    snapshot = json.loads(SOURCE_SNAPSHOT.read_text(encoding="utf-8"))
    if snapshot.get("doi") != DOI or snapshot.get("version") != 1:
        raise AuditBlocked("Figshare DOI 或版本漂移")
    if snapshot.get("license", {}).get("name") != LICENSE:
        raise AuditBlocked("Figshare 许可漂移")
    if snapshot.get("status") != "public" or snapshot.get("download_disabled"):
        raise AuditBlocked("Figshare 公开或下载状态漂移")
    official_files = {entry["name"]: entry for entry in snapshot.get("files", [])}
    for official_name, local_name in (
        ("SLS_TPU_Lattice_Cushion_Metrics_Master.xlsx", "原始数据.xlsx"),
        ("Supplementary Material.docx", "补充材料.docx"),
    ):
        entry = official_files.get(official_name)
        if not entry:
            raise AuditBlocked(f"官方元数据缺少：{official_name}")
        expected_bytes, _, expected_md5 = FROZEN_FILES[local_name]
        if (entry.get("size"), entry.get("supplied_md5")) != (
            expected_bytes,
            expected_md5,
        ):
            raise AuditBlocked(f"官方文件身份漂移：{official_name}")
    return rows


def _docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))
    fragments = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
    return " ".join(" ".join(fragments).split())


def verify_protocol() -> dict[str, object]:
    text = _docx_text(SOURCE_DOCX)
    required = (
        "L25 (5³) Orthogonal array",
        "Each run was repeated three times, resulting in 75 specimens",
        "layer thickness (LT) was fixed at 0.1 mm",
        "ED = LP / (SS × HD)",
        "expressed in J/mm²",
        "seven levels of SS and two replicates per level",
    )
    missing = [phrase for phrase in required if phrase not in text]
    if missing:
        raise AuditBlocked(f"补充材料协议证据漂移：{missing}")
    return {
        "l25_conditions": EXPECTED_CONDITIONS,
        "replicates_per_condition": EXPECTED_REPLICATES,
        "fixed_layer_thickness_mm": 0.1,
        "official_energy_density_formula": "LP/(SS*HD)",
        "official_energy_density_unit": "J/mm^2",
        "additional_scan_speed_validation_levels": 7,
        "additional_scan_speed_validation_replicates_per_level": 2,
        "additional_scan_speed_validation_specimens_without_rows": 14,
        "additional_validation_in_scalar_manifest": False,
    }


def _number(value: object, *, label: str, excel_row: int) -> float:
    if isinstance(value, bool):
        raise AuditBlocked(f"{label} 第{excel_row}行不是数值")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise AuditBlocked(f"{label} 第{excel_row}行无法解析：{value!r}") from exc
    if not math.isfinite(result):
        raise AuditBlocked(f"{label} 第{excel_row}行不是有限数")
    return result


def _maximum_abs(values: list[float]) -> float:
    return max((abs(value) for value in values), default=0.0)


def parse_workbook() -> tuple[list[dict[str, object]], dict[str, object]]:
    # 该工作簿很小且审计需要大量随机单元格访问；普通只读语义加载比
    # openpyxl 的流式 read_only 模式更快，同时脚本从不调用 save()。
    formulas = load_workbook(SOURCE_XLSX, read_only=False, data_only=False)
    values = load_workbook(SOURCE_XLSX, read_only=False, data_only=True)
    try:
        if formulas.sheetnames != [RAW_SHEET, PROCESSED_SHEET]:
            raise AuditBlocked(f"工作表漂移：{formulas.sheetnames}")
        raw = values[RAW_SHEET]
        processed = values[PROCESSED_SHEET]
        processed_formulas = formulas[PROCESSED_SHEET]
        headers = tuple(raw.cell(1, column).value for column in range(1, 10))
        if headers != RAW_HEADERS:
            raise AuditBlocked(f"原始表表头漂移：{headers}")
        if processed.cell(1, 5).value != "ED[J/mm2]" or processed.cell(1, 6).value != "LT":
            raise AuditBlocked("Processed_Metrics 的 ED/LT 表头漂移")

        processed_rows: dict[tuple[float, float, float], dict[str, object]] = {}
        condition_ids: dict[tuple[float, float, float], str] = {}
        average_errors: list[float] = []
        ed_errors: list[float] = []
        duplicate_load_errors: list[float] = []
        for excel_row in range(2, 27):
            condition_no = int(_number(processed.cell(excel_row, 1).value, label="工况", excel_row=excel_row))
            lp = _number(processed.cell(excel_row, 2).value, label="LP", excel_row=excel_row)
            ss = _number(processed.cell(excel_row, 3).value, label="SS", excel_row=excel_row)
            hd = _number(processed.cell(excel_row, 4).value, label="HD", excel_row=excel_row)
            lt = _number(processed.cell(excel_row, 6).value, label="LT", excel_row=excel_row)
            key = (lp, ss, hd)
            if key in processed_rows:
                raise AuditBlocked(f"Processed_Metrics 重复工况：{key}")
            if condition_no != excel_row - 1 or not math.isclose(lt, 0.1, abs_tol=1e-12):
                raise AuditBlocked(f"工况编号或层厚漂移：第{excel_row}行")
            formula = str(processed_formulas.cell(excel_row, 5).value).replace("$", "")
            if not re.fullmatch(r"=\(?B\d+\)?/\(C\d+\*D\d+\)", formula):
                raise AuditBlocked(f"ED公式漂移：{processed_formulas.cell(excel_row, 5).value}")
            ed = _number(processed.cell(excel_row, 5).value, label="ED", excel_row=excel_row)
            ed_errors.append(ed - lp / (ss * hd))
            replicate_columns = {
                1: (8, 13, 18, 23, 28),
                2: (9, 14, 19, 24, 29),
                3: (10, 15, 20, 25, 30),
            }
            replicates: dict[int, dict[str, float]] = {}
            for replicate, columns in replicate_columns.items():
                load25_col, load65_col, sag_col, hlr_col, weight_col = columns
                replicates[replicate] = {
                    "load25": _number(processed.cell(excel_row, load25_col).value, label="Load25", excel_row=excel_row),
                    "load65": _number(processed.cell(excel_row, load65_col).value, label="Load65", excel_row=excel_row),
                    "sag": _number(processed.cell(excel_row, sag_col).value, label="SAG", excel_row=excel_row),
                    "hlr": _number(processed.cell(excel_row, hlr_col).value, label="HLR", excel_row=excel_row),
                    "weight": _number(processed.cell(excel_row, weight_col).value, label="Weight", excel_row=excel_row),
                }
            for cols, average_col in (
                ((8, 9, 10), 11),
                ((13, 14, 15), 16),
                ((18, 19, 20), 21),
                ((23, 24, 25), 26),
                ((28, 29, 30), 31),
            ):
                observed_average = _number(processed.cell(excel_row, average_col).value, label="平均值", excel_row=excel_row)
                source_values = [_number(processed.cell(excel_row, col).value, label="重复值", excel_row=excel_row) for col in cols]
                average_errors.append(observed_average - statistics.mean(source_values))
            for source_col, duplicate_col in ((8, 36), (9, 37), (10, 38), (13, 40), (14, 41), (15, 42)):
                duplicate_load_errors.append(
                    _number(processed.cell(excel_row, duplicate_col).value, label="重复载荷列", excel_row=excel_row)
                    - _number(processed.cell(excel_row, source_col).value, label="载荷列", excel_row=excel_row)
                )
            processed_rows[key] = {
                "condition_no": condition_no,
                "layer_thickness": lt,
                "ed_areal": ed,
                "replicates": replicates,
            }
            condition_ids[key] = f"condition_{condition_no:02d}"

        if len(processed_rows) != EXPECTED_CONDITIONS:
            raise AuditBlocked("Processed_Metrics 工况数漂移")

        scalar_rows: list[dict[str, object]] = []
        raw_processed_errors: list[float] = []
        sag_errors: list[float] = []
        output_vectors: list[tuple[float, float, float, float]] = []
        specimen_ids: set[int] = set()
        condition_replicates: dict[str, set[int]] = defaultdict(set)
        set_counts: Counter[int] = Counter()
        endpoint_definitions = (
            ("load25", "compressive_load_at_25_percent_deflection", "N", "closed", 0.35, "admitted_reference"),
            ("load65", "compressive_load_at_65_percent_deflection", "N", "closed", 0.35, "admitted_reference"),
            ("sag", "sag_factor", "1", "closed", 0.15, "admitted_reference"),
            ("hlr", "hysteresis_loss_ratio", "1", "closed", 0.35, "admitted_reference"),
            ("weight", "specimen_weight", "", "unresolved", 0.10, "conditional_reference"),
        )
        raw_column_by_key = {"load25": "E", "load65": "F", "sag": "G", "hlr": "H"}
        processed_weight_column = {1: "AB", 2: "AC", 3: "AD"}

        for excel_row in range(2, 77):
            row = [raw.cell(excel_row, column).value for column in range(1, 10)]
            if any(value is None for value in row):
                raise AuditBlocked(f"L25_OA table 第{excel_row}行存在缺字段：{row}")
            specimen_no = int(_number(row[0], label="Specimen", excel_row=excel_row))
            lp = _number(row[1], label="LP", excel_row=excel_row)
            ss = _number(row[2], label="SS", excel_row=excel_row)
            hd = _number(row[3], label="HD", excel_row=excel_row)
            replicate = int(_number(row[8], label="set", excel_row=excel_row))
            if specimen_no in specimen_ids or replicate not in (1, 2, 3):
                raise AuditBlocked(f"试件ID重复或set非法：第{excel_row}行")
            specimen_ids.add(specimen_no)
            set_counts[replicate] += 1
            key = (lp, ss, hd)
            if key not in processed_rows:
                raise AuditBlocked(f"原始表工况无法映射到Processed_Metrics：{key}")
            processed_record = processed_rows[key]
            condition_id = condition_ids[key]
            condition_replicates[condition_id].add(replicate)
            source_values = {
                "load25": _number(row[4], label="Load25", excel_row=excel_row),
                "load65": _number(row[5], label="Load65", excel_row=excel_row),
                "sag": _number(row[6], label="SAG", excel_row=excel_row),
                "hlr": _number(row[7], label="HLR", excel_row=excel_row),
                "weight": processed_record["replicates"][replicate]["weight"],
            }
            for endpoint in ("load25", "load65", "sag", "hlr"):
                raw_processed_errors.append(
                    source_values[endpoint]
                    - processed_record["replicates"][replicate][endpoint]
                )
            sag_errors.append(source_values["sag"] - source_values["load65"] / source_values["load25"])
            output_vectors.append(tuple(source_values[name] for name in ("load25", "load65", "sag", "hlr")))
            lt = float(processed_record["layer_thickness"])
            ed_areal = float(processed_record["ed_areal"])
            ed_volumetric = ed_areal / lt
            specimen_id = f"figshare_31550614_v1_specimen_{specimen_no:03d}"
            repeat_group = f"doi:{DOI}|{condition_id}"
            material_group = f"doi:{DOI}|commercial_tpu_single_lattice_system"
            for endpoint, observable, unit, unit_status, weight_ceiling, admission in endpoint_definitions:
                location = (
                    f"原始数据.xlsx#{RAW_SHEET}!{raw_column_by_key[endpoint]}{excel_row}"
                    if endpoint != "weight"
                    else f"原始数据.xlsx#{PROCESSED_SHEET}!{processed_weight_column[replicate]}{int(processed_record['condition_no']) + 1}"
                )
                notes = (
                    "Weight原工作簿未声明单位，仅作同源条件参考，不跨来源换算。"
                    if endpoint == "weight"
                    else (
                        "SAG与Load@65%/Load@25%高度一致但并非逐行完全相等，保留作者原值并在独立端点计数中降权。"
                        if endpoint == "sag"
                        else "逐试件实验端点；同一工况的三个重复必须同折。"
                    )
                )
                scalar_rows.append(
                    {
                        "scalar_id": f"{specimen_id}_{endpoint}",
                        "specimen_id": specimen_id,
                        "source_row": excel_row,
                        "source_location": location,
                        "condition_id": condition_id,
                        "replicate_index": replicate,
                        "split_group": repeat_group,
                        "repeat_leakage_group": repeat_group,
                        "material_geometry_leakage_group": material_group,
                        "observable": observable,
                        "value": source_values[endpoint],
                        "unit": unit,
                        "unit_status": unit_status,
                        "target_origin": "experimental",
                        "data_origin": "experimental_primary_workbook",
                        "record_granularity": "specimen_endpoint",
                        "quality_gate": admission,
                        "gold_layer": "Gold-E",
                        "gold_admission_status": admission,
                        "future_weight_ceiling": weight_ceiling,
                        "laser_power_w": lp,
                        "scan_speed_mm_s": ss,
                        "hatch_distance_mm": hd,
                        "layer_thickness_mm": lt,
                        "energy_density_areal_j_mm2": ed_areal,
                        "energy_density_volumetric_j_mm3": ed_volumetric,
                        "energy_density_areal_provenance": "official_workbook_formula_LP_over_SS_HD",
                        "energy_density_volumetric_provenance": "derived_areal_over_fixed_LT_not_direct_observation",
                        "chemistry_resolution": "commercial_TPU_identity_unresolved",
                        "geometry_resolution": "single_lattice_system_shared_by_source",
                        "notes": notes,
                    }
                )
    finally:
        formulas.close()
        values.close()

    if specimen_ids != set(range(1, EXPECTED_SPECIMENS + 1)):
        raise AuditBlocked("试件ID并非1至75完整序列")
    if set_counts != Counter({1: 25, 2: 25, 3: 25}):
        raise AuditBlocked(f"set计数漂移：{dict(set_counts)}")
    if len(scalar_rows) != EXPECTED_SCALARS:
        raise AuditBlocked(f"标量记录数漂移：{len(scalar_rows)}")
    if any(replicates != {1, 2, 3} for replicates in condition_replicates.values()):
        raise AuditBlocked("至少一个工况未包含完整的三个重复")

    unit_closed = sum(row["unit_status"] == "closed" for row in scalar_rows)
    unit_unresolved = sum(row["unit_status"] == "unresolved" for row in scalar_rows)
    if (unit_closed, unit_unresolved) != (EXPECTED_UNIT_CLOSED, EXPECTED_UNIT_UNRESOLVED):
        raise AuditBlocked("单位闭合计数漂移")

    property_stats: dict[str, dict[str, object]] = {}
    for observable in sorted({row["observable"] for row in scalar_rows}):
        endpoint_values = [float(row["value"]) for row in scalar_rows if row["observable"] == observable]
        property_stats[observable] = {
            "count": len(endpoint_values),
            "min": min(endpoint_values),
            "max": max(endpoint_values),
            "unique": len(set(endpoint_values)),
        }

    summary = {
        "audit_version": AUDIT_VERSION,
        "source": {
            "title": "Dataset and Supplementary Files for Influence of SLS Process Parameters on Cushion Properties and Sustainability of TPU Lattice Structures for Automotive Seating",
            "doi": DOI,
            "version": 1,
            "url": SOURCE_URL,
            "license": LICENSE,
            "source_reliability": "R1",
        },
        "counts": {
            "independent_specimens": EXPECTED_SPECIMENS,
            "process_conditions": EXPECTED_CONDITIONS,
            "replicates_per_condition": EXPECTED_REPLICATES,
            "scalar_records": len(scalar_rows),
            "unit_closed_scalar_records": unit_closed,
            "unit_unresolved_scalar_records": unit_unresolved,
            "admitted_reference_scalar_records": sum(row["gold_admission_status"] == "admitted_reference" for row in scalar_rows),
            "conditional_reference_scalar_records": sum(row["gold_admission_status"] == "conditional_reference" for row in scalar_rows),
            "nonredundant_unit_closed_scalar_records": EXPECTED_SPECIMENS * 3,
            "repeat_leakage_groups": len({row["repeat_leakage_group"] for row in scalar_rows}),
            "material_geometry_leakage_groups": len({row["material_geometry_leakage_group"] for row in scalar_rows}),
            "resolved_formulations": 0,
            "base_material_systems": 1,
            "lattice_topology_geometry_systems": 1,
        },
        "protocol_evidence": verify_protocol(),
        "property_statistics": property_stats,
        "energy_density_semantics": {
            "official_formula": "LP/(SS*HD)",
            "official_formula_dimension": "areal",
            "official_unit": "J/mm^2",
            "supplement_calls_formula_volumetric": True,
            "supplement_wording_conflict": True,
            "derived_volumetric_formula": "LP/(SS*HD*LT)",
            "derived_volumetric_unit": "J/mm^3",
            "fixed_layer_thickness_mm": 0.1,
            "derived_volumetric_values_are_observations": False,
        },
        "sag_semantics": {
            "near_derived_formula": "Load@65%/Load@25%",
            "author_values_preserved_without_recalculation": True,
            "max_abs_difference_from_ratio": _maximum_abs(sag_errors),
            "counted_as_independent_endpoint_for_effective_sample_size": False,
        },
        "reconciliation_checks": {
            "max_abs_raw_processed_endpoint_error": _maximum_abs(raw_processed_errors),
            "max_abs_processed_average_error": _maximum_abs(average_errors),
            "max_abs_sag_ratio_error": _maximum_abs(sag_errors),
            "max_abs_areal_energy_density_error": _maximum_abs(ed_errors),
            "max_abs_duplicate_load_column_error": _maximum_abs(duplicate_load_errors),
            "exact_duplicate_four_endpoint_vectors": len(output_vectors) - len(set(output_vectors)),
        },
        "scientific_classification": {
            "gold_layer": "Gold-E",
            "gold_admission_status": "mixed_by_endpoint",
            "scientific_role": "SLS-TPU工艺-缓冲性能迁移与过程模型参考",
            "direct_process_property_supervision": True,
            "direct_chemistry_property_supervision": False,
            "maximum_future_weight": 0.35,
            "minimum_split_unit": "process_condition",
            "strongest_novel_material_split_unit": "shared_commercial_TPU_and_lattice_system",
        },
        "limitations": [
            "工作簿没有公开商业TPU牌号、单体结构、配方、SMILES、分子量或硬段含量。",
            "75个试件共享一个商业TPU和一个晶格体系，不能当作75种独立材料。",
            "同一工况的三个重复必须同折；做新材料外推时整个来源应作为一个材料-几何泄漏组。",
            "Weight有75个逐试件数值但没有声明单位，仅进入条件参考层。",
            "SAG与Load@65%/Load@25%高度一致但有最高约0.00258的非零差；保留作者原值，375条记录中按近确定性派生口径仅225条是单位闭合且非冗余端点。",
            "补充材料另述7个扫描速度乘2重复的14个验证试件，但没有公开逐试件端点，因此未并入清单。",
            "官方补充材料对能量密度的volumetric称谓与LP/(SS*HD)、J/mm²不一致；脚本并列保存面/体定义。",
        ],
    }
    return scalar_rows, summary


def _tsv(rows: list[dict[str, object]], columns: tuple[str, ...]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=columns, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in columns})
    return stream.getvalue().encode("utf-8")


def _json(value: object) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8")


def atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def run_audit(*, write_outputs: bool = True) -> dict[str, object]:
    files = verify_files()
    scalar_rows, summary = parse_workbook()
    outputs = {
        "内容审计摘要.json": _json(summary),
        "标量审计清单.tsv": _tsv(scalar_rows, SCALAR_COLUMNS),
        "文件校验清单.tsv": _tsv(
            files, ("file", "bytes", "sha256", "md5", "verification")
        ),
    }
    if write_outputs:
        for name, payload in outputs.items():
            atomic_write(SOURCE_DIR / name, payload)
    return {"summary": summary, "scalars": scalar_rows, "files": files, "outputs": outputs}


if __name__ == "__main__":
    result = run_audit(write_outputs=True)
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
