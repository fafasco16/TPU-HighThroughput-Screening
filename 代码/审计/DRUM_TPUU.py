"""对两套 DRUM TPUU 原始数据执行可复现的只读内容审计。

本脚本以自身路径推导项目根目录，只读取 ``数据/原始`` 下两个既定来源。
唯一允许写入/覆盖的文件，是每个来源根目录内的以下四个审计产物：

* ``内容审计摘要.json``
* ``文件校验清单.tsv``
* ``曲线审计清单.tsv``
* ``批次审计清单.tsv``

原始 ZIP、解包文件、README 与工作簿均以只读方式访问；低天花板来源的旧版
``.xls`` DMTA 文件由同目录 ``读取低天花板DMTA.ps1`` 通过 Excel COM 只读解析。
运行 ``python DRUM_TPUU.py --check-layout`` 可只检查布局和错误门，不生成产物。
四个审计产物使用固定审计规范基准日，避免连续复跑产生无意义的时间戳差异。
"""

from __future__ import annotations

import argparse
import binascii
import csv
import hashlib
import io
import json
import math
import os
import re
import subprocess
import tempfile
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook

from 审计.第十批ACS表格物化 import RECORD_COLUMNS


SCRIPT_PATH = Path(__file__).resolve()
SCRIPT_DIR = SCRIPT_PATH.parent
PROJECT_ROOT = SCRIPT_PATH.parents[2]
BASE = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
MECH_DIR = BASE / "DRUM_TPUU_机械回收"
LOW_DIR = BASE / "DRUM_TPUU_低天花板"
PS_DMTA = SCRIPT_DIR / "读取低天花板DMTA.ps1"

# 本项目首次正式DRUM审计的规范基准日。它是审计协议版本的一部分，不表示运行时刻。
AUDIT_BASELINE_UTC = "2026-07-20T00:00:00+00:00"

AUDIT_OUTPUTS = {
    "内容审计摘要.json",
    "文件校验清单.tsv",
    "曲线审计清单.tsv",
    "批次审计清单.tsv",
}

EXPECTED_CONTAINERS = {
    "机械回收": {
        "根目录": {
            "RAW_Chemical_recycling.zip",
            "RAW_P4PrCL_kinetics_and_Me.zip",
            "RAW_Polyol_characterization.zip",
            "RAW_Thermodynamics.zip",
            "RAW_TPUU_characterization.zip",
        },
        "解包根目录": set(),
    },
    "低天花板": {
        "根目录": {"PrimaryData.zip"},
        "解包根目录": {
            "Chemdraws.zip",
            "Copolymerization_Code.zip",
            "Raw_DSC.zip",
            "Raw_FTIR.zip",
            "Raw_GCMS.zip",
            "Raw_Mechanical_Testing.zip",
            "Raw_NMR.zip",
            "Raw_SEC.zip",
            "Raw_TGA.zip",
        },
    },
}

FILE_COLUMNS = [
    "记录类型",
    "来源ID",
    "容器相对路径",
    "容器层级",
    "容器字节数",
    "容器SHA256",
    "容器CRC整体状态",
    "成员相对路径",
    "成员字节数",
    "成员CRC32声明",
    "成员CRC32实算",
    "成员SHA256",
    "解包相对路径",
    "解包存在",
    "解包字节数",
    "解包SHA256",
    "解包匹配状态",
    "扩展名",
    "数据模态",
    "是否机械候选",
    "路径安全",
    "成员名重复数",
    "内容重复组大小",
    "备注",
]

CURVE_COLUMNS = [
    "来源ID",
    "DOI",
    "文件相对路径",
    "工作表",
    "材料代码",
    "配方键",
    "生命周期状态",
    "批次键",
    "批次键依据",
    "试样键",
    "试样标签",
    "试验类型",
    "曲线层级",
    "保真度",
    "模型准入层",
    "自变量",
    "自变量单位",
    "因变量",
    "因变量单位",
    "点数",
    "数据行数",
    "数值单元数",
    "主轴缺失行数",
    "内部空行数",
    "次要通道缺失数",
    "循环数",
    "期望重复数",
    "曲线SHA256",
    "曲线精确重复组大小",
    "自变量最小值",
    "自变量最大值",
    "因变量最小值",
    "因变量最大值",
    "几何参数可用",
    "试验速度",
    "方法依据",
    "泄漏分组键",
    "准入结论",
    "排除或降权原因",
    "备注",
]

BATCH_COLUMNS = [
    "来源ID",
    "DOI",
    "批次键",
    "批次键依据",
    "材料代码",
    "配方键",
    "生命周期状态",
    "模型准入层",
    "试验类型",
    "曲线数",
    "独立试样数",
    "曲线点数",
    "数值单元数",
    "泄漏分组键",
    "备注",
]


DATASETS = {
    "机械回收": {
        "dir": MECH_DIR,
        "source_id": "DRUM_TPUU_机械回收_10.13020_05ek-6k60",
        "doi": "10.13020/05ek-6k60",
        "title": "Data for Alkyl Substituted Polycaprolactone Poly(Urethane-Urea)s as Mechanically-Competitive and Chemically-Recyclable Materials",
        "landing": "https://hdl.handle.net/11299/264123",
        "dataset_citation": "Pfau-Cloud, M. R.; Batiste, D. C.; Kim, H. J.; Ellison, C. J.; Hillmyer, M. A. (2024). Supporting data for Alkyl Substituted Polycaprolactone Poly(Urethane-Urea)s as Mechanically-Competitive and Chemically-Recyclable Materials. Data Repository for the University of Minnesota (DRUM). https://doi.org/10.13020/05ek-6k60",
        "article_citation": "Batiste, D. C.; Pfau-Cloud, M. R.; Kim, H. J.; Ellison, C. J.; Hillmyer, M. A. Alkyl-Substituted Polycaprolactone Poly(urethane-urea)s as Mechanically Competitive and Chemically Recyclable Materials. ACS Macro Letters 2024, 13 (11), 1449–1455. https://doi.org/10.1021/acsmacrolett.4c00474",
        "license": "CC0 1.0 Universal",
        "license_url": "https://creativecommons.org/publicdomain/zero/1.0/legalcode",
    },
    "低天花板": {
        "dir": LOW_DIR,
        "source_id": "DRUM_TPUU_低天花板_10.13020_zf53-w893",
        "doi": "10.13020/zf53-w893",
        "title": "Supporting Information for Tackling the thermodynamic stability of low-ceiling temperature polymers in the preparation of tough and chemically recyclable thermoplastic polyurethane-urea elastomers",
        "landing": "https://conservancy.umn.edu/handle/11299/263305",
        "dataset_citation": "Meyersohn, M. S.; Block, A.; Bates, F. S.; Hillmyer, M. A. (2024). Supporting Information for Tackling the thermodynamic stability of low-ceiling temperature polymers in the preparation of tough and chemically recyclable thermoplastic polyurethane-urea elastomers. University Digital Conservancy / DRUM. https://doi.org/10.13020/zf53-w893",
        "article_citation": "Meyersohn, M. S.; Block, A.; Bates, F. S.; Hillmyer, M. A. Tackling the Thermodynamic Stability of Low-Ceiling Temperature Polymers for the Preparation of Tough and Chemically Recyclable Thermoplastic Polyurethane-Urea Elastomers. Macromolecules 2024, 57 (19), 9230–9240. https://doi.org/10.1021/acs.macromol.4c01431",
        "license": "CC0 1.0 Universal",
        "license_url": "https://creativecommons.org/publicdomain/zero/1.0/legalcode",
    },
}


LOW_GOLD_E_OPTIONAL_COLUMNS = (
    "batch_id",
    "curve_id",
    "point_index",
    "secondary_condition_name",
    "secondary_condition_value",
    "secondary_condition_unit",
    "auxiliary_value_name",
    "auxiliary_value",
    "auxiliary_unit",
    "sample_identity_status",
    "global_structure_family_key",
    "family_leakage_group",
    "curve_points_are_independent_samples",
    "duplicate_status",
)
LOW_MECHANICAL_DIR = LOW_DIR / "解包内容" / "Raw_Mechanical_Testing"
LOW_CURVE_AUDIT = LOW_DIR / "曲线审计清单.tsv"
LOW_MATERIALIZED_FILE_SPECS = {
    "TPUU-C_hysteresis.csv": (
        1_293_131,
        "3a589865feb3248ba11d767fe1a345fecd24e4ffa24a7e194d434bfc3ae38ad8",
    ),
    "TPUU-C_tensile.csv": (
        56_143,
        "a3e6049dbba7def5fd3f37e4d8d4045b977a818ed567893f97dadefc222d770c",
    ),
    "TPUU-D_hysteresis.csv": (
        1_616_402,
        "0dde58a3db8dc663259d282feb1193b8987c2faee65d662a687db88cf6d8599f",
    ),
    "TPUU-D_tensile.csv": (
        54_537,
        "84fa631b586911c6fb2a7b53b9a2c7147615ef05e6389db7148bf57cbcec7137",
    ),
    "TPUU-R_hysteresis.csv": (
        1_485_410,
        "3ec1625d45bfc9f025ba1b70cb4b62878c0402d27f581d1e5e7d5135e1eca44d",
    ),
    "TPUU-R_tensile.csv": (
        46_380,
        "83fb582175300bdd21d8a440d1f96f24f21cf834ab830d100497fb6c0b44bce2",
    ),
    "TPUU-S_hysteresis.csv": (
        1_179_193,
        "d6dfd0840017caaf8ffce4a3aba4cd840aabffd4a5591dc685455c0074e4a30c",
    ),
    "TPUU-S_tensile.csv": (
        63_079,
        "8d4c3816735df5342f30bafc675506c6aa5f9557df6fd8843662eae312fc45e6",
    ),
}
LOW_EXPECTED_TENSILE_CURVES = 20
LOW_EXPECTED_CYCLIC_CURVES = 4
LOW_EXPECTED_TENSILE_POINTS = 4_369
LOW_EXPECTED_CYCLIC_POINTS = 105_912
LOW_EXPECTED_RAW_POINTS = 110_281
LOW_EXPECTED_DERIVED_SCALARS = 60


def _is_relative_to(path: Path, parent: Path) -> bool:
    """兼容性明确的路径包含判断，不依赖字符串前缀。"""

    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def audit_output_path(root: Path, name: str) -> Path:
    """返回唯一允许写入的审计文件路径，并拒绝越界或符号链接。"""

    if name not in AUDIT_OUTPUTS:
        raise RuntimeError(f"拒绝写入未授权文件: {name!r}")
    if root.is_symlink() or (hasattr(root, "is_junction") and root.is_junction()):
        raise RuntimeError(f"拒绝通过符号链接或联接点写入DRUM来源目录: {root}")
    resolved_root = root.resolve(strict=True)
    allowed_roots = {Path(meta["dir"]).resolve(strict=True) for meta in DATASETS.values()}
    if resolved_root not in allowed_roots:
        raise RuntimeError(f"拒绝写入非DRUM来源目录: {resolved_root}")
    candidate = root / name
    if candidate.is_symlink() or (
        hasattr(candidate, "is_junction") and candidate.is_junction()
    ):
        raise RuntimeError(f"拒绝覆盖符号链接或联接点审计文件: {candidate}")
    if candidate.exists() and not candidate.is_file():
        raise RuntimeError(f"审计输出目标不是普通文件: {candidate}")
    if candidate.parent.resolve(strict=True) != resolved_root:
        raise RuntimeError(f"审计输出路径越界: {candidate}")
    return candidate


def validate_runtime_layout(selected: list[str]) -> dict[str, Any]:
    """在任何写入前验证项目、来源容器、解包目录与输出边界。"""

    if SCRIPT_DIR.name != "审计" or SCRIPT_DIR.parent.name != "代码":
        raise RuntimeError(f"脚本必须位于 <项目根>/代码/审计，当前为: {SCRIPT_PATH}")
    if not (PROJECT_ROOT / "pyproject.toml").is_file():
        raise RuntimeError(f"无法确认项目根目录（缺少pyproject.toml）: {PROJECT_ROOT}")
    if not PS_DMTA.is_file():
        raise RuntimeError(f"缺少同目录XLS只读解析器: {PS_DMTA}")
    if not BASE.is_dir():
        raise RuntimeError(f"缺少DRUM来源基目录: {BASE}")

    resolved_project = PROJECT_ROOT.resolve(strict=True)
    resolved_base = BASE.resolve(strict=True)
    if not _is_relative_to(resolved_base, resolved_project):
        raise RuntimeError(f"原始数据目录越出项目根: {resolved_base}")

    report: dict[str, Any] = {
        "脚本": str(SCRIPT_PATH),
        "项目根": str(PROJECT_ROOT),
        "XLS解析器": str(PS_DMTA),
        "审计规范基准日UTC": AUDIT_BASELINE_UTC,
        "数据集": {},
        "允许覆盖输出": sorted(AUDIT_OUTPUTS),
    }
    for key in selected:
        root = Path(DATASETS[key]["dir"])
        if not root.is_dir():
            raise RuntimeError(f"[{key}] 缺少来源目录: {root}")
        resolved_root = root.resolve(strict=True)
        if not _is_relative_to(resolved_root, resolved_base):
            raise RuntimeError(f"[{key}] 来源目录越出DRUM基目录: {resolved_root}")
        unpacked = root / "解包内容"
        if not unpacked.is_dir():
            raise RuntimeError(f"[{key}] 缺少解包内容目录: {unpacked}")

        actual_root_zips = {path.name for path in root.glob("*.zip") if path.is_file()}
        actual_unpacked_zips = {path.name for path in unpacked.glob("*.zip") if path.is_file()}
        expected = EXPECTED_CONTAINERS[key]
        if actual_root_zips != expected["根目录"]:
            raise RuntimeError(
                f"[{key}] 根目录ZIP集合不符；期望={sorted(expected['根目录'])}，"
                f"实际={sorted(actual_root_zips)}"
            )
        if actual_unpacked_zips != expected["解包根目录"]:
            raise RuntimeError(
                f"[{key}] 解包根目录ZIP集合不符；期望={sorted(expected['解包根目录'])}，"
                f"实际={sorted(actual_unpacked_zips)}"
            )

        for output_name in AUDIT_OUTPUTS:
            audit_output_path(root, output_name)
        report["数据集"][key] = {
            "来源目录": str(root),
            "根目录ZIP数": len(actual_root_zips),
            "内嵌ZIP数": len(actual_unpacked_zips),
            "解包文件数": sum(path.is_file() for path in unpacked.rglob("*")),
        }
    return report


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """解析命令行；数据集为空时默认审计两套来源。"""

    parser = argparse.ArgumentParser(
        description="只读审计两套DRUM TPUU来源，并仅覆盖四个固定JSON/TSV审计产物。",
    )
    parser.add_argument(
        "datasets",
        nargs="*",
        metavar="数据集",
        help="可选：机械回收、低天花板；省略时两者都审计。",
    )
    parser.add_argument(
        "--check-layout",
        "--检查环境",
        action="store_true",
        help="只检查项目布局、输入容器和输出边界；不解析数据、不写文件。",
    )
    args = parser.parse_args(argv)
    selected = args.datasets or list(DATASETS)
    unknown = [key for key in selected if key not in DATASETS]
    if unknown:
        parser.error(f"未知数据集: {unknown}; 可选: {list(DATASETS)}")
    if len(selected) != len(set(selected)):
        parser.error("同一数据集不能重复指定")
    args.datasets = selected
    return args


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def hash_zip_member(zf: zipfile.ZipFile, info: zipfile.ZipInfo) -> tuple[str, str]:
    digest = hashlib.sha256()
    crc = 0
    with zf.open(info, "r") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
            crc = binascii.crc32(chunk, crc)
    return digest.hexdigest(), f"{crc & 0xFFFFFFFF:08x}"


def path_is_safe(name: str) -> bool:
    normalized = name.replace("\\", "/")
    pure = PurePosixPath(normalized)
    return not pure.is_absolute() and ".." not in pure.parts and not re.match(r"^[A-Za-z]:", normalized)


def extension_of(path: str) -> str:
    suffix = PurePosixPath(path.replace("\\", "/")).suffix
    return suffix.lower() if suffix else "[无扩展名]"


def modality_for(path: str) -> str:
    p = path.replace("\\", "/").lower()
    name = PurePosixPath(p).name
    if name.endswith(".zip"):
        return "压缩容器"
    if "mechanical" in p or any(k in name for k in ("tensile", "hysteresis", "dmta")):
        return "机械/动态力学"
    if "/raw_nmr/" in f"/{p}" or "nmr" in name:
        return "核磁共振"
    if "/raw_dsc/" in f"/{p}" or "dsc" in name or "calorim" in name or "combustion" in name:
        return "量热/DSC"
    if "/raw_tga/" in f"/{p}" or "tga" in name or "compost" in name:
        return "热重/降解"
    if "/raw_ftir/" in f"/{p}" or any(k in name for k in ("ftir", "_ir", "ir_")):
        return "红外光谱"
    if "/raw_sec/" in f"/{p}" or any(k in name for k in ("sec", "gpc")):
        return "SEC/GPC摩尔质量"
    if "/raw_gcms/" in f"/{p}" or any(k in name for k in ("gcms", "gc_ms", "_ms")):
        return "GC-MS/质谱"
    if "saxs" in name or "waxs" in name:
        return "X射线散射"
    if "kinetic" in p or "copolymerization_code" in p or extension_of(name) in {".in", ".out", ".dat", ".matrix"}:
        return "聚合动力学/模型代码"
    if "thermodynamic" in p or "density" in name or "yield" in name:
        return "热力学/密度/收率"
    if extension_of(name) in {".cdxml", ".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
        return "结构图/实验照片"
    if "recycl" in p:
        return "化学回收表征"
    return "其他化学/材料表征"


def container_specs(key: str, root: Path) -> list[tuple[Path, str, Path]]:
    if key == "机械回收":
        return [(p, "顶层数据包", root / "解包内容") for p in sorted(root.glob("*.zip"))]
    specs = [(root / "PrimaryData.zip", "顶层数据包", root / "解包内容")]
    specs.extend((p, "内嵌分模态数据包", root / "解包内容") for p in sorted((root / "解包内容").glob("*.zip")))
    return specs


def audit_containers(key: str, meta: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    root: Path = meta["dir"]
    source_id = meta["source_id"]
    rows: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    mapped_paths: set[Path] = set()
    extracted_hashes: dict[Path, str] = {}

    for container, level, map_root in container_specs(key, root):
        container_rel = container.relative_to(root).as_posix()
        container_sha = sha256_file(container)
        container_size = container.stat().st_size
        with zipfile.ZipFile(container, "r") as zf:
            bad_member = zf.testzip()
            infos = [info for info in zf.infolist() if not info.is_dir()]
            name_counts = Counter(info.filename.replace("\\", "/") for info in infos)
            unsafe_count = sum(not path_is_safe(info.filename) for info in infos)
            encrypted_count = sum(bool(info.flag_bits & 0x1) for info in infos)
            matched = 0
            missing = 0
            mismatched = 0
            rows.append(
                {
                    "记录类型": "容器",
                    "来源ID": source_id,
                    "容器相对路径": container_rel,
                    "容器层级": level,
                    "容器字节数": container_size,
                    "容器SHA256": container_sha,
                    "容器CRC整体状态": "通过" if bad_member is None else f"失败:{bad_member}",
                    "成员相对路径": "",
                    "成员字节数": sum(info.file_size for info in infos),
                    "成员CRC32声明": "",
                    "成员CRC32实算": "",
                    "成员SHA256": "",
                    "解包相对路径": "",
                    "解包存在": "",
                    "解包字节数": "",
                    "解包SHA256": "",
                    "解包匹配状态": "",
                    "扩展名": ".zip",
                    "数据模态": "压缩容器",
                    "是否机械候选": "否",
                    "路径安全": "是" if unsafe_count == 0 else "否",
                    "成员名重复数": sum(v - 1 for v in name_counts.values() if v > 1),
                    "内容重复组大小": "",
                    "备注": f"成员文件{len(infos)}个；加密成员{encrypted_count}个",
                }
            )
            for info in infos:
                member_name = info.filename.replace("\\", "/")
                member_sha, actual_crc = hash_zip_member(zf, info)
                extracted = map_root.joinpath(*PurePosixPath(member_name).parts)
                exists = extracted.is_file()
                extracted_rel = extracted.relative_to(root).as_posix()
                extracted_size: int | str = ""
                extracted_sha = ""
                if exists:
                    mapped_paths.add(extracted.resolve())
                    extracted_size = extracted.stat().st_size
                    extracted_sha = extracted_hashes.get(extracted)
                    if not extracted_sha:
                        extracted_sha = sha256_file(extracted)
                        extracted_hashes[extracted] = extracted_sha
                    status = "完全一致" if extracted_size == info.file_size and extracted_sha == member_sha else "不一致"
                    if status == "完全一致":
                        matched += 1
                    else:
                        mismatched += 1
                else:
                    status = "缺失"
                    missing += 1
                modality = modality_for(member_name)
                rows.append(
                    {
                        "记录类型": "成员文件",
                        "来源ID": source_id,
                        "容器相对路径": container_rel,
                        "容器层级": level,
                        "容器字节数": container_size,
                        "容器SHA256": container_sha,
                        "容器CRC整体状态": "通过" if bad_member is None else f"失败:{bad_member}",
                        "成员相对路径": member_name,
                        "成员字节数": info.file_size,
                        "成员CRC32声明": f"{info.CRC:08x}",
                        "成员CRC32实算": actual_crc,
                        "成员SHA256": member_sha,
                        "解包相对路径": extracted_rel,
                        "解包存在": "是" if exists else "否",
                        "解包字节数": extracted_size,
                        "解包SHA256": extracted_sha,
                        "解包匹配状态": status,
                        "扩展名": extension_of(member_name),
                        "数据模态": modality,
                        "是否机械候选": "是" if modality == "机械/动态力学" else "否",
                        "路径安全": "是" if path_is_safe(member_name) else "否",
                        "成员名重复数": name_counts[member_name],
                        "内容重复组大小": "",
                        "备注": "",
                    }
                )
            summaries.append(
                {
                    "容器": container_rel,
                    "层级": level,
                    "字节数": container_size,
                    "SHA256": container_sha,
                    "CRC测试": "通过" if bad_member is None else f"失败:{bad_member}",
                    "文件成员数": len(infos),
                    "目录成员数": sum(info.is_dir() for info in zf.infolist()),
                    "解压后字节数": sum(info.file_size for info in infos),
                    "映射完全一致数": matched,
                    "映射缺失数": missing,
                    "映射不一致数": mismatched,
                    "不安全路径数": unsafe_count,
                    "重复成员名额外数": sum(v - 1 for v in name_counts.values() if v > 1),
                    "加密成员数": encrypted_count,
                }
            )

    member_rows = [row for row in rows if row["记录类型"] == "成员文件" and row["解包SHA256"]]
    duplicate_groups = Counter(row["解包SHA256"] for row in member_rows)
    for row in member_rows:
        row["内容重复组大小"] = duplicate_groups[row["解包SHA256"]]

    sidecars = []
    for path in sorted(root.iterdir()):
        if not path.is_file() or path.suffix.lower() == ".zip" or path.name in AUDIT_OUTPUTS:
            continue
        digest = sha256_file(path)
        sidecars.append(path.name)
        rows.append(
            {
                "记录类型": "旁车文件",
                "来源ID": source_id,
                "容器相对路径": "",
                "容器层级": "容器外",
                "容器字节数": "",
                "容器SHA256": "",
                "容器CRC整体状态": "不适用",
                "成员相对路径": "",
                "成员字节数": "",
                "成员CRC32声明": "",
                "成员CRC32实算": "",
                "成员SHA256": "",
                "解包相对路径": path.name,
                "解包存在": "是",
                "解包字节数": path.stat().st_size,
                "解包SHA256": digest,
                "解包匹配状态": "容器外旁车",
                "扩展名": extension_of(path.name),
                "数据模态": "README/结构说明" if path.suffix.lower() == ".txt" else modality_for(path.name),
                "是否机械候选": "否",
                "路径安全": "是",
                "成员名重复数": "",
                "内容重复组大小": 1,
                "备注": "未包含于下载ZIP；保留为数据集旁车文件",
            }
        )

    extracted_root = root / "解包内容"
    extracted_files = [p for p in extracted_root.rglob("*") if p.is_file()]
    orphan = [p.relative_to(root).as_posix() for p in extracted_files if p.resolve() not in mapped_paths]
    inventory = {
        "解包文件数": len(extracted_files),
        "解包总字节数": sum(p.stat().st_size for p in extracted_files),
        "已映射解包文件数": len(mapped_paths),
        "未映射解包文件数": len(orphan),
        "未映射解包文件示例": orphan[:20],
        "容器外旁车文件": sidecars,
        "精确重复内容组数": sum(1 for count in duplicate_groups.values() if count > 1),
        "处于精确重复组的成员记录数": sum(count for count in duplicate_groups.values() if count > 1),
        "最大精确重复组大小": max(duplicate_groups.values(), default=1),
    }
    return rows, summaries, inventory


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        value = float(value)
        return value if math.isfinite(value) else None
    try:
        result = float(str(value).strip())
        return result if math.isfinite(result) else None
    except (TypeError, ValueError):
        return None


def fmt_num(value: float) -> str:
    return format(value, ".15g")


@dataclass
class CurveAccumulator:
    channel_count: int
    primary_x: int
    primary_y: int
    cycle_index: int | None = None
    point_count: int = 0
    data_rows: int = 0
    numeric_value_count: int = 0
    missing_primary_rows: int = 0
    secondary_missing_cells: int = 0
    internal_blank_rows: int = 0
    blank_run: int = 0
    seen_data: bool = False
    cycles: set[int] = field(default_factory=set)
    min_x: float | None = None
    max_x: float | None = None
    min_y: float | None = None
    max_y: float | None = None
    digest: Any = field(default_factory=hashlib.sha256)

    def add(self, values: Iterable[Any]) -> None:
        nums = [number(v) for v in values]
        nonempty = sum(v is not None for v in nums)
        if nonempty == 0:
            if self.seen_data:
                self.blank_run += 1
            return
        if self.seen_data and self.blank_run:
            self.internal_blank_rows += self.blank_run
            self.blank_run = 0
        self.seen_data = True
        self.data_rows += 1
        self.numeric_value_count += nonempty
        x = nums[self.primary_x]
        y = nums[self.primary_y]
        if x is not None and y is not None:
            self.point_count += 1
            self.min_x = x if self.min_x is None else min(self.min_x, x)
            self.max_x = x if self.max_x is None else max(self.max_x, x)
            self.min_y = y if self.min_y is None else min(self.min_y, y)
            self.max_y = y if self.max_y is None else max(self.max_y, y)
            self.digest.update(f"{fmt_num(x)},{fmt_num(y)}\n".encode("ascii"))
        else:
            self.missing_primary_rows += 1
        for idx, value in enumerate(nums):
            if idx not in (self.primary_x, self.primary_y) and value is None:
                self.secondary_missing_cells += 1
        if self.cycle_index is not None and nums[self.cycle_index] is not None:
            self.cycles.add(int(round(nums[self.cycle_index])))

    @property
    def curve_hash(self) -> str:
        return self.digest.hexdigest()


def batch_from_text(text: str) -> tuple[str, str]:
    value = clean(text)
    patterns = [
        r"(?<![A-Za-z0-9])MSM[-_ ]\d+[-_ ]\d+(?!\d)",
        r"(?<![A-Za-z0-9])MRP[-_ ]\d+[-_ ]\d+[A-Za-z]?(?![A-Za-z0-9])",
        r"(?<![A-Za-z0-9])DCB[-_ ]?\d+[-_ ]\d+[A-Za-z]?(?![A-Za-z0-9])",
    ]
    for pattern in patterns:
        match = re.search(pattern, value, re.I)
        if match:
            return re.sub(r"[ _]", "-", match.group(0).upper()), "仪器测试名/试样标签解析"
    return "", "未解析"


def lifecycle_for(filename: str, material: str) -> str:
    lower = filename.lower()
    if "repolymerized" in lower:
        return "回收单体再聚合"
    if "elastollan" in lower:
        return "商业对照"
    if "thermoset" in lower:
        return "热固性/橡胶桥接对照"
    if material in {"TPUU-R", "TPUU-S", "TPUU-C", "TPUU-D"}:
        return "论文代码态（C/D/R/S需由正文配方表解码）"
    return "原生实验材料"


def tier_for(filename: str, material: str, dataset_key: str) -> tuple[str, str, str]:
    lower = filename.lower()
    if dataset_key == "低天花板":
        return "核心实验层", "准入", "四个论文配方代码均为IPDI/水扩链TPUU；配方代码需先解码"
    if "elastollan" in lower:
        return "迁移/外部力学层", "降权准入", "商业牌号缺少完整可计算配方"
    if "comparison_to_14bdo" in lower:
        return "迁移/机理桥接层", "降权准入", "14BDO扩链TPU，不是水扩链TPUU"
    if "proof_of_concept" in lower:
        if "rubber" in material.lower():
            return "外域对照层", "排除核心训练", "商业橡皮筋且化学结构未知"
        return "迁移/机理桥接层", "降权准入", "热固性PU，不属于目标热塑性TPUU"
    return "核心实验层", "准入", "水扩链IPDI-TPUU，材料代码可映射软段Mn与硬段质量分数"


def canonical_formula(material: str, filename: str) -> str:
    value = clean(material)
    lower = filename.lower()
    if "elastollan" in lower:
        return "Elastollan-C60A10WH"
    value = value.replace("%", "HS")
    value = re.sub(r"\s+", " ", value).strip()
    aliases = {
        "P4MCL-1k-45HS": "P4MCL-1k-46HS",
        "PMCL-1k-44HS": "PMCL-1k-46HS",
        "P4MCL thermoset PU": "P4MCL Thermoset PU",
    }
    return aliases.get(value, value)


def make_curve_row(
    meta: dict[str, Any],
    dataset_key: str,
    file_path: Path,
    sheet: str,
    material: str,
    sample_label: str,
    test_type: str,
    acc: CurveAccumulator,
    x_field: str,
    x_unit: str,
    y_field: str,
    y_unit: str,
    batch_text: str = "",
    geometry: bool = False,
    speed: str = "",
    method: str = "",
    expected_replicates: int | str = "",
    notes: str = "",
) -> dict[str, Any]:
    filename = file_path.name
    formula = canonical_formula(material, filename)
    tier, decision, reason = tier_for(filename, material, dataset_key)
    raw_batch, basis = batch_from_text(batch_text or sample_label)
    if raw_batch:
        batch = f"{formula}|{raw_batch}"
    else:
        batch = f"{formula}|未解析批次"
        basis = "材料代码代理；待与实验记录/正文配方表连接"
    lifecycle = lifecycle_for(filename, material)
    specimen_key = f"{meta['doi']}|{batch}|{material}|{sample_label or test_type}"
    leakage_key = f"{meta['doi']}|{batch}"
    return {
        "来源ID": meta["source_id"],
        "DOI": meta["doi"],
        "文件相对路径": file_path.relative_to(meta["dir"]).as_posix(),
        "工作表": sheet,
        "材料代码": material,
        "配方键": formula,
        "生命周期状态": lifecycle,
        "批次键": batch,
        "批次键依据": basis,
        "试样键": specimen_key,
        "试样标签": sample_label,
        "试验类型": test_type,
        "曲线层级": "独立试样曲线" if test_type in {"单轴拉伸", "循环滞回"} and "Proof" not in filename and "Comparison" not in filename else "材料表征曲线",
        "保真度": "实验",
        "模型准入层": tier,
        "自变量": x_field,
        "自变量单位": x_unit,
        "因变量": y_field,
        "因变量单位": y_unit,
        "点数": acc.point_count,
        "数据行数": acc.data_rows,
        "数值单元数": acc.numeric_value_count,
        "主轴缺失行数": acc.missing_primary_rows,
        "内部空行数": acc.internal_blank_rows,
        "次要通道缺失数": acc.secondary_missing_cells,
        "循环数": len(acc.cycles) if acc.cycles else "",
        "期望重复数": expected_replicates,
        "曲线SHA256": acc.curve_hash,
        "曲线精确重复组大小": "",
        "自变量最小值": acc.min_x if acc.min_x is not None else "",
        "自变量最大值": acc.max_x if acc.max_x is not None else "",
        "因变量最小值": acc.min_y if acc.min_y is not None else "",
        "因变量最大值": acc.max_y if acc.max_y is not None else "",
        "几何参数可用": "是" if geometry else "否",
        "试验速度": speed,
        "方法依据": method,
        "泄漏分组键": leakage_key,
        "准入结论": decision,
        "排除或降权原因": reason,
        "备注": notes,
    }


def header_positions(row: tuple[Any, ...], expected: list[str]) -> list[int]:
    normalized = [clean(v).lower() for v in row]
    positions = []
    width = len(expected)
    for idx in range(0, len(normalized) - width + 1):
        if normalized[idx : idx + width] == [v.lower() for v in expected]:
            positions.append(idx)
    return positions


def metadata_value(row1: tuple[Any, ...], row2: tuple[Any, ...], label: str) -> str:
    for idx, value in enumerate(row1):
        if clean(value).lower() == label.lower() and idx < len(row2):
            return clean(row2[idx])
    return ""


def audit_xlsx_mechanical(meta: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    root = meta["dir"] / "解包内容"
    files = sorted(
        p
        for p in root.rglob("*.xlsx")
        if any(k in p.name.lower() for k in ("tensile", "hysteresis", "dmta"))
    )
    curves: list[dict[str, Any]] = []
    issues: list[str] = []
    method_tensile = "README: Shimadzu AGS-X；室温；50 mm/min；至少5个拉伸重复；模量拟合0–5%应变"
    method_hyst = "README: Shimadzu AGS-X；50%应变加载/卸载；每材料1个dogbone；通常10循环"
    method_dmta = "README: RSA-G2；约-90°C起始；5°C/min；6.28 rad/s (1 Hz)"

    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            iterator = ws.iter_rows(values_only=True)
            row1 = tuple(next(iterator, ()))
            row2 = tuple(next(iterator, ()))
            row3 = tuple(next(iterator, ()))
            lower_name = path.name.lower()
            if "hysteresis" in lower_name:
                if [clean(v).lower() for v in row2[:4]] != ["time", "force", "stroke", "cycle"]:
                    issues.append(f"{path.name}/{ws.title}: 未识别循环滞回四通道")
                    continue
                acc = CurveAccumulator(4, 0, 1, 3)
                geometry = any(clean(v).lower() == "thickness" for row in (row1, row2, row3) for v in row)
                for row in iterator:
                    acc.add(row[:4])
                test_name = metadata_value(row1, row2, "Test Name")
                speed = metadata_value(row1, row2, "Speed")
                curves.append(
                    make_curve_row(
                        meta,
                        "机械回收",
                        path,
                        ws.title,
                        ws.title,
                        "滞回试样1",
                        "循环滞回",
                        acc,
                        "时间",
                        "s",
                        "力",
                        "N",
                        batch_text=test_name,
                        geometry=geometry,
                        speed=speed,
                        method=method_hyst,
                        expected_replicates=1,
                        notes="原始表直接给Time/Force/Stroke/Cycle；应力-应变需用试样几何转换",
                    )
                )
                continue

            if "dmta" in lower_name and ws.title.lower() == "dmta" or path.name == "TPUU_DMTA.xlsx":
                positions = header_positions(row2, ["Temperature (～C)", "E' (MPa)", "tan 汛", 'E" (MPa)'])
                if not positions:
                    positions = [idx for idx in range(0, len(row1), 4) if clean(row1[idx])]
                accumulators = {pos: CurveAccumulator(4, 0, 1) for pos in positions}
                for row in iterator:
                    for pos, acc in accumulators.items():
                        acc.add(row[pos : pos + 4])
                for pos, acc in accumulators.items():
                    material = clean(row1[pos]) or ws.title
                    curves.append(
                        make_curve_row(
                            meta,
                            "机械回收",
                            path,
                            ws.title,
                            material,
                            "DMTA试样1",
                            "DMTA温度扫描",
                            acc,
                            "温度",
                            "°C",
                            "储能模量E'",
                            "MPa",
                            geometry=False,
                            method=method_dmta,
                            expected_replicates=1,
                            notes="同时含tanδ与损耗模量E''",
                        )
                    )
                continue

            positions5 = header_positions(row2, ["Time", "Force", "Stroke", "Stress", "Strain"])
            if positions5:
                accs = {pos: CurveAccumulator(5, 4, 3) for pos in positions5}
                for row in iterator:
                    for pos, acc in accs.items():
                        acc.add(row[pos : pos + 5])
                test_name = metadata_value(row1, row2, "Test Name")
                speed = metadata_value(row1, row2, "Speed")
                material = ws.title
                for idx, (pos, acc) in enumerate(accs.items(), start=1):
                    sample = clean(row1[pos]) or f"拉伸重复{idx}"
                    curves.append(
                        make_curve_row(
                            meta,
                            "机械回收",
                            path,
                            ws.title,
                            material,
                            sample,
                            "单轴拉伸",
                            acc,
                            "应变",
                            "%",
                            "应力",
                            "MPa",
                            batch_text=test_name,
                            geometry=True,
                            speed=speed,
                            method=method_tensile,
                            expected_replicates=5,
                            notes="独立列组视为独立dogbone试样",
                        )
                    )
                continue

            positions2 = header_positions(row2, ["Strain (%)", "Stress (MPa)"])
            if positions2:
                accs = {pos: CurveAccumulator(2, 0, 1) for pos in positions2}
                for row in iterator:
                    for pos, acc in accs.items():
                        acc.add(row[pos : pos + 2])
                for idx, (pos, acc) in enumerate(accs.items(), start=1):
                    material = clean(row1[pos]) or ws.title
                    curves.append(
                        make_curve_row(
                            meta,
                            "机械回收",
                            path,
                            ws.title,
                            material,
                            f"汇总曲线{idx}",
                            "单轴拉伸",
                            acc,
                            "应变",
                            "%",
                            "应力",
                            "MPa",
                            geometry=False,
                            method="工作簿给出二维应力-应变曲线；未附独立重复层级",
                            expected_replicates="未声明",
                            notes="材料表征汇总曲线，不可冒充独立重复",
                        )
                    )
                continue

            if any(k in ws.title.lower() for k in ("tensile", "dmta")) or any(k in lower_name for k in ("tensile", "dmta")):
                if not any(clean(v) for v in row1 + row2 + row3):
                    issues.append(f"{path.name}/{ws.title}: 空工作表")
                elif ws.title.lower() not in {"ir", "sec"}:
                    issues.append(f"{path.name}/{ws.title}: 存在机械相关名称但未识别曲线结构")
        wb.close()
    return curves, issues


def read_csv(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            return list(csv.reader(io.StringIO(text)))
        except UnicodeDecodeError:
            pass
    raise UnicodeError(path)


def audit_low_csv(meta: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    directory = meta["dir"] / "解包内容" / "Raw_Mechanical_Testing"
    curves: list[dict[str, Any]] = []
    issues: list[str] = []
    method_tensile = "README: Shimadzu AGS-X；室温；50 mm/min；每材料5个拉伸重复；模量拟合0–1%应变"
    method_hyst = "README: Shimadzu AGS-X；每材料1个dogbone；50%应变；20循环；Origin平滑"
    for path in sorted(directory.glob("*.csv")):
        rows = read_csv(path)
        if len(rows) < 4:
            issues.append(f"{path.name}: 少于4行")
            continue
        material = path.name.split("_")[0]
        if "_tensile" in path.name:
            positions = header_positions(tuple(rows[1]), ["Time", "Force", "Stroke", "Stress", "Strain"])
            accs = {pos: CurveAccumulator(5, 4, 3) for pos in positions}
            for row in rows[3:]:
                for pos, acc in accs.items():
                    acc.add(row[pos : pos + 5])
            for idx, (pos, acc) in enumerate(accs.items(), start=1):
                sample = clean(rows[0][pos]) or f"拉伸重复{idx}"
                curves.append(
                    make_curve_row(
                        meta,
                        "低天花板",
                        path,
                        "CSV",
                        material,
                        sample,
                        "单轴拉伸",
                        acc,
                        "应变",
                        "%",
                        "应力",
                        "MPa",
                        batch_text=sample,
                        geometry=True,
                        speed="50 mm/min",
                        method=method_tensile,
                        expected_replicates=5,
                        notes="五个列组分别为独立试样",
                    )
                )
        elif "_hysteresis" in path.name:
            headers = [clean(v) for v in rows[1]]
            try:
                strain_idx = headers.index("Strain")
                stress_idx = headers.index("Stress")
                cycle_idx = headers.index("Cycle")
            except ValueError:
                issues.append(f"{path.name}: 缺少Strain/Stress/Cycle字段")
                continue
            acc = CurveAccumulator(len(headers), strain_idx, stress_idx, cycle_idx)
            for row in rows[3:]:
                acc.add(row[: len(headers)])
            sample = clean(rows[0][0]) or "滞回试样1"
            curves.append(
                make_curve_row(
                    meta,
                    "低天花板",
                    path,
                    "CSV",
                    material,
                    sample,
                    "循环滞回",
                    acc,
                    "应变",
                    "%",
                    "应力",
                    "kPa",
                    batch_text=sample,
                    geometry=True,
                    speed="50 mm/min",
                    method=method_hyst,
                    expected_replicates=1,
                    notes="README声明原始噪声数据经Origin平滑；保留曲线但需记录处理史",
                )
            )
    return curves, issues


def audit_low_xls_dmta(meta: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    directory = meta["dir"] / "解包内容" / "Raw_Mechanical_Testing"
    command = [
        "pwsh.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(PS_DMTA),
        "-目录",
        str(directory),
    ]
    completed = subprocess.run(command, check=True, capture_output=True, text=True, encoding="utf-8")
    payload = json.loads(completed.stdout.lstrip("\ufeff"))
    if isinstance(payload, dict):
        payload = [payload]
    curves = []
    issues = []
    for item in payload:
        path = directory / item["file"]
        material = path.name.split("_")[0]
        acc = CurveAccumulator(8, 2, 6)
        acc.point_count = int(item["point_count"])
        acc.data_rows = int(item["data_rows"])
        acc.numeric_value_count = int(item["numeric_value_count"])
        acc.missing_primary_rows = int(item["missing_primary_rows"])
        acc.secondary_missing_cells = int(item["secondary_missing_cells"])
        acc.min_x = item["min_temperature_c"]
        acc.max_x = item["max_temperature_c"]
        acc.min_y = item["min_storage_pa"]
        acc.max_y = item["max_storage_pa"]
        row = make_curve_row(
            meta,
            "低天花板",
            path,
            item["sheet"],
            material,
            "DMTA试样1",
            "DMTA温度扫描",
            acc,
            "温度",
            "°C",
            "储能模量E'",
            "Pa",
            batch_text=item.get("test_name", ""),
            geometry=True,
            method="README: RSA-G2；-80至200°C；5°C/min；0.05%振荡应变；6.28 rad/s (1 Hz)",
            expected_replicates=1,
            notes="XLS经本机Excel COM只读审计；同时含tanδ、损耗模量、应力与应变通道",
        )
        row["曲线SHA256"] = item["curve_sha256"]
        curves.append(row)
    return curves, issues


def reconcile_batches(curves: list[dict[str, Any]]) -> None:
    candidates: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    for row in curves:
        if not str(row["批次键"]).endswith("|未解析批次"):
            candidates[(row["来源ID"], row["配方键"])][row["批次键"]] += 1
    for row in curves:
        if str(row["批次键"]).endswith("|未解析批次"):
            counts = candidates.get((row["来源ID"], row["配方键"]))
            if counts:
                batch = counts.most_common(1)[0][0]
                row["批次键"] = batch
                row["批次键依据"] = "由同材料代码的另一机械模态回填"
                row["试样键"] = f"{row['DOI']}|{batch}|{row['材料代码']}|{row['试样标签'] or row['试验类型']}"
                row["泄漏分组键"] = f"{row['DOI']}|{batch}"


def mark_curve_duplicates(curves: list[dict[str, Any]]) -> None:
    groups = Counter(
        (row["试验类型"], row["自变量"], row["自变量单位"], row["因变量"], row["因变量单位"], row["曲线SHA256"])
        for row in curves
        if int(row["点数"]) > 0
    )
    for row in curves:
        key = (row["试验类型"], row["自变量"], row["自变量单位"], row["因变量"], row["因变量单位"], row["曲线SHA256"])
        row["曲线精确重复组大小"] = groups.get(key, 0)


def make_batch_rows(curves: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in curves:
        key = (
            row["来源ID"],
            row["DOI"],
            row["批次键"],
            row["配方键"],
            row["生命周期状态"],
            row["模型准入层"],
        )
        groups[key].append(row)
    output = []
    for key, rows in sorted(groups.items()):
        output.append(
            {
                "来源ID": key[0],
                "DOI": key[1],
                "批次键": key[2],
                "批次键依据": "; ".join(sorted(set(row["批次键依据"] for row in rows))),
                "材料代码": "; ".join(sorted(set(row["材料代码"] for row in rows))),
                "配方键": key[3],
                "生命周期状态": key[4],
                "模型准入层": key[5],
                "试验类型": "; ".join(sorted(set(row["试验类型"] for row in rows))),
                "曲线数": len(rows),
                "独立试样数": len(set(row["试样键"] for row in rows)),
                "曲线点数": sum(int(row["点数"]) for row in rows),
                "数值单元数": sum(int(row["数值单元数"]) for row in rows),
                "泄漏分组键": f"{key[1]}|{key[2]}",
                "备注": "批次键可能为仪器测试名解析或材料代码代理；同批次全部模态必须在同一数据划分",
            }
        )
    return output


def atomic_write(path: Path, payload: bytes) -> None:
    """以同目录普通临时文件、fsync 和原子替换提交一个审计产物。"""

    parent = path.parent.resolve(strict=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def write_tsv(root: Path, name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    """原子覆盖一个获授权TSV审计产物。"""

    path = audit_output_path(root, name)
    handle = io.StringIO(newline="")
    writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t", extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: clean(row.get(column, "")) for column in columns})
    atomic_write(path, handle.getvalue().encode("utf-8-sig"))


def write_json(root: Path, name: str, payload: dict[str, Any]) -> None:
    """原子覆盖一个获授权JSON审计产物。"""

    path = audit_output_path(root, name)
    atomic_write(path, json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))


def mode_counts(file_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    members = [row for row in file_rows if row["记录类型"] == "成员文件" and row["容器层级"] != "顶层数据包"]
    if not members:
        members = [row for row in file_rows if row["记录类型"] == "成员文件"]
    counts: dict[str, dict[str, int]] = defaultdict(lambda: {"文件数": 0, "字节数": 0})
    for row in members:
        counts[row["数据模态"]]["文件数"] += 1
        counts[row["数据模态"]]["字节数"] += int(row["成员字节数"])
    return [{"模态": key, **value} for key, value in sorted(counts.items())]


def curve_summary(curves: list[dict[str, Any]]) -> dict[str, Any]:
    by_type = Counter(row["试验类型"] for row in curves)
    points_by_type = Counter()
    values_by_type = Counter()
    for row in curves:
        points_by_type[row["试验类型"]] += int(row["点数"])
        values_by_type[row["试验类型"]] += int(row["数值单元数"])
    exact_duplicate_groups = Counter(
        (row["试验类型"], row["曲线SHA256"])
        for row in curves
        if int(row["曲线精确重复组大小"] or 0) > 1
    )
    return {
        "曲线总数": len(curves),
        "按试验类型曲线数": dict(sorted(by_type.items())),
        "曲线点总数": sum(int(row["点数"]) for row in curves),
        "按试验类型点数": dict(sorted(points_by_type.items())),
        "机械曲线数值单元总数": sum(int(row["数值单元数"]) for row in curves),
        "按试验类型数值单元数": dict(sorted(values_by_type.items())),
        "材料/配方代码数": len(set(row["配方键"] for row in curves)),
        "批次键数": len(set(row["批次键"] for row in curves)),
        "试样键数": len(set(row["试样键"] for row in curves)),
        "核心实验层曲线数": sum(row["模型准入层"] == "核心实验层" for row in curves),
        "迁移或对照层曲线数": sum(row["模型准入层"] != "核心实验层" for row in curves),
        "含主轴缺失的曲线数": sum(int(row["主轴缺失行数"]) > 0 for row in curves),
        "主轴缺失行总数": sum(int(row["主轴缺失行数"]) for row in curves),
        "含内部空行的曲线数": sum(int(row["内部空行数"]) > 0 for row in curves),
        "曲线精确重复组数": len(exact_duplicate_groups),
        "精确重复曲线记录数": sum(int(row["曲线精确重复组大小"]) > 1 for row in curves),
    }


def dataset_conclusion(key: str) -> dict[str, Any]:
    if key == "机械回收":
        return {
            "总判定": "核心准入（实验多模态TPUU）；商业/14BDO/热固性对照分层降权",
            "适用任务": [
                "软段化学结构、软段Mn、硬段质量分数到拉伸/滞回/DMTA的多任务建模",
                "原生与回收单体再聚合材料的成对保持率/循环性分析",
                "实验曲线微调与模拟数据校准",
            ],
            "不可直接做": [
                "把每个曲线点当独立样本随机切分",
                "把同一配方的5个dogbone随机分散到训练和测试",
                "把Elastollan、橡皮筋或二维汇总曲线当作等权核心TPUU配方",
            ],
            "推荐权重": {
                "核心实验独立试样曲线": 1.0,
                "同配方重复试样": "组内保留用于估计噪声；损失按配方/批次归一，避免重复数支配训练",
                "14BDO线性TPU发表汇总": 0.65,
                "热固性PU桥接": 0.25,
                "Elastollan商业对照": 0.15,
                "商业橡皮筋": 0.0,
                "模拟数据": "仅协议完整、输出可审计、收敛且按目标任务映射的MD/AIMD可在未来预训练中取0.2–0.4上限；输入、未收敛、未解析或未映射结果为0/仅描述符；实验校准1.0，禁止伪装成实验重复",
            },
        }
    return {
        "总判定": "核心准入（四个实验TPUU配方、多模态曲线）；仅可作小样本/外部验证或与其他来源联合训练",
        "适用任务": [
            "低天花板聚酯软段处理路线到拉伸/滞回/DMTA的多任务比较",
            "C/D/R/S材料的严格成组验证和回收路线证据链",
            "NMR/SEC/FTIR/DSC/TGA与机械性质的材料级多模态连接",
        ],
        "不可直接做": [
            "仅凭C/D/R/S代码训练化学结构QSPR；需先从正文/补充材料解码定量配方",
            "将同一材料的5条拉伸重复当作5种配方",
            "将Origin平滑后的滞回点当作新增独立实验",
        ],
        "推荐权重": {
            "核心实验独立试样曲线": 1.0,
            "同配方重复试样": "组内噪声建模；按材料/批次归一",
            "模拟数据": "仅协议完整、输出可审计、收敛且按目标任务映射的MD/AIMD可在未来联合预训练中取0.2–0.4上限；输入、未收敛、未解析或未映射结果为0/仅描述符；实验1.0并做分层校准",
        },
    }


def build_summary(
    key: str,
    meta: dict[str, Any],
    file_rows: list[dict[str, Any]],
    containers: list[dict[str, Any]],
    inventory: dict[str, Any],
    curves: list[dict[str, Any]],
    issues: list[str],
) -> dict[str, Any]:
    extension_counts = Counter(
        row["扩展名"]
        for row in file_rows
        if row["记录类型"] == "成员文件" and row["容器层级"] != "顶层数据包"
    )
    if not extension_counts:
        extension_counts = Counter(row["扩展名"] for row in file_rows if row["记录类型"] == "成员文件")
    quality = list(issues)
    if inventory["未映射解包文件数"]:
        quality.append(f"存在{inventory['未映射解包文件数']}个未映射解包文件")
    if any(c["CRC测试"] != "通过" for c in containers):
        quality.append("至少一个容器CRC失败")
    if any(c["映射缺失数"] or c["映射不一致数"] for c in containers):
        quality.append("至少一个容器存在解包缺失或不一致")
    for row in curves:
        if int(row["主轴缺失行数"]) > 0:
            quality.append(
                f"{row['文件相对路径']}/{row['工作表']}/{row['试样标签']}: "
                f"主轴缺失{row['主轴缺失行数']}行（完整点{row['点数']}个）。"
            )
    formula_batches: dict[str, set[str]] = defaultdict(set)
    for row in curves:
        formula_batches[row["配方键"]].add(row["批次键"])
    for formula, batches in sorted(formula_batches.items()):
        if len(batches) > 1:
            quality.append(
                f"{formula}: 机械模态中出现{len(batches)}个批次键（{'; '.join(sorted(batches))}）；"
                "可能是真实多批次，也可能是工作簿复制元数据，入库前需人工核对。"
            )
    if key == "机械回收":
        quality.extend(
            [
                "Polyol-2k-18HS_Tensile_Prop.xlsx 的 PMCL-2k-18HS 工作表为空，不可计作曲线或负样本。",
                "部分循环滞回工作表只含Time/Force/Stroke/Cycle；应力-应变需结合厚度、宽度和标距转换，且不同文件记录速度含10与50 mm/min，不能混为同一条件。",
                "拉伸工作簿的仪器Qty/Batch可大于可见列组数；本审计仅把实际存在的曲线列组计为独立试样。",
                "14BDO、热固性PU、商业Elastollan与橡皮筋为迁移/对照域，不与水扩链TPUU等权。",
            ]
        )
    else:
        quality.extend(
            [
                "README只说明C/D/R/S沿用论文命名，未在原始文件内提供可直接机器读取的定量配方表；化学结构模型入库前需补正文/SI映射。",
                "四条滞回CSV在字段顺序上不一致（Stress/Strain/Force/Stroke排列变化），必须按列名而非列号解析。",
                "README明确滞回噪声数据经Origin smoothing function处理；应记录为processed experimental curve，不可作为额外独立重复。",
                "旧版XLS DMTA需Excel/兼容BIFF读取器；本审计使用本机Excel COM只读解析。",
            ]
        )
    return {
        "审计版本": "DRUM正式审计_v1.0",
        "审计规范基准日UTC": AUDIT_BASELINE_UTC,
        "审计基准说明": "固定为本项目首次正式DRUM审计日期；运行时刻不进入产物，保证相同输入连续复跑字节确定。",
        "来源": {
            "来源ID": meta["source_id"],
            "标题": meta["title"],
            "DOI": meta["doi"],
            "永久链接": meta["landing"],
            "发布者": "Data Repository for the University of Minnesota (DRUM)",
            "发表年": 2024,
            "许可": meta["license"],
            "许可链接": meta["license_url"],
            "DataCite元数据自链接": f"https://api.datacite.org/dois/{meta['doi']}",
            "元数据固化方式": "标题、DOI、出版者、年份、许可和引用固化于审计脚本；不以运行时网络状态改变产物。",
        },
        "容器与解包": {
            "容器数": len(containers),
            "容器全部CRC通过": all(c["CRC测试"] == "通过" for c in containers),
            "容器全部安全路径": all(c["不安全路径数"] == 0 for c in containers),
            "容器全部与解包完全映射": all(c["映射缺失数"] == 0 and c["映射不一致数"] == 0 for c in containers),
            "逐容器": containers,
            **inventory,
            "扩展名计数（避免顶层嵌套ZIP重复计数）": dict(sorted(extension_counts.items())),
        },
        "实验层级": {
            "研究数": 1,
            "配方/材料代码数": len(set(row["配方键"] for row in curves)),
            "批次键数": len(set(row["批次键"] for row in curves)),
            "试样键数": len(set(row["试样键"] for row in curves)),
            "曲线数": len(curves),
            "曲线点数": sum(int(row["点数"]) for row in curves),
            "说明": "配方/材料代码不等于完全定量配方；批次键优先由仪器测试名解析，缺失时使用材料代码代理。",
        },
        "机械曲线审计": curve_summary(curves),
        "非机械表征分类": mode_counts(file_rows),
        "质量问题": quality,
        "泄漏控制": {
            "最低分组键": "DOI|批次键；同一批次的拉伸、滞回、DMTA、SEC/DSC/FTIR/NMR必须同折",
            "配方级任务": "进一步按DOI|材料代码分组；同配方的5个dogbone不得跨训练/测试",
            "外部验证": "整篇研究/整DOI留出，避免同实验室、同配方族和同仪器处理链泄漏",
            "回收谱系": "原生、回收单体、再聚合材料及其所有表征建立lineage_group并同折",
            "曲线点": "曲线点只是同一试样内的相关观测，绝不是独立样本",
        },
        "准入与多保真权重": dataset_conclusion(key),
        "参考文献": [meta["dataset_citation"], meta["article_citation"]],
        "方法与可复现性": {
            "ZIP": "Python zipfile逐成员解码；testzip CRC；成员CRC32实算；成员/解包SHA256逐字节比较；路径穿越与重复成员名检查",
            "XLSX/CSV": "openpyxl只读data_only解析；CSV按UTF-8-SIG并按字段名识别；曲线哈希基于主轴数值序列",
            "XLS": "Windows Excel COM只读解析Temperature Ramp - 1，不保存原工作簿",
            "输出": ["内容审计摘要.json", "文件校验清单.tsv", "曲线审计清单.tsv", "批次审计清单.tsv"],
        },
    }


def _low_gold_e_base(**updates: Any) -> dict[str, Any]:
    row: dict[str, Any] = {
        column: "" for column in (*RECORD_COLUMNS, *LOW_GOLD_E_OPTIONAL_COLUMNS)
    }
    row.update(
        {
            "source_directory": LOW_DIR.name,
            "target_origin": "experimental",
            "data_origin": "source_raw_curve",
            "reduction_level": "raw_point",
            "fidelity_level": "direct_experimental_tpuu",
            "gold_admission_status": "admitted_reference",
            "mapping_status": (
                "formulation_code_resolved_to_tpuu_family_exact_composition_pending"
            ),
            "current_weight_materialized": "false",
            "training_weight": "",
            "license": "CC0-1.0",
            "citation_keys": (
                "ledger-055-meyersohn-2024-low-ceiling-tpuu-data;"
                "ledger-056-meyersohn-2024-low-ceiling-tpuu"
            ),
            "sample_identity_status": "instrument_specimen_label",
            "curve_points_are_independent_samples": "false",
            "duplicate_status": "unique_curve_payload",
        }
    )
    row.update(updates)
    return row


def _low_curve_audit_index() -> dict[tuple[str, str, str], dict[str, Any]]:
    curves, issues = audit_low_csv(DATASETS["低天花板"])
    if issues:
        raise RuntimeError(f"第十七批DRUM机械CSV审计出现问题: {issues}")
    reconcile_batches(curves)
    mark_curve_duplicates(curves)
    selected = [
        row
        for row in curves
        if row["试验类型"] in {"单轴拉伸", "循环滞回"}
    ]
    if Counter(row["试验类型"] for row in selected) != {
        "单轴拉伸": LOW_EXPECTED_TENSILE_CURVES,
        "循环滞回": LOW_EXPECTED_CYCLIC_CURVES,
    }:
        raise RuntimeError("第十七批DRUM曲线类型数量漂移")
    if sum(int(row["点数"]) for row in selected) != LOW_EXPECTED_RAW_POINTS:
        raise RuntimeError("第十七批DRUM曲线点数量漂移")
    if any(int(row["曲线精确重复组大小"]) != 1 for row in selected):
        raise RuntimeError("第十七批DRUM出现未治理的精确重复曲线")

    with LOW_CURVE_AUDIT.open(
        "r", encoding="utf-8-sig", newline=""
    ) as handle:
        frozen = [
            row
            for row in csv.DictReader(handle, delimiter="\t")
            if row["试验类型"] in {"单轴拉伸", "循环滞回"}
        ]

    def fingerprint(
        row: dict[str, Any],
    ) -> tuple[str, str, str, str, str, str, int, str]:
        return (
            PurePosixPath(str(row["文件相对路径"]).replace("\\", "/")).name,
            str(row["试样标签"]),
            str(row["试验类型"]),
            str(row["配方键"]),
            str(row["批次键"]),
            str(row["泄漏分组键"]),
            int(row["点数"]),
            str(row["曲线SHA256"]),
        )

    if sorted(map(fingerprint, selected)) != sorted(map(fingerprint, frozen)):
        raise RuntimeError("第十七批DRUM重算曲线与冻结审计清单不一致")
    index = {
        (
            PurePosixPath(str(row["文件相对路径"]).replace("\\", "/")).name,
            str(row["试样标签"]),
            str(row["试验类型"]),
        ): row
        for row in selected
    }
    if len(index) != LOW_EXPECTED_TENSILE_CURVES + LOW_EXPECTED_CYCLIC_CURVES:
        raise RuntimeError("第十七批DRUM曲线审计键不唯一")
    return index


def _low_curve_common(
    *,
    path: Path,
    material: str,
    sample: str,
    test_type: str,
    audit_row: dict[str, Any],
) -> dict[str, str]:
    kind = "tensile" if test_type == "单轴拉伸" else "cyclic-tensile"
    curve_id = f"drum-low-ceiling:{material}:{sample}:{kind}"
    batch_id = str(audit_row["批次键"])
    split_group = f"{DATASETS['低天花板']['doi']}|{audit_row['配方键']}"
    return {
        "curve_id": curve_id,
        "source_record_id": curve_id,
        "formulation_id": material,
        "batch_id": batch_id,
        "sample_id": sample,
        "split_group": split_group,
        "family_leakage_group": split_group,
        "global_structure_family_key": f"family_drum_low_ceiling_tpuu|{material}",
        "file_sha256": sha256_file(path),
    }


def _low_tensile_rows(
    path: Path,
    audit_index: dict[tuple[str, str, str], dict[str, Any]],
) -> tuple[list[dict[str, Any]], int, int]:
    source_rows = read_csv(path)
    if len(source_rows) < 4:
        raise RuntimeError(f"第十七批DRUM拉伸CSV行数不足: {path.name}")
    positions = header_positions(
        tuple(source_rows[1]), ["Time", "Force", "Stroke", "Stress", "Strain"]
    )
    if len(positions) != 5:
        raise RuntimeError(f"第十七批DRUM拉伸列组数量漂移: {path.name}")
    material = path.name.split("_")[0]
    output: list[dict[str, Any]] = []
    curve_count = 0
    derived_count = 0
    for column_group, position in enumerate(positions, start=1):
        sample = clean(source_rows[0][position])
        if not sample:
            raise RuntimeError(f"第十七批DRUM拉伸试样标签为空: {path.name}")
        audit_row = audit_index[(path.name, sample, "单轴拉伸")]
        common = _low_curve_common(
            path=path,
            material=material,
            sample=sample,
            test_type="单轴拉伸",
            audit_row=audit_row,
        )
        points: list[tuple[float, float]] = []
        digest = hashlib.sha256()
        point_index = 0
        for physical_row, raw in enumerate(source_rows[3:], start=4):
            values = [
                number(raw[index]) if index < len(raw) else None
                for index in range(position, position + 5)
            ]
            if all(value is None for value in values):
                continue
            time_value, force_value, stroke_value, stress_value, strain_value = values
            if stress_value is None or strain_value is None:
                continue
            if time_value is None or force_value is None or stroke_value is None:
                raise RuntimeError(
                    f"第十七批DRUM拉伸次要通道缺失: {path.name}:row={physical_row}"
                )
            point_index += 1
            points.append((strain_value, stress_value))
            digest.update(
                f"{fmt_num(strain_value)},{fmt_num(stress_value)}\n".encode("ascii")
            )
            output.append(
                _low_gold_e_base(
                    **common,
                    observation_id=f"{common['curve_id']}:point={point_index:06d}",
                    record_kind="curve_point",
                    property_name="tensile_stress",
                    value=fmt_num(stress_value),
                    unit="MPa",
                    condition_name="tensile_strain",
                    condition_value=fmt_num(strain_value),
                    condition_unit="%",
                    secondary_condition_name="elapsed_time",
                    secondary_condition_value=fmt_num(time_value),
                    secondary_condition_unit="s",
                    method_or_test_protocol=(
                        "Shimadzu AGS-X; room temperature; 50 mm/min; "
                        "at least five dogbone replicates per formulation"
                    ),
                    protocol_status="source_protocol_and_axes_complete",
                    potential_weight_ceiling="0.85",
                    source_locator=(
                        f"{path.relative_to(LOW_DIR).as_posix()}#row={physical_row};"
                        f"column_group={column_group}"
                    ),
                    point_index=str(point_index),
                    notes=(
                        "原始单轴拉伸点；材料代码属于IPDI/水扩链TPUU，"
                        "精确定量配方仍待正文映射。"
                    ),
                )
            )
        if point_index != int(audit_row["点数"]):
            raise RuntimeError(
                f"第十七批DRUM拉伸点数与审计不一致: {path.name}/{sample}"
            )
        if digest.hexdigest() != str(audit_row["曲线SHA256"]):
            raise RuntimeError(
                f"第十七批DRUM拉伸曲线哈希与审计不一致: {path.name}/{sample}"
            )
        if not points:
            raise RuntimeError(f"第十七批DRUM拉伸曲线为空: {path.name}/{sample}")

        stresses = [stress for _, stress in points]
        strains = [strain for strain, _ in points]
        toughness = sum(
            (left_stress + right_stress)
            * 0.5
            * (right_strain - left_strain)
            / 100.0
            for (left_strain, left_stress), (right_strain, right_stress) in zip(
                points, points[1:]
            )
            if right_strain >= left_strain
        )
        derived_specs = (
            ("tensile_strength", max(stresses), "MPa"),
            ("elongation_at_break", max(strains), "%"),
            ("tensile_toughness", toughness, "MJ/m^3"),
        )
        for property_name, value, unit in derived_specs:
            output.append(
                _low_gold_e_base(
                    **common,
                    observation_id=f"{common['curve_id']}:derived={property_name}",
                    record_kind="derived_scalar",
                    property_name=property_name,
                    value=fmt_num(value),
                    unit=unit,
                    data_origin="deterministically_derived_from_source_curve",
                    reduction_level="derived",
                    method_or_test_protocol=(
                        "peak/max or trapezoidal integration of ordered source "
                        "stress-strain points; strain converted from percent to fraction"
                    ),
                    fidelity_level="deterministic_curve_derived_experimental",
                    protocol_status="source_protocol_complete_deterministic_derivation",
                    potential_weight_ceiling="0.75",
                    source_locator=(
                        f"{path.relative_to(LOW_DIR).as_posix()}#sample={sample};"
                        f"derived={property_name}"
                    ),
                    notes=(
                        "与母曲线共享试样和泄漏组；派生端点不增加独立试样。"
                    ),
                )
            )
            derived_count += 1
        curve_count += 1
    return output, curve_count, derived_count


def _low_hysteresis_rows(
    path: Path,
    audit_index: dict[tuple[str, str, str], dict[str, Any]],
) -> tuple[list[dict[str, Any]], int]:
    source_rows = read_csv(path)
    if len(source_rows) < 4:
        raise RuntimeError(f"第十七批DRUM循环CSV行数不足: {path.name}")
    headers = [clean(value) for value in source_rows[1]]
    required = {"Time", "Strain", "Stress", "Cycle"}
    if not required.issubset(headers):
        raise RuntimeError(f"第十七批DRUM循环字段漂移: {path.name}")
    time_index = headers.index("Time")
    strain_index = headers.index("Strain")
    stress_index = headers.index("Stress")
    cycle_index = headers.index("Cycle")
    material = path.name.split("_")[0]
    sample = clean(source_rows[0][0])
    audit_row = audit_index[(path.name, sample, "循环滞回")]
    common = _low_curve_common(
        path=path,
        material=material,
        sample=sample,
        test_type="循环滞回",
        audit_row=audit_row,
    )
    output: list[dict[str, Any]] = []
    digest = hashlib.sha256()
    point_index = 0
    for physical_row, raw in enumerate(source_rows[3:], start=4):
        values = [number(raw[index]) if index < len(raw) else None for index in range(len(headers))]
        if all(value is None for value in values):
            continue
        time_value = values[time_index]
        strain_value = values[strain_index]
        stress_value = values[stress_index]
        cycle_value = values[cycle_index]
        if stress_value is None or strain_value is None:
            continue
        if time_value is None or cycle_value is None:
            raise RuntimeError(
                f"第十七批DRUM循环次要通道缺失: {path.name}:row={physical_row}"
            )
        point_index += 1
        digest.update(
            f"{fmt_num(strain_value)},{fmt_num(stress_value)}\n".encode("ascii")
        )
        output.append(
            _low_gold_e_base(
                **common,
                observation_id=f"{common['curve_id']}:point={point_index:06d}",
                record_kind="curve_point",
                property_name="cyclic_tensile_stress",
                value=fmt_num(stress_value),
                unit="kPa",
                condition_name="tensile_strain",
                condition_value=fmt_num(strain_value),
                condition_unit="%",
                secondary_condition_name="cycle_count",
                secondary_condition_value=fmt_num(cycle_value),
                secondary_condition_unit="dimensionless",
                auxiliary_value_name="elapsed_time",
                auxiliary_value=fmt_num(time_value),
                auxiliary_unit="s",
                method_or_test_protocol=(
                    "Shimadzu AGS-X; 50% strain loading-unloading; 20 cycles; "
                    "50 mm/min; source README declares Origin smoothing"
                ),
                data_origin="source_processed_experimental_curve",
                fidelity_level="processed_experimental_tpuu_cyclic_curve",
                protocol_status="source_declared_origin_smoothing_retained",
                potential_weight_ceiling="0.65",
                source_locator=(
                    f"{path.relative_to(LOW_DIR).as_posix()}#row={physical_row}"
                ),
                point_index=str(point_index),
                notes=(
                    "来源明确声明使用Origin平滑；保留为同一试样的循环响应，"
                    "不扩增为独立材料。"
                ),
            )
        )
    if point_index != int(audit_row["点数"]):
        raise RuntimeError(f"第十七批DRUM循环点数与审计不一致: {path.name}")
    if digest.hexdigest() != str(audit_row["曲线SHA256"]):
        raise RuntimeError(f"第十七批DRUM循环曲线哈希与审计不一致: {path.name}")
    return output, 1


def build_low_ceiling_gold_e_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """物化低天花板TPUU的CSV力学曲线；DMTA旧XLS保留为待接入审计项。"""

    actual_files = {path.name for path in LOW_MECHANICAL_DIR.glob("*.csv")}
    if actual_files != set(LOW_MATERIALIZED_FILE_SPECS):
        raise RuntimeError(
            "第十七批DRUM机械CSV集合漂移: "
            f"missing={sorted(set(LOW_MATERIALIZED_FILE_SPECS) - actual_files)}, "
            f"extra={sorted(actual_files - set(LOW_MATERIALIZED_FILE_SPECS))}"
        )
    for filename, expected in LOW_MATERIALIZED_FILE_SPECS.items():
        path = LOW_MECHANICAL_DIR / filename
        actual = (path.stat().st_size, sha256_file(path))
        if actual != expected:
            raise RuntimeError(
                f"第十七批DRUM原始文件漂移: {filename}; actual={actual}"
            )

    audit_index = _low_curve_audit_index()
    rows: list[dict[str, Any]] = []
    tensile_curves = cyclic_curves = derived_scalars = 0
    for path in sorted(LOW_MECHANICAL_DIR.glob("*.csv")):
        if path.name.endswith("_tensile.csv"):
            source_rows, curve_count, derived_count = _low_tensile_rows(
                path, audit_index
            )
            tensile_curves += curve_count
            derived_scalars += derived_count
        elif path.name.endswith("_hysteresis.csv"):
            source_rows, curve_count = _low_hysteresis_rows(path, audit_index)
            cyclic_curves += curve_count
        else:
            raise RuntimeError(f"第十七批DRUM出现未治理CSV: {path.name}")
        rows.extend(source_rows)

    point_count = sum(row["record_kind"] == "curve_point" for row in rows)
    if (
        tensile_curves,
        cyclic_curves,
        point_count,
        derived_scalars,
        len(rows),
    ) != (
        LOW_EXPECTED_TENSILE_CURVES,
        LOW_EXPECTED_CYCLIC_CURVES,
        LOW_EXPECTED_RAW_POINTS,
        LOW_EXPECTED_DERIVED_SCALARS,
        LOW_EXPECTED_RAW_POINTS + LOW_EXPECTED_DERIVED_SCALARS,
    ):
        raise RuntimeError("第十七批DRUM物化规模漂移")
    if len({str(row["observation_id"]) for row in rows}) != len(rows):
        raise RuntimeError("第十七批DRUM observation_id不唯一")
    if any(str(row["training_weight"]) for row in rows):
        raise RuntimeError("第十七批DRUM提前物化训练权重")
    summary = {
        "audit_version": "batch17-drum-low-ceiling-tpuu-v1",
        "source_directory": LOW_DIR.name,
        "dataset_doi": DATASETS["低天花板"]["doi"],
        "article_doi": "10.1021/acs.macromol.4c01431",
        "license": "CC0-1.0",
        "material_count": 4,
        "tensile_curve_count": tensile_curves,
        "cyclic_curve_count": cyclic_curves,
        "dmta_curve_count_audited_not_materialized": 4,
        "raw_curve_point_count": point_count,
        "derived_scalar_count": derived_scalars,
        "gold_e_numeric_row_count": len(rows),
        "admitted_reference_count": len(rows),
        "conditional_reference_count": 0,
        "source_file_count": len(LOW_MATERIALIZED_FILE_SPECS),
        "source_file_sha256": {
            filename: digest
            for filename, (_, digest) in sorted(
                LOW_MATERIALIZED_FILE_SPECS.items()
            )
        },
        "known_gap": (
            "4条DMTA温度扫描位于旧版XLS，已经审计但本批不依赖Excel COM"
            "写入跨平台Gold-E长表"
        ),
        "training_state": {
            "current_weight_materialized": False,
            "training_split_created": False,
            "model_ready_record_count": 0,
        },
    }
    return rows, summary


def validate_outputs(root: Path, file_rows: list[dict[str, Any]], curves: list[dict[str, Any]], batches: list[dict[str, Any]]) -> None:
    """回读四个获授权产物，并核对JSON可解析性和TSV行数。"""

    with audit_output_path(root, "内容审计摘要.json").open("r", encoding="utf-8") as handle:
        json.load(handle)
    expected = {
        "文件校验清单.tsv": len(file_rows),
        "曲线审计清单.tsv": len(curves),
        "批次审计清单.tsv": len(batches),
    }
    for name, row_count in expected.items():
        with audit_output_path(root, name).open("r", encoding="utf-8-sig", newline="") as handle:
            actual = sum(1 for _ in csv.reader(handle, delimiter="\t")) - 1
        if actual != row_count:
            raise RuntimeError(f"{name}: expected {row_count}, got {actual}")


def main(argv: list[str] | None = None) -> int:
    """执行布局检查或正式审计；返回进程退出码。"""

    args = parse_args(argv)
    selected: list[str] = args.datasets
    layout = validate_runtime_layout(selected)
    if args.check_layout:
        print(json.dumps(layout, ensure_ascii=False, indent=2))
        return 0

    all_results = {}
    for key in selected:
        meta = DATASETS[key]
        print(f"[{key}] 容器哈希、CRC与解包映射...", flush=True)
        file_rows, containers, inventory = audit_containers(key, meta)
        print(f"[{key}] 机械曲线解析...", flush=True)
        if key == "机械回收":
            curves, issues = audit_xlsx_mechanical(meta)
        else:
            curves, issues = audit_low_csv(meta)
            xls_curves, xls_issues = audit_low_xls_dmta(meta)
            curves.extend(xls_curves)
            issues.extend(xls_issues)
        reconcile_batches(curves)
        mark_curve_duplicates(curves)
        batches = make_batch_rows(curves)
        summary = build_summary(key, meta, file_rows, containers, inventory, curves, issues)
        root: Path = meta["dir"]
        write_tsv(root, "文件校验清单.tsv", FILE_COLUMNS, file_rows)
        write_tsv(root, "曲线审计清单.tsv", CURVE_COLUMNS, curves)
        write_tsv(root, "批次审计清单.tsv", BATCH_COLUMNS, batches)
        write_json(root, "内容审计摘要.json", summary)
        validate_outputs(root, file_rows, curves, batches)
        all_results[key] = {
            "files": len(file_rows),
            "curves": len(curves),
            "points": sum(int(row["点数"]) for row in curves),
            "numeric_values": sum(int(row["数值单元数"]) for row in curves),
            "batches": len(batches),
            "summary": curve_summary(curves),
        }
        print(json.dumps({key: all_results[key]}, ensure_ascii=False), flush=True)
    print(json.dumps(all_results, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
