"""只读深审第三批两个 Zenodo TPU 实验来源。

本脚本不联网、不修改原始科学文件、不反序列化任意对象，也不创建训练集。
它对固定 25 个科学文件、下载器生成的官方元数据与清单执行完整性检查，实际
解析 14 个 CSV、10 个 XLSX 和 1 个 ZIP，并按材料—协议—曲线—点复算主要
力学/电学统计与已知异常。只有输入、容器和科学语义全部通过后，才原子覆盖
每个来源三个固定白名单审计输出。

运行：

    python 代码/审计/新增开放数据第三批Zenodo实验双源.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import struct
import sys
import tempfile
import zipfile
from collections import Counter
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Iterable, Sequence
from urllib.parse import urlsplit

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"
AUDIT_VERSION = "1.1"

COMMERCIAL = "Zenodo_商业TPU多材料打印传感"
TECOFLEX = "Zenodo_Tecoflex药物复合TPU"
SOURCE_NAMES = (COMMERCIAL, TECOFLEX)
SOURCE_RECORDS = {
    COMMERCIAL: (5_841_610, "10.5281/zenodo.5841610"),
    TECOFLEX: (6_128_356, "10.5281/zenodo.6128356"),
}

OUTPUT_NAMES = ("内容审计摘要.json", "文件校验清单.tsv", "曲线审计清单.tsv")
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source in SOURCE_NAMES
    for filename in OUTPUT_NAMES
)

EXPECTED_SCIENTIFIC_FILES: dict[str, dict[str, tuple[int, str]]] = {
    COMMERCIAL: {
        "Figure_2a.csv": (1_572_196, "0f66f8e66ce09a07de416e75c1a44394a0e72c406951e4bede0c6276b4668e3e"),
        "Figure_2b.csv": (81, "c7c7527363680f9f75a3ee8e8b8c4a45f5a1cd4fc2327240ada082b522422cc1"),
        "Figure_3.csv": (6_021, "ead3f64369fcbd641b6d98a02201e9dff2152c958200a67e463ef7bfb2bc29c9"),
        "Figure_4.csv": (781_439, "170fa88bc743244711e39e13fa5fe2d80fc48fa2d68a5c3b87a6ab7eeee35347"),
        "Figure_5.csv": (2_507_642, "01c5fd854d700b869ea4cce33f29f4148cf80c29d609d3bdfdb79222b417ac27"),
        "Figure_6.csv": (9_795, "cce42a2792f106e0ffc078432a3eaea3a1f871477c47dca98e2c1313e5ed965d"),
        "Figure_7.csv": (590_856, "e1722a25dd3fed80fd6efbba347bc1fe06c000c69031d3ed38199e438626837d"),
        "Figure_9a.csv": (53_235, "953b9d9496db2b836d66bf858d29951cbee32ab4cb4ce340cfc108d260840f44"),
        "Figure_9b.csv": (8_622, "c40f79ed3e873e808d0423c69773e3ccf6330f3589f9d267c6f54afd1172dd86"),
        "Figure_9c.csv": (8_010, "d6e1d1451f272b15367b7360f156f700a1f2ea9ae5a1f9cadc9c2dacdb639451"),
        "Figure_S4.csv": (2_471_108, "59f8b3f3ffbf501e572efbbed4c063c0b9dd8eb5d2de2b480cb0692b46c2a99e"),
        "Figure_S5.csv": (7_096_012, "e05a0a9320778ea42c9aee8f0fe36cf73d79120be55b9956d4bdf8a04004f5d3"),
        "Figure_S6.csv": (2_600_886, "921df56563192fbea3cc4e0c9b3320fcbf8c29f9dab85ed9044491e42b61dadb"),
        "Figure_S7.csv": (1_857_048, "f5d39d6f04f7cca5c842189c3fd7cf7101fd65d0c56ebe2bdb9b7731b640348a"),
    },
    TECOFLEX: {
        "contact angle.xlsx": (20_132, "8510b636dc027191b6dbba767620ba4de108671e412fa1390841d5fa14062a34"),
        "FTIR.xlsx": (1_205_634, "e278d4c2023e78e1b65d4bad33f2e41c83f075f2c7e5fbb3444003f15b9ad6ce"),
        "In vitro antibacterial evaluation.xlsx": (35_763, "7c68c729b916dfad433b14e7dec52dbcbae3631a839c93d278149d2ace5e06b3"),
        "In vitro release and disk diffusion.xlsx": (66_679, "4451604aca8d996341e176bbedf634e12e01b5e1475869ec37652c20a97bac8c"),
        "In vivo antibacterial evaluation.xlsx": (16_460, "9fef24c04f0649182d6d65795d881643bbca8cf1e6fc18a5c1ce1a69a9f5e6b9"),
        "mechanical testing sup fig 1.xlsx": (41_340, "53cc6a8a94a4d873ba556b8a3f0a1fdcb8f9c7f6842b9d79bd59ddee785cc371"),
        "mechanical testing.xlsx": (350_279, "48961b352585271486cbb387fc8d94b65c58b8f2ab78cb552ec01607348ba5c4"),
        "Results in vivo Release.xlsx": (20_390, "6308c9b7b881f7bddb9464ff841914836c24b382a4cc18bdd39002e2472956a1"),
        "TGA thermal analysis.xlsx": (819_167, "a536b40877642fefbc0ae8a476da53317d39696541d96733ba70ac40cfd75153"),
        "tube diameter.xlsx": (24_143, "31c87ac3310d2ca0aa3d05796fb24e3708df41017ec67647f53b5c05c26b580b"),
        "XRD.zip": (61_005, "c327fe8fbff9502591b628539aedbc6e47a22049a31d9b07222f60f061d22dfb"),
    },
}

EXPECTED_CSV_SHAPES = {
    "Figure_2a.csv": (19_880, 17),
    "Figure_2b.csv": (7, 2),
    "Figure_3.csv": (77, 17),
    "Figure_4.csv": (22_583, 20),
    "Figure_5.csv": (66_717, 20),
    "Figure_6.csv": (134, 20),
    "Figure_7.csv": (16_014, 20),
    "Figure_9a.csv": (1_062, 5),
    "Figure_9b.csv": (196, 5),
    "Figure_9c.csv": (250, 3),
    "Figure_S4.csv": (22_586, 20),
    "Figure_S5.csv": (66_720, 20),
    "Figure_S6.csv": (33_444, 20),
    "Figure_S7.csv": (16_029, 20),
}

EXPECTED_WORKBOOK_SHAPES: dict[str, dict[str, tuple[int, int]]] = {
    "contact angle.xlsx": {"Contact Angle": (87, 26)},
    "FTIR.xlsx": {
        "TPU": (7318, 2), "NIC": (7469, 2), "NIC180C": (7469, 2),
        "2%NIC": (7470, 2), "5%NIC": (7470, 2), "10%NIC": (7470, 2),
    },
    "In vitro antibacterial evaluation.xlsx": {
        "Analgesic ": (17, 5), "S.aureus ATCC25923": (61, 8),
        "S.aureus ATCC33591": (57, 10), "S.epidermidis ATCC35984": (57, 10),
        "S.epidermidis  O47": (64, 9),
    },
    "In vitro release and disk diffusion.xlsx": {
        "disk diffusion": (64, 26), "release NIC": (201, 27),
    },
    "In vivo antibacterial evaluation.xlsx": {
        "S. aureus ATCC25923": (55, 11), "S. aureus ATCC33591": (55, 26),
    },
    "mechanical testing sup fig 1.xlsx": {
        "protocol": (12, 7), "Sheet1 (2)": (286, 18),
    },
    "mechanical testing.xlsx": {
        "protocol": (12, 7), "figures": (1, 1), "Elastic modulus": (24, 19),
        "stress @ 100% strain": (18, 21), "Sheet1": (286, 8),
    },
    "Results in vivo Release.xlsx": {"Hoja1": (61, 12)},
    "TGA thermal analysis.xlsx": {"TPU-NIC": (5926, 19)},
    "tube diameter.xlsx": {
        "summary": (13, 7), "TPU": (95, 11), "nic2": (50, 14),
        "nic5-n": (82, 14), "nic5": (35, 14), "nic10": (60, 14),
    },
}

COMMERCIAL_MATERIALS = ("70A", "82A", "85A", "95A", "40D", "98A")
TECOFLEX_MATERIALS = ("TPU", "NIC-2", "NIC-5", "NIC-10")

MECHANICAL_LAYOUTS = {
    "Figure_2a.csv": (COMMERCIAL_MATERIALS, tuple((3 * i, 3 * i + 1) for i in range(6)), "quasi_static_tension", "strain_percent", "stress_mpa"),
    "Figure_S4.csv": (COMMERCIAL_MATERIALS, tuple((3 + 3 * i, 4 + 3 * i) for i in range(6)), "cyclic_10_percent", "time_s", "stress_mpa"),
    "Figure_S5.csv": (COMMERCIAL_MATERIALS, tuple((3 + 3 * i, 4 + 3 * i) for i in range(6)), "cyclic_100_percent", "time_s", "stress_mpa"),
    "Figure_S6.csv": (("82A", "85A", "95A", "40D"), tuple((2 + 3 * i, 3 + 3 * i) for i in range(4)), "high_strain_tension", "time_s", "stress_mpa"),
    "Figure_S7.csv": (COMMERCIAL_MATERIALS, tuple((3 + 3 * i, 4 + 3 * i) for i in range(6)), "repeated_cyclic", "time_s", "stress_mpa"),
}

ELECTRICAL_LAYOUTS = {
    "Figure_3.csv": (COMMERCIAL_MATERIALS, tuple((3 * i, 3 * i + 1) for i in range(6)), "quasi_static_tension", "strain_percent", "relative_resistance"),
    "Figure_4.csv": (COMMERCIAL_MATERIALS, tuple((3 + 3 * i, 4 + 3 * i) for i in range(6)), "cyclic_10_percent", "time_s", "relative_resistance"),
    "Figure_5.csv": (COMMERCIAL_MATERIALS, tuple((3 + 3 * i, 4 + 3 * i) for i in range(6)), "cyclic_100_percent", "time_s", "relative_resistance"),
    "Figure_6.csv": (("82A", "85A", "95A", "40D"), tuple((3 + 3 * i, 4 + 3 * i) for i in range(4)), "high_strain_tension", "time_s", "relative_resistance"),
    "Figure_7.csv": (COMMERCIAL_MATERIALS, tuple((3 + 3 * i, 4 + 3 * i) for i in range(6)), "repeated_cyclic", "time_s", "relative_resistance"),
}


class AuditBlocked(RuntimeError):
    """输入、容器或科学语义不满足固定审计协议。"""


def _same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def _is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & flag) or bool(is_junction())


def _assert_plain_chain(path: Path, stop: Path) -> None:
    if path != stop and stop not in path.parents:
        raise AuditBlocked(f"路径越出审计根：{path}")
    cursor = path
    while True:
        if _is_reparse_point(cursor):
            raise AuditBlocked(f"拒绝符号链接或重解析点：{cursor}")
        if cursor == stop:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"目录缺失、不是普通目录或经链接解析：{path}")
    _assert_plain_chain(path, PROJECT_ROOT)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not _same_path(resolved, path.absolute()):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    _assert_plain_chain(path, PROJECT_ROOT)


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    parent = path.parent
    require_directory(parent)
    if path.exists() and (not path.is_file() or _is_reparse_point(path)):
        raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if _is_reparse_point(temporary) or not temporary.is_file():
            raise AuditBlocked(f"审计临时输出不是普通文件：{temporary}")
        assert_output_allowed(path)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_json(path: Path, payload: dict[str, object]) -> None:
    rendered = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False
    )
    atomic_write(path, (rendered + "\n").encode("utf-8"))


def render_tsv(rows: list[dict[str, object]], columns: list[str]) -> bytes:
    if not rows:
        raise AuditBlocked("拒绝渲染空 TSV")
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer, fieldnames=columns, delimiter="\t", lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _hash_stream(handle: BinaryIO, algorithm: str = "sha256") -> str:
    digest = hashlib.new(algorithm)
    for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
        digest.update(block)
    return digest.hexdigest()


def file_hash(path: Path, algorithm: str = "sha256") -> str:
    require_file(path)
    with path.open("rb") as handle:
        return _hash_stream(handle, algorithm)


def _manifest_digest(items: Iterable[tuple[str, int, str]]) -> str:
    digest = hashlib.sha256()
    for relative, size, checksum in sorted(items):
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(str(size).encode("ascii"))
        digest.update(b"\0")
        digest.update(checksum.encode("ascii"))
        digest.update(b"\n")
    return digest.hexdigest()


def scientific_input_snapshot() -> dict[str, tuple[int, str]]:
    snapshot: dict[str, tuple[int, str]] = {}
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        require_directory(base)
        expected_names = set(EXPECTED_SCIENTIFIC_FILES[source]) | {
            "官方API元数据.json", "官方文件清单.tsv",
        }
        actual_names: set[str] = set()
        for path in sorted(base.iterdir(), key=lambda item: item.name.casefold()):
            if _is_reparse_point(path):
                raise AuditBlocked(f"来源根含符号链接或重解析点：{path}")
            if path in OUTPUT_WHITELIST:
                continue
            if not path.is_file():
                raise AuditBlocked(f"来源根出现未登记目录或特殊对象：{path}")
            if path.name.endswith((".part", ".tmp", ".audit.tmp")):
                raise AuditBlocked(f"来源根出现未完成临时文件：{path}")
            actual_names.add(path.name)
            snapshot[f"{source}/{path.name}"] = (path.stat().st_size, file_hash(path))
        if actual_names != expected_names:
            raise AuditBlocked(
                f"来源根文件集合漂移：{source}; "
                f"missing={sorted(expected_names-actual_names)}; "
                f"extra={sorted(actual_names-expected_names)}"
            )
    return snapshot


def validate_official_capture(source: str) -> None:
    base = DATA_ROOT / source
    record_id, doi = SOURCE_RECORDS[source]
    metadata_path = base / "官方API元数据.json"
    manifest_path = base / "官方文件清单.tsv"
    require_file(metadata_path)
    require_file(manifest_path)
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    if int(metadata.get("record_id", -1)) != record_id:
        raise AuditBlocked(f"官方元数据记录 ID 不符：{source}")
    if str(metadata.get("doi", "")).casefold() != doi.casefold():
        raise AuditBlocked(f"官方元数据 DOI 不符：{source}")
    if str(metadata.get("license", "")).casefold() != "cc-by-4.0":
        raise AuditBlocked(f"官方元数据许可证不符：{source}")
    if str(metadata.get("provider", "")).casefold() != "zenodo":
        raise AuditBlocked(f"官方元数据 provider 不符：{source}")
    api_url = urlsplit(str(metadata.get("api_url", "")))
    if (
        api_url.scheme != "https"
        or api_url.hostname != "zenodo.org"
        or api_url.path.rstrip("/") != f"/api/records/{record_id}"
        or api_url.query
        or api_url.fragment
    ):
        raise AuditBlocked(f"官方元数据 API URL 不符：{source}")
    download_policy = metadata.get("download_policy") or {}
    if not bool(download_policy.get("all_record_files_downloaded")):
        raise AuditBlocked(f"官方元数据未声明全记录下载：{source}")

    expected = EXPECTED_SCIENTIFIC_FILES[source]
    if (
        int(download_policy.get("required_file_count", -1)) != len(expected)
        or int(download_policy.get("excluded_file_count", -1)) != 0
    ):
        raise AuditBlocked(f"官方元数据下载范围不符：{source}")
    with manifest_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    if len(rows) != len(expected) or {row["filename"] for row in rows} != set(expected):
        raise AuditBlocked(f"官方文件清单集合不符：{source}")
    metadata_files = metadata.get("files")
    if not isinstance(metadata_files, list):
        raise AuditBlocked(f"官方元数据 files 不是列表：{source}")
    metadata_by_name = {
        str(row.get("filename", "")): row
        for row in metadata_files
        if isinstance(row, dict)
    }
    if len(metadata_by_name) != len(metadata_files) or set(metadata_by_name) != set(expected):
        raise AuditBlocked(f"官方元数据文件集合不符或存在重复：{source}")
    for row in rows:
        name = row["filename"]
        path = base / name
        if (
            row.get("source_directory") != source
            or row.get("provider", "").casefold() != "zenodo"
            or int(row.get("record_id", -1)) != record_id
            or row.get("doi", "").casefold() != doi.casefold()
        ):
            raise AuditBlocked(f"官方文件清单来源身份不符：{source}/{name}")
        if row.get("decision") != "download" or row.get("local_state") != "verified_present":
            raise AuditBlocked(f"官方文件清单本地状态不符：{source}/{name}")
        if int(row["bytes"]) != expected[name][0]:
            raise AuditBlocked(f"官方文件清单大小不符：{source}/{name}")
        if row.get("local_sha256", "").lower() != expected[name][1]:
            raise AuditBlocked(f"官方文件清单 SHA-256 不符：{source}/{name}")
        md5 = row.get("md5", "").lower()
        if re.fullmatch(r"[0-9a-f]{32}", md5) is None:
            raise AuditBlocked(f"官方文件清单 MD5 格式不符：{source}/{name}")
        download_url = urlsplit(row.get("download_url", ""))
        if (
            download_url.scheme != "https"
            or download_url.hostname != "zenodo.org"
            or not download_url.path.startswith(f"/api/records/{record_id}/files/")
            or not download_url.path.endswith("/content")
            or download_url.query
            or download_url.fragment
        ):
            raise AuditBlocked(f"官方文件清单下载 URL 不符：{source}/{name}")
        captured = metadata_by_name[name]
        comparable = {
            "filename": name,
            "bytes": int(row["bytes"]),
            "md5": md5,
            "download_url": row["download_url"],
            "decision": row["decision"],
        }
        if captured != comparable:
            raise AuditBlocked(f"官方元数据与 TSV 文件字段不一致：{source}/{name}")
        if (
            path.stat().st_size != expected[name][0]
            or file_hash(path) != expected[name][1]
            or file_hash(path, "md5") != md5
        ):
            raise AuditBlocked(f"科学文件完整性不符：{source}/{name}")


def _safe_zip_name(name: str) -> None:
    if "\\" in name or "\x00" in name:
        raise AuditBlocked(f"ZIP 成员名含非法字符：{name!r}")
    pure = PurePosixPath(name)
    if pure.is_absolute() or any(part in {"", ".", ".."} for part in pure.parts):
        raise AuditBlocked(f"ZIP 成员路径越界：{name!r}")
    if re.match(r"^[A-Za-z]:", name):
        raise AuditBlocked(f"ZIP 成员含盘符：{name!r}")


def inspect_zip_container(path: Path) -> dict[str, int]:
    require_file(path)
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        names = [item.filename for item in infos]
        if len(names) != len(set(names)):
            raise AuditBlocked(f"ZIP 含重复成员名：{path.name}")
        total = 0
        for info in infos:
            _safe_zip_name(info.filename)
            if info.flag_bits & 0x1:
                raise AuditBlocked(f"ZIP 含加密成员：{path.name}/{info.filename}")
            mode = (info.external_attr >> 16) & 0o170000
            if mode == stat.S_IFLNK:
                raise AuditBlocked(f"ZIP 含符号链接成员：{path.name}/{info.filename}")
            total += info.file_size
            if info.compress_size == 0 and info.file_size > 0:
                raise AuditBlocked(f"ZIP 成员压缩大小为零：{path.name}/{info.filename}")
            if info.compress_size and info.file_size / info.compress_size > 1_000:
                raise AuditBlocked(f"ZIP 成员压缩比异常：{path.name}/{info.filename}")
        if total > 250_000_000:
            raise AuditBlocked(f"ZIP 总解压量超上限：{path.name}={total}")
        corrupt = archive.testzip()
        if corrupt is not None:
            raise AuditBlocked(f"ZIP CRC 失败：{path.name}/{corrupt}")
    return {"archive_entries": len(infos), "uncompressed_bytes": total}


def _finite(value: object) -> float | None:
    if value is None:
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def read_semicolon_csv(path: Path) -> list[list[str]]:
    require_file(path)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    expected_rows, expected_columns = EXPECTED_CSV_SHAPES[path.name]
    if len(rows) != expected_rows or {len(row) for row in rows} != {expected_columns}:
        raise AuditBlocked(
            f"CSV 形状漂移：{path.name}; rows={len(rows)}; "
            f"widths={sorted({len(row) for row in rows})}"
        )
    return rows


def numeric_pairs(
    rows: Sequence[Sequence[object]], x_column: int, y_column: int, *, start: int = 2
) -> tuple[list[tuple[float, float]], int, int]:
    pairs: list[tuple[float, float]] = []
    x_only = 0
    y_only = 0
    for row in rows[start:]:
        x = _finite(row[x_column])
        y = _finite(row[y_column])
        if x is not None and y is not None:
            pairs.append((x, y))
        elif x is not None:
            x_only += 1
        elif y is not None:
            y_only += 1
    return pairs, x_only, y_only


def sequence_digest(points: Sequence[tuple[float, float]]) -> str:
    digest = hashlib.sha256()
    for x, y in points:
        digest.update(struct.pack(">dd", x, y))
    return digest.hexdigest()


def _curve_row(
    *, source: str, file: str, sheet: str, material: str, protocol: str,
    x: str, y: str, points: int, lineage: str, decision: str,
    reason: str, future_weight_ceiling: float | None,
) -> dict[str, object]:
    record_id, doi = SOURCE_RECORDS[source]
    return {
        "source": source,
        "record_id": record_id,
        "doi": doi,
        "file": file,
        "sheet": sheet,
        "material": material,
        "protocol": protocol,
        "x_observable": x,
        "y_observable": y,
        "raw_points": points,
        "usable_points": (
            0 if decision.startswith(("exclude_", "hold_")) else points
        ),
        "lineage_group": lineage,
        # 协议是观测身份，不是材料筛选任务的拆分维度。同一 DOI 下同一牌号/
        # 配方的全部协议、表征和派生量必须同折，避免跨协议材料身份泄漏。
        "split_group": f"{doi}|{material}",
        "decision": decision,
        "decision_reason": reason,
        "future_weight_ceiling": "" if future_weight_ceiling is None else future_weight_ceiling,
        "training_split_created": False,
        "training_weight_materialized": False,
    }


def audit_commercial() -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    base = DATA_ROOT / COMMERCIAL
    parsed: dict[str, list[list[str]]] = {}
    file_rows: list[dict[str, object]] = []
    for name, (size, sha256) in sorted(EXPECTED_SCIENTIFIC_FILES[COMMERCIAL].items()):
        path = base / name
        rows = read_semicolon_csv(path)
        parsed[name] = rows
        numeric_cells = sum(_finite(value) is not None for row in rows for value in row)
        dash_cells = sum(str(value).strip() == "--" for row in rows for value in row)
        file_rows.append({
            "source": COMMERCIAL, "filename": name, "bytes": size, "sha256": sha256,
            "container_type": "csv_semicolon", "parsed_units": len(rows),
            "parse_metrics": json.dumps({
                "rows": len(rows), "columns": len(rows[0]),
                "numeric_cells": numeric_cells, "double_dash_cells": dash_cells,
            }, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
            "decision": "parsed", "anomaly": "",
        })

    curve_rows: list[dict[str, object]] = []
    mechanical_points = 0
    mechanical_histories = 0
    mechanical_sequences: dict[tuple[str, str], list[tuple[float, float]]] = {}
    expected_mechanical_counts = {
        "Figure_2a.csv": (8097, 19005, 10609, 10057, 19878, 10793),
        "Figure_S4.csv": (22582, 22577, 22576, 22584, 22202, 22580),
        "Figure_S5.csv": (66715, 66718, 66715, 66715, 66697, 66718),
        "Figure_S6.csv": (25941, 25947, 33442, 25904),
        "Figure_S7.csv": (16011, 16017, 16011, 16002, 16013, 16027),
    }
    for name, (materials, columns, protocol, x_name, y_name) in MECHANICAL_LAYOUTS.items():
        observed: list[int] = []
        for material, (x_column, y_column) in zip(materials, columns, strict=True):
            points, x_only, y_only = numeric_pairs(parsed[name], x_column, y_column)
            observed.append(len(points))
            if name == "Figure_S4.csv" and material == "40D":
                if (x_only, y_only) != (29, 0):
                    raise AuditBlocked(f"S4/40D 尾部缺失口径漂移：{x_only}/{y_only}")
            elif x_only or y_only:
                raise AuditBlocked(f"机械曲线出现未登记孤值：{name}/{material}")
            mechanical_points += len(points)
            mechanical_histories += 1
            mechanical_sequences[(name, material)] = points
            identity_conflict = (
                name == "Figure_S7.csv" and material in {"70A", "85A"}
            )
            curve_rows.append(_curve_row(
                source=COMMERCIAL, file=name, sheet="", material=material,
                protocol=protocol, x=x_name, y=y_name, points=len(points),
                lineage=f"{name}|{material}",
                decision=(
                    "hold_material_identity_conflict"
                    if identity_conflict else "candidate_experimental_history"
                ),
                reason=(
                    "Figure_S7 的70A与85A逐点完全相同；来源不能判定真实牌号，二者均隔离"
                    if identity_conflict else "文件内直接实验历史；尚未创建训练切分"
                ),
                future_weight_ceiling=0.0 if identity_conflict else None,
            ))
        if tuple(observed) != expected_mechanical_counts[name]:
            raise AuditBlocked(f"机械曲线点数漂移：{name}={observed}")
    if mechanical_histories != 28 or mechanical_points != 821_133:
        raise AuditBlocked(
            f"机械历史总量漂移：histories={mechanical_histories}; points={mechanical_points}"
        )

    s7_70 = mechanical_sequences[("Figure_S7.csv", "70A")]
    s7_85 = mechanical_sequences[("Figure_S7.csv", "85A")]
    if s7_70 != s7_85 or len(s7_70) != 16_011:
        raise AuditBlocked("Figure_S7 的 70A/85A 精确重复关系漂移")
    s7_duplicate_digest = sequence_digest(s7_70)

    electrical_points = 0
    electrical_histories = 0
    expected_electrical_counts = {
        "Figure_3.csv": (32, 70, 40, 26, 75, 39),
        "Figure_4.csv": (91, 92, 92, 90, 92, 91),
        "Figure_5.csv": (295, 298, 294, 299, 297, 297),
        "Figure_6.csv": (102, 105, 132, 103),
        "Figure_7.csv": (382, 388, 380, 372, 322, 399),
    }
    for name, (materials, columns, protocol, x_name, y_name) in ELECTRICAL_LAYOUTS.items():
        observed: list[int] = []
        for material, (x_column, y_column) in zip(materials, columns, strict=True):
            points, x_only, y_only = numeric_pairs(parsed[name], x_column, y_column)
            if x_only or y_only:
                raise AuditBlocked(f"电学曲线出现未登记孤值：{name}/{material}")
            observed.append(len(points))
            electrical_points += len(points)
            electrical_histories += 1
            curve_rows.append(_curve_row(
                source=COMMERCIAL, file=name, sheet="", material=material,
                protocol=protocol, x=x_name, y=y_name, points=len(points),
                lineage=f"{name}|{material}", decision="candidate_electrical_history",
                reason="直接电学历史；与力学采样率不同，后续只按协议/材料分组对齐",
                future_weight_ceiling=None,
            ))
        if tuple(observed) != expected_electrical_counts[name]:
            raise AuditBlocked(f"电学曲线点数漂移：{name}={observed}")
    if electrical_histories != 28 or electrical_points != 5_295:
        raise AuditBlocked(
            f"电学历史总量漂移：histories={electrical_histories}; points={electrical_points}"
        )

    # Figure 2b 是 6 个模量—硬度标量对；Figure 9a-c 是 10 条原始电阻历史。
    scalar_pairs, scalar_x_only, scalar_y_only = numeric_pairs(parsed["Figure_2b.csv"], 0, 1, start=1)
    if len(scalar_pairs) != 6 or scalar_x_only or scalar_y_only:
        raise AuditBlocked("Figure_2b 模量—硬度标量口径漂移")
    resistance_layouts = {
        "Figure_9a.csv": (("82A", "85A", "40D", "95A"), (897, 877, 870, 877)),
        "Figure_9b.csv": (("82A", "85A", "40D", "95A"), (136, 134, 137, 140)),
        "Figure_9c.csv": (("40D", "95A"), (231, 231)),
    }
    resistance_points = 0
    resistance_histories = 0
    resistance_x_only = 0
    resistance_y_only = 0
    expected_resistance_orphans = {
        "Figure_9a.csv": ((0, 163), (20, 0), (27, 0), (20, 0)),
        "Figure_9b.csv": ((39, 0), (41, 0), (38, 0), (35, 0)),
        "Figure_9c.csv": ((0, 1), (0, 17)),
    }
    for name, (materials, expected_counts) in resistance_layouts.items():
        observed = []
        observed_orphans = []
        for index, material in enumerate(materials, start=1):
            points, x_only, y_only = numeric_pairs(parsed[name], 0, index)
            observed_orphans.append((x_only, y_only))
            resistance_x_only += x_only
            resistance_y_only += y_only
            observed.append(len(points))
            resistance_points += len(points)
            resistance_histories += 1
            curve_rows.append(_curve_row(
                source=COMMERCIAL, file=name, sheet="", material=material,
                protocol="baseline_resistance", x="time_s", y="resistance_ohm",
                points=len(points), lineage=f"{name}|{material}",
                decision="candidate_auxiliary_electrical_history",
                reason="绝对电阻辅助历史，不与相对电阻混列", future_weight_ceiling=None,
            ))
        if tuple(observed) != expected_counts:
            raise AuditBlocked(f"绝对电阻点数漂移：{name}={observed}")
        if tuple(observed_orphans) != expected_resistance_orphans[name]:
            raise AuditBlocked(f"绝对电阻共享时间列孤值口径漂移：{name}={observed_orphans}")
    if resistance_histories != 10 or resistance_points != 4_530:
        raise AuditBlocked("绝对电阻历史总量漂移")

    # S5 A:B 在两级表头后只剩 25,903 对，时间/应变范围与 303 s/100% 的
    # 六条 S5 应力史不符；其时间列与 S6 的 40D 时间通道错开一文件行相同。
    s5 = parsed["Figure_S5.csv"]
    s6 = parsed["Figure_S6.csv"]
    s5_reference, x_only, y_only = numeric_pairs(s5, 0, 1, start=2)
    if x_only or y_only or len(s5_reference) != 25_903:
        raise AuditBlocked("S5 A:B 协议列完整数值对口径漂移")
    if max(x for x, _ in s5_reference) != 106.152 or max(y for _, y in s5_reference) != 119.99815:
        raise AuditBlocked("S5 A:B 协议列范围漂移")
    figure5_reference, fx_only, fy_only = numeric_pairs(parsed["Figure_5.csv"], 0, 1, start=1)
    if fx_only or fy_only or len(figure5_reference) != 66_715:
        raise AuditBlocked("Figure_5 正常协议列口径漂移")
    if max(x for x, _ in figure5_reference) != 303.26 or max(y for _, y in figure5_reference) != 99.99825:
        raise AuditBlocked("Figure_5 正常协议列范围漂移")
    shifted_s5 = [_finite(row[0]) for row in s5[3:25_905]]       # 文件行 4..25905
    shifted_s6 = [_finite(row[11]) for row in s6[4:25_906]]      # 文件行 5..25906
    if len(shifted_s5) != 25_902 or shifted_s5 != shifted_s6 or any(v is None for v in shifted_s5):
        raise AuditBlocked("S5 A列与S6/40D时间列的一行错位精确相同关系漂移")

    figure5_dash_padding = sum(
        value.strip() == "--" for row in parsed["Figure_5.csv"] for value in row
    )
    if figure5_dash_padding != 66_426:
        raise AuditBlocked(f"Figure_5 双横线填充口径漂移：{figure5_dash_padding}")

    summary: dict[str, object] = {
        "audit_date": AUDIT_DATE,
        "audit_version": AUDIT_VERSION,
        "source": COMMERCIAL,
        "record_id": SOURCE_RECORDS[COMMERCIAL][0],
        "doi": SOURCE_RECORDS[COMMERCIAL][1],
        "license": "CC BY 4.0",
        "scientific_file_count": 14,
        "physical_specimen_count": None,
        "physical_specimen_count_reason": "图表级材料—协议历史没有逐试样ID，禁止把6个牌号或28条历史伪装为试样数",
        "mechanical": {
            "material_protocol_history_count": mechanical_histories,
            "finite_point_count": mechanical_points,
        },
        "electrical_relative_resistance": {
            "material_protocol_history_count": electrical_histories,
            "finite_point_count": electrical_points,
        },
        "electrical_absolute_resistance_auxiliary": {
            "history_count": resistance_histories,
            "finite_point_count": resistance_points,
            "shared_time_without_resistance_count": resistance_x_only,
            "resistance_without_shared_time_count": resistance_y_only,
            "decision": "仅完整坐标对可用；孤值隔离",
        },
        "scalar_modulus_hardness_pair_count": len(scalar_pairs),
        "anomalies": {
            "figure_s7_exact_duplicate": {
                "conflicting_material_labels": ["70A", "85A"],
                "point_count_each": len(s7_70),
                "sequence_sha256": s7_duplicate_digest,
                "decision": "hold_both_until_material_identity_is_resolved",
                "future_weight_ceiling_each": 0,
            },
            "figure_s4_40d_missing_stress_tail": {
                "time_without_stress_count": 29, "decision": "quarantine_tail_only",
            },
            "figure_s5_reference_misalignment": {
                "embedded_pair_count_after_two_header_rows": len(s5_reference),
                "time_max_s": max(x for x, _ in s5_reference),
                "strain_max_percent": max(y for _, y in s5_reference),
                "expected_protocol_pair_count": len(figure5_reference),
                "expected_time_max_s": max(x for x, _ in figure5_reference),
                "expected_strain_max_percent": max(y for _, y in figure5_reference),
                "s6_40d_one_row_shift_exact_match_count": len(shifted_s5),
                "decision": "quarantine_S5_columns_A_B_only_keep_six_stress_histories",
            },
            "figure_5_double_dash_padding_count": figure5_dash_padding,
        },
        "training_split_created": False,
        "training_weight_materialized": False,
        "training_rows_created": 0,
    }
    return summary, file_rows, curve_rows


def canonical_cell_digest(workbook_path: Path) -> tuple[dict[str, tuple[int, int]], dict[str, int | str]]:
    expected_shapes = EXPECTED_WORKBOOK_SHAPES[workbook_path.name]
    container = inspect_zip_container(workbook_path)
    digest = hashlib.sha256()
    counts: Counter[str] = Counter()
    workbook = load_workbook(
        workbook_path, read_only=True, data_only=False, keep_links=False
    )
    try:
        shapes = {sheet.title: (sheet.max_row, sheet.max_column) for sheet in workbook.worksheets}
        if shapes != expected_shapes:
            raise AuditBlocked(f"工作簿工作表形状漂移：{workbook_path.name}={shapes}")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    if cell.data_type == "f":
                        kind = "formula"
                    elif isinstance(value, bool):
                        kind = "boolean"
                    elif isinstance(value, (int, float)) and math.isfinite(float(value)):
                        kind = "numeric"
                    else:
                        kind = "text"
                    counts[kind] += 1
                    payload = json.dumps(
                        [sheet.title, cell.row, cell.column, kind, value],
                        ensure_ascii=False, separators=(",", ":"), default=str,
                    )
                    digest.update(payload.encode("utf-8"))
                    digest.update(b"\n")
    finally:
        workbook.close()
    metrics: dict[str, int | str] = {
        **container,
        "nonempty_cells": sum(counts.values()),
        "numeric_cells": counts["numeric"],
        "formula_cells": counts["formula"],
        "text_cells": counts["text"],
        "boolean_cells": counts["boolean"],
        "parsed_cell_sha256": digest.hexdigest(),
    }
    return shapes, metrics


def audit_xrd(path: Path) -> tuple[dict[str, int], list[tuple[str, int, str]]]:
    container = inspect_zip_container(path)
    expected_entries = {
        "TPU+2%Niclosamide.xy": (46_214, 3_866_757_921),
        "TPU+5%Niclosamide.xy": (46_081, 3_108_175_240),
        "TPU+10%Niclosamide.xy": (46_153, 4_216_482_463),
        "TPU control.xy": (46_236, 549_318_356),
    }
    rows: list[tuple[str, int, str]] = []
    grids: list[list[float]] = []
    with zipfile.ZipFile(path) as archive:
        actual = {item.filename: (item.file_size, item.CRC) for item in archive.infolist()}
        if actual != expected_entries:
            raise AuditBlocked(f"XRD ZIP 精确成员漂移：{actual}")
        for name in sorted(expected_entries):
            text = archive.read(name).decode("utf-8-sig")
            points: list[tuple[float, float]] = []
            for line_number, line in enumerate(text.splitlines(), start=1):
                fields = line.split()
                if len(fields) != 2:
                    raise AuditBlocked(f"XRD 行字段数异常：{name}:{line_number}")
                x = _finite(fields[0])
                y = _finite(fields[1])
                if x is None or y is None:
                    raise AuditBlocked(f"XRD 非有限数：{name}:{line_number}")
                points.append((x, y))
            if len(points) != 2_573 or any(b[0] <= a[0] for a, b in zip(points, points[1:])):
                raise AuditBlocked(f"XRD 点数或角度单调性异常：{name}")
            grids.append([point[0] for point in points])
            rows.append((name, len(points), sequence_digest(points)))
    if any(grid != grids[0] for grid in grids[1:]):
        raise AuditBlocked("XRD 四条曲线不再共享角度网格")
    if container != {"archive_entries": 4, "uncompressed_bytes": 184_684}:
        raise AuditBlocked(f"XRD ZIP 形状漂移：{container}")
    return container, rows


def _xlsx_values(path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook[sheet_name]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def audit_tecoflex() -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    base = DATA_ROOT / TECOFLEX
    file_rows: list[dict[str, object]] = []
    workbook_profiles: dict[str, dict[str, object]] = {}
    for name, (size, sha256) in sorted(EXPECTED_SCIENTIFIC_FILES[TECOFLEX].items()):
        path = base / name
        if name.endswith(".xlsx"):
            shapes, metrics = canonical_cell_digest(path)
            profile = {
                "sheets": {key: list(value) for key, value in sorted(shapes.items())},
                **metrics,
            }
            workbook_profiles[name] = profile
            file_rows.append({
                "source": TECOFLEX, "filename": name, "bytes": size, "sha256": sha256,
                "container_type": "xlsx", "parsed_units": len(shapes),
                "parse_metrics": json.dumps(profile, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
                "decision": "parsed", "anomaly": "",
            })

    xrd_container, xrd_rows = audit_xrd(base / "XRD.zip")
    file_rows.append({
        "source": TECOFLEX, "filename": "XRD.zip",
        "bytes": EXPECTED_SCIENTIFIC_FILES[TECOFLEX]["XRD.zip"][0],
        "sha256": EXPECTED_SCIENTIFIC_FILES[TECOFLEX]["XRD.zip"][1],
        "container_type": "zip_xy", "parsed_units": len(xrd_rows),
        "parse_metrics": json.dumps({
            **xrd_container, "curve_count": len(xrd_rows),
            "finite_points": sum(row[1] for row in xrd_rows),
        }, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
        "decision": "parsed", "anomaly": "",
    })

    sup_path = base / "mechanical testing sup fig 1.xlsx"
    main_path = base / "mechanical testing.xlsx"
    sup = _xlsx_values(sup_path, "Sheet1 (2)")
    main = _xlsx_values(main_path, "Sheet1")
    if sup[0][10:18:2] != list(TECOFLEX_MATERIALS):
        # 工作簿标签含 TPU-NIC-x%，规范标签需单独映射，原始首标签仍必须锁定。
        if sup[0][10:18:2] != ["TPU", "TPU-NIC-2%", "TPU-NIC-5%", "TPU-NIC-10%"]:
            raise AuditBlocked("Tecoflex 补表材料标签漂移")
    if main[0][0:8:2] != ["TPU", "TPU-NIC-2%", "TPU-NIC-5%", "TPU-NIC-10%"]:
        raise AuditBlocked("Tecoflex 主表材料标签漂移")
    if sup[1][10:18] != ["Stress (MPa)", "Strain (%)"] * 4:
        raise AuditBlocked("Tecoflex 补表力学列头漂移")
    if main[1][0:8] != ["Stress (MPa)", "Strain (%)"] * 4:
        raise AuditBlocked("Tecoflex 主表力学列头漂移")

    curve_rows: list[dict[str, object]] = []
    strain_maxima: dict[str, float] = {}
    stress_offsets: dict[str, float] = {}
    sup_point_total = 0
    main_point_total = 0
    for index, material in enumerate(TECOFLEX_MATERIALS):
        sup_points: list[tuple[float, float]] = []
        main_points: list[tuple[float, float]] = []
        for row in sup[2:286]:
            stress = _finite(row[10 + 2 * index])
            strain = _finite(row[11 + 2 * index])
            if stress is None or strain is None:
                raise AuditBlocked(f"Tecoflex 补表曲线非完整数值：{material}")
            sup_points.append((strain, stress))
        for row in main[2:286]:
            stress = _finite(row[2 * index])
            strain = _finite(row[1 + 2 * index])
            if stress is None or strain is None:
                raise AuditBlocked(f"Tecoflex 主表曲线非完整数值：{material}")
            main_points.append((strain, stress))
        if len(sup_points) != 284 or len(main_points) != 284:
            raise AuditBlocked(f"Tecoflex 曲线点数漂移：{material}")
        if [point[0] for point in sup_points] != [point[0] for point in main_points]:
            raise AuditBlocked(f"Tecoflex 主/补表应变网格不再逐点相同：{material}")
        deltas = [main_y - sup_y for (_, main_y), (_, sup_y) in zip(main_points, sup_points)]
        rounded = {round(delta, 12) for delta in deltas}
        if len(rounded) != 1:
            raise AuditBlocked(f"Tecoflex 主/补表应力不再是固定平移：{material}")
        offset = next(iter(rounded))
        expected_offset = (1.592409869, 1.208398665, 1.034507082, 1.072387592)[index]
        if not math.isclose(offset, expected_offset, rel_tol=0.0, abs_tol=1e-12):
            raise AuditBlocked(f"Tecoflex 应力平移常数漂移：{material}={offset}")
        strain_maxima[material] = max(point[0] for point in sup_points)
        stress_offsets[material] = offset
        sup_point_total += len(sup_points)
        main_point_total += len(main_points)
        curve_rows.append(_curve_row(
            source=TECOFLEX, file="mechanical testing sup fig 1.xlsx", sheet="Sheet1 (2)",
            material=material, protocol="uniaxial_tension_published_zero_baseline",
            x="strain_header_percent_values_fraction_like", y="stress_mpa",
            points=len(sup_points), lineage=f"sup|{material}",
            decision="hold_unit_resolution",
            reason="原始列头为Strain (%)，但数值约0–2且100%标量位于约1.0；入模前须确认并显式换算",
            future_weight_ceiling=None,
        ))
        curve_rows.append(_curve_row(
            source=TECOFLEX, file="mechanical testing.xlsx", sheet="Sheet1",
            material=material, protocol="uniaxial_tension_shifted_copy",
            x="strain_header_percent_values_fraction_like", y="stress_mpa",
            points=len(main_points), lineage=f"main_shifted_copy_of_sup|{material}",
            decision="exclude_derived_copy",
            reason=f"应变逐点相同；主表应力=补表应力+固定常数{offset:.9f} MPa",
            future_weight_ceiling=0.0,
        ))
    if sup_point_total != 1_136 or main_point_total != 1_136:
        raise AuditBlocked("Tecoflex 主/补表曲线总点数漂移")

    # 补表 B8:F11：4 配方 × 5 报告试样；NIC-2 的第4/5值完全相同。
    scalar_values: list[tuple[str, int, float]] = []
    for material_index, material in enumerate(TECOFLEX_MATERIALS, start=7):
        row = sup[material_index]
        if str(row[0]).strip() != material:
            raise AuditBlocked(f"Tecoflex 补表标量材料标签漂移：{row[0]!r}/{material}")
        for replicate, value in enumerate(row[1:6], start=1):
            numeric = _finite(value)
            if numeric is None:
                raise AuditBlocked(f"Tecoflex 补表标量缺失：{material}/{replicate}")
            scalar_values.append((material, replicate, numeric))
    scalar_counter = Counter(value for _, _, value in scalar_values)
    duplicate_values = {value: count for value, count in scalar_counter.items() if count > 1}
    if len(scalar_values) != 20 or len(scalar_counter) != 19:
        raise AuditBlocked("Tecoflex 补表应力@100%标量计数漂移")
    if duplicate_values != {10.937081187: 2}:
        raise AuditBlocked(f"Tecoflex 补表标量重复关系漂移：{duplicate_values}")
    seen_values: set[float] = set()
    for material, replicate, value in scalar_values:
        duplicate = value in seen_values
        seen_values.add(value)
        curve_rows.append(_curve_row(
            source=TECOFLEX, file="mechanical testing sup fig 1.xlsx", sheet="Sheet1 (2)",
            material=material, protocol=f"reported_specimen_{replicate}",
            x="strain_target_100_percent", y="stress_at_100_percent_mpa",
            points=1, lineage=f"scalar|{material}|replicate_{replicate}",
            decision="exclude_duplicate" if duplicate else "candidate_reported_scalar",
            reason=("数值与同配方前一报告试样完全相同；保留reported_specimen_count但副本不得计权"
                    if duplicate else "补表直接报告的试样级标量"),
            future_weight_ceiling=0.0 if duplicate else None,
        ))

    # 主工作簿另有 22 个直接报告的试样槽位，每个槽位同时给出直径、弹性模量
    # 和 100% 应变应力。它们与补表的 20 个槽位数值集合不同，但来源没有提供
    # 跨工作簿试样 ID，故分别登记血缘且禁止简单相加为 42 个物理试样。
    main_modulus = _xlsx_values(main_path, "Elastic modulus")
    main_stress = _xlsx_values(main_path, "stress @ 100% strain")
    if str(main_modulus[0][1]).strip() != "Elastic Modulus (Mpa)":
        raise AuditBlocked("Tecoflex 主表弹性模量标题漂移")
    if str(main_stress[0][1]).strip() != "stress @ 100% strain (Mpa)":
        raise AuditBlocked("Tecoflex 主表100%应变应力标题漂移")
    expected_replicate_counts = {
        "TPU": 5, "NIC-2": 6, "NIC-5": 6, "NIC-10": 5,
    }
    main_scalar_records: list[dict[str, object]] = []
    for material_index, material in enumerate(TECOFLEX_MATERIALS):
        modulus_row = main_modulus[2 + material_index]
        diameter_row = main_modulus[10 + material_index]
        stress_row = main_stress[2 + material_index]
        if (
            str(modulus_row[1]).strip() != material
            or str(diameter_row[1]).strip() != material
            or str(stress_row[1]).strip() != material
        ):
            raise AuditBlocked(f"Tecoflex 主表试样标量材料标签漂移：{material}")
        triples: list[tuple[int, float, float, float]] = []
        for replicate, column in enumerate(range(2, 8), start=1):
            modulus = _finite(modulus_row[column])
            diameter = _finite(diameter_row[column])
            stress100 = _finite(stress_row[column])
            availability = (modulus is not None, diameter is not None, stress100 is not None)
            if len(set(availability)) != 1:
                raise AuditBlocked(
                    f"Tecoflex 主表同一试样槽位的三项标量缺失不一致：{material}/{replicate}"
                )
            if modulus is not None and diameter is not None and stress100 is not None:
                triples.append((replicate, modulus, diameter, stress100))
        if len(triples) != expected_replicate_counts[material]:
            raise AuditBlocked(
                f"Tecoflex 主表直接试样槽位数漂移：{material}={len(triples)}"
            )
        for replicate, modulus, diameter, stress100 in triples:
            lineage = f"main_scalar|{material}|replicate_{replicate}"
            main_scalar_records.append({
                "material": material,
                "replicate": replicate,
                "elastic_modulus_mpa": modulus,
                "diameter_mm": diameter,
                "stress_at_100_percent_mpa": stress100,
            })
            for observable, value, decision in (
                ("elastic_modulus_mpa", modulus, "candidate_direct_specimen_scalar"),
                ("stress_at_100_percent_mpa", stress100, "candidate_direct_specimen_scalar"),
                ("diameter_mm", diameter, "candidate_specimen_geometry_feature"),
            ):
                curve_rows.append(_curve_row(
                    source=TECOFLEX,
                    file="mechanical testing.xlsx",
                    sheet=(
                        "Elastic modulus"
                        if observable in {"elastic_modulus_mpa", "diameter_mm"}
                        else "stress @ 100% strain"
                    ),
                    material=material,
                    protocol=f"main_reported_specimen_{replicate}",
                    x="specimen_slot",
                    y=observable,
                    points=1,
                    lineage=lineage,
                    decision=decision,
                    reason=(
                        f"主工作簿直接报告值={value:.12g}；三项属性共享同一槽位血缘，"
                        "与补表20个槽位的跨表身份尚未解析"
                    ),
                    future_weight_ceiling=None,
                ))
    if len(main_scalar_records) != 22:
        raise AuditBlocked(f"Tecoflex 主表直接试样槽位总数漂移：{len(main_scalar_records)}")

    # 列头声称百分数，但四曲线最大值均约 2；结合100%标量表，数值更像应变比。
    if any(not (1.8 < value < 2.1) for value in strain_maxima.values()):
        raise AuditBlocked(f"Tecoflex 应变单位冲突判据漂移：{strain_maxima}")

    for entry_name, point_count, digest in xrd_rows:
        material = {
            "TPU control.xy": "TPU", "TPU+2%Niclosamide.xy": "NIC-2",
            "TPU+5%Niclosamide.xy": "NIC-5", "TPU+10%Niclosamide.xy": "NIC-10",
        }[entry_name]
        curve_rows.append(_curve_row(
            source=TECOFLEX, file="XRD.zip", sheet=entry_name, material=material,
            protocol="XRD", x="two_theta_degree", y="intensity_count",
            points=point_count, lineage=f"XRD|{material}|{digest}",
            decision="candidate_auxiliary_characterization",
            reason="ZIP内直接解析的共享角度网格XRD曲线", future_weight_ceiling=None,
        ))

    summary: dict[str, object] = {
        "audit_date": AUDIT_DATE,
        "audit_version": AUDIT_VERSION,
        "source": TECOFLEX,
        "record_id": SOURCE_RECORDS[TECOFLEX][0],
        "doi": SOURCE_RECORDS[TECOFLEX][1],
        "license": "CC BY 4.0",
        "scientific_file_count": 11,
        "xlsx_file_count": 10,
        "zip_file_count": 1,
        "all_workbooks_parsed": True,
        "workbook_profiles": workbook_profiles,
        "physical_specimen_count": None,
        "physical_specimen_count_reason": (
            "补表有20个报告槽位，主表有22个直接试样槽位，但来源未提供跨工作簿试样ID；"
            "两者不能相加或假定一一对应"
        ),
        "reported_specimen_slot_count_by_table": {
            "mechanical_testing_sup_fig_1_stress_at_100_percent": 20,
            "mechanical_testing_main_direct_specimen_rows": 22,
        },
        "reported_stress_at_100_percent_scalar_count": len(scalar_values),
        "numerically_distinguishable_scalar_count": len(scalar_counter),
        "main_workbook_direct_specimen_slot_count": len(main_scalar_records),
        "main_workbook_direct_observation_count": len(main_scalar_records) * 3,
        "main_workbook_direct_qois": [
            "elastic_modulus_mpa", "stress_at_100_percent_mpa", "diameter_mm",
        ],
        "duplicate_scalar": {
            "material": "NIC-2", "value_mpa": 10.937081187,
            "occurrence_count": 2, "duplicate_future_weight_ceiling": 0,
        },
        "published_zero_baseline_curves": {
            "curve_count": 4, "finite_point_count": sup_point_total,
        },
        "main_workbook_shifted_copies": {
            "curve_count": 4, "finite_point_count": main_point_total,
            "strain_grid_exactly_equal": True,
            "stress_offset_mpa_by_material": stress_offsets,
            "future_weight_ceiling": 0,
        },
        "strain_unit_conflict": {
            "header": "Strain (%)", "curve_maxima": strain_maxima,
            "evidence": "100%应力标量对应曲线横坐标约1.0，故数值更像无量纲应变比而非百分数",
            "decision": "hold_until_explicit_unit_normalization",
        },
        "xrd": {
            "curve_count": len(xrd_rows), "finite_point_count": sum(row[1] for row in xrd_rows),
            "shared_angle_grid": True,
        },
        "training_split_created": False,
        "training_weight_materialized": False,
        "training_rows_created": 0,
    }
    return summary, file_rows, curve_rows


FILE_COLUMNS = [
    "source", "filename", "bytes", "sha256", "container_type", "parsed_units",
    "parse_metrics", "decision", "anomaly",
]
CURVE_COLUMNS = [
    "source", "record_id", "doi", "file", "sheet", "material", "protocol",
    "x_observable", "y_observable", "raw_points", "usable_points", "lineage_group",
    "split_group", "decision", "decision_reason", "future_weight_ceiling",
    "training_split_created", "training_weight_materialized",
]


def main() -> int:
    before = scientific_input_snapshot()
    for source in SOURCE_NAMES:
        validate_official_capture(source)

    commercial_summary, commercial_files, commercial_curves = audit_commercial()
    tecoflex_summary, tecoflex_files, tecoflex_curves = audit_tecoflex()
    summaries = {
        COMMERCIAL: commercial_summary,
        TECOFLEX: tecoflex_summary,
    }
    file_rows = {
        COMMERCIAL: commercial_files,
        TECOFLEX: tecoflex_files,
    }
    curve_rows = {
        COMMERCIAL: commercial_curves,
        TECOFLEX: tecoflex_curves,
    }

    input_manifest_sha256 = _manifest_digest(
        (relative, size, checksum) for relative, (size, checksum) in before.items()
    )
    for source in SOURCE_NAMES:
        summaries[source]["scientific_and_capture_input_manifest_sha256"] = input_manifest_sha256
        summaries[source]["audit_output_whitelist"] = list(OUTPUT_NAMES)
        base = DATA_ROOT / source
        write_json(base / "内容审计摘要.json", summaries[source])
        atomic_write(
            base / "文件校验清单.tsv",
            render_tsv(sorted(file_rows[source], key=lambda row: str(row["filename"]).casefold()), FILE_COLUMNS),
        )
        atomic_write(
            base / "曲线审计清单.tsv",
            render_tsv(
                sorted(
                    curve_rows[source],
                    key=lambda row: (
                        str(row["file"]).casefold(), str(row["sheet"]).casefold(),
                        str(row["material"]).casefold(), str(row["protocol"]).casefold(),
                    ),
                ),
                CURVE_COLUMNS,
            ),
        )

    after = scientific_input_snapshot()
    if before != after:
        raise AuditBlocked("审计运行前后科学输入或官方捕获文件发生变化")

    output_hashes: dict[str, str] = {}
    for source in SOURCE_NAMES:
        for name in OUTPUT_NAMES:
            path = DATA_ROOT / source / name
            output_hashes[f"{source}/{name}"] = file_hash(path)
    print(json.dumps({
        "input_file_count": len(before),
        "input_manifest_sha256": input_manifest_sha256,
        "output_hashes": output_hashes,
        "training_split_created": False,
        "training_weight_materialized": False,
    }, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"BLOCKED: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc
