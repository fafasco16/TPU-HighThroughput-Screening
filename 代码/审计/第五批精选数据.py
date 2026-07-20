"""只读深审第五批两个开放的多模态聚氨酯数据来源。

本脚本不联网、不修改科学原件、不构建训练集，也不物化训练权重。它对固定
字节数和 SHA-256 的原始载荷做失败关闭核验，安全解析 TXT、XML 与 OOXML，
逐曲线建立来源、配方、运行和重复关系，并把右删失值、缺失样品标签、跨表
标签碰撞及下游碳纤维任务显式隔离。

运行：

    python 代码/审计/第五批精选数据.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import tempfile
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
from xml.etree import ElementTree

import yaml
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
REGISTRY = PROJECT_ROOT / "配置" / "候选数据源.yaml"
AUDIT_DATE = "2026-07-20"
AUDIT_VERSION = "1.0"

FISHER_NAME = "DataInBrief_聚氨酯形状记忆多模态原始数据"
BIOBASED_NAME = "Zenodo_木质素_TPU多模态数据"
SOURCE_NAMES = (FISHER_NAME, BIOBASED_NAME)
OUTPUT_NAMES = (
    "内容审计摘要.json",
    "文件校验清单.tsv",
    "曲线审计清单.tsv",
    "标量审计清单.tsv",
    "配方审计清单.tsv",
)
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source in SOURCE_NAMES
    for filename in OUTPUT_NAMES
)

FISHER_FILES = {
    "mmc1.txt": (
        238_259,
        "c12f6d47ba1316018b2217e04a8151f8fcb0e57e5c6d32c00e79830e84c713ca",
        "DMA_storage_modulus",
    ),
    "mmc2.txt": (
        263_010,
        "9645491e8ccc3bb648eae032abce6bfeb0285196dbcb0ce402186964bdbaf911",
        "DMA_tan_delta",
    ),
    "mmc3.txt": (
        42_273,
        "54833a24480665882af959633bfacad41a639df01381b77e0a00ba09dc9a1e53",
        "DSC",
    ),
    "mmc4.xlsx": (
        5_447_590,
        "95602f21bead7323df157a1bcbd84d6d1c638a33ec8a53268670ef92253037ec",
        "failure_tensile",
    ),
    "mmc5.txt": (
        275_021,
        "8ca6f66a29ea10d8e363a983a687753238b795844a8f08cd3ea49debfbb2b044",
        "TGA",
    ),
    "mmc6.xlsx": (
        41_022_679,
        "687f2837d4e4922c1a76981e031ba704c60c36de71566a85cc786f0c7f71e032",
        "uniaxial_cyclic_tensile",
    ),
    "mmc7.xml": (
        358,
        "96c71fac9d41a4b5f27be0fcd501ace626629b913b9bbf2104101f764233de69",
        "data_availability_statement",
    ),
}
BIOBASED_FILE = (
    "Biobased.xlsx",
    1_318_528,
    "5dd712d854f56a50946e195039d875e1dc22fb755b309f99e9516789d120c6d4",
)

MECHANICAL_HEADERS = (
    "Time",
    "Total cycle count",
    "Extension",
    "Load",
    "Tensile stress",
    "Tensile strain (Extension)",
)
MECHANICAL_UNITS = ("(s)", None, "(mm)", "(N)", "(MPa)", "(mm/mm)")
SHEET_LABEL_RE = re.compile(r"^SMP-?(\d+)[_-](\d+)$", re.IGNORECASE)

FILE_COLUMNS = (
    "source_directory",
    "path",
    "role",
    "bytes",
    "sha256",
    "integrity",
    "parser_state",
    "license",
    "training_split_materialized",
    "training_weight_materialized",
)
CURVE_COLUMNS = (
    "source_directory",
    "record_id",
    "source_file",
    "source_location",
    "formulation_id",
    "material_scope",
    "modality",
    "test_type",
    "point_count",
    "usable_point_count",
    "contamination_point_count",
    "quality_status",
    "axis_fields",
    "unit_status",
    "sample_mapping_status",
    "instance_key",
    "split_group",
    "dedup_status",
    "duplicate_of",
    "decision",
    "future_weight_ceiling",
    "training_split",
    "training_weight",
    "notes",
)
SCALAR_COLUMNS = (
    "source_directory",
    "record_id",
    "source_file",
    "source_location",
    "formulation_id",
    "task_role",
    "material_state",
    "result_names",
    "direct_numeric_result_count",
    "right_censored_result_count",
    "raw_censored_value",
    "unit_status",
    "decision",
    "future_weight_ceiling",
    "split_group",
    "training_split",
    "training_weight",
    "notes",
)
FORMULATION_COLUMNS = (
    "source_directory",
    "formulation_id",
    "material_family",
    "component_1",
    "component_1_fraction",
    "component_2",
    "component_2_fraction",
    "component_3",
    "component_3_fraction",
    "fraction_basis",
    "identity_mapping_status",
    "evidence",
    "split_group",
    "future_weight_ceiling",
    "training_split",
    "training_weight",
    "notes",
)


class AuditBlocked(RuntimeError):
    """原件、解析、内容结构或输出安全门禁失败。"""


@dataclass(frozen=True)
class AuditBundle:
    source_directory: str
    summary: dict[str, Any]
    files: list[dict[str, Any]]
    curves: list[dict[str, Any]]
    scalars: list[dict[str, Any]]
    formulations: list[dict[str, Any]]


def _is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(flag and attributes & flag) or bool(is_junction())


def _require_plain_directory(path: Path) -> None:
    if not path.is_dir() or _is_reparse_point(path):
        raise AuditBlocked(f"来源目录不是普通目录：{path}")
    if path.resolve(strict=True) != path.absolute():
        raise AuditBlocked(f"来源目录解析发生漂移：{path}")


def _require_plain_file(path: Path) -> None:
    if not path.is_file() or _is_reparse_point(path):
        raise AuditBlocked(f"输入不是普通文件：{path}")
    _require_plain_directory(path.parent)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _verify_frozen(path: Path, expected_size: int, expected_sha256: str) -> None:
    _require_plain_file(path)
    if path.stat().st_size != expected_size:
        raise AuditBlocked(
            f"输入字节数漂移：{path.name}，期望{expected_size}，实际{path.stat().st_size}"
        )
    actual = _sha256(path)
    if actual != expected_sha256:
        raise AuditBlocked(
            f"输入SHA-256漂移：{path.name}，期望{expected_sha256}，实际{actual}"
        )


def _safe_xlsx(path: Path) -> dict[str, int]:
    """在交给 openpyxl 之前检查 OOXML ZIP 路径、加密与活动载荷。"""

    _require_plain_file(path)
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if not infos or archive.testzip() is not None:
                raise AuditBlocked(f"XLSX ZIP完整性失败：{path.name}")
            names: set[str] = set()
            total = 0
            max_member = 0
            for info in infos:
                name = info.filename
                pure = PurePosixPath(name)
                if (
                    not name
                    or "\\" in name
                    or name.startswith("/")
                    or pure.is_absolute()
                    or any(part in {"", ".", ".."} for part in pure.parts)
                ):
                    raise AuditBlocked(f"XLSX含不安全成员路径：{path.name}:{name}")
                folded = name.casefold()
                if folded in names:
                    raise AuditBlocked(f"XLSX含大小写折叠重复成员：{path.name}:{name}")
                names.add(folded)
                if info.flag_bits & 0x1:
                    raise AuditBlocked(f"XLSX含加密成员：{path.name}:{name}")
                if info.file_size > 64 * 1024 * 1024:
                    raise AuditBlocked(f"XLSX单成员异常大：{path.name}:{name}")
                if info.file_size / max(info.compress_size, 1) > 100:
                    raise AuditBlocked(f"XLSX成员压缩比异常：{path.name}:{name}")
                total += info.file_size
                max_member = max(max_member, info.file_size)
            if total > 320 * 1024 * 1024:
                raise AuditBlocked(f"XLSX解压总量超过固定上限：{path.name}")
            forbidden = (
                "xl/vbaproject.bin",
                "xl/externallinks/",
                "xl/activex/",
                "xl/embeddings/",
                "customui/",
            )
            if any(
                name == forbidden[0] or any(name.startswith(p) for p in forbidden[1:])
                for name in names
            ):
                raise AuditBlocked(f"XLSX含宏、外链或嵌入活动对象：{path.name}")
            return {
                "zip_member_count": len(infos),
                "uncompressed_bytes": total,
                "largest_member_bytes": max_member,
            }
    except zipfile.BadZipFile as exc:
        raise AuditBlocked(f"XLSX不是有效ZIP：{path.name}") from exc


def _finite_number(value: Any, context: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise AuditBlocked(f"预期有限数值：{context}={value!r}")
    result = float(value)
    if not math.isfinite(result):
        raise AuditBlocked(f"发现非有限数值：{context}={value!r}")
    return result


def _load_registry_candidate(candidate_id: str) -> dict[str, Any]:
    _require_plain_file(REGISTRY)
    payload = yaml.safe_load(REGISTRY.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise AuditBlocked("第五批候选登记根节点不是映射")
    candidates = [
        row
        for row in payload.get("candidates", [])
        if isinstance(row, dict) and row.get("candidate_id") == candidate_id
    ]
    if len(candidates) != 1:
        raise AuditBlocked(f"候选登记不唯一：{candidate_id}")
    return candidates[0]


def _validate_source_metadata(
    directory: Path,
    candidate_id: str,
    canonical_identifier: str,
    expected_files: dict[str, tuple[int, str]],
) -> dict[str, Any]:
    metadata_path = directory / "来源元数据.json"
    manifest_path = directory / "下载清单.tsv"
    _require_plain_file(metadata_path)
    _require_plain_file(manifest_path)
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    if (
        metadata.get("candidate_id") != candidate_id
        or metadata.get("canonical_identifier") != canonical_identifier
        or metadata.get("license_spdx") != "CC-BY-4.0"
        or metadata.get("training_split_created") is not False
        or metadata.get("training_weight_materialized") is not False
    ):
        raise AuditBlocked(f"来源元数据身份、许可或训练状态错误：{directory.name}")
    rows = list(
        csv.DictReader(
            manifest_path.read_text(encoding="utf-8").splitlines(), delimiter="\t"
        )
    )
    observed = {
        row["filename"]: (int(row["size_bytes"]), row["sha256"]) for row in rows
    }
    if observed != expected_files:
        raise AuditBlocked(f"下载清单与固定原件不一致：{directory.name}")
    metadata_files = {
        row["filename"]: (int(row["size_bytes"]), row["sha256"])
        for row in metadata.get("files", [])
    }
    if metadata_files != expected_files:
        raise AuditBlocked(f"来源元数据文件清单不一致：{directory.name}")
    return metadata


def _file_row(
    source: str,
    path: Path,
    role: str,
    parser_state: str,
) -> dict[str, Any]:
    _require_plain_file(path)
    return {
        "source_directory": source,
        "path": path.name,
        "role": role,
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
        "integrity": "verified",
        "parser_state": parser_state,
        "license": "CC-BY-4.0",
        "training_split_materialized": False,
        "training_weight_materialized": False,
    }


def _curve_row(
    *,
    source: str,
    record_id: str,
    source_file: str,
    source_location: str,
    formulation_id: str,
    material_scope: str,
    modality: str,
    test_type: str,
    point_count: int,
    axis_fields: str,
    unit_status: str,
    sample_mapping_status: str,
    instance_key: str,
    split_group: str,
    dedup_status: str = "unique",
    duplicate_of: str = "",
    decision: str = "transfer_candidate",
    future_weight_ceiling: float = 0.25,
    usable_point_count: int | None = None,
    contamination_point_count: int = 0,
    quality_status: str | None = None,
    notes: str = "",
) -> dict[str, Any]:
    if usable_point_count is None:
        usable_point_count = 0 if dedup_status == "exact_duplicate" else point_count
    if quality_status is None:
        if dedup_status == "exact_duplicate":
            quality_status = "exact_duplicate"
        elif decision == "hold_missing_sample_label":
            quality_status = "mapping_hold"
        else:
            quality_status = "observed_unique"
    return {
        "source_directory": source,
        "record_id": record_id,
        "source_file": source_file,
        "source_location": source_location,
        "formulation_id": formulation_id,
        "material_scope": material_scope,
        "modality": modality,
        "test_type": test_type,
        "point_count": point_count,
        "usable_point_count": usable_point_count,
        "contamination_point_count": contamination_point_count,
        "quality_status": quality_status,
        "axis_fields": axis_fields,
        "unit_status": unit_status,
        "sample_mapping_status": sample_mapping_status,
        "instance_key": instance_key,
        "split_group": split_group,
        "dedup_status": dedup_status,
        "duplicate_of": duplicate_of,
        "decision": decision,
        "future_weight_ceiling": f"{future_weight_ceiling:.2f}",
        "training_split": "false",
        "training_weight": "false",
        "notes": notes,
    }


def _parse_fisher_text(
    path: Path, test_type: str
) -> tuple[list[dict[str, Any]], int]:
    grouped: dict[int, list[tuple[float, float]]] = defaultdict(list)
    with path.open(encoding="utf-8", newline="") as handle:
        for line_number, line in enumerate(handle, start=1):
            fields = line.split()
            if len(fields) != 3:
                raise AuditBlocked(f"Fisher文本列数错误：{path.name}:{line_number}")
            try:
                formulation = int(fields[0])
                x_value = float(fields[1])
                y_value = float(fields[2])
            except ValueError as exc:
                raise AuditBlocked(
                    f"Fisher文本数值解析失败：{path.name}:{line_number}"
                ) from exc
            if formulation not in range(1, 13) or not (
                math.isfinite(x_value) and math.isfinite(y_value)
            ):
                raise AuditBlocked(f"Fisher文本值越界：{path.name}:{line_number}")
            grouped[formulation].append((x_value, y_value))
    if set(grouped) != set(range(1, 13)):
        raise AuditBlocked(f"Fisher文本缺配方：{path.name}")
    unit_map = {
        "DMA_storage_modulus": "temperature:°C inferred from article; storage_modulus:Pa",
        "DMA_tan_delta": "temperature:°C inferred from article; tan_delta:dimensionless",
        "DSC": "temperature_and_signal_units_not_embedded_in_file",
        "TGA": "temperature:°C inferred from article; mass:%",
    }
    rows = [
        _curve_row(
            source=FISHER_NAME,
            record_id=f"fisher_{test_type.lower()}_smp_{formulation:02d}",
            source_file=path.name,
            source_location=f"column1_formula={formulation}",
            formulation_id=f"SMP-{formulation}",
            material_scope="crosslinked_aliphatic_polyurethane_shape_memory_network",
            modality="thermal",
            test_type=test_type,
            point_count=len(grouped[formulation]),
            axis_fields="temperature;response",
            unit_status=unit_map[test_type],
            sample_mapping_status="formulation_level_no_physical_specimen_key",
            instance_key=f"{path.name}|SMP-{formulation}",
            split_group=f"doi:10.1016/j.dib.2020.106294|SMP-{formulation}",
            future_weight_ceiling=0.20,
            notes="按配方编号分组；同一配方序列点不得跨训练/测试拆分",
        )
        for formulation in sorted(grouped)
    ]
    return rows, sum(len(points) for points in grouped.values())


def _normalized_sheet_label(value: str) -> tuple[int, int, str]:
    match = SHEET_LABEL_RE.fullmatch(value.strip())
    if not match:
        raise AuditBlocked(f"不能解析Fisher工作表标签：{value!r}")
    formulation = int(match.group(1))
    replicate = int(match.group(2))
    if formulation not in range(1, 13) or replicate < 1:
        raise AuditBlocked(f"Fisher工作表标签越界：{value!r}")
    return formulation, replicate, f"SMP-{formulation}_{replicate}"


def _parse_fisher_workbook(
    path: Path, test_type: str
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    archive_stats = _safe_xlsx(path)
    workbook = load_workbook(
        path, read_only=True, data_only=False, keep_links=False
    )
    rows: list[dict[str, Any]] = []
    labels: list[str] = []
    formulation_counts: Counter[int] = Counter()
    point_total = 0
    formula_count = 0
    error_count = 0
    formatted_blank_rows = 0
    leading_blank_header_rows = 0
    duplicate_counter: Counter[str] = Counter()
    sheet_payloads: dict[str, list[tuple[float, ...]]] = {}
    curve_by_sheet: dict[str, dict[str, Any]] = {}
    missing_terminal_marker_runs: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            formulation, _replicate, normalized = _normalized_sheet_label(sheet.title)
            duplicate_counter[normalized] += 1
            occurrence = duplicate_counter[normalized]
            labels.append(normalized)
            formulation_counts[formulation] += 1
            leading_rows = [
                tuple(row)
                for row in sheet.iter_rows(min_row=1, max_row=min(4, sheet.max_row), values_only=True)
            ]
            header_positions = [
                index for index, row in enumerate(leading_rows, start=1) if row == MECHANICAL_HEADERS
            ]
            if len(header_positions) != 1:
                raise AuditBlocked(f"Fisher机械表头不唯一：{path.name}:{sheet.title}")
            header_row = header_positions[0]
            leading_blank_header_rows += header_row - 1
            header = leading_rows[header_row - 1]
            units = leading_rows[header_row] if header_row < len(leading_rows) else ()
            if header != MECHANICAL_HEADERS or units != MECHANICAL_UNITS:
                raise AuditBlocked(f"Fisher机械表头漂移：{path.name}:{sheet.title}")
            values = sheet.iter_rows(min_row=header_row + 2, values_only=True)
            sheet_points = 0
            cycle_values: set[float] = set()
            sheet_payload: list[tuple[float, ...]] = []
            for row_number, value_row in enumerate(values, start=header_row + 2):
                if len(value_row) != 6:
                    raise AuditBlocked(
                        f"Fisher机械列数错误：{path.name}:{sheet.title}:{row_number}"
                    )
                if all(value is None for value in value_row):
                    formatted_blank_rows += 1
                    continue
                if any(value is None for value in value_row):
                    raise AuditBlocked(
                        f"Fisher机械存在部分空行：{path.name}:{sheet.title}:{row_number}"
                    )
                numeric = [
                    _finite_number(value, f"{path.name}:{sheet.title}:{row_number}")
                    for value in value_row
                ]
                sheet_payload.append(tuple(numeric))
                cycle_values.add(numeric[1])
                sheet_points += 1
                for value in value_row:
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                    if isinstance(value, str) and value.startswith("#"):
                        error_count += 1
            if sheet_points <= 0:
                raise AuditBlocked(f"Fisher机械工作表无数据：{path.name}:{sheet.title}")
            if test_type == "failure_tensile" and cycle_values != {0.0}:
                raise AuditBlocked(f"失效拉伸循环列应全为0：{path.name}:{sheet.title}")
            if test_type == "uniaxial_cyclic_tensile":
                expected_max = 12.5 if sheet.title == "SMP-8_2" else 13.0
                if min(cycle_values) != 0.0 or max(cycle_values) != expected_max:
                    raise AuditBlocked(f"循环拉伸半周期范围错误：{path.name}:{sheet.title}")
                if any(abs(value * 2 - round(value * 2)) > 1e-12 for value in cycle_values):
                    raise AuditBlocked(f"循环拉伸存在非半周期值：{path.name}:{sheet.title}")
            point_total += sheet_points
            sheet_payloads[sheet.title] = sheet_payload
            instance_key = f"{path.name}|{sheet.title}"
            if occurrence > 1:
                instance_key += f"|normalized_collision_{occurrence}"
            curve = _curve_row(
                    source=FISHER_NAME,
                    record_id=(
                        f"fisher_{test_type}_{path.stem}_{len(rows) + 1:03d}"
                    ),
                    source_file=path.name,
                    source_location=sheet.title,
                    formulation_id=f"SMP-{formulation}",
                    material_scope="crosslinked_aliphatic_polyurethane_shape_memory_network",
                    modality="mechanical",
                    test_type=test_type,
                    point_count=sheet_points,
                    axis_fields=(
                        "time;cycle_count;extension;load;tensile_stress;tensile_strain"
                    ),
                    unit_status="time:s;extension:mm;load:N;stress:MPa;strain:mm/mm",
                    sample_mapping_status=(
                        "normalized_label_collision_within_workbook"
                        if occurrence > 1
                        else "label_only_cross_workbook_identity_unresolved"
                    ),
                    instance_key=instance_key,
                    split_group=f"doi:10.1016/j.dib.2020.106294|SMP-{formulation}",
                    future_weight_ceiling=0.35,
                    notes=(
                        "原始工作表是一条测试运行；跨工作簿同名不证明同一物理试样"
                    ),
                )
            rows.append(curve)
            curve_by_sheet[sheet.title] = curve
            if test_type == "uniaxial_cyclic_tensile" and max(cycle_values) < 13.0:
                curve["quality_status"] = "missing_terminal_marker"
                curve["decision"] = "transfer_candidate_with_protocol_flag"
                curve["future_weight_ceiling"] = "0.30"
                curve["notes"] += "；完成10次正式循环但缺cycle_count=13终止标记，不视为缺半周期"
                missing_terminal_marker_runs.append(
                    {
                        "record_id": curve["record_id"],
                        "source_sheet": sheet.title,
                        "observed_max_cycle_count": max(cycle_values),
                        "expected_max_cycle_count": 13.0,
                    }
                )
    finally:
        workbook.close()
    expected = (
        {number: (4 if number == 2 else 3) for number in range(1, 13)}
        if test_type == "failure_tensile"
        else {number: 2 for number in range(1, 13)}
    )
    if dict(sorted(formulation_counts.items())) != expected:
        raise AuditBlocked(f"Fisher机械配方重复结构漂移：{path.name}")
    expected_runs = 37 if test_type == "failure_tensile" else 24
    expected_points = 112_141 if test_type == "failure_tensile" else 838_957
    if len(rows) != expected_runs or point_total != expected_points:
        raise AuditBlocked(
            f"Fisher机械运行或点数漂移：{path.name}={len(rows)}/{point_total}"
        )
    if formula_count or error_count:
        raise AuditBlocked(f"Fisher机械工作簿出现公式或错误单元：{path.name}")
    contamination_rows: list[dict[str, Any]] = []
    if test_type == "failure_tensile":
        if formatted_blank_rows != 266:
            raise AuditBlocked(
                f"Failure工作簿带格式空行数漂移：{formatted_blank_rows}"
            )
        base = sheet_payloads.get("SMP-8_1")
        if base is None:
            raise AuditBlocked("Failure工作簿缺SMP-8_1污染参照")
        for target, expected_suffix in (("SMP-8_3", 696), ("SMP-8_4", 1_006)):
            payload = sheet_payloads.get(target)
            if payload is None:
                raise AuditBlocked(f"Failure工作簿缺污染目标：{target}")
            suffix = _common_suffix_length(base, payload)
            if suffix != expected_suffix:
                raise AuditBlocked(
                    f"Failure复制尾段长度漂移：{target}={suffix}"
                )
            curve = curve_by_sheet[target]
            curve["usable_point_count"] = int(curve["point_count"]) - suffix
            curve["contamination_point_count"] = suffix
            curve["quality_status"] = "tail_contamination_excluded"
            curve["decision"] = "transfer_candidate_after_tail_exclusion"
            curve["future_weight_ceiling"] = "0.30"
            curve["notes"] += (
                f"；末尾{suffix}点逐单元格复制自SMP-8_1尾段，原始层保留，监督视图排除"
            )
            clean = payload[:-suffix]
            contamination_rows.append(
                {
                    "record_id": curve["record_id"],
                    "source_sheet": target,
                    "copied_from_sheet": "SMP-8_1",
                    "contamination_point_count": suffix,
                    "clean_point_count": len(clean),
                    "clean_peak_tensile_stress_MPa": max(row[4] for row in clean),
                }
            )
        peaks = [
            max(row[4] for row in sheet_payloads["SMP-8_1"]),
            contamination_rows[0]["clean_peak_tensile_stress_MPa"],
            contamination_rows[1]["clean_peak_tensile_stress_MPa"],
        ]
        expected_peaks = (5.27427, 4.26426, 3.81412)
        if any(
            not math.isclose(actual, expected, abs_tol=1e-8)
            for actual, expected in zip(peaks, expected_peaks, strict=True)
        ):
            raise AuditBlocked(f"Failure SMP-8去污染峰值漂移：{peaks}")
    elif formatted_blank_rows != 0:
        raise AuditBlocked(f"循环工作簿出现意外空行：{formatted_blank_rows}")
    if test_type == "failure_tensile" and leading_blank_header_rows != 0:
        raise AuditBlocked(f"Failure工作簿出现意外前导空行：{leading_blank_header_rows}")
    if test_type == "uniaxial_cyclic_tensile" and leading_blank_header_rows != 1:
        raise AuditBlocked(f"循环工作簿前导空行数漂移：{leading_blank_header_rows}")
    expected_missing_terminal = 1 if test_type == "uniaxial_cyclic_tensile" else 0
    if len(missing_terminal_marker_runs) != expected_missing_terminal:
        raise AuditBlocked(
            f"循环缺终止标记运行数漂移：{len(missing_terminal_marker_runs)}"
        )
    return rows, {
        **archive_stats,
        "labels": labels,
        "normalized_labels": sorted(set(labels)),
        "within_workbook_normalized_collisions": sum(
            count - 1 for count in duplicate_counter.values() if count > 1
        ),
        "formula_cell_count": formula_count,
        "error_cell_count": error_count,
        "formatted_blank_row_count": formatted_blank_rows,
        "leading_blank_header_row_count": leading_blank_header_rows,
        "copied_tail_contamination": contamination_rows,
        "missing_terminal_marker_runs": missing_terminal_marker_runs,
        "missing_terminal_marker_run_count": len(missing_terminal_marker_runs),
        "copied_tail_contamination_point_count": sum(
            row["contamination_point_count"] for row in contamination_rows
        ),
        "usable_point_count": sum(int(row["usable_point_count"]) for row in rows),
        "point_count": point_total,
        "run_count": len(rows),
    }


def _common_suffix_length(
    left: list[tuple[float, ...]], right: list[tuple[float, ...]]
) -> int:
    count = 0
    for left_row, right_row in zip(reversed(left), reversed(right)):
        if left_row != right_row:
            break
        count += 1
    return count


FISHER_COMPOSITIONS = (
    (1, 53.5, 46.5, 0.0),
    (2, 53.9, 44.5, 1.6),
    (3, 54.3, 42.5, 3.2),
    (4, 55.1, 38.4, 6.5),
    (5, 56.0, 34.1, 9.9),
    (6, 56.9, 29.7, 13.4),
    (7, 57.8, 25.1, 17.1),
    (8, 58.8, 20.4, 20.8),
    (9, 59.7, 15.6, 24.7),
    (10, 60.7, 10.6, 28.7),
    (11, 61.8, 5.4, 32.8),
    (12, 62.3, 2.7, 35.0),
)


def _fisher_formulations() -> list[dict[str, Any]]:
    rows = []
    for formulation, hdi, hped, tea in FISHER_COMPOSITIONS:
        if not math.isclose(hdi + hped + tea, 100.0, abs_tol=1e-9):
            raise AuditBlocked(f"Fisher配方不闭合：SMP-{formulation}")
        rows.append(
            {
                "source_directory": FISHER_NAME,
                "formulation_id": f"SMP-{formulation}",
                "material_family": "crosslinked_shape_memory_polyurethane",
                "component_1": "HDI",
                "component_1_fraction": hdi,
                "component_2": "HPED",
                "component_2_fraction": hped,
                "component_3": "TEA",
                "component_3_fraction": tea,
                "fraction_basis": "molar_percent",
                "identity_mapping_status": "external_companion_article_mapping",
                "evidence": "doi:10.1016/j.jmbbm.2018.08.037; doi:10.1016/j.dib.2020.106294",
                "split_group": f"doi:10.1016/j.dib.2020.106294|SMP-{formulation}",
                "future_weight_ceiling": "0.35",
                "training_split": "false",
                "training_weight": "false",
                "notes": "交联脂肪族PU形状记忆网络，不是热塑性TPU",
            }
        )
    return rows


def audit_fisher() -> AuditBundle:
    directory = DATA_ROOT / FISHER_NAME
    _require_plain_directory(directory)
    candidate = _load_registry_candidate("fisher_2020_pu_shape_memory_raw")
    if (
        candidate.get("canonical_identifier") != "doi:10.1016/j.dib.2020.106294"
        or candidate.get("scientific_role") != "polyurethane_adjacent_transfer"
    ):
        raise AuditBlocked("Fisher候选身份或科学角色漂移")
    expected_files = {
        name: (spec[0], spec[1]) for name, spec in FISHER_FILES.items()
    }
    metadata = _validate_source_metadata(
        directory,
        "fisher_2020_pu_shape_memory_raw",
        "doi:10.1016/j.dib.2020.106294",
        expected_files,
    )
    file_rows: list[dict[str, Any]] = []
    curves: list[dict[str, Any]] = []
    thermal_points = 0
    workbook_stats: dict[str, dict[str, Any]] = {}
    for name, (size, digest, role) in FISHER_FILES.items():
        path = directory / name
        _verify_frozen(path, size, digest)
        if name.endswith(".txt"):
            rows, points = _parse_fisher_text(path, role)
            curves.extend(rows)
            thermal_points += points
            parser_state = f"parsed_{role}"
        elif name.endswith(".xlsx"):
            rows, stats = _parse_fisher_workbook(path, role)
            curves.extend(rows)
            workbook_stats[name] = stats
            parser_state = f"parsed_{role}"
        else:
            try:
                root = ElementTree.parse(path).getroot()
            except ElementTree.ParseError as exc:
                raise AuditBlocked("mmc7.xml解析失败") from exc
            comment = "".join(root.itertext())
            if "raw data is available as a supplementary data file" not in comment:
                raise AuditBlocked("mmc7.xml不是预期的数据可用性说明")
            parser_state = "parsed_data_availability_only_not_analysis_script"
        file_rows.append(_file_row(FISHER_NAME, path, role, parser_state))
    for name, role in (
        ("来源元数据.json", "normalized_acquisition_metadata"),
        ("下载清单.tsv", "verified_acquisition_manifest"),
    ):
        file_rows.append(_file_row(FISHER_NAME, directory / name, role, "validated"))

    mechanical_points = sum(
        stats["point_count"] for stats in workbook_stats.values()
    )
    mechanical_labels = [
        label for stats in workbook_stats.values() for label in stats["labels"]
    ]
    normalized_labels = set(mechanical_labels)
    within_collision = sum(
        stats["within_workbook_normalized_collisions"]
        for stats in workbook_stats.values()
    )
    mechanical_usable_points = sum(
        stats["usable_point_count"] for stats in workbook_stats.values()
    )
    contamination_points = sum(
        stats["copied_tail_contamination_point_count"]
        for stats in workbook_stats.values()
    )
    formatted_blank_rows = sum(
        stats["formatted_blank_row_count"] for stats in workbook_stats.values()
    )
    missing_terminal_marker_runs = sum(
        stats["missing_terminal_marker_run_count"] for stats in workbook_stats.values()
    )
    if thermal_points != 24_805 or mechanical_points != 951_098:
        raise AuditBlocked("Fisher热/机械点数总量漂移")
    if (
        mechanical_usable_points != 949_396
        or contamination_points != 1_702
        or formatted_blank_rows != 266
        or missing_terminal_marker_runs != 1
    ):
        raise AuditBlocked("Fisher污染排除或带格式空行口径漂移")
    if len(curves) != 109:
        raise AuditBlocked("Fisher曲线容器数漂移")
    summary = {
        "audit_date": AUDIT_DATE,
        "audit_version": AUDIT_VERSION,
        "source_directory": FISHER_NAME,
        "title": metadata["title"],
        "canonical_identifier": "doi:10.1016/j.dib.2020.106294",
        "publication_identifier": "doi:10.1016/j.dib.2020.106294",
        "companion_publication_identifier": "doi:10.1016/j.jmbbm.2018.08.037",
        "license": "CC-BY-4.0",
        "origin_kind": "experiment",
        "scientific_role": "polyurethane_adjacent_transfer",
        "thermoplastic_tpu_core": False,
        "material_definition": "crosslinked HDI/HPED/TEA aliphatic polyurethane shape-memory network",
        "formulation_count": 12,
        "mechanical_run_count": 61,
        "failure_tensile_run_count": workbook_stats["mmc4.xlsx"]["run_count"],
        "cyclic_tensile_run_count": workbook_stats["mmc6.xlsx"]["run_count"],
        "thermal_curve_count": 48,
        "curve_container_count": len(curves),
        "mechanical_point_row_count": mechanical_points,
        "thermal_point_row_count": thermal_points,
        "point_row_count": mechanical_points + thermal_points,
        "copied_tail_contamination_point_count": contamination_points,
        "high_confidence_usable_point_row_count": (
            mechanical_usable_points + thermal_points
        ),
        "formatted_blank_row_count": formatted_blank_rows,
        "missing_terminal_marker_run_count": missing_terminal_marker_runs,
        "normalized_cross_workbook_label_count": len(normalized_labels),
        "within_workbook_normalized_label_collision_count": within_collision,
        "distinguishable_instance_count": len(normalized_labels) + within_collision,
        "physical_specimen_count": None,
        "physical_specimen_count_reason": (
            "跨failure与cyclic工作簿同名标签不证明同一物件；mmc6内部另有一个归一化同名碰撞"
        ),
        "analysis_script_publicly_present": False,
        "known_missing_public_item": (
            "SMP_plots_calculations.m在文章中被提及，但下载清单只有mmc1-mmc7，且mmc7为数据可用性XML"
        ),
        "workbook_safety": workbook_stats,
        "formula_cell_count": 0,
        "error_cell_count": 0,
        "independent_sample_warning": (
            "61是测试运行，不是已证明的61个独立物理试样；975903是数值行，去除复制污染后高置信点为974201"
        ),
        "split_policy": "按12个配方及来源家族分组，禁止按曲线点随机拆分",
        "training_split_materialized": False,
        "training_weight_materialized": False,
        "training_state": "transfer_hold_pending_entity_and_group_split",
        "registry_sha256": _sha256(REGISTRY),
    }
    if (
        summary["normalized_cross_workbook_label_count"] != 45
        or summary["distinguishable_instance_count"] != 46
    ):
        raise AuditBlocked("Fisher标签实体口径漂移")
    return AuditBundle(
        FISHER_NAME,
        summary,
        file_rows,
        sorted(curves, key=lambda row: row["record_id"]),
        [],
        _fisher_formulations(),
    )


def _scan_biobased_workbook(path: Path) -> dict[str, Any]:
    archive = _safe_xlsx(path)
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    nonempty = 0
    numeric = 0
    formulas = 0
    errors = 0
    charts = 0
    hidden_rows = 0
    hidden_columns = 0
    try:
        if workbook.sheetnames != [f"Sheet{number}" for number in range(1, 9)]:
            raise AuditBlocked("Biobased.xlsx工作表清单漂移")
        for sheet in workbook.worksheets:
            charts += len(sheet._charts)
            hidden_rows += sum(
                1 for details in sheet.row_dimensions.values() if details.hidden
            )
            hidden_columns += sum(
                1 for details in sheet.column_dimensions.values() if details.hidden
            )
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    nonempty += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric += 1
                    if cell.data_type == "f" or (
                        isinstance(value, str) and value.startswith("=")
                    ):
                        formulas += 1
                    if cell.data_type == "e":
                        errors += 1
        external_links = len(getattr(workbook, "_external_links", []))
    finally:
        workbook.close()
    return {
        **archive,
        "sheet_count": 8,
        "nonempty_cell_count": nonempty,
        "numeric_cell_count": numeric,
        "formula_cell_count": formulas,
        "error_cell_count": errors,
        "chart_count": charts,
        "external_link_count": external_links,
        "hidden_row_count": hidden_rows,
        "hidden_column_count": hidden_columns,
        "macro_present": False,
    }


def _numeric_pairs(
    sheet,
    x_column: int,
    y_column: int,
    start_row: int,
    end_row: int | None = None,
) -> list[tuple[float, float]]:
    pairs: list[tuple[float, float]] = []
    final = end_row or sheet.max_row
    for row in range(start_row, final + 1):
        x_value = sheet.cell(row, x_column).value
        y_value = sheet.cell(row, y_column).value
        if x_value is None and y_value is None:
            if pairs:
                break
            continue
        if not isinstance(x_value, (int, float)) or isinstance(x_value, bool):
            if pairs:
                break
            continue
        if not isinstance(y_value, (int, float)) or isinstance(y_value, bool):
            if pairs:
                break
            continue
        pairs.append(
            (
                _finite_number(x_value, f"{sheet.title}!R{row}C{x_column}"),
                _finite_number(y_value, f"{sheet.title}!R{row}C{y_column}"),
            )
        )
    if not pairs:
        raise AuditBlocked(
            f"没有解析到曲线：{sheet.title}!C{x_column}:C{y_column}"
        )
    return pairs


def _bio_curve(
    record_id: str,
    source_location: str,
    formulation_id: str,
    modality: str,
    points: int,
    *,
    unit_status: str,
    mapping: str,
    decision: str = "transfer_candidate",
    ceiling: float = 0.20,
    dedup: str = "unique",
    duplicate_of: str = "",
    notes: str = "",
) -> dict[str, Any]:
    return _curve_row(
        source=BIOBASED_NAME,
        record_id=record_id,
        source_file="Biobased.xlsx",
        source_location=source_location,
        formulation_id=formulation_id,
        material_scope="lignin_commercial_tpu_carbon_fiber_precursor_blend",
        modality=modality,
        test_type=modality,
        point_count=points,
        axis_fields={
            "FTIR": "wavenumber;transmittance",
            "DSC": "temperature;heat_flow",
            "rheology": "temperature_scan;11_instrument_channels",
            "TGA": "temperature;time;mass_percent",
            "XRD": "angle;intensity",
        }[modality],
        unit_status=unit_status,
        sample_mapping_status=mapping,
        instance_key=f"doi:10.5281/zenodo.3631551|{record_id}",
        split_group=f"doi:10.5281/zenodo.3631551|{formulation_id or 'unmapped'}",
        dedup_status=dedup,
        duplicate_of=duplicate_of,
        decision=decision,
        future_weight_ceiling=ceiling,
        notes=notes,
    )


def _parse_biobased_curves(workbook) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    curves: list[dict[str, Any]] = []
    unique_counts: Counter[str] = Counter()
    unique_points: Counter[str] = Counter()

    sheet = workbook["Sheet1"]
    ftir_pairs: dict[str, list[tuple[float, float]]] = {}
    for axis_col, label_row, data_row, response_cols, block in (
        (1, 5, 6, range(2, 8), "TcA"),
        (9, 6, 7, range(10, 17), "TcC"),
    ):
        for column in response_cols:
            label = str(sheet.cell(label_row, column).value).strip()
            pairs = _numeric_pairs(sheet, axis_col, column, data_row)
            if len(pairs) != 6_701:
                raise AuditBlocked(f"FTIR点数漂移：{block}:{label}")
            key = re.sub(r"\s+", "", label).replace("%", "pct").replace("-", "_")
            record_id = f"bio_ftir_{block.lower()}_{key.lower()}"
            duplicate_of = ""
            dedup = "unique"
            decision = "transfer_candidate"
            ceiling = 0.20
            if label == "100% PU" and "100% PU" in ftir_pairs:
                if pairs != ftir_pairs["100% PU"]:
                    raise AuditBlocked("FTIR两个100% PU参考不再逐点相同")
                duplicate_of = "bio_ftir_tca_100pctpu"
                dedup = "exact_duplicate"
                decision = "hold_exact_duplicate"
                ceiling = 0.0
            else:
                ftir_pairs[label] = pairs
                unique_counts["FTIR"] += 1
                unique_points["FTIR"] += len(pairs)
            formulation = _normalize_bio_label(label)
            curves.append(
                _bio_curve(
                    record_id,
                    f"Sheet1:{block}:{label}",
                    formulation,
                    "FTIR",
                    len(pairs),
                    unit_status="wavenumber:cm^-1;transmittance:%T",
                    mapping="workbook_label_explicit",
                    decision=decision,
                    ceiling=ceiling,
                    dedup=dedup,
                    duplicate_of=duplicate_of,
                    notes="%T超过100的点保留为基线/归一化质量标志",
                )
            )

    sheet = workbook["Sheet2"]
    dsc_specs = (
        ("bio_dsc_tca_unknown_1", 1, 2, "TcA_unknown_1"),
        ("bio_dsc_tca_unknown_2", 1, 3, "TcA_unknown_2"),
        ("bio_dsc_tca_unknown_3", 1, 4, "TcA_unknown_3"),
        ("bio_dsc_tca_shared_reference", 5, 6, "shared_reference_unknown"),
        ("bio_dsc_tcc_unknown_1", 9, 10, "TcC_unknown_1"),
        ("bio_dsc_tcc_unknown_2", 9, 11, "TcC_unknown_2"),
        ("bio_dsc_tcc_unknown_3", 9, 12, "TcC_unknown_3"),
        ("bio_dsc_tcc_shared_reference", 13, 14, "shared_reference_unknown"),
    )
    dsc_pairs: dict[str, list[tuple[float, float]]] = {}
    for record_id, x_col, y_col, label in dsc_specs:
        pairs = _numeric_pairs(sheet, x_col, y_col, 3)
        dsc_pairs[record_id] = pairs
        duplicate_of = ""
        dedup = "unique"
        if record_id == "bio_dsc_tcc_shared_reference":
            reference = dsc_pairs["bio_dsc_tca_shared_reference"]
            if pairs != reference[6:]:
                raise AuditBlocked("DSC TcC参考不再是TcA参考去掉前6点后的精确副本")
            duplicate_of = "bio_dsc_tca_shared_reference"
            dedup = "exact_duplicate"
        else:
            unique_counts["DSC"] += 1
            unique_points["DSC"] += len(pairs)
        curves.append(
            _bio_curve(
                record_id,
                f"Sheet2:C{x_col}:C{y_col}",
                "",
                "DSC",
                len(pairs),
                unit_status="temperature_and_heat_flow_units_missing",
                mapping="family_only_sample_label_missing",
                decision="hold_missing_sample_label",
                ceiling=0.0,
                dedup=dedup,
                duplicate_of=duplicate_of,
                notes=f"工作簿只提供家族块；暂用标签{label}，不得猜配方",
            )
        )

    rheology_validation: list[dict[str, float]] = []
    for sheet_name, family in (("Sheet3", "TcA"), ("Sheet4", "TcC")):
        sheet = workbook[sheet_name]
        for block_number, start_column in enumerate((1, 13, 26, 38), start=1):
            measurement_rows: list[list[float]] = []
            for row_number in range(1, sheet.max_row + 1):
                values = [sheet.cell(row_number, start_column + offset).value for offset in range(11)]
                if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
                    measurement_rows.append(
                        [
                            _finite_number(
                                value,
                                f"{sheet_name}:block{block_number}:row{row_number}",
                            )
                            for value in values
                        ]
                    )
            if not measurement_rows:
                raise AuditBlocked(f"流变块为空：{sheet_name}:block{block_number}")
            tan_errors: list[float] = []
            viscosity_errors: list[float] = []
            for values in measurement_rows:
                storage, loss, tan_delta, omega = values[:4]
                if omega != 1.0 or values[5] != values[9]:
                    raise AuditBlocked(f"流变控制/时间冗余关系漂移：{sheet_name}")
                tan_errors.append(abs(tan_delta - loss / storage))
                viscosity_errors.append(
                    abs(values[10] - math.hypot(storage, loss) / omega)
                    / max(abs(values[10]), 1.0)
                )
            if max(tan_errors) > 1e-4 or max(viscosity_errors) > 1e-4:
                raise AuditBlocked(f"流变派生通道关系漂移：{sheet_name}")
            record_id = f"bio_rheology_{family.lower()}_unknown_{block_number}"
            curves.append(
                _bio_curve(
                    record_id,
                    f"{sheet_name}:block{block_number}:C{start_column}-C{start_column + 10}",
                    "",
                    "rheology",
                    len(measurement_rows),
                    unit_status=(
                        "G/G_loss:Pa;omega:rad/s;torque:uN.m;time:s;temperature:°C;"
                        "phase:°;displacement:rad;complex_viscosity:Pa.s"
                    ),
                    mapping="family_only_sample_label_missing",
                    decision="hold_missing_sample_label",
                    ceiling=0.0,
                    notes="tan_delta、phase、complex_viscosity为仪器派生/冗余通道；Step time与Time重复",
                )
            )
            unique_counts["rheology"] += 1
            unique_points["rheology"] += len(measurement_rows)
            rheology_validation.append(
                {
                    "record_id": record_id,
                    "max_tan_delta_absolute_error": max(tan_errors),
                    "max_complex_viscosity_relative_error": max(viscosity_errors),
                }
            )

    sheet = workbook["Sheet6"]
    for family, base_column in (("TcC", 1), ("TcA", 11)):
        for offset, ratio in ((0, "70-30"), (3, "60-40"), (6, "50-50")):
            start = base_column + offset
            temperature_mass = _numeric_pairs(sheet, start, start + 2, 6)
            temperature_time = _numeric_pairs(sheet, start, start + 1, 6)
            if len(temperature_mass) != 1_355 or len(temperature_time) != 1_355:
                raise AuditBlocked(f"TGA点数漂移：{family}:{ratio}")
            formulation = f"{family}-TPU-{ratio}"
            record_id = f"bio_tga_{family.lower()}_{ratio.replace('-', '_')}"
            curves.append(
                _bio_curve(
                    record_id,
                    f"Sheet6:{family}:{ratio}:C{start}-C{start + 2}",
                    formulation,
                    "TGA",
                    len(temperature_mass),
                    unit_status="temperature_unit_missing;time_unit_missing;mass:%",
                    mapping="workbook_ratio_with_external_direction_mapping",
                    notes=(
                        "原始7-30和60-470表头保留；canonical解释为70-30和60-40并记录修正原因"
                    ),
                )
            )
            unique_counts["TGA"] += 1
            unique_points["TGA"] += len(temperature_mass)

    sheet = workbook["Sheet7"]
    for start, family, lignin_percent in (
        (1, "TcA", 60),
        (4, "TcA", 50),
        (7, "TcC", 50),
        (10, "TcC", 60),
    ):
        pairs = _numeric_pairs(sheet, start, start + 1, 8)
        if len(pairs) != 951:
            raise AuditBlocked(f"XRD点数漂移：{family}:{lignin_percent}")
        formulation = f"{family}-TPU-{lignin_percent}-{100 - lignin_percent}"
        record_id = f"bio_xrd_{family.lower()}_{lignin_percent}_{100 - lignin_percent}"
        curves.append(
            _bio_curve(
                record_id,
                f"Sheet7:C{start}:C{start + 1}",
                formulation,
                "XRD",
                len(pairs),
                unit_status="axis_and_intensity_units_missing",
                mapping="workbook_ratio_with_external_direction_mapping",
                notes="工作表原名XDR；canonical模态为XRD，原始文字保留",
            )
        )
        unique_counts["XRD"] += 1
        unique_points["XRD"] += len(pairs)

    expected_counts = {"FTIR": 12, "DSC": 7, "rheology": 8, "TGA": 6, "XRD": 4}
    expected_points = {
        "FTIR": 80_412,
        "DSC": 5_615,
        "rheology": 609,
        "TGA": 8_130,
        "XRD": 3_804,
    }
    if dict(unique_counts) != expected_counts or dict(unique_points) != expected_points:
        raise AuditBlocked(
            f"Biobased唯一曲线或点数漂移：{dict(unique_counts)}/{dict(unique_points)}"
        )
    if len(curves) != 39:
        raise AuditBlocked("Biobased展示曲线数漂移")
    return sorted(curves, key=lambda row: row["record_id"]), {
        "curve_counts_by_modality": dict(sorted(unique_counts.items())),
        "unique_points_by_modality": dict(sorted(unique_points.items())),
        "displayed_curve_count": len(curves),
        "unique_curve_count": sum(unique_counts.values()),
        "duplicate_curve_count": 2,
        "unique_point_row_count": sum(unique_points.values()),
        "rheology_derived_channel_validation": rheology_validation,
    }


def _normalize_bio_label(label: str) -> str:
    compact = re.sub(r"\s+", "", label).replace("%", "")
    if compact.upper() == "100PU":
        return "TPU-100"
    if compact.upper() == "100TCA":
        return "TcA-100"
    if compact.upper() == "100TCC":
        return "TcC-100"
    match = re.fullmatch(r"(\d+)(TcA|TcC)-(\d+)PU", compact, re.IGNORECASE)
    if not match:
        raise AuditBlocked(f"不能规范化Biobased配方标签：{label!r}")
    lignin = int(match.group(1))
    family = "TcA" if match.group(2).lower() == "tca" else "TcC"
    tpu = int(match.group(3))
    if lignin + tpu != 100:
        raise AuditBlocked(f"Biobased配方比例不闭合：{label!r}")
    return f"{family}-TPU-{lignin}-{tpu}"


def _biobased_scalars(workbook) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    sheet = workbook["Sheet5"]
    for row_number in range(3, 7):
        tpu = int(_finite_number(sheet.cell(row_number, 1).value, "Sheet5 TPU%"))
        for family, columns in (("TcC", (2, 4, 6)), ("TcA", (3, 5, 7))):
            values = [sheet.cell(row_number, column).value for column in columns]
            numeric = sum(
                isinstance(value, (int, float)) and not isinstance(value, bool)
                for value in values
            )
            censored_values = [
                value
                for value in values
                if isinstance(value, str) and re.fullmatch(r">\s*200", value.strip())
            ]
            if numeric + len(censored_values) != 3:
                raise AuditBlocked(f"Sheet5标量类型异常：R{row_number}:{family}")
            lignin = 100 - tpu
            formulation = f"{family}-TPU-{lignin}-{tpu}"
            rows.append(
                {
                    "source_directory": BIOBASED_NAME,
                    "record_id": f"bio_precursor_mechanical_{family.lower()}_{lignin}_{tpu}",
                    "source_file": "Biobased.xlsx",
                    "source_location": f"Sheet5:R{row_number}:{family}",
                    "formulation_id": formulation,
                    "task_role": "precursor_fiber_mechanical",
                    "material_state": "melt_spun_lignin_tpu_precursor_fiber",
                    "result_names": "modulus_MPa;tensile_strength_MPa;strain_percent",
                    "direct_numeric_result_count": numeric,
                    "right_censored_result_count": len(censored_values),
                    "raw_censored_value": censored_values[0] if censored_values else "",
                    "unit_status": "explicit_but_header_typographical_errors_preserved",
                    "decision": "low_weight_transfer_candidate",
                    "future_weight_ceiling": "0.15",
                    "split_group": f"doi:10.5281/zenodo.3631551|{formulation}",
                    "training_split": "false",
                    "training_weight": "false",
                    "notes": (
                        "无试样号、批次、重复数、误差条或完整曲线；>200按右删失下界保存"
                    ),
                }
            )
    sheet = workbook["Sheet8"]
    for row_number in range(3, 7):
        raw_label = str(sheet.cell(row_number, 1).value).strip()
        match = re.fullmatch(r"(TcA|TcC)-TPU\s+(\d+)-(\d+)", raw_label)
        if not match:
            raise AuditBlocked(f"Sheet8配方标签异常：{raw_label}")
        formulation = f"{match.group(1)}-TPU-{match.group(2)}-{match.group(3)}"
        values = [
            _finite_number(sheet.cell(row_number, column).value, f"Sheet8:R{row_number}")
            for column in range(2, 6)
        ]
        if len(values) != 4:
            raise AuditBlocked("Sheet8结果列数漂移")
        rows.append(
            {
                "source_directory": BIOBASED_NAME,
                "record_id": f"bio_carbon_fiber_{match.group(1).lower()}_{match.group(2)}_{match.group(3)}",
                "source_file": "Biobased.xlsx",
                "source_location": f"Sheet8:R{row_number}",
                "formulation_id": formulation,
                "task_role": "downstream_carbon_fiber",
                "material_state": "carbonized_fiber_downstream_product",
                "result_names": "tensile_strength_MPa;modulus_GPa;elongation_percent;diameter_raw",
                "direct_numeric_result_count": 4,
                "right_censored_result_count": 0,
                "raw_censored_value": "",
                "unit_status": "diameter_header_mm_physically_suspect_requires_primary_article_confirmation",
                "decision": "downstream_task_only",
                "future_weight_ceiling": "0.00",
                "split_group": f"doi:10.5281/zenodo.3631551|{formulation}",
                "training_split": "false",
                "training_weight": "false",
                "notes": "直径25-31与mm不相容；原始层保留，主论文确认前不得静默改为μm",
            }
        )
    return sorted(rows, key=lambda row: row["record_id"])


def _biobased_formulations() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    def append(
        formulation: str,
        family: str,
        component_1: str,
        fraction_1: float,
        component_2: str,
        fraction_2: float,
        notes: str,
    ) -> None:
        rows.append(
            {
                "source_directory": BIOBASED_NAME,
                "formulation_id": formulation,
                "material_family": family,
                "component_1": component_1,
                "component_1_fraction": fraction_1,
                "component_2": component_2,
                "component_2_fraction": fraction_2,
                "component_3": "",
                "component_3_fraction": "",
                "fraction_basis": "weight_percent",
                "identity_mapping_status": "workbook_plus_primary_article_and_patent_metadata",
                "evidence": (
                    "doi:10.5281/zenodo.3631551; doi:10.1021/acssuschemeng.8b01170; "
                    "US20200347232A1"
                ),
                "split_group": f"doi:10.5281/zenodo.3631551|{formulation}",
                "future_weight_ceiling": "0.20",
                "training_split": "false",
                "training_weight": "false",
                "notes": notes,
            }
        )

    append("TPU-100", "commercial_tpu_control", "Pearlthane ECO 12T95", 100.0, "", 0.0, "商业TPU对照")
    append("TcA-100", "lignin_control", "TcA", 100.0, "", 0.0, "Alcell organosolv hardwood lignin")
    append("TcC-100", "lignin_control", "TcC", 100.0, "", 0.0, "hydroxypropyl-modified Kraft hardwood lignin")
    for family, lignin_name in (
        ("TcA", "Alcell organosolv hardwood lignin"),
        ("TcC", "hydroxypropyl-modified Kraft hardwood lignin"),
    ):
        for lignin in (90, 85, 80, 70, 65, 60, 55, 50):
            tpu = 100 - lignin
            append(
                f"{family}-TPU-{lignin}-{tpu}",
                "lignin_commercial_tpu_blend",
                lignin_name,
                float(lignin),
                "Pearlthane ECO 12T95",
                float(tpu),
                "碳纤维前驱体迁移域；不是TPU合成配方",
            )
    return sorted(rows, key=lambda row: row["formulation_id"])


def audit_biobased() -> AuditBundle:
    directory = DATA_ROOT / BIOBASED_NAME
    _require_plain_directory(directory)
    candidate = _load_registry_candidate("zenodo_3631551_lignin_tpu_blends")
    if (
        candidate.get("canonical_identifier") != "doi:10.5281/zenodo.3631551"
        or candidate.get("publication_identifier")
        != "doi:10.1021/acssuschemeng.8b01170"
        or candidate.get("publication_year") != 2018
        or candidate.get("scientific_role") != "tpu_composite_transfer"
    ):
        raise AuditBlocked("Biobased候选身份、年份或科学角色漂移")
    filename, size, digest = BIOBASED_FILE
    expected_files = {filename: (size, digest)}
    metadata = _validate_source_metadata(
        directory,
        "zenodo_3631551_lignin_tpu_blends",
        "doi:10.5281/zenodo.3631551",
        expected_files,
    )
    path = directory / filename
    _verify_frozen(path, size, digest)
    workbook_stats = _scan_biobased_workbook(path)
    if (
        workbook_stats["nonempty_cell_count"] != 150_907
        or workbook_stats["numeric_cell_count"] != 150_638
        or any(
            workbook_stats[key]
            for key in (
                "formula_cell_count",
                "error_cell_count",
                "chart_count",
                "external_link_count",
                "hidden_row_count",
                "hidden_column_count",
            )
        )
    ):
        raise AuditBlocked(f"Biobased工作簿结构统计漂移：{workbook_stats}")
    # 后续需要按列做确定性随机访问；read_only 工作表的 cell() 会反复流式扫描，
    # 因而这里使用已通过 ZIP 安全门且仅约 6.4 MB 解压量的普通只读语义对象。
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        curves, curve_stats = _parse_biobased_curves(workbook)
        scalars = _biobased_scalars(workbook)
    finally:
        workbook.close()
    formulations = _biobased_formulations()
    if len(formulations) != 19 or len(scalars) != 12:
        raise AuditBlocked("Biobased配方或标量记录数漂移")
    numeric_results = sum(int(row["direct_numeric_result_count"]) for row in scalars)
    censored_results = sum(int(row["right_censored_result_count"]) for row in scalars)
    summary = {
        "audit_date": AUDIT_DATE,
        "audit_version": AUDIT_VERSION,
        "source_directory": BIOBASED_NAME,
        "title": metadata["title"],
        "canonical_identifier": "doi:10.5281/zenodo.3631551",
        "publication_identifier": "doi:10.1021/acssuschemeng.8b01170",
        "publication_year": 2018,
        "license": "CC-BY-4.0",
        "origin_kind": "experiment",
        "scientific_role": "tpu_composite_transfer",
        "thermoplastic_tpu_core": False,
        "commercial_tpu_grade": "Pearlthane ECO 12T95",
        "lignin_identities": {
            "TcA": "Alcell organosolv hardwood lignin",
            "TcC": "hydroxypropyl-modified Kraft hardwood lignin",
        },
        "distinct_material_or_blend_identity_count": len(formulations),
        **curve_stats,
        "precursor_fiber_mechanical_record_count": sum(
            row["task_role"] == "precursor_fiber_mechanical" for row in scalars
        ),
        "downstream_carbon_fiber_record_count": sum(
            row["task_role"] == "downstream_carbon_fiber" for row in scalars
        ),
        "numeric_scalar_result_count": numeric_results,
        "right_censored_scalar_result_count": censored_results,
        "workbook_safety": workbook_stats,
        "formula_cell_count": workbook_stats["formula_cell_count"],
        "error_cell_count": workbook_stats["error_cell_count"],
        "major_holds": [
            "DSC 8条展示曲线没有样品标签且轴/热流单位缺失",
            "流变8个块没有样品标签，只能保留为无监督序列",
            "Sheet8直径表头mm与25-31的物理尺度冲突，主论文确认前不改单位",
            "Sheet5两个>200应变是右删失下界，不能写成200点值",
        ],
        "independent_sample_warning": (
            "98570是37个唯一运行内的序列点；前驱体机械只有8条配方记录"
        ),
        "split_policy": "整来源/配方/共享参考/多模态同组，禁止按点随机拆分",
        "training_split_materialized": False,
        "training_weight_materialized": False,
        "training_state": "composite_transfer_hold_pending_mapping_and_group_split",
        "registry_sha256": _sha256(REGISTRY),
    }
    return AuditBundle(
        BIOBASED_NAME,
        summary,
        [
            _file_row(BIOBASED_NAME, path, "scientific_workbook", "parsed_multimodal"),
            _file_row(
                BIOBASED_NAME,
                directory / "来源元数据.json",
                "normalized_acquisition_metadata",
                "validated",
            ),
            _file_row(
                BIOBASED_NAME,
                directory / "下载清单.tsv",
                "verified_acquisition_manifest",
                "validated",
            ),
        ],
        curves,
        scalars,
        formulations,
    )


def _tsv(rows: Iterable[dict[str, Any]], columns: tuple[str, ...]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def render_outputs(bundle: AuditBundle) -> dict[str, bytes]:
    return {
        "内容审计摘要.json": (
            json.dumps(bundle.summary, ensure_ascii=False, indent=2, sort_keys=True)
            + "\n"
        ).encode("utf-8"),
        "文件校验清单.tsv": _tsv(
            sorted(bundle.files, key=lambda row: str(row["path"]).casefold()),
            FILE_COLUMNS,
        ),
        "曲线审计清单.tsv": _tsv(
            sorted(bundle.curves, key=lambda row: str(row["record_id"]).casefold()),
            CURVE_COLUMNS,
        ),
        "标量审计清单.tsv": _tsv(
            sorted(bundle.scalars, key=lambda row: str(row["record_id"]).casefold()),
            SCALAR_COLUMNS,
        ),
        "配方审计清单.tsv": _tsv(
            sorted(
                bundle.formulations,
                key=lambda row: str(row["formulation_id"]).casefold(),
            ),
            FORMULATION_COLUMNS,
        ),
    }


def _assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"输出路径不在白名单：{path}")
    _require_plain_directory(path.parent)
    if path.exists() and (not path.is_file() or _is_reparse_point(path)):
        raise AuditBlocked(f"拒绝覆盖非普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    _assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_bundle(bundle: AuditBundle) -> None:
    directory = DATA_ROOT / bundle.source_directory
    for filename, payload in render_outputs(bundle).items():
        atomic_write(directory / filename, payload)


def main() -> int:
    bundles = (audit_fisher(), audit_biobased())
    for bundle in bundles:
        write_bundle(bundle)
        print(
            json.dumps(
                {
                    "source": bundle.source_directory,
                    "curves": len(bundle.curves),
                    "scalars": len(bundle.scalars),
                    "formulations": len(bundle.formulations),
                    "training_split_materialized": False,
                    "training_weight_materialized": False,
                },
                ensure_ascii=False,
                sort_keys=True,
            )
        )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"BLOCKED: {exc}", file=os.sys.stderr)
        raise SystemExit(2) from exc
