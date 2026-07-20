"""离线复算第七批玻纤增强聚氨酯加速老化数据的科学审计。

该数据是热固性玻纤增强聚氨酯，不是 TPU 核心配方真值；它只适合作为
老化、DMA 和弯曲性能的迁移参考。脚本不联网、不改写官方归档，也不把
RAR 内容留在数据目录。只有官方 ZIP 身份、内嵌 RAR 和 23 个科学文件均
通过冻结校验，且全部关键计数复算一致后，才逐个原子替换七个白名单审计
输出。

运行：

    python 代码/审计/第七批玻纤PU老化.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import shutil
import stat
import subprocess
import tempfile
import zipfile
from contextlib import contextmanager
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Iterator, Sequence

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第七批实验_玻纤PU加速老化"
)
ARCHIVE = SOURCE_DIR / "jp5fztws54-1.zip"
ARCHIVE_SIZE = 3_694_147
ARCHIVE_SHA256 = "84dcc881697d9ead4baa25c659f4d64587f7fcedc1ef056bf333c3e7b5b6325b"
NESTED_RAR_PATH = (
    "Accelerated aging of a glass fiberpolyurethane composite for automotive "
    "applications/Data.rar"
)
NESTED_RAR_SIZE = 3_692_762
NESTED_RAR_SHA256 = "5fa18cc739d0db166ece9a06bca186d2f033bde72a2f2bbaa487b6a5be02d39a"
NESTED_RAR_CRC32 = 0x66B76262

AUDIT_VERSION = "batch7-experimental-audit-v1-reproducible"
CAPTURED_AT = "2026-07-21T18:30:00+08:00"

OUTPUT_NAMES = (
    "内容审计摘要.json",
    "文件校验清单.tsv",
    "三点弯曲曲线审计清单.tsv",
    "DMA运行审计清单.tsv",
    "字段字典.tsv",
    "官方元数据快照.json",
    "第七批候选源对比.tsv",
)
OUTPUT_WHITELIST = frozenset(SOURCE_DIR / name for name in OUTPUT_NAMES)

EXPECTED_3PB_CURVE_OCCURRENCES = 38
EXPECTED_3PB_UNIQUE_CURVES = 29
EXPECTED_3PB_DUPLICATE_OCCURRENCES = 9
EXPECTED_3PB_UNIQUE_POINTS = 96_255
EXPECTED_3PB_DISPLAYED_POINTS = 126_459
EXPECTED_3PB_PLACEHOLDERS = 2_668
EXPECTED_DMA_RUNS = 17
EXPECTED_DMA_POINTS = 39_097
EXPECTED_DMA_COMPLETE_POINTS = 36_800
EXPECTED_DMA_PARTIAL_POINTS = 2_297
EXPECTED_DMA_SENTINELS = 1
EXPECTED_DMA_PARTIAL_FILE = "W70_10d.txt"
MAX_TAN_DELTA_IDENTITY_ERROR = 2.5e-7

# 精确成员集合把 7z 解包限制在已冻结官方字节所对应的 23 个文件。
EXPECTED_PAYLOAD: dict[str, tuple[int, str]] = {
    "Data/3PB/3PB_Bf40.xlsx": (384_688, "48a5ce5be182d2d05fd7b548a8db00aa50c7c7363285e640c4ec75b9d913aae3"),
    "Data/3PB/3PB_Bf55.xlsx": (448_258, "b992386489064bb19dcd21a08ebaec05330a6e7b71cb1760c2bb642552d4c9e4"),
    "Data/3PB/3PB_Bf70.xlsx": (784_538, "a469d5f2a6cfd9e88fefe469c2942943e10e58f5dd3be66be307d0fcb537fbbb"),
    "Data/3PB/3PB_Bf_Gasoline_70.xlsx": (485_998, "bfb149d6866618f6343112709f97624c6cdc8181ae8e12fec2227633c1683043"),
    "Data/3PB/3PB_Bf_Silicone_70.xlsx": (384_850, "5a934e6e491877ad7223d6d0f419d49ba803d5ed25c63f5f19a1724e704cdc8e"),
    "Data/3PB/3PB_Bf_Water_70.xlsx": (549_333, "ba51ce8a013429c99ddfe1704f36c6294295d4837261187b210c710347b3d087"),
    "Data/DMA/Bf40_10d.txt": (110_702, "5171502712da9aecf85e4593ed210677e91e24bdec2b8133a18abd0f129b54e3"),
    "Data/DMA/Bf40_20d.txt": (110_707, "64e4a3440d64fca69d71c749c8e4e0d961292f1cbbb4def81b0aae737f79f6ef"),
    "Data/DMA/Bf40_30d.txt": (110_712, "77d7a4251134ab2bf2350a0090674299cf69c9ade1016ce58ec47e187bebd731"),
    "Data/DMA/Bf40_40d.txt": (110_718, "09f1477c73387a7555795837ffb82865185c4dc1d8ecd48053fd12aa5e2234c9"),
    "Data/DMA/Bf70_10d.txt": (110_746, "5096b1f066580b18a166151848ec8b571a7c7550f964deb46ba8d318287082ff"),
    "Data/DMA/Bf70_1yr.txt": (110_534, "b1d63d92420c997934138cbaef31b4afd470c45528cf3012d475ec7c1b61ec3e"),
    "Data/DMA/Bf70_1yr_core.txt": (110_605, "215ba5b30a29046cc4372e14f867f5bff7b54a1a55777a4a50e2dc6e3b5aacec"),
    "Data/DMA/Bf70_20d.txt": (110_745, "3cf91e2dd7c2c47dfbb43b85ed0eac6b4d5fec7327ae187c0691b9f633e81dae"),
    "Data/DMA/Bf70_30d.txt": (110_590, "f6f07520134d08f20941e9c6968ee14709f3336d6396af4862e2fc7c0614ee97"),
    "Data/DMA/Bf70_40d.txt": (110_531, "678b41f50384675f2babf3a0d2d5f1269dae522decd78148324bfc26187b790f"),
    "Data/DMA/Bf70_40d_core.txt": (110_674, "b5800ec53a6bd4612aedb0cbd9f9fea4bd1ee1a6ca8acaf2b13bcc90ea78c582"),
    "Data/DMA/W70_10d.txt": (89_945, "9218349e954449c598876b3b178b7bf3228ee613b7426b1245fb6846e86f81a8"),
    "Data/DMA/W70_1yr.txt": (111_748, "f2932b35c4ede331cf7e5a7f8445294378146fdfe0115aeba61ccf67d4e8c9ce"),
    "Data/DMA/W70_1yr_dried.txt": (111_793, "2eb6b060f8a712358cf0ca1a495d38eaa5b30321c96e8db722857267d201a37b"),
    "Data/DMA/W70_20d.txt": (110_721, "aba75001193183d8b60a4e2e25f211ad23bb856d2791913fbd6b132f2452b749"),
    "Data/DMA/W70_30d.txt": (110_738, "f7f86356ef32ad9b94f0702637e815ce6aa983242504ec2ca10198020efbb73f"),
    "Data/DMA/W70_40d.txt": (110_729, "6cd898516a24464c7e9e8a2fc00165d6fc8964441217f658336a27dd753efd4b"),
}


class AuditBlocked(RuntimeError):
    """归档身份、文件安全或科学事实不满足冻结协议。"""


def same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & flag) or bool(is_junction())


def assert_plain_chain(path: Path, stop: Path) -> None:
    if path != stop and stop not in path.parents:
        raise AuditBlocked(f"路径越出项目根：{path}")
    cursor = path
    while True:
        if is_reparse_point(cursor):
            raise AuditBlocked(f"拒绝符号链接或重解析点：{cursor}")
        if cursor == stop:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not same_path(path, resolved):
        raise AuditBlocked(f"目录缺失、不是普通目录或经链接解析：{path}")
    assert_plain_chain(path, PROJECT_ROOT)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not same_path(path, resolved):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    assert_plain_chain(path, PROJECT_ROOT)


def assert_output_allowed(path: Path) -> None:
    if not any(same_path(path, allowed) for allowed in OUTPUT_WHITELIST):
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    require_directory(path.parent)
    if path.exists() and (not path.is_file() or is_reparse_point(path)):
        raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    """在目标同卷临时写入、fsync 后以 os.replace 原子替换。"""
    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if not temporary.is_file() or is_reparse_point(temporary):
            raise AuditBlocked(f"审计临时输出异常：{temporary}")
        assert_output_allowed(path)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def hash_stream(handle: BinaryIO) -> str:
    digest = hashlib.sha256()
    for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
        digest.update(block)
    return digest.hexdigest()


def file_hash(path: Path) -> str:
    with path.open("rb") as handle:
        return hash_stream(handle)


def _safe_zip_member(name: str) -> bool:
    normalized = name.replace("\\", "/")
    pure = PurePosixPath(normalized)
    return not (
        pure.is_absolute()
        or ".." in pure.parts
        or re.match(r"^[A-Za-z]:", normalized)
        or normalized.startswith("//")
    )


def locate_7z() -> str:
    candidates = [shutil.which("7z"), shutil.which("7zz")]
    program_files = os.environ.get("ProgramFiles")
    if program_files:
        candidates.append(str(Path(program_files) / "7-Zip/7z.exe"))
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return candidate
    raise AuditBlocked("未找到 7z；无法只读解开官方 ZIP 内的 RAR 科学载荷")


@contextmanager
def extracted_payload() -> Iterator[Path]:
    """校验官方 ZIP 后，把固定 RAR 暂时解到系统临时目录。"""
    require_directory(SOURCE_DIR)
    require_file(ARCHIVE)
    if ARCHIVE.stat().st_size != ARCHIVE_SIZE or file_hash(ARCHIVE) != ARCHIVE_SHA256:
        raise AuditBlocked("官方 ZIP 大小或 SHA256 漂移")

    with tempfile.TemporaryDirectory(prefix="tpu_batch7_audit_") as temporary_name:
        temporary = Path(temporary_name)
        rar_path = temporary / "Data.rar"
        payload_root = temporary / "payload"
        payload_root.mkdir()
        with zipfile.ZipFile(ARCHIVE) as archive:
            members = archive.infolist()
            if len(members) != 1 or members[0].filename != NESTED_RAR_PATH:
                raise AuditBlocked("官方 ZIP 成员集合漂移")
            member = members[0]
            if not _safe_zip_member(member.filename) or member.is_dir():
                raise AuditBlocked("官方 ZIP 成员路径不安全")
            if member.flag_bits & 0x1:
                raise AuditBlocked("官方 ZIP 成员被加密")
            if (
                member.file_size != NESTED_RAR_SIZE
                or member.CRC != NESTED_RAR_CRC32
                or archive.testzip() is not None
            ):
                raise AuditBlocked("内嵌 RAR 的大小、CRC 或 ZIP 完整性漂移")
            with archive.open(member, "r") as source, rar_path.open("wb") as target:
                shutil.copyfileobj(source, target, length=4 * 1024 * 1024)
        if file_hash(rar_path) != NESTED_RAR_SHA256:
            raise AuditBlocked("内嵌 RAR SHA256 漂移")

        command = [locate_7z(), "x", str(rar_path), f"-o{payload_root}", "-y", "-bd", "-bb0"]
        completed = subprocess.run(
            command,
            shell=False,
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if completed.returncode != 0:
            raise AuditBlocked(f"7z 解包失败：{completed.stderr.strip()}")

        actual_files: dict[str, Path] = {}
        for path in payload_root.rglob("*"):
            if is_reparse_point(path):
                raise AuditBlocked(f"RAR 解包产生链接或重解析点：{path}")
            if path.is_file():
                relative = path.relative_to(payload_root).as_posix()
                if not _safe_zip_member(relative):
                    raise AuditBlocked(f"RAR 解包路径不安全：{relative}")
                actual_files[relative] = path
        if set(actual_files) != set(EXPECTED_PAYLOAD):
            missing = sorted(set(EXPECTED_PAYLOAD) - set(actual_files))
            extra = sorted(set(actual_files) - set(EXPECTED_PAYLOAD))
            raise AuditBlocked(f"RAR 科学文件集合漂移；缺失={missing}，新增={extra}")
        for relative, (expected_size, expected_hash) in EXPECTED_PAYLOAD.items():
            path = actual_files[relative]
            if path.stat().st_size != expected_size or file_hash(path) != expected_hash:
                raise AuditBlocked(f"RAR 科学文件身份漂移：{relative}")
        yield payload_root


def _number(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    return None


def _float_text(value: float | None) -> str:
    if value is None:
        return ""
    return format(value, ".15g")


def _curve_context(filename: str, sheet: str) -> tuple[str, int, int, str]:
    if "Bf40" in filename:
        temperature = 40
    elif "Bf55" in filename:
        temperature = 55
    else:
        temperature = 70
    if sheet.startswith("Virgin"):
        return "none_unaged_reference", temperature, 0, "unaged_reference"
    if sheet.startswith("Oven"):
        return "dry_oven_control", temperature, 0, "thermal_control"
    duration = 365 if sheet == "1year" else int(re.search(r"\d+", sheet).group())
    if "Gasoline" in filename:
        medium = "gasoline"
    elif "Silicone" in filename:
        medium = "DOT5_silicone_brake_fluid"
    elif "Water" in filename:
        medium = "water"
    else:
        medium = "DOT4_glycol_brake_fluid"
    return medium, temperature, duration, "aged_exposure"


def audit_three_point_bending(payload_root: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    first_by_hash: dict[str, str] = {}
    xlsx_paths = sorted((payload_root / "Data/3PB").glob("*.xlsx"), key=lambda p: p.name)
    for path in xlsx_paths:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            for sheet in workbook.worksheets:
                labels = [sheet.cell(1, column).value for column in (1, 2)]
                units = [sheet.cell(2, column).value for column in (1, 2)]
                if labels != ["Strain", "Stress"] or units != ["%", "MPa"]:
                    raise AuditBlocked(f"3PB 字段头或单位漂移：{path.name}/{sheet.title}")
                points: list[tuple[float, float]] = []
                placeholders = invalid = formulas = 0
                for cells in sheet.iter_rows(min_row=3, max_col=2):
                    if any(cell.data_type == "f" for cell in cells):
                        formulas += 1
                    raw_strain, raw_stress = (cell.value for cell in cells)
                    if raw_strain is None and raw_stress is None:
                        continue
                    if str(raw_strain).strip() == "--" and str(raw_stress).strip() == "--":
                        placeholders += 1
                        continue
                    strain, stress = _number(raw_strain), _number(raw_stress)
                    if strain is None or stress is None:
                        invalid += 1
                        continue
                    points.append((strain, stress))
                if not points or invalid or formulas:
                    raise AuditBlocked(
                        f"3PB 非数值或公式异常：{path.name}/{sheet.title}; "
                        f"points={len(points)}, invalid={invalid}, formulas={formulas}"
                    )
                # 无末尾换行的 17 位有效数字表示是本批首次人工审计采用的
                # 曲线指纹协议；保持它可让新脚本复现既有 curve_sha256。
                canonical = "\n".join(f"{a:.17g}\t{b:.17g}" for a, b in points).encode("ascii")
                curve_hash = hashlib.sha256(canonical).hexdigest()
                occurrence_id = f"jp5f_3pb_{path.stem}_{sheet.title}"
                duplicate_of = first_by_hash.get(curve_hash, "")
                if not duplicate_of:
                    first_by_hash[curve_hash] = occurrence_id
                medium, temperature, duration, state = _curve_context(path.name, sheet.title)
                rows.append(
                    {
                        "curve_occurrence_id": occurrence_id,
                        "source_file": path.name,
                        "sheet": sheet.title,
                        "raw_condition_label": f"{path.stem}|{sheet.title}",
                        "exposure_medium": medium,
                        "aging_temperature_C": temperature,
                        "aging_duration_days": duration,
                        "condition_state": state,
                        "strain_unit": "percent",
                        "stress_unit": "MPa",
                        "point_count": len(points),
                        "strain_min": min(a for a, _ in points),
                        "strain_max": max(a for a, _ in points),
                        "stress_min": min(b for _, b in points),
                        "stress_max": max(b for _, b in points),
                        "placeholder_tail_rows": placeholders,
                        "other_invalid_rows": invalid,
                        "formula_cell_count": formulas,
                        "curve_sha256": curve_hash,
                        "decision": "duplicate_reference" if duplicate_of else "candidate_reference",
                        "duplicate_of": duplicate_of,
                    }
                )
        finally:
            workbook.close()
    return rows


def _parse_dma_lines(path: Path) -> tuple[list[list[float]], int, bool]:
    text = path.read_bytes().decode("cp1252", errors="strict")
    rows: list[list[float]] = []
    sentinels = 0
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        tokens = [token for token in re.split(r"[\t ;]+", stripped.replace(",", ".")) if token]
        try:
            values = [float(token) for token in tokens]
        except ValueError:
            continue
        if not values:
            continue
        if len(values) == 4 and values == [-1.0, -1.0, 0.0, 0.0]:
            sentinels += 1
            continue
        if len(values) not in (4, 5) or not all(math.isfinite(value) for value in values):
            raise AuditBlocked(f"DMA 数值行列数或有限性异常：{path.name}: {stripped[:120]}")
        rows.append(values)
    widths = {len(row) for row in rows}
    if not rows or len(widths) != 1:
        raise AuditBlocked(f"DMA 数据为空或列宽混合：{path.name}")
    return rows, sentinels, text.lstrip().startswith("CLOSED")


def _dma_context(filename: str) -> tuple[str, int, int, str, str]:
    stem = Path(filename).stem
    medium = "water" if stem.startswith("W") else "DOT4_glycol_brake_fluid"
    temperature = 40 if stem.startswith("Bf40") else 70
    duration = 365 if "1yr" in stem else int(re.search(r"_(\d+)d", stem).group(1))
    location = "core" if "_core" in stem else "bulk"
    post = "dried_100C" if "_dried" in stem else "as_aged"
    return medium, temperature, duration, location, post


def audit_dma(payload_root: Path) -> list[dict[str, object]]:
    audited: list[dict[str, object]] = []
    for path in sorted((payload_root / "Data/DMA").glob("*.txt"), key=lambda p: p.name):
        rows, sentinels, has_header = _parse_dma_lines(path)
        width = len(rows[0])
        if width == 4 and path.name != EXPECTED_DMA_PARTIAL_FILE:
            raise AuditBlocked(f"意外的 DMA 四通道文件：{path.name}")
        if width == 5 and path.name == EXPECTED_DMA_PARTIAL_FILE:
            raise AuditBlocked("W70_10d 不再是冻结的部分四通道结构")
        temperatures = [row[1] for row in rows]
        storage = [row[2] for row in rows]
        tan_delta = [row[4] if width == 5 else row[3] for row in rows]
        global_index = max(range(len(rows)), key=lambda index: tan_delta[index])
        high_indices = [index for index, value in enumerate(temperatures) if 80 <= value <= 160]
        high_index = max(high_indices, key=lambda index: tan_delta[index])
        identity_error: float | None = None
        if width == 5:
            errors = [abs(row[4] - row[3] / row[2]) for row in rows if row[2] != 0]
            identity_error = max(errors)
            if identity_error >= MAX_TAN_DELTA_IDENTITY_ERROR:
                raise AuditBlocked(f"tanδ=loss/storage 恒等式不成立：{path.name}")
        expected_header = path.name in {"W70_1yr.txt", "W70_1yr_dried.txt"}
        if has_header != expected_header:
            raise AuditBlocked(f"DMA 内嵌字段头集合漂移：{path.name}")
        medium, temperature, duration, location, post = _dma_context(path.name)
        audited.append(
            {
                "run_id": f"jp5f_dma_{path.stem}",
                "source_file": path.name,
                "exposure_medium": medium,
                "aging_temperature_C": temperature,
                "aging_duration_days": duration,
                "sampling_location": location,
                "post_condition": post,
                "point_count": len(rows),
                "column_count": width,
                "field_mapping": (
                    "time_min|temperature_C|storage_modulus_MPa|loss_modulus_MPa|tan_delta"
                    if width == 5
                    else "time_min|temperature_C|storage_modulus_MPa|tan_delta"
                ),
                "mapping_evidence": "embedded_header" if has_header else "cross_file_schema_plus_tan_delta_identity",
                "temperature_min_C": min(temperatures),
                "temperature_max_C": max(temperatures),
                "storage_modulus_min_MPa": min(storage),
                "storage_modulus_max_MPa": max(storage),
                "global_tan_delta_peak_temperature_C": temperatures[global_index],
                "global_tan_delta_peak": tan_delta[global_index],
                "high_temp_peak_80_160C_temperature_C": temperatures[high_index],
                "high_temp_peak_80_160C_tan_delta": tan_delta[high_index],
                "tan_delta_identity_max_abs_error": identity_error,
                "sentinel_rows_excluded": sentinels,
                "file_sha256": EXPECTED_PAYLOAD[f"Data/DMA/{path.name}"][1],
                "decision": "candidate_partial_missing_loss_modulus" if width == 4 else "candidate_reference",
                "note": (
                    "全局峰可能是降解诱发次峰，不能未经峰分离直接等同Tg"
                    if path.name.startswith("Bf70")
                    else ""
                ),
            }
        )
    return audited


def _validate_frozen_facts(curves: Sequence[dict[str, object]], dma: Sequence[dict[str, object]]) -> dict[str, int]:
    duplicate_count = sum(bool(row["duplicate_of"]) for row in curves)
    unique_points = sum(int(row["point_count"]) for row in curves if not row["duplicate_of"])
    displayed_points = sum(int(row["point_count"]) for row in curves)
    placeholders = sum(int(row["placeholder_tail_rows"]) for row in curves)
    complete_points = sum(int(row["point_count"]) for row in dma if row["column_count"] == 5)
    partial_points = sum(int(row["point_count"]) for row in dma if row["column_count"] == 4)
    sentinels = sum(int(row["sentinel_rows_excluded"]) for row in dma)
    facts = {
        "three_point_bending_curve_occurrences": len(curves),
        "three_point_bending_unique_curves": len(curves) - duplicate_count,
        "three_point_bending_duplicate_occurrences": duplicate_count,
        "three_point_bending_unique_points": unique_points,
        "three_point_bending_displayed_points": displayed_points,
        "three_point_bending_placeholder_rows_excluded": placeholders,
        "dma_runs": len(dma),
        "dma_points": sum(int(row["point_count"]) for row in dma),
        "dma_complete_five_channel_points": complete_points,
        "dma_partial_four_channel_points": partial_points,
        "dma_sentinel_rows_excluded": sentinels,
    }
    expected = {
        "three_point_bending_curve_occurrences": EXPECTED_3PB_CURVE_OCCURRENCES,
        "three_point_bending_unique_curves": EXPECTED_3PB_UNIQUE_CURVES,
        "three_point_bending_duplicate_occurrences": EXPECTED_3PB_DUPLICATE_OCCURRENCES,
        "three_point_bending_unique_points": EXPECTED_3PB_UNIQUE_POINTS,
        "three_point_bending_displayed_points": EXPECTED_3PB_DISPLAYED_POINTS,
        "three_point_bending_placeholder_rows_excluded": EXPECTED_3PB_PLACEHOLDERS,
        "dma_runs": EXPECTED_DMA_RUNS,
        "dma_points": EXPECTED_DMA_POINTS,
        "dma_complete_five_channel_points": EXPECTED_DMA_COMPLETE_POINTS,
        "dma_partial_four_channel_points": EXPECTED_DMA_PARTIAL_POINTS,
        "dma_sentinel_rows_excluded": EXPECTED_DMA_SENTINELS,
    }
    if facts != expected:
        raise AuditBlocked(f"关键科学统计漂移：actual={facts}, expected={expected}")
    partial_files = [str(row["source_file"]) for row in dma if row["column_count"] == 4]
    if partial_files != [EXPECTED_DMA_PARTIAL_FILE]:
        raise AuditBlocked(f"DMA 部分通道文件漂移：{partial_files}")
    return facts


def _tsv(rows: Sequence[dict[str, object]], columns: Sequence[str]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=list(columns), delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                column: _float_text(value) if isinstance(value, float) else value
                for column, value in row.items()
            }
        )
    return buffer.getvalue().encode("utf-8")


def _json(value: object) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=False) + "\n").encode("utf-8")


def _official_metadata() -> dict[str, object]:
    return {
        "captured_at": CAPTURED_AT,
        "snapshot_type": "frozen_official_metadata_for_offline_reproduction",
        "urls": {
            "landing_page": "https://data.mendeley.com/datasets/jp5fztws54/1",
            "versions": "https://data.mendeley.com/public-api/datasets/jp5fztws54/versions",
            "snapshot": "https://data.mendeley.com/public-api/datasets/jp5fztws54/snapshot/1",
            "zip_metadata": "https://data.mendeley.com/api/datasets-v2/datasets/jp5fztws54/zip?version=1",
            "stable_download": "https://data.mendeley.com/public-api/zip/jp5fztws54/download/1",
            "article": "https://doi.org/10.1016/j.polymertesting.2019.01.008",
            "institutional_bibliography": "https://repositum.tuwien.at/handle/20.500.12708/142637",
        },
        "dataset": {
            "id": "jp5fztws54",
            "version": 1,
            "doi": "10.17632/jp5fztws54.1",
            "title": "Accelerated aging of a glass fiber/polyurethane composite for automotive applications",
            "description": "Data from Dynamic-Mechanical and Three-Point-Bending flexural tests of virgin and aged specimens.",
            "published": "2018-08-30T07:41:05.491Z",
            "license": "CC BY 4.0",
            "contributors": ["Ileana Panaitescu", "Thomas Koch", "Vasiliki-Maria Archodoulaki"],
            "categories": ["Thermal Analysis", "Flexural Testing"],
        },
        "zip_metadata": {
            "size": ARCHIVE_SIZE,
            "sha256_hash": ARCHIVE_SHA256,
            "status": "FINISH",
            "created_on": "2020-02-14T14:47:33.506",
            "modified_on": "2022-12-05T22:58:32.571667",
        },
    }


def _candidate_rows() -> list[dict[str, object]]:
    return [
        {
            "rank": 1,
            "candidate_id": "mendeley_jp5fztws54_v1",
            "title": "Accelerated aging of a glass fiber/polyurethane composite for automotive applications",
            "dataset_doi": "10.17632/jp5fztws54.1",
            "article_doi": "10.1016/j.polymertesting.2019.01.008",
            "repository": "Mendeley Data",
            "license": "CC BY 4.0",
            "direct_download": "https://data.mendeley.com/public-api/zip/jp5fztws54/download/1",
            "format": "ZIP>RAR>6 XLSX+17 TXT",
            "verified_size_bytes": ARCHIVE_SIZE,
            "verified_sha256": ARCHIVE_SHA256,
            "verified_scale": "29条唯一3PB曲线/96255点；17条DMA温扫/39097点",
            "value": "补温度/介质老化、DMA、弯曲；配方为polyether polyol+MDI 100:132与E-glass RTM",
            "risk": "热固性GFRPU而非TPU；3PB无重复ID；15个DMA文件字段由同格式推断",
            "recommendation": "接入可靠参考/迁移层；按来源和基体分组并降低核心TPU任务权重",
        },
        {
            "rank": 2,
            "candidate_id": "figshare_31550614_v1",
            "title": "Influence of SLS Process Parameters on Cushion Properties and Sustainability of TPU Lattice Structures for Automotive Seating",
            "dataset_doi": "10.6084/m9.figshare.31550614.v1",
            "article_doi": "",
            "repository": "Figshare",
            "license": "CC BY 4.0",
            "direct_download": "https://ndownloader.figshare.com/files/62459923",
            "format": "XLSX+DOCX",
            "verified_size_bytes": 63_965,
            "verified_sha256": "ed327722f815fbf171a5a9b126c3438a8a85fb492429b0fb7016c0c286a7747a",
            "verified_scale": "25组SLS参数×3重复=75试样；Load25/65、SAG、HLR、质量",
            "value": "TPU工艺—缓冲应用性能；输入LP/SS/HD/LT/ED",
            "risk": "商业TPU牌号未公开；只有处理后指标而非完整压缩曲线",
            "recommendation": "接入工艺应用可靠参考层，权重低于化学身份闭合数据",
        },
        {
            "rank": 3,
            "candidate_id": "acs_figshare_32019436_s002",
            "title": "Machine Learning-Driven Prediction and Interpretation of Glass Transition Temperature in Polyurethanes",
            "dataset_doi": "10.1021/acsapm.5c04524.s002",
            "article_doi": "10.1021/acsapm.5c04524",
            "repository": "ACS Figshare",
            "license": "CC BY-NC 4.0",
            "direct_download": "https://ndownloader.figshare.com/files/63743962",
            "format": "XLSX",
            "verified_size_bytes": 33_447,
            "verified_sha256": "5662d58b87b5f49aabdfc023c4d861c5601873cdae5f63065a3c233df63d920e",
            "verified_scale": "83条非空PU Tg记录；9类HS、9类SS",
            "value": "结构/组成—Tg监督，含HS/SS、MW、HS wt%、密度和分子描述符",
            "risk": "二次文献汇编；链扩剂身份不明；CC BY-NC；需逐行去重",
            "recommendation": "条件接入二次汇编参考层，低权重且按原文献同折",
        },
    ]


def build_outputs(
    curves: Sequence[dict[str, object]],
    dma: Sequence[dict[str, object]],
    facts: dict[str, int],
) -> tuple[dict[str, object], dict[str, bytes]]:
    summary = {
        "audit_version": AUDIT_VERSION,
        "captured_at": CAPTURED_AT,
        "source_title": "Accelerated aging of a glass fiber/polyurethane composite for automotive applications",
        "dataset_doi": "10.17632/jp5fztws54.1",
        "dataset_version": 1,
        "article_doi": "10.1016/j.polymertesting.2019.01.008",
        "license": "CC BY 4.0",
        "authors": ["Ileana Panaitescu", "Thomas Koch", "Vasiliki-Maria Archodoulaki"],
        "material": {
            "family": "E-glass-fiber-reinforced thermoset polyurethane composite",
            "matrix_components": ["polyether polyol", "diphenylmethane diisocyanate (MDI)"],
            "component_mixing_ratio": "Part A:Part B = 100:132",
            "manufacturing": "resin transfer molding (RTM)",
            "pure_resin_post_cure": "2 h at 150 degC",
            "tpu_core_eligible": False,
            "role": "reliable_polyurethane_composite_aging_reference",
        },
        "payload": {
            "official_zip_bytes": ARCHIVE_SIZE,
            "official_zip_sha256": ARCHIVE_SHA256,
            "nested_scientific_files": len(EXPECTED_PAYLOAD),
            "xlsx_files": 6,
            "dma_txt_files": 17,
            **facts,
        },
        "admission_recommendation": {
            "scope": "可靠参考/迁移层",
            "status": "admitted_reference_after_governance_registration",
            "weight_policy": "任务相关动态权重；不得把热固性GFRPU当作TPU核心化学真值",
            "grouping_key": "dataset DOI + material family + exposure medium + temperature; same base material remains in one outer fold",
        },
        "risks": [
            "不是热塑性TPU，严禁作为TPU核心化学结构真值",
            "三条基准曲线在4个工作簿各重复4次，已隔离9个重复出现",
            "3PB无试样/重复编号，29条是唯一曲线通道而非29个独立试样",
            "2668行显式--尾部占位已排除；未擅自校正应力偏置",
            "15个DMA文件无字段头，映射由两个完整导出文件和tanδ=loss/storage共同验证",
            "W70_10d缺loss modulus列，只用于storage modulus和tan delta任务",
        ],
        "references": [
            "10.17632/jp5fztws54.1",
            "10.1016/j.polymertesting.2019.01.008",
            "https://data.mendeley.com/datasets/jp5fztws54/1",
            "https://repositum.tuwien.at/handle/20.500.12708/142637",
        ],
    }

    file_rows: list[dict[str, object]] = [
        {
            "archive_path": ARCHIVE.name,
            "role": "official_mendeley_zip",
            "bytes": ARCHIVE_SIZE,
            "sha256": ARCHIVE_SHA256,
            "decision": "retained_original",
            "note": "Mendeley固定v1下载字节；与官方zip metadata一致",
        },
        {
            "archive_path": NESTED_RAR_PATH,
            "role": "nested_rar",
            "bytes": NESTED_RAR_SIZE,
            "sha256": NESTED_RAR_SHA256,
            "decision": "container_only",
            "note": "ZIP CRC32=66b76262；RAR内含6 XLSX和17 TXT",
        },
    ]
    for relative, (size, checksum) in EXPECTED_PAYLOAD.items():
        file_rows.append(
            {
                "archive_path": relative,
                "role": "three_point_bending_xlsx" if relative.endswith(".xlsx") else "dma_txt",
                "bytes": size,
                "sha256": checksum,
                "decision": "scientific_payload",
                "note": "从官方ZIP内嵌RAR无损解出，仅用于审计；未改写原文件",
            }
        )

    curve_columns = list(curves[0])
    dma_columns = list(dma[0])
    field_rows = [
        {"record_family": "3PB", "raw_field": "Strain", "unit": "%", "canonical_field": "engineering_strain_percent", "status": "explicit_header", "note": "保持百分数，不除以100写回原证据"},
        {"record_family": "3PB", "raw_field": "Stress", "unit": "MPa", "canonical_field": "flexural_stress_MPa", "status": "explicit_header", "note": "存在初始偏置；不擅自归零"},
        {"record_family": "DMA", "raw_field": "column_1", "unit": "min", "canonical_field": "time_min", "status": "two_files_explicit_15_files_schema_inferred", "note": "W70_1yr及dried文件含Sig1头"},
        {"record_family": "DMA", "raw_field": "column_2", "unit": "degC", "canonical_field": "temperature_C", "status": "two_files_explicit_15_files_schema_inferred", "note": "同仪器同点数温扫"},
        {"record_family": "DMA", "raw_field": "column_3", "unit": "MPa", "canonical_field": "storage_modulus_MPa", "status": "two_files_explicit_15_files_schema_inferred", "note": ""},
        {"record_family": "DMA", "raw_field": "column_4", "unit": "MPa", "canonical_field": "loss_modulus_MPa", "status": "present_in_16_runs_missing_in_W70_10d", "note": "W70_10d只能用于E-prime与tan-delta任务"},
        {"record_family": "DMA", "raw_field": "column_5_or_4", "unit": "dimensionless", "canonical_field": "tan_delta", "status": "validated", "note": "16个五列文件满足tan_delta≈loss/storage，最大绝对误差<2.5e-7"},
    ]
    candidate_columns = list(_candidate_rows()[0])
    outputs = {
        "内容审计摘要.json": _json(summary),
        "文件校验清单.tsv": _tsv(file_rows, list(file_rows[0])),
        "三点弯曲曲线审计清单.tsv": _tsv(curves, curve_columns),
        "DMA运行审计清单.tsv": _tsv(dma, dma_columns),
        "字段字典.tsv": _tsv(field_rows, list(field_rows[0])),
        "官方元数据快照.json": _json(_official_metadata()),
        "第七批候选源对比.tsv": _tsv(_candidate_rows(), candidate_columns),
    }
    if set(outputs) != set(OUTPUT_NAMES):
        raise AuditBlocked("审计输出集合与白名单不一致")
    return summary, outputs


def run_audit(write_outputs: bool = True) -> dict[str, object]:
    """执行完整离线审计；可关闭写入用于测试和只读复核。"""
    with extracted_payload() as payload_root:
        curves = audit_three_point_bending(payload_root)
        dma = audit_dma(payload_root)
        facts = _validate_frozen_facts(curves, dma)
        summary, outputs = build_outputs(curves, dma, facts)
    if write_outputs:
        for name in OUTPUT_NAMES:
            atomic_write(SOURCE_DIR / name, outputs[name])
    return {"summary": summary, "curves": curves, "dma": dma, "outputs": outputs}


def main() -> None:
    result = run_audit(write_outputs=True)
    payload = result["summary"]["payload"]
    print(
        json.dumps(
            {
                "status": "ok",
                "source": str(ARCHIVE),
                "outputs": list(OUTPUT_NAMES),
                "three_point_bending_unique_curves": payload["three_point_bending_unique_curves"],
                "dma_points": payload["dma_points"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
