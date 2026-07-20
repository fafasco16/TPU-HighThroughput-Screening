"""离线复算第七批孢子填充 TPU Source Data 的力学审计。

该脚本只读取出版社原始 ``源数据.xlsx``，并原子写入三个轻量审计文件：

* ``内容审计摘要.json``
* ``曲线审计清单.tsv``
* ``标量审计清单.tsv``

它不会联网，不改写原始工作簿，也不修改项目配置、总账或文档。工作簿中
曲线表与 Figure 4 标量表的重复顺序并非处处一致，因此两个清单分别保留
各自在源表中的重复顺序，不凭行号构造未经证明的跨表样本配对。

运行：

    python 代码/审计/第七批孢子填充TPU.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import stat
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = (
    PROJECT_ROOT
    / "数据/原始/外部数据/新增开放数据/第七批补充材料_孢子填充TPU"
)
SOURCE_XLSX = SOURCE_DIR / "源数据.xlsx"
SOURCE_XLSX_BYTES = 6_539_183
SOURCE_XLSX_SHA256 = "39ed4045fd71f89547d6c54977a69838b0895c815cb7b7b68bc0dee039a012a3"

DOI = "10.1038/s41467-024-47132-8"
LICENSE = "CC BY 4.0"
AUDIT_VERSION = "batch7-spore-filled-tpu-v1"
CURVE_SHEET = "Supplmentary Figure 9A, B"
SCALAR_SHEET = "Figure 4A-H"
OUTPUT_NAMES = ("内容审计摘要.json", "曲线审计清单.tsv", "标量审计清单.tsv")

EXPECTED_CURVES = 36
EXPECTED_POINTS = 280_288
EXPECTED_FORMULATION_CONDITIONS = 12
EXPECTED_REPLICATES_PER_CONDITION = 3
EXPECTED_SCALARS = 144
EXPECTED_CURVE_POINT_COUNTS = (
    7_076, 7_783, 7_193, 7_663, 7_862, 8_030,
    7_682, 8_354, 7_779, 8_128, 8_073, 8_215,
    7_314, 7_537, 7_612, 7_500, 6_985, 7_313,
    7_529, 7_179, 7_174, 7_705, 7_679, 7_887,
    7_789, 8_152, 7_933, 7_485, 8_600, 8_459,
    8_179, 8_122, 8_210, 8_162, 7_904, 8_041,
)

CURVE_COLUMNS = (
    "curve_id",
    "source_sheet",
    "source_figure",
    "spore_type",
    "spore_wt_pct",
    "formulation_id",
    "replicate_source_order",
    "strain_column",
    "stress_column",
    "first_excel_row",
    "last_excel_row",
    "point_count",
    "partial_pair_rows",
    "strain_min_pct",
    "strain_max_pct",
    "stress_min_MPa",
    "stress_max_MPa",
    "trapezoid_toughness_MJ_m3",
)

SCALAR_COLUMNS = (
    "scalar_id",
    "source_sheet",
    "source_figure",
    "spore_type",
    "spore_wt_pct",
    "formulation_id",
    "replicate_source_order",
    "metric",
    "value",
    "unit",
    "source_cell",
)

LOADINGS = (0.0, 0.2, 0.4, 0.6, 0.8, 1.0)

# (figure, spore type, first value row, metric, unit)
SCALAR_BLOCKS = (
    ("Figure 4A", "WT", 4, "toughness", "MJ/m3"),
    ("Figure 4B", "WT", 11, "tensile_stress", "MPa"),
    ("Figure 4C", "WT", 18, "elongation_at_break", "%"),
    ("Figure 4D", "WT", 25, "young_modulus", "MPa"),
    ("Figure 4E", "HST", 32, "toughness", "MJ/m3"),
    ("Figure 4F", "HST", 39, "tensile_stress", "MPa"),
    ("Figure 4G", "HST", 46, "elongation_at_break", "%"),
    ("Figure 4H", "HST", 53, "young_modulus", "MPa"),
)


class AuditBlocked(RuntimeError):
    """原始文件身份、版式或科学计数不满足冻结协议。"""


def _same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def _is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & flag) or bool(is_junction())


def _require_plain_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not _same_path(path, resolved) or _is_reparse_point(path):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    if PROJECT_ROOT.resolve() not in path.resolve().parents:
        raise AuditBlocked(f"文件越出项目根：{path}")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_source(path: Path = SOURCE_XLSX) -> dict[str, object]:
    _require_plain_file(path)
    actual_bytes = path.stat().st_size
    actual_sha256 = sha256(path)
    if actual_bytes != SOURCE_XLSX_BYTES:
        raise AuditBlocked(
            f"源数据.xlsx 字节数漂移：{actual_bytes} != {SOURCE_XLSX_BYTES}"
        )
    if actual_sha256.lower() != SOURCE_XLSX_SHA256:
        raise AuditBlocked(
            f"源数据.xlsx SHA256 漂移：{actual_sha256} != {SOURCE_XLSX_SHA256}"
        )
    return {"bytes": actual_bytes, "sha256": actual_sha256}


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def _format_number(value: object) -> str:
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            raise AuditBlocked(f"拒绝写出非有限数：{value}")
        return format(value, ".15g")
    return str(value)


def _curve_specs() -> list[dict[str, object]]:
    specs: list[dict[str, object]] = []
    # 第一面板 B:AP 为 WT，第二面板 AS:CG 为 HST；每个配方占 7 列，
    # 内含 3 对 strain/stress，末列为空白分隔。
    for spore_type, section_start, figure in (
        ("WT", 2, "Supplementary Figure 9A"),
        ("HST", 45, "Supplementary Figure 9B"),
    ):
        for loading_index, loading in enumerate(LOADINGS):
            group_start = section_start + 7 * loading_index
            for replicate in range(1, 4):
                strain_column = group_start + 2 * (replicate - 1)
                stress_column = strain_column + 1
                formulation_id = f"{spore_type}_{loading:g}wtpct"
                specs.append(
                    {
                        "spore_type": spore_type,
                        "spore_wt_pct": loading,
                        "formulation_id": formulation_id,
                        "replicate": replicate,
                        "figure": figure,
                        "strain_column": strain_column,
                        "stress_column": stress_column,
                        "points": [],
                        "first_row": None,
                        "last_row": None,
                        "partial_pair_rows": 0,
                    }
                )
    return specs


def parse_curves(workbook_path: Path = SOURCE_XLSX) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if CURVE_SHEET not in workbook.sheetnames:
            raise AuditBlocked(f"缺失曲线工作表：{CURVE_SHEET}")
        sheet = workbook[CURVE_SHEET]
        specs = _curve_specs()

        header = next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))
        for spec in specs:
            x_index = int(spec["strain_column"]) - 1
            y_index = int(spec["stress_column"]) - 1
            if header[x_index] != "Strain (%)" or header[y_index] != "Stress (MPa)":
                raise AuditBlocked(
                    "曲线列头漂移："
                    f"{get_column_letter(x_index + 1)}={header[x_index]!r}, "
                    f"{get_column_letter(y_index + 1)}={header[y_index]!r}"
                )

        for excel_row, values in enumerate(
            sheet.iter_rows(min_row=5, values_only=True), start=5
        ):
            for spec in specs:
                x_index = int(spec["strain_column"]) - 1
                y_index = int(spec["stress_column"]) - 1
                strain = values[x_index] if x_index < len(values) else None
                stress = values[y_index] if y_index < len(values) else None
                if _is_number(strain) and _is_number(stress):
                    point = (float(strain), float(stress))
                    points = spec["points"]
                    assert isinstance(points, list)
                    points.append(point)
                    if spec["first_row"] is None:
                        spec["first_row"] = excel_row
                    spec["last_row"] = excel_row
                elif strain is not None or stress is not None:
                    spec["partial_pair_rows"] = int(spec["partial_pair_rows"]) + 1

        rows: list[dict[str, object]] = []
        for index, spec in enumerate(specs, start=1):
            points = spec.pop("points")
            assert isinstance(points, list)
            if not points:
                raise AuditBlocked(f"曲线无有效点：{spec}")
            toughness = math.fsum(
                (points[position - 1][1] + points[position][1])
                * 0.5
                * (points[position][0] - points[position - 1][0])
                / 100.0
                for position in range(1, len(points))
            )
            rows.append(
                {
                    "curve_id": f"curve_{index:02d}",
                    "source_sheet": CURVE_SHEET,
                    "source_figure": spec["figure"],
                    "spore_type": spec["spore_type"],
                    "spore_wt_pct": spec["spore_wt_pct"],
                    "formulation_id": spec["formulation_id"],
                    "replicate_source_order": spec["replicate"],
                    "strain_column": get_column_letter(int(spec["strain_column"])),
                    "stress_column": get_column_letter(int(spec["stress_column"])),
                    "first_excel_row": spec["first_row"],
                    "last_excel_row": spec["last_row"],
                    "point_count": len(points),
                    "partial_pair_rows": spec["partial_pair_rows"],
                    "strain_min_pct": min(point[0] for point in points),
                    "strain_max_pct": max(point[0] for point in points),
                    "stress_min_MPa": min(point[1] for point in points),
                    "stress_max_MPa": max(point[1] for point in points),
                    "trapezoid_toughness_MJ_m3": toughness,
                }
            )
        return rows
    finally:
        workbook.close()


def parse_scalars(workbook_path: Path = SOURCE_XLSX) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SCALAR_SHEET not in workbook.sheetnames:
            raise AuditBlocked(f"缺失标量工作表：{SCALAR_SHEET}")
        sheet = workbook[SCALAR_SHEET]
        rows: list[dict[str, object]] = []
        for figure, spore_type, first_row, metric, unit in SCALAR_BLOCKS:
            if sheet.cell(first_row - 3, 1).value != figure:
                raise AuditBlocked(
                    f"标量区块标题漂移：{sheet.cell(first_row - 3, 1).coordinate}"
                )
            for replicate in range(1, 4):
                excel_row = first_row + replicate - 1
                for loading_index, loading in enumerate(LOADINGS):
                    column = 3 + loading_index
                    value = sheet.cell(excel_row, column).value
                    if not _is_number(value):
                        raise AuditBlocked(
                            f"标量缺失或非数值：{sheet.cell(excel_row, column).coordinate}={value!r}"
                        )
                    rows.append(
                        {
                            "scalar_id": f"scalar_{len(rows) + 1:03d}",
                            "source_sheet": SCALAR_SHEET,
                            "source_figure": figure,
                            "spore_type": spore_type,
                            "spore_wt_pct": loading,
                            "formulation_id": f"{spore_type}_{loading:g}wtpct",
                            "replicate_source_order": replicate,
                            "metric": metric,
                            "value": float(value),
                            "unit": unit,
                            "source_cell": sheet.cell(excel_row, column).coordinate,
                        }
                    )
        return rows
    finally:
        workbook.close()


def _validate(curves: Sequence[dict[str, object]], scalars: Sequence[dict[str, object]]) -> None:
    if len(curves) != EXPECTED_CURVES:
        raise AuditBlocked(f"曲线数漂移：{len(curves)} != {EXPECTED_CURVES}")
    point_counts = tuple(int(row["point_count"]) for row in curves)
    if point_counts != EXPECTED_CURVE_POINT_COUNTS:
        raise AuditBlocked("逐曲线点数或列顺序漂移")
    if sum(point_counts) != EXPECTED_POINTS:
        raise AuditBlocked(f"总点数漂移：{sum(point_counts)} != {EXPECTED_POINTS}")
    if any(int(row["partial_pair_rows"]) for row in curves):
        raise AuditBlocked("发现只有 strain 或只有 stress 的半缺失曲线行")

    condition_counts = Counter(str(row["formulation_id"]) for row in curves)
    if len(condition_counts) != EXPECTED_FORMULATION_CONDITIONS:
        raise AuditBlocked("配方条件数漂移")
    if set(condition_counts.values()) != {EXPECTED_REPLICATES_PER_CONDITION}:
        raise AuditBlocked("曲线重复数不再是每条件 3 次")

    if len(scalars) != EXPECTED_SCALARS:
        raise AuditBlocked(f"标量数漂移：{len(scalars)} != {EXPECTED_SCALARS}")
    scalar_counts = Counter(
        (str(row["formulation_id"]), str(row["metric"])) for row in scalars
    )
    if len(scalar_counts) != EXPECTED_FORMULATION_CONDITIONS * 4:
        raise AuditBlocked("配方×指标组合数漂移")
    if set(scalar_counts.values()) != {EXPECTED_REPLICATES_PER_CONDITION}:
        raise AuditBlocked("标量重复数不再是每条件每指标 3 次")


def _metric_summary(scalars: Sequence[dict[str, object]]) -> dict[str, object]:
    grouped: dict[str, list[float]] = defaultdict(list)
    units: dict[str, str] = {}
    for row in scalars:
        metric = str(row["metric"])
        grouped[metric].append(float(row["value"]))
        units[metric] = str(row["unit"])
    return {
        metric: {
            "unit": units[metric],
            "count": len(values),
            "min": min(values),
            "max": max(values),
            "mean": math.fsum(values) / len(values),
        }
        for metric, values in sorted(grouped.items())
    }


def build_summary(
    identity: dict[str, object],
    curves: Sequence[dict[str, object]],
    scalars: Sequence[dict[str, object]],
) -> dict[str, object]:
    return {
        "schema_version": AUDIT_VERSION,
        "dataset": "Biocomposite thermoplastic polyurethanes containing evolved bacterial spores",
        "doi": DOI,
        "license": LICENSE,
        "source_file": SOURCE_XLSX.name,
        "source_identity": identity,
        "scientific_classification": {
            "gold_layer": "Gold-实验/条件化力学曲线层",
            "evidence_type": "experiment",
            "polymer": "BASF Elastollan BCF45 polyester TPU",
            "training_guidance": "工艺、填料、曲线和力学标签高权重；分子结构任务不使用",
            "split_guidance": "按论文来源或 formulation_id 分组，禁止同配方重复跨训练/测试泄漏",
            "cross_table_replicate_warning": "曲线表与标量表的重复源顺序独立，不凭顺序强制配对",
        },
        "field_coverage": {
            "repeat_unit_smiles": "missing",
            "monomer_identity": "partial_commercial_grade_and_polyester_type_only",
            "nco_oh_ratio": "missing",
            "hard_segment_fraction": "missing",
            "filler_identity_and_loading": "complete",
            "melt_processing": "complete",
            "tensile_test_condition": "complete",
            "stress_strain_curve": "complete_with_replicates",
            "mechanical_scalars": "complete_with_replicates",
        },
        "counts": {
            "curve_sheets": 1,
            "curves": len(curves),
            "curve_points": sum(int(row["point_count"]) for row in curves),
            "partial_curve_pair_rows": sum(
                int(row["partial_pair_rows"]) for row in curves
            ),
            "formulation_conditions": len(
                {str(row["formulation_id"]) for row in curves}
            ),
            "replicates_per_condition": EXPECTED_REPLICATES_PER_CONDITION,
            "scalar_measurements": len(scalars),
            "scalar_metrics": len({str(row["metric"]) for row in scalars}),
        },
        "curve_point_count_range": {
            "min": min(int(row["point_count"]) for row in curves),
            "max": max(int(row["point_count"]) for row in curves),
        },
        "metric_summary": _metric_summary(scalars),
        "outputs": list(OUTPUT_NAMES),
    }


def _json_bytes(payload: object) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def _tsv_bytes(rows: Sequence[dict[str, object]], columns: Sequence[str]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(
        stream,
        fieldnames=list(columns),
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _format_number(row[column]) for column in columns})
    return stream.getvalue().encode("utf-8")


def _assert_output_path(path: Path) -> None:
    allowed = {SOURCE_DIR / name for name in OUTPUT_NAMES}
    if not any(_same_path(path, candidate) for candidate in allowed):
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    if not SOURCE_DIR.is_dir() or _is_reparse_point(SOURCE_DIR):
        raise AuditBlocked(f"输出目录缺失或是重解析点：{SOURCE_DIR}")
    if path.exists() and (not path.is_file() or _is_reparse_point(path)):
        raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    _assert_output_path(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=SOURCE_DIR
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if not temporary.is_file() or _is_reparse_point(temporary):
            raise AuditBlocked(f"临时输出异常：{temporary}")
        _assert_output_path(path)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def render_outputs(
    summary: dict[str, object],
    curves: Sequence[dict[str, object]],
    scalars: Sequence[dict[str, object]],
) -> dict[str, bytes]:
    return {
        "内容审计摘要.json": _json_bytes(summary),
        "曲线审计清单.tsv": _tsv_bytes(curves, CURVE_COLUMNS),
        "标量审计清单.tsv": _tsv_bytes(scalars, SCALAR_COLUMNS),
    }


def run_audit(*, write_outputs: bool = True) -> dict[str, object]:
    identity = verify_source()
    curves = parse_curves()
    scalars = parse_scalars()
    _validate(curves, scalars)
    summary = build_summary(identity, curves, scalars)
    outputs = render_outputs(summary, curves, scalars)
    if write_outputs:
        for name in OUTPUT_NAMES:
            atomic_write(SOURCE_DIR / name, outputs[name])
    return {
        "summary": summary,
        "curves": curves,
        "scalars": scalars,
        "outputs": outputs,
    }


def main() -> None:
    result = run_audit(write_outputs=True)
    counts = result["summary"]["counts"]
    print(
        "审计完成："
        f"{counts['curves']} 条曲线，{counts['curve_points']} 个点，"
        f"{counts['scalar_measurements']} 个标量；输出 {len(OUTPUT_NAMES)} 个文件。"
    )


if __name__ == "__main__":
    main()
