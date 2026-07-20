"""只读复算第三批 Mendeley TPU 三源的归档完整性与科学曲线规模。

脚本不联网、不把 ZIP 解压到原始数据树、不生成训练集。它先冻结并复核三个
来源目录的输入集合，再对归档执行 SHA256、CRC、成员路径、重复名、加密、
符号链接、压缩比和总解压量硬门；随后直接从归档内工作簿/CSV 复算试样、
历史、曲线、点数、仿真家族和数据异常。全部检查通过后，仅原子替换每个
来源目录中的三个白名单审计输出。

运行：

    python 代码/审计/新增开放数据第三批Mendeley三源.py
"""

from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import math
import os
import re
import stat
import struct
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from itertools import zip_longest
from numbers import Real
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Callable, Iterable
from urllib.parse import urlsplit
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = PROJECT_ROOT / "数据/原始" / "外部数据" / "新增开放数据"
AUDIT_DATE = "2026-07-20"
AUDIT_VERSION = "1.2"

FATIGUE = "Mendeley_商业TPU温度疲劳多工况"
FDM = "Mendeley_FDM_TPU晶格与基材力学"
EXPERIMENT_SIMULATION = "Mendeley_TPU实验仿真曲线"
SOURCE_NAMES = (FATIGUE, FDM, EXPERIMENT_SIMULATION)

OUTPUT_NAMES = ("内容审计摘要.json", "文件校验清单.tsv", "曲线审计清单.tsv")
FDM_SCALAR_OUTPUT_NAME = "标量审计清单.tsv"
OUTPUT_WHITELIST = frozenset(
    DATA_ROOT / source / filename
    for source in SOURCE_NAMES
    for filename in OUTPUT_NAMES
) | frozenset({DATA_ROOT / FDM / FDM_SCALAR_OUTPUT_NAME})

EXPECTED_ARCHIVES: dict[str, tuple[str, int, str]] = {
    FATIGUE: (
        "hc6npzvw3m-1.zip",
        41_218_060,
        "1a47a26b3c5ac93a7b56ef8e94c2e2b0308a5ceec34ed407435586e1744a65ab",
    ),
    FDM: (
        "dbzdkz95f8-1.zip",
        210_709_465,
        "3cf82a71f83cfa46925556b5e0e9a901e5d184aeeb415ed2d823cca1d4674d3c",
    ),
    EXPERIMENT_SIMULATION: (
        "kysnxmy7xw-1.zip",
        4_543_043,
        "3585c67dac25988b651999d4a9b25ca3fb55da1a25b05386fbbf8fa8a87cf55e",
    ),
}

EXPECTED_SOURCE_IDENTITIES = {
    FATIGUE: ("hc6npzvw3m", 1, "10.17632/hc6npzvw3m.1"),
    FDM: ("dbzdkz95f8", 1, "10.17632/dbzdkz95f8.1"),
    EXPERIMENT_SIMULATION: ("kysnxmy7xw", 1, "10.17632/kysnxmy7xw.1"),
}
MENDELEY_HOST = "data.mendeley.com"
ZIP_CACHE_HOST = "prod-dcd-datasets-cache-zipfiles.s3.eu-west-1.amazonaws.com"

EXPECTED_ZIP_SHAPES = {
    FATIGUE: {
        "entry_count": 336,
        "uncompressed_bytes": 44_252_151,
        "extensions": {
            ".csv": 27,
            ".ino": 1,
            ".m": 9,
            ".mat": 17,
            ".pdf": 11,
            ".txt": 1,
            ".xlsx": 269,
            ".zip": 1,
        },
    },
    FDM: {
        "entry_count": 40,
        "uncompressed_bytes": 221_356_461,
        "extensions": {".stl": 6, ".txt": 1, ".xlsx": 29, ".zs2": 4},
    },
    EXPERIMENT_SIMULATION: {
        "entry_count": 2,
        "uncompressed_bytes": 4_776_992,
        "extensions": {".xlsx": 2},
    },
}

MAX_COMPRESSION_RATIO = 10.0
MAX_UNCOMPRESSED_BYTES = 300_000_000

FILE_COLUMNS = ["归档", "成员", "未压缩字节", "压缩字节", "CRC32", "扩展名", "角色"]
CURVE_COLUMNS = [
    "来源",
    "材料",
    "试验类型",
    "条件",
    "数据角色",
    "曲线ID",
    "点数",
    "曲线SHA256",
    "试样或家族组",
    "source_summary_state",
    "source_summary_evidence",
    "source_display_id",
    "formula_target_specimen_id",
    "quality_gate",
    "training_split",
    "weight",
    "备注",
]

SCALAR_COLUMNS = [
    "来源",
    "工作簿",
    "试样组",
    "试样ID",
    "observable",
    "value",
    "unit",
    "definition_id",
    "scalar_lineage_class",
    "source_summary_state",
    "source_summary_evidence",
    "source_display_id",
    "formula_target_specimen_id",
    "quality_gate",
    "formula_evidence",
    "training_split",
    "weight",
    "备注",
]


class AuditBlocked(RuntimeError):
    """输入、归档完整性或科学语义不满足冻结审计协议。"""


def same_path(left: Path, right: Path) -> bool:
    return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(
        os.path.abspath(str(right))
    )


def is_reparse_point(path: Path) -> bool:
    try:
        details = path.lstat()
    except FileNotFoundError:
        return False
    attributes = getattr(details, "st_file_attributes", 0)
    flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    is_junction = getattr(path, "is_junction", lambda: False)
    return path.is_symlink() or bool(attributes & flag) or bool(is_junction())


def assert_plain_chain(path: Path, stop: Path) -> None:
    if path != stop and stop not in path.parents:
        raise AuditBlocked(f"路径越出项目根：{path}")
    cursor = path
    while True:
        if is_reparse_point(cursor):
            raise AuditBlocked(f"拒绝符号链接或重解析点：{cursor}")
        if cursor == stop:
            return
        cursor = cursor.parent


def require_directory(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_dir() or not same_path(path, resolved):
        raise AuditBlocked(f"目录缺失、不是普通目录或经链接解析：{path}")
    assert_plain_chain(path, PROJECT_ROOT)


def require_file(path: Path) -> None:
    resolved = path.resolve(strict=True)
    if not path.is_file() or not same_path(path, resolved):
        raise AuditBlocked(f"文件缺失、不是普通文件或经链接解析：{path}")
    assert_plain_chain(path, PROJECT_ROOT)


def assert_output_allowed(path: Path) -> None:
    if path not in OUTPUT_WHITELIST:
        raise AuditBlocked(f"拒绝写入白名单以外路径：{path}")
    require_directory(path.parent)
    if path.exists() and (not path.is_file() or is_reparse_point(path)):
        raise AuditBlocked(f"审计输出不是普通文件：{path}")


def atomic_write(path: Path, payload: bytes) -> None:
    assert_output_allowed(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".audit.tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        if not temporary.is_file() or is_reparse_point(temporary):
            raise AuditBlocked(f"审计临时输出异常：{temporary}")
        if not same_path(temporary, temporary.resolve(strict=True)):
            raise AuditBlocked(f"审计临时输出经链接解析：{temporary}")
        assert_output_allowed(path)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def hash_stream(handle: BinaryIO, algorithm: str = "sha256") -> str:
    value = hashlib.new(algorithm)
    for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
        value.update(block)
    return value.hexdigest()


def file_hash(path: Path, algorithm: str = "sha256") -> str:
    require_file(path)
    with path.open("rb") as handle:
        return hash_stream(handle, algorithm)


def manifest_hash(items: Iterable[tuple[str, int, str]]) -> str:
    value = hashlib.sha256()
    for relative, size, checksum in sorted(items):
        value.update(relative.encode("utf-8"))
        value.update(b"\0")
        value.update(str(size).encode("ascii"))
        value.update(b"\0")
        value.update(checksum.encode("ascii"))
        value.update(b"\n")
    return value.hexdigest()


def scientific_input_snapshot() -> dict[str, tuple[int, str]]:
    snapshot: dict[str, tuple[int, str]] = {}
    for source in SOURCE_NAMES:
        base = DATA_ROOT / source
        require_directory(base)
        archive_name, expected_size, expected_sha256 = EXPECTED_ARCHIVES[source]
        expected_names = {archive_name, "官方API元数据.json", "官方文件清单.tsv"}
        actual_names: set[str] = set()
        for path in sorted(base.iterdir(), key=lambda item: item.name.casefold()):
            if is_reparse_point(path):
                raise AuditBlocked(f"来源根含符号链接或重解析点：{path}")
            if path in OUTPUT_WHITELIST:
                continue
            if not path.is_file():
                raise AuditBlocked(f"来源根出现未登记目录或特殊对象：{path}")
            if path.name.endswith((".part", ".tmp", ".audit.tmp")):
                raise AuditBlocked(f"来源根出现未完成临时文件：{path}")
            actual_names.add(path.name)
            relative = path.relative_to(PROJECT_ROOT).as_posix()
            snapshot[relative] = (path.stat().st_size, file_hash(path))
        if actual_names != expected_names:
            raise AuditBlocked(
                f"{source} 输入集合漂移：缺失={sorted(expected_names-actual_names)}，"
                f"多余={sorted(actual_names-expected_names)}"
            )
        archive_path = base / archive_name
        if (
            archive_path.stat().st_size != expected_size
            or file_hash(archive_path) != expected_sha256
        ):
            raise AuditBlocked(f"冻结归档大小或 SHA256 不匹配：{source}/{archive_name}")
    return snapshot


def validate_official_capture(source: str) -> None:
    """闭合本地官方捕获、清单、冻结归档与来源身份。"""
    base = DATA_ROOT / source
    dataset_id, version, doi = EXPECTED_SOURCE_IDENTITIES[source]
    archive_name, archive_size, archive_sha256 = EXPECTED_ARCHIVES[source]
    metadata_path = base / "官方API元数据.json"
    manifest_path = base / "官方文件清单.tsv"
    require_file(metadata_path)
    require_file(manifest_path)
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    if (
        metadata.get("provider") != "Mendeley Data"
        or metadata.get("dataset_id") != dataset_id
        or int(metadata.get("version", -1)) != version
        or str(metadata.get("doi", "")).casefold() != doi.casefold()
        or str((metadata.get("license") or {}).get("short_name", "")).casefold()
        != "cc by 4.0"
    ):
        raise AuditBlocked(f"Mendeley 官方元数据来源身份不符：{source}")
    expected_endpoints = {
        "snapshot": f"https://{MENDELEY_HOST}/public-api/datasets/{dataset_id}/snapshot/{version}",
        "versions": f"https://{MENDELEY_HOST}/public-api/datasets/{dataset_id}/versions",
        "zip_metadata": f"https://{MENDELEY_HOST}/api/datasets-v2/datasets/{dataset_id}/zip?version={version}",
        "stable_download": f"https://{MENDELEY_HOST}/public-api/zip/{dataset_id}/download/{version}",
    }
    if metadata.get("official_endpoints") != expected_endpoints:
        raise AuditBlocked(f"Mendeley 官方端点捕获不符：{source}")
    for endpoint in expected_endpoints.values():
        parsed = urlsplit(endpoint)
        if parsed.scheme != "https" or parsed.hostname != MENDELEY_HOST:
            raise AuditBlocked(f"Mendeley 官方端点越出白名单：{source}/{endpoint}")

    if metadata.get("raw_api_capture_format") != "exact_response_bytes_base64_with_sha256":
        raise AuditBlocked(f"Mendeley 官方 API 精确响应格式缺失或漂移：{source}")
    captures = metadata.get("raw_api_captures")
    if not isinstance(captures, list) or len(captures) != 3:
        raise AuditBlocked(f"Mendeley 官方 API 精确响应数量漂移：{source}")
    expected_capture_urls = {
        expected_endpoints["snapshot"],
        expected_endpoints["versions"],
        expected_endpoints["zip_metadata"],
    }
    actual_capture_urls: set[str] = set()
    for capture in captures:
        if not isinstance(capture, dict):
            raise AuditBlocked(f"Mendeley 官方 API 精确响应结构异常：{source}")
        request_url = str(capture.get("request_url", ""))
        if request_url in actual_capture_urls:
            raise AuditBlocked(f"Mendeley 官方 API 精确响应重复：{source}/{request_url}")
        actual_capture_urls.add(request_url)
        final_url = str(capture.get("final_url", ""))
        final_parsed = urlsplit(final_url)
        if final_parsed.scheme != "https" or final_parsed.hostname != MENDELEY_HOST:
            raise AuditBlocked(f"Mendeley 官方 API 最终主机越界：{source}/{final_url}")
        if int(capture.get("status", -1)) != 200:
            raise AuditBlocked(f"Mendeley 官方 API 响应状态异常：{source}/{request_url}")
        encoded = capture.get("payload_base64")
        if not isinstance(encoded, str):
            raise AuditBlocked(f"Mendeley 官方 API 响应缺少 Base64：{source}/{request_url}")
        try:
            payload = base64.b64decode(encoded.encode("ascii"), validate=True)
        except (UnicodeEncodeError, ValueError) as exc:
            raise AuditBlocked(
                f"Mendeley 官方 API Base64 非严格编码：{source}/{request_url}"
            ) from exc
        if (
            len(payload) != int(capture.get("payload_bytes", -1))
            or hashlib.sha256(payload).hexdigest()
            != str(capture.get("payload_sha256", "")).lower()
        ):
            raise AuditBlocked(f"Mendeley 官方 API 响应字节或 SHA256 不符：{source}/{request_url}")
        try:
            json.loads(payload.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise AuditBlocked(f"Mendeley 官方 API 响应不是严格 UTF-8 JSON：{source}/{request_url}") from exc
    if actual_capture_urls != expected_capture_urls:
        raise AuditBlocked(
            f"Mendeley 官方 API 请求集合漂移：{source}/"
            f"缺失={sorted(expected_capture_urls-actual_capture_urls)}/"
            f"多余={sorted(actual_capture_urls-expected_capture_urls)}"
        )
    redirect_policy = metadata.get("redirect_policy") or {}
    if (
        redirect_policy.get("initial_host") != MENDELEY_HOST
        or redirect_policy.get("allowed_zip_cache_host") != ZIP_CACHE_HOST
        or redirect_policy.get("signed_redirect_url_persisted") is not False
    ):
        raise AuditBlocked(f"Mendeley 重定向证据不符：{source}")
    archive = metadata.get("archive") or {}
    if (
        archive.get("filename") != archive_name
        or int(archive.get("bytes", -1)) != archive_size
        or str(archive.get("sha256", "")).lower() != archive_sha256
        or archive.get("zip_status") != "FINISH"
    ):
        raise AuditBlocked(f"Mendeley 元数据归档证据不符：{source}")

    with manifest_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    if len(rows) != 1:
        raise AuditBlocked(f"Mendeley 官方文件清单不是单一冻结归档：{source}")
    row = rows[0]
    expected_row = {
        "source_directory": source,
        "provider": "Mendeley Data",
        "dataset_id": dataset_id,
        "version": str(version),
        "doi": doi,
        "filename": archive_name,
        "bytes": str(archive_size),
        "sha256": archive_sha256,
        "stable_download_url": expected_endpoints["stable_download"],
        "redirect_cache_host": ZIP_CACHE_HOST,
        "local_state": "verified_present",
        "local_sha256": archive_sha256,
    }
    if row != expected_row:
        raise AuditBlocked(f"Mendeley 清单与冻结来源证据不一致：{source}")
    archive_path = base / archive_name
    if (
        archive_path.stat().st_size != archive_size
        or file_hash(archive_path) != archive_sha256
    ):
        raise AuditBlocked(f"Mendeley 归档与官方清单不一致：{source}")


def snapshot_by_source(snapshot: dict[str, tuple[int, str]]) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for source in SOURCE_NAMES:
        marker = f"/新增开放数据/{source}/"
        items = [
            (path, size, checksum)
            for path, (size, checksum) in snapshot.items()
            if marker in f"/{path}"
        ]
        result[source] = {
            "输入文件数": len(items),
            "输入总字节数": sum(size for _, size, _ in items),
            "输入清单SHA256": manifest_hash(items),
        }
    return result


def safe_zip_name(name: str) -> str:
    if "\x00" in name or re.match(r"^[A-Za-z]:", name):
        raise AuditBlocked(f"ZIP 含危险成员路径：{name!r}")
    candidate = PurePosixPath(name.replace("\\", "/"))
    if candidate.is_absolute() or ".." in candidate.parts:
        raise AuditBlocked(f"ZIP 含路径穿越成员：{name!r}")
    normalized = candidate.as_posix()
    if normalized in {"", "."}:
        raise AuditBlocked(f"ZIP 含空成员路径：{name!r}")
    return normalized


def fatigue_file_role(name: str) -> str:
    lower = name.lower()
    if "/load_frame_measurements/" in lower:
        if any(
            f"/{material.lower()}/" in lower
            for material in (
                "Elastollan1154D",
                "Elastollan1164D",
                "Elastollan1174D",
                "Elastollan1195A",
                "Texin245",
            )
        ):
            return "TPU原始力学曲线"
        if "/santoprene" in lower:
            return "TPV类比曲线_不纳入TPU"
    if name.endswith("Final_data_all_experiments.xlsx"):
        return "实验元数据与派生汇总"
    if lower.endswith("readme.txt"):
        return "说明文件"
    if lower.endswith(".pdf"):
        return "材料或项目文档"
    return "项目辅助文件"


def fdm_file_role(name: str) -> str:
    lower = name.lower()
    parts = PurePosixPath(name).parts
    basename = parts[-1].lower()
    if "/tpu/" in lower and basename.endswith(".xlsx"):
        if "summary" in basename:
            return "TPU派生汇总_不作原始曲线"
        return "TPU原始力学工作簿"
    if "/tpu/" in lower and basename.endswith(".zs2"):
        return "TPU专有试验原始文件"
    if "/pla/" in lower:
        return "PLA类比数据_不纳入TPU"
    if basename.endswith(".stl"):
        return "晶格几何"
    if basename == "readme.txt":
        return "说明文件"
    return "辅助文件"


def experiment_simulation_file_role(name: str) -> str:
    if name.endswith("Raw Data Experiment TPU.xlsx"):
        return "TPU实验原始曲线"
    if name.endswith("Comparison Excel Experiment and Simulation.xlsx"):
        return "TPU仿真比较工作簿"
    return "未分类"


def audit_zip(
    source: str, archive_path: Path, role_function: Callable[[str], str]
) -> tuple[dict[str, object], list[dict[str, object]]]:
    expected = EXPECTED_ZIP_SHAPES[source]
    rows: list[dict[str, object]] = []
    with zipfile.ZipFile(archive_path) as archive:
        infos = archive.infolist()
        if len(infos) != expected["entry_count"]:
            raise AuditBlocked(f"{source} ZIP 条目数漂移：{len(infos)}")
        normalized_names: set[str] = set()
        folded_names: set[str] = set()
        extension_counts: Counter[str] = Counter()
        total_uncompressed = 0
        max_ratio = 0.0
        directory_count = 0
        for info in infos:
            normalized = safe_zip_name(info.filename)
            folded = normalized.casefold()
            if normalized in normalized_names or folded in folded_names:
                raise AuditBlocked(f"{source} ZIP 含重复或大小写冲突路径：{normalized}")
            normalized_names.add(normalized)
            folded_names.add(folded)
            unix_mode = (info.external_attr >> 16) & 0xFFFF
            if stat.S_ISLNK(unix_mode):
                raise AuditBlocked(f"{source} ZIP 含符号链接成员：{normalized}")
            if info.flag_bits & 0x1:
                raise AuditBlocked(f"{source} ZIP 含加密成员：{normalized}")
            if info.is_dir():
                directory_count += 1
                continue
            extension = PurePosixPath(normalized).suffix.lower()
            extension_counts[extension] += 1
            total_uncompressed += info.file_size
            if info.compress_size == 0:
                ratio = 1.0 if info.file_size == 0 else math.inf
            else:
                ratio = info.file_size / info.compress_size
            max_ratio = max(max_ratio, ratio)
            if ratio > MAX_COMPRESSION_RATIO:
                raise AuditBlocked(f"{source} ZIP 成员压缩比过高：{normalized}={ratio}")
            rows.append(
                {
                    "归档": archive_path.name,
                    "成员": normalized,
                    "未压缩字节": info.file_size,
                    "压缩字节": info.compress_size,
                    "CRC32": f"{info.CRC:08x}",
                    "扩展名": extension,
                    "角色": role_function(normalized),
                }
            )
        if directory_count != 0:
            raise AuditBlocked(f"{source} ZIP 目录条目数漂移：{directory_count}")
        if total_uncompressed != expected["uncompressed_bytes"]:
            raise AuditBlocked(
                f"{source} ZIP 解压总字节漂移：{total_uncompressed}/"
                f"{expected['uncompressed_bytes']}"
            )
        if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
            raise AuditBlocked(f"{source} ZIP 超过解压总量硬门")
        if dict(sorted(extension_counts.items())) != expected["extensions"]:
            raise AuditBlocked(
                f"{source} ZIP 扩展名构成漂移：{dict(sorted(extension_counts.items()))}"
            )
        bad_member = archive.testzip()
        if bad_member is not None:
            raise AuditBlocked(f"{source} ZIP CRC 失败：{bad_member}")
    rows.sort(key=lambda item: str(item["成员"]).casefold())
    return (
        {
            "归档": archive_path.name,
            "归档字节数": archive_path.stat().st_size,
            "归档SHA256": file_hash(archive_path),
            "成员数": len(rows),
            "目录条目数": directory_count,
            "总解压字节数": total_uncompressed,
            "扩展名计数": dict(sorted(extension_counts.items())),
            "最大压缩比": round(max_ratio, 6),
            "CRC全部通过": True,
            "路径安全": True,
            "无重复路径": True,
            "无加密或符号链接": True,
        },
        rows,
    )


def is_finite_number(value: object) -> bool:
    return isinstance(value, Real) and not isinstance(value, bool) and math.isfinite(float(value))


def update_numeric_digest(value: hashlib._Hash, numbers: Iterable[float]) -> None:  # type: ignore[name-defined]
    for number in numbers:
        value.update(struct.pack("!d", float(number)))


def open_workbook_member(archive: zipfile.ZipFile, member: str):
    return load_workbook(
        io.BytesIO(archive.read(member)),
        read_only=True,
        data_only=True,
        keep_links=False,
    )


def parse_fatigue_xlsx(data: bytes, member: str) -> tuple[int, str, bool]:
    workbook = load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False
    )
    try:
        if len(workbook.worksheets) != 1:
            raise AuditBlocked(f"原始历史工作簿工作表数异常：{member}")
        worksheet = workbook.worksheets[0]
        value = hashlib.sha256()
        points = 0
        previous_time: float | None = None
        monotonic = True
        for row in worksheet.iter_rows(
            min_row=4, min_col=1, max_col=4, values_only=True
        ):
            if not any(cell is not None for cell in row):
                continue
            if not all(is_finite_number(cell) for cell in row):
                raise AuditBlocked(f"原始历史含不完整或非数值四通道行：{member}")
            numbers = tuple(float(cell) for cell in row)
            if previous_time is not None and numbers[0] < previous_time:
                monotonic = False
            previous_time = numbers[0]
            update_numeric_digest(value, numbers)
            points += 1
        if points == 0:
            raise AuditBlocked(f"原始历史无数值点：{member}")
        return points, value.hexdigest(), monotonic
    finally:
        workbook.close()


def parse_fatigue_csv(data: bytes, member: str) -> tuple[int, str, bool]:
    text = io.TextIOWrapper(io.BytesIO(data), encoding="utf-8-sig", newline="")
    reader = csv.reader(text)
    value = hashlib.sha256()
    points = 0
    previous_time: float | None = None
    monotonic = True
    for index, row in enumerate(reader, start=1):
        if index <= 3:
            continue
        if not row or not any(cell.strip() for cell in row[:4]):
            continue
        if len(row) < 4:
            raise AuditBlocked(f"CSV 原始历史列数不足：{member}:{index}")
        try:
            numbers = tuple(float(cell.strip()) for cell in row[:4])
        except ValueError as exc:
            raise AuditBlocked(f"CSV 原始历史含非数值四通道行：{member}:{index}") from exc
        if not all(math.isfinite(number) for number in numbers):
            raise AuditBlocked(f"CSV 原始历史含非有限值：{member}:{index}")
        if previous_time is not None and numbers[0] < previous_time:
            monotonic = False
        previous_time = numbers[0]
        update_numeric_digest(value, numbers)
        points += 1
    if points == 0:
        raise AuditBlocked(f"CSV 原始历史无数值点：{member}")
    return points, value.hexdigest(), monotonic


def parse_metadata_anomalies(
    archive: zipfile.ZipFile,
    main_member: str,
    raw_histories: list[dict[str, object]],
) -> dict[str, object]:
    workbook = open_workbook_member(archive, main_member)
    try:
        experiment_sheets = workbook.sheetnames[:4]
        metadata_rows: list[dict[str, str]] = []
        for sheet_name in experiment_sheets:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=7, values_only=True):
                short_id, long_id, material = row[:3]
                test_date = row[6]
                if not isinstance(short_id, str) or not isinstance(long_id, str):
                    continue
                metadata_rows.append(
                    {
                        "sheet": sheet_name,
                        "short_id": short_id.strip(),
                        "long_id": long_id.strip(),
                        "material": str(material).strip() if material is not None else "",
                        "test_date": (
                            test_date.isoformat()
                            if hasattr(test_date, "isoformat")
                            else str(test_date).strip() if test_date is not None else ""
                        ),
                    }
                )

        material_sheet = workbook["Material"]
        material_labels = [
            str(row[0]).strip()
            for row in material_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
            if row[0] is not None
        ]
        raw_material_labels = sorted({str(item["材料"]) for item in raw_histories})
        material_grade_conflicts = [
            label
            for label in material_labels
            if label.endswith("1195D")
            and any(raw.endswith("1195A") for raw in raw_material_labels)
        ]

        identity_conflicts = [
            row
            for row in metadata_rows
            if re.match(r"^1194A", row["short_id"])
            and re.match(r"^1195A", row["long_id"])
        ]
        recovery_alias_rows = [
            row
            for row in metadata_rows
            if row["short_id"].endswith("_2") and "recovery" in row["long_id"].lower()
        ]

        non_recovery = [
            item for item in raw_histories if item["数据角色"] != "恢复曲线"
        ]
        raw_ids = {str(item["试样或家族组"]) for item in non_recovery}
        metadata_base_ids: set[str] = set()
        for row in metadata_rows:
            short_id = re.sub(r"_2$", "", row["short_id"])
            # 身份冲突行以 Long ID 中可由实际文件复核的 1195A24 为归一化键，
            # 但原始 1194A24 拼写仍单独保留在异常清单中。
            long_match = re.match(r"^(1195A\d+)_", row["long_id"])
            if short_id.startswith("1194A") and long_match:
                short_id = long_match.group(1)
            if re.match(r"^(?:1154D|1164D|1174D|1195A|Texin)\d+$", short_id):
                metadata_base_ids.add(short_id)
        metadata_without_raw = sorted(metadata_base_ids - raw_ids)

        by_short: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
        for row in metadata_rows:
            by_short[row["short_id"]].append(row)
        long_id_filename_mismatches: list[dict[str, str]] = []
        for item in non_recovery:
            short_id = str(item["试样或家族组"])
            stem = str(item["条件"])
            candidates = by_short.get(short_id, [])
            if not candidates:
                candidates = [row for row in metadata_rows if row["long_id"].startswith(short_id + "_")]
            if len(candidates) != 1:
                continue
            long_id = candidates[0]["long_id"]
            if stem != long_id:
                long_id_filename_mismatches.append(
                    {"short_id": short_id, "filename_stem": stem, "metadata_long_id": long_id}
                )

        tmin_abbreviations = sorted(
            str(item["条件"])
            for item in non_recovery
            if re.search(r"_Tmin_", str(item["条件"]))
        )
        texin_year_conflicts = sorted(
            (
                {
                    "short_id": row["short_id"],
                    "metadata_long_id": row["long_id"],
                    "metadata_test_date": row["test_date"],
                }
                for row in metadata_rows
                if row["short_id"].startswith("Texin")
                and "_2023_" in row["long_id"]
                and row["test_date"].startswith("2024-")
            ),
            key=lambda item: item["short_id"],
        )

        if material_grade_conflicts != ["Elastollan 1195D"]:
            raise AuditBlocked(f"1195A/1195D 材料标签异常形态漂移：{material_grade_conflicts}")
        if len(identity_conflicts) != 1:
            raise AuditBlocked(f"1194A24/1195A24 身份异常数量漂移：{len(identity_conflicts)}")
        if len(recovery_alias_rows) != 6:
            raise AuditBlocked(f"恢复曲线元数据别名数量漂移：{len(recovery_alias_rows)}")
        if metadata_without_raw != ["1154D17", "1174D1", "1174D2"]:
            raise AuditBlocked(f"元数据有记录但归档无原始曲线的试样漂移：{metadata_without_raw}")
        if len(tmin_abbreviations) != 2:
            raise AuditBlocked(f"Tmin 文件名缩写数量漂移：{len(tmin_abbreviations)}")
        if len(texin_year_conflicts) != 6:
            raise AuditBlocked(f"Texin 日期年份冲突数量漂移：{len(texin_year_conflicts)}")

        return {
            "元数据实验记录数": len(metadata_rows),
            "Material工作表材料标签": material_labels,
            "材料牌号冲突": material_grade_conflicts,
            "Short_ID与Long_ID身份冲突": identity_conflicts,
            "恢复曲线Short_ID别名行": recovery_alias_rows,
            "元数据有记录但无归档原始历史": metadata_without_raw,
            "Long_ID与文件名差异": sorted(
                long_id_filename_mismatches, key=lambda item: item["short_id"]
            ),
            "Long_ID与文件名差异数": len(long_id_filename_mismatches),
            "Tmin省略20的文件名": tmin_abbreviations,
            "Texin年份冲突": texin_year_conflicts,
        }
    finally:
        workbook.close()


def audit_fatigue(archive_path: Path) -> tuple[dict[str, object], list[dict[str, object]]]:
    material_map = {
        "Elastollan1154D": "Elastollan 1154D",
        "Elastollan1164D": "Elastollan 1164D",
        "Elastollan1174D": "Elastollan 1174D",
        "Elastollan1195A": "Elastollan 1195A",
        "Texin245": "Texin 245",
    }
    curves: list[dict[str, object]] = []
    with zipfile.ZipFile(archive_path) as archive:
        names = archive.namelist()
        main_members = [name for name in names if name.endswith("Final_data_all_experiments.xlsx")]
        if len(main_members) != 1:
            raise AuditBlocked(f"主实验元数据工作簿数量异常：{len(main_members)}")
        target_members: list[tuple[str, str]] = []
        for name in names:
            path = PurePosixPath(name)
            if len(path.parts) < 4 or "/Load_frame_measurements/" not in name:
                continue
            material_dir = path.parts[-2]
            if material_dir in material_map and path.suffix.lower() in {".xlsx", ".csv"}:
                target_members.append((name, material_dir))
        target_members.sort(key=lambda item: item[0].casefold())
        for member, material_dir in target_members:
            data = archive.read(member)
            if member.lower().endswith(".xlsx"):
                points, curve_sha256, monotonic = parse_fatigue_xlsx(data, member)
            else:
                points, curve_sha256, monotonic = parse_fatigue_csv(data, member)
            stem = PurePosixPath(member).stem
            match = re.match(r"^(1154D\d+|1164D\d+|1174D\d+|1195A\d+|Texin\d+)", stem)
            if not match:
                raise AuditBlocked(f"无法从曲线文件名解析试样 ID：{member}")
            short_id = match.group(1)
            lower = stem.lower()
            role = "恢复曲线" if "recovery" in lower else ("离轴曲线" if "offaxis" in lower else "轴向曲线")
            curves.append(
                {
                    "来源": FATIGUE,
                    "材料": material_map[material_dir],
                    "试验类型": "压缩温度疲劳多工况",
                    "条件": stem,
                    "数据角色": role,
                    "曲线ID": f"{short_id}:{role}:{PurePosixPath(member).suffix.lower()}",
                    "点数": points,
                    "曲线SHA256": curve_sha256,
                    "试样或家族组": short_id,
                    "training_split": "false",
                    "weight": "false",
                    "备注": "Time/Force/Stroke/Epsilon四通道；时间单调" if monotonic else "时间不单调",
                }
            )
        santoprene_histories = [
            name
            for name in names
            if "/Load_frame_measurements/Santoprene" in name
            and PurePosixPath(name).suffix.lower() in {".xlsx", ".csv"}
        ]
        anomalies = parse_metadata_anomalies(archive, main_members[0], curves)
        main_sha256 = hashlib.sha256(archive.read(main_members[0])).hexdigest()

    by_material: dict[str, dict[str, int]] = {}
    for material in sorted({str(item["材料"]) for item in curves}):
        selected = [item for item in curves if item["材料"] == material]
        by_material[material] = {
            "历史数": len(selected),
            "独立试样数": len({str(item["试样或家族组"]) for item in selected}),
            "完整点数": sum(int(item["点数"]) for item in selected),
        }
    role_counts = Counter(str(item["数据角色"]) for item in curves)
    duplicate_short_ids = sorted(
        short_id
        for short_id, count in Counter(str(item["试样或家族组"]) for item in curves).items()
        if count > 1
    )
    expected_materials = {
        "Elastollan 1154D": {"历史数": 44, "独立试样数": 42, "完整点数": 71_440},
        "Elastollan 1164D": {"历史数": 35, "独立试样数": 35, "完整点数": 53_732},
        "Elastollan 1174D": {"历史数": 33, "独立试样数": 33, "完整点数": 64_028},
        "Elastollan 1195A": {"历史数": 42, "独立试样数": 40, "完整点数": 72_281},
        "Texin 245": {"历史数": 42, "独立试样数": 40, "完整点数": 72_011},
    }
    if by_material != expected_materials:
        raise AuditBlocked(f"商业 TPU 材料分组复算漂移：{by_material}")
    if len(curves) != 196 or sum(int(item["点数"]) for item in curves) != 333_492:
        raise AuditBlocked("商业 TPU 历史数或完整点数漂移")
    if len({str(item["试样或家族组"]) for item in curves}) != 190:
        raise AuditBlocked("商业 TPU 独立试样数漂移")
    if role_counts != Counter({"轴向曲线": 169, "离轴曲线": 21, "恢复曲线": 6}):
        raise AuditBlocked(f"商业 TPU 曲线角色计数漂移：{role_counts}")
    if duplicate_short_ids != ["1154D13", "1154D14", "1195A11", "1195A12", "Texin11", "Texin12"]:
        raise AuditBlocked(f"恢复曲线复用试样 ID 集合漂移：{duplicate_short_ids}")
    if any("时间不单调" in str(item["备注"]) for item in curves):
        raise AuditBlocked("商业 TPU 原始历史出现时间非单调")
    if len(santoprene_histories) != 99:
        raise AuditBlocked(f"Santoprene TPV 类比历史数量漂移：{len(santoprene_histories)}")

    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": FATIGUE,
        "主元数据工作簿": {
            "成员": main_members[0],
            "成员SHA256": main_sha256,
        },
        "TPU历史数": len(curves),
        "TPU独立物理试样数": len({str(item["试样或家族组"]) for item in curves}),
        "TPU完整四通道点数": sum(int(item["点数"]) for item in curves),
        "曲线角色计数": dict(sorted(role_counts.items())),
        "按材料复算": by_material,
        "恢复曲线复用原试样ID": duplicate_short_ids,
        "Santoprene_TPV类比历史数": len(santoprene_histories),
        "Santoprene处理": "保留作类比与外部验证，不计入TPU主训练语义",
        "数据异常复算": anomalies,
        "训练状态": {"training_split": False, "weight": False},
    }
    curves.sort(key=lambda item: (str(item["材料"]), str(item["条件"])))
    return summary, curves


def parse_pair_sheet(
    worksheet,
    *,
    x_column: int,
    y_column: int,
    x_scale: float = 1.0,
) -> tuple[int, str]:
    first_column = min(x_column, y_column)
    last_column = max(x_column, y_column)
    x_offset = x_column - first_column
    y_offset = y_column - first_column
    value = hashlib.sha256()
    points = 0
    for row in worksheet.iter_rows(
        min_row=4,
        min_col=first_column,
        max_col=last_column,
        values_only=True,
    ):
        x_cell = row[x_offset]
        y_cell = row[y_offset]
        if x_cell is None and y_cell is None:
            continue
        if not is_finite_number(x_cell) or not is_finite_number(y_cell):
            raise AuditBlocked(
                f"工作表含不完整或非数值曲线行：{worksheet.title}"
            )
        update_numeric_digest(value, (float(x_cell) * x_scale, float(y_cell)))
        points += 1
    if points == 0:
        raise AuditBlocked(f"工作表无曲线点：{worksheet.title}")
    return points, value.hexdigest()


def validate_fdm_curve_lineage(
    worksheet,
    cached_worksheet,
    kind: str,
    expected_points: int,
    *,
    cached_results: dict[int, tuple[object, object]] | None = None,
) -> str:
    """复核标准化曲线、公式、缓存值与原始力/位移之间的血缘。"""
    if cached_results is not None and not isinstance(cached_results, dict):
        cached_results = {
            row_number: tuple(row)
            for row_number, row in enumerate(
                cached_results.iter_rows(
                    min_row=1, min_col=9, max_col=10, values_only=True
                ),
                start=1,
            )
        }
    header_rows = list(
        worksheet.iter_rows(
            min_row=2, max_row=3, min_col=1, max_col=5, values_only=True
        )
    )
    if len(header_rows) != 2:
        raise AuditBlocked(f"FDM曲线列头行数漂移：{worksheet.title}")
    header, units = (tuple(row) for row in header_rows)
    count = 0
    missing_raw_channels: Counter[str] = Counter()
    formula_rows = worksheet.iter_rows(
        min_row=4, min_col=1, max_col=5, values_only=True
    )
    cached_rows = cached_worksheet.iter_rows(
        min_row=4, min_col=1, max_col=5, values_only=True
    )
    for row_number, paired_rows in enumerate(
        zip_longest(formula_rows, cached_rows, fillvalue=None),
        start=4,
    ):
        formula_row, cached_row = paired_rows
        if formula_row is None or cached_row is None:
            raise AuditBlocked(f"FDM公式/缓存工作表行数不一致：{worksheet.title}")
        values = list(formula_row)
        cached_values = list(cached_row)
        if all(value is None for value in values):
            continue
        # 这些工作簿把原始位移/力与标准化应力-应变并排保存，部分试样的
        # 原始通道长于标准化曲线。血缘复核只应逐点约束实际纳入的标准曲线，
        # 同时显式记录原始通道的额外尾段，不能把它误报成“标准曲线缺失”。
        target_indexes = (2, 3) if kind == "晶格弯曲" else (0, 1)
        target_values = [values[index] for index in target_indexes]
        if all(value is None for value in target_values):
            missing_raw_channels["raw_only_tail_rows"] += 1
            continue
        if any(value is None for value in target_values):
            raise AuditBlocked(
                f"FDM标准曲线成对列不完整：{worksheet.title}:{row_number}"
            )
        if kind == "晶格弯曲":
            if not all(is_finite_number(values[index]) for index in (1, 2)):
                raise AuditBlocked(f"FDM晶格弯曲曲线源列不完整：{worksheet.title}:{row_number}")
            if not is_finite_number(values[0]):
                missing_raw_channels["deformation_mm_A"] += 1
            stress_formula = str(values[3]).replace(" ", "")
            strain_formula = str(values[4]).replace(" ", "")
            stress_match = re.fullmatch(
                rf"=B{row_number}/\(Results!I\$(\d+)\*Results!J\$(\d+)\)",
                stress_formula,
            )
            if stress_match is None or strain_formula != f"=C{row_number}/100":
                raise AuditBlocked(
                    f"FDM晶格弯曲公式血缘漂移：{worksheet.title}:{row_number}"
                )
            if cached_results is None:
                raise AuditBlocked("FDM晶格弯曲缺少Results缓存表用于公式复算")
            area_row_i, area_row_j = (int(value) for value in stress_match.groups())
            area_i = cached_results.get(area_row_i, (None, None))[0]
            area_j = cached_results.get(area_row_j, (None, None))[1]
            if not all(
                is_finite_number(value)
                for value in (area_i, area_j, cached_values[1], cached_values[2], cached_values[3], cached_values[4])
            ):
                raise AuditBlocked(f"FDM晶格弯曲公式缓存值不完整：{worksheet.title}:{row_number}")
            expected_stress = float(cached_values[1]) / (float(area_i) * float(area_j))
            expected_strain = float(cached_values[2]) / 100.0
            if not math.isclose(float(cached_values[3]), expected_stress, rel_tol=1e-10, abs_tol=1e-12):
                raise AuditBlocked(f"FDM晶格弯曲应力缓存与公式不符：{worksheet.title}:{row_number}")
            if not math.isclose(float(cached_values[4]), expected_strain, rel_tol=1e-10, abs_tol=1e-12):
                raise AuditBlocked(f"FDM晶格弯曲应变缓存与公式不符：{worksheet.title}:{row_number}")
        elif kind == "晶格压缩":
            if not all(is_finite_number(values[index]) for index in (0, 1)):
                raise AuditBlocked(f"FDM晶格压缩标准曲线列不完整：{worksheet.title}:{row_number}")
            if not is_finite_number(values[2]):
                missing_raw_channels["deformation_mm_C"] += 1
            if not is_finite_number(values[3]):
                missing_raw_channels["force_N_D"] += 1
        elif kind == "基材弯曲":
            if not all(is_finite_number(values[index]) for index in (0, 1)):
                raise AuditBlocked(f"FDM基材弯曲标准曲线列不完整：{worksheet.title}:{row_number}")
            if str(values[2]).replace(" ", "") != f"=A{row_number}/100":
                raise AuditBlocked(f"FDM基材弯曲应变公式漂移：{worksheet.title}:{row_number}")
            if not is_finite_number(cached_values[2]) or not math.isclose(
                float(cached_values[2]), float(cached_values[0]) / 100.0,
                rel_tol=1e-10, abs_tol=1e-12,
            ):
                raise AuditBlocked(f"FDM基材弯曲应变缓存与公式不符：{worksheet.title}:{row_number}")
            if not is_finite_number(values[3]):
                missing_raw_channels["deformation_mm_D"] += 1
            if not is_finite_number(values[4]):
                missing_raw_channels["force_N_E"] += 1
        else:
            if not all(is_finite_number(values[index]) for index in (0, 1)):
                raise AuditBlocked(f"FDM基材拉伸标准曲线列不完整：{worksheet.title}:{row_number}")
            if not is_finite_number(values[2]):
                missing_raw_channels["force_N_C"] += 1
            if header[4] is not None:
                if str(values[3]).replace(" ", "") != f"=A{row_number}/100":
                    raise AuditBlocked(f"FDM基材拉伸应变公式漂移：{worksheet.title}:{row_number}")
                if not is_finite_number(cached_values[3]) or not math.isclose(
                    float(cached_values[3]), float(cached_values[0]) / 100.0,
                    rel_tol=1e-10, abs_tol=1e-12,
                ):
                    raise AuditBlocked(f"FDM基材拉伸应变缓存与公式不符：{worksheet.title}:{row_number}")
                if not is_finite_number(values[4]):
                    missing_raw_channels["deformation_mm_E"] += 1
            elif not is_finite_number(values[3]):
                missing_raw_channels["deformation_mm_D"] += 1
        count += 1
    if count != expected_points:
        raise AuditBlocked(
            f"FDM公式/原始血缘行数与缓存曲线不符：{worksheet.title}={count}/{expected_points}"
        )

    if kind == "晶格弯曲":
        if header != (
            "Deformation", "Standard force", "Deformation", "Standard force", "Deformation"
        ) or units != ("mm", "N", "%", "MPa", "mm/mm"):
            raise AuditBlocked(f"FDM晶格弯曲列头/单位漂移：{worksheet.title}")
        description = "D=B/(Results!I*Results!J)，E=C/100；曲线采用C/100与D，B/C原始力和百分应变保留"
        return description + (f"；原始通道缺失计数={dict(missing_raw_channels)}" if missing_raw_channels else "")
    if kind == "晶格压缩":
        if header != ("Deformation", "Standard force", "Deformation", "Standard force", None) or units != ("%", "MPa", "mm", "N", None):
            raise AuditBlocked(f"FDM晶格压缩列头/单位漂移：{worksheet.title}")
        description = "曲线采用A/100与B；C/D原始位移和力保留"
        return description + (f"；原始通道缺失计数={dict(missing_raw_channels)}" if missing_raw_channels else "")
    if kind == "基材弯曲":
        if (
            header[0:2] != ("Deformación", "Fuerza estándar")
            or header[2] not in {"Deformación", "Fuerza estándar"}
            or header[3:5] != ("Deformación", "Fuerza estándar")
            or units != ("%", "MPa", "mm/mm", "mm", "N")
        ):
            raise AuditBlocked(f"FDM基材弯曲列头/单位漂移：{worksheet.title}")
        header_note = "；C列标题误写为Fuerza estándar" if header[2] == "Fuerza estándar" else ""
        description = f"曲线采用A/100与B；C=A/100，D/E原始位移和力保留{header_note}"
        return description + (f"；原始通道缺失计数={dict(missing_raw_channels)}" if missing_raw_channels else "")
    if (
        header[0:3] != ("Deformación", "Fuerza estándar", "Fuerza estándar")
        or units[0:3] != ("%", "MPa", "N")
    ):
        raise AuditBlocked(f"FDM基材拉伸核心列头/单位漂移：{worksheet.title}")
    if header[4] is None:
        if header[3] != "Deformación" or units[3:] != ("mm", None):
            raise AuditBlocked(f"FDM基材拉伸位移列头漂移：{worksheet.title}")
        description = "曲线采用A/100与B；C为原始力，D为原始位移；E列缺省"
        return description + (f"；原始通道缺失计数={dict(missing_raw_channels)}" if missing_raw_channels else "")
    if (
        header[3:5] != ("Deformación", "Deformación")
        or units[3] not in {"mm/mm", "%"}
        or units[4] != "mm"
    ):
        raise AuditBlocked(f"FDM基材拉伸派生/位移列头漂移：{worksheet.title}")
    unit_note = "；D公式=A/100但列单位误标为%" if units[3] == "%" else ""
    description = f"曲线采用A/100与B；C/E原始力和位移保留，D=A/100{unit_note}"
    return description + (f"；原始通道缺失计数={dict(missing_raw_channels)}" if missing_raw_channels else "")


def primary_fdm_workbook_kind(member: str) -> str | None:
    lower = member.lower()
    basename = PurePosixPath(member).name.lower()
    if "/tpu/" not in lower or not basename.endswith(".xlsx") or "summary" in basename:
        return None
    if "/bending/tpu/" in lower and basename.startswith("ef "):
        return "晶格弯曲"
    if "/compression/tpu/" in lower and basename.startswith("ec "):
        return "晶格压缩"
    if "/material characterization/tpu/" in lower and basename.startswith("bending "):
        return "基材弯曲"
    if "/material characterization/tpu/" in lower and basename.startswith("tensile "):
        return "基材拉伸"
    return None


EXPECTED_FDM_SELECTION_STATES: dict[str, dict[str, set[str]]] = {
    "EF Cubic TPU": {
        "selected": {"EF C1", "EF C3", "EF C4", "EF C5", "EF C6"},
        "not_selected": {"EF C2"},
        "conflict": set(),
    },
    "EF BCC TPU": {
        "selected": {"EF B1", "EF B3", "EF B4", "EF B5"},
        "not_selected": set(),
        "conflict": {"EF B2", "EF B6"},
    },
    "EF ARA TPU": {
        "selected": {"EF A1", "EF A2", "EF A3", "EF A5", "EF A6"},
        "not_selected": {"EF A4"},
        "conflict": set(),
    },
    "EC Cubic TPU MPa": {
        "selected": {"EC C1", "EC C2", "EC C3", "EC C4", "EC C5"},
        "not_selected": {"EC C6"},
        "conflict": set(),
    },
    "EC BCC TPU MPa": {
        "selected": {"EC B1", "EC B2", "EC B3", "EC B4", "EC B5"},
        "not_selected": {"EC B6"},
        "conflict": set(),
    },
    "EC ARA TPU MPa": {
        "selected": {"EC A1", "EC A3", "EC A4", "EC A5"},
        "not_selected": set(),
        "conflict": {"EC A2", "EC A6"},
    },
    "Bending TPU +45 X": {
        "selected": {f"X{index}" for index in range(1, 7)},
        "not_selected": set(),
        "conflict": set(),
    },
    "Bending TPU +45 Y": {
        "selected": {f"Y{index}" for index in range(1, 7)},
        "not_selected": set(),
        "conflict": set(),
    },
    "Bending TPU +45 Z": {
        "selected": set(),
        "not_selected": set(),
        "conflict": {f"Z{index}" for index in range(1, 7)},
    },
    "Tensile TPU +45 X": {
        "selected": {"X1", "X3", "X4", "X6", "X8"},
        "not_selected": {"X2", "X5", "X7"},
        "conflict": set(),
    },
    "Tensile TPU +45 Y": {
        "selected": {"Y1", "Y2", "Y3", "Y4", "Y7"},
        "not_selected": {"Y5", "Y6"},
        "conflict": set(),
    },
    "Tensile TPU +45 Z": {
        "selected": {f"Z{index}" for index in range(1, 8)},
        "not_selected": set(),
        "conflict": set(),
    },
}


def quality_gate_for_summary_state(state: str) -> str:
    gates = {
        "selected": "pass_source_summary_selected",
        "not_selected": "hold_source_summary_not_selected",
        "conflict": "hold_source_summary_conflict",
    }
    if state not in gates:
        raise AuditBlocked(f"未知来源汇总状态：{state}")
    return gates[state]


def validate_expected_selection(
    workbook_stem: str, selection: dict[str, dict[str, str]]
) -> None:
    expected = EXPECTED_FDM_SELECTION_STATES.get(workbook_stem)
    if expected is None:
        raise AuditBlocked(f"FDM 工作簿未登记来源内选择预期：{workbook_stem}")
    actual = {
        state: {
            specimen
            for specimen, information in selection.items()
            if information["source_summary_state"] == state
        }
        for state in ("selected", "not_selected", "conflict")
    }
    if actual != expected:
        raise AuditBlocked(
            f"FDM 来源内选择状态漂移：{workbook_stem}/"
            f"actual={actual}/expected={expected}"
        )


def workbook_value_matrix(worksheet, *, max_column: int) -> list[tuple[object, ...]]:
    return [
        tuple(row)
        for row in worksheet.iter_rows(
            min_row=1, min_col=1, max_col=max_column, values_only=True
        )
    ]


def matrix_value(
    rows: list[tuple[object, ...]], row_number: int, column_number: int
) -> object:
    if row_number < 1 or row_number > len(rows):
        return None
    row = rows[row_number - 1]
    if column_number < 1 or column_number > len(row):
        return None
    return row[column_number - 1]


def finite_float(value: object, context: str) -> float:
    if not is_finite_number(value):
        raise AuditBlocked(f"FDM 标量不是有限数：{context}={value!r}")
    return float(value)


def require_formula_close(actual: object, expected: float, context: str) -> float:
    result = finite_float(actual, context)
    if not math.isclose(result, expected, rel_tol=1e-9, abs_tol=1e-12):
        raise AuditBlocked(
            f"FDM 公式缓存值与逐公式复算不符：{context}={result}/{expected}"
        )
    return result


def fdm_specimen_group(kind: str, workbook_stem: str, specimen_id: str) -> str:
    return f"{kind}|{workbook_stem}|{specimen_id}"


def make_scalar_row(
    *,
    workbook_stem: str,
    specimen_group: str,
    specimen_id: str,
    observable: str,
    value: object,
    unit: object,
    definition_id: str,
    lineage_class: str,
    formula_evidence: str,
    note: str,
) -> dict[str, object]:
    return {
        "来源": FDM,
        "工作簿": workbook_stem,
        "试样组": specimen_group,
        "试样ID": specimen_id,
        "observable": observable,
        "value": value,
        "unit": "" if unit is None else str(unit).replace("\xa0", "").strip(),
        "definition_id": definition_id,
        "scalar_lineage_class": lineage_class,
        "source_summary_state": "",
        "source_summary_evidence": "",
        "source_display_id": "",
        "formula_target_specimen_id": "",
        "quality_gate": "",
        "formula_evidence": formula_evidence,
        "training_split": "false",
        "weight": "false",
        "备注": note,
    }


def cached_sheet_values(
    workbook,
    cache: dict[tuple[str, int], dict[int, object]],
    sheet_name: str,
    column_number: int,
    row_numbers: set[int],
) -> dict[int, object]:
    if sheet_name not in workbook.sheetnames:
        raise AuditBlocked(f"FDM 公式目标工作表缺失：{sheet_name}")
    key = (sheet_name, column_number)
    cached = cache.setdefault(key, {})
    missing = row_numbers - cached.keys()
    if missing:
        max_row = max(missing)
        for row_number, row in enumerate(
            workbook[sheet_name].iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=column_number,
                max_col=column_number,
                values_only=True,
            ),
            start=1,
        ):
            if row_number in missing:
                cached[row_number] = row[0]
        if missing - cached.keys():
            raise AuditBlocked(
                f"FDM 公式目标单元格缺失：{sheet_name}/col={column_number}/"
                f"rows={sorted(missing-cached.keys())}"
            )
    return {row_number: cached[row_number] for row_number in row_numbers}


def extract_lattice_results_scalars(
    cached_workbook,
    formula_workbook,
    *,
    kind: str,
    workbook_stem: str,
) -> tuple[
    list[dict[str, object]],
    dict[str, str],
    set[str],
    dict[str, object],
]:
    """抽取晶格 Results 原始量与显式派生量，并返回 Summary 单元格目标映射。"""
    max_column = 20 if kind == "晶格弯曲" else 19
    cached_rows = workbook_value_matrix(cached_workbook["Results"], max_column=max_column)
    formula_rows = workbook_value_matrix(formula_workbook["Results"], max_column=max_column)
    if len(cached_rows) != len(formula_rows):
        raise AuditBlocked(f"FDM Results 公式/缓存行数不一致：{workbook_stem}")
    scalar_rows: list[dict[str, object]] = []
    result_cell_targets: dict[str, str] = {}
    specimen_ids: set[str] = set()
    sheet_cache: dict[tuple[str, int], dict[int, object]] = {}
    raw_numeric_count = 0
    formula_count = 0

    for source_row in range(3, 9):
        specimen_id = str(matrix_value(cached_rows, source_row, 1) or "")
        expected_pattern = r"EF [ABC][1-6]" if kind == "晶格弯曲" else r"EC [ABC][1-6]"
        if re.fullmatch(expected_pattern, specimen_id) is None:
            raise AuditBlocked(f"FDM Results 试样 ID 漂移：{workbook_stem}!A{source_row}")
        specimen_ids.add(specimen_id)
        specimen_group = fdm_specimen_group(kind, workbook_stem, specimen_id)
        for column_number in range(2, max_column + 1):
            cached_value = matrix_value(cached_rows, source_row, column_number)
            formula_value = matrix_value(formula_rows, source_row, column_number)
            if isinstance(formula_value, str) and formula_value.startswith("="):
                raise AuditBlocked(
                    f"FDM Results 原始字段意外变为公式：{workbook_stem}!"
                    f"{get_column_letter(column_number)}{source_row}"
                )
            if cached_value is None or not is_finite_number(cached_value):
                continue
            header = matrix_value(cached_rows, 1, column_number)
            if not isinstance(header, str) or not header.strip():
                raise AuditBlocked(
                    f"FDM Results 数值字段缺少列头：{workbook_stem}!"
                    f"{get_column_letter(column_number)}{source_row}"
                )
            scalar_rows.append(
                make_scalar_row(
                    workbook_stem=workbook_stem,
                    specimen_group=specimen_group,
                    specimen_id=specimen_id,
                    observable=f"Results::{header.strip()}",
                    value=float(cached_value),
                    unit=matrix_value(cached_rows, 2, column_number),
                    definition_id=(
                        f"fdm_{'lattice_bending' if kind == '晶格弯曲' else 'lattice_compression'}"
                        f".results.literal.{get_column_letter(column_number)}.{header.strip()}"
                    ),
                    lineage_class="source_literal_results",
                    formula_evidence=(
                        f"Results!{get_column_letter(column_number)}{source_row}=literal"
                    ),
                    note="Results 逐试样有限数原始字段；未使用全局平均",
                )
            )
            raw_numeric_count += 1

    if len(specimen_ids) != 6:
        raise AuditBlocked(f"FDM Results 独立试样数漂移：{workbook_stem}={len(specimen_ids)}")

    if kind == "晶格弯曲":
        derived_specs = (
            ("maximum_flexural_stress", "最大弯曲应力", range(14, 20), "formula_recomputed_results"),
            ("flexural_modulus", "弯曲弹性模量", range(24, 30), "formula_recomputed_selected_curve_row"),
            ("core_shear_stress", "芯层最大剪切应力", range(34, 40), "formula_recomputed_results"),
            ("face_flexural_stress", "面层弯曲应力", range(44, 50), "formula_recomputed_results"),
        )
        for definition_suffix, observable, derived_rows, lineage_class in derived_specs:
            for derived_row in derived_rows:
                formula = str(matrix_value(formula_rows, derived_row, 4) or "").strip()
                if definition_suffix == "maximum_flexural_stress":
                    match = re.fullmatch(r"=\(3\*B(\d+)\*H\1\)/\(2\*I\1\*J\1\^2\)", formula)
                elif definition_suffix == "flexural_modulus":
                    match = re.fullmatch(
                        r"=\(B(\d+)\*H\1\^3\)/\(4\*I\1\*J\1\^3\*'([^']+)'!A(\d+)\)",
                        formula,
                    )
                elif definition_suffix == "core_shear_stress":
                    match = re.fullmatch(r"=\(B(\d+)\)/\(\(J\1\+K\1\)\*I\1\)", formula)
                else:
                    match = re.fullmatch(
                        r"=\(B(\d+)\*H\1\)/\(2\*L\1\*\(J\1\+K\1\)\*I\1\)",
                        formula,
                    )
                if match is None:
                    raise AuditBlocked(
                        f"FDM 晶格弯曲标量公式漂移：{workbook_stem}!D{derived_row}={formula}"
                    )
                source_row = int(match.group(1))
                specimen_id = str(matrix_value(cached_rows, source_row, 1) or "")
                if specimen_id not in specimen_ids:
                    raise AuditBlocked(
                        f"FDM 晶格弯曲标量公式目标试样异常：{workbook_stem}!D{derived_row}"
                    )
                b = finite_float(matrix_value(cached_rows, source_row, 2), f"{workbook_stem}!B{source_row}")
                h = finite_float(matrix_value(cached_rows, source_row, 8), f"{workbook_stem}!H{source_row}")
                i = finite_float(matrix_value(cached_rows, source_row, 9), f"{workbook_stem}!I{source_row}")
                j = finite_float(matrix_value(cached_rows, source_row, 10), f"{workbook_stem}!J{source_row}")
                if definition_suffix == "maximum_flexural_stress":
                    expected_value = (3.0 * b * h) / (2.0 * i * j**2)
                    selected_row_note = ""
                elif definition_suffix == "flexural_modulus":
                    target_sheet = match.group(2)
                    target_row = int(match.group(3))
                    if target_sheet != specimen_id:
                        raise AuditBlocked(
                            f"FDM 晶格弯曲模量公式跨试样：{workbook_stem}!D{derived_row}/"
                            f"target={target_sheet!r}/specimen={specimen_id!r}"
                        )
                    deflection = finite_float(
                        cached_sheet_values(
                            cached_workbook, sheet_cache, target_sheet, 1, {target_row}
                        )[target_row],
                        f"{workbook_stem}!{target_sheet}!A{target_row}",
                    )
                    expected_value = (b * h**3) / (4.0 * i * j**3 * deflection)
                    selected_row_note = f"；显式曲线行选择={target_sheet}!A{target_row}"
                elif definition_suffix == "core_shear_stress":
                    k = finite_float(matrix_value(cached_rows, source_row, 11), f"{workbook_stem}!K{source_row}")
                    expected_value = b / ((j + k) * i)
                    selected_row_note = ""
                else:
                    k = finite_float(matrix_value(cached_rows, source_row, 11), f"{workbook_stem}!K{source_row}")
                    face_thickness = finite_float(
                        matrix_value(cached_rows, source_row, 12),
                        f"{workbook_stem}!L{source_row}",
                    )
                    expected_value = (b * h) / (2.0 * face_thickness * (j + k) * i)
                    selected_row_note = ""
                cached_value = require_formula_close(
                    matrix_value(cached_rows, derived_row, 4),
                    expected_value,
                    f"{workbook_stem}!D{derived_row}",
                )
                if str(matrix_value(cached_rows, derived_row, 5) or "").casefold() != "mpa":
                    raise AuditBlocked(f"FDM 晶格弯曲派生标量单位漂移：{workbook_stem}!E{derived_row}")
                result_cell_targets[f"D{derived_row}"] = specimen_id
                scalar_rows.append(
                    make_scalar_row(
                        workbook_stem=workbook_stem,
                        specimen_group=fdm_specimen_group(kind, workbook_stem, specimen_id),
                        specimen_id=specimen_id,
                        observable=observable,
                        value=cached_value,
                        unit="MPa",
                        definition_id=f"fdm_lattice_bending.results.formula.{definition_suffix}",
                        lineage_class=lineage_class,
                        formula_evidence=f"Results!D{derived_row}{formula}",
                        note=f"逐公式复算并与 Excel cached 数值一致{selected_row_note}",
                    )
                )
                formula_count += 1
    else:
        for derived_row in range(11, 17):
            formula = str(matrix_value(formula_rows, derived_row, 3) or "").strip()
            match = re.fullmatch(r"=H(\d+)/S\1", formula)
            if match is None:
                raise AuditBlocked(
                    f"FDM 晶格压缩应力公式漂移：{workbook_stem}!C{derived_row}={formula}"
                )
            source_row = int(match.group(1))
            specimen_id = str(matrix_value(cached_rows, source_row, 1) or "")
            expected_value = finite_float(
                matrix_value(cached_rows, source_row, 8), f"{workbook_stem}!H{source_row}"
            ) / finite_float(
                matrix_value(cached_rows, source_row, 19), f"{workbook_stem}!S{source_row}"
            )
            cached_value = require_formula_close(
                matrix_value(cached_rows, derived_row, 3), expected_value, f"{workbook_stem}!C{derived_row}"
            )
            result_cell_targets[f"C{derived_row}"] = specimen_id
            scalar_rows.append(
                make_scalar_row(
                    workbook_stem=workbook_stem,
                    specimen_group=fdm_specimen_group(kind, workbook_stem, specimen_id),
                    specimen_id=specimen_id,
                    observable="压缩应力",
                    value=cached_value,
                    unit="MPa",
                    definition_id="fdm_lattice_compression.results.formula.compressive_stress",
                    lineage_class="formula_recomputed_results",
                    formula_evidence=f"Results!C{derived_row}{formula}",
                    note="逐公式复算 H/S 并与 Excel cached 数值一致",
                )
            )
            formula_count += 1
        for derived_row in range(21, 27):
            formula = str(matrix_value(formula_rows, derived_row, 3) or "").strip()
            match = re.fullmatch(
                r"=\('([^']+)'!G(\d+)\*Results!R(\d+)\)/Results!S\3", formula
            )
            if match is None:
                raise AuditBlocked(
                    f"FDM 晶格压缩模量公式漂移：{workbook_stem}!C{derived_row}={formula}"
                )
            target_sheet = match.group(1)
            target_row = int(match.group(2))
            source_row = int(match.group(3))
            specimen_id = str(matrix_value(cached_rows, source_row, 1) or "")
            if target_sheet != specimen_id or specimen_id not in specimen_ids:
                raise AuditBlocked(
                    f"FDM 晶格压缩模量公式跨试样：{workbook_stem}!C{derived_row}"
                )
            slope_value = finite_float(
                cached_sheet_values(
                    cached_workbook, sheet_cache, target_sheet, 7, {target_row}
                )[target_row],
                f"{workbook_stem}!{target_sheet}!G{target_row}",
            )
            expected_value = (
                slope_value
                * finite_float(matrix_value(cached_rows, source_row, 18), f"{workbook_stem}!R{source_row}")
                / finite_float(matrix_value(cached_rows, source_row, 19), f"{workbook_stem}!S{source_row}")
            )
            cached_value = require_formula_close(
                matrix_value(cached_rows, derived_row, 3), expected_value, f"{workbook_stem}!C{derived_row}"
            )
            result_cell_targets[f"C{derived_row}"] = specimen_id
            scalar_rows.append(
                make_scalar_row(
                    workbook_stem=workbook_stem,
                    specimen_group=fdm_specimen_group(kind, workbook_stem, specimen_id),
                    specimen_id=specimen_id,
                    observable="压缩模量",
                    value=cached_value,
                    unit="MPa",
                    definition_id="fdm_lattice_compression.results.formula.compressive_modulus",
                    lineage_class="formula_recomputed_manual_row_selection",
                    formula_evidence=f"Results!C{derived_row}{formula}",
                    note=f"模量采用人工选定斜率行 {target_sheet}!G{target_row}；逐公式复算 cached 数值",
                )
            )
            formula_count += 1

    expected_formula_count = 24 if kind == "晶格弯曲" else 12
    if formula_count != expected_formula_count:
        raise AuditBlocked(
            f"FDM 晶格派生标量数量漂移：{workbook_stem}={formula_count}/{expected_formula_count}"
        )
    n10_by_specimen = {
        specimen_id: finite_float(
            cached_sheet_values(
                cached_workbook, sheet_cache, specimen_id, 14, {10}
            )[10],
            f"{workbook_stem}!{specimen_id}!N10",
        )
        for specimen_id in sorted(specimen_ids)
    }
    return (
        scalar_rows,
        result_cell_targets,
        specimen_ids,
        {
            "原始有限数标量行": raw_numeric_count,
            "显式公式派生标量行": formula_count,
            "聚合量处理": "Results 中 Prom/AVERAGE 为全局聚合，仅留在来源证据，不作为试样标量",
            "逐试样N10缓存": n10_by_specimen,
        },
    )


def extract_material_scalars_and_selection(
    cached_workbook,
    formula_workbook,
    *,
    kind: str,
    workbook_stem: str,
) -> tuple[
    list[dict[str, object]],
    dict[str, dict[str, str]],
    set[str],
    dict[str, object],
]:
    """抽取基材 Resultados literal 与 Resumen 可复算逐试样量。"""
    max_column = 11 if kind == "基材弯曲" else 13
    cached_rows = workbook_value_matrix(cached_workbook["Resultados"], max_column=max_column)
    formula_rows = workbook_value_matrix(formula_workbook["Resultados"], max_column=max_column)
    if len(cached_rows) != len(formula_rows):
        raise AuditBlocked(f"FDM Resultados 公式/缓存行数不一致：{workbook_stem}")
    scalar_rows: list[dict[str, object]] = []
    specimen_ids: set[str] = set()
    raw_numeric_count = 0
    missing_literal_count = 0
    expected_id_pattern = r"[XYZ][1-8]" if kind == "基材拉伸" else r"[XYZ][1-6]"
    for row_number in range(3, len(cached_rows) + 1):
        specimen_id = str(matrix_value(cached_rows, row_number, 1) or "")
        if re.fullmatch(expected_id_pattern, specimen_id) is None:
            raise AuditBlocked(f"FDM Resultados 试样 ID 漂移：{workbook_stem}!A{row_number}")
        if specimen_id in specimen_ids:
            raise AuditBlocked(f"FDM Resultados 试样 ID 重复：{workbook_stem}/{specimen_id}")
        specimen_ids.add(specimen_id)
        specimen_group = fdm_specimen_group(kind, workbook_stem, specimen_id)
        for column_number in range(2, max_column + 1):
            formula_value = matrix_value(formula_rows, row_number, column_number)
            if isinstance(formula_value, str) and formula_value.startswith("="):
                raise AuditBlocked(
                    f"FDM Resultados literal 字段意外变为公式：{workbook_stem}!"
                    f"{get_column_letter(column_number)}{row_number}"
                )
            cached_value = matrix_value(cached_rows, row_number, column_number)
            if cached_value is None:
                missing_literal_count += 1
                continue
            if not is_finite_number(cached_value):
                raise AuditBlocked(
                    f"FDM Resultados 逐试样字段非有限数：{workbook_stem}!"
                    f"{get_column_letter(column_number)}{row_number}={cached_value!r}"
                )
            header = matrix_value(cached_rows, 1, column_number)
            if not isinstance(header, str) or not header.strip():
                raise AuditBlocked(
                    f"FDM Resultados 数值字段缺少列头：{workbook_stem}!"
                    f"{get_column_letter(column_number)}{row_number}"
                )
            scalar_rows.append(
                make_scalar_row(
                    workbook_stem=workbook_stem,
                    specimen_group=specimen_group,
                    specimen_id=specimen_id,
                    observable=f"Resultados::{header.strip()}",
                    value=float(cached_value),
                    unit=matrix_value(cached_rows, 2, column_number),
                    definition_id=(
                        f"fdm_{'base_bending' if kind == '基材弯曲' else 'base_tensile'}"
                        f".resultados.literal.{get_column_letter(column_number)}.{header.strip()}"
                    ),
                    lineage_class="source_literal_resultados",
                    formula_evidence=(
                        f"Resultados!{get_column_letter(column_number)}{row_number}=literal"
                    ),
                    note="Resultados 逐试样 literal；与 Resumen 近义量采用独立 definition_id",
                )
            )
            raw_numeric_count += 1

    expected_count = 6 if kind == "基材弯曲" else (
        8 if workbook_stem.endswith(" X") else 7
    )
    if len(specimen_ids) != expected_count:
        raise AuditBlocked(
            f"FDM Resultados 试样数漂移：{workbook_stem}={len(specimen_ids)}/{expected_count}"
        )
    expected_missing = 24 if kind == "基材弯曲" else 0
    if missing_literal_count != expected_missing:
        raise AuditBlocked(
            f"FDM Resultados literal 缺失单元格计数漂移："
            f"{workbook_stem}={missing_literal_count}/{expected_missing}"
        )

    resumen_cached = workbook_value_matrix(cached_workbook["Resumen"], max_column=16)
    resumen_formula = workbook_value_matrix(formula_workbook["Resumen"], max_column=16)
    if len(resumen_cached) != len(resumen_formula):
        raise AuditBlocked(f"FDM Resumen 公式/缓存行数不一致：{workbook_stem}")
    expected_headers = (
        "Designación probeta",
        "Módulo de Young (MPa)",
        "Esfuerzo (MPa)",
        "Límite elástico (MPa)",
        "Fuerza max (N)",
        "Deformación (%)",
        "Deformación (mm/mm)",
    )
    actual_headers = tuple(matrix_value(resumen_cached, 2, column) for column in range(10, 17))
    if actual_headers != expected_headers:
        raise AuditBlocked(f"FDM Resumen 六类逐试样标签漂移：{workbook_stem}/{actual_headers}")

    target_cell_rows = (7, 10, 6, 8, 9)
    target_cache: dict[tuple[str, int], dict[int, object]] = {}
    records_by_target: dict[str, dict[str, object]] = {}
    for row_number in range(3, len(resumen_cached) + 1):
        display_id = matrix_value(resumen_cached, row_number, 10)
        if not isinstance(display_id, str) or re.fullmatch(r"[XYZ][1-8]", display_id) is None:
            continue
        formulas = [str(matrix_value(resumen_formula, row_number, column) or "") for column in range(11, 16)]
        targets: list[str] = []
        for column_number, (formula, target_cell_row) in enumerate(
            zip(formulas, target_cell_rows, strict=True), start=11
        ):
            match = re.fullmatch(r"='([^']+)'!\$?N\$?(\d+)", formula.replace(" ", ""))
            if match is None or int(match.group(2)) != target_cell_row:
                raise AuditBlocked(
                    f"FDM Resumen 逐试样公式漂移：{workbook_stem}!"
                    f"{get_column_letter(column_number)}{row_number}={formula}"
                )
            targets.append(match.group(1))
        if len(set(targets)) != 1:
            raise AuditBlocked(f"FDM Resumen 单行公式跨试样：{workbook_stem}!{row_number}")
        target_id = targets[0]
        if target_id not in specimen_ids or target_id in records_by_target:
            raise AuditBlocked(
                f"FDM Resumen 公式目标缺失或重复：{workbook_stem}/{target_id}"
            )
        target_values = cached_sheet_values(
            cached_workbook,
            target_cache,
            target_id,
            14,
            set(target_cell_rows),
        )
        cached_values: list[float] = []
        for offset, target_cell_row in enumerate(target_cell_rows, start=11):
            expected_value = finite_float(
                target_values[target_cell_row],
                f"{workbook_stem}!{target_id}!N{target_cell_row}",
            )
            cached_values.append(
                require_formula_close(
                    matrix_value(resumen_cached, row_number, offset),
                    expected_value,
                    f"{workbook_stem}!Resumen!{get_column_letter(offset)}{row_number}",
                )
            )
        ratio_formula = str(matrix_value(resumen_formula, row_number, 16) or "").replace(" ", "")
        if ratio_formula != f"=O{row_number}/100":
            raise AuditBlocked(
                f"FDM Resumen 应变比公式漂移：{workbook_stem}!P{row_number}={ratio_formula}"
            )
        ratio_value = require_formula_close(
            matrix_value(resumen_cached, row_number, 16),
            cached_values[4] / 100.0,
            f"{workbook_stem}!Resumen!P{row_number}",
        )
        records_by_target[target_id] = {
            "row": row_number,
            "display_id": display_id,
            "target_id": target_id,
            "formulas": formulas + [ratio_formula],
            "values": cached_values + [ratio_value],
        }

    selection: dict[str, dict[str, str]] = {}
    for specimen_id in sorted(specimen_ids):
        record = records_by_target.get(specimen_id)
        if record is None:
            state = "not_selected"
            evidence = (
                f"{workbook_stem}.xlsx|Resumen 无公式指向 {specimen_id}；"
                "来源汇总主动未选用（非异常）"
            )
            display_id = ""
            formula_target = ""
        else:
            state = "selected" if record["display_id"] == specimen_id else "conflict"
            row_number = int(record["row"])
            evidence = (
                f"{workbook_stem}.xlsx|Resumen!J{row_number}={record['display_id']};"
                + ";".join(
                    f"{get_column_letter(column)}{row_number}={formula}"
                    for column, formula in zip(
                        range(11, 17), record["formulas"], strict=True
                    )
                )
            )
            display_id = str(record["display_id"])
            formula_target = specimen_id
        selection[specimen_id] = {
            "source_summary_state": state,
            "source_summary_evidence": evidence,
            "source_display_id": display_id,
            "formula_target_specimen_id": formula_target,
            "quality_gate": quality_gate_for_summary_state(state),
        }
    validate_expected_selection(workbook_stem, selection)

    for row in scalar_rows:
        information = selection[str(row["试样ID"])]
        for field in (
            "source_summary_state",
            "source_summary_evidence",
            "source_display_id",
            "formula_target_specimen_id",
            "quality_gate",
        ):
            row[field] = information[field]

    resumen_definitions = (
        ("young_modulus", "Resumen::Módulo de Young", "MPa", "internal_formula_link"),
        ("maximum_stress", "Resumen::Esfuerzo máximo", "MPa", "internal_formula_link"),
        ("yield_limit", "Resumen::Límite elástico", "MPa", "internal_formula_link"),
        ("maximum_force", "Resumen::Fuerza máxima", "N", "internal_formula_link"),
        ("deformation_percent", "Resumen::Deformación", "%", "internal_formula_link"),
        ("deformation_ratio", "Resumen::Deformación", "mm/mm", "percent_to_ratio_formula"),
    )
    resumen_scalar_count = 0
    blocked_conflict_scalar_count = 0
    for specimen_id, record in sorted(records_by_target.items()):
        information = selection[specimen_id]
        row_number = int(record["row"])
        for offset, (definition_suffix, observable, unit, lineage) in enumerate(
            resumen_definitions
        ):
            conflict = information["source_summary_state"] == "conflict"
            value: object = "" if conflict else record["values"][offset]
            scalar = make_scalar_row(
                workbook_stem=workbook_stem,
                specimen_group=fdm_specimen_group(kind, workbook_stem, specimen_id),
                specimen_id=specimen_id,
                observable=observable,
                value=value,
                unit=unit,
                definition_id=(
                    f"fdm_{'base_bending' if kind == '基材弯曲' else 'base_tensile'}"
                    f".resumen.{definition_suffix}"
                ),
                lineage_class=(
                    "blocked_summary_conflict_evidence" if conflict else f"resumen_{lineage}"
                ),
                formula_evidence=(
                    f"Resumen!{get_column_letter(11 + offset)}{row_number}="
                    f"{record['formulas'][offset]}"
                ),
                note=(
                    "Resumen 显示 ID 与公式目标冲突，仅保留公式证据，不发布数值"
                    if conflict
                    else "Resumen 逐试样公式已重算；与 Resultados 近义量 definition_id 分开"
                ),
            )
            for field in (
                "source_summary_state",
                "source_summary_evidence",
                "source_display_id",
                "formula_target_specimen_id",
                "quality_gate",
            ):
                scalar[field] = information[field]
            scalar_rows.append(scalar)
            resumen_scalar_count += 1
            blocked_conflict_scalar_count += int(conflict)

    aggregate_start_rows = [
        row_number
        for row_number in range(1, len(resumen_cached) + 1)
        if matrix_value(resumen_cached, row_number, 10) == "Resultados"
    ]
    if len(aggregate_start_rows) != 1:
        raise AuditBlocked(f"FDM Resumen 聚合标签块数量漂移：{workbook_stem}")
    aggregate_start = aggregate_start_rows[0]
    aggregate_labels: list[dict[str, object]] = []
    for row_number in range(aggregate_start + 1, aggregate_start + 7):
        label = matrix_value(resumen_cached, row_number, 10)
        formula = matrix_value(resumen_formula, row_number, 12)
        unit = matrix_value(resumen_cached, row_number, 13)
        if not isinstance(label, str) or not isinstance(formula, str) or not formula.startswith("="):
            raise AuditBlocked(f"FDM Resumen 六类聚合标签或公式漂移：{workbook_stem}!{row_number}")
        aggregate_labels.append(
            {
                "标签": label.strip(),
                "单元格": f"Resumen!L{row_number}",
                "公式": formula,
                "单位": unit,
                "处理": "全局聚合，仅记录证据，不作为试样标量",
            }
        )

    return (
        scalar_rows,
        selection,
        specimen_ids,
        {
            "Resultados_literal标量行": raw_numeric_count,
            "Resultados_literal缺失字段证据数": missing_literal_count,
            "Resumen逐试样公式证据行": resumen_scalar_count,
            "Resumen冲突仅证据空值行": blocked_conflict_scalar_count,
            "Resumen全局聚合六类": aggregate_labels,
        },
    )


def extract_lattice_summary_selection(
    archive: zipfile.ZipFile,
    *,
    summary_member: str,
    kind: str,
    result_targets: dict[str, dict[str, str]],
    result_values: dict[str, dict[str, float]],
    specimen_ids_by_stem: dict[str, set[str]],
    n10_by_stem: dict[str, dict[str, float]],
) -> tuple[dict[str, dict[str, dict[str, str]]], dict[str, object]]:
    """用 Summary 的 Datos 与 C. Chauvenet 双证据解析源内选择状态。"""
    data = archive.read(summary_member)
    cached_workbook = load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False
    )
    formula_workbook = load_workbook(
        io.BytesIO(data), read_only=True, data_only=False, keep_links=False
    )
    try:
        datos_cached = workbook_value_matrix(cached_workbook["Datos"], max_column=7)
        datos_formula = workbook_value_matrix(formula_workbook["Datos"], max_column=7)
        chauvenet_cached = workbook_value_matrix(cached_workbook["C. Chauvenet"], max_column=15)
        chauvenet_formula = workbook_value_matrix(formula_workbook["C. Chauvenet"], max_column=15)
        if kind == "晶格弯曲":
            summary_stems = {
                1: ("EF Cubic TPU", "C"),
                2: ("EF BCC TPU", "B"),
                3: ("EF ARA TPU", "A"),
            }
            datos_blocks = (range(4, 9), range(15, 20), range(26, 31), range(37, 42))
            expected_result_column = "D"
            expected_test_prefix = "EF"
        else:
            summary_stems = {
                1: ("EC Cubic TPU MPa", "C"),
                2: ("EC BCC TPU MPa", "B"),
                3: ("EC ARA TPU MPa", "A"),
            }
            datos_blocks = (range(4, 9), range(15, 20))
            expected_result_column = "C"
            expected_test_prefix = "EC"

        datos_by_stem: dict[str, dict[str, dict[str, object]]] = {}
        datos_evidence: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
        for external_index, (workbook_stem, geometry) in summary_stems.items():
            column_number = external_index + 4
            reference_mapping: dict[str, dict[str, object]] | None = None
            for block in datos_blocks:
                block_mapping: dict[str, dict[str, object]] = {}
                for row_number in block:
                    display_number = matrix_value(datos_formula, row_number, 4)
                    if display_number not in {1, 2, 3, 4, 5}:
                        raise AuditBlocked(
                            f"FDM Summary Datos 显示编号漂移：{summary_member}!D{row_number}"
                        )
                    formula = str(matrix_value(datos_formula, row_number, column_number) or "").replace(" ", "")
                    match = re.fullmatch(
                        r"=\+\[(\d+)\]Results!\$?([A-Z]+)\$?(\d+)", formula
                    )
                    if (
                        match is None
                        or int(match.group(1)) != external_index
                        or match.group(2) != expected_result_column
                    ):
                        raise AuditBlocked(
                            f"FDM Summary Datos 公式漂移：{summary_member}!"
                            f"{get_column_letter(column_number)}{row_number}={formula}"
                        )
                    target_cell = f"{match.group(2)}{match.group(3)}"
                    target_id = result_targets[workbook_stem].get(target_cell)
                    target_value = result_values[workbook_stem].get(target_cell)
                    if target_id is None or target_value is None:
                        raise AuditBlocked(
                            f"FDM Summary Datos 无法解析主工作簿公式目标："
                            f"{workbook_stem}/{target_cell}"
                        )
                    require_formula_close(
                        matrix_value(datos_cached, row_number, column_number),
                        target_value,
                        f"{PurePosixPath(summary_member).name}!Datos!"
                        f"{get_column_letter(column_number)}{row_number}",
                    )
                    display_id = f"{expected_test_prefix} {geometry}{display_number}"
                    if target_id in block_mapping:
                        raise AuditBlocked(
                            f"FDM Summary Datos 单块重复指向试样：{workbook_stem}/{target_id}"
                        )
                    evidence = (
                        f"{PurePosixPath(summary_member).name}|Datos!"
                        f"{get_column_letter(column_number)}{row_number}={formula}"
                    )
                    block_mapping[target_id] = {
                        "display_id": display_id,
                        "target_id": target_id,
                    }
                    datos_evidence[workbook_stem][target_id].append(evidence)
                if reference_mapping is None:
                    reference_mapping = block_mapping
                elif reference_mapping != block_mapping:
                    raise AuditBlocked(
                        f"FDM Summary Datos 各标量块试样映射不一致：{workbook_stem}"
                    )
            if reference_mapping is None:
                raise AuditBlocked(f"FDM Summary Datos 未形成选择映射：{workbook_stem}")
            datos_by_stem[workbook_stem] = reference_mapping

        chauvenet_by_stem: dict[str, dict[str, dict[str, object]]] = {
            stem: {} for stem, _ in summary_stems.values()
        }
        chauvenet_evidence: dict[str, dict[str, str]] = defaultdict(dict)
        for external_index, (workbook_stem, geometry) in summary_stems.items():
            column_number = external_index + 2
            for row_number in range(6, 11):
                display_number = matrix_value(chauvenet_formula, row_number, 2)
                if display_number not in {1, 2, 3, 4, 5}:
                    raise AuditBlocked(
                        f"FDM C. Chauvenet 显示编号漂移：{summary_member}!B{row_number}"
                    )
                formula = str(matrix_value(chauvenet_formula, row_number, column_number) or "").strip()
                match = re.fullmatch(
                    r"^=\+'\[(\d+)\](EF|EC) ([ABC])(\d+)'!\$N\$10$", formula
                )
                if (
                    match is None
                    or int(match.group(1)) != external_index
                    or match.group(2) != expected_test_prefix
                    or match.group(3) != geometry
                ):
                    raise AuditBlocked(
                        f"FDM C. Chauvenet 公式漂移：{summary_member}!"
                        f"{get_column_letter(column_number)}{row_number}={formula}"
                    )
                target_id = f"{match.group(2)} {match.group(3)}{match.group(4)}"
                if target_id not in specimen_ids_by_stem[workbook_stem]:
                    raise AuditBlocked(
                        f"FDM C. Chauvenet 公式指向未知试样：{workbook_stem}/{target_id}"
                    )
                if target_id in chauvenet_by_stem[workbook_stem]:
                    raise AuditBlocked(
                        f"FDM C. Chauvenet 重复指向试样：{workbook_stem}/{target_id}"
                    )
                require_formula_close(
                    matrix_value(chauvenet_cached, row_number, column_number),
                    n10_by_stem[workbook_stem][target_id],
                    f"{PurePosixPath(summary_member).name}!C. Chauvenet!"
                    f"{get_column_letter(column_number)}{row_number}",
                )
                display_id = f"{expected_test_prefix} {geometry}{display_number}"
                chauvenet_by_stem[workbook_stem][target_id] = {
                    "display_id": display_id,
                    "target_id": target_id,
                }
                chauvenet_evidence[workbook_stem][target_id] = (
                    f"{PurePosixPath(summary_member).name}|C. Chauvenet!"
                    f"{get_column_letter(column_number)}{row_number}={formula}"
                )

        selection_by_stem: dict[str, dict[str, dict[str, str]]] = {}
        for _, (workbook_stem, _) in summary_stems.items():
            selection: dict[str, dict[str, str]] = {}
            datos_targets = datos_by_stem[workbook_stem]
            chauvenet_targets = chauvenet_by_stem[workbook_stem]
            for specimen_id in sorted(specimen_ids_by_stem[workbook_stem]):
                in_datos = specimen_id in datos_targets
                in_chauvenet = specimen_id in chauvenet_targets
                expected_states = EXPECTED_FDM_SELECTION_STATES[workbook_stem]
                matching_states = [
                    state
                    for state in ("selected", "not_selected", "conflict")
                    if specimen_id in expected_states[state]
                ]
                if len(matching_states) != 1:
                    raise AuditBlocked(
                        f"FDM 来源内选择预期未唯一覆盖试样：{workbook_stem}/{specimen_id}"
                    )
                state = matching_states[0]
                if state == "selected" and not (in_datos and in_chauvenet):
                    raise AuditBlocked(
                        f"FDM selected 试样未被两套汇总共同指向：{workbook_stem}/{specimen_id}"
                    )
                if state == "not_selected" and (in_datos or in_chauvenet):
                    raise AuditBlocked(
                        f"FDM not_selected 试样仍被汇总公式指向：{workbook_stem}/{specimen_id}"
                    )
                display_parts: list[str] = []
                target_parts: list[str] = []
                evidence_parts: list[str] = []
                if in_datos:
                    display_parts.append(str(datos_targets[specimen_id]["display_id"]))
                    target_parts.append(specimen_id)
                    evidence_parts.extend(datos_evidence[workbook_stem][specimen_id])
                else:
                    evidence_parts.append(
                        f"{PurePosixPath(summary_member).name}|Datos 未引用 {specimen_id}"
                    )
                if in_chauvenet:
                    display_parts.append(str(chauvenet_targets[specimen_id]["display_id"]))
                    target_parts.append(specimen_id)
                    evidence_parts.append(chauvenet_evidence[workbook_stem][specimen_id])
                else:
                    evidence_parts.append(
                        f"{PurePosixPath(summary_member).name}|C. Chauvenet 未引用 {specimen_id}"
                    )
                # 未被公式实际指向的显示编号仍是重要证据。例如 EC ARA 的
                # display A2 同时指向 formula target A6；按来源语义 A2/A6 均
                # 必须保持 conflict，不能把 A2 简化成普通缺测。
                if not target_parts:
                    for target_id, record in datos_targets.items():
                        if record["display_id"] == specimen_id:
                            display_parts.append(specimen_id)
                            target_parts.append(target_id)
                            evidence_parts.extend(datos_evidence[workbook_stem][target_id])
                            evidence_parts.append(
                                f"Datos 显示 {specimen_id} 但公式目标={target_id}"
                            )
                    for target_id, record in chauvenet_targets.items():
                        if record["display_id"] == specimen_id:
                            display_parts.append(specimen_id)
                            target_parts.append(target_id)
                            evidence_parts.append(chauvenet_evidence[workbook_stem][target_id])
                            evidence_parts.append(
                                f"C. Chauvenet 显示 {specimen_id} 但公式目标={target_id}"
                            )
                if state == "not_selected":
                    evidence_parts.append(
                        "两套来源汇总公式均未指向该实际试样；not_selected 是主动未选用（非异常）"
                    )
                unique_display = list(dict.fromkeys(display_parts))
                unique_targets = list(dict.fromkeys(target_parts))
                selection[specimen_id] = {
                    "source_summary_state": state,
                    "source_summary_evidence": ";".join(evidence_parts),
                    "source_display_id": (
                        "" if not unique_display else "|".join(unique_display)
                    ),
                    "formula_target_specimen_id": "|".join(unique_targets),
                    "quality_gate": quality_gate_for_summary_state(state),
                }
            validate_expected_selection(workbook_stem, selection)
            selection_by_stem[workbook_stem] = selection

        blocked_external: list[dict[str, str]] = []
        for worksheet in formula_workbook.worksheets:
            for row in worksheet.iter_rows(values_only=False):
                for cell in row:
                    formula = cell.value
                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue
                    if "[4]" in formula or "#REF!" in formula:
                        blocked_external.append(
                            {
                                "单元格": f"{worksheet.title}!{cell.coordinate}",
                                "公式": formula,
                                "处理": "blocked_missing_external；仅记录证据，不发布值或权重",
                            }
                        )
        if not blocked_external:
            raise AuditBlocked(f"FDM Summary 缺失外部依赖证据意外消失：{summary_member}")
        return (
            selection_by_stem,
            {
                "成员": summary_member,
                "成员SHA256": hashlib.sha256(data).hexdigest(),
                "双来源选择证据": "Datos 与 C. Chauvenet 逐公式交叉；差异记 conflict",
                "blocked_missing_external": blocked_external,
                "聚合量处理": "Average/Promedio/Desv. Est. 为全局聚合，不作为试样",
            },
        )
    finally:
        cached_workbook.close()
        formula_workbook.close()


def audit_fdm(
    archive_path: Path,
) -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    curves: list[dict[str, object]] = []
    scalar_rows: list[dict[str, object]] = []
    curve_bindings: list[tuple[dict[str, object], str, str]] = []
    primary_member_hashes: list[tuple[str, int, str]] = []
    result_targets_by_stem: dict[str, dict[str, str]] = {}
    result_values_by_stem: dict[str, dict[str, float]] = {}
    specimen_ids_by_stem: dict[str, set[str]] = {}
    n10_by_stem: dict[str, dict[str, float]] = {}
    selection_by_stem: dict[str, dict[str, dict[str, str]]] = {}
    scalar_diagnostics: dict[str, dict[str, object]] = {}
    summary_workbook_evidence: list[dict[str, object]] = []

    with zipfile.ZipFile(archive_path) as archive:
        names = archive.namelist()
        primary = [
            (name, kind)
            for name in names
            if (kind := primary_fdm_workbook_kind(name)) is not None
        ]
        primary.sort(key=lambda item: item[0].casefold())
        if len(primary) != 12:
            raise AuditBlocked(f"FDM TPU 主工作簿数量漂移：{len(primary)}")
        for member, kind in primary:
            data = archive.read(member)
            member_hash = hashlib.sha256(data).hexdigest()
            primary_member_hashes.append((member, len(data), member_hash))
            workbook_stem = PurePosixPath(member).stem
            workbook = load_workbook(
                io.BytesIO(data), read_only=True, data_only=True, keep_links=False
            )
            formula_workbook = load_workbook(
                io.BytesIO(data), read_only=True, data_only=False, keep_links=False
            )
            try:
                if kind in {"晶格弯曲", "晶格压缩"}:
                    (
                        workbook_scalars,
                        result_targets,
                        specimen_ids,
                        diagnostics,
                    ) = extract_lattice_results_scalars(
                        workbook,
                        formula_workbook,
                        kind=kind,
                        workbook_stem=workbook_stem,
                    )
                    result_targets_by_stem[workbook_stem] = result_targets
                    result_values_by_stem[workbook_stem] = {}
                    for scalar in workbook_scalars:
                        evidence = str(scalar["formula_evidence"])
                        match = re.match(r"^Results!([A-Z]+\d+)=", evidence)
                        if match is not None:
                            result_values_by_stem[workbook_stem][match.group(1)] = finite_float(
                                scalar["value"], f"{workbook_stem}/{match.group(1)}"
                            )
                    n10_by_stem[workbook_stem] = {
                        str(key): float(value)
                        for key, value in dict(diagnostics["逐试样N10缓存"]).items()
                    }
                else:
                    (
                        workbook_scalars,
                        workbook_selection,
                        specimen_ids,
                        diagnostics,
                    ) = extract_material_scalars_and_selection(
                        workbook,
                        formula_workbook,
                        kind=kind,
                        workbook_stem=workbook_stem,
                    )
                    selection_by_stem[workbook_stem] = workbook_selection
                specimen_ids_by_stem[workbook_stem] = specimen_ids
                scalar_rows.extend(workbook_scalars)
                scalar_diagnostics[workbook_stem] = diagnostics

                if kind == "晶格弯曲":
                    pattern = re.compile(r"^EF ([ABC])([1-6])$")
                    x_column, y_column, x_scale = 3, 4, 0.01
                    cached_results = {
                        row_number: tuple(row)
                        for row_number, row in enumerate(
                            workbook["Results"].iter_rows(
                                min_row=1, min_col=9, max_col=10, values_only=True
                            ),
                            start=1,
                        )
                    }
                elif kind == "晶格压缩":
                    pattern = re.compile(r"^EC ([ABC])([1-6])$")
                    x_column, y_column, x_scale = 1, 2, 0.01
                    cached_results = None
                elif kind == "基材弯曲":
                    pattern = re.compile(r"^[XYZ][1-6]$")
                    x_column, y_column, x_scale = 1, 2, 0.01
                    cached_results = None
                else:
                    pattern = re.compile(r"^(?:X[1-8]|Y[1-7]|Z[1-7])$")
                    x_column, y_column, x_scale = 1, 2, 0.01
                    cached_results = None
                specimen_sheets = [
                    sheet for sheet in workbook.worksheets if pattern.fullmatch(sheet.title)
                ]
                expected_count = 6 if kind != "基材拉伸" else (
                    8 if " +45 x.xlsx" in member.lower() else 7
                )
                if len(specimen_sheets) != expected_count:
                    raise AuditBlocked(
                        f"FDM 试样工作表数量漂移：{PurePosixPath(member).name}="
                        f"{len(specimen_sheets)}/{expected_count}"
                    )
                for worksheet in specimen_sheets:
                    points, curve_sha256 = parse_pair_sheet(
                        worksheet,
                        x_column=x_column,
                        y_column=y_column,
                        x_scale=x_scale,
                    )
                    lineage_note = validate_fdm_curve_lineage(
                        formula_workbook[worksheet.title],
                        worksheet,
                        kind,
                        points,
                        cached_results=cached_results,
                    )
                    specimen_group = fdm_specimen_group(
                        kind, workbook_stem, worksheet.title
                    )
                    curve = {
                        "来源": FDM,
                        "材料": "FDM打印TPU",
                        "试验类型": kind,
                        "条件": workbook_stem,
                        "数据角色": "实验工作簿内标准化应力应变曲线",
                        "曲线ID": f"{workbook_stem}:{worksheet.title}",
                        "点数": points,
                        "曲线SHA256": curve_sha256,
                        "试样或家族组": specimen_group,
                        "source_summary_state": "",
                        "source_summary_evidence": "",
                        "source_display_id": "",
                        "formula_target_specimen_id": "",
                        "quality_gate": "",
                        "training_split": "false",
                        "weight": "false",
                        "备注": lineage_note,
                    }
                    curves.append(curve)
                    curve_bindings.append((curve, workbook_stem, worksheet.title))
            finally:
                workbook.close()
                formula_workbook.close()

        expected_summary_members = {
            "晶格弯曲": [
                name
                for name in names
                if name.endswith("/Bending/TPU/Summary F TPU.xlsx")
            ],
            "晶格压缩": [
                name
                for name in names
                if name.endswith("/Compression/TPU/Summary C TPU.xlsx")
            ],
        }
        if any(len(items) != 1 for items in expected_summary_members.values()):
            raise AuditBlocked(f"FDM TPU Summary 工作簿集合漂移：{expected_summary_members}")
        for lattice_kind, summary_members in expected_summary_members.items():
            lattice_selection, evidence = extract_lattice_summary_selection(
                archive,
                summary_member=summary_members[0],
                kind=lattice_kind,
                result_targets=result_targets_by_stem,
                result_values=result_values_by_stem,
                specimen_ids_by_stem=specimen_ids_by_stem,
                n10_by_stem=n10_by_stem,
            )
            overlap = set(selection_by_stem) & set(lattice_selection)
            if overlap:
                raise AuditBlocked(f"FDM 来源内选择工作簿重复：{sorted(overlap)}")
            selection_by_stem.update(lattice_selection)
            summary_workbook_evidence.append(evidence)

        fea_extensions = {
            ".inp", ".odb", ".dat", ".msg", ".sta", ".cae", ".k", ".d3plot",
            ".vtk", ".vtu", ".h5", ".hdf5", ".cdb", ".rst",
        }
        solver_segments = {
            "fea", "simulation", "simulations", "numerical", "abaqus", "ansys", "comsol"
        }
        fea_candidates: list[str] = []
        for name in names:
            path = PurePosixPath(name)
            relevant_parts = [part.casefold() for part in path.parts[1:]]
            if path.suffix.casefold() in fea_extensions or any(
                part in solver_segments for part in relevant_parts[:-1]
            ):
                fea_candidates.append(name)
        readme_members = [
            name for name in names if PurePosixPath(name).name.casefold() == "readme.txt"
        ]
        if len(readme_members) != 1:
            raise AuditBlocked("FDM README 数量漂移")
        readme_text = archive.read(readme_members[0]).decode("utf-8", errors="replace")
        description_claims_numerical = "numerical" in readme_text.casefold()
        zs2_count = sum(PurePosixPath(name).suffix.casefold() == ".zs2" for name in names)
        stl_count = sum(PurePosixPath(name).suffix.casefold() == ".stl" for name in names)

    if set(selection_by_stem) != set(EXPECTED_FDM_SELECTION_STATES):
        raise AuditBlocked(
            f"FDM 来源内选择工作簿集合漂移：{sorted(selection_by_stem)}"
        )
    for curve, workbook_stem, specimen_id in curve_bindings:
        information = selection_by_stem[workbook_stem][specimen_id]
        for field in (
            "source_summary_state",
            "source_summary_evidence",
            "source_display_id",
            "formula_target_specimen_id",
            "quality_gate",
        ):
            curve[field] = information[field]
    for scalar in scalar_rows:
        information = selection_by_stem[str(scalar["工作簿"])][str(scalar["试样ID"])]
        for field in (
            "source_summary_state",
            "source_summary_evidence",
            "source_display_id",
            "formula_target_specimen_id",
            "quality_gate",
        ):
            scalar[field] = information[field]

    selection_counts = Counter(str(item["source_summary_state"]) for item in curves)
    if selection_counts != Counter({"selected": 57, "not_selected": 9, "conflict": 10}):
        raise AuditBlocked(f"FDM 76 曲线来源内选择计数漂移：{selection_counts}")
    counts_by_kind = Counter(str(item["试验类型"]) for item in curves)
    points_by_kind = {
        kind: sum(int(item["点数"]) for item in curves if item["试验类型"] == kind)
        for kind in sorted(counts_by_kind)
    }
    expected_counts = Counter({"晶格弯曲": 18, "晶格压缩": 18, "基材弯曲": 18, "基材拉伸": 22})
    expected_points = {"晶格弯曲": 416_583, "晶格压缩": 101_599, "基材弯曲": 126_604, "基材拉伸": 67_454}
    if counts_by_kind != expected_counts:
        raise AuditBlocked(f"FDM TPU 试样曲线数漂移：{counts_by_kind}")
    if points_by_kind != expected_points:
        raise AuditBlocked(f"FDM TPU 曲线点数漂移：{points_by_kind}")
    if len(curves) != 76 or sum(int(item["点数"]) for item in curves) != 712_240:
        raise AuditBlocked("FDM TPU 总试样数或总点数漂移")
    specimen_groups = {str(item["试样或家族组"]) for item in curves}
    if len(specimen_groups) != 76:
        raise AuditBlocked(f"FDM 独立试样组发生碰撞：{len(specimen_groups)}/76")
    scalar_groups = {str(item["试样组"]) for item in scalar_rows}
    if scalar_groups != specimen_groups:
        raise AuditBlocked(
            f"FDM 标量试样组未与76条曲线一一复用："
            f"标量组={len(scalar_groups)}/曲线组={len(specimen_groups)}"
        )
    if len(scalar_rows) != 1_206:
        raise AuditBlocked(f"FDM 标量审计总行数漂移：{len(scalar_rows)}/1206")
    for scalar in scalar_rows:
        value = scalar["value"]
        if value == "":
            if (
                scalar["source_summary_state"] != "conflict"
                or scalar["scalar_lineage_class"] != "blocked_summary_conflict_evidence"
            ):
                raise AuditBlocked(f"FDM 标量出现未登记空值：{scalar}")
        elif not is_finite_number(value):
            raise AuditBlocked(f"FDM 标量出现非有限值：{scalar}")
        if scalar["training_split"] != "false" or scalar["weight"] != "false":
            raise AuditBlocked("FDM 标量审计阶段意外分配拆分或权重")
    lineage_counts = Counter(str(item["scalar_lineage_class"]) for item in scalar_rows)
    formula_definition_counts = Counter(
        str(item["definition_id"])
        for item in scalar_rows
        if ".formula." in str(item["definition_id"])
    )
    expected_formula_definition_counts = Counter(
        {
            "fdm_lattice_bending.results.formula.maximum_flexural_stress": 18,
            "fdm_lattice_bending.results.formula.flexural_modulus": 18,
            "fdm_lattice_bending.results.formula.core_shear_stress": 18,
            "fdm_lattice_bending.results.formula.face_flexural_stress": 18,
            "fdm_lattice_compression.results.formula.compressive_stress": 18,
            "fdm_lattice_compression.results.formula.compressive_modulus": 18,
        }
    )
    if formula_definition_counts != expected_formula_definition_counts:
        raise AuditBlocked(
            f"FDM 显式公式标量定义计数漂移：{formula_definition_counts}"
        )
    resumen_rows = [
        item for item in scalar_rows if ".resumen." in str(item["definition_id"])
    ]
    if len(resumen_rows) != 210:
        raise AuditBlocked(f"FDM Resumen 六类逐试样证据行漂移：{len(resumen_rows)}/210")
    blocked_resumen = [item for item in resumen_rows if item["value"] == ""]
    if len(blocked_resumen) != 36:
        raise AuditBlocked(f"FDM Resumen 冲突空值证据行漂移：{len(blocked_resumen)}/36")

    duplicate_hashes = sorted(
        checksum
        for checksum, count in Counter(str(item["曲线SHA256"]) for item in curves).items()
        if count > 1
    )
    if duplicate_hashes:
        raise AuditBlocked(f"FDM TPU 出现完全相同曲线：{duplicate_hashes}")
    if fea_candidates:
        raise AuditBlocked(f"FDM 归档出现未登记 FEA 文件：{fea_candidates}")
    if zs2_count != 4 or stl_count != 6:
        raise AuditBlocked(
            f"FDM 专有试验文件或 STL 数量漂移：zs2={zs2_count}, stl={stl_count}"
        )

    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": FDM,
        "TPU实验试样曲线数": len(curves),
        "TPU实验完整点数": sum(int(item["点数"]) for item in curves),
        "独立试样组数": len(specimen_groups),
        "来源内选择状态计数": dict(sorted(selection_counts.items())),
        "来源内选择解释": {
            "selected": "来源汇总明确且两套证据一致地选入",
            "not_selected": "来源汇总主动未选用；这不是数据异常",
            "conflict": "显示编号、公式目标或两套汇总选择不一致；质量门保持关闭",
        },
        "按工作簿来源内选择": {
            stem: dict(
                sorted(
                    Counter(
                        information["source_summary_state"]
                        for information in selection.values()
                    ).items()
                )
            )
            for stem, selection in sorted(selection_by_stem.items())
        },
        "曲线数据角色": "工作簿内标准化应力应变；原始力/位移与公式血缘保留并逐行复核",
        "未来泄漏拆分家族": "dataset_doi|material_identity|print_batch；当前未创建拆分",
        "按试验类型曲线数": dict(sorted(counts_by_kind.items())),
        "按试验类型点数": points_by_kind,
        "完全重复曲线数": 0,
        "主工作簿数": len(primary_member_hashes),
        "主工作簿清单SHA256": manifest_hash(primary_member_hashes),
        "Summary工作簿证据": summary_workbook_evidence,
        "标量审计": {
            "标量行数": len(scalar_rows),
            "复用曲线试样组数": len(scalar_groups),
            "新增独立样本数": 0,
            "血缘类别计数": dict(sorted(lineage_counts.items())),
            "工作簿诊断": scalar_diagnostics,
            "全局聚合处理": "只记录来源证据，不把 Average/Promedio/Resumen 总结行当作试样",
            "缺失外部依赖处理": "blocked_missing_external 只保留证据，不给值或权重",
        },
        "专有ZS2文件数": zs2_count,
        "STL晶格几何数": stl_count,
        "README宣称含numerical": description_claims_numerical,
        "归档内可识别FEA文件": fea_candidates,
        "simulation_run_count": len(fea_candidates),
        "FEA处理结论": "归档无可识别求解器输入或输出，不能把README表述当作可训练仿真运行",
        "训练状态": {"training_split": False, "weight": False},
    }
    curves.sort(key=lambda item: (str(item["试验类型"]), str(item["曲线ID"])))
    scalar_rows.sort(
        key=lambda item: (
            str(item["工作簿"]),
            str(item["试样ID"]),
            str(item["definition_id"]),
        )
    )
    return summary, curves, scalar_rows


CELL_RANGE_PATTERN = re.compile(
    r"^(?:'((?:[^']|'')+)'|([^!]+))!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)$"
)


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def formula_under(series, container_names: set[str]) -> str | None:
    for element in series.iter():
        if local_name(element.tag) not in container_names:
            continue
        for child in element.iter():
            if local_name(child.tag) == "f" and child.text:
                return child.text
    return None


def parse_formula_range(formula: str) -> tuple[str, int, int, int, int] | None:
    match = CELL_RANGE_PATTERN.fullmatch(formula)
    if not match:
        return None
    sheet = (match.group(1) or match.group(2)).replace("''", "'")
    first_column = column_index_from_string(match.group(3))
    first_row = int(match.group(4))
    last_column = column_index_from_string(match.group(5))
    last_row = int(match.group(6))
    return sheet, first_column, first_row, last_column, last_row


def chart_series_ranges(workbook_bytes: bytes) -> tuple[set[tuple[str, int, int, int, int, int]], set[str]]:
    ranges: set[tuple[str, int, int, int, int, int]] = set()
    external_references: set[str] = set()
    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as package:
        chart_members = sorted(
            name
            for name in package.namelist()
            if name.startswith("xl/charts/chart") and name.endswith(".xml")
        )
        for member in chart_members:
            root = ElementTree.fromstring(package.read(member))
            for series in root.iter():
                if local_name(series.tag) != "ser":
                    continue
                x_formula = formula_under(series, {"xVal", "cat"})
                y_formula = formula_under(series, {"yVal", "val"})
                if not x_formula or not y_formula:
                    continue
                x_range = parse_formula_range(x_formula)
                y_range = parse_formula_range(y_formula)
                if x_range is None or y_range is None:
                    external_references.update(
                        formula for formula in (x_formula, y_formula) if formula.startswith("[")
                    )
                    continue
                if x_range[0].startswith("[") or y_range[0].startswith("["):
                    external_references.update((x_formula, y_formula))
                    continue
                if x_range[0] != y_range[0] or x_range[2] != y_range[2] or x_range[4] != y_range[4]:
                    raise AuditBlocked(f"仿真图表 X/Y 引用行范围不一致：{x_formula}; {y_formula}")
                if x_range[1] != x_range[3] or y_range[1] != y_range[3]:
                    raise AuditBlocked(f"仿真图表引用不是单列：{x_formula}; {y_formula}")
                ranges.add(
                    (
                        x_range[0],
                        x_range[1],
                        y_range[1],
                        x_range[2],
                        x_range[4],
                        len(chart_members),
                    )
                )
    return ranges, external_references


def audit_experiment_simulation(
    archive_path: Path,
) -> tuple[dict[str, object], list[dict[str, object]]]:
    curves: list[dict[str, object]] = []
    with zipfile.ZipFile(archive_path) as archive:
        raw_members = [name for name in archive.namelist() if name.endswith("Raw Data Experiment TPU.xlsx")]
        comparison_members = [
            name for name in archive.namelist() if name.endswith("Comparison Excel Experiment and Simulation.xlsx")
        ]
        if len(raw_members) != 1 or len(comparison_members) != 1:
            raise AuditBlocked("实验或仿真比较工作簿数量漂移")
        raw_bytes = archive.read(raw_members[0])
        comparison_bytes = archive.read(comparison_members[0])

    raw_workbook = load_workbook(
        io.BytesIO(raw_bytes), read_only=True, data_only=True, keep_links=False
    )
    try:
        if len(raw_workbook.worksheets) != 1:
            raise AuditBlocked("实验原始工作簿工作表数量漂移")
        worksheet = raw_workbook.worksheets[0]
        digests = [hashlib.sha256() for _ in range(3)]
        points = [0, 0, 0]
        common_overlap = 0
        potential_rows = 0
        for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=8, values_only=True):
            pairs = [(row[index], row[index + 5]) for index in range(3)]
            if not any(x is not None or y is not None for x, y in pairs):
                continue
            potential_rows += 1
            complete_mask: list[bool] = []
            for index, (x_cell, y_cell) in enumerate(pairs):
                if x_cell is None and y_cell is None:
                    complete_mask.append(False)
                    continue
                if not is_finite_number(x_cell) or not is_finite_number(y_cell):
                    raise AuditBlocked(f"实验重复{index+1}含不完整或非数值点")
                update_numeric_digest(digests[index], (float(x_cell), float(y_cell)))
                points[index] += 1
                complete_mask.append(True)
            if all(complete_mask):
                common_overlap += 1
        for index in range(3):
            curves.append(
                {
                    "来源": EXPERIMENT_SIMULATION,
                    "材料": "TPU_牌号未报告",
                    "试验类型": "单轴应力应变实验",
                    "条件": f"实验重复{index+1}",
                    "数据角色": "独立实验试样曲线",
                    "曲线ID": f"experiment_replicate_{index+1}",
                    "点数": points[index],
                    "曲线SHA256": digests[index].hexdigest(),
                    "试样或家族组": f"experiment_specimen_{index+1}",
                    "training_split": "false",
                    "weight": "false",
                    "备注": "Average列为派生列，未计入独立试样",
                }
            )
    finally:
        raw_workbook.close()

    series_ranges, external_references = chart_series_ranges(comparison_bytes)
    comparison_workbook = load_workbook(
        io.BytesIO(comparison_bytes), read_only=True, data_only=True, keep_links=False
    )
    try:
        valid_ranges = [item for item in series_ranges if item[0] in comparison_workbook.sheetnames]
        # 同一工作表在多个图表重复引用；规范化后每个仿真子运行只保留一个范围。
        normalized_ranges = sorted(
            {(sheet, x_col, y_col, start, end) for sheet, x_col, y_col, start, end, _ in valid_ranges},
            key=lambda item: item[0].casefold(),
        )
        if len(normalized_ranges) != 13:
            raise AuditBlocked(f"图表引用复算的仿真子运行数漂移：{len(normalized_ranges)}")
        referenced_sheets = {item[0] for item in normalized_ranges}
        chart_only_sheets = sorted(set(comparison_workbook.sheetnames) - referenced_sheets)
        for sheet_name, x_column, y_column, start_row, end_row in normalized_ranges:
            worksheet = comparison_workbook[sheet_name]
            value = hashlib.sha256()
            point_count = 0
            for row_number in range(start_row, end_row + 1):
                x_cell = worksheet.cell(row=row_number, column=x_column).value
                y_cell = worksheet.cell(row=row_number, column=y_column).value
                if not is_finite_number(x_cell) or not is_finite_number(y_cell):
                    raise AuditBlocked(f"仿真曲线引用范围含非数值点：{sheet_name}!{row_number}")
                update_numeric_digest(value, (float(x_cell), float(y_cell)))
                point_count += 1
            curves.append(
                {
                    "来源": EXPERIMENT_SIMULATION,
                    "材料": "TPU_牌号未报告",
                    "试验类型": "应力应变仿真",
                    "条件": sheet_name,
                    "数据角色": "仿真子运行曲线",
                    "曲线ID": f"simulation:{sheet_name}",
                    "点数": point_count,
                    "曲线SHA256": value.hexdigest(),
                    "试样或家族组": "simulation_family_comparison_workbook_1",
                    "training_split": "false",
                    "weight": "false",
                    "备注": f"图表实际引用范围{start_row}:{end_row}；缺少求解器与参数元数据",
                }
            )
    finally:
        comparison_workbook.close()

    experiment_curves = [item for item in curves if item["数据角色"] == "独立实验试样曲线"]
    simulation_curves = [item for item in curves if item["数据角色"] == "仿真子运行曲线"]
    if [int(item["点数"]) for item in experiment_curves] != [48_003, 48_002, 48_004]:
        raise AuditBlocked(f"实验三重复点数漂移：{[item['点数'] for item in experiment_curves]}")
    if sum(int(item["点数"]) for item in experiment_curves) != 144_009 or common_overlap != 48_002:
        raise AuditBlocked("实验三重复总点数或公共重叠长度漂移")
    if potential_rows != 48_004:
        raise AuditBlocked(f"实验原始潜在行数漂移：{potential_rows}")
    if len(simulation_curves) != 13 or sum(int(item["点数"]) for item in simulation_curves) != 6_453:
        raise AuditBlocked("仿真子运行数或图表实际引用点数漂移")
    family_count = len({str(item["试样或家族组"]) for item in simulation_curves})
    if family_count != 1:
        raise AuditBlocked(f"仿真家族数漂移：{family_count}")
    if chart_only_sheets != ["Compare M1, M2, M3", "Comparison", "Comparison Exp Sim"]:
        raise AuditBlocked(f"空白/图表专用工作表集合漂移：{chart_only_sheets}")

    summary = {
        "审计日期": AUDIT_DATE,
        "审计版本": AUDIT_VERSION,
        "来源": EXPERIMENT_SIMULATION,
        "实验原始工作簿": {
            "成员": raw_members[0],
            "成员SHA256": hashlib.sha256(raw_bytes).hexdigest(),
        },
        "仿真比较工作簿": {
            "成员": comparison_members[0],
            "成员SHA256": hashlib.sha256(comparison_bytes).hexdigest(),
        },
        "独立实验试样数": len(experiment_curves),
        "实验重复点数": [int(item["点数"]) for item in experiment_curves],
        "实验完整点数": sum(int(item["点数"]) for item in experiment_curves),
        "三重复公共重叠点数": common_overlap,
        "尾部缺失点数": [potential_rows - int(item["点数"]) for item in experiment_curves],
        "simulation_run_count": len(simulation_curves),
        "simulation_curve_points": sum(int(item["点数"]) for item in simulation_curves),
        "simulation_family_count": family_count,
        "仿真家族ID": "simulation_family_comparison_workbook_1",
        "空白或图表专用工作表": chart_only_sheets,
        "外部实验平均曲线图表引用": sorted(external_references),
        "仿真适用限制": "13条曲线属于同一比较工作簿家族；缺少求解器、模型、网格、材料参数与运行条件，不能当作13个独立物理体系",
        "训练状态": {"training_split": False, "weight": False},
    }
    curves.sort(key=lambda item: (str(item["数据角色"]), str(item["曲线ID"])))
    return summary, curves


def render_json(payload: dict[str, object]) -> bytes:
    return (
        json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False) + "\n"
    ).encode("utf-8")


def render_tsv(rows: list[dict[str, object]], columns: list[str]) -> bytes:
    if not rows:
        raise AuditBlocked("拒绝渲染空 TSV")
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        delimiter="\t",
        lineterminator="\n",
        extrasaction="raise",
    )
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def main() -> int:
    before = scientific_input_snapshot()
    for source in SOURCE_NAMES:
        validate_official_capture(source)
    input_summaries = snapshot_by_source(before)
    audit_functions = {
        FATIGUE: (fatigue_file_role, audit_fatigue),
        FDM: (fdm_file_role, audit_fdm),
        EXPERIMENT_SIMULATION: (experiment_simulation_file_role, audit_experiment_simulation),
    }
    rendered_outputs: dict[Path, bytes] = {}
    report: dict[str, object] = {}
    for source in SOURCE_NAMES:
        archive_name, _, _ = EXPECTED_ARCHIVES[source]
        archive_path = DATA_ROOT / source / archive_name
        role_function, scientific_audit = audit_functions[source]
        zip_summary, file_rows = audit_zip(source, archive_path, role_function)
        scientific_result = scientific_audit(archive_path)
        if source == FDM:
            scientific_summary, curve_rows, scalar_rows = scientific_result
            scalar_payload = render_tsv(scalar_rows, SCALAR_COLUMNS)
            scientific_summary["标量审计清单输出"] = {
                "文件名": FDM_SCALAR_OUTPUT_NAME,
                "行数": len(scalar_rows),
                "SHA256": hashlib.sha256(scalar_payload).hexdigest(),
                "training_split": False,
                "weight": False,
            }
        else:
            scientific_summary, curve_rows = scientific_result
            scalar_rows = []
            scalar_payload = b""
        scientific_summary["ZIP安全与完整性"] = zip_summary
        scientific_summary["输入快照"] = input_summaries[source]
        scientific_summary["治理状态"] = {
            "原始归档保留": True,
            "training_split": False,
            "weight": False,
            "自动训练": False,
        }
        base = DATA_ROOT / source
        rendered_outputs[base / "内容审计摘要.json"] = render_json(scientific_summary)
        rendered_outputs[base / "文件校验清单.tsv"] = render_tsv(file_rows, FILE_COLUMNS)
        rendered_outputs[base / "曲线审计清单.tsv"] = render_tsv(curve_rows, CURVE_COLUMNS)
        if source == FDM:
            rendered_outputs[base / FDM_SCALAR_OUTPUT_NAME] = scalar_payload
        report[source] = {
            "文件成员数": len(file_rows),
            "曲线数": len(curve_rows),
            "曲线点数": sum(int(item["点数"]) for item in curve_rows),
            "标量审计行数": len(scalar_rows),
        }

    after = scientific_input_snapshot()
    if before != after:
        raise AuditBlocked("审计过程中科学输入发生变化，拒绝写入输出")
    if set(rendered_outputs) != OUTPUT_WHITELIST:
        raise AuditBlocked("待写输出集合不等于固定白名单")
    for path in sorted(rendered_outputs, key=lambda item: str(item).casefold()):
        atomic_write(path, rendered_outputs[path])

    output_hashes = {
        path.relative_to(PROJECT_ROOT).as_posix(): hashlib.sha256(payload).hexdigest()
        for path, payload in sorted(rendered_outputs.items(), key=lambda item: str(item[0]).casefold())
    }
    print(
        json.dumps(
            {
                "status": "verified",
                "sources": report,
                "scientific_inputs_unchanged": True,
                "training_split": False,
                "weight": False,
                "output_sha256": output_hashes,
            },
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditBlocked as exc:
        print(f"BLOCKED: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc
