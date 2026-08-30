"""提取DRUM机械回收来源TPUU逐循环物理端点。"""

from __future__ import annotations
import argparse
import hashlib
import json
import tempfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from 提取TPUU循环端点 import extract_cycle_endpoints
from 接入DRUM机械回收 import parse_material_code

ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = (
    ROOT / "数据" / "原始" / "外部数据" / "新增开放数据" / "DRUM_TPUU_机械回收"
)
AUDIT = SOURCE_ROOT / "曲线审计清单.tsv"
OUT = ROOT / "结果" / "定向筛选" / "DRUM机械回收循环端点.csv"
MAN = ROOT / "结果" / "定向筛选" / "DRUM机械回收循环发布清单.json"


def sha(p):
    d = hashlib.sha256()
    d.update(p.read_bytes())
    return d.hexdigest()


def build_release():
    audit = pd.read_csv(AUDIT, sep="\t")
    selected = audit[(audit["试验类型"] == "循环滞回") & (audit["准入结论"] == "准入")]
    rows = []
    for rel, group in selected.groupby("文件相对路径"):
        path = SOURCE_ROOT / rel
        wb = load_workbook(path, read_only=True, data_only=True)
        for item in group.itertuples(index=False):
            ws = wb[str(getattr(item, "工作表"))]
            data = list(ws.iter_rows(values_only=True))
            thickness, width, gauge = data[7][6:9]
            raw = (
                pd.DataFrame(
                    [(r[0], r[1], r[2], r[3]) for r in data[3:] if len(r) >= 4],
                    columns=["time", "force", "stroke", "cycle"],
                )
                .apply(pd.to_numeric, errors="coerce")
                .dropna()
            )
            stress = raw.force / (float(thickness) * float(width))
            strain = 100 * raw.stroke / float(gauge)
            material = str(getattr(item, "配方键"))
            for cycle_no, idx in raw.groupby("cycle").groups.items():
                ep = extract_cycle_endpoints(
                    strain.loc[idx].to_numpy(), stress.loc[idx].to_numpy() * 1000
                )
                if not ep:
                    continue
                rows.append(
                    {
                        "source_id": "source_drum_tpuu_recycling_05ek6k60",
                        "formulation_id": material,
                        **parse_material_code(material),
                        "cycle_number": int(cycle_no),
                        **{k: v for k, v in ep[0].items() if k != "cycle_number"},
                        "chemistry_mapping_status": "family_mn_hard_segment_mapped",
                        "usage_mode": "primary_train",
                        "source_locator": f"{path.relative_to(ROOT).as_posix()}#sheet={ws.title};cycle={int(cycle_no)}",
                        "license": "CC0-1.0",
                        "citation_keys": "reference-53;reference-54",
                    }
                )
        wb.close()
    return (
        pd.DataFrame(rows)
        .sort_values(["formulation_id", "cycle_number"])
        .reset_index(drop=True)
    )


def write(f):
    f.to_csv(OUT, index=False, encoding="utf-8-sig", lineterminator="\n")
    MAN.write_text(
        json.dumps(
            {
                "counts": {
                    "curve_count": 22,
                    "formulation_count": f.formulation_id.nunique(),
                    "cycle_rows": len(f),
                    "valid_rows": int((f.quality_status == "valid").sum()),
                },
                "audit_sha256": sha(AUDIT),
                "output_sha256": sha(OUT),
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def check(f):
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / OUT.name
        f.to_csv(p, index=False, encoding="utf-8-sig", lineterminator="\n")
        if sha(p) != sha(OUT):
            raise SystemExit("DRUM循环端点不一致")
    print("DRUM循环端点检查通过")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--检查", action="store_true")
    a = p.parse_args()
    f = build_release()
    if a.检查:
        check(f)
    else:
        write(f)
        print(json.dumps({"cycles": len(f)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
