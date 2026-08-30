"""从Mendeley归档流式物化商业TPU温度/冲击疲劳与恢复端点。"""

from __future__ import annotations

import argparse
import csv
import functools
import hashlib
import json
import math
import re
import tempfile
from datetime import date
from io import BytesIO, TextIOWrapper
from pathlib import Path, PurePosixPath
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "Mendeley_商业TPU温度疲劳多工况"
)
ARCHIVE = SOURCE_DIR / "hc6npzvw3m-1.zip"
SOURCE_METADATA = SOURCE_DIR / "官方API元数据.json"
SOURCE_AUDIT = SOURCE_DIR / "内容审计摘要.json"
CURVE_AUDIT = SOURCE_DIR / "曲线审计清单.tsv"
FILE_AUDIT = SOURCE_DIR / "文件校验清单.tsv"
DIRECTED = ROOT / "结果" / "定向筛选"
HISTORIES = DIRECTED / "商业TPU温度疲劳端点.csv"
RECOVERY = DIRECTED / "商业TPU恢复配对端点.csv"
MANIFEST = DIRECTED / "商业TPU温度疲劳发布清单.json"
RELEASE_ID = "commercial-tpu-impact-fatigue-2024-v1"
DATASET_DOI = "10.17632/hc6npzvw3m.1"
LICENSE = "CC-BY-4.0"
CITATIONS = "reference-186"
SOURCE_SAMPLE_HEIGHT_MM = 13.5
SOURCE_PAD_RADIUS_M = 13.2522e-3
SOURCE_PAD_AREA_M2 = math.pi * SOURCE_PAD_RADIUS_M**2


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _entry(path: Path) -> dict[str, object]:
    return {
        "path": path.relative_to(ROOT).as_posix(),
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def _finite(value: object) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _normalise_material(value: object) -> str:
    text = str(value).strip()
    aliases = {
        "Elastollan 1195D": "Elastollan 1195A",
        "Elastolan 1154D": "Elastollan 1154D",
        "Texin-245": "Texin 245",
    }
    return aliases.get(text, text)


def _field(mapping: dict[str, object], token: str) -> object | None:
    token = token.lower()
    for key, value in mapping.items():
        if token in key.lower():
            return value
    return None


def _field_exact(mapping: dict[str, object], *names: str) -> object | None:
    normalised = {key.strip().lower(): value for key, value in mapping.items()}
    for name in names:
        if name.strip().lower() in normalised:
            return normalised[name.strip().lower()]
    return None


def _date_from_long_id(long_id: str) -> date:
    match = re.search(r"_(\d{2})_(\d{2})_(\d{4})_", long_id)
    if not match:
        raise ValueError(f"无法从Long ID解析日期: {long_id}")
    month, day, year = (int(value) for value in match.groups())
    return date(year, month, day)


def _metadata_rows(archive: ZipFile) -> tuple[list[dict[str, object]], dict[str, dict[str, object]]]:
    member = next(
        name
        for name in archive.namelist()
        if name.endswith("Final_data_all_experiments.xlsx")
    )
    workbook = load_workbook(
        BytesIO(archive.read(member)), read_only=True, data_only=True
    )
    rows = []
    try:
        for sheet_name in workbook.sheetnames[:4]:
            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            headers = next(iterator)
            for values in iterator:
                short_id = values[0] if values else None
                long_id = values[1] if len(values) > 1 else None
                material = values[2] if len(values) > 2 else None
                if not short_id or not long_id or not material:
                    continue
                normalised_material = _normalise_material(material)
                if normalised_material not in {
                    "Elastollan 1154D",
                    "Elastollan 1164D",
                    "Elastollan 1174D",
                    "Elastollan 1195A",
                    "Texin 245",
                } or not re.search(r"_(\d{2})_(\d{2})_(\d{4})_", str(long_id)):
                    continue
                mapping = {
                    str(headers[index]): value
                    for index, value in enumerate(values)
                    if index < len(headers) and headers[index] is not None
                }
                recovery = "recovery" in str(long_id).lower()
                temperature_value = _field_exact(
                    mapping, "Temperature (Celsius)", "Temperature"
                )
                temperature = _finite(temperature_value)
                if temperature is None and str(temperature_value).lower() == "ambient":
                    temperature = 20.0
                fatigue = _finite(
                    _field_exact(mapping, "Fatigue cycles (-)", "Fatigue cycles")
                )
                rows.append(
                    {
                        "source_sheet": sheet_name,
                        "source_short_id": str(short_id),
                        "source_long_id": str(long_id),
                        "short_base": str(short_id).split("_")[0],
                        "long_base": str(long_id).split("_")[0],
                        "material_source_label": str(material),
                        "material_grade": normalised_material,
                        "history_role": (
                            "ambient_recovery_retest"
                            if recovery
                            else "off_axis_fatigue_response"
                            if "off axis" in sheet_name.lower()
                            else "on_axis_fatigue_response"
                        ),
                        "orientation": (
                            "off_axis" if "off axis" in sheet_name.lower() else "on_axis"
                        ),
                        "fatigue_cycles": fatigue,
                        "temperature_C": temperature,
                        "test_date": _date_from_long_id(str(long_id)),
                        "repetition": _finite(
                            _field_exact(mapping, "Repetition (-)", "Repetition")
                        ),
                        "weight_g": _finite(_field(mapping, "weight")),
                        "height_mm": _finite(_field(mapping, "height")),
                        "diameter_mm": _finite(_field(mapping, "diameter")),
                        "strain_rate_s_1": _finite(_field(mapping, "strain rate")),
                        "deformation_speed_mm_min": _finite(
                            _field(mapping, "deformation speed")
                        ),
                        "energy_absorption_slope_source_J_m3": _finite(
                            _field(mapping, "energy absorbed slope")
                        ),
                        "strain_at_slope_threshold": _finite(
                            _field(mapping, "strain value energy absorption")
                        ),
                        "energy_absorption_50_source_J_m3": _finite(
                            _field(mapping, "energy absorbed 50")
                        ),
                    }
                )
        material_sheet = workbook["Material"]
        grade_properties = {}
        for values in material_sheet.iter_rows(min_row=2, values_only=True):
            if not values[0]:
                continue
            material = _normalise_material(values[0])
            if material not in {
                "Elastollan 1154D",
                "Elastollan 1164D",
                "Elastollan 1174D",
                "Elastollan 1195A",
                "Texin 245",
            }:
                continue
            grade_properties[material] = {
                "hardness_Shore_D": _finite(values[1]),
                "stress_at_100pct_MPa_TDS": _finite(values[2]),
                "density_g_cm3_TDS": _finite(values[4]),
                "Tg_C_TDS": _finite(values[6]),
                "brittleness_temperature_C_TDS": _finite(values[9]),
            }
    finally:
        workbook.close()
    return rows, grade_properties


def _read_curve(archive: ZipFile, member: str) -> list[tuple[float, float, float, float]]:
    points = []
    if member.lower().endswith(".xlsx"):
        workbook = load_workbook(
            BytesIO(archive.read(member)), read_only=True, data_only=True
        )
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            rows = worksheet.iter_rows(min_row=4, values_only=True)
            for values in rows:
                parsed = tuple(_finite(values[index]) for index in range(4))
                if all(value is not None for value in parsed):
                    points.append(parsed)
        finally:
            workbook.close()
    else:
        handle = TextIOWrapper(
            BytesIO(archive.read(member)), encoding="utf-8-sig", errors="replace"
        )
        for index, values in enumerate(csv.reader(handle)):
            if index < 3 or len(values) < 4:
                continue
            parsed = tuple(_finite(values[column]) for column in range(4))
            if all(value is not None for value in parsed):
                points.append(parsed)
    return points


def _trapz(points: list[tuple[float, float]]) -> float:
    return sum(
        (x1 - x0) * (y0 + y1) / 2
        for (x0, y0), (x1, y1) in zip(points, points[1:], strict=False)
    )


def _curve_endpoints(points: list[tuple[float, float, float, float]]) -> dict[str, object]:
    strains = [epsilon / SOURCE_SAMPLE_HEIGHT_MM for _, _, _, epsilon in points]
    stresses = [force / SOURCE_PAD_AREA_M2 for _, force, _, _ in points]
    start = next(
        (index for index, point in enumerate(points) if point[2] > 0), None
    )
    end = next(
        (index for index, strain in enumerate(strains) if strain > 0.5), None
    )
    energy_50 = math.nan
    if start is not None and end is not None and end > start:
        energy_50 = _trapz(
            [(strains[index], stresses[index]) for index in range(start, end + 1)]
        )
    peak_stroke_index = max(range(len(points)), key=lambda index: points[index][2])
    force_stroke = [
        (points[index][2] / 1000.0, max(points[index][1], 0.0))
        for index in range(peak_stroke_index + 1)
    ]
    return {
        "maximum_force_N": max(point[1] for point in points),
        "maximum_stroke_mm": max(point[2] for point in points),
        "maximum_compression_strain_percent": 100.0 * max(strains),
        "force_stroke_work_to_peak_J": _trapz(force_stroke),
        "energy_absorption_50_recomputed_J_m3": energy_50,
    }


def _common(material: str, grade_properties: dict[str, dict[str, object]]) -> dict[str, object]:
    supplier = "BASF" if material.startswith("Elastollan") else "Covestro"
    return {
        "release_id": RELEASE_ID,
        "source_id": f"doi:{DATASET_DOI}",
        "material_grade": material,
        "supplier": supplier,
        **grade_properties[material],
        "polymer_family": "commercial_thermoplastic_polyurethane",
        "thermoplastic_tpu_core": True,
        "chemistry_mapping_status": "commercial_grade_identity_only",
        "model_admission_layer": "core_tpu_application_experimental",
        "usage_mode": "compression_fatigue_and_recovery_supervision",
        "future_weight_ceiling": 0.60,
        "split_group": f"{DATASET_DOI}|{material}",
        "license": LICENSE,
        "citation_keys": CITATIONS,
    }


def _build_histories() -> pd.DataFrame:
    audit = pd.read_csv(CURVE_AUDIT, sep="\t")
    with ZipFile(ARCHIVE) as archive:
        metadata_rows, grade_properties = _metadata_rows(archive)
        lookup = {}
        for metadata in metadata_rows:
            recovery = metadata["history_role"] == "ambient_recovery_retest"
            lookup[(metadata["short_base"], recovery)] = metadata
            lookup[(metadata["long_base"], recovery)] = metadata
        members = [
            name
            for name in archive.namelist()
            if "/Load_frame_measurements/" in name
            and name.lower().endswith((".xlsx", ".csv"))
        ]
        member_by_stem = {PurePosixPath(name).stem: name for name in members}
        records = []
        for row in audit.to_dict(orient="records"):
            condition = str(row["条件"])
            physical_specimen = str(row["试样或家族组"])
            recovery = row["数据角色"] == "恢复曲线"
            metadata = lookup.get((physical_specimen, recovery))
            if metadata is None:
                raise ValueError(f"{condition}缺少元数据映射")
            member = member_by_stem.get(condition)
            if member is None:
                raise ValueError(f"{condition}缺少归档曲线成员")
            points = _read_curve(archive, member)
            if len(points) != int(row["点数"]):
                raise ValueError(f"{condition}点数与审计不一致")
            endpoints = _curve_endpoints(points)
            source_energy = metadata["energy_absorption_50_source_J_m3"]
            recomputed_energy = endpoints["energy_absorption_50_recomputed_J_m3"]
            relative_error = math.nan
            if source_energy is not None and math.isfinite(recomputed_energy):
                relative_error = 100.0 * abs(recomputed_energy - source_energy) / abs(
                    source_energy
                )
            records.append(
                {
                    **_common(str(row["材料"]), grade_properties),
                    "history_id": f"mendeley_hc6npzvw3m_{condition}",
                    "physical_specimen_id": physical_specimen,
                    "history_role": metadata["history_role"],
                    "orientation": metadata["orientation"],
                    "independent_physical_specimen": not recovery,
                    "fatigue_cycles": metadata["fatigue_cycles"],
                    "temperature_C": metadata["temperature_C"],
                    "test_date": metadata["test_date"].isoformat(),
                    "repetition": metadata["repetition"],
                    "strain_rate_s_1": metadata["strain_rate_s_1"],
                    "deformation_speed_mm_min": metadata[
                        "deformation_speed_mm_min"
                    ],
                    "weight_g": metadata["weight_g"],
                    "height_mm": metadata["height_mm"],
                    "diameter_mm": metadata["diameter_mm"],
                    "energy_absorption_slope_source_J_m3": metadata[
                        "energy_absorption_slope_source_J_m3"
                    ],
                    "strain_at_slope_threshold": metadata[
                        "strain_at_slope_threshold"
                    ],
                    "energy_absorption_50_source_J_m3": source_energy,
                    **endpoints,
                    "energy_recalculation_relative_error_percent": relative_error,
                    "endpoint_quality": (
                        "gold_source_recomputed_agree"
                        if math.isfinite(relative_error) and relative_error <= 0.5
                        else "silver_source_recomputed_within_1_5pct"
                        if math.isfinite(relative_error) and relative_error <= 1.5
                        else "silver_recomputed_only"
                        if source_energy is None and math.isfinite(recomputed_energy)
                        else "conditional_incomplete_or_discrepant"
                    ),
                    "source_material_label": metadata["material_source_label"],
                    "source_material_label_conflict": (
                        _normalise_material(metadata["material_source_label"])
                        != str(row["材料"])
                    ),
                    "source_long_id": metadata["source_long_id"],
                    "source_member": member,
                    "source_curve_sha256": str(row["曲线SHA256"]),
                    "source_point_count": len(points),
                    "energy_method": (
                        "source_MATLAB_fixed_height_13.5mm_fixed_radius_13.2522mm_"
                        "integral_to_first_strain_above_0.5"
                    ),
                }
            )
    frame = pd.DataFrame(records)
    baseline_rows = frame.loc[
        frame["history_role"].ne("ambient_recovery_retest")
        & frame["fatigue_cycles"].eq(0)
        & frame["energy_absorption_50_source_J_m3"].notna()
    ]
    baselines = (
        baseline_rows.groupby(["material_grade", "temperature_C", "orientation"])[
            "energy_absorption_50_source_J_m3"
        ]
        .mean()
        .to_dict()
    )
    frame["baseline_N0_group_mean_energy_J_m3"] = frame.apply(
        lambda row: baselines.get(
            (row.material_grade, row.temperature_C, row.orientation), math.nan
        ),
        axis=1,
    )
    frame["energy_retention_vs_N0_group_mean_percent"] = 100.0 * frame[
        "energy_absorption_50_source_J_m3"
    ] / frame["baseline_N0_group_mean_energy_J_m3"]
    return frame.sort_values(["material_grade", "physical_specimen_id", "history_role"]).reset_index(
        drop=True
    )


def _build_recovery(histories: pd.DataFrame) -> pd.DataFrame:
    records = []
    for recovered in histories.loc[
        histories["history_role"].eq("ambient_recovery_retest")
    ].itertuples(index=False):
        paired = histories.loc[
            histories["physical_specimen_id"].eq(recovered.physical_specimen_id)
            & histories["history_role"].eq("on_axis_fatigue_response")
            & histories["fatigue_cycles"].eq(100)
        ]
        if len(paired) != 1:
            raise ValueError(
                f"{recovered.physical_specimen_id}的100次疲劳配对数量不是1"
            )
        fatigued = paired.iloc[0]
        recovered_energy = recovered.energy_absorption_50_source_J_m3
        fatigued_energy = fatigued.energy_absorption_50_source_J_m3
        records.append(
            {
                "release_id": RELEASE_ID,
                "source_id": f"doi:{DATASET_DOI}",
                "material_grade": recovered.material_grade,
                "physical_specimen_id": recovered.physical_specimen_id,
                "paired_fatigued_history_id": fatigued.history_id,
                "recovery_history_id": recovered.history_id,
                "prior_fatigue_cycles": 100,
                "fatigued_test_date": fatigued.test_date,
                "recovery_test_date": recovered.test_date,
                "recovery_elapsed_days": (
                    date.fromisoformat(recovered.test_date)
                    - date.fromisoformat(fatigued.test_date)
                ).days,
                "fatigued_energy_absorption_50_J_m3": fatigued_energy,
                "recovered_energy_absorption_50_J_m3": recovered_energy,
                "recovery_energy_ratio_vs_fatigued_percent": 100.0
                * recovered_energy
                / fatigued_energy,
                "recovery_energy_change_percent": 100.0
                * (recovered_energy - fatigued_energy)
                / fatigued_energy,
                "baseline_N0_group_mean_energy_J_m3": recovered.baseline_N0_group_mean_energy_J_m3,
                "recovered_energy_retention_vs_N0_percent": recovered.energy_retention_vs_N0_group_mean_percent,
                "direct_property_recovery_available": True,
                "recovery_property": "compression_energy_absorption_at_50pct_strain",
                "direct_shape_recovery_available": False,
                "shape_recovery_ratio_percent": math.nan,
                "recovery_metric_status": (
                    "exact_same_specimen_ambient_energy_recovery_retest_not_shape_recovery"
                ),
                "thermoplastic_tpu_core": True,
                "model_admission_layer": "core_tpu_application_experimental",
                "chemistry_mapping_status": "commercial_grade_identity_only",
                "split_group": recovered.split_group,
                "license": LICENSE,
                "citation_keys": CITATIONS,
            }
        )
    return pd.DataFrame(records).sort_values(
        ["material_grade", "physical_specimen_id"]
    ).reset_index(drop=True)


@functools.lru_cache(maxsize=1)
def _build_cached() -> tuple[pd.DataFrame, pd.DataFrame]:
    histories = _build_histories()
    recovery = _build_recovery(histories)
    return histories, recovery


def build_release() -> tuple[pd.DataFrame, pd.DataFrame]:
    return tuple(frame.copy() for frame in _build_cached())


def _write_frames(
    histories: pd.DataFrame, recovery: pd.DataFrame, directory: Path
) -> dict[str, Path]:
    paths = {
        "histories": directory / HISTORIES.name,
        "recovery": directory / RECOVERY.name,
    }
    histories.to_csv(
        paths["histories"], index=False, encoding="utf-8-sig", lineterminator="\n"
    )
    recovery.to_csv(
        paths["recovery"], index=False, encoding="utf-8-sig", lineterminator="\n"
    )
    return paths


def write_release(histories: pd.DataFrame, recovery: pd.DataFrame) -> None:
    DIRECTED.mkdir(parents=True, exist_ok=True)
    paths = _write_frames(histories, recovery, DIRECTED)
    payload = {
        "release_id": RELEASE_ID,
        "counts": {
            "commercial_tpu_grade_count": int(histories["material_grade"].nunique()),
            "independent_physical_specimen_count": int(
                histories["physical_specimen_id"].nunique()
            ),
            "curve_history_count": len(histories),
            "on_axis_history_count": int(
                histories["history_role"].eq("on_axis_fatigue_response").sum()
            ),
            "off_axis_history_count": int(
                histories["history_role"].eq("off_axis_fatigue_response").sum()
            ),
            "recovery_retest_history_count": int(
                histories["history_role"].eq("ambient_recovery_retest").sum()
            ),
            "recovery_pair_count": len(recovery),
            "raw_curve_point_count": int(histories["source_point_count"].sum()),
            "source_energy_summary_row_count": int(
                histories["energy_absorption_50_source_J_m3"].notna().sum()
            ),
            "recomputed_energy_row_count": int(
                histories["energy_absorption_50_recomputed_J_m3"].notna().sum()
            ),
            "source_recomputed_comparable_row_count": int(
                histories["energy_recalculation_relative_error_percent"].notna().sum()
            ),
            "published_compact_row_count": len(histories) + len(recovery),
        },
        "counting_note": (
            "196条曲线历史属于190个独立TPU物理试样；6条恢复复测复用原试样ID。"
            "333492个曲线点不作为独立样本。"
        ),
        "source": {
            "dataset_doi": DATASET_DOI,
            "license": LICENSE,
            "archive": _entry(ARCHIVE),
            "metadata": _entry(SOURCE_METADATA),
            "source_audit": _entry(SOURCE_AUDIT),
            "curve_audit": _entry(CURVE_AUDIT),
            "file_audit": _entry(FILE_AUDIT),
        },
        "policy": {
            "raw_curves_republished": False,
            "raw_curve_reason": (
                "以196条紧凑历史和6条恢复配对替代333492行曲线点；归档由SHA256复现"
            ),
            "commercial_grade_not_exact_chemistry": True,
            "direct_shape_recovery_available": False,
            "direct_energy_recovery_pair_available": True,
            "santoprene_excluded_from_tpu_package": True,
        },
        "outputs": {key: _entry(path) for key, path in paths.items()},
    }
    MANIFEST.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def check_release(histories: pd.DataFrame, recovery: pd.DataFrame) -> None:
    with tempfile.TemporaryDirectory(prefix="commercial-tpu-fatigue-check-") as directory:
        temporary = _write_frames(histories, recovery, Path(directory))
        published = {"histories": HISTORIES, "recovery": RECOVERY}
        mismatches = [
            key
            for key in published
            if _sha256(temporary[key]) != _sha256(published[key])
        ]
        if mismatches:
            raise SystemExit(f"商业TPU温度疲劳输出不一致: {','.join(mismatches)}")
    print("商业TPU温度疲劳数据检查通过")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--检查", action="store_true")
    args = parser.parse_args()
    histories, recovery = build_release()
    if args.检查:
        check_release(histories, recovery)
    else:
        write_release(histories, recovery)
        print(
            json.dumps(
                {
                    "history_rows": len(histories),
                    "physical_specimens": int(
                        histories["physical_specimen_id"].nunique()
                    ),
                    "recovery_pairs": len(recovery),
                },
                ensure_ascii=False,
            )
        )


if __name__ == "__main__":
    main()
