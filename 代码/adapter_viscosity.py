"""预聚体温度—黏度工作簿的显式 sheet 适配器。"""

from __future__ import annotations

import math
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ids import stable_id
from units import temperature_to_k, viscosity_to_pa_s


SCHEMA_VERSION = "v0.1"
SHEET_PATTERN = re.compile(r"^(?P<polyol>[A-Za-z]+)_(?P<isocyanate>[0-9A-Za-z]+)_(?P<pnco>[0-9]+)$")
EXPECTED_SHEETS = (
    "P_44M_4",
    "P_44M_6",
    "P_44M_8",
    "P_44M_10",
    "D_44M_4",
    "D_44M_6",
    "D_44M_8",
    "D_44M_10",
    "D_MLQ_4",
    "D_MLQ_6",
    "D_MLQ_8",
    "D_MLQ_10",
    "P_MLQ_4",
    "P_MLQ_6",
    "P_MLQ_8",
    "P_MLQ_10",
    "C_44M_6",
    "C_44M_8",
    "C_44M_10",
    "C_MLQ_4",
    "C_MLQ_6",
    "C_MLQ_8",
    "C_MLQ_10",
    "P_TD80_4",
    "P_TD80_7",
    "P_TD80_10",
    "D_TD80_7",
    "D_TD80_10",
    "P_I_5",
    "P_I_9",
    "P_W_5",
    "P_W_9",
    "D_W_5",
    "D_W_9",
    "D_I_5",
    "D_I_9",
    "P_TDS_4",
    "P_TDS_7",
    "P_TDS_10",
)

CURVE_COLUMNS = (
    "record_id",
    "curve_id",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "curve_type",
    "polyol_code",
    "isocyanate_code",
    "p_nco_percent",
    "temperature_unit_raw",
    "temperature_unit",
    "viscosity_unit_raw",
    "viscosity_unit",
    "unit_evidence",
    "point_count",
)
POINT_COLUMNS = (
    "record_id",
    "curve_id",
    "point_index",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "temperature_raw",
    "temperature_unit_raw",
    "temperature_k",
    "temperature_unit",
    "viscosity_raw",
    "viscosity_unit_raw",
    "viscosity_pa_s",
    "viscosity_unit",
    "unit_evidence",
)
AUDIT_COLUMNS = (
    "record_id",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "sheet_name",
    "status",
    "reason",
    "nonempty_cell_count",
)


def _require_provenance(source_id: str, source_file_id: str) -> None:
    for name, value in (("source_id", source_id), ("source_file_id", source_file_id)):
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"{name} must be a non-empty string")


def _number(value: Any, locator: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"Expected numeric value at {locator}, got boolean")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Expected numeric value at {locator}, got {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"Expected finite value at {locator}, got {value!r}")
    return number


def _nonempty_cell_count(sheet: Any) -> int:
    return sum(
        value is not None
        for row in sheet.iter_rows(values_only=True)
        for value in row
    )


def _audit_row(
    sheet_name: str,
    status: str,
    reason: str,
    nonempty_count: int,
    source_id: str,
    source_file_id: str,
) -> dict[str, Any]:
    return {
        "record_id": stable_id("audit", source_file_id, sheet_name, status),
        "source_id": source_id,
        "source_file_id": source_file_id,
        "source_locator": f"sheet={sheet_name}",
        "extraction_method": "openpyxl:sheet_inventory",
        "fidelity": "audit_metadata",
        "schema_version": SCHEMA_VERSION,
        "sheet_name": sheet_name,
        "status": status,
        "reason": reason,
        "nonempty_cell_count": nonempty_count,
    }


def adapt_viscosity(
    path: str | Path,
    *,
    source_id: str,
    source_file_id: str,
    expected_sheets: Iterable[str] = EXPECTED_SHEETS,
) -> dict[str, pd.DataFrame]:
    """Extract one viscosity-temperature curve per explicitly expected sheet."""

    _require_provenance(source_id, source_file_id)
    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Viscosity source must be XLSX: {source_path}")

    expected = tuple(expected_sheets)
    if len(expected) != len(set(expected)) or not expected:
        raise ValueError("expected_sheets must contain unique, non-empty sheet names")
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    missing = [sheet for sheet in expected if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"Viscosity sheet fingerprint mismatch: missing {missing!r}")

    curves: list[dict[str, Any]] = []
    points: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    for sheet_name in expected:
        match = SHEET_PATTERN.fullmatch(sheet_name)
        if match is None:
            raise ValueError(
                f"Viscosity sheet fingerprint mismatch: unsupported expected name {sheet_name!r}"
            )
        sheet = workbook[sheet_name]
        header = (sheet.cell(1, 1).value, sheet.cell(1, 2).value)
        if header != ("Viscosity", "Temperature") or any(
            sheet.cell(1, column).value is not None
            for column in range(3, sheet.max_column + 1)
        ):
            raise ValueError(
                f"Viscosity header fingerprint mismatch in {sheet_name}: got {header!r}"
            )

        second_row = (sheet.cell(2, 1).value, sheet.cell(2, 2).value)
        if any(isinstance(value, str) for value in second_row):
            if second_row != ("Pa.s", "\N{DEGREE SIGN}C"):
                raise ValueError(
                    f"Viscosity unit-row fingerprint mismatch in {sheet_name}: {second_row!r}"
                )
            start_row = 3
            unit_evidence = "sheet_unit_row"
        else:
            start_row = 2
            unit_evidence = "upstream_workbook_contract"

        raw_points: list[tuple[int, float, float]] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=start_row, max_col=2, values_only=True),
            start=start_row,
        ):
            viscosity, temperature = row
            if viscosity is None and temperature is None:
                continue
            if viscosity is None or temperature is None:
                raise ValueError(
                    f"Incomplete viscosity-temperature pair at {sheet_name}!A{row_number}:B{row_number}"
                )
            viscosity_value = _number(viscosity, f"{sheet_name}!A{row_number}")
            temperature_value = _number(temperature, f"{sheet_name}!B{row_number}")
            if viscosity_value <= 0:
                raise ValueError(f"Viscosity must be positive at {sheet_name}!A{row_number}")
            if temperature_to_k(temperature_value, "degC") <= 0:
                raise ValueError(f"Temperature is below absolute zero at {sheet_name}!B{row_number}")
            raw_points.append((row_number, temperature_value, viscosity_value))
        if not raw_points:
            raise ValueError(f"Viscosity sheet {sheet_name!r} contains no data points")

        polyol = match.group("polyol")
        isocyanate = match.group("isocyanate")
        p_nco = int(match.group("pnco"))
        curve_id = stable_id("curve", source_file_id, sheet_name, "viscosity_temperature")
        common = {
            "source_id": source_id,
            "source_file_id": source_file_id,
            "extraction_method": "openpyxl:explicit_sheet_and_column_map",
            "fidelity": "measured_raw",
            "schema_version": SCHEMA_VERSION,
        }
        curves.append(
            {
                "record_id": curve_id,
                "curve_id": curve_id,
                **common,
                "source_locator": f"sheet={sheet_name};columns=A:B",
                "curve_type": "viscosity_temperature",
                "polyol_code": polyol,
                "isocyanate_code": isocyanate,
                "p_nco_percent": p_nco,
                "temperature_unit_raw": "degC",
                "temperature_unit": "K",
                "viscosity_unit_raw": "Pa.s",
                "viscosity_unit": "Pa·s",
                "unit_evidence": unit_evidence,
                "point_count": len(raw_points),
            }
        )
        for point_index, (row_number, temperature, viscosity) in enumerate(raw_points):
            points.append(
                {
                    "record_id": stable_id("curve_point", curve_id, point_index),
                    "curve_id": curve_id,
                    "point_index": point_index,
                    **common,
                    "source_locator": f"sheet={sheet_name};cells=A{row_number}:B{row_number}",
                    "temperature_raw": temperature,
                    "temperature_unit_raw": "degC",
                    "temperature_k": temperature_to_k(temperature, "degC"),
                    "temperature_unit": "K",
                    "viscosity_raw": viscosity,
                    "viscosity_unit_raw": "Pa.s",
                    "viscosity_pa_s": viscosity_to_pa_s(viscosity, "Pa.s"),
                    "viscosity_unit": "Pa·s",
                    "unit_evidence": unit_evidence,
                }
            )

    expected_set = set(expected)
    for sheet_name in workbook.sheetnames:
        if sheet_name in expected_set:
            continue
        nonempty_count = _nonempty_cell_count(workbook[sheet_name])
        status = (
            "unrecognized_nonempty_sheet" if nonempty_count else "unrecognized_empty_sheet"
        )
        audit.append(
            _audit_row(
                sheet_name,
                status,
                "sheet is outside the explicit v0.1 source mapping",
                nonempty_count,
                source_id,
                source_file_id,
            )
        )

    workbook.close()
    return {
        "curves": pd.DataFrame(curves, columns=CURVE_COLUMNS),
        "curve_points": pd.DataFrame(points, columns=POINT_COLUMNS),
        "audit": pd.DataFrame(audit, columns=AUDIT_COLUMNS),
    }


__all__ = ["EXPECTED_SHEETS", "SCHEMA_VERSION", "adapt_viscosity"]
