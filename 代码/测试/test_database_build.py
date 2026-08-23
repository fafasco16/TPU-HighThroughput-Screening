from __future__ import annotations

import json
import shutil
from pathlib import Path

import duckdb
import pandas as pd
import pytest
import yaml
from openpyxl import Workbook

from database_build import (
    PIPELINE_VERSION,
    REQUIRED_SOURCES,
    SCHEMA_VERSION,
    DatabaseBuildResult,
    _public_view,
    build_database,
    derive_curve_metrics,
    extract_staging_tables,
    normalize_staging_tables,
    pipeline_signature,
    run_quality_checks,
)
from qc import QualityIssue
from snapshot import sha256_file


FIXTURES = Path(__file__).parent / "夹具"
PROJECT_ROOT = Path(__file__).resolve().parents[2]


def _write_hbond(path: Path) -> None:
    spec = json.loads((FIXTURES / "xlsx_最小结构.json").read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in spec["hbond"].items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def _write_viscosity(path: Path) -> None:
    spec = json.loads((FIXTURES / "xlsx_最小结构.json").read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in spec["viscosity"].items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def _source_row(
    root: Path,
    source_id: str,
    relative_path: str,
    *,
    license_spdx: str,
    derivatives_allowed: bool | None,
    redistribution_allowed: bool | None,
    material_scope: str,
    status: str,
) -> dict[str, object]:
    path = root / relative_path
    return {
        "source_id": source_id,
        "source_file_id": f"file_{source_id}",
        "raw_path": relative_path,
        "original_filename": path.name,
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "doi": "",
        "url": f"https://example.test/{source_id}",
        "accessed_at": "2026-07-18",
        "license_spdx": license_spdx,
        "derivatives_allowed": derivatives_allowed,
        "redistribution_allowed": redistribution_allowed,
        "access_restriction": "open",
        "evidence_grade": (
            "candidate_structure"
            if source_id == "ds_smipoly_monomers"
            else "measured_raw"
        ),
        "material_scope": material_scope,
        "status": status,
        "notes": "test fixture",
    }


def _project(tmp_path: Path) -> tuple[Path, list[dict[str, object]], dict[str, dict[str, object]]]:
    root = tmp_path / "项目"
    for directory in (
        "数据/原始",
        "数据/暂存",
        "数据/规范",
        "数据/派生",
        "数据/快照",
        "结果",
        "配置/结构定义",
    ):
        (root / directory).mkdir(parents=True, exist_ok=True)
    shutil.copyfile(
        PROJECT_ROOT / "配置/结构定义" / "v0.1枚举.yaml",
        root / "配置/结构定义" / "v0.1枚举.yaml",
    )

    smipoly_path = root / "数据/原始" / "smipoly.csv"
    pue_path = root / "数据/原始" / "pue.csv"
    hbond_path = root / "数据/原始" / "hbond.xlsx"
    viscosity_path = root / "数据/原始" / "viscosity.xlsx"
    shutil.copyfile(FIXTURES / "smipoly_最小.csv", smipoly_path)
    shutil.copyfile(FIXTURES / "pue326_最小.csv", pue_path)
    _write_hbond(hbond_path)
    _write_viscosity(viscosity_path)

    rows = [
        _source_row(
            root,
            "ds_smipoly_monomers",
            "数据/原始/smipoly.csv",
            license_spdx="BSD-3-Clause",
            derivatives_allowed=True,
            redistribution_allowed=True,
            material_scope="virtual_candidate",
            status="available",
        ),
        _source_row(
            root,
            "ds_pue326_dq",
            "数据/原始/pue.csv",
            license_spdx="UNKNOWN",
            derivatives_allowed=None,
            redistribution_allowed=None,
            material_scope="crosslinked_pue",
            status="review_required",
        ),
        _source_row(
            root,
            "ds_eom_hbond_2021",
            "数据/原始/hbond.xlsx",
            license_spdx="CC-BY-4.0",
            derivatives_allowed=True,
            redistribution_allowed=True,
            material_scope="linear_tpu",
            status="available",
        ),
        _source_row(
            root,
            "ds_prepolymer_viscosity",
            "数据/原始/viscosity.xlsx",
            license_spdx="UNKNOWN",
            derivatives_allowed=None,
            redistribution_allowed=None,
            material_scope="polyurethane_prepolymer",
            status="review_required",
        ),
    ]
    options = {
        "ds_eom_hbond_2021": {"expected_sheets": ("Figure 1b",)},
        "ds_prepolymer_viscosity": {"expected_sheets": ("P_44M_4",)},
    }
    return root, rows, options


def test_four_source_build_is_layered_gated_and_reproducible(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    result = build_database(root, rows, adapter_options=options)

    assert isinstance(result, DatabaseBuildResult)
    assert result.snapshot_id.startswith("snapshot_")
    assert result.has_errors is False
    assert result.issues == ()
    assert result.row_counts["staging_smipoly_chemical"] == 2
    assert result.row_counts["normalized_chemical_candidate"] == 1
    assert result.row_counts["staging_pue_transformed"] == 2
    assert result.row_counts["normalized_curve"] == 8
    assert result.row_counts["normalized_curve_point"] == 16
    assert result.row_counts["derived_property"] == 21
    assert result.row_counts["public_pue_transformed_auxiliary"] == 0
    assert result.row_counts["public_derived_property"] == 21
    assert result.to_dict()["has_errors"] is False

    required_output_kinds = {
        "duckdb",
        "snapshot_json",
        "qc_csv",
        "qc_json",
        "field_coverage_csv",
        "normalized_curve",
        "snapshot_normalized_curve",
    }
    assert required_output_kinds <= set(result.outputs)
    for name, metadata in result.outputs.items():
        assert (root / str(metadata["path"])).is_file()
        if name == "duckdb":
            assert metadata["byte_reproducible"] is False
            assert metadata["content_basis"] == "snapshot_parquet_sha256_and_row_counts"
            assert "sha256" not in metadata
        else:
            assert len(str(metadata["sha256"])) == 64

    database_path = root / result.outputs["duckdb"]["path"]
    with duckdb.connect(str(database_path), read_only=True) as connection:
        public_curve_sources = {
            row[0]
            for row in connection.execute(
                "SELECT DISTINCT source_id FROM public_curve"
            ).fetchall()
        }
        public_pue_count = connection.execute(
            "SELECT count(*) FROM public_pue_transformed_auxiliary"
        ).fetchone()[0]
        mapped_types = {
            row[0]
            for row in connection.execute(
                "SELECT DISTINCT curve_type FROM normalized_curve"
            ).fetchall()
        }
    assert public_curve_sources == {"ds_eom_hbond_2021"}
    assert public_pue_count == 0
    assert mapped_types == {"tensile_monotonic", "viscosity_temperature"}

    coverage = pd.read_csv(root / result.outputs["field_coverage_csv"]["path"])
    assert set(coverage.columns) == {
        "table_name",
        "column_name",
        "dtype",
        "row_count",
        "present_count",
        "missing_count",
        "missing_fraction",
    }
    curve_id_coverage = coverage[
        (coverage["table_name"] == "curve")
        & (coverage["column_name"] == "curve_id")
    ].iloc[0]
    assert curve_id_coverage["missing_fraction"] == 0.0

    first_snapshot = (root / result.outputs["snapshot_json"]["path"]).read_bytes()
    snapshot_payload = json.loads(first_snapshot)
    assert snapshot_payload["pipeline"]["pipeline_version"] == PIPELINE_VERSION
    assert snapshot_payload["pipeline"]["pipeline_id"].startswith("pipeline_")
    assert len(snapshot_payload["pipeline"]["files"]) >= 11
    assert pipeline_signature(root) == snapshot_payload["pipeline"]
    assert snapshot_payload["build_parameters"] == {
        "adapter_options": {
            "ds_eom_hbond_2021": {"expected_sheets": ["Figure 1b"]},
            "ds_prepolymer_viscosity": {"expected_sheets": ["P_44M_4"]},
        }
    }
    first_parquet_hash = result.outputs["snapshot_normalized_curve"]["sha256"]
    second = build_database(root, pd.DataFrame(rows), adapter_options=options)
    assert second.snapshot_id == result.snapshot_id
    assert second.outputs["snapshot_normalized_curve"]["sha256"] == first_parquet_hash
    assert (root / second.outputs["snapshot_json"]["path"]).read_bytes() == first_snapshot

    list_options = {
        source_id: {"expected_sheets": list(values["expected_sheets"])}
        for source_id, values in options.items()
    }
    equivalent = build_database(root, rows, adapter_options=list_options)
    assert equivalent.snapshot_id == result.snapshot_id

    changed_options = {
        **options,
        "ds_eom_hbond_2021": {"expected_sheets": ("Figure 1b", "Mystery")},
    }
    changed_build = build_database(root, rows, adapter_options=changed_options)
    assert changed_build.snapshot_id != result.snapshot_id

    changed_metadata = [dict(row) for row in rows]
    changed_metadata[0]["url"] = "https://example.test/revised-provenance"
    revised = build_database(root, changed_metadata, adapter_options=options)
    assert revised.snapshot_id != result.snapshot_id


def test_normalized_records_use_v01_schema_and_declared_enums(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    staging, selected = extract_staging_tables(root, rows, adapter_options=options)
    normalized = normalize_staging_tables(staging, selected)

    enums = yaml.safe_load(
        (PROJECT_ROOT / "配置/结构定义" / "v0.1枚举.yaml").read_text(encoding="utf-8")
    )["enums"]
    fields = yaml.safe_load(
        (PROJECT_ROOT / "配置/结构定义" / "v0.1字段字典.yaml").read_text(
            encoding="utf-8"
        )
    )
    assert fields["schema_version"] == SCHEMA_VERSION == "v0.1"
    assert fields["tables"]["chemical"]["primary_key"] == ["chemical_id"]
    assert fields["tables"]["source"]["fields"]["status"]["enum"] == (
        "source_status"
    )
    assert normalized["source"]["status"].isin(enums["source_status"]).all()
    assert normalized["source"]["access_restriction"].isin(
        enums["access_restriction"]
    ).all()
    assert normalized["chemical_candidate"]["chemical_id"].is_unique
    assert normalized["chemical_candidate"]["fidelity"].eq(
        "candidate_structure"
    ).all()
    assert normalized["chemical_candidate"]["extraction_method"].eq(
        "direct_table"
    ).all()
    assert normalized["chemical_candidate"]["preferred_name"].isna().all()
    assert normalized["pue_transformed_auxiliary"]["table_role"].eq(
        "transformed_feature_auxiliary"
    ).all()
    assert normalized["curve"]["curve_type"].isin(enums["curve_type"]).all()
    assert normalized["curve"]["unit_status"].isin(enums["unit_status"]).all()
    assert normalized["curve"]["extraction_method"].isin(
        enums["extraction_method"]
    ).all()
    assert normalized["curve_point"]["unit_status"].isin(
        enums["unit_status"]
    ).all()
    assert normalized["curve"]["test_link_status"].str.contains(
        "no_specimen"
    ).all()
    for frame in normalized.values():
        if "schema_version" in frame.columns:
            assert frame["schema_version"].eq("v0.1").all()


def test_manifest_guards_missing_sources_hashes_and_paths(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    with pytest.raises(ValueError, match="缺少来源"):
        extract_staging_tables(root, rows[:-1], adapter_options=options)

    bad_hash = [dict(row) for row in rows]
    bad_hash[0]["sha256"] = "0" * 64
    with pytest.raises(ValueError, match="哈希不匹配"):
        extract_staging_tables(root, bad_hash, adapter_options=options)

    for invalid_hash in ("", " ", None, "0" * 63, "0" * 65, "g" * 64):
        invalid = [dict(row) for row in rows]
        invalid[0]["sha256"] = invalid_hash
        with pytest.raises(ValueError, match="64位十六进制"):
            extract_staging_tables(root, invalid, adapter_options=options)

    duplicate = [dict(row) for row in rows]
    duplicate[1]["source_file_id"] = duplicate[0]["source_file_id"]
    with pytest.raises(ValueError, match="必须非空且唯一"):
        extract_staging_tables(root, duplicate, adapter_options=options)

    missing_column = pd.DataFrame(rows).drop(columns=["status"])
    with pytest.raises(ValueError, match="缺少必需字段"):
        extract_staging_tables(root, missing_column, adapter_options=options)

    escaped = [dict(row) for row in rows]
    escaped[0]["raw_path"] = "../outside.csv"
    (tmp_path / "outside.csv").write_text("x", encoding="utf-8")
    escaped[0]["sha256"] = sha256_file(tmp_path / "outside.csv")
    with pytest.raises(ValueError, match="越出项目目录"):
        extract_staging_tables(root, escaped, adapter_options=options)

    invalid_status = [dict(row) for row in rows]
    invalid_status[0]["status"] = "not_declared"
    with pytest.raises(ValueError, match="source_status"):
        extract_staging_tables(root, invalid_status, adapter_options=options)


def test_adapter_options_reject_unknown_or_ambiguous_values(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    with pytest.raises(ValueError, match="未知来源"):
        build_database(root, rows, adapter_options={"ds_unknown": {}})
    with pytest.raises(ValueError, match="未知参数"):
        build_database(
            root,
            rows,
            adapter_options={"ds_eom_hbond_2021": {"guess_units": True}},
        )
    with pytest.raises(ValueError, match="不支持"):
        build_database(
            root,
            rows,
            adapter_options={
                "ds_eom_hbond_2021": {"expected_sheets": {"Figure 1b"}}
            },
        )


def test_public_policy_blocks_withdrawn_and_non_open_sources(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    staging, selected = extract_staging_tables(root, rows, adapter_options=options)
    for field, value in (
        ("status", "withdrawn"),
        ("access_restriction", "author_request"),
    ):
        changed = selected.copy()
        changed.loc[
            changed["source_id"].eq("ds_eom_hbond_2021"), field
        ] = value
        normalized = normalize_staging_tables(staging, changed)
        assert _public_view(normalized["curve"])["source_id"].ne(
            "ds_eom_hbond_2021"
        ).all()


def test_mixed_chemical_provenance_is_fail_closed(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    staging, selected = extract_staging_tables(root, rows, adapter_options=options)
    restricted_manifest = selected.iloc[[0]].copy()
    restricted_manifest["source_id"] = "ds_restricted_candidate"
    restricted_manifest["source_file_id"] = "file_restricted_candidate"
    restricted_manifest["license_spdx"] = "UNKNOWN"
    restricted_manifest["derivatives_allowed"] = None
    restricted_manifest["redistribution_allowed"] = None
    restricted_manifest["status"] = "review_required"
    selected = pd.concat([selected, restricted_manifest], ignore_index=True)

    restricted_record = staging["smipoly_chemical"].iloc[[0]].copy()
    restricted_record["source_id"] = "ds_restricted_candidate"
    restricted_record["source_file_id"] = "file_restricted_candidate"
    restricted_record["record_id"] = "record_restricted_candidate"
    restricted_record["source_record_id"] = "restricted-record"
    restricted_record["iupac_name_raw"] = "BlockedName"
    staging = dict(staging)
    staging["smipoly_chemical"] = pd.concat(
        [staging["smipoly_chemical"], restricted_record], ignore_index=True
    )

    normalized = normalize_staging_tables(staging, selected)
    candidate = normalized["chemical_candidate"].iloc[0]
    assert candidate["license_resolution_status"] == "mixed_or_blocked"
    assert bool(candidate["may_publish"]) is False
    assert candidate["license_spdx"] == "NOASSERTION"
    assert "file_restricted_candidate" in candidate["source_file_ids_json"]
    assert _public_view(normalized["chemical_candidate"]).empty

    derived, derivation_issues = derive_curve_metrics(
        normalized["curve"], normalized["curve_point"]
    )
    public_views = {
        f"public_{name}": _public_view(frame)
        for name, frame in normalized.items()
        if "may_publish" in frame.columns
    }
    public_views["public_derived_property"] = _public_view(derived)
    issues = run_quality_checks(
        staging, normalized, derived, public_views, derivation_issues
    )
    assert "license.mixed_provenance_requires_review" in {
        issue.rule_id for issue in issues
    }


def test_same_source_id_cannot_hide_conflicting_file_policy(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    staging, selected = extract_staging_tables(root, rows, adapter_options=options)
    conflicting = selected.iloc[[0]].copy()
    conflicting["source_file_id"] = "file_conflicting"
    conflicting["license_spdx"] = "UNKNOWN"
    conflicting["derivatives_allowed"] = None
    conflicting["redistribution_allowed"] = None
    with pytest.raises(ValueError, match="许可策略必须一致"):
        normalize_staging_tables(
            staging, pd.concat([selected, conflicting], ignore_index=True)
        )


def test_derivation_and_quality_checks_report_structured_failures(tmp_path: Path):
    root, rows, options = _project(tmp_path)
    staging, selected = extract_staging_tables(root, rows, adapter_options=options)
    normalized = normalize_staging_tables(staging, selected)

    curve = normalized["curve"].copy()
    points = normalized["curve_point"].copy()
    tensile_id = curve.loc[curve["curve_type"] == "tensile_monotonic", "curve_id"].iloc[0]
    only_one = points[
        (points["curve_id"] != tensile_id)
        | ((points["curve_id"] == tensile_id) & (points["point_index"] == 0))
    ]
    derived, derivation_issues = derive_curve_metrics(curve, only_one)
    assert "curve.insufficient_points" in {
        issue.rule_id for issue in derivation_issues
    }

    pue = normalized["pue_transformed_auxiliary"].copy()
    duplicated = pd.concat([pue, pue.iloc[[0]].copy()], ignore_index=True)
    duplicated.loc[len(duplicated) - 1, "record_id"] = "duplicate_variant"
    duplicated.loc[len(duplicated) - 1, "split_group"] = "wrong_split"
    tampered = dict(normalized)
    tampered["pue_transformed_auxiliary"] = duplicated
    tampered["curve_point"] = only_one
    blocked_public = {"public_curve": curve.iloc[[0]].assign(may_publish=False)}
    issues = run_quality_checks(
        staging,
        tampered,
        derived,
        blocked_public,
        derivation_issues,
    )
    rule_ids = {issue.rule_id for issue in issues}
    assert {
        "curve.insufficient_points",
        "curve.point_count_mismatch",
        "leakage.lineage_cross_split",
        "license.public_release_blocked",
        "derived.coverage",
    } <= rule_ids

    result = DatabaseBuildResult(
        snapshot_id="snapshot_test",
        row_counts={},
        outputs={},
        issues=(QualityIssue("x", "error", "t", "r", "m"),),
    )
    assert result.has_errors is True
    assert result.to_dict()["issues"][0]["rule_id"] == "x"
