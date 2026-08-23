from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from adapter_hbond import EXPECTED_SHEETS as HBOND_EXPECTED_SHEETS
from adapter_hbond import adapt_hbond
from adapter_pue326 import LINEAGE_FAMILY, adapt_pue326
from adapter_smipoly import adapt_smipoly
from adapter_viscosity import adapt_viscosity


FIXTURES = Path(__file__).parent / "夹具"
PROVENANCE = {
    "record_id",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
}


def _write_workbook(tmp_path: Path, section: str) -> Path:
    spec = json.loads((FIXTURES / "xlsx_最小结构.json").read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in spec[section].items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    path = tmp_path / f"{section}.xlsx"
    workbook.save(path)
    return path


def _write_full_hbond_workbook(tmp_path: Path) -> Path:
    path = _write_workbook(tmp_path, "hbond")
    workbook = load_workbook(path)
    del workbook["Mystery"]

    sheet = workbook.create_sheet("Figure 1c")
    values = {
        "A1": "Figure 1c",
        "B2": "Elongation at break (%)",
        "B8": "Tensile strength (MPa)",
        "B14": "Toughness MJ m-3",
        "C3": "C-IP-SS",
        "D3": "E-IP-SS",
        "E3": "Es-MD",
        "B4": "Virgin",
        "B5": "Healed",
    }
    for cell, value in values.items():
        sheet[cell] = value
    for row, row_values in {
        4: (480, 923, 880),
        5: (397, 919, "-"),
        10: (42.88, 6.76, 35.8),
        11: (33.09, 5.96, "-"),
        16: (75.054, 26.93, 115),
        17: (48.348, 20.75, "-"),
    }.items():
        for column, value in enumerate(row_values, start=3):
            sheet.cell(row, column, value)

    sheet = workbook.create_sheet("Figure 4b")
    sheet["A1"] = "Figure 4b"
    sheet["C2"] = "% of streching (C-IP-SS)"
    sheet["B3"] = "Wavenumber cm-1"
    for column, stretching in enumerate((0, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4), start=3):
        sheet.cell(3, column, stretching)
        sheet.cell(4, column, float(column))
    sheet["B4"] = 1609.306

    sheet = workbook.create_sheet("Figure 5c")
    sheet["A1"] = "Figure 5c"
    temperatures = (25, 35, 45, 55, 65, 75, 85)
    for material, header_row, data_row in (("E-IP-SS", 2, 4), ("C-IP-SS", 43, 45)):
        for index, column in enumerate(range(2, 23, 3)):
            spacing = "  " if material == "E-IP-SS" and temperatures[index] == 35 else " "
            sheet.cell(header_row, column, f"{material} {temperatures[index]}{spacing}oC")
            sheet.cell(header_row + 1, column, "Angular Frequency")
            sheet.cell(header_row + 1, column + 1, "Storage\n Modulus")
            sheet.cell(header_row + 1, column + 2, "Loss \nModulus")
            sheet.cell(data_row, column, 0.05)
            sheet.cell(data_row, column + 1, 1000.0 + column)
            sheet.cell(data_row, column + 2, 500.0 + column)

    sheet = workbook.create_sheet("Figure 5e")
    sheet["A1"] = "Figure 5e"
    sheet["D2"] = "Flow relaxation time tf, sec"
    sheet["D7"] = "Segmental relaxation time ts, sec"
    for row, material, value in (
        (3, "C-IP-SS", 74_000_000),
        (4, "E-IP-SS", 112),
        (8, "C-IP-SS", 8.52),
        (9, "E-IP-SS", 7.35),
    ):
        sheet.cell(row, 2, material)
        sheet.cell(row, 4, value)

    sheet = workbook.create_sheet("Figure 6c")
    sheet["A1"] = "Figure 6c"
    for column, temperature in zip(
        range(2, 19, 2), (-30, -20, -10, 0, 10, 20, 30, 35, 40), strict=True
    ):
        sheet.cell(2, column, f"{temperature} oC")
        sheet.cell(3, column, 100.0)
        sheet.cell(3, column + 1, 5.0)

    for sheet_name in HBOND_EXPECTED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            sheet["A1"] = sheet_name
    workbook.save(path)
    return path


def _assert_provenance(frame: pd.DataFrame) -> None:
    assert PROVENANCE <= set(frame.columns)
    assert not frame[list(PROVENANCE)].isna().any().any()
    assert frame["record_id"].is_unique
    assert set(frame["schema_version"]) == {"v0.1"}


def test_smipoly_preserves_raw_identity_without_inventing_roles():
    result = adapt_smipoly(
        FIXTURES / "smipoly_最小.csv",
        source_id="ds_smipoly_2023",
        source_file_id="source_file_fixture",
    )

    _assert_provenance(result)
    assert set(result["role_status"]) == {"unclassified"}
    assert result["tpu_role"].isna().all()
    assert result["functionality"].isna().all()
    assert result["raw_smiles"].tolist() == ["C(C(C(CS)O)O)S"] * 2
    assert result["duplicate_group"].nunique() == 1
    assert result["duplicate_count"].tolist() == [2, 2]
    assert set(result["normalization_status"]) == {"raw_smiles_unvalidated"}


def test_smipoly_header_change_fails_fast(tmp_path: Path):
    frame = pd.read_csv(FIXTURES / "smipoly_最小.csv")
    frame = frame.rename(columns={"SMILES": "smiles"})
    path = tmp_path / "bad.csv"
    frame.to_csv(path, index=False)

    with pytest.raises(ValueError, match="header fingerprint"):
        adapt_smipoly(path, source_id="source", source_file_id="file")


def test_pue_lineage_and_split_are_stable_across_source_variants(tmp_path: Path):
    original = adapt_pue326(
        FIXTURES / "pue326_最小.csv",
        source_id="ds_pue_original",
        source_file_id="source_file_original",
    )
    variant_frame = pd.read_csv(FIXTURES / "pue326_最小.csv")
    variant_frame.loc[0, "logTS"] = -999.0
    variant_path = tmp_path / "variant.csv"
    variant_frame.to_csv(variant_path, index=False)
    variant = adapt_pue326(
        variant_path,
        source_id="ds_pue_missing_variant",
        source_file_id="source_file_variant",
    )

    _assert_provenance(original)
    _assert_provenance(variant)
    assert set(original["lineage_family"]) == {LINEAGE_FAMILY}
    assert original.set_index("SSID")["lineage_record_id"].to_dict() == variant.set_index(
        "SSID"
    )["lineage_record_id"].to_dict()
    assert original.set_index("SSID")["split_group"].to_dict() == variant.set_index(
        "SSID"
    )["split_group"].to_dict()
    assert set(original["record_id"]).isdisjoint(set(variant["record_id"]))


def test_pue_header_change_and_duplicate_ssid_fail_fast(tmp_path: Path):
    frame = pd.read_csv(FIXTURES / "pue326_最小.csv")
    missing = tmp_path / "missing.csv"
    frame.drop(columns=["logTS"]).to_csv(missing, index=False)
    with pytest.raises(ValueError, match="header fingerprint"):
        adapt_pue326(missing, source_id="source", source_file_id="file")

    duplicate = tmp_path / "duplicate.csv"
    pd.concat([frame.iloc[[0]], frame.iloc[[0]]], ignore_index=True).to_csv(
        duplicate, index=False
    )
    with pytest.raises(ValueError, match="SSID"):
        adapt_pue326(duplicate, source_id="source", source_file_id="file")


def test_hbond_explicit_curve_mapping_and_unknown_sheet_audit(tmp_path: Path):
    path = _write_workbook(tmp_path, "hbond")
    result = adapt_hbond(
        path,
        source_id="ds_eom_2021",
        source_file_id="source_file_hbond",
        expected_sheets=("Figure 1b",),
    )

    _assert_provenance(result["curves"])
    _assert_provenance(result["curve_points"])
    assert len(result["curves"]) == 7
    assert len(result["curve_points"]) == 14
    assert result["curves"]["curve_type"].eq("tensile_stress_strain").all()
    assert result["curve_points"]["x_value"].max() == pytest.approx(1.0)
    assert result["curve_points"]["y_unit"].eq("MPa").all()
    assert result["audit"].loc[0, "sheet_name"] == "Mystery"
    assert result["audit"].loc[0, "status"] == "unrecognized_sheet"


def test_hbond_known_header_change_fails_fast(tmp_path: Path):
    path = _write_workbook(tmp_path, "hbond")
    workbook = load_workbook(path)
    workbook["Figure 1b"]["B2"] = "renamed"
    workbook.save(path)

    with pytest.raises(ValueError, match="header fingerprint"):
        adapt_hbond(
            path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("Figure 1b",),
        )


def test_hbond_all_explicit_mappings_and_deferred_sheets_are_accounted_for(
    tmp_path: Path,
):
    result = adapt_hbond(
        _write_full_hbond_workbook(tmp_path),
        source_id="ds_eom_2021",
        source_file_id="source_file_hbond_full",
    )

    for table in ("curves", "curve_points", "properties", "audit"):
        _assert_provenance(result[table])
    assert len(result["curves"]) == 53
    assert len(result["curve_points"]) == 60
    assert len(result["properties"]) == 19
    assert len(result["audit"]) == 6
    assert set(result["audit"]["status"]) == {"recognized_unmapped_sheet"}
    unresolved = result["curves"].query("figure == 'Figure 5c'")
    assert unresolved["unit_status"].eq("unresolved").all()
    assert result["properties"]["normalized_value"].notna().all()


def test_viscosity_normalizes_units_and_audits_unknown_sheet(tmp_path: Path):
    path = _write_workbook(tmp_path, "viscosity")
    result = adapt_viscosity(
        path,
        source_id="ds_prepolymer_viscosity",
        source_file_id="source_file_viscosity",
        expected_sheets=("P_44M_4",),
    )

    _assert_provenance(result["curves"])
    _assert_provenance(result["curve_points"])
    assert result["curves"].loc[0, "curve_type"] == "viscosity_temperature"
    assert result["curves"].loc[0, "polyol_code"] == "P"
    assert result["curves"].loc[0, "isocyanate_code"] == "44M"
    assert result["curves"].loc[0, "p_nco_percent"] == 4
    points = result["curve_points"]
    assert points.loc[0, "temperature_k"] == pytest.approx(313.15)
    assert points.loc[0, "viscosity_pa_s"] == pytest.approx(10.0)
    assert points.loc[0, "temperature_unit_raw"] == "degC"
    assert points.loc[0, "viscosity_unit_raw"] == "Pa.s"
    assert result["audit"].loc[0, "status"] == "unrecognized_nonempty_sheet"


def test_viscosity_header_change_and_missing_provenance_fail_fast(tmp_path: Path):
    path = _write_workbook(tmp_path, "viscosity")
    workbook = load_workbook(path)
    workbook["P_44M_4"]["B1"] = "Temp"
    workbook.save(path)
    with pytest.raises(ValueError, match="header fingerprint"):
        adapt_viscosity(
            path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("P_44M_4",),
        )

    with pytest.raises(ValueError, match="source_id"):
        adapt_smipoly(
            FIXTURES / "smipoly_最小.csv",
            source_id="",
            source_file_id="file",
        )


def test_csv_adapters_reject_invalid_files_and_structural_values(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        adapt_smipoly(
            tmp_path / "missing.csv", source_id="source", source_file_id="file"
        )
    wrong_suffix = tmp_path / "source.txt"
    wrong_suffix.write_text("not csv", encoding="utf-8")
    with pytest.raises(ValueError, match="must be CSV"):
        adapt_smipoly(wrong_suffix, source_id="source", source_file_id="file")

    smipoly = pd.read_csv(FIXTURES / "smipoly_最小.csv")
    smipoly_cases = (
        (smipoly.iloc[0:0], "no records"),
        (smipoly.assign(SMILES=""), "blank"),
        (smipoly.assign(comID="same"), "unique"),
        (smipoly.assign(MolecularWeight="bad"), "finite"),
        (smipoly.assign(MolecularWeight=0), "positive"),
    )
    for index, (frame, message) in enumerate(smipoly_cases):
        path = tmp_path / f"bad_smipoly_{index}.csv"
        frame.to_csv(path, index=False)
        with pytest.raises(ValueError, match=message):
            adapt_smipoly(path, source_id="source", source_file_id="file")

    with pytest.raises(ValueError, match="source_file_id"):
        adapt_pue326(
            FIXTURES / "pue326_最小.csv", source_id="source", source_file_id=""
        )
    with pytest.raises(FileNotFoundError):
        adapt_pue326(
            tmp_path / "missing_pue.csv", source_id="source", source_file_id="file"
        )
    with pytest.raises(ValueError, match="must be CSV"):
        adapt_pue326(wrong_suffix, source_id="source", source_file_id="file")

    pue = pd.read_csv(FIXTURES / "pue326_最小.csv")
    pue_cases = (
        (pue.iloc[0:0], "no records"),
        (pue.assign(SSID=""), "blank"),
        (pue.assign(ZS_CHS="bad"), "finite numeric"),
    )
    for index, (frame, message) in enumerate(pue_cases):
        path = tmp_path / f"bad_pue_{index}.csv"
        frame.to_csv(path, index=False)
        with pytest.raises(ValueError, match=message):
            adapt_pue326(path, source_id="source", source_file_id="file")


def test_xlsx_adapters_reject_missing_sheets_and_bad_curve_rows(tmp_path: Path):
    viscosity_path = _write_workbook(tmp_path, "viscosity")
    with pytest.raises(ValueError, match="expected_sheets"):
        adapt_viscosity(
            viscosity_path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("P_44M_4", "P_44M_4"),
        )
    with pytest.raises(ValueError, match="missing"):
        adapt_viscosity(
            viscosity_path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("P_MISSING_4",),
        )

    workbook = load_workbook(viscosity_path)
    workbook["P_44M_4"]["A3"] = -1
    workbook.save(viscosity_path)
    with pytest.raises(ValueError, match="positive"):
        adapt_viscosity(
            viscosity_path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("P_44M_4",),
        )

    hbond_path = _write_workbook(tmp_path, "hbond")
    with pytest.raises(ValueError, match="missing"):
        adapt_hbond(
            hbond_path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("Figure missing",),
        )
    workbook = load_workbook(hbond_path)
    workbook["Figure 1b"]["C4"] = None
    workbook.save(hbond_path)
    with pytest.raises(ValueError, match="Incomplete curve pair"):
        adapt_hbond(
            hbond_path,
            source_id="source",
            source_file_id="file",
            expected_sheets=("Figure 1b",),
        )
