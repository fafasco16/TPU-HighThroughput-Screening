import base64
import csv
import hashlib
import importlib.util
import io
import json
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RAW = ROOT / "01_原始数据" / "外部数据" / "新增开放数据"


def _json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def _tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def _load_script(relative_path: str):
    path = ROOT / relative_path
    module_name = f"tpu_regression_{path.stem}_{abs(hash(path))}"
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def test_commercial_tpu_identity_conflict_is_held_on_both_labels() -> None:
    directory = RAW / "Zenodo_商业TPU多材料打印传感"
    if not directory.is_dir():
        pytest.skip("原始开放数据未在当前检出中分发")
    rows = _tsv(directory / "曲线审计清单.tsv")
    held = [
        row
        for row in rows
        if row["file"] == "Figure_S7.csv" and row["material"] in {"70A", "85A"}
    ]
    assert {row["material"] for row in held} == {"70A", "85A"}
    assert all(row["decision"] == "hold_material_identity_conflict" for row in held)
    assert all(int(row["usable_points"]) == 0 for row in held)
    assert all(float(row["future_weight_ceiling"]) == 0.0 for row in held)
    assert all(
        row["split_group"] == f"10.5281/zenodo.5841610|{row['material']}"
        for row in rows
    )


def test_tecoflex_does_not_inflate_cross_workbook_specimen_count() -> None:
    path = RAW / "Zenodo_Tecoflex药物复合TPU" / "内容审计摘要.json"
    if not path.is_file():
        pytest.skip("原始开放数据未在当前检出中分发")
    summary = _json(path)
    assert summary["physical_specimen_count"] is None
    assert summary["main_workbook_direct_specimen_slot_count"] == 22
    assert summary["main_workbook_direct_observation_count"] == 66
    assert summary["main_workbook_direct_qois"] == [
        "elastic_modulus_mpa",
        "stress_at_100_percent_mpa",
        "diameter_mm",
    ]


def test_fdm_curves_have_unique_specimen_groups_and_standardized_role() -> None:
    directory = RAW / "Mendeley_FDM_TPU晶格与基材力学"
    if not directory.is_dir():
        pytest.skip("原始开放数据未在当前检出中分发")
    summary = _json(directory / "内容审计摘要.json")
    rows = _tsv(directory / "曲线审计清单.tsv")
    assert summary["审计版本"] == "1.2"
    assert len(rows) == summary["TPU实验试样曲线数"] == 76
    assert len({row["试样或家族组"] for row in rows}) == 76
    assert {row["数据角色"] for row in rows} == {"实验工作簿内标准化应力应变曲线"}
    assert all(row["training_split"] == "false" and row["weight"] == "false" for row in rows)


def test_fdm_lattice_formula_and_cached_values_are_recomputed() -> None:
    directory = RAW / "Mendeley_FDM_TPU晶格与基材力学"
    archive_path = directory / "dbzdkz95f8-1.zip"
    if not archive_path.is_file():
        pytest.skip("FDM 原始归档未在当前检出中分发")
    module = _load_script("代码/审计/新增开放数据第三批Mendeley三源.py")
    member = (
        "Experimental and numerical data for FDM-printed PL/"
        "Bending/TPU/EF ARA TPU.xlsx"
    )
    with zipfile.ZipFile(archive_path) as archive:
        payload = archive.read(member)
    cached = load_workbook(io.BytesIO(payload), read_only=True, data_only=True, keep_links=False)
    formulas = load_workbook(io.BytesIO(payload), read_only=True, data_only=False, keep_links=False)
    try:
        points, _ = module.parse_pair_sheet(
            cached["EF A1"], x_column=3, y_column=4, x_scale=0.01
        )
        note = module.validate_fdm_curve_lineage(
            formulas["EF A1"],
            cached["EF A1"],
            "晶格弯曲",
            points,
            cached_results=cached["Results"],
        )
    finally:
        cached.close()
        formulas.close()
    assert points == 25_624
    assert "D=B/(Results!I*Results!J)" in note


def test_pcl_real_trajectory_payloads_and_lfs_pointers_are_separated() -> None:
    path = RAW / "Zenodo_PCL软段构象粗粒化MD" / "内容审计摘要.json"
    if not path.is_file():
        pytest.skip("PCL 原始开放数据未在当前检出中分发")
    summary = _json(path)
    trajectories = summary["真实TRR轨迹"]
    pointers = summary["Git_LFS"]
    assert trajectories["总帧数"] == 6_396
    assert trajectories["解压TRR总字节"] == 149_898_708
    assert trajectories["正常完成运行数"] == 5
    assert trajectories["TERM终止运行数"] == 3
    assert len(trajectories["运行"]) == 8
    assert pointers["指针文件数"] == 10
    assert pointers["声明载荷总字节"] == 2_313_207_356
    tau_run = next(run for run in trajectories["运行"] if "/tau0.5/" in run["path"])
    assert tau_run["dt_fs"] == pytest.approx(0.5)
    assert tau_run["tau_t_ps"] == pytest.approx(0.1)
    assert any(run["ref_t_K"] == pytest.approx(600.0) for run in trajectories["运行"])


def test_impact_piston_velocity_uses_angstrom_per_femtosecond_conversion() -> None:
    path = RAW / "Zenodo_PTMO_MDI_BDO聚氨酯冲击MD" / "内容审计摘要.json"
    if not path.is_file():
        pytest.skip("冲击 MD 原始开放数据未在当前检出中分发")
    recipe = _json(path)["单一运行配方"]
    assert recipe["piston_A_per_fs"] == pytest.approx(0.02)
    assert recipe["piston_km_s"] == pytest.approx(2.0)


def test_pcl_lfs_supplement_counts_runs_not_frames_as_independent_units() -> None:
    path = RAW / "PCL_GitLFS轨迹补采" / "内容审计摘要.json"
    if not path.is_file():
        pytest.skip("PCL Git LFS 补采未在当前检出中分发")
    summary = _json(path)
    assert summary["对象数"] == summary["独立模拟运行家族上限"] == 10
    assert summary["压缩总字节"] == 2_313_207_356
    assert summary["总帧数"] == 10_569
    assert summary["解压TRR总字节"] == 2_578_712_040
    assert summary["完成状态计数"] == {
        "finished": 7,
        "terminated_after_continuation_beyond_declared_nsteps": 1,
        "terminated_before_declared_nsteps": 1,
        "terminated_by_second_int_term_signal": 1,
    }
    assert summary["训练许可"] is False and summary["训练权重"] is None
    second_signal = next(
        row
        for row in summary["轨迹"]
        if row["完成状态"] == "terminated_by_second_int_term_signal"
    )
    assert second_signal["日志终止信号"] == "second_int_term_signal"
    assert second_signal["末步"] <= second_signal["日志终止步"]


def test_official_api_captures_retain_exact_bytes_and_sha256() -> None:
    directories = [
        "Mendeley_商业TPU温度疲劳多工况",
        "Mendeley_FDM_TPU晶格与基材力学",
        "Mendeley_TPU实验仿真曲线",
        "Zenodo_反应型粗粒化聚脲固化",
        "Zenodo_NIPU反应路径DFT与MD",
        "Zenodo_PCL软段构象粗粒化MD",
        "Zenodo_PTMO_MDI_BDO聚氨酯冲击MD",
        "Mendeley_TPU压缩打印DOE",
        "ACS_Figshare_TPU退火硬段聚集",
        "ACS_Figshare_双相演化聚氨酯",
        "ACS_Figshare_PLA立构复合TPU",
        "ACS_Figshare_呋喃高强聚氨酯",
        "ACS_Figshare_聚酰亚胺回收链扩剂PU",
        "ACS_Figshare_二氧化碳共聚酯聚氨酯",
        "ACS_Figshare_聚碳酸酯大分子二醇TPU",
        "ACS_Figshare_氢键纳米结构TPU",
    ]
    present = 0
    for directory in directories:
        path = RAW / directory / "官方API元数据.json"
        if not path.is_file():
            continue
        present += 1
        metadata = _json(path)
        assert metadata["raw_api_capture_format"] == "exact_response_bytes_base64_with_sha256"
        captures = metadata["raw_api_captures"]
        assert captures
        assert len({capture["request_url"] for capture in captures}) == len(captures)
        for capture in captures:
            payload = base64.b64decode(capture["payload_base64"], validate=True)
            assert len(payload) == capture["payload_bytes"]
            assert hashlib.sha256(payload).hexdigest() == capture["payload_sha256"]
            json.loads(payload.decode("utf-8"))
            assert capture["status"] == 200
    if present == 0:
        pytest.skip("官方API捕获元数据未在当前检出中分发")
    assert present == len(directories)


def test_simulation_downloader_enforces_exact_range_headers_and_redirects() -> None:
    module = _load_script("代码/获取/下载第三批模拟四源.py")
    good_headers = {
        "Content-Length": "900",
        "Content-Range": "bytes 100-999/1000",
        "ETag": '"frozen"',
    }
    assert module.validate_download_response_headers(
        status=206,
        headers=good_headers,
        offset=100,
        expected_size=1_000,
        url="https://zenodo.org/file",
    ) == '"frozen"'
    for headers in (
        {**good_headers, "Content-Range": "bytes 99-999/1000"},
        {**good_headers, "Content-Range": "bytes 100-999/1001"},
        {**good_headers, "Content-Length": "899"},
    ):
        with pytest.raises(module.AcquisitionBlocked):
            module.validate_download_response_headers(
                status=206,
                headers=headers,
                offset=100,
                expected_size=1_000,
                url="https://zenodo.org/file",
            )

    handler = module.WhitelistRedirectHandler()
    request = module.Request("https://zenodo.org/api/records/1")
    with pytest.raises(module.AcquisitionBlocked, match="重定向跨越固定主机"):
        handler.redirect_request(
            request,
            None,
            302,
            "Found",
            {},
            "https://raw.githubusercontent.com/other/file",
        )


def test_simulation_downloader_clears_oversize_dedicated_part(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    module = _load_script("代码/获取/下载第三批模拟四源.py")
    target = tmp_path / "payload.bin"
    partial = tmp_path / "payload.bin.part"
    partial.write_bytes(b"x" * 11)

    def stop_before_network(*_args, **_kwargs):
        raise module.AcquisitionBlocked("injected stop")

    monkeypatch.setattr(module, "open_request", stop_before_network)
    with pytest.raises(module.AcquisitionBlocked, match="injected stop"):
        module.download_file(
            "https://zenodo.org/file",
            target,
            expected_size=10,
            expected_md5=None,
            expected_sha256=None,
        )
    assert not partial.exists()
