"""Eom 2021 TPU 氢键主文 Source Data 工作簿适配器。"""

from __future__ import annotations

import math
from collections.abc import Callable, Iterable
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ids import stable_id
from units import strain_to_fraction, stress_to_mpa, temperature_to_k, time_to_seconds


SCHEMA_VERSION = "v0.1"
EXPECTED_SHEETS = (
    "Figure 1b",
    "Figure 1c",
    "Figure 3a",
    "Figure 4a",
    "Figure 4b",
    "Figure 5a",
    "Figure 5b",
    "Figure 5c",
    "Figure 5e",
    "Figure 6b",
    "Figure 6c",
    "Figure 6d",
)
MAPPED_SHEETS = {
    "Figure 1b",
    "Figure 1c",
    "Figure 4b",
    "Figure 5c",
    "Figure 5e",
    "Figure 6c",
}

CURVE_COLUMNS = (
    "record_id",
    "curve_id",
    "test_id",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "figure",
    "curve_type",
    "material_name",
    "material_status",
    "condition_label",
    "temperature_raw_deg_c",
    "temperature_k",
    "x_quantity",
    "x_unit_raw",
    "x_unit",
    "y_quantity",
    "y_unit_raw",
    "y_unit",
    "unit_status",
    "point_count",
)
POINT_COLUMNS = (
    "record_id",
    "curve_id",
    "point_index",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "x_raw",
    "x_unit_raw",
    "x_value",
    "x_unit",
    "y_raw",
    "y_unit_raw",
    "y_value",
    "y_unit",
    "unit_status",
)
PROPERTY_COLUMNS = (
    "record_id",
    "property_id",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "figure",
    "material_name",
    "condition_label",
    "property_name",
    "raw_value",
    "raw_unit",
    "normalized_value",
    "normalized_unit",
    "unit_status",
)
AUDIT_COLUMNS = (
    "record_id",
    "source_id",
    "source_file_id",
    "source_locator",
    "extraction_method",
    "fidelity",
    "schema_version",
    "sheet_name",
    "status",
    "reason",
    "nonempty_cell_count",
)


def _require_provenance(source_id: str, source_file_id: str) -> None:
    for name, value in (("source_id", source_id), ("source_file_id", source_file_id)):
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"{name} must be a non-empty string")


def _number(value: Any, locator: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"Expected numeric value at {locator}, got boolean")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Expected numeric value at {locator}, got {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"Expected finite value at {locator}, got {value!r}")
    return number


def _require_cells(sheet: Any, expected: dict[str, Any]) -> None:
    mismatches = {
        cell: (expected_value, sheet[cell].value)
        for cell, expected_value in expected.items()
        if sheet[cell].value != expected_value
    }
    if mismatches:
        raise ValueError(
            f"HBond header fingerprint mismatch in {sheet.title}: {mismatches!r}"
        )


def _pair_points(sheet: Any, start_row: int, x_column: int, y_column: int) -> list[tuple[int, float, float]]:
    points: list[tuple[int, float, float]] = []
    x_letter = get_column_letter(x_column)
    y_letter = get_column_letter(y_column)
    for row_number in range(start_row, sheet.max_row + 1):
        x_raw = sheet.cell(row_number, x_column).value
        y_raw = sheet.cell(row_number, y_column).value
        if x_raw is None and y_raw is None:
            continue
        if x_raw is None or y_raw is None:
            raise ValueError(
                f"Incomplete curve pair at {sheet.title}!{x_letter}{row_number}:{y_letter}{row_number}"
            )
        points.append(
            (
                row_number,
                _number(x_raw, f"{sheet.title}!{x_letter}{row_number}"),
                _number(y_raw, f"{sheet.title}!{y_letter}{row_number}"),
            )
        )
    if not points:
        raise ValueError(
            f"Mapped curve contains no points: {sheet.title}!{x_letter}:{y_letter}"
        )
    return points


def _common(source_id: str, source_file_id: str, fidelity: str) -> dict[str, str]:
    return {
        "source_id": source_id,
        "source_file_id": source_file_id,
        "extraction_method": "openpyxl:explicit_sheet_and_cell_map",
        "fidelity": fidelity,
        "schema_version": SCHEMA_VERSION,
    }


def _append_curve(
    curves: list[dict[str, Any]],
    points: list[dict[str, Any]],
    *,
    source_id: str,
    source_file_id: str,
    figure: str,
    curve_type: str,
    material_name: str | None,
    condition_label: str,
    x_column: int,
    y_column: int,
    raw_points: list[tuple[int, float, float]],
    x_quantity: str,
    x_unit_raw: str | None,
    x_unit: str | None,
    y_quantity: str,
    y_unit_raw: str | None,
    y_unit: str | None,
    x_converter: Callable[[float], float] | None,
    y_converter: Callable[[float], float] | None,
    temperature_raw_deg_c: float | None = None,
) -> None:
    curve_id = stable_id(
        "curve", source_file_id, figure, curve_type, material_name, condition_label
    )
    test_id = stable_id(
        "test", source_file_id, figure, material_name, condition_label, curve_type
    )
    x_letter = get_column_letter(x_column)
    y_letter = get_column_letter(y_column)
    unit_status = "resolved" if x_unit is not None and y_unit is not None else "unresolved"
    common = _common(source_id, source_file_id, "measured_raw")
    curves.append(
        {
            "record_id": curve_id,
            "curve_id": curve_id,
            "test_id": test_id,
            **common,
            "source_locator": f"sheet={figure};columns={x_letter}:{y_letter}",
            "figure": figure,
            "curve_type": curve_type,
            "material_name": material_name,
            "material_status": "declared" if material_name else "not_declared_in_sheet",
            "condition_label": condition_label,
            "temperature_raw_deg_c": temperature_raw_deg_c,
            "temperature_k": (
                temperature_to_k(temperature_raw_deg_c, "degC")
                if temperature_raw_deg_c is not None
                else None
            ),
            "x_quantity": x_quantity,
            "x_unit_raw": x_unit_raw,
            "x_unit": x_unit,
            "y_quantity": y_quantity,
            "y_unit_raw": y_unit_raw,
            "y_unit": y_unit,
            "unit_status": unit_status,
            "point_count": len(raw_points),
        }
    )
    for point_index, (row_number, x_raw, y_raw) in enumerate(raw_points):
        points.append(
            {
                "record_id": stable_id("curve_point", curve_id, point_index),
                "curve_id": curve_id,
                "point_index": point_index,
                **common,
                "source_locator": (
                    f"sheet={figure};cells={x_letter}{row_number}:{y_letter}{row_number}"
                ),
                "x_raw": x_raw,
                "x_unit_raw": x_unit_raw,
                "x_value": x_converter(x_raw) if x_converter else None,
                "x_unit": x_unit,
                "y_raw": y_raw,
                "y_unit_raw": y_unit_raw,
                "y_value": y_converter(y_raw) if y_converter else None,
                "y_unit": y_unit,
                "unit_status": unit_status,
            }
        )


def _property_row(
    *,
    source_id: str,
    source_file_id: str,
    figure: str,
    cell: str,
    material: str,
    condition: str,
    property_name: str,
    value: float,
    raw_unit: str,
    normalized_unit: str,
    converter: Callable[[float], float],
) -> dict[str, Any]:
    property_id = stable_id(
        "property", source_file_id, figure, cell, material, condition, property_name
    )
    return {
        "record_id": property_id,
        "property_id": property_id,
        **_common(source_id, source_file_id, "measured_summary"),
        "source_locator": f"sheet={figure};cell={cell}",
        "figure": figure,
        "material_name": material,
        "condition_label": condition,
        "property_name": property_name,
        "raw_value": value,
        "raw_unit": raw_unit,
        "normalized_value": converter(value),
        "normalized_unit": normalized_unit,
        "unit_status": "resolved",
    }


def _parse_figure_1b(sheet: Any, curves: list[dict[str, Any]], points: list[dict[str, Any]], source_id: str, source_file_id: str) -> None:
    _require_cells(
        sheet,
        {
            "A1": "Figure 1b",
            "F1": "C-IP-SS",
            "B2": "E-IP-SS",
            "D2": "Es-MD",
            "F2": "Virgin",
            "H2": "1 h",
            "J2": "6 h",
            "L2": "24 h",
            "N2": "48 h",
        },
    )
    mappings = (
        (2, 3, "E-IP-SS", "Virgin"),
        (4, 5, "Es-MD", "Virgin"),
        (6, 7, "C-IP-SS", "Virgin"),
        (8, 9, "C-IP-SS", "1 h"),
        (10, 11, "C-IP-SS", "6 h"),
        (12, 13, "C-IP-SS", "24 h"),
        (14, 15, "C-IP-SS", "48 h"),
    )
    for x_column, y_column, material, condition in mappings:
        _append_curve(
            curves,
            points,
            source_id=source_id,
            source_file_id=source_file_id,
            figure="Figure 1b",
            curve_type="tensile_stress_strain",
            material_name=material,
            condition_label=condition,
            x_column=x_column,
            y_column=y_column,
            raw_points=_pair_points(sheet, 3, x_column, y_column),
            x_quantity="engineering_strain",
            x_unit_raw="%",
            x_unit="1",
            y_quantity="engineering_stress",
            y_unit_raw="MPa",
            y_unit="MPa",
            x_converter=lambda value: strain_to_fraction(value, "%"),
            y_converter=lambda value: stress_to_mpa(value, "MPa"),
        )


def _parse_figure_1c(sheet: Any, properties: list[dict[str, Any]], source_id: str, source_file_id: str) -> None:
    _require_cells(
        sheet,
        {
            "A1": "Figure 1c",
            "B2": "Elongation at break (%)",
            "B8": "Tensile strength (MPa)",
            "B14": "Toughness MJ m-3",
            "C3": "C-IP-SS",
            "D3": "E-IP-SS",
            "E3": "Es-MD",
            "B4": "Virgin",
            "B5": "Healed",
        },
    )
    blocks = (
        (4, 5, "elongation_at_break", "%", "1", lambda value: strain_to_fraction(value, "%")),
        (10, 11, "tensile_strength", "MPa", "MPa", lambda value: stress_to_mpa(value, "MPa")),
        (16, 17, "toughness", "MJ m-3", "MJ·m^-3", float),
    )
    materials = {3: "C-IP-SS", 4: "E-IP-SS", 5: "Es-MD"}
    for virgin_row, healed_row, property_name, raw_unit, normalized_unit, converter in blocks:
        for row, condition in ((virgin_row, "Virgin"), (healed_row, "Healed")):
            for column, material in materials.items():
                value = sheet.cell(row, column).value
                if value in (None, "-"):
                    continue
                cell = f"{get_column_letter(column)}{row}"
                properties.append(
                    _property_row(
                        source_id=source_id,
                        source_file_id=source_file_id,
                        figure="Figure 1c",
                        cell=cell,
                        material=material,
                        condition=condition,
                        property_name=property_name,
                        value=_number(value, f"Figure 1c!{cell}"),
                        raw_unit=raw_unit,
                        normalized_unit=normalized_unit,
                        converter=converter,
                    )
                )


def _parse_figure_4b(sheet: Any, curves: list[dict[str, Any]], points: list[dict[str, Any]], source_id: str, source_file_id: str) -> None:
    _require_cells(
        sheet,
        {
            "A1": "Figure 4b",
            "C2": "% of streching (C-IP-SS)",
            "B3": "Wavenumber cm-1",
            "C3": 0,
            "D3": 0.5,
            "E3": 1,
            "F3": 1.5,
            "G3": 2,
            "H3": 2.5,
            "I3": 3,
            "J3": 3.5,
            "K3": 4,
        },
    )
    for y_column in range(3, 12):
        stretching = sheet.cell(3, y_column).value
        _append_curve(
            curves,
            points,
            source_id=source_id,
            source_file_id=source_file_id,
            figure="Figure 4b",
            curve_type="ftir_absorbance_spectrum",
            material_name="C-IP-SS",
            condition_label=f"stretching_raw={stretching}",
            x_column=2,
            y_column=y_column,
            raw_points=_pair_points(sheet, 4, 2, y_column),
            x_quantity="wavenumber",
            x_unit_raw="cm^-1",
            x_unit="cm^-1",
            y_quantity="absorbance",
            y_unit_raw="a.u.",
            y_unit="a.u.",
            x_converter=float,
            y_converter=float,
        )


def _parse_figure_5c(sheet: Any, curves: list[dict[str, Any]], points: list[dict[str, Any]], source_id: str, source_file_id: str) -> None:
    _require_cells(
        sheet,
        {
            "A1": "Figure 5c",
            "B2": "E-IP-SS 25 oC",
            "E2": "E-IP-SS 35  oC",
            "H2": "E-IP-SS 45 oC",
            "K2": "E-IP-SS 55 oC",
            "N2": "E-IP-SS 65 oC",
            "Q2": "E-IP-SS 75 oC",
            "T2": "E-IP-SS 85 oC",
            "B43": "C-IP-SS 25 oC",
            "E43": "C-IP-SS 35 oC",
            "H43": "C-IP-SS 45 oC",
            "K43": "C-IP-SS 55 oC",
            "N43": "C-IP-SS 65 oC",
            "Q43": "C-IP-SS 75 oC",
            "T43": "C-IP-SS 85 oC",
            "B3": "Angular Frequency",
            "C3": "Storage\n Modulus",
            "D3": "Loss \nModulus",
            "B44": "Angular Frequency",
            "C44": "Storage\n Modulus",
            "D44": "Loss \nModulus",
        },
    )
    for material, header_row, start_row, end_row in (
        ("E-IP-SS", 2, 4, 40),
        ("C-IP-SS", 43, 45, 81),
    ):
        for group_index, x_column in enumerate(range(2, 23, 3)):
            temperature = float((25, 35, 45, 55, 65, 75, 85)[group_index])
            storage_column = x_column + 1
            loss_column = x_column + 2
            header_value = sheet.cell(header_row, x_column).value
            if material not in str(header_value):
                raise ValueError(
                    f"HBond header fingerprint mismatch in Figure 5c at {get_column_letter(x_column)}{header_row}"
                )
            for y_column, curve_type, y_quantity in (
                (storage_column, "rheology_frequency_sweep_storage_modulus", "storage_modulus"),
                (loss_column, "rheology_frequency_sweep_loss_modulus", "loss_modulus"),
            ):
                raw_points: list[tuple[int, float, float]] = []
                for row in range(start_row, end_row + 1):
                    x_raw = sheet.cell(row, x_column).value
                    y_raw = sheet.cell(row, y_column).value
                    if x_raw is None and y_raw is None:
                        continue
                    if x_raw is None or y_raw is None:
                        raise ValueError(
                            f"Incomplete Figure 5c triple at row {row}, columns {x_column}:{y_column}"
                        )
                    raw_points.append(
                        (
                            row,
                            _number(x_raw, f"Figure 5c!{get_column_letter(x_column)}{row}"),
                            _number(y_raw, f"Figure 5c!{get_column_letter(y_column)}{row}"),
                        )
                    )
                _append_curve(
                    curves,
                    points,
                    source_id=source_id,
                    source_file_id=source_file_id,
                    figure="Figure 5c",
                    curve_type=curve_type,
                    material_name=material,
                    condition_label=f"temperature={temperature:g} degC",
                    x_column=x_column,
                    y_column=y_column,
                    raw_points=raw_points,
                    x_quantity="angular_frequency",
                    x_unit_raw=None,
                    x_unit=None,
                    y_quantity=y_quantity,
                    y_unit_raw=None,
                    y_unit=None,
                    x_converter=None,
                    y_converter=None,
                    temperature_raw_deg_c=temperature,
                )


def _parse_figure_5e(sheet: Any, properties: list[dict[str, Any]], source_id: str, source_file_id: str) -> None:
    _require_cells(
        sheet,
        {
            "A1": "Figure 5e",
            "D2": "Flow relaxation time tf, sec",
            "D7": "Segmental relaxation time ts, sec",
        },
    )
    for row, property_name in (
        (3, "flow_relaxation_time"),
        (4, "flow_relaxation_time"),
        (8, "segmental_relaxation_time"),
        (9, "segmental_relaxation_time"),
    ):
        material = str(sheet.cell(row, 2).value)
        cell = f"D{row}"
        value = _number(sheet[cell].value, f"Figure 5e!{cell}")
        properties.append(
            _property_row(
                source_id=source_id,
                source_file_id=source_file_id,
                figure="Figure 5e",
                cell=cell,
                material=material,
                condition="reported",
                property_name=property_name,
                value=value,
                raw_unit="sec",
                normalized_unit="s",
                converter=lambda raw: time_to_seconds(raw, "sec"),
            )
        )


def _parse_figure_6c(sheet: Any, curves: list[dict[str, Any]], points: list[dict[str, Any]], source_id: str, source_file_id: str) -> None:
    temperatures = (-30, -20, -10, 0, 10, 20, 30, 35, 40)
    expected = {"A1": "Figure 6c"}
    for column, temperature in zip(range(2, 19, 2), temperatures, strict=True):
        expected[f"{get_column_letter(column)}2"] = f"{temperature} oC"
    _require_cells(sheet, expected)
    for x_column, temperature in zip(range(2, 19, 2), temperatures, strict=True):
        y_column = x_column + 1
        _append_curve(
            curves,
            points,
            source_id=source_id,
            source_file_id=source_file_id,
            figure="Figure 6c",
            curve_type="temperature_dependent_tensile_stress_strain",
            material_name=None,
            condition_label=f"temperature={temperature} degC",
            x_column=x_column,
            y_column=y_column,
            raw_points=_pair_points(sheet, 3, x_column, y_column),
            x_quantity="engineering_strain",
            x_unit_raw="%",
            x_unit="1",
            y_quantity="engineering_stress",
            y_unit_raw="MPa",
            y_unit="MPa",
            x_converter=lambda value: strain_to_fraction(value, "%"),
            y_converter=lambda value: stress_to_mpa(value, "MPa"),
            temperature_raw_deg_c=float(temperature),
        )


PARSERS = {
    "Figure 1b": lambda sheet, curves, points, properties, source_id, source_file_id: _parse_figure_1b(sheet, curves, points, source_id, source_file_id),
    "Figure 1c": lambda sheet, curves, points, properties, source_id, source_file_id: _parse_figure_1c(sheet, properties, source_id, source_file_id),
    "Figure 4b": lambda sheet, curves, points, properties, source_id, source_file_id: _parse_figure_4b(sheet, curves, points, source_id, source_file_id),
    "Figure 5c": lambda sheet, curves, points, properties, source_id, source_file_id: _parse_figure_5c(sheet, curves, points, source_id, source_file_id),
    "Figure 5e": lambda sheet, curves, points, properties, source_id, source_file_id: _parse_figure_5e(sheet, properties, source_id, source_file_id),
    "Figure 6c": lambda sheet, curves, points, properties, source_id, source_file_id: _parse_figure_6c(sheet, curves, points, source_id, source_file_id),
}


def _nonempty_cell_count(sheet: Any) -> int:
    return sum(
        value is not None
        for row in sheet.iter_rows(values_only=True)
        for value in row
    )


def _audit_row(sheet_name: str, status: str, reason: str, sheet: Any, source_id: str, source_file_id: str) -> dict[str, Any]:
    return {
        "record_id": stable_id("audit", source_file_id, sheet_name, status),
        **_common(source_id, source_file_id, "audit_metadata"),
        "source_locator": f"sheet={sheet_name}",
        "sheet_name": sheet_name,
        "status": status,
        "reason": reason,
        "nonempty_cell_count": _nonempty_cell_count(sheet),
    }


def adapt_hbond(
    path: str | Path,
    *,
    source_id: str,
    source_file_id: str,
    expected_sheets: Iterable[str] = EXPECTED_SHEETS,
) -> dict[str, pd.DataFrame]:
    """Extract explicitly mapped figures and inventory every remaining sheet."""

    _require_provenance(source_id, source_file_id)
    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError(f"HBond source must be XLSX: {source_path}")

    expected = tuple(expected_sheets)
    if len(expected) != len(set(expected)) or not expected:
        raise ValueError("expected_sheets must contain unique, non-empty sheet names")
    workbook = load_workbook(source_path, read_only=False, data_only=True)
    missing = [sheet for sheet in expected if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"HBond sheet fingerprint mismatch: missing {missing!r}")

    curves: list[dict[str, Any]] = []
    points: list[dict[str, Any]] = []
    properties: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    for sheet_name in expected:
        sheet = workbook[sheet_name]
        parser = PARSERS.get(sheet_name)
        if parser is None:
            audit.append(
                _audit_row(
                    sheet_name,
                    "recognized_unmapped_sheet",
                    "sheet is present, but v0.1 defers extraction because its axis or unit semantics are not explicit enough",
                    sheet,
                    source_id,
                    source_file_id,
                )
            )
            continue
        parser(sheet, curves, points, properties, source_id, source_file_id)

    expected_set = set(expected)
    for sheet_name in workbook.sheetnames:
        if sheet_name not in expected_set:
            audit.append(
                _audit_row(
                    sheet_name,
                    "unrecognized_sheet",
                    "sheet is outside the explicit v0.1 source mapping",
                    workbook[sheet_name],
                    source_id,
                    source_file_id,
                )
            )

    workbook.close()
    return {
        "curves": pd.DataFrame(curves, columns=CURVE_COLUMNS),
        "curve_points": pd.DataFrame(points, columns=POINT_COLUMNS),
        "properties": pd.DataFrame(properties, columns=PROPERTY_COLUMNS),
        "audit": pd.DataFrame(audit, columns=AUDIT_COLUMNS),
    }


__all__ = ["EXPECTED_SHEETS", "MAPPED_SHEETS", "SCHEMA_VERSION", "adapt_hbond"]
