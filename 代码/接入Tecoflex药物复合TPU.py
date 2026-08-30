"""物化Tecoflex EG-60D/尼可刹米复合物的拉伸下界与TGA端点。"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import statistics
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "Zenodo_Tecoflex药物复合TPU"
)
TGA_SOURCE = SOURCE_DIR / "TGA thermal analysis.xlsx"
TENSILE_SOURCE = SOURCE_DIR / "mechanical testing sup fig 1.xlsx"
MECHANICAL_SOURCE = SOURCE_DIR / "mechanical testing.xlsx"
SOURCE_METADATA = SOURCE_DIR / "官方API元数据.json"
SOURCE_AUDIT = SOURCE_DIR / "内容审计摘要.json"
DIRECTED = ROOT / "结果" / "定向筛选"
OUTPUT = DIRECTED / "Tecoflex药物复合TPU多性能端点.csv"
MANIFEST = DIRECTED / "Tecoflex药物复合TPU发布清单.json"
RELEASE_ID = "tecoflex-eg60d-nic-multiperformance-2022-v1"
DATASET_DOI = "10.5281/zenodo.6128356"
ARTICLE_DOI = "10.1038/s41598-022-16107-4"
LICENSE = "CC-BY-4.0"
CITATIONS = "reference-188;reference-189"
SPECS = (
    ("Tecoflex_EG60D", "TPU", 0.0, 12, 10, 11),
    ("Tecoflex_EG60D_NIC2", "NIC-2", 2.0, 0, 12, 13),
    ("Tecoflex_EG60D_NIC5", "NIC-5", 5.0, 4, 14, 15),
    ("Tecoflex_EG60D_NIC10", "NIC-10", 10.0, 8, 16, 17),
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _entry(path: Path) -> dict[str, object]:
    return {
        "path": path.relative_to(ROOT).as_posix(),
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def _finite(value: object) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _crossing_temperature(
    points: list[tuple[float, float]], threshold: float
) -> float:
    for (temperature0, weight0), (temperature1, weight1) in zip(
        points, points[1:], strict=False
    ):
        if weight0 > threshold >= weight1 and weight0 != weight1:
            return temperature0 + (threshold - weight0) * (
                temperature1 - temperature0
            ) / (weight1 - weight0)
    raise ValueError(f"TGA曲线未跨越{threshold}%")


def _trapz_positive(points: list[tuple[float, float]]) -> float:
    area = 0.0
    for (x0, y0), (x1, y1) in zip(points, points[1:], strict=False):
        if x1 > x0:
            area += (x1 - x0) * (max(y0, 0.0) + max(y1, 0.0)) / 2
    return area


def _mechanical_summary() -> dict[str, dict[str, object]]:
    workbook = load_workbook(MECHANICAL_SOURCE, read_only=True, data_only=True)
    result: dict[str, dict[str, object]] = {}
    try:
        tables = {
            "Elastic modulus": (
                "elastic_modulus",
                "elastic_modulus_mean_MPa",
                "elastic_modulus_std_MPa",
            ),
            "stress @ 100% strain": (
                "stress_at_100_percent",
                "stress_at_100_percent_mean_MPa",
                "stress_at_100_percent_std_MPa",
            ),
        }
        for sheet_name, (prefix, mean_name, std_name) in tables.items():
            worksheet = workbook[sheet_name]
            for values in worksheet.iter_rows(min_row=3, max_row=6, values_only=True):
                label = str(values[1])
                replicates = [
                    number
                    for number in (_finite(value) for value in values[2:8])
                    if number is not None
                ]
                result.setdefault(label, {}).update(
                    {
                        f"{prefix}_replicate_count": len(replicates),
                        mean_name: _finite(values[8]),
                        std_name: _finite(values[9]),
                    }
                )
    finally:
        workbook.close()
    return result


def build_release() -> pd.DataFrame:
    tga_workbook = load_workbook(TGA_SOURCE, read_only=True, data_only=True)
    tensile_workbook = load_workbook(
        TENSILE_SOURCE, read_only=True, data_only=True
    )
    mechanical = _mechanical_summary()
    records = []
    try:
        tga_sheet = tga_workbook["TPU-NIC"]
        tga_rows = list(tga_sheet.iter_rows(min_row=3, values_only=True))
        tensile_sheet = tensile_workbook["Sheet1 (2)"]
        tensile_rows = list(tensile_sheet.iter_rows(min_row=3, values_only=True))
        for material_id, source_label, nic_fraction, tga_column, stress_column, strain_column in SPECS:
            raw_tga = []
            for values in tga_rows:
                temperature = _finite(values[tga_column])
                weight = _finite(values[tga_column + 1])
                if temperature is not None and weight is not None:
                    raw_tga.append((temperature, weight))
            baseline = statistics.median(weight for _, weight in raw_tga[:100])
            normalised_tga = [
                (temperature, 100.0 * weight / baseline)
                for temperature, weight in raw_tga
                if temperature >= 100
            ]
            tensile_points = []
            for values in tensile_rows:
                stress = _finite(values[stress_column])
                strain_fraction = _finite(values[strain_column])
                if stress is not None and strain_fraction is not None:
                    tensile_points.append((strain_fraction, stress))
            records.append(
                {
                    "release_id": RELEASE_ID,
                    "source_id": f"doi:{DATASET_DOI}",
                    "material_id": material_id,
                    "source_material_label": source_label,
                    "material_family": "Tecoflex_EG60D_niclosamide_composite",
                    "TPU_grade": "Tecoflex EG-60D",
                    "TPU_supplier": "Lubrizol",
                    "TPU_soft_to_hard_segment_ratio": "3:1",
                    "TPU_Tg_C": 23.0,
                    "niclosamide_wt_percent": nic_fraction,
                    "TPU_nominal_wt_percent": 100.0 - nic_fraction,
                    "chemistry_mapping_status": (
                        "commercial_grade_additive_fraction_mapped"
                    ),
                    "T5_C": _crossing_temperature(normalised_tga, 95.0),
                    "T10_C": _crossing_temperature(normalised_tga, 90.0),
                    "T50_C": _crossing_temperature(normalised_tga, 50.0),
                    "residue_at_max_temperature_percent": normalised_tga[-1][1],
                    "TGA_max_temperature_C": normalised_tga[-1][0],
                    "TGA_point_count": len(raw_tga),
                    "TGA_baseline_weight_percent": baseline,
                    "TGA_heating_rate_C_min": 20.0,
                    "TGA_atmosphere": "nitrogen",
                    "TGA_nitrogen_flow_mL_min": 60.0,
                    "TGA_sample_mass_mg_approx": 10.0,
                    "maximum_observed_stress_MPa": max(
                        stress for _, stress in tensile_points
                    ),
                    "maximum_observed_strain_percent": 100.0
                    * max(strain for strain, _ in tensile_points),
                    "partial_tensile_curve_area_MJ_m3": _trapz_positive(
                        tensile_points
                    ),
                    "tensile_curve_point_count": len(tensile_points),
                    "strain_unit_resolution": (
                        "source_header_percent_but_values_fraction; normalized_x100"
                    ),
                    "toughness_evidence_level": (
                        "partial_tensile_curve_area_lower_bound"
                    ),
                    **mechanical[source_label],
                    "tensile_standard": "ISO 527-1",
                    "tensile_rate_mm_min": 20.0,
                    "grip_distance_mm": 20.0,
                    "solvent_cast_TPU_chloroform_w_v_percent": 12.5,
                    "vacuum_dry_days": 3.0,
                    "vacuum_dry_temperature_C": 25.0,
                    "extrusion_temperature_C": 180.0,
                    "extrusion_speed_rpm": 75.0,
                    "thermoplastic_tpu_core": True,
                    "model_admission_layer": "core_tpu_composite_experimental",
                    "usage_mode": "partial_toughness_and_thermal_supervision",
                    "future_weight_ceiling": 0.75,
                    "split_group": f"{DATASET_DOI}|{material_id}",
                    "source_locator": (
                        f"{TGA_SOURCE.relative_to(ROOT).as_posix()}#TPU-NIC;"
                        f"{TENSILE_SOURCE.relative_to(ROOT).as_posix()}#Sheet1 (2)"
                    ),
                    "license": LICENSE,
                    "citation_keys": CITATIONS,
                }
            )
    finally:
        tga_workbook.close()
        tensile_workbook.close()
    return pd.DataFrame(records)


def write_release(frame: pd.DataFrame) -> None:
    DIRECTED.mkdir(parents=True, exist_ok=True)
    frame.to_csv(OUTPUT, index=False, encoding="utf-8-sig", lineterminator="\n")
    payload = {
        "release_id": RELEASE_ID,
        "counts": {
            "formulation_count": len(frame),
            "tga_curve_count": len(frame),
            "tga_source_point_count": int(frame["TGA_point_count"].sum()),
            "tensile_curve_count": len(frame),
            "tensile_source_point_count": int(
                frame["tensile_curve_point_count"].sum()
            ),
            "mechanical_direct_specimen_slot_count": int(
                frame["elastic_modulus_replicate_count"].sum()
            ),
            "published_compact_row_count": len(frame),
        },
        "source": {
            "dataset_doi": DATASET_DOI,
            "article_doi": ARTICLE_DOI,
            "license": LICENSE,
            "inputs": [
                _entry(path)
                for path in (
                    TGA_SOURCE,
                    TENSILE_SOURCE,
                    MECHANICAL_SOURCE,
                    SOURCE_METADATA,
                    SOURCE_AUDIT,
                )
            ],
        },
        "policy": {
            "pure_niclosamide_tga_excluded": True,
            "raw_curves_republished": False,
            "raw_curve_reason": (
                "以4条紧凑多性能记录替代22463个TGA点和1136个拉伸点"
            ),
            "tensile_curve_is_partial_not_break_toughness": True,
            "strain_header_conflict_resolved_from_article_and_100pct_scalars": True,
        },
        "output": _entry(OUTPUT),
    }
    MANIFEST.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def check_release(frame: pd.DataFrame) -> None:
    with tempfile.TemporaryDirectory(prefix="tecoflex-nic-check-") as directory:
        temporary = Path(directory) / OUTPUT.name
        frame.to_csv(
            temporary, index=False, encoding="utf-8-sig", lineterminator="\n"
        )
        if _sha256(temporary) != _sha256(OUTPUT):
            raise SystemExit("Tecoflex药物复合TPU输出不一致")
    print("Tecoflex药物复合TPU数据检查通过")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--检查", action="store_true")
    args = parser.parse_args()
    frame = build_release()
    if args.检查:
        check_release(frame)
    else:
        write_release(frame)
        print(
            json.dumps(
                {
                    "formulations": len(frame),
                    "tga_points_summarized": int(frame["TGA_point_count"].sum()),
                    "tensile_points_summarized": int(
                        frame["tensile_curve_point_count"].sum()
                    ),
                },
                ensure_ascii=False,
            )
        )


if __name__ == "__main__":
    main()
