"""Read-only profiling of computational and virtual-candidate TPU data assets.

The counts produced here are *candidates* for later adapters.  No row is
promoted to an admitted scientific observation and no training split is made.
"""

from __future__ import annotations

import csv
import hashlib
import math
import re
import zlib
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Mapping, Sequence


DEFAULT_POLYOMICS_QOI_COLUMNS = (
    "density",
    "Rg",
    "self-diffusion",
    "Cp",
    "Cv",
    "compressibility",
    "isentropic_compressibility",
    "bulk_modulus",
    "isentropic_bulk_modulus",
    "volume_expansion",
    "linear_expansion",
    "static_dielectric_const",
    "dielectric_const_dc",
    "nematic_order_parameter",
    "refractive_index",
    "thermal_conductivity",
    "thermal_diffusivity",
    "tg",
    "sp_ced",
    "sp_total",
    "sp_vdw",
    "sp_ele",
)

_FIDELITY_SUFFIX = re.compile(r"_(DFT|MD|GC)\.csv$", re.IGNORECASE)
_IDENTITY_SENTINELS = frozenset({"nan", "na", "n/a", "none", "null", "-", "--"})
_PUE_RESPONSE_COLUMNS = frozenset({"logEB", "logYM", "logTS"})
_MISSING_RATIO_TOKENS = (
    "0.1",
    "0.2",
    "0.30000000000000004",
    "0.4",
    "0.5",
    "0.6000000000000001",
    "0.7000000000000001",
    "0.8",
    "0.9",
)
_FILLED_RESULT_METHODS = frozenset(
    {
        "fill_with_dt",
        "fill_with_et",
        "fill_with_gbr",
        "fill_with_lgb",
        "fill_with_rf",
        "fill_with_ridge",
        "fill_with_xgb",
    }
)
_ET_METRIC_COLUMNS = (
    "Evaluated: et",
    "mean",
    "hyperimpute",
    "missforest",
    "gain",
    "sinkhorn",
)
_MEAN_STD_PATTERN = re.compile(
    r"^\s*(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?\s+\+/-\s+"
    r"(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?\s*$"
)
_REFERENCE_BY_SOURCE = {
    "pi1m": "[6]",
    "smipoly": "[7]",
    "adept": "[8]",
    "polygraphmt": "[8]",
    "dq_matimpute": "[1], [12]",
    "polyomics": "[20]",
}


class ComputationalAdmissionError(ValueError):
    """Structured profiling failure that blocks an admission claim."""

    def __init__(self, code: str, message: str, *, path: str | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.path = path

    def as_dict(self) -> dict[str, str]:
        payload = {"code": self.code, "message": self.message}
        if self.path is not None:
            payload["path"] = self.path
        return payload


@dataclass(frozen=True)
class ComputationalAdmissionProfile:
    source_key: str
    evidence_class: str
    file_count: int
    source_record_candidate_count: int | None
    unique_system_candidate_count: int | None
    computational_activity_candidate_count: int | None
    computational_observation_candidate_count: int | None
    admitted_observation_count: int = 0
    fidelity_counts: Mapping[str, int] = field(default_factory=dict)
    diagnostics: Mapping[str, int | str] = field(default_factory=dict)
    blocking_reason: str = "逐记录体系、协议、条件、单位、QoI 与不确定度适配尚未完成"

    def __post_init__(self) -> None:
        if not self.source_key.strip() or not self.evidence_class.strip():
            raise ComputationalAdmissionError("invalid_profile", "profile keys must be non-empty")
        counts = (
            self.file_count,
            self.source_record_candidate_count,
            self.unique_system_candidate_count,
            self.computational_activity_candidate_count,
            self.computational_observation_candidate_count,
            self.admitted_observation_count,
        )
        if any(
            value is not None
            and (isinstance(value, bool) or not isinstance(value, int) or value < 0)
            for value in counts
        ):
            raise ComputationalAdmissionError(
                "invalid_profile", "profile counts must be non-negative integers or null"
            )
        if any(
            not isinstance(key, str)
            or not key.strip()
            or isinstance(value, bool)
            or not isinstance(value, int)
            or value < 0
            for key, value in self.fidelity_counts.items()
        ):
            raise ComputationalAdmissionError(
                "invalid_profile", "fidelity counts require non-empty keys and non-negative integers"
            )
        if self.admitted_observation_count != 0:
            raise ComputationalAdmissionError(
                "premature_admission", "this profiling stage cannot admit computational observations"
            )


@dataclass(frozen=True)
class ExactStructureOverlapProfile:
    """Case-sensitive exact-string overlap lower bounds for leakage control."""

    source_exact_structure_counts: Mapping[str, int]
    pair_overlap_counts: Mapping[str, int]
    diagnostics: Mapping[str, int | str]

    def __post_init__(self) -> None:
        expected_sources = {"pi1m", "adept", "polyomics", "polygraphmt"}
        expected_pairs = {
            "pi1m__adept",
            "pi1m__polyomics",
            "pi1m__polygraphmt",
            "adept__polyomics",
            "adept__polygraphmt",
            "polyomics__polygraphmt",
        }
        if set(self.source_exact_structure_counts) != expected_sources:
            raise ComputationalAdmissionError(
                "invalid_overlap_profile",
                "overlap profile must contain the four frozen source keys",
            )
        if set(self.pair_overlap_counts) != expected_pairs:
            raise ComputationalAdmissionError(
                "invalid_overlap_profile",
                "overlap profile must contain the six frozen pair keys",
            )
        for mapping in (
            self.source_exact_structure_counts,
            self.pair_overlap_counts,
        ):
            if any(
                not isinstance(key, str)
                or not key.strip()
                or isinstance(value, bool)
                or not isinstance(value, int)
                or value < 0
                for key, value in mapping.items()
            ):
                raise ComputationalAdmissionError(
                    "invalid_overlap_profile",
                    "overlap counts require non-empty keys and non-negative integers",
                )
        for pair, overlap in self.pair_overlap_counts.items():
            left, right = pair.split("__", 1)
            if overlap > min(
                self.source_exact_structure_counts[left],
                self.source_exact_structure_counts[right],
            ):
                raise ComputationalAdmissionError(
                    "invalid_overlap_profile",
                    "pair overlap cannot exceed either source cardinality",
                )
        if any(
            not isinstance(key, str)
            or not key.strip()
            or isinstance(value, bool)
            or not isinstance(value, (int, str))
            or (isinstance(value, int) and value < 0)
            for key, value in self.diagnostics.items()
        ):
            raise ComputationalAdmissionError(
                "invalid_overlap_profile",
                "overlap diagnostics require non-empty keys and integer or string values",
            )


def _dict_rows(path: str | Path) -> tuple[tuple[str, ...], Iterable[dict[str, str]]]:
    source = Path(path)
    try:
        stream = source.open(encoding="utf-8-sig", newline="")
    except OSError as error:
        raise ComputationalAdmissionError(
            "file_unreadable", "candidate CSV cannot be opened", path=source.as_posix()
        ) from error
    reader = csv.DictReader(stream)
    try:
        columns = tuple(reader.fieldnames or ())
    except UnicodeError as error:
        stream.close()
        raise ComputationalAdmissionError(
            "file_decode_failed", "candidate CSV is not valid UTF-8", path=source.as_posix()
        ) from error
    if len(columns) != len(set(columns)):
        stream.close()
        raise ComputationalAdmissionError(
            "duplicate_columns",
            "candidate CSV contains duplicate header names",
            path=source.as_posix(),
        )

    def iterator() -> Iterable[dict[str, str]]:
        try:
            for row in reader:
                if None in row:
                    raise ComputationalAdmissionError(
                        "invalid_row_width", "CSV row has more values than its header", path=source.as_posix()
                    )
                yield {key: value if value is not None else "" for key, value in row.items()}
        except UnicodeError as error:
            raise ComputationalAdmissionError(
                "file_decode_failed", "candidate CSV is not valid UTF-8", path=source.as_posix()
            ) from error
        finally:
            stream.close()

    return columns, iterator()


def _require_columns(columns: Sequence[str], required: Sequence[str], path: Path) -> None:
    missing = [column for column in required if column not in columns]
    if missing:
        raise ComputationalAdmissionError(
            "invalid_columns",
            f"candidate CSV is missing columns: {', '.join(missing)}",
            path=path.as_posix(),
        )


def _identity_is_missing(value: str) -> bool:
    return not value or value.casefold() in _IDENTITY_SENTINELS


def profile_polygraphmt(raw_directory: str | Path) -> ComputationalAdmissionProfile:
    """Profile DFT/MD/group-contribution property rows without admitting them."""

    root = Path(raw_directory)
    files = sorted(root.glob("*.csv"), key=lambda path: (path.name.casefold(), path.name))
    if not files:
        raise ComputationalAdmissionError(
            "no_candidate_files", "PolyGraphMT raw directory contains no CSV files", path=root.as_posix()
        )
    unique_systems: set[str] = set()
    fidelity_counts: defaultdict[str, int] = defaultdict(int)
    total_rows = 0
    valid_observation_rows = 0
    invalid_identity_rows = 0
    duplicate_groups = 0
    duplicate_extra_rows = 0
    conflicting_groups = 0
    conflicting_extra_rows = 0
    redundant_extra_rows = 0
    invalid_numeric_rows = 0
    nonfinite_numeric_rows = 0

    for path in files:
        match = _FIDELITY_SUFFIX.search(path.name)
        if match is None:
            raise ComputationalAdmissionError(
                "unknown_fidelity", "filename does not freeze DFT, MD, or GC fidelity", path=path.as_posix()
            )
        fidelity = match.group(1).lower()
        columns, rows = _dict_rows(path)
        if len(columns) != 2 or columns[0] != "SMILES":
            raise ComputationalAdmissionError(
                "invalid_columns", "PolyGraphMT table must have SMILES and one target", path=path.as_posix()
            )
        target = columns[1]
        values_by_system: defaultdict[str, list[str]] = defaultdict(list)
        for row in rows:
            system = row["SMILES"].strip()
            value = row[target].strip()
            if not value:
                raise ComputationalAdmissionError(
                    "missing_value", "PolyGraphMT row has blank target value", path=path.as_posix()
                )
            total_rows += 1
            if not system:
                raise ComputationalAdmissionError(
                    "missing_identity", "PolyGraphMT row has blank SMILES", path=path.as_posix()
                )
            if _identity_is_missing(system):
                invalid_identity_rows += 1
                continue
            try:
                numeric = float(value)
            except ValueError:
                invalid_numeric_rows += 1
                continue
            if not math.isfinite(numeric):
                nonfinite_numeric_rows += 1
                continue
            valid_observation_rows += 1
            fidelity_counts[fidelity] += 1
            unique_systems.add(system)
            values_by_system[system].append(value)
        for values in values_by_system.values():
            if len(values) > 1:
                duplicate_groups += 1
                duplicate_extra_rows += len(values) - 1
                if len(set(values)) > 1:
                    conflicting_groups += 1
                    conflicting_extra_rows += len(values) - 1
                else:
                    redundant_extra_rows += len(values) - 1

    casefold_systems: defaultdict[str, set[str]] = defaultdict(set)
    for system in unique_systems:
        casefold_systems[system.casefold()].add(system)
    casefold_collision_groups = sum(len(values) > 1 for values in casefold_systems.values())
    casefold_collision_extra = sum(len(values) - 1 for values in casefold_systems.values())

    return ComputationalAdmissionProfile(
        source_key="polygraphmt",
        evidence_class="multi_fidelity_computational_observation_candidate",
        file_count=len(files),
        source_record_candidate_count=total_rows,
        unique_system_candidate_count=len(unique_systems),
        computational_activity_candidate_count=None,
        computational_observation_candidate_count=valid_observation_rows,
        fidelity_counts=dict(sorted(fidelity_counts.items())),
        diagnostics={
            "duplicate_key_group_count": duplicate_groups,
            "duplicate_extra_row_count": duplicate_extra_rows,
            "conflicting_target_group_count": conflicting_groups,
            "conflicting_extra_row_count": conflicting_extra_rows,
            "redundant_extra_row_count": redundant_extra_rows,
            "invalid_identity_row_count": invalid_identity_rows,
            "invalid_numeric_row_count": invalid_numeric_rows,
            "nonfinite_numeric_row_count": nonfinite_numeric_rows,
            "raw_source_row_count": total_rows,
            "valid_identity_property_occurrence_count": valid_observation_rows,
            "unique_raw_property_key_count": valid_observation_rows - duplicate_extra_rows,
            "casefold_collision_group_count": casefold_collision_groups,
            "casefold_collision_extra_structure_count": casefold_collision_extra,
            "identity_basis": "case-sensitive exact SMILES string",
        },
        blocking_reason=(
            "含无效身份哨兵且缺少逐行计算活动、方法版本、条件和不确定度；冲突重复未裁决前不得按 SMILES 聚合"
        ),
    )


def _read_polyomics(
    path: Path,
    *,
    qoi_columns: Sequence[str],
    collect_qoi: bool,
    capture_uids: set[str] | None = None,
    capture_all: bool = False,
) -> tuple[
    int,
    set[str],
    set[str],
    int,
    tuple[str, ...],
    dict[str, tuple[str, ...]],
]:
    columns, rows = _dict_rows(path)
    required = ["UUID", "smiles_list", *qoi_columns] if collect_qoi else ["UUID", "smiles_list"]
    _require_columns(columns, required, path)
    count = 0
    uuids: set[str] = set()
    systems: set[str] = set()
    observations = 0
    captured_rows: dict[str, tuple[str, ...]] = {}
    for row in rows:
        uid = row["UUID"].strip()
        system = row["smiles_list"].strip()
        if _identity_is_missing(uid) or _identity_is_missing(system):
            raise ComputationalAdmissionError(
                "missing_identity", "PolyOmics row lacks UUID or smiles_list", path=path.as_posix()
            )
        if uid in uuids:
            raise ComputationalAdmissionError(
                "duplicate_source_record_id", "PolyOmics UUID is not unique", path=path.as_posix()
            )
        uuids.add(uid)
        systems.add(system)
        if capture_all or (capture_uids is not None and uid in capture_uids):
            captured_rows[uid] = tuple(row[column] for column in columns)
        count += 1
        if collect_qoi:
            for column in qoi_columns:
                raw_value = row[column].strip()
                if not raw_value:
                    continue
                try:
                    numeric = float(raw_value)
                except ValueError as error:
                    raise ComputationalAdmissionError(
                        "invalid_qoi_value",
                        f"PolyOmics QoI {column} is not numeric",
                        path=path.as_posix(),
                    ) from error
                if not math.isfinite(numeric):
                    raise ComputationalAdmissionError(
                        "invalid_qoi_value",
                        f"PolyOmics QoI {column} is not finite",
                        path=path.as_posix(),
                    )
                observations += 1
    return count, uuids, systems, observations, columns, captured_rows


def _numeric_format_equivalent(left: str, right: str) -> bool:
    try:
        left_value = float(left)
        right_value = float(right)
    except ValueError:
        return False
    return (
        math.isfinite(left_value)
        and math.isfinite(right_value)
        and math.isclose(left_value, right_value, rel_tol=1e-12, abs_tol=0.0)
    )


_POLYOMICS_CONTEXT_COLUMNS = (
    "monomer_ID",
    "copoly_ratio_list",
    "copoly_type",
    "temp",
    "press",
    "tacticity",
    "qm_method",
    "forcefield",
)


def _boolean_bucket(value: str, *, column: str, path: Path) -> str:
    normalized = value.strip().casefold()
    if not normalized:
        return "blank"
    if normalized in {"true", "false"}:
        return normalized
    raise ComputationalAdmissionError(
        "invalid_quality_flag",
        f"PolyOmics {column} is not True, False, or blank",
        path=path.as_posix(),
    )


def _polyomics_governance_diagnostics(
    path: Path, *, qoi_columns: Sequence[str]
) -> tuple[set[str], dict[str, int]]:
    columns, rows = _dict_rows(path)
    required = (
        "UUID",
        "smiles_list",
        "class_PURT",
        "check_eq",
        "check_tc",
        "do_TC",
        "remarks",
        *_POLYOMICS_CONTEXT_COLUMNS,
        *qoi_columns,
    )
    _require_columns(columns, required, path)
    purt_uuids: set[str] = set()
    system_counts: Counter[str] = Counter()
    contexts: defaultdict[str, set[tuple[str, ...]]] = defaultdict(set)
    flag_counts: Counter[str] = Counter()
    remarks_nonempty = 0
    missing_qoi = 0
    for row in rows:
        uid = row["UUID"].strip()
        system = row["smiles_list"].strip()
        class_purt = _boolean_bucket(row["class_PURT"], column="class_PURT", path=path)
        if class_purt == "true":
            purt_uuids.add(uid)
        for column in ("check_eq", "check_tc", "do_TC"):
            flag_counts[f"{column.casefold()}_{_boolean_bucket(row[column], column=column, path=path)}_count"] += 1
        remarks_nonempty += bool(row["remarks"].strip())
        missing_qoi += sum(not row[column].strip() for column in qoi_columns)
        system_counts[system] += 1
        contexts[system].add(tuple(row[column].strip() for column in _POLYOMICS_CONTEXT_COLUMNS))
    diagnostics = {
        "duplicate_exact_system_group_count": sum(value > 1 for value in system_counts.values()),
        "duplicate_exact_system_extra_record_count": sum(value - 1 for value in system_counts.values()),
        "duplicate_system_context_variant_group_count": sum(
            system_counts[system] > 1 and len(values) > 1
            for system, values in contexts.items()
        ),
        "remarks_nonempty_count": remarks_nonempty,
        "qoi_missing_cell_count": missing_qoi,
        **dict(sorted(flag_counts.items())),
    }
    return purt_uuids, diagnostics


def _validate_purt_material_semantics(
    columns: Sequence[str], rows: Mapping[str, tuple[str, ...]], path: Path
) -> dict[str, int]:
    required = ("smiles_1", "smiles_2", "smiles_3", "smiles_4", "copoly_ratio_list", "copoly_type", "class_PURT")
    _require_columns(columns, required, path)
    index = {column: columns.index(column) for column in required}
    one_component = 0
    for values in rows.values():
        component_count = sum(bool(values[index[f"smiles_{number}"]].strip()) for number in range(1, 5))
        try:
            ratio = float(values[index["copoly_ratio_list"]].strip())
        except ValueError as error:
            raise ComputationalAdmissionError(
                "subset_material_semantics_mismatch",
                "PolyOmics PURT copoly_ratio_list is not numeric",
                path=path.as_posix(),
            ) from error
        if (
            component_count != 1
            or not math.isclose(ratio, 1.0, rel_tol=0.0, abs_tol=0.0)
            or values[index["copoly_type"]].strip()
            or values[index["class_PURT"]].strip().casefold() != "true"
        ):
            raise ComputationalAdmissionError(
                "subset_material_semantics_mismatch",
                "PolyOmics PURT is not uniformly a one-component class_PURT view",
                path=path.as_posix(),
            )
        one_component += 1
    return {
        "purt_one_component_row_count": one_component,
        "purt_multicomponent_row_count": len(rows) - one_component,
    }


def profile_polyomics(
    general_path: str | Path,
    purt_path: str | Path,
    *,
    qoi_columns: Sequence[str] = DEFAULT_POLYOMICS_QOI_COLUMNS,
) -> ComputationalAdmissionProfile:
    """Profile the parent computational table and prove PURT is only a subset view."""

    if not qoi_columns or len(qoi_columns) != len(set(qoi_columns)):
        raise ComputationalAdmissionError("invalid_qoi_catalog", "QoI catalog must be unique and non-empty")
    general = Path(general_path)
    purt = Path(purt_path)
    purt_rows, purt_uuids, purt_systems, _, purt_columns, purt_values = _read_polyomics(
        purt, qoi_columns=(), collect_qoi=False, capture_all=True
    )
    rows, general_uuids, systems, observations, general_columns, general_values = _read_polyomics(
        general,
        qoi_columns=qoi_columns,
        collect_qoi=True,
        capture_uids=purt_uuids,
    )
    if purt_columns != general_columns:
        raise ComputationalAdmissionError(
            "subset_schema_mismatch",
            "PolyOmics PURT schema differs from the parent table",
            path=purt.as_posix(),
        )
    outside = purt_uuids - general_uuids
    if outside:
        raise ComputationalAdmissionError(
            "subset_not_contained",
            "PolyOmics PURT contains UUIDs absent from the parent table",
            path=purt.as_posix(),
        )
    raw_changed = 0
    format_equivalent = 0
    materially_changed: list[str] = []
    for uid in sorted(purt_uuids):
        parent_values = general_values[uid]
        subset_values = purt_values[uid]
        if parent_values == subset_values:
            continue
        raw_changed += 1
        differences = [
            (left, right)
            for left, right in zip(parent_values, subset_values, strict=True)
            if left != right
        ]
        if differences and all(_numeric_format_equivalent(left, right) for left, right in differences):
            format_equivalent += 1
        else:
            materially_changed.append(uid)
    if materially_changed:
        raise ComputationalAdmissionError(
            "subset_content_mismatch",
            "PolyOmics PURT has materially different rows from the parent table",
            path=purt.as_posix(),
        )
    expected_purt_uuids, governance_diagnostics = _polyomics_governance_diagnostics(
        general, qoi_columns=qoi_columns
    )
    if expected_purt_uuids != purt_uuids:
        raise ComputationalAdmissionError(
            "subset_predicate_mismatch",
            "PolyOmics PURT UUIDs are not exactly the general class_PURT=True set",
            path=purt.as_posix(),
        )
    material_diagnostics = _validate_purt_material_semantics(
        purt_columns, purt_values, purt
    )
    return ComputationalAdmissionProfile(
        source_key="polyomics",
        evidence_class="md_qm_aggregate_with_subset_view",
        file_count=2,
        source_record_candidate_count=rows,
        unique_system_candidate_count=len(systems),
        computational_activity_candidate_count=None,
        computational_observation_candidate_count=observations,
        fidelity_counts={"md_qm_aggregate": rows},
        diagnostics={
            "purt_subset_rows": purt_rows,
            "purt_subset_unique_systems": len(purt_systems),
            "purt_uuid_not_in_general": len(outside),
            "purt_raw_content_mismatch_count": raw_changed,
            "purt_numeric_format_equivalent_count": format_equivalent,
            "purt_material_content_mismatch_count": len(materially_changed),
            "qoi_column_count": len(qoi_columns),
            "unique_source_record_uuid_count": len(general_uuids),
            "identity_basis": "UUID=source_record; smiles_list=exact_structure_string_only",
            **governance_diagnostics,
            **material_diagnostics,
        },
        blocking_reason=(
            "QoI 单元格仅为数值候选；需逐字段绑定单位、协议、质量标志、完整体系、完成状态和不确定度，PURT 不重复计数且不能仅凭名称认定为TPU"
        ),
    )


def profile_structure_candidates(
    path: str | Path,
    *,
    source_key: str,
    identity_column: str,
    evidence_class: str,
    system_column: str | None = None,
    require_unique_identity: bool = False,
) -> ComputationalAdmissionProfile:
    """Count virtual structures or reaction records without treating them as properties."""

    source = Path(path)
    columns, rows = _dict_rows(source)
    system_field = system_column or identity_column
    _require_columns(columns, [identity_column, system_field], source)
    count = 0
    identity_counts: Counter[str] = Counter()
    system_counts: Counter[str] = Counter()
    for row in rows:
        identity = row[identity_column].strip()
        system = row[system_field].strip()
        if _identity_is_missing(identity) or _identity_is_missing(system):
            raise ComputationalAdmissionError(
                "missing_identity",
                f"blank {identity_column} or {system_field} in candidate table",
                path=source.as_posix(),
            )
        count += 1
        identity_counts[identity] += 1
        system_counts[system] += 1
    duplicate_identity_rows = sum(value - 1 for value in identity_counts.values())
    if require_unique_identity and duplicate_identity_rows:
        raise ComputationalAdmissionError(
            "duplicate_record_identity",
            f"{identity_column} is not unique in candidate table",
            path=source.as_posix(),
        )
    duplicate_system_groups = sum(value > 1 for value in system_counts.values())
    duplicate_system_rows = sum(value - 1 for value in system_counts.values())
    return ComputationalAdmissionProfile(
        source_key=source_key,
        evidence_class=evidence_class,
        file_count=1,
        source_record_candidate_count=count,
        unique_system_candidate_count=len(system_counts),
        computational_activity_candidate_count=0,
        computational_observation_candidate_count=0,
        diagnostics={
            "unique_record_identity_count": len(identity_counts),
            "duplicate_identity_rows": duplicate_identity_rows,
            "duplicate_system_group_count": duplicate_system_groups,
            "duplicate_system_rows": duplicate_system_rows,
        },
        blocking_reason="仅提供结构空间或反应约束，没有实验或计算性能观测",
    )


def profile_adept_candidates(
    smiles_path: str | Path, *, simulation_input_file_count: int
) -> ComputationalAdmissionProfile:
    """Profile ADEPT structures and inputs while requiring real outputs for observations."""

    if (
        isinstance(simulation_input_file_count, bool)
        or not isinstance(simulation_input_file_count, int)
        or simulation_input_file_count < 0
    ):
        raise ComputationalAdmissionError(
            "invalid_input_count", "simulation input file count must be non-negative"
        )
    source = Path(smiles_path)
    columns, rows = _dict_rows(source)
    _require_columns(columns, ["PID", "SMILES"], source)
    row_count = 0
    pids: set[str] = set()
    systems: set[str] = set()
    for row in rows:
        pid = row["PID"].strip()
        system = row["SMILES"].strip()
        if _identity_is_missing(pid) or _identity_is_missing(system):
            raise ComputationalAdmissionError(
                "missing_identity", "ADEPT row lacks PID or SMILES", path=source.as_posix()
            )
        if pid in pids:
            raise ComputationalAdmissionError(
                "duplicate_source_record_id", "ADEPT PID is not unique", path=source.as_posix()
            )
        pids.add(pid)
        systems.add(system)
        row_count += 1
    return ComputationalAdmissionProfile(
        source_key="adept",
        evidence_class="virtual_candidate_and_simulation_workflow_input",
        file_count=1 + simulation_input_file_count,
        source_record_candidate_count=row_count,
        unique_system_candidate_count=len(systems),
        computational_activity_candidate_count=0,
        computational_observation_candidate_count=0,
        diagnostics={
            "unique_pid_count": len(pids),
            "simulation_input_file_count": simulation_input_file_count,
        },
        blocking_reason="已发现候选结构和输入模板，但未发现可回连协议、体系与输出的本地计算活动结果",
    )


@dataclass(frozen=True)
class PueParentAudit:
    columns: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]
    content_sha256: str


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError as error:
        raise ComputationalAdmissionError(
            "file_unreadable", "candidate file cannot be read", path=path.as_posix()
        ) from error
    return digest.hexdigest()


def _table_rows(path: Path) -> tuple[tuple[str, ...], tuple[tuple[str, ...], ...]]:
    columns, row_iterator = _dict_rows(path)
    rows = tuple(tuple(row[column] for column in columns) for row in row_iterator)
    return columns, rows


def _finite_decimal(value: str, *, code: str, path: Path) -> Decimal:
    try:
        number = Decimal(value.strip())
    except (InvalidOperation, ValueError) as error:
        raise ComputationalAdmissionError(
            code, "expected a finite numeric value", path=path.as_posix()
        ) from error
    if not number.is_finite():
        raise ComputationalAdmissionError(
            code, "expected a finite numeric value", path=path.as_posix()
        )
    return number


def audit_pue_parent_occurrences(
    dq_parent_path: str | Path,
    matimpute_parent_path: str | Path,
    *,
    expected_row_count: int = 326,
    expected_column_count: int = 24,
) -> PueParentAudit:
    """Verify the byte-identical PUE parents and their numeric source rows."""

    for label, value in (
        ("expected_row_count", expected_row_count),
        ("expected_column_count", expected_column_count),
    ):
        if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
            raise ComputationalAdmissionError(
                "invalid_input_count", f"{label} must be a positive integer"
            )
    dq = Path(dq_parent_path)
    matimpute = Path(matimpute_parent_path)
    dq_hash = _sha256_path(dq)
    matimpute_hash = _sha256_path(matimpute)
    if dq_hash != matimpute_hash:
        raise ComputationalAdmissionError(
            "parent_dataset_mismatch",
            "DQ and MatImpute PUE parent files are not byte-identical",
        )
    columns, rows = _table_rows(dq)
    required = {"SSID", *_PUE_RESPONSE_COLUMNS}
    if (
        len(columns) != expected_column_count
        or len(rows) != expected_row_count
        or not required.issubset(columns)
    ):
        raise ComputationalAdmissionError(
            "pue_parent_schema_mismatch",
            "PUE parent dimensions or required identity/response columns differ",
            path=dq.as_posix(),
        )
    identity_index = columns.index("SSID")
    identities: set[str] = set()
    for row in rows:
        identity = row[identity_index].strip()
        if _identity_is_missing(identity) or identity in identities:
            raise ComputationalAdmissionError(
                "pue_parent_content_invalid",
                "PUE parent SSID must be present and unique",
                path=dq.as_posix(),
            )
        identities.add(identity)
        for index, value in enumerate(row):
            if index == identity_index:
                continue
            if not value.strip():
                raise ComputationalAdmissionError(
                    "pue_parent_content_invalid",
                    "PUE parent numeric cells cannot be blank",
                    path=dq.as_posix(),
                )
            _finite_decimal(value, code="pue_parent_content_invalid", path=dq)
    return PueParentAudit(columns=columns, rows=rows, content_sha256=dq_hash)


def audit_dq_pue_projections(
    parent: PueParentAudit,
    projections: Mapping[str, str | Path],
) -> Mapping[str, int]:
    """Require the two DQ tables to be exact, row-preserving column projections."""

    expected_targets = {"logTS", "logYM"}
    if set(projections) != expected_targets:
        raise ComputationalAdmissionError(
            "pue_projection_set_mismatch",
            "DQ PUE projections must contain exactly logTS and logYM",
        )
    parent_index = {column: index for index, column in enumerate(parent.columns)}
    for target in sorted(expected_targets):
        path = Path(projections[target])
        expected_columns = tuple(
            column
            for column in parent.columns
            if column != "SSID"
            and (column not in _PUE_RESPONSE_COLUMNS or column == target)
        )
        columns, rows = _table_rows(path)
        if columns != expected_columns or len(rows) != len(parent.rows):
            raise ComputationalAdmissionError(
                "pue_projection_content_mismatch",
                "DQ projection columns or row count differ from the parent projection",
                path=path.as_posix(),
            )
        for row_index, row in enumerate(rows):
            expected_row = tuple(
                parent.rows[row_index][parent_index[column]] for column in columns
            )
            if row != expected_row:
                raise ComputationalAdmissionError(
                    "pue_projection_content_mismatch",
                    "DQ projection changed a retained parent value or row order",
                    path=path.as_posix(),
                )
    return {
        "dq_projection_file_count": len(projections),
        "dq_projection_row_count_each": len(parent.rows),
        "dq_projection_column_count_each": len(parent.columns) - 3,
        "dq_projection_material_difference_count": 0,
    }


def _parse_missing_variant_name(path: Path) -> tuple[str, str] | None:
    for token in sorted(_MISSING_RATIO_TOKENS, key=len, reverse=True):
        suffix = f"_{token}.csv"
        if path.name.endswith(suffix):
            return path.name[: -len(suffix)], token
    return None


def audit_pue_missing_variants(
    parent: PueParentAudit,
    variants_directory: str | Path,
) -> Mapping[str, int]:
    """Verify the complete 23-column by 9-ratio artificial-missingness family."""

    root = Path(variants_directory)
    files = sorted(root.glob("*.csv"), key=lambda path: (path.name.casefold(), path.name))
    target_columns = tuple(column for column in parent.columns if column != "SSID")
    expected_keys = {
        (column, token) for column in target_columns for token in _MISSING_RATIO_TOKENS
    }
    actual: dict[tuple[str, str], Path] = {}
    for path in files:
        parsed = _parse_missing_variant_name(path)
        if parsed is None or parsed in actual:
            raise ComputationalAdmissionError(
                "pue_missing_variant_set_mismatch",
                "PUE missing-variant filename is unknown or duplicated",
                path=path.as_posix(),
            )
        actual[parsed] = path
    if set(actual) != expected_keys:
        raise ComputationalAdmissionError(
            "pue_missing_variant_set_mismatch",
            "PUE missing variants do not form the complete column by ratio grid",
            path=root.as_posix(),
        )
    parent_index = {column: index for index, column in enumerate(parent.columns)}
    identity_index = parent_index["SSID"]
    total_missing = 0
    non_target_format_coordinates: set[tuple[int, str]] = set()
    for target, token in sorted(actual):
        path = actual[(target, token)]
        columns, rows = _table_rows(path)
        if columns != parent.columns or len(rows) != len(parent.rows):
            raise ComputationalAdmissionError(
                "pue_missing_variant_content_mismatch",
                "PUE missing variant changed the parent shape or header",
                path=path.as_posix(),
            )
        target_index = parent_index[target]
        expected_missing = round(len(parent.rows) * float(token))
        observed_missing = 0
        for row_index, row in enumerate(rows):
            if row[identity_index] != parent.rows[row_index][identity_index]:
                raise ComputationalAdmissionError(
                    "pue_missing_variant_content_mismatch",
                    "PUE missing variant changed SSID identity or row order",
                    path=path.as_posix(),
                )
            for column_index, value in enumerate(row):
                parent_value = parent.rows[row_index][column_index]
                if column_index == identity_index:
                    continue
                if not value.strip():
                    if column_index != target_index:
                        raise ComputationalAdmissionError(
                            "pue_missing_variant_content_mismatch",
                            "PUE missing variant blanked a non-target column",
                            path=path.as_posix(),
                        )
                    observed_missing += 1
                    continue
                if _finite_decimal(
                    value, code="pue_missing_variant_content_mismatch", path=path
                ) != _finite_decimal(
                    parent_value,
                    code="pue_missing_variant_content_mismatch",
                    path=path,
                ):
                    raise ComputationalAdmissionError(
                        "pue_missing_variant_content_mismatch",
                        "PUE missing variant materially changed a retained value",
                        path=path.as_posix(),
                    )
                if column_index != target_index and value != parent_value:
                    non_target_format_coordinates.add(
                        (row_index, parent.columns[column_index])
                    )
        if observed_missing != expected_missing:
            raise ComputationalAdmissionError(
                "pue_missing_variant_content_mismatch",
                "PUE missing count does not equal the rounded declared ratio",
                path=path.as_posix(),
            )
        total_missing += observed_missing
    return {
        "missing_variant_file_count": len(actual),
        "missing_variant_target_column_count": len(target_columns),
        "missing_variant_ratio_count": len(_MISSING_RATIO_TOKENS),
        "missing_variant_intentional_blank_cell_count": total_missing,
        "missing_variant_material_difference_count": 0,
        "missing_variant_non_target_format_coordinate_count": len(
            non_target_format_coordinates
        ),
    }


def _audit_npy_output(path: Path) -> None:
    try:
        import numpy as np

        array = np.load(path, allow_pickle=False)
    except (OSError, ValueError) as error:
        raise ComputationalAdmissionError(
            "pue_model_output_invalid", "PUE RDF array cannot be loaded", path=path.as_posix()
        ) from error
    expected_centers = np.arange(3.0, 166.0, 6.0, dtype=np.float64)
    if (
        array.shape != (4, 28)
        or array.dtype != np.dtype("float64")
        or not np.isfinite(array).all()
        or not np.array_equal(array[3], expected_centers)
        or (array[:3] < 0).any()
    ):
        raise ComputationalAdmissionError(
            "pue_model_output_invalid",
            "PUE RDF array shape, dtype, finiteness, bins, or values differ",
            path=path.as_posix(),
        )


def _audit_filled_metrics(path: Path, expected_columns: set[str]) -> int:
    columns, rows = _dict_rows(path)
    expected_header = (
        "dataset",
        "column",
        "miss_ratio",
        "method",
        "RMSE",
        "Wasserstein",
        "time",
    )
    if columns != expected_header:
        raise ComputationalAdmissionError(
            "pue_model_output_invalid", "filled-results header differs", path=path.as_posix()
        )
    keys: set[tuple[str, str, str]] = set()
    for row in rows:
        if row["dataset"] != "PUE" or row["column"] not in expected_columns:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "filled-results scope differs", path=path.as_posix()
            )
        ratio = row["miss_ratio"]
        ratio_token = next(
            (
                token
                for token in _MISSING_RATIO_TOKENS
                if Decimal(token) == _finite_decimal(
                    ratio, code="pue_model_output_invalid", path=path
                )
            ),
            None,
        )
        if ratio_token is None or row["method"] not in _FILLED_RESULT_METHODS:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "filled-results grid value differs", path=path.as_posix()
            )
        key = (row["column"], ratio_token, row["method"])
        if key in keys:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "filled-results key is duplicated", path=path.as_posix()
            )
        keys.add(key)
        for metric in ("RMSE", "Wasserstein", "time"):
            if _finite_decimal(
                row[metric], code="pue_model_output_invalid", path=path
            ) < 0:
                raise ComputationalAdmissionError(
                    "pue_model_output_invalid", "filled-results metric is negative", path=path.as_posix()
                )
    expected_keys = {
        (column, token, method)
        for column in expected_columns
        for token in _MISSING_RATIO_TOKENS
        for method in _FILLED_RESULT_METHODS
    }
    if keys != expected_keys:
        raise ComputationalAdmissionError(
            "pue_model_output_invalid", "filled-results grid is incomplete", path=path.as_posix()
        )
    return len(keys)


def _audit_et_metrics(path: Path) -> int:
    columns, rows = _dict_rows(path)
    expected_header = ("Scenario", "miss_pct [0, 1]", *_ET_METRIC_COLUMNS)
    if columns != expected_header:
        raise ComputationalAdmissionError(
            "pue_model_output_invalid", "Et metric header differs", path=path.as_posix()
        )
    keys: set[tuple[str, Decimal]] = set()
    allowed_ratios = {Decimal(str(index / 10)) for index in range(1, 6)}
    for row in rows:
        ratio = _finite_decimal(
            row["miss_pct [0, 1]"], code="pue_model_output_invalid", path=path
        )
        key = (row["Scenario"], ratio)
        if row["Scenario"] not in {"MAR", "MCAR", "MNAR"} or ratio not in allowed_ratios:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "Et metric grid value differs", path=path.as_posix()
            )
        if key in keys or any(not _MEAN_STD_PATTERN.fullmatch(row[column]) for column in _ET_METRIC_COLUMNS):
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "Et metric key or mean +/- std value differs", path=path.as_posix()
            )
        keys.add(key)
    expected_keys = {
        (scenario, ratio)
        for scenario in ("MAR", "MCAR", "MNAR")
        for ratio in allowed_ratios
    }
    if keys != expected_keys:
        raise ComputationalAdmissionError(
            "pue_model_output_invalid", "Et metric grid is incomplete", path=path.as_posix()
        )
    return len(keys)


def _audit_rmse_workbook(path: Path) -> tuple[int, int]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError) as error:
        raise ComputationalAdmissionError(
            "pue_model_output_invalid", "PUE RMSE workbook cannot be loaded", path=path.as_posix()
        ) from error
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "PUE RMSE workbook sheet set differs", path=path.as_posix()
            )
        sheet = workbook["Sheet1"]
        values = list(sheet.iter_rows())
        if len(values) != 181 or len(values[0]) != 4:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "PUE RMSE workbook dimensions differ", path=path.as_posix()
            )
        if tuple(cell.value for cell in values[0]) != ("RMSE", "Method", "Range", "Col"):
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "PUE RMSE workbook header differs", path=path.as_posix()
            )
        if any(cell.data_type == "f" for row in values for cell in row):
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "PUE RMSE workbook contains formulas", path=path.as_posix()
            )
        expected_methods = {"Gain", "HyperImpute", "MatImpute", "Mean", "MissForest", "Sinkhorn"}
        expected_ranges = {"≤1σ", "1~2σ", ">2σ"}
        expected_columns = {
            "Form_Method",
            "ZS_CHS",
            "ZS_HS_BertzCT",
            "ZS_SS_PEOE_VSA8",
            "ZS_SS_TPSA_norm",
            "ZS_SS_VSA_EState8",
            "ZS_log_FCVm",
            "ZS_log_Fchi",
            "ZS_log_StrainRate",
            "ZS_log_Tr2K",
        }
        keys: set[tuple[str, str, str]] = set()
        missing_rmse = 0
        for cells in values[1:]:
            rmse, method, value_range, column = (cell.value for cell in cells)
            key = (str(method), str(value_range), str(column))
            if (
                key in keys
                or method not in expected_methods
                or value_range not in expected_ranges
                or column not in expected_columns
            ):
                raise ComputationalAdmissionError(
                    "pue_model_output_invalid", "PUE RMSE workbook grid differs", path=path.as_posix()
                )
            keys.add(key)
            if rmse is None:
                missing_rmse += 1
            elif _finite_decimal(
                str(rmse), code="pue_model_output_invalid", path=path
            ) < 0:
                raise ComputationalAdmissionError(
                    "pue_model_output_invalid", "PUE RMSE workbook metric is negative", path=path.as_posix()
                )
        expected_keys = {
            (method, value_range, column)
            for method in expected_methods
            for value_range in expected_ranges
            for column in expected_columns
        }
        if keys != expected_keys or missing_rmse != 18:
            raise ComputationalAdmissionError(
                "pue_model_output_invalid", "PUE RMSE workbook grid or blank count differs", path=path.as_posix()
            )
        return len(keys), missing_rmse
    finally:
        workbook.close()


def audit_pue_model_outputs(
    parent: PueParentAudit,
    outputs: Mapping[str, str | Path],
) -> Mapping[str, int]:
    """Verify six PUE benchmark artifacts without treating them as observations."""

    expected_keys = {
        "rdf_ratio",
        "rdf_type",
        "filled_metrics",
        "distance_metrics",
        "rmse_metrics",
        "rmse_workbook",
    }
    if set(outputs) != expected_keys:
        raise ComputationalAdmissionError(
            "pue_model_output_set_mismatch",
            "PUE model outputs must contain the six frozen artifact roles",
        )
    paths = {key: Path(value) for key, value in outputs.items()}
    _audit_npy_output(paths["rdf_ratio"])
    _audit_npy_output(paths["rdf_type"])
    expected_columns = set(parent.columns) - {"SSID"}
    filled_rows = _audit_filled_metrics(paths["filled_metrics"], expected_columns)
    distance_rows = _audit_et_metrics(paths["distance_metrics"])
    rmse_rows = _audit_et_metrics(paths["rmse_metrics"])
    workbook_rows, workbook_missing = _audit_rmse_workbook(paths["rmse_workbook"])
    return {
        "model_output_file_count": len(outputs),
        "rdf_array_file_count": 2,
        "rdf_array_value_count": 224,
        "filled_metric_row_count": filled_rows,
        "et_distance_metric_row_count": distance_rows,
        "et_rmse_metric_row_count": rmse_rows,
        "stratified_rmse_row_count": workbook_rows,
        "stratified_rmse_missing_metric_count": workbook_missing,
        "model_output_material_observation_count": 0,
    }


def _audit_repository_benchmark_csv(path: Path) -> int:
    columns, rows = _dict_rows(path)
    if (
        len(columns) not in {7, 8, 10}
        or columns[:2] != ("Scenario", "miss_pct [0, 1]")
        or len(set(columns)) != len(columns)
    ):
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute benchmark CSV schema differs",
            path=path.as_posix(),
        )
    metric_columns = columns[2:]
    keys: set[tuple[str, Decimal]] = set()
    allowed_ratios = {Decimal(str(index / 10)) for index in range(1, 6)}
    metric_cells = 0
    for row in rows:
        ratio = _finite_decimal(
            row["miss_pct [0, 1]"],
            code="repository_model_output_invalid",
            path=path,
        )
        key = (row["Scenario"], ratio)
        if (
            key in keys
            or row["Scenario"] not in {"MAR", "MCAR", "MNAR"}
            or ratio not in allowed_ratios
            or any(
                not _MEAN_STD_PATTERN.fullmatch(row[column])
                for column in metric_columns
            )
        ):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute benchmark CSV grid or metric differs",
                path=path.as_posix(),
            )
        keys.add(key)
        metric_cells += len(metric_columns)
    expected_keys = {
        (scenario, ratio)
        for scenario in ("MAR", "MCAR", "MNAR")
        for ratio in allowed_ratios
    }
    if keys != expected_keys:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute benchmark CSV grid is incomplete",
            path=path.as_posix(),
        )
    return metric_cells


def _audit_repository_filled_metrics(path: Path) -> int:
    columns, rows = _dict_rows(path)
    expected_header = (
        "dataset",
        "column",
        "miss_ratio",
        "method",
        "RMSE",
        "Wasserstein",
        "time",
    )
    if columns != expected_header:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute filled-results schema differs",
            path=path.as_posix(),
        )
    materialized = list(rows)
    datasets = {row["dataset"] for row in materialized}
    target_columns = {row["column"] for row in materialized}
    if len(datasets) != 1 or "" in datasets or "" in target_columns:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute filled-results dataset or target identity differs",
            path=path.as_posix(),
        )
    keys: set[tuple[str, str, str]] = set()
    for row in materialized:
        ratio_token = next(
            (
                token
                for token in _MISSING_RATIO_TOKENS
                if Decimal(token)
                == _finite_decimal(
                    row["miss_ratio"],
                    code="repository_model_output_invalid",
                    path=path,
                )
            ),
            None,
        )
        key = (row["column"], str(ratio_token), row["method"])
        if (
            ratio_token is None
            or row["method"] not in _FILLED_RESULT_METHODS
            or key in keys
        ):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute filled-results grid value differs",
                path=path.as_posix(),
            )
        keys.add(key)
        for metric in ("RMSE", "Wasserstein", "time"):
            if _finite_decimal(
                row[metric], code="repository_model_output_invalid", path=path
            ) < 0:
                raise ComputationalAdmissionError(
                    "repository_model_output_invalid",
                    "MatImpute filled-results metric is negative",
                    path=path.as_posix(),
                )
    expected_keys = {
        (column, token, method)
        for column in target_columns
        for token in _MISSING_RATIO_TOKENS
        for method in _FILLED_RESULT_METHODS
    }
    if keys != expected_keys:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute filled-results grid is incomplete",
            path=path.as_posix(),
        )
    return len(keys)


def _audit_repository_rmse_workbook(path: Path) -> tuple[int, int]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError) as error:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute RMSE workbook cannot be loaded",
            path=path.as_posix(),
        ) from error
    try:
        if len(workbook.sheetnames) != 1:
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute RMSE workbook must contain one sheet",
                path=path.as_posix(),
            )
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows())
        if not rows or tuple(cell.value for cell in rows[0]) != (
            "RMSE",
            "Method",
            "Range",
            "Col",
        ):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute RMSE workbook header differs",
                path=path.as_posix(),
            )
        if any(cell.data_type == "f" for row in rows for cell in row):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute RMSE workbook contains formulas",
                path=path.as_posix(),
            )
        values = [tuple(cell.value for cell in row) for row in rows[1:]]
        methods = {str(row[1]) for row in values}
        ranges = {str(row[2]) for row in values}
        target_columns = {str(row[3]) for row in values}
        keys: set[tuple[str, str, str]] = set()
        missing = 0
        finite = 0
        for rmse, method, value_range, column in values:
            key = (str(method), str(value_range), str(column))
            if key in keys or None in (method, value_range, column):
                raise ComputationalAdmissionError(
                    "repository_model_output_invalid",
                    "MatImpute RMSE workbook key differs",
                    path=path.as_posix(),
                )
            keys.add(key)
            if rmse is None:
                missing += 1
            else:
                if _finite_decimal(
                    str(rmse), code="repository_model_output_invalid", path=path
                ) < 0:
                    raise ComputationalAdmissionError(
                        "repository_model_output_invalid",
                        "MatImpute RMSE workbook metric is negative",
                        path=path.as_posix(),
                    )
                finite += 1
        expected_keys = {
            (method, value_range, column)
            for method in methods
            for value_range in ranges
            for column in target_columns
        }
        if (
            methods
            != {"Gain", "HyperImpute", "MatImpute", "Mean", "MissForest", "Sinkhorn"}
            or ranges != {"≤1σ", "1~2σ", ">2σ"}
            or not target_columns
            or keys != expected_keys
        ):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute RMSE workbook grid is incomplete",
                path=path.as_posix(),
            )
        return finite, missing
    finally:
        workbook.close()


def _audit_png(path: Path) -> tuple[int, int]:
    try:
        payload = path.read_bytes()
    except OSError as error:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute PNG cannot be read",
            path=path.as_posix(),
        ) from error
    if len(payload) < 45 or payload[:8] != b"\x89PNG\r\n\x1a\n":
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute image is not a complete PNG",
            path=path.as_posix(),
        )
    offset = 8
    width = 0
    height = 0
    chunk_index = 0
    saw_idat = False
    saw_iend = False
    while offset < len(payload):
        if offset + 12 > len(payload):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute PNG chunk is truncated",
                path=path.as_posix(),
            )
        length = int.from_bytes(payload[offset : offset + 4], "big")
        chunk_type = payload[offset + 4 : offset + 8]
        data_start = offset + 8
        data_end = data_start + length
        crc_end = data_end + 4
        if crc_end > len(payload):
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute PNG chunk length exceeds the file",
                path=path.as_posix(),
            )
        stored_crc = int.from_bytes(payload[data_end:crc_end], "big")
        if zlib.crc32(chunk_type + payload[data_start:data_end]) & 0xFFFFFFFF != stored_crc:
            raise ComputationalAdmissionError(
                "repository_model_output_invalid",
                "MatImpute PNG chunk CRC differs",
                path=path.as_posix(),
            )
        if chunk_index == 0:
            if chunk_type != b"IHDR" or length != 13:
                raise ComputationalAdmissionError(
                    "repository_model_output_invalid",
                    "MatImpute PNG does not begin with a valid IHDR",
                    path=path.as_posix(),
                )
            width = int.from_bytes(payload[data_start : data_start + 4], "big")
            height = int.from_bytes(payload[data_start + 4 : data_start + 8], "big")
        elif chunk_type == b"IDAT":
            saw_idat = True
        elif chunk_type == b"IEND":
            if length != 0 or crc_end != len(payload):
                raise ComputationalAdmissionError(
                    "repository_model_output_invalid",
                    "MatImpute PNG IEND is malformed or not final",
                    path=path.as_posix(),
                )
            saw_iend = True
            offset = crc_end
            break
        chunk_index += 1
        offset = crc_end
    if width <= 0 or height <= 0 or not saw_idat or not saw_iend or offset != len(payload):
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute PNG dimensions or required chunks are invalid",
            path=path.as_posix(),
        )
    return width, height


def audit_matimpute_repository_model_outputs(
    model_output_paths: Sequence[str | Path],
) -> Mapping[str, int]:
    """Content-check all 61 MatImpute model-output assets by frozen class."""

    paths = tuple(Path(path) for path in model_output_paths)
    normalized = [path.resolve(strict=False) for path in paths]
    if len(paths) != len(set(normalized)):
        raise ComputationalAdmissionError(
            "repository_model_output_set_mismatch",
            "MatImpute model-output paths must be unique",
        )
    classes: defaultdict[str, list[Path]] = defaultdict(list)
    for path in paths:
        suffix = path.suffix.casefold()
        if path.parent.name == "filled_results" and suffix in {".csv", ".png"}:
            classes[f"filled_{suffix[1:]}"].append(path)
        elif path.parent.name == "experiment" and suffix in {".npy", ".csv", ".xlsx", ".png"}:
            classes[f"root_{suffix[1:]}"] .append(path)
        else:
            raise ComputationalAdmissionError(
                "repository_model_output_set_mismatch",
                "MatImpute model-output path is outside the frozen classes",
                path=path.as_posix(),
            )
    expected_counts = {
        "root_npy": 6,
        "root_csv": 20,
        "root_xlsx": 8,
        "root_png": 1,
        "filled_csv": 11,
        "filled_png": 15,
    }
    if {key: len(value) for key, value in classes.items()} != expected_counts:
        raise ComputationalAdmissionError(
            "repository_model_output_set_mismatch",
            "MatImpute model-output class counts differ from the frozen 61-file inventory",
        )
    try:
        import numpy as np

        shape_counts: Counter[tuple[int, ...]] = Counter()
        for path in classes["root_npy"]:
            array = np.load(path, allow_pickle=False)
            if (
                array.dtype != np.dtype("float64")
                or array.ndim != 2
                or array.shape[0] != 4
                or not np.isfinite(array).all()
                or (array[:3] < 0).any()
                or not (np.diff(array[3]) > 0).all()
            ):
                raise ComputationalAdmissionError(
                    "repository_model_output_invalid",
                    "MatImpute repository RDF array differs",
                    path=path.as_posix(),
                )
            shape_counts[tuple(array.shape)] += 1
    except ComputationalAdmissionError:
        raise
    except (OSError, ValueError) as error:
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute repository RDF array cannot be loaded",
        ) from error
    if shape_counts != Counter({(4, 28): 2, (4, 27): 2, (4, 25): 2}):
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute repository RDF shape inventory differs",
        )
    benchmark_metric_cells = sum(
        _audit_repository_benchmark_csv(path) for path in classes["root_csv"]
    )
    workbook_finite = 0
    workbook_missing = 0
    workbook_rows = 0
    for path in classes["root_xlsx"]:
        finite, missing = _audit_repository_rmse_workbook(path)
        workbook_finite += finite
        workbook_missing += missing
        workbook_rows += finite + missing
    filled_rows = sum(
        _audit_repository_filled_metrics(path) for path in classes["filled_csv"]
    )
    png_dimensions = [
        _audit_png(path)
        for key in ("root_png", "filled_png")
        for path in classes[key]
    ]
    if (
        benchmark_metric_cells != 1800
        or workbook_rows != 882
        or workbook_finite != 786
        or workbook_missing != 96
        or filled_rows != 8316
        or len(png_dimensions) != 16
    ):
        raise ComputationalAdmissionError(
            "repository_model_output_invalid",
            "MatImpute repository output aggregate counts differ",
        )
    return {
        "repository_model_output_file_count": len(paths),
        "repository_rdf_array_file_count": 6,
        "repository_benchmark_csv_file_count": 20,
        "repository_benchmark_metric_cell_count": benchmark_metric_cells,
        "repository_rmse_workbook_file_count": 8,
        "repository_rmse_workbook_row_count": workbook_rows,
        "repository_rmse_finite_metric_count": workbook_finite,
        "repository_rmse_missing_metric_count": workbook_missing,
        "repository_root_png_file_count": 1,
        "repository_filled_metric_csv_file_count": 11,
        "repository_filled_metric_row_count": filled_rows,
        "repository_filled_png_file_count": 15,
        "repository_model_output_material_observation_count": 0,
    }


def profile_dq_matimpute(
    dq_parent_path: str | Path,
    matimpute_parent_path: str | Path,
    *,
    dq_projection_paths: Mapping[str, str | Path],
    missing_variants_directory: str | Path,
    model_output_paths: Mapping[str, str | Path],
    repository_model_output_paths: Sequence[str | Path],
    expected_parent_row_count: int = 326,
    expected_parent_column_count: int = 24,
) -> ComputationalAdmissionProfile:
    """Content-verify one PUE parent plus projections, missingness, and outputs."""

    parent = audit_pue_parent_occurrences(
        dq_parent_path,
        matimpute_parent_path,
        expected_row_count=expected_parent_row_count,
        expected_column_count=expected_parent_column_count,
    )
    projection_diagnostics = audit_dq_pue_projections(parent, dq_projection_paths)
    missing_diagnostics = audit_pue_missing_variants(
        parent, missing_variants_directory
    )
    model_diagnostics = audit_pue_model_outputs(parent, model_output_paths)
    repository_model_diagnostics = audit_matimpute_repository_model_outputs(
        repository_model_output_paths
    )
    derived_count = (
        projection_diagnostics["dq_projection_file_count"]
        + missing_diagnostics["missing_variant_file_count"]
    )
    model_count = model_diagnostics["model_output_file_count"]
    diagnostics: dict[str, int | str] = {
        "parent_file_occurrence_count": 2,
        "canonical_parent_dataset_count": 1,
        "parent_source_row_count": len(parent.rows),
        "parent_column_count": len(parent.columns),
        "parent_unique_ssid_count": len(parent.rows),
        "parent_missing_numeric_cell_count": 0,
        "parent_nonfinite_numeric_cell_count": 0,
        "parent_content_sha256": parent.content_sha256,
        "profile_file_scope": "PUE lineage files only; all repository model outputs audited separately",
        "derived_container_file_count": derived_count,
        "pue_model_output_file_count": model_count,
        "repository_non_pue_model_output_file_count": (
            repository_model_diagnostics["repository_model_output_file_count"]
            - model_count
        ),
        **projection_diagnostics,
        **missing_diagnostics,
        **model_diagnostics,
        **repository_model_diagnostics,
    }
    return ComputationalAdmissionProfile(
        source_key="dq_matimpute",
        evidence_class="experimental_parent_with_verified_synthetic_missingness_and_ml_outputs",
        file_count=2 + derived_count + model_count,
        source_record_candidate_count=len(parent.rows),
        unique_system_candidate_count=None,
        computational_activity_candidate_count=0,
        computational_observation_candidate_count=0,
        diagnostics=diagnostics,
        blocking_reason=(
            "326行母数据只有一个来源基数且无法恢复原始配方；2个投影、207个缺失变体和6个PUE评估输出均已验明为派生容器；MatImpute全仓61个模型输出均为数组、聚合指标或图件，不得成为新的实验或计算观测"
        ),
    )


def _exact_values(path: Path, column: str) -> set[str]:
    columns, rows = _dict_rows(path)
    _require_columns(columns, [column], path)
    values: set[str] = set()
    for row in rows:
        value = row[column].strip()
        if _identity_is_missing(value):
            raise ComputationalAdmissionError(
                "missing_identity",
                f"{path.name} contains a missing {column} identity",
                path=path.as_posix(),
            )
        values.add(value)
    return values


def profile_exact_structure_overlaps(
    pi1m_path: str | Path,
    adept_path: str | Path,
    polyomics_path: str | Path,
    polygraphmt_directory: str | Path,
) -> ExactStructureOverlapProfile:
    """Recompute exact-string overlap lower bounds without chemical canonicalization."""

    pi1m = _exact_values(Path(pi1m_path), "SMILES")
    polyomics = _exact_values(Path(polyomics_path), "smiles_list")

    adept_source = Path(adept_path)
    columns, rows = _dict_rows(adept_source)
    _require_columns(columns, ["PID", "SMILES"], adept_source)
    adept_links: defaultdict[str, set[str]] = defaultdict(set)
    seen_pids: set[str] = set()
    for row in rows:
        pid = row["PID"].strip()
        structure = row["SMILES"].strip()
        if _identity_is_missing(pid) or _identity_is_missing(structure):
            raise ComputationalAdmissionError(
                "missing_identity", "ADEPT overlap source lacks PID or SMILES", path=adept_source.as_posix()
            )
        if pid in seen_pids:
            raise ComputationalAdmissionError(
                "duplicate_source_record_id", "ADEPT overlap source has duplicate PID", path=adept_source.as_posix()
            )
        seen_pids.add(pid)
        adept_links[structure].add(pid)
    adept = set(adept_links)

    polygraph_root = Path(polygraphmt_directory)
    polygraph: set[str] = set()
    polygraph_invalid_identity = 0
    files = sorted(polygraph_root.glob("*.csv"), key=lambda path: (path.name.casefold(), path.name))
    if not files:
        raise ComputationalAdmissionError(
            "no_candidate_files", "PolyGraphMT overlap source has no CSV files", path=polygraph_root.as_posix()
        )
    for path in files:
        columns, rows = _dict_rows(path)
        _require_columns(columns, ["SMILES"], path)
        for row in rows:
            structure = row["SMILES"].strip()
            if not structure:
                raise ComputationalAdmissionError(
                    "missing_identity", "PolyGraphMT overlap source has blank SMILES", path=path.as_posix()
                )
            if _identity_is_missing(structure):
                polygraph_invalid_identity += 1
                continue
            polygraph.add(structure)
    outside = polygraph - adept
    if outside:
        raise ComputationalAdmissionError(
            "polygraphmt_not_adept_subset",
            "PolyGraphMT exact structures are not all contained in ADEPT",
            path=polygraph_root.as_posix(),
        )
    sources = {
        "pi1m": pi1m,
        "adept": adept,
        "polyomics": polyomics,
        "polygraphmt": polygraph,
    }
    pairs = (
        ("pi1m", "adept"),
        ("pi1m", "polyomics"),
        ("pi1m", "polygraphmt"),
        ("adept", "polyomics"),
        ("adept", "polygraphmt"),
        ("polyomics", "polygraphmt"),
    )
    return ExactStructureOverlapProfile(
        source_exact_structure_counts={name: len(values) for name, values in sources.items()},
        pair_overlap_counts={
            f"{left}__{right}": len(sources[left] & sources[right])
            for left, right in pairs
        },
        diagnostics={
            "identity_basis": "trimmed, case-sensitive exact structure string; lower bound only",
            "polygraphmt_invalid_identity_record_count": polygraph_invalid_identity,
            "polygraphmt_outside_adept_count": len(outside),
            "adept_multi_pid_structure_count": sum(len(pids) > 1 for pids in adept_links.values()),
            "adept_extra_pid_link_count": sum(len(pids) - 1 for pids in adept_links.values()),
            "split_requirement": "group by study/source family plus canonical structure",
        },
    )


def _count_text(value: int | None) -> str:
    return "不可由现有证据确定" if value is None else f"{value:,}"


def render_computational_admission_markdown(
    profiles: Iterable[ComputationalAdmissionProfile],
    *,
    overlap_profile: ExactStructureOverlapProfile | None = None,
    ledger_link: str = "../TPU_数据来源与研究路线台账.md",
) -> str:
    """Render a deterministic, citation-ready provisional admission report."""

    if not isinstance(ledger_link, str) or not ledger_link.strip():
        raise ComputationalAdmissionError(
            "invalid_ledger_link", "master-ledger link must be a non-empty string"
        )
    ordered = sorted(profiles, key=lambda profile: profile.source_key)
    admitted_total = sum(profile.admitted_observation_count for profile in ordered)
    lines = [
        "# TPU 数据库 v0.2 计算数据准入报告",
        "",
        "> 状态：`provisional_profile_only`。本报告不构成训练集、冻结数据库或模型就绪声明。",
        "",
        "## 计数口径",
        "",
        "CSV 行数不等于独立实验样本数。来源记录、唯一体系、计算活动、计算观测和已准入观测分别计数；镜像、子集视图、缺失率变体与模型输出容器不得虚增独立配方。",
        "",
        f"- 已准入计算观测 | {admitted_total}",
        "- 本阶段训练/验证划分 | 未创建",
        "- 本阶段模型权重 | 未设置",
        "",
        "## 文件与候选画像",
        "",
        "| 来源 | 证据类别 | 文件 | 来源记录 occurrence | 精确结构串候选（不等于完整体系） | 计算活动候选（可空） | 数值性质 occurrence 候选 | 已准入计算观测 | 引用 |",
        "|---|---|---:|---:|---:|---:|---:|---:|---|",
    ]
    for profile in ordered:
        lines.append(
            "| "
            + " | ".join(
                (
                    f"`{profile.source_key}`",
                    profile.evidence_class,
                    f"{profile.file_count:,}",
                    _count_text(profile.source_record_candidate_count),
                    _count_text(profile.unique_system_candidate_count),
                    _count_text(profile.computational_activity_candidate_count),
                    _count_text(profile.computational_observation_candidate_count),
                    f"{profile.admitted_observation_count:,}",
                    _REFERENCE_BY_SOURCE.get(profile.source_key, "见主台账"),
                )
            )
            + " |"
        )
    lines.extend(["", "## 阻断条件", ""])
    for profile in ordered:
        lines.append(f"- `{profile.source_key}`：{profile.blocking_reason}")
    lines.extend(["", "## 可复算诊断", ""])
    for profile in ordered:
        lines.extend(
            [
                f"### `{profile.source_key}`",
                "",
                "| 指标 | 值 |",
                "|---|---:|",
            ]
        )
        for key, value in sorted(profile.diagnostics.items()):
            rendered = f"{value:,}" if isinstance(value, int) and not isinstance(value, bool) else str(value)
            lines.append(f"| `{key}` | {rendered} |")
        if not profile.diagnostics:
            lines.append("| `none` | 0 |")
        lines.append("")
    if overlap_profile is not None:
        source_labels = {
            "pi1m": "PI1M",
            "adept": "ADEPT",
            "polyomics": "PolyOmics",
            "polygraphmt": "PolyGraphMT",
        }
        lines.extend(
            [
                "## 跨库精确结构重叠（泄漏控制下界）",
                "",
                "> 这里使用区分大小写、未经化学标准化的原始结构字符串；所得重叠只能作为下界，不能替代后续结构规范化与图同构去重。",
                "",
                "| 来源 | 唯一精确结构串 |",
                "|---|---:|",
            ]
        )
        for source_key in ("pi1m", "adept", "polyomics", "polygraphmt"):
            lines.append(
                f"| {source_labels[source_key]} | "
                f"{overlap_profile.source_exact_structure_counts[source_key]:,} |"
            )
        lines.extend(["", "| 来源对 | 精确重叠结构串 |", "|---|---:|"])
        for pair in (
            "pi1m__adept",
            "pi1m__polyomics",
            "pi1m__polygraphmt",
            "adept__polyomics",
            "adept__polygraphmt",
            "polyomics__polygraphmt",
        ):
            left, right = pair.split("__", 1)
            lines.append(
                f"| {source_labels[left]} ↔ {source_labels[right]} | "
                f"{overlap_profile.pair_overlap_counts[pair]:,} |"
            )
        lines.extend(["", "### 重叠诊断", "", "| 指标 | 值 |", "|---|---:|"])
        for key, value in sorted(overlap_profile.diagnostics.items()):
            rendered = f"{value:,}" if isinstance(value, int) and not isinstance(value, bool) else str(value)
            lines.append(f"| `{key}` | {rendered} |")
        lines.extend(
            [
                "",
                "训练拆分必须在同一研究/来源家族与规范化结构联合分组后进行；禁止按文件行随机拆分。",
                "",
            ]
        )
    lines.extend(
        [
            "",
            "## 参考文献",
            "",
            f"以下编号与[数据来源与研究路线主台账]({ledger_link})一致。",
            "",
            "[1] Ding, F.; Liu, L.-Y.; Liu, T.-L.; Li, Y.-Q.; Li, J.-P.; Sun, Z.-Y. Predicting the Mechanical Properties of Polyurethane Elastomers Using Machine Learning. *Chinese Journal of Polymer Science* **2023**, *41*, 422–431. https://doi.org/10.1007/s10118-022-2838-6.",
            "",
            "[6] Ma, R.; Luo, T. PI1M: A Benchmark Database for Polymer Informatics. *Journal of Chemical Information and Modeling* **2020**, *60*, 4684–4690. https://doi.org/10.1021/acs.jcim.0c00726.",
            "",
            "[7] Ohno, M.; Hayashi, Y.; Zhang, Q.; Kaneko, Y.; Yoshida, R. SMiPoly: Generation of a Synthesizable Polymer Virtual Library Using Rule-Based Polymerization Reactions. *Journal of Chemical Information and Modeling* **2023**, *63*, 5539–5548. https://doi.org/10.1021/acs.jcim.3c00329.",
            "",
            "[8] Alosious, S.; Liu, Y.; Xu, J.; Liu, G.; Zhang, R.; Jiang, M.; Luo, T. ADEPT-PolyGraphMT: Automated Molecular Simulation and Multi-Task Multi-Fidelity Machine Learning for Polymer Property Generation and Prediction. *Digital Discovery* **2026**, advance article. https://doi.org/10.1039/D6DD00206D.",
            "",
            "[12] Xie, C.; Li, R.; Li, Y.; Xie, H.; Liu, Q. Imputation of Missing Data in Materials Science through Nearest Neighbors and Iterative Predictions. *Journal of Chemical Theory and Computation* **2025**, *21*, 70–78. https://doi.org/10.1021/acs.jctc.4c01237.",
            "",
            "[20] Yoshida, R.; Hayashi, Y.; Furuya, H.; Hosoya, R.; Kaneko, K.; Sugisawa, H.; Kaneko, Y.; Takahashi, A.; Noguchi, Y.; Nanjo, S.; et al. Omics-Scale Polymer Computational Database Transferable to Real-World Artificial Intelligence Applications. *arXiv* **2025**, arXiv:2511.11626. https://doi.org/10.48550/arXiv.2511.11626. Dataset: https://doi.org/10.57967/hf/7475.",
            "",
        ]
    )
    return "\n".join(lines)
