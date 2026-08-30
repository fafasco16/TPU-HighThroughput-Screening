"""物化Data in Brief交联形状记忆PU的拉伸、循环代理与热端点。"""

from __future__ import annotations

import argparse
import functools
import hashlib
import json
import math
import statistics
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "DataInBrief_聚氨酯形状记忆多模态原始数据"
)
CURVE_AUDIT = SOURCE_DIR / "曲线审计清单.tsv"
FORMULATION_AUDIT = SOURCE_DIR / "配方审计清单.tsv"
SOURCE_AUDIT = SOURCE_DIR / "内容审计摘要.json"
SOURCE_METADATA = SOURCE_DIR / "来源元数据.json"
FAILURE_BOOK = SOURCE_DIR / "mmc4.xlsx"
CYCLIC_BOOK = SOURCE_DIR / "mmc6.xlsx"
TGA_SOURCE = SOURCE_DIR / "mmc5.txt"
TAN_DELTA_SOURCE = SOURCE_DIR / "mmc2.txt"
DIRECTED = ROOT / "结果" / "定向筛选"
TENSILE = DIRECTED / "DataInBrief形状记忆PU拉伸端点.csv"
CYCLIC = DIRECTED / "DataInBrief形状记忆PU循环端点.csv"
THERMAL = DIRECTED / "DataInBrief形状记忆PU热稳定端点.csv"
MANIFEST = DIRECTED / "DataInBrief形状记忆PU发布清单.json"
RELEASE_ID = "pu-shape-memory-dib-2020-v1"
DATASET_DOI = "10.1016/j.dib.2020.106294"
COMPANION_DOI = "10.1016/j.jmbbm.2018.08.037"
LICENSE = "CC-BY-4.0"
CITATIONS = "reference-184;reference-185"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _entry(path: Path) -> dict[str, object]:
    return {
        "path": path.relative_to(ROOT).as_posix(),
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def _finite(value: object) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _read_text_triplets(path: Path) -> pd.DataFrame:
    records = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        fields = line.split()
        if len(fields) < 3:
            continue
        values = [_finite(field) for field in fields[:3]]
        if all(value is not None for value in values):
            records.append(values)
    return pd.DataFrame(records, columns=["formula_number", "x", "y"])


def _load_formulations() -> dict[str, dict[str, object]]:
    frame = pd.read_csv(FORMULATION_AUDIT, sep="\t")
    result = {}
    for row in frame.itertuples(index=False):
        result[row.formulation_id] = {
            "polymer_family": row.material_family,
            "HDI_mol_percent": float(row.component_1_fraction),
            "HPED_mol_percent": float(row.component_2_fraction),
            "TEA_mol_percent": float(row.component_3_fraction),
            "fraction_basis": row.fraction_basis,
            "chemistry_mapping_status": "monomer_set_molar_composition_mapped",
            "identity_mapping_evidence": row.evidence,
            "split_group": row.split_group,
        }
    return result


def _common(
    formulation: str,
    source_file: Path,
    formulations: dict[str, dict[str, object]],
    weight_ceiling: float,
) -> dict[str, object]:
    return {
        "release_id": RELEASE_ID,
        "source_id": f"doi:{DATASET_DOI}",
        "formulation_id": formulation,
        **formulations[formulation],
        "HDI_name": "1,6-hexamethylene diisocyanate",
        "HDI_smiles": "O=C=NCCCCCCN=C=O",
        "HPED_name": "N,N,N',N'-tetrakis(2-hydroxypropyl)ethylenediamine",
        "TEA_name": "triethanolamine",
        "TEA_smiles": "OCCN(CCO)CCO",
        "thermoplastic_tpu_core": False,
        "model_admission_layer": "polyurethane_transfer",
        "usage_mode": "transfer_learning_only",
        "future_weight_ceiling": weight_ceiling,
        "source_locator": source_file.relative_to(ROOT).as_posix(),
        "license": LICENSE,
        "citation_keys": CITATIONS,
    }


def _path_area(points: list[tuple[float, float]], direction: str) -> float:
    area = 0.0
    for (strain0, stress0), (strain1, stress1) in zip(
        points, points[1:], strict=False
    ):
        delta = strain1 - strain0
        if (direction == "positive" and delta > 0) or (
            direction == "negative" and delta < 0
        ):
            area += abs(delta) * (max(stress0, 0.0) + max(stress1, 0.0)) / 2
    return area


def _sheet_points(
    worksheet: object, maximum_points: int | None = None
) -> list[tuple[float, float, float]]:
    points = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        cycle = _finite(row[1] if len(row) > 1 else None)
        stress = _finite(row[4] if len(row) > 4 else None)
        strain = _finite(row[5] if len(row) > 5 else None)
        if cycle is None or stress is None or strain is None:
            continue
        points.append((cycle, stress, strain))
        if maximum_points is not None and len(points) >= maximum_points:
            break
    return points


def _crossing_temperature(
    points: list[tuple[float, float]], threshold: float
) -> float:
    for (temperature0, weight0), (temperature1, weight1) in zip(
        points, points[1:], strict=False
    ):
        if weight0 > threshold >= weight1 and weight0 != weight1:
            return temperature0 + (threshold - weight0) * (
                temperature1 - temperature0
            ) / (weight1 - weight0)
    raise ValueError(f"TGA曲线未跨越{threshold}%")


def _build_thermal(
    formulations: dict[str, dict[str, object]],
) -> tuple[pd.DataFrame, dict[str, float]]:
    tga = _read_text_triplets(TGA_SOURCE)
    tan_delta = _read_text_triplets(TAN_DELTA_SOURCE)
    records = []
    tg_map = {}
    for formula_number in range(1, 13):
        formulation = f"SMP-{formula_number}"
        tga_rows = tga.loc[tga["formula_number"].eq(formula_number), ["x", "y"]]
        raw_points = list(tga_rows.itertuples(index=False, name=None))
        baseline = statistics.median(weight for _, weight in raw_points[:50])
        normalised = [
            (temperature, 100.0 * weight / baseline)
            for temperature, weight in raw_points
            if temperature >= 100
        ]
        tan_rows = tan_delta.loc[
            tan_delta["formula_number"].eq(formula_number), ["x", "y"]
        ]
        tan_points = list(tan_rows.itertuples(index=False, name=None))
        tg_temperature, tan_peak = max(tan_points, key=lambda point: point[1])
        tg_map[formulation] = tg_temperature
        records.append(
            {
                **_common(formulation, TGA_SOURCE, formulations, 0.20),
                "T5_C": _crossing_temperature(normalised, 95.0),
                "T10_C": _crossing_temperature(normalised, 90.0),
                "T50_C": _crossing_temperature(normalised, 50.0),
                "Tg_DMA_C": tg_temperature,
                "tan_delta_peak": tan_peak,
                "TGA_baseline_weight_percent": baseline,
                "TGA_point_count": len(raw_points),
                "TGA_atmosphere": "nitrogen",
                "TGA_heating_rate_C_min": 10.0,
                "TGA_temperature_range_C": "31-600",
                "Tg_method": "temperature_at_maximum_tan_delta",
            }
        )
    return pd.DataFrame(records), tg_map


def _build_tensile(
    audit: pd.DataFrame, formulations: dict[str, dict[str, object]]
) -> pd.DataFrame:
    selected = audit.loc[audit["source_file"].eq("mmc4.xlsx")]
    workbook = load_workbook(FAILURE_BOOK, read_only=True, data_only=True)
    records = []
    try:
        for row in selected.itertuples(index=False):
            usable = int(row.usable_point_count)
            points = _sheet_points(workbook[row.source_location], usable)
            if len(points) != usable:
                raise ValueError(f"{row.record_id}可用点数与审计不一致")
            stress_strain = [
                (strain, stress) for _, stress, strain in points if strain >= 0
            ]
            records.append(
                {
                    **_common(
                        row.formulation_id,
                        FAILURE_BOOK,
                        formulations,
                        float(row.future_weight_ceiling),
                    ),
                    "test_run_id": row.record_id,
                    "source_sheet": row.source_location,
                    "observation_unit": "test_run_not_proven_unique_specimen",
                    "tensile_strength_MPa": max(
                        stress for _, stress in stress_strain
                    ),
                    "elongation_at_break_percent": 100.0
                    * max(strain for strain, _ in stress_strain),
                    "toughness_MJ_m3": _path_area(stress_strain, "positive"),
                    "source_point_count": int(row.point_count),
                    "used_point_count": usable,
                    "excluded_tail_contamination_point_count": int(
                        row.contamination_point_count
                    ),
                    "quality_status": row.quality_status,
                    "sample_mapping_status": row.sample_mapping_status,
                    "test_temperature_rule": "Tg_plus_10_C",
                    "crosshead_speed_mm_min": 2.0,
                    "toughness_method": "positive_strain_increment_trapezoid",
                }
            )
    finally:
        workbook.close()
    return pd.DataFrame(records)


def _build_cyclic(
    audit: pd.DataFrame,
    formulations: dict[str, dict[str, object]],
    tg_map: dict[str, float],
) -> pd.DataFrame:
    selected = audit.loc[audit["source_file"].eq("mmc6.xlsx")]
    workbook = load_workbook(CYCLIC_BOOK, read_only=True, data_only=True)
    records = []
    try:
        for row in selected.itertuples(index=False):
            points = _sheet_points(workbook[row.source_location])
            if len(points) != int(row.usable_point_count):
                raise ValueError(f"{row.record_id}循环点数与审计不一致")
            cycle_payloads = []
            for cycle_number, raw_cycle in enumerate(range(3, 13), start=1):
                loading = [
                    (strain, stress)
                    for cycle, stress, strain in points
                    if cycle == float(raw_cycle)
                ]
                unloading = [
                    (strain, stress)
                    for cycle, stress, strain in points
                    if cycle == raw_cycle + 0.5
                ]
                if not loading or not unloading:
                    raise ValueError(
                        f"{row.record_id}缺少测量循环{cycle_number}加载或卸载段"
                    )
                peak_stress = max(stress for _, stress in loading + unloading)
                loading_energy = _path_area(loading, "positive")
                unloading_energy = _path_area(unloading, "negative")
                cycle_payloads.append(
                    {
                        "cycle_number": cycle_number,
                        "peak_stress_MPa": peak_stress,
                        "maximum_strain_percent": 100.0
                        * max(strain for strain, _ in loading + unloading),
                        "loading_energy_MJ_m3": loading_energy,
                        "unloading_energy_MJ_m3": unloading_energy,
                        "hysteresis_energy_MJ_m3": max(
                            loading_energy - unloading_energy, 0.0
                        ),
                    }
                )
            first_peak = cycle_payloads[0]["peak_stress_MPa"]
            for payload in cycle_payloads:
                retention = 100.0 * payload["peak_stress_MPa"] / first_peak
                records.append(
                    {
                        **_common(
                            row.formulation_id,
                            CYCLIC_BOOK,
                            formulations,
                            float(row.future_weight_ceiling),
                        ),
                        "test_run_id": row.record_id,
                        "source_sheet": row.source_location,
                        "observation_unit": "test_run_cycle_dependent",
                        **payload,
                        "peak_stress_retention_percent": retention,
                        "peak_stress_reduction_percent": 100.0 - retention,
                        "dissipation_fraction": (
                            payload["hysteresis_energy_MJ_m3"]
                            / payload["loading_energy_MJ_m3"]
                            if payload["loading_energy_MJ_m3"] > 0
                            else math.nan
                        ),
                        "preconditioning_cycle_count": 3,
                        "measurement_cycle_count": 10,
                        "measurement_strain_fraction_of_failure": 0.50,
                        "test_temperature_rule": "Tg_plus_10_C",
                        "nominal_test_temperature_C": tg_map[row.formulation_id]
                        + 10.0,
                        "crosshead_speed_mm_min": 2.0,
                        "direct_shape_recovery_available": False,
                        "shape_recovery_ratio_percent": math.nan,
                        "shape_fixity_ratio_percent": math.nan,
                        "cyclic_target_role": (
                            "stress_retention_and_hysteresis_transfer_proxy"
                        ),
                        "recovery_metric_status": (
                            "not_measured_by_this_isothermal_cyclic_protocol"
                        ),
                        "source_run_point_count": int(row.point_count),
                        "quality_status": row.quality_status,
                        "sample_mapping_status": row.sample_mapping_status,
                    }
                )
    finally:
        workbook.close()
    return pd.DataFrame(records)


@functools.lru_cache(maxsize=1)
def _build_cached() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    audit = pd.read_csv(CURVE_AUDIT, sep="\t")
    formulations = _load_formulations()
    thermal, tg_map = _build_thermal(formulations)
    tensile = _build_tensile(audit, formulations)
    cyclic = _build_cyclic(audit, formulations, tg_map)
    return tensile, cyclic, thermal


def build_release() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    return tuple(frame.copy() for frame in _build_cached())


def _write_frames(
    tensile: pd.DataFrame,
    cyclic: pd.DataFrame,
    thermal: pd.DataFrame,
    directory: Path,
) -> dict[str, Path]:
    paths = {
        "tensile": directory / TENSILE.name,
        "cyclic": directory / CYCLIC.name,
        "thermal": directory / THERMAL.name,
    }
    tensile.to_csv(
        paths["tensile"], index=False, encoding="utf-8-sig", lineterminator="\n"
    )
    cyclic.to_csv(
        paths["cyclic"], index=False, encoding="utf-8-sig", lineterminator="\n"
    )
    thermal.to_csv(
        paths["thermal"], index=False, encoding="utf-8-sig", lineterminator="\n"
    )
    return paths


def write_release(
    tensile: pd.DataFrame, cyclic: pd.DataFrame, thermal: pd.DataFrame
) -> None:
    DIRECTED.mkdir(parents=True, exist_ok=True)
    paths = _write_frames(tensile, cyclic, thermal, DIRECTED)
    source_audit = json.loads(SOURCE_AUDIT.read_text(encoding="utf-8"))
    payload = {
        "release_id": RELEASE_ID,
        "counts": {
            "formulation_count": 12,
            "failure_test_run_count": len(tensile),
            "cyclic_test_run_count": int(cyclic["test_run_id"].nunique()),
            "measurement_cycle_endpoint_count": len(cyclic),
            "tga_curve_count": len(thermal),
            "raw_source_point_rows": int(source_audit["point_row_count"]),
            "published_endpoint_rows": len(tensile) + len(cyclic) + len(thermal),
        },
        "counting_note": (
            "975903是来源曲线点数；发布层只保留37次失效测试、24次循环测试的"
            "240个依赖循环端点和12条TGA端点。测试运行不自动等于独立物理试样。"
        ),
        "source": {
            "dataset_doi": DATASET_DOI,
            "companion_article_doi": COMPANION_DOI,
            "license": LICENSE,
            "inputs": [
                _entry(path)
                for path in (
                    SOURCE_AUDIT,
                    SOURCE_METADATA,
                    CURVE_AUDIT,
                    FORMULATION_AUDIT,
                    FAILURE_BOOK,
                    CYCLIC_BOOK,
                    TGA_SOURCE,
                    TAN_DELTA_SOURCE,
                )
            ],
        },
        "policy": {
            "thermoplastic_tpu_core": False,
            "model_admission_layer": "polyurethane_transfer",
            "direct_shape_recovery_available": False,
            "cycle_role": "stress_retention_and_hysteresis_transfer_proxy",
            "curves_not_republished_reason": (
                "用确定性端点代替975903行曲线长表，原始文件由SHA256和来源定位复现"
            ),
        },
        "outputs": {key: _entry(path) for key, path in paths.items()},
    }
    MANIFEST.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def check_release(
    tensile: pd.DataFrame, cyclic: pd.DataFrame, thermal: pd.DataFrame
) -> None:
    with tempfile.TemporaryDirectory(prefix="dib-shape-memory-check-") as directory:
        temporary = _write_frames(tensile, cyclic, thermal, Path(directory))
        published = {"tensile": TENSILE, "cyclic": CYCLIC, "thermal": THERMAL}
        mismatches = [
            key
            for key in published
            if _sha256(temporary[key]) != _sha256(published[key])
        ]
        if mismatches:
            raise SystemExit(f"DataInBrief形状记忆PU输出不一致: {','.join(mismatches)}")
    print("DataInBrief形状记忆PU数据检查通过")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--检查", action="store_true")
    args = parser.parse_args()
    tensile, cyclic, thermal = build_release()
    if args.检查:
        check_release(tensile, cyclic, thermal)
    else:
        write_release(tensile, cyclic, thermal)
        print(
            json.dumps(
                {
                    "failure_test_runs": len(tensile),
                    "cyclic_test_runs": int(cyclic["test_run_id"].nunique()),
                    "cycle_endpoints": len(cyclic),
                    "tga_curves": len(thermal),
                },
                ensure_ascii=False,
            )
        )


if __name__ == "__main__":
    main()
