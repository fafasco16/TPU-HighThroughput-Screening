"""物化Figshare碳酸酯TPU强韧与自愈力学数据。"""

from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "数据" / "原始" / "外部数据" / "新增开放数据" / "Figshare_碳酸酯TPU强韧自愈"
SOURCE = SOURCE_DIR / "Source-data_Main Figures.xlsx"
SOURCE_MANIFEST = SOURCE_DIR / "来源清单.json"
DIRECTED = ROOT / "结果" / "定向筛选"
SUMMARY = DIRECTED / "Figshare强韧自愈端点.csv"
CURVES = DIRECTED / "Figshare强韧自愈曲线.csv.gz"
MANIFEST = DIRECTED / "Figshare强韧自愈发布清单.json"
RELEASE_ID = "tpu-figshare-mechano-responsive-2021-v1"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _entry(path: Path) -> dict[str, object]:
    return {"path": str(path.relative_to(ROOT)).replace("\\", "/"), "bytes": path.stat().st_size, "sha256": _sha256(path)}


def build_release() -> tuple[pd.DataFrame, pd.DataFrame]:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Figure 1b"]
    rows = list(sheet.iter_rows(values_only=True))
    specs = [(1, 2, "E-IP-SS", "Virgin"), (3, 4, "Es-MD", "Virgin"), (5, 6, "C-IP-SS", "Virgin"), (7, 8, "C-IP-SS", "Healed_1h"), (9, 10, "C-IP-SS", "Healed_6h"), (11, 12, "C-IP-SS", "Healed_24h"), (13, 14, "C-IP-SS", "Healed_48h")]
    curve_rows = []
    for strain_col, stress_col, material, state in specs:
        curve_id = f"figshare12936989_{material}_{state}"
        point = 0
        for row in rows[2:]:
            if strain_col >= len(row) or stress_col >= len(row):
                continue
            strain, stress = row[strain_col], row[stress_col]
            if not isinstance(strain, (int, float)) or not isinstance(stress, (int, float)):
                continue
            point += 1
            curve_rows.append({"release_id": RELEASE_ID, "curve_id": curve_id, "material_code": material, "state": state, "point_index": point, "strain_percent": float(strain), "stress_MPa": float(stress), "source_locator": f"{SOURCE.relative_to(ROOT).as_posix()}#sheet=Figure 1b", "license": "CC-BY-4.0", "citation_keys": "reference-180"})
    curves = pd.DataFrame(curve_rows)
    summary_values = {
        ("C-IP-SS", "Virgin"): (42.88, 480.0, 75.054),
        ("E-IP-SS", "Virgin"): (6.76, 923.0, 26.93),
        ("Es-MD", "Virgin"): (35.8, 880.0, 115.0),
        ("C-IP-SS", "Healed"): (33.09, 397.0, 48.348),
        ("E-IP-SS", "Healed"): (5.96, 919.0, 20.75),
    }
    records = []
    for (material, state), values in summary_values.items():
        strength, elongation, toughness = values
        virgin = summary_values[(material, "Virgin")]
        records.append({"release_id": RELEASE_ID, "source_id": "source_figshare_12936989_v1", "formulation_id": material, "material_code": material, "state": state, "tensile_strength_MPa": strength, "elongation_at_break_percent": elongation, "toughness_MJ_m3": toughness, "strength_retention_percent": 100 * strength / virgin[0], "elongation_retention_percent": 100 * elongation / virgin[1], "toughness_retention_percent": 100 * toughness / virgin[2], "chemistry_mapping_status": "material_code_only_structure_mapping_pending", "usage_mode": "auxiliary_train", "source_locator": f"{SOURCE.relative_to(ROOT).as_posix()}#sheet=Figure 1c", "license": "CC-BY-4.0", "citation_keys": "reference-180"})
    workbook.close()
    return pd.DataFrame(records), curves


def write_release(summary: pd.DataFrame, curves: pd.DataFrame) -> None:
    summary.to_csv(SUMMARY, index=False, encoding="utf-8-sig", lineterminator="\n")
    curves.to_csv(CURVES, index=False, encoding="utf-8", lineterminator="\n", compression={"method": "gzip", "mtime": 0})
    payload = {"release_id": RELEASE_ID, "counts": {"summary_rows": len(summary), "stress_strain_curve_count": curves["curve_id"].nunique(), "curve_point_rows": len(curves)}, "source": {"doi": "10.6084/m9.figshare.12936989.v1", "license": "CC-BY-4.0", "workbook": _entry(SOURCE), "source_manifest": _entry(SOURCE_MANIFEST)}, "outputs": {"summary": _entry(SUMMARY), "curves": _entry(CURVES)}}
    MANIFEST.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def check_release(summary: pd.DataFrame, curves: pd.DataFrame) -> None:
    with tempfile.TemporaryDirectory(prefix="figshare-tpu-check-") as directory:
        tmp = Path(directory)
        s, c = tmp / SUMMARY.name, tmp / CURVES.name
        summary.to_csv(s, index=False, encoding="utf-8-sig", lineterminator="\n")
        curves.to_csv(c, index=False, encoding="utf-8", lineterminator="\n", compression={"method": "gzip", "mtime": 0})
        if _sha256(s) != _sha256(SUMMARY) or _sha256(c) != _sha256(CURVES):
            raise SystemExit("Figshare强韧自愈输出不一致")
    print("Figshare强韧自愈数据检查通过")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--检查", action="store_true")
    args = parser.parse_args()
    summary, curves = build_release()
    if args.检查:
        check_release(summary, curves)
    else:
        write_release(summary, curves)
        print(
            json.dumps(
                {
                    "summary_rows": len(summary),
                    "curve_count": curves["curve_id"].nunique(),
                    "curve_points": len(curves),
                },
                ensure_ascii=False,
            )
        )


if __name__ == "__main__":
    main()
