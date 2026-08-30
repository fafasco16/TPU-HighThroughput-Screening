"""物化Zenodo多孔导电TPU纳米复合膜的25条拉伸曲线端点。"""

from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from 接入DRUM机械回收 import _derive_endpoints


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = (
    ROOT
    / "数据"
    / "原始"
    / "外部数据"
    / "新增开放数据"
    / "Zenodo_多孔导电TPU纳米复合膜"
)
SOURCE = SOURCE_DIR / "Supronics_Porous-TPU-Nanocomposites Dataset.xlsx"
SOURCE_MANIFEST = SOURCE_DIR / "来源清单.json"
OUTPUT = ROOT / "结果" / "定向筛选" / "Zenodo多孔TPU拉伸端点.csv"
MANIFEST = ROOT / "结果" / "定向筛选" / "Zenodo多孔TPU发布清单.json"
RELEASE_ID = "tpu-zenodo-porous-conductive-2017-v1"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _entry(path: Path) -> dict[str, object]:
    return {
        "path": str(path.relative_to(ROOT)).replace("\\", "/"),
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def build_release() -> pd.DataFrame:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["2. Tensile Properties"]
    rows = list(sheet.iter_rows(values_only=True))
    material_starts = {4: "TPU", 15: "TPU10", 26: "TPU20", 37: "TPU30", 48: "TPU40"}
    records = []
    source_hash = _sha256(SOURCE)
    for start, material in material_starts.items():
        for replicate in range(1, 6):
            strain_index = start + 2 * (replicate - 1)
            stress_index = strain_index + 1
            values = [
                (
                    row[strain_index] if strain_index < len(row) else None,
                    row[stress_index] if stress_index < len(row) else None,
                )
                for row in rows[5:]
            ]
            curve = pd.DataFrame(values, columns=["strain", "stress"])
            curve = curve.apply(pd.to_numeric, errors="coerce").dropna()
            if len(curve) < 4:
                raise ValueError(f"拉伸曲线点不足：{material} replicate {replicate}")
            endpoints = _derive_endpoints(
                curve.rename(columns={"strain": "strain", "stress": "stress"})[
                    ["stress", "strain"]
                ]
            )
            records.append(
                {
                    "release_id": RELEASE_ID,
                    "source_id": "source_zenodo_1098206_v1",
                    "source_family_id": "family_porous_conductive_tpu_2017",
                    "observation_id": f"zenodo1098206_{material.lower()}_r{replicate}",
                    "formulation_id": material,
                    "sample_id": f"{material}_replicate_{replicate}",
                    "replicate_index": replicate,
                    "sample_code_level": int(material.replace("TPU", "") or 0),
                    "sample_code_level_interpretation": (
                        "numeric_sample_code_not_asserted_as_weight_fraction"
                    ),
                    "polymer_family": "TPU_nanocomposite",
                    "chemistry_mapping_status": "commercial_TPU_identity_unresolved",
                    "usage_mode": "auxiliary_train",
                    "recommended_loss_weight_ceiling": 0.5,
                    **endpoints,
                    "test_system": "Lloyd universal testing machine (Ametek Inc.)",
                    "specimen_geometry": "rectangular_50x2x2_mm",
                    "tensile_speed_mm_min": 100.0,
                    "grip_distance_mm": 30.0,
                    "source_file_sha256": source_hash,
                    "source_locator": (
                        f"{SOURCE.relative_to(ROOT).as_posix()}"
                        f"#sheet=2. Tensile Properties;material={material};replicate={replicate}"
                    ),
                    "license": "CC-BY-4.0",
                    "citation_keys": "reference-181",
                }
            )
    workbook.close()
    return pd.DataFrame(records).sort_values("observation_id").reset_index(drop=True)


def _write_csv(frame: pd.DataFrame, path: Path) -> None:
    frame.to_csv(path, index=False, encoding="utf-8-sig", lineterminator="\n")


def _counts(frame: pd.DataFrame) -> dict[str, int]:
    return {
        "tensile_curve_rows": len(frame),
        "formulation_count": frame["formulation_id"].nunique(),
        "replicates_per_formulation": int(
            frame.groupby("formulation_id")["replicate_index"].nunique().min()
        ),
    }


def _manifest(frame: pd.DataFrame, output: Path) -> dict[str, object]:
    return {
        "release_id": RELEASE_ID,
        "counts": _counts(frame),
        "source": {
            "doi": "10.1038/s41598-017-17647-w",
            "zenodo_record": "https://zenodo.org/records/1098206",
            "license": "CC-BY-4.0",
            "workbook": _entry(SOURCE),
            "source_manifest": _entry(SOURCE_MANIFEST),
        },
        "output_file": _entry(output),
        "mapping_limit": (
            "TPU/TPU10/TPU20/TPU30/TPU40 are preserved as source sample codes; "
            "the numeric suffix is not asserted as a weight fraction"
        ),
    }


def write_release(frame: pd.DataFrame) -> None:
    _write_csv(frame, OUTPUT)
    MANIFEST.write_text(
        json.dumps(_manifest(frame, OUTPUT), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def check_release(frame: pd.DataFrame) -> None:
    if not OUTPUT.is_file() or not MANIFEST.is_file():
        raise SystemExit("缺少Zenodo多孔TPU发布；请先运行生成模式")
    with tempfile.TemporaryDirectory(prefix="zenodo-porous-tpu-check-") as directory:
        candidate = Path(directory) / OUTPUT.name
        _write_csv(frame, candidate)
        if _sha256(candidate) != _sha256(OUTPUT):
            raise SystemExit("Zenodo多孔TPU端点与当前原件或算法不一致")
    if json.loads(MANIFEST.read_text(encoding="utf-8")) != _manifest(frame, OUTPUT):
        raise SystemExit("Zenodo多孔TPU发布清单不一致")
    print("Zenodo多孔TPU拉伸端点检查通过")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--检查", action="store_true")
    args = parser.parse_args()
    frame = build_release()
    if args.检查:
        check_release(frame)
    else:
        write_release(frame)
        print(json.dumps(_counts(frame), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
